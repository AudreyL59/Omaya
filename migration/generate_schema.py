"""Generateur de schema PostgreSQL depuis les descriptifs WinDev (.xlsx).

Entree  : <input>/<base>/Table <Nom>.xlsx
          (1 sous-dossier par base HFSQL = 1 schema PostgreSQL)
Sorties : <out>/<base>.sql        (CREATE SCHEMA + CREATE TABLE pgt_*)
          <mapping>               (CSV correspondance HFSQL -> snake_case)
          + rapport console (types inconnus, PK ambigues, ALL-CAPS a revoir)

Conventions (cf. memory project_db_strategy + migration/schema/ticket.sql) :
  - noms snake_case ; tables prefixees "pgt_" (PG target)
  - IDs 8 octets -> bigint ; PK = cle metier (pas l'Identifiant auto)
  - dates -> date/timestamp ; memos binaires -> bytea ; memos texte -> text
  - PAS de FK (HFSQL ne les impose pas) ; index sur les "cles avec doublon"

Usage :
  python migration/generate_schema.py
  python migration/generate_schema.py --input "D:\\Claude\\Table HFSQL" \
      --out migration/schema --mapping migration/mapping/columns.csv
"""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from pathlib import Path

# ----------------------------------------------------------------------------
#  Exceptions de nommage : ALL-CAPS composes que la regle ne sait pas couper.
#  A enrichir au fur et a mesure (le rapport signale les candidats).
# ----------------------------------------------------------------------------
OVERRIDES: dict[str, str] = {
    # Composes ALL-CAPS coupes manuellement (revue 2026-05-27).
    # Les abreviations trop ambigues restent collees (DNAISS->dnaiss, LNAISS->lnaiss).
    "DATECREA": "date_crea",
    "OPCREA": "op_crea",
    "OPDEST": "op_dest",
    "OPSAISIE": "op_saisie",
    "DATEPAIEMENT": "date_paiement",
    "RECUDATE": "recu_date",
    "DATEFIN": "date_fin",
    "DATEDEB": "date_deb",
    "MUTDATE": "mut_date",
    "DATENAISS": "date_naiss",
    "DEPNAISS": "dep_naiss",
    "ADRBAT": "adr_bat",
    "TELGSM1": "tel_gsm1",
    "URGNOM": "urg_nom",
    "URGLIEN": "urg_lien",
    "URGTEL": "urg_tel",
    "FORFAITKM": "forfait_km",
    "KMMENSUEL": "km_mensuel",
    "ALERTEREL": "alerte_rel",
    "FORMEJURI": "forme_juri",
    "NUMSS": "num_ss",
    "NUMCIN": "num_cin",
    "SSFAM": "ss_fam",
}


# Mots reserves PostgreSQL interdits comme identifiant non quote -> suffixe '_'
# (on garde snake_case sans quoting ; ex. desc -> desc_, order -> order_).
PG_RESERVED: set[str] = {
    "all", "analyse", "analyze", "and", "any", "array", "as", "asc",
    "asymmetric", "authorization", "binary", "both", "case", "cast", "check",
    "collate", "collation", "column", "concurrently", "constraint", "create",
    "cross", "current_catalog", "current_date", "current_role", "current_schema",
    "current_time", "current_timestamp", "current_user", "default", "deferrable",
    "desc", "distinct", "do", "else", "end", "except", "false", "fetch", "for",
    "foreign", "freeze", "from", "full", "grant", "group", "having", "ilike",
    "in", "initially", "inner", "intersect", "into", "is", "isnull", "join",
    "lateral", "leading", "left", "like", "limit", "localtime", "localtimestamp",
    "natural", "not", "notnull", "null", "offset", "on", "only", "or", "order",
    "outer", "overlaps", "placing", "primary", "references", "returning", "right",
    "select", "session_user", "similar", "some", "symmetric", "system_user",
    "table", "tablesample", "then", "to", "trailing", "true", "union", "unique",
    "user", "using", "variadic", "verbose", "when", "where", "window", "with",
}


def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )


def to_snake(name: str) -> str:
    """Nom HFSQL -> snake_case (sans accents, prefixe ID isole)."""
    raw = (name or "").strip()
    if raw in OVERRIDES:
        return OVERRIDES[raw]
    s = strip_accents(raw)
    prefix = ""
    # Prefixe "ID" traite comme token : IDSalarie -> id_salarie,
    # IDsalarie_progevo -> id_salarie_progevo (ID + lettre/underscore).
    if s[:2] == "ID" and len(s) > 2 and (s[2].isalpha() or s[2] == "_"):
        prefix = "id_"
        s = s[2:].lstrip("_")
    s = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", s)   # xY... -> x_Y...
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s)   # aB -> a_B
    s = s.replace("-", "_").replace(" ", "_")
    out = (prefix + s).lower()
    out = re.sub(r"_+", "_", out).strip("_")
    if out in PG_RESERVED:   # desc -> desc_, order -> order_, user -> user_...
        out += "_"
    return out


def pg_type(type_str: str, size) -> tuple[str, bool]:
    """(type PG, est_inconnu). type_str = colonne 'Type' du descriptif."""
    t = strip_accents(type_str or "").lower().strip()
    sz = int(size) if str(size).strip().isdigit() else 0
    if "identifiant automatique" in t:
        return "bigint", False
    if "memo" in t and any(k in t for k in ("binaire", "image", "son", "ole")):
        return "bytea", False
    if "memo" in t:
        return "text", False
    if "booleen" in t:
        return "boolean", False
    if "date+heure" in t or "dateheure" in t or "date + heure" in t:
        return "timestamp", False
    if t == "date":
        return "date", False
    if t == "heure":
        return "time", False
    if "monetaire" in t:
        return "numeric(19,4)", False
    if "numerique" in t or "decimal" in t:
        return "numeric", False
    if "reel" in t or "flottant" in t:
        return "double precision", False
    if "unicode" in t:   # "Texte Unicode" HFSQL
        return (f"varchar({sz})" if sz > 0 else "text"), False
    if "uuid" in t:      # WinDev "UUID (256 bits)" -> pas le type uuid PG (128 bits)
        return "varchar(64)", False
    if "entier" in t:
        if sz <= 2:
            return "smallint", False
        if sz <= 4:
            return "integer", False
        return "bigint", False
    if "texte" in t or "chaine" in t:
        return (f"varchar({sz})" if sz > 0 else "text"), False
    return "text", True  # inconnu -> fallback + signale


# ----------------------------------------------------------------------------
#  Lecture d'un descriptif .xlsx
# ----------------------------------------------------------------------------
def read_descriptor(path: Path) -> list[dict]:
    """Retourne [{name, type, size, key}] pour chaque colonne de la table."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # En-tete : reperer les colonnes par leur libelle (accent/casse-insensible)
    header = [strip_accents(str(c or "")).lower().strip() for c in rows[0]]

    def col_idx(*labels: str) -> int:
        for i, h in enumerate(header):
            if any(lbl in h for lbl in labels):
                return i
        return -1

    i_name = col_idx("rubrique")
    i_type = col_idx("type") if col_idx("type") != col_idx("type de cle") else -1
    # 'type' peut matcher 'type de cle' : on prend la 1ere colonne 'type' pure
    i_type = next(
        (i for i, h in enumerate(header) if h == "type" or h.startswith("type ")
         and "cle" not in h),
        col_idx("type"),
    )
    i_size = col_idx("taille")
    i_key = col_idx("type de cle", "cle")
    cols = []
    for r in rows[1:]:
        if i_name < 0 or i_name >= len(r):
            continue
        nm = r[i_name]
        if nm in (None, ""):
            continue
        cols.append({
            "name": str(nm).strip(),
            "type": str(r[i_type]).strip() if 0 <= i_type < len(r) and r[i_type] else "",
            "size": r[i_size] if 0 <= i_size < len(r) else "",
            "key": strip_accents(str(r[i_key]).lower()) if 0 <= i_key < len(r) and r[i_key] else "",
        })
    return cols


def table_name_from_file(path: Path) -> str:
    stem = path.stem  # "Table TK_Liste"
    return re.sub(r"^Table\s+", "", stem).strip()


def folder_to_schema(folder: str) -> str:
    """Nom de dossier (= base HFSQL) -> nom de schema PG = cle logique du
    code. 'Bdd_Omaya_Ticket_BO' -> 'ticket_bo' ; 'ulease' -> 'ulease'."""
    return re.sub(r"^Bdd_Omaya_", "", folder, flags=re.IGNORECASE).lower()


# ----------------------------------------------------------------------------
#  Generation du DDL d'une table
# ----------------------------------------------------------------------------
def build_table(schema: str, hf_table: str, cols: list[dict],
                report: list[str], map_rows: list[dict]) -> str:
    pg_table = "pgt_" + to_snake(hf_table)
    auto_col = next((c for c in cols if "identifiant automatique"
                     in strip_accents(c["type"]).lower()), None)
    unique_cols = [c for c in cols if "cle unique" in c["key"]]
    non_auto_unique = [c for c in unique_cols if c is not auto_col]

    if len(non_auto_unique) == 1:
        pk = non_auto_unique[0]
    elif non_auto_unique:
        # plusieurs : prefere ID<Table>, sinon la 1ere
        want = "ID" + hf_table
        pk = next((c for c in non_auto_unique if c["name"].lower() == want.lower()),
                  non_auto_unique[0])
        report.append(f"[PK ambigue] {hf_table}: {[c['name'] for c in non_auto_unique]}"
                      f" -> choix {pk['name']}")
    elif auto_col:
        pk = auto_col
    else:
        # Pas de cle unique declaree : on deduit la PK de la 1ere colonne si
        # c'est un identifiant "ID<...>" (cas des tables ou l'auto-id n'est pas
        # marque unique dans le descriptif). Sinon table sans PK.
        c0 = cols[0] if cols else None
        if c0 and c0["name"].lower().startswith("id") and len(c0["name"]) > 2:
            pk = c0
            report.append(f"[PK deduite] {hf_table}: {c0['name']} (1ere colonne ID, non marquee unique)")
        else:
            pk = None
            report.append(f"[PK absente] {hf_table}: aucune cle -> table sans PK")

    width = max((len(to_snake(c["name"])) for c in cols), default=10)
    items: list[tuple[str, str]] = []   # (sql colonne/contrainte, commentaire HFSQL)
    indexed = []
    for c in cols:
        sc = to_snake(c["name"])
        pgt, unknown = pg_type(c["type"], c["size"])
        if unknown:
            report.append(f"[type inconnu] {hf_table}.{c['name']} : '{c['type']}' -> text")
        # ALL-CAPS compose non override potentiellement mal coupe
        if (c["name"].isupper() and len(c["name"]) > 4
                and "_" not in c["name"] and c["name"] not in OVERRIDES
                and "_" not in sc):
            report.append(f"[ALL-CAPS a verifier] {hf_table}.{c['name']} -> {sc}")
        notnull = " NOT NULL" if (pk and c is pk) else ""
        items.append((f"{sc:<{width}}  {pgt}{notnull}", c["name"]))
        # Index : "cles avec doublon" metier + modif_date (synchro). On
        # exclut modif_op/modif_elem (housekeeping, jamais filtres seuls).
        if (("cle avec doublon" in c["key"]) or sc == "modif_date") \
                and sc not in ("modif_op", "modif_elem"):
            indexed.append(sc)
        map_rows.append({
            "schema": schema, "hfsql_table": hf_table, "pg_table": pg_table,
            "hfsql_column": c["name"], "pg_column": sc, "pg_type": pgt,
            "pk": "1" if (pk and c is pk) else "",
        })
    if pk:
        items.append((f"CONSTRAINT pk_{pg_table} PRIMARY KEY ({to_snake(pk['name'])})", ""))
    if auto_col and auto_col is not pk:
        items.append((f"CONSTRAINT uq_{pg_table}_auto UNIQUE ({to_snake(auto_col['name'])})", ""))

    # Assemble : virgule sur tous les items sauf le dernier (colonnes ET
    # contraintes), commentaire APRES la virgule -> pas de virgule trainante.
    lines = [f"CREATE TABLE {schema}.{pg_table} ("]
    for i, (part, comment) in enumerate(items):
        comma = "," if i < len(items) - 1 else ""
        line = f"    {part}{comma}"
        if comment:
            line += f"  -- {comment}"
        lines.append(line)
    lines.append(");")
    sql = "\n".join(lines)
    # Index
    idx_sql = []
    for sc in dict.fromkeys(indexed):  # dedup en gardant l'ordre
        idx_sql.append(f"CREATE INDEX ix_{pg_table}_{sc} ON {schema}.{pg_table} ({sc});")
    if idx_sql:
        sql += "\n" + "\n".join(idx_sql)
    return sql


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default=r"D:\Claude\Table HFSQL")
    ap.add_argument("--out", default="migration/schema")
    ap.add_argument("--mapping", default="migration/mapping/columns.csv")
    args = ap.parse_args()

    root = Path(args.input)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    map_path = Path(args.mapping)
    map_path.parent.mkdir(parents=True, exist_ok=True)

    report: list[str] = []
    map_rows: list[dict] = []
    n_tables = 0

    bases = sorted(d for d in root.iterdir() if d.is_dir())
    if not bases:
        print(f"Aucun sous-dossier (base) dans {root}. "
              "Range les .xlsx par base (1 dossier/base).")
        return

    for base_dir in bases:
        schema = folder_to_schema(base_dir.name)
        files = sorted(base_dir.glob("Table *.xlsx"))
        if not files:
            continue
        blocks = [f"CREATE SCHEMA IF NOT EXISTS {schema};\n"]
        for f in files:
            hf_table = table_name_from_file(f)
            cols = read_descriptor(f)
            if not cols:
                report.append(f"[vide] {f.name}")
                continue
            blocks.append(build_table(schema, hf_table, cols, report, map_rows))
            n_tables += 1
        (out_dir / f"{schema}.sql").write_text(
            "\n\n".join(blocks) + "\n", encoding="utf-8")
        print(f"  {schema}: {len(files)} tables -> {schema}.sql")

    # CSV de correspondance
    with map_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=[
            "schema", "hfsql_table", "pg_table", "hfsql_column",
            "pg_column", "pg_type", "pk"])
        w.writeheader()
        w.writerows(map_rows)

    print(f"\n{n_tables} tables, {len(map_rows)} colonnes -> {map_path}")
    if report:
        print(f"\n--- A REVOIR ({len(report)}) ---")
        for line in report:
            print("  " + line)


if __name__ == "__main__":
    main()
