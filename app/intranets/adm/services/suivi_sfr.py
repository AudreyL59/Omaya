"""Services pour Fen_SuiviSFR (sous-fonctionnalites ADM > Suivi SFR).

1. Fen_SFRCttaRacc (Ctts à raccorder) :
   - list_ctts_a_raccorder : SELECT JOIN avec preselect Choix
   - send_mails_to_bos : envoi mail aux BO SFR cluster pour les lignes cochees
"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from typing import Optional

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection


def _new_id() -> int:
    """ID 8 octets timestamp (cf idEntierDateHeureSys WinDev)."""
    n = datetime.now()
    return int(n.strftime("%Y%m%d%H%M%S")) * 1000 + n.microsecond // 1000


# ====================================================================
# 1. CTTS A RACCORDER (Fen_SFRCttaRacc)
# ====================================================================


class CttARaccorderItem(BaseModel):
    id_contrat: str
    num_bs: str = ""
    nom: str = ""
    prenom: str = ""
    cp: str = ""
    ville: str = ""
    code_vad: str = ""
    nom_cluster: str = ""
    nom_sa: str = ""               # nom commercial
    prenom_sa: str = ""
    date_signature: str = ""
    date_validation: str = ""
    date_raccordement: str = ""
    date_rdv_tech: str = ""
    lib_etat: str = ""
    type_etat: int = 0
    type_install: int = 0           # SelfInstall
    type_offre: int = 0             # IDproduit
    type_vente: int = 0
    technologie: int = 0
    option_dec: str = ""
    box8: bool = False
    nb_pts_payes: float = 0.0
    mail_bo_envoye: bool = False
    mail_bo_date_envoi: str = ""
    choix: int = 0                  # 1 = preselect a envoyer mail


def _capitalize(p: str) -> str:
    return (p[:1].upper() + p[1:].lower()) if p else ""


def _date_str(d) -> str:
    if not d:
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%Y-%m-%d")
    return str(d)


def list_ctts_a_raccorder(
    du: date, au: date,
) -> list[CttARaccorderItem]:
    """Charge les contrats SFR entre Du et Au avec un etat 'planifie' ou
    'paye par employeur'. Pre-coche Choix=1 si :
      - lib_etat = 'Le raccordement est planifié dans le créneau - '
      - ET date_rdv_tech vide
      - ET (mail_bo_envoye=False OU mail_bo_date_envoi > 7 jours)
    """
    db = get_pg_connection("adv")
    rows = db.query(
        """
        SELECT c.id_contrat, c.num_bs, c.date_signature, c.date_validation,
               c.date_racc_activ, c.date_rdv_tech, c.id_sfr_cluster,
               c.id_etat_sfr, c.technologie, c.id_produit, c.self_install,
               c.type_vente, c.box8, c.option_dec, c.nb_pts_payes_va,
               c.mail_bo_envoye, c.mail_bo_date_envoi, c.info_interne,
               cli.nom AS cli_nom, cli.prenom AS cli_prenom,
               cli.cp AS cli_cp, cli.ville AS cli_ville,
               clu.code_vad, clu.nom_cluster,
               s.nom AS sa_nom, s.prenom AS sa_prenom,
               e.lib_etat, e.id_type_etat
          FROM adv.pgt_sfr_contrat c
          LEFT JOIN adv.pgt_client cli ON cli.id_client = c.id_client
          LEFT JOIN adv.pgt_sfr_cluster clu ON clu.id_sfr_cluster = c.id_sfr_cluster
          LEFT JOIN adv.pgt_sfr_etat_contrat e ON e.id_etat = c.id_etat_sfr
          LEFT JOIN rh.pgt_salarie s ON s.id_salarie = c.id_salarie
         WHERE c.date_signature BETWEEN ? AND ?
           AND (c.modif_elem IS NULL OR c.modif_elem NOT LIKE '%suppr%')
           AND (
                e.lib_etat ILIKE '%Le raccordement est planifié%'
             OR e.lib_etat ILIKE '%Payé par employeur - Validation%'
             OR c.date_rdv_tech BETWEEN ? AND ?
           )
         ORDER BY c.date_signature DESC
         LIMIT 5000
        """,
        (du, au, du, au),
    ) or []

    today = date.today()
    out: list[CttARaccorderItem] = []
    for r in rows:
        lib_etat = r.get("lib_etat") or ""
        date_rdv = _date_str(r.get("date_rdv_tech"))
        mail_envoye = bool(r.get("mail_bo_envoye"))
        mail_date = r.get("mail_bo_date_envoi")

        choix = 0
        # Preselect Choix=1 si etat exact 'Le raccordement est planifié...'
        # ET pas de RDV Tech ET (pas encore envoye OU envoye > 7j)
        if "Le raccordement est planifié" in lib_etat and not date_rdv:
            if not mail_envoye:
                choix = 1
            elif mail_date and isinstance(mail_date, (date, datetime)):
                md = mail_date.date() if isinstance(mail_date, datetime) else mail_date
                if (today - md).days > 7:
                    choix = 1

        out.append(CttARaccorderItem(
            id_contrat=str(r["id_contrat"]),
            num_bs=r.get("num_bs") or "",
            nom=(r.get("cli_nom") or "").strip(),
            prenom=_capitalize((r.get("cli_prenom") or "").strip()),
            cp=r.get("cli_cp") or "",
            ville=r.get("cli_ville") or "",
            code_vad=r.get("code_vad") or "",
            nom_cluster=r.get("nom_cluster") or "",
            nom_sa=(r.get("sa_nom") or "").strip(),
            prenom_sa=_capitalize((r.get("sa_prenom") or "").strip()),
            date_signature=_date_str(r.get("date_signature")),
            date_validation=_date_str(r.get("date_validation")),
            date_raccordement=_date_str(r.get("date_racc_activ")),
            date_rdv_tech=date_rdv,
            lib_etat=lib_etat,
            type_etat=int(r.get("id_type_etat") or 0),
            type_install=int(r.get("self_install") or 0),
            type_offre=int(r.get("id_produit") or 0),
            type_vente=int(r.get("type_vente") or 0),
            technologie=int(r.get("technologie") or 0),
            option_dec=r.get("option_dec") or "",
            box8=bool(r.get("box8")),
            nb_pts_payes=float(r.get("nb_pts_payes_va") or 0),
            mail_bo_envoye=mail_envoye,
            mail_bo_date_envoi=_date_str(mail_date),
            choix=choix,
        ))
    return out


# --------------------------------------------------------------------
# Envoi mail BO SFR
# --------------------------------------------------------------------


class SendMailsResult(BaseModel):
    id_contrat: str
    num_bs: str = ""
    ok: bool = False
    message: str = ""


# ====================================================================
# 2. REMUNERATIONS SFR (Fen_RemInterneSFR)
# ====================================================================


class RemunItem(BaseModel):
    id_sfr_remun: str
    categorie: str = ""              # 'FIBRE' ou 'MOBILE'
    id_produit: int = 0
    lib_produit: str = ""
    type_vente: int = 0
    date_debut: str = ""
    date_fin: str = ""
    montant_va: float = 0.0
    montant_va_remise: float = 0.0
    montant_ra: float = 0.0
    montant_ra_remise: float = 0.0
    prime_volumique: float = 0.0
    abonnement_tv: float = 0.0
    type_repart_rem: int = 0


class RemunPayload(BaseModel):
    categorie: str                      # 'FIBRE' ou 'MOBILE'
    id_produit: int
    type_vente: int = 0
    date_debut: Optional[date] = None
    date_fin: Optional[date] = None
    montant_va: float = 0.0
    montant_va_remise: float = 0.0
    montant_ra: float = 0.0
    montant_ra_remise: float = 0.0
    prime_volumique: float = 0.0
    abonnement_tv: float = 0.0
    type_repart_rem: int = 0


class ProduitSfrItem(BaseModel):
    id_produit: int
    lib_produit: str


def list_remunerations(categorie: str) -> list[RemunItem]:
    """Liste les rémunérations SFR pour une catégorie ('FIBRE' ou 'MOBILE'),
    JOIN sfr_produit pour le libellé. Tri date_debut DESC, lib_produit ASC.
    """
    db = get_pg_connection("adv")
    rows = db.query(
        """SELECT r.id_sfr_remun, r.categorie, r.id_produit,
                  p.lib_produit,
                  r.type_vente, r.date_debut, r.date_fin,
                  r.montant_va, r.montant_va_remise,
                  r.montant_ra, r.montant_ra_remise,
                  r.prime_volumique, r.abonnement_tv, r.type_repart_rem
             FROM adv.pgt_sfr_remun r
             LEFT JOIN adv.pgt_sfr_produit p ON p.id_produit = r.id_produit
            WHERE r.categorie = ?
              AND (r.modif_elem IS NULL OR r.modif_elem NOT LIKE '%suppr%')
            ORDER BY r.date_debut DESC NULLS LAST, p.lib_produit""",
        (categorie.upper(),),
    ) or []
    return [RemunItem(
        id_sfr_remun=str(r["id_sfr_remun"]),
        categorie=r.get("categorie") or "",
        id_produit=int(r.get("id_produit") or 0),
        lib_produit=r.get("lib_produit") or "",
        type_vente=int(r.get("type_vente") or 0),
        date_debut=_date_str(r.get("date_debut")),
        date_fin=_date_str(r.get("date_fin")),
        montant_va=float(r.get("montant_va") or 0),
        montant_va_remise=float(r.get("montant_va_remise") or 0),
        montant_ra=float(r.get("montant_ra") or 0),
        montant_ra_remise=float(r.get("montant_ra_remise") or 0),
        prime_volumique=float(r.get("prime_volumique") or 0),
        abonnement_tv=float(r.get("abonnement_tv") or 0),
        type_repart_rem=int(r.get("type_repart_rem") or 0),
    ) for r in rows]


def list_sfr_produits(categorie: Optional[str] = None) -> list[ProduitSfrItem]:
    """Liste les produits SFR, optionnellement filtres par categorie.
    Categorie SFR : 1=Fibre, 2=Mobile (smallint pgt_sfr_produit.categorie)."""
    db = get_pg_connection("adv")
    where = ["(modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')",
             "COALESCE(pro_actif, 0) = 1"]
    params: list = []
    if categorie:
        cat_int = 1 if categorie.upper() == "FIBRE" else 2
        where.append("categorie = ?")
        params.append(cat_int)
    rows = db.query(
        f"""SELECT id_produit, lib_produit FROM adv.pgt_sfr_produit
            WHERE {' AND '.join(where)}
            ORDER BY lib_produit""",
        tuple(params),
    ) or []
    return [ProduitSfrItem(
        id_produit=int(r["id_produit"]),
        lib_produit=r.get("lib_produit") or "",
    ) for r in rows]


def get_remun(id_sfr_remun: int) -> Optional[RemunItem]:
    """Lit une remuneration pour edition."""
    db = get_pg_connection("adv")
    r = db.query_one(
        """SELECT r.id_sfr_remun, r.categorie, r.id_produit, p.lib_produit,
                  r.type_vente, r.date_debut, r.date_fin,
                  r.montant_va, r.montant_va_remise,
                  r.montant_ra, r.montant_ra_remise,
                  r.prime_volumique, r.abonnement_tv, r.type_repart_rem
             FROM adv.pgt_sfr_remun r
             LEFT JOIN adv.pgt_sfr_produit p ON p.id_produit = r.id_produit
            WHERE r.id_sfr_remun = ? LIMIT 1""",
        (int(id_sfr_remun),),
    )
    if not r:
        return None
    return RemunItem(
        id_sfr_remun=str(r["id_sfr_remun"]),
        categorie=r.get("categorie") or "",
        id_produit=int(r.get("id_produit") or 0),
        lib_produit=r.get("lib_produit") or "",
        type_vente=int(r.get("type_vente") or 0),
        date_debut=_date_str(r.get("date_debut")),
        date_fin=_date_str(r.get("date_fin")),
        montant_va=float(r.get("montant_va") or 0),
        montant_va_remise=float(r.get("montant_va_remise") or 0),
        montant_ra=float(r.get("montant_ra") or 0),
        montant_ra_remise=float(r.get("montant_ra_remise") or 0),
        prime_volumique=float(r.get("prime_volumique") or 0),
        abonnement_tv=float(r.get("abonnement_tv") or 0),
        type_repart_rem=int(r.get("type_repart_rem") or 0),
    )


def create_remun(p: RemunPayload, op_id: int) -> int:
    db = get_pg_connection("adv")
    id_new = _new_id()
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_sfr_remun_auto), 0) + 1 AS n FROM adv.pgt_sfr_remun"
    )
    auto_n = int(auto["n"]) if auto else 1
    db.query(
        """INSERT INTO adv.pgt_sfr_remun
              (id_sfr_remun_auto, id_sfr_remun, categorie, id_produit,
               type_vente, date_debut, date_fin,
               montant_va, montant_va_remise, montant_ra, montant_ra_remise,
               prime_volumique, abonnement_tv, type_repart_rem,
               modif_date, modif_op, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NOW(), ?, 'new')""",
        (auto_n, id_new, p.categorie.upper(), int(p.id_produit),
         int(p.type_vente), p.date_debut, p.date_fin,
         float(p.montant_va), float(p.montant_va_remise),
         float(p.montant_ra), float(p.montant_ra_remise),
         float(p.prime_volumique), float(p.abonnement_tv),
         int(p.type_repart_rem), int(op_id)),
    )
    return id_new


def update_remun(id_sfr_remun: int, p: RemunPayload, op_id: int) -> bool:
    db = get_pg_connection("adv")
    db.query(
        """UPDATE adv.pgt_sfr_remun
              SET categorie=?, id_produit=?, type_vente=?,
                  date_debut=?, date_fin=?,
                  montant_va=?, montant_va_remise=?,
                  montant_ra=?, montant_ra_remise=?,
                  prime_volumique=?, abonnement_tv=?, type_repart_rem=?,
                  modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_sfr_remun=?""",
        (p.categorie.upper(), int(p.id_produit), int(p.type_vente),
         p.date_debut, p.date_fin,
         float(p.montant_va), float(p.montant_va_remise),
         float(p.montant_ra), float(p.montant_ra_remise),
         float(p.prime_volumique), float(p.abonnement_tv),
         int(p.type_repart_rem), int(op_id), int(id_sfr_remun)),
    )
    return True


def duplicate_remun(id_sfr_remun: int, op_id: int) -> int:
    """Duplique une remuneration : copie tous les champs sauf l'ID."""
    db = get_pg_connection("adv")
    src = db.query_one(
        """SELECT categorie, id_produit, type_vente, date_debut, date_fin,
                  montant_va, montant_va_remise, montant_ra, montant_ra_remise,
                  prime_volumique, abonnement_tv, type_repart_rem
             FROM adv.pgt_sfr_remun WHERE id_sfr_remun = ?""",
        (int(id_sfr_remun),),
    )
    if not src:
        return 0
    id_new = _new_id()
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_sfr_remun_auto), 0) + 1 AS n FROM adv.pgt_sfr_remun"
    )
    auto_n = int(auto["n"]) if auto else 1
    db.query(
        """INSERT INTO adv.pgt_sfr_remun
              (id_sfr_remun_auto, id_sfr_remun, categorie, id_produit,
               type_vente, date_debut, date_fin,
               montant_va, montant_va_remise, montant_ra, montant_ra_remise,
               prime_volumique, abonnement_tv, type_repart_rem,
               modif_date, modif_op, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NOW(), ?, 'new')""",
        (auto_n, id_new, src.get("categorie"), src.get("id_produit"),
         src.get("type_vente"), src.get("date_debut"), src.get("date_fin"),
         src.get("montant_va"), src.get("montant_va_remise"),
         src.get("montant_ra"), src.get("montant_ra_remise"),
         src.get("prime_volumique"), src.get("abonnement_tv"),
         src.get("type_repart_rem"), int(op_id)),
    )
    return id_new


def delete_remun(id_sfr_remun: int, op_id: int) -> bool:
    db = get_pg_connection("adv")
    db.query(
        """UPDATE adv.pgt_sfr_remun
              SET modif_elem='suppr', modif_date=NOW(), modif_op=?
            WHERE id_sfr_remun=?""",
        (int(op_id), int(id_sfr_remun)),
    )
    return True


# ====================================================================
# 3. TICKET CALL SFR (Fen_TicketCallSFR)
# ====================================================================


class TicketCallItem(BaseModel):
    id_tk_call_sfr: str
    id_tk_liste: str
    nb_ctt: int = 0                      # nb lignes panier (NB Ctt dans le panier)
    nb_num_rens: int = 0                 # nb panier avec vrai NumBS (sans 'TK')
    nb_ctt_avec_num: int = 0             # nb lignes panier avec NumBS (TK ou pas)
    contenu_panier: str = ""             # multi-ligne 'TYPE LibOffre NUM (date)'
    date_crea: str = ""
    nom_vendeur: str = ""
    agence: str = ""                     # affectation cf vendeur (TODO)
    equipe: str = ""
    nom_client: str = ""
    prenom_client: str = ""
    cp: str = ""
    ville: str = ""
    nom_operateur: str = ""              # OpeAppel (qui a pris l'appel)
    date_deb_prise_en_charge: str = ""
    date_fin_prise_en_charge: str = ""
    delai_av_prise_charge_min: float = 0.0   # date_deb - date_crea en minutes
    duree_appel_min: float = 0.0         # date_fin - date_deb en minutes
    parcours_chaines: bool = False       # FIBRE + MOBILE dans le panier
    row_color_alert: bool = False        # au moins un panier saisi > 1h apres crea
    lib_statut: str = ""
    cloturee: bool = False
    nb_valide: int = 0                   # nb panier statut=1 ou 3


class AnalyseTrancheItem(BaseModel):
    tranche_horaire: str
    nb_ticket: int = 0
    moins_3min: int = 0
    entre_3_5min: int = 0
    entre_5_7min: int = 0
    plus_de_7min: int = 0


class AnalyseVentesItem(BaseModel):
    delai: str
    ventes_valides: int = 0
    ventes_annulees: int = 0


class AnalyseVentesTotaux(BaseModel):
    pas_encore_statuees: int = 0
    ventes_validees: int = 0
    ventes_annulees: int = 0
    par_delai: list[AnalyseVentesItem] = []


def _delai_label(min_value: float) -> str:
    """Categorie de delai (cf code WinDev) : <3m / 3-5m / 5-7m / >7m."""
    if min_value < 3:
        return "< 3 min"
    if min_value < 5:
        return "Entre 3 et 5 min"
    if min_value < 7:
        return "Entre 5 et 7 min"
    return "> 7 min"


def _load_ticket_call_sfr(
    du: date, au: date, etat: str,
) -> list[dict]:
    """Helper : execute la requete principale TK_CallSFR + JOIN.
    etat : 'ouverts' / 'clotures' / 'tous'."""
    where_cloture = ""
    if etat == "ouverts":
        where_cloture = "AND tl.cloturee = FALSE"
    elif etat == "clotures":
        where_cloture = "AND tl.cloturee = TRUE"

    db = get_pg_connection("ticket_bo")
    rows = db.query(
        f"""
        SELECT DISTINCT
               tc.id_tk_call_sfr, tc.id_tk_liste, tc.id_salarie,
               tc.nom_client, tc.prenom_client,
               tc.cp, tc.ville,
               tc.date_h_appel, tc.ope_appel,
               tc.date_deb_prise_en_charge, tc.date_fin_prise_en_charge,
               tl.date_crea, tl.id_tk_statut, tl.cloturee,
               tl.date_cloture, tl.op_crea,
               ts.lib_statut,
               s.nom AS sa_nom, s.prenom AS sa_prenom,
               sa.nom AS oa_nom, sa.prenom AS oa_prenom
          FROM ticket_bo.pgt_tk_call_sfr tc
          JOIN ticket.pgt_tk_liste tl ON tl.id_tk_liste = tc.id_tk_liste
          LEFT JOIN ticket.pgt_tk_statut ts ON ts.id_tk_statut = tl.id_tk_statut
          LEFT JOIN rh.pgt_salarie s ON s.id_salarie = tl.op_crea
          LEFT JOIN rh.pgt_salarie sa ON sa.id_salarie = tc.ope_appel
         WHERE (tc.modif_elem IS NULL OR tc.modif_elem NOT LIKE '%suppr%')
           -- date_crea est un timestamp : on filtre [du 00:00, au+1 00:00[
           -- pour inclure toute la journee 'au' (sinon BETWEEN du au ne
           -- match que minuit pile quand du = au).
           AND tl.date_crea >= ?
           AND tl.date_crea <  (?::date + INTERVAL '1 day')
           AND tl.op_crea <> 6
           {where_cloture}
         ORDER BY tl.date_crea DESC
         LIMIT 5000
        """,
        (du, au),
    ) or []
    return rows


def _load_paniers_by_tickets(ids_tickets: list[int]) -> dict[int, list[dict]]:
    """Charge tous les paniers d'un coup, retourne dict {id_tk_call_sfr: [paniers]}."""
    if not ids_tickets:
        return {}
    db = get_pg_connection("ticket_bo")
    ids_sql = ",".join(str(i) for i in ids_tickets)
    rows = db.query(
        f"""SELECT p.id_tk_call_sfr_panier, p.id_tk_call_sfr, p.num,
                   p.id_offres_sfr, p.type, p.statut_prod, p.num_date_saisie,
                   o.lib_offre
              FROM ticket_bo.pgt_tk_call_sfr_panier p
              LEFT JOIN adv.pgt_sfr_offres_provad o
                     ON o.id_offres_sfr = p.id_offres_sfr
             WHERE p.id_tk_call_sfr IN ({ids_sql})
               AND (p.modif_elem IS NULL OR p.modif_elem NOT LIKE '%suppr%')""",
    ) or []
    out: dict[int, list[dict]] = {}
    for r in rows:
        out.setdefault(int(r["id_tk_call_sfr"]), []).append(r)
    return out


def _statut_prod_label(s: int) -> str:
    return {1: "Validé", 2: "Annulé", 3: "Num BS ajouté"}.get(int(s or 0), "Pas statué")


def list_ticket_call_sfr(
    du: date, au: date, etat: str = "tous",
) -> list[TicketCallItem]:
    """Onglet Liste : un ligne par ticket TK_CallSFR + resume des paniers."""
    rows = _load_ticket_call_sfr(du, au, etat)
    ids = [int(r["id_tk_call_sfr"]) for r in rows]
    paniers_by_id = _load_paniers_by_tickets(ids)

    out: list[TicketCallItem] = []
    for r in rows:
        id_tc = int(r["id_tk_call_sfr"])
        paniers = paniers_by_id.get(id_tc, [])

        # Compteurs panier
        nb_valide = sum(1 for p in paniers if int(p.get("statut_prod") or 0) in (1, 3))
        nb_avec_num = sum(1 for p in paniers if (p.get("num") or "").strip())

        # Resume panier multiligne + NB Num Rens + parcours chaines + alerte
        # color (au moins un panier saisi > 1h apres crea ticket)
        crea = r.get("date_crea")
        nb_num_rens = 0
        test_fibre = False
        test_mob = False
        alert = False
        contenu_lines = []
        for p in paniers:
            num = (p.get("num") or "").strip()
            tp = (p.get("type") or "").strip()
            lib = (p.get("lib_offre") or "").strip()
            line = f"{tp} {lib} {num}".strip()
            d = p.get("num_date_saisie")
            if d:
                line += f" ({_date_str(d)})"
                if isinstance(d, datetime) and isinstance(crea, datetime):
                    if (d - crea).total_seconds() >= 3600:
                        alert = True
            contenu_lines.append(line)
            # NB num rens : NUM != "" et ne commence pas par 'TK'
            if num and not num.upper().startswith("TK"):
                nb_num_rens += 1
                if tp.upper() == "FIBRE": test_fibre = True
                else: test_mob = True

        # Delai prise en charge (date_deb_prise_en_charge - date_crea) en min
        # NB : code WinDev fait DateHeureDifference(DH_crea, DH_Deb), donc
        # delai = deb - crea (et pas deb - date_h_appel).
        delai = 0.0
        deb = r.get("date_deb_prise_en_charge")
        if isinstance(deb, datetime) and isinstance(crea, datetime):
            delai = (deb - crea).total_seconds() / 60.0
            if delai < 0: delai = 0.0

        # Duree appel (date_fin - date_deb)
        duree_appel = 0.0
        fin = r.get("date_fin_prise_en_charge")
        if isinstance(deb, datetime) and isinstance(fin, datetime):
            duree_appel = (fin - deb).total_seconds() / 60.0
            if duree_appel < 0: duree_appel = 0.0

        # Nom operateur d'appel (OpeAppel)
        nom_ope = " ".join(filter(None, [
            _capitalize((r.get("oa_prenom") or "").strip()),
            (r.get("oa_nom") or "").strip(),
        ]))

        out.append(TicketCallItem(
            id_tk_call_sfr=str(id_tc),
            id_tk_liste=str(r.get("id_tk_liste") or ""),
            nb_ctt=len(paniers),
            nb_num_rens=nb_num_rens,
            nb_ctt_avec_num=nb_avec_num,
            contenu_panier="\n".join(contenu_lines),
            date_crea=_date_str(crea),
            nom_vendeur=" ".join(filter(None, [
                (r.get("sa_nom") or "").strip(),
                _capitalize((r.get("sa_prenom") or "").strip()),
            ])),
            nom_client=(r.get("nom_client") or "").strip(),
            prenom_client=_capitalize((r.get("prenom_client") or "").strip()),
            cp=r.get("cp") or "",
            ville=r.get("ville") or "",
            nom_operateur=nom_ope,
            date_deb_prise_en_charge=_date_str(deb),
            date_fin_prise_en_charge=_date_str(fin),
            delai_av_prise_charge_min=round(delai, 2),
            duree_appel_min=round(duree_appel, 2),
            parcours_chaines=test_fibre and test_mob,
            row_color_alert=alert,
            lib_statut=r.get("lib_statut") or "",
            cloturee=bool(r.get("cloturee")),
            nb_valide=nb_valide,
        ))
    return out


def analyse_tk_call_sfr(
    du: date, au: date, etat: str = "tous",
) -> list[AnalyseTrancheItem]:
    """Onglet Analyse : regroupement par 'tranche horaire' (cree le).
    Format tranche : 'Jjj JJ AAAA HH:mm' (cf WinDev DateHeureVersChaine)."""
    rows = _load_ticket_call_sfr(du, au, etat)
    jours_fr = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]

    buckets: dict[str, AnalyseTrancheItem] = {}
    for r in rows:
        deb = r.get("date_deb_prise_en_charge")
        app = r.get("date_h_appel")
        crea = r.get("date_crea")
        delai = 0.0
        if isinstance(deb, datetime) and isinstance(app, datetime):
            delai = (deb - app).total_seconds() / 60.0
            if delai < 0: delai = 0.0
        if not isinstance(crea, datetime):
            continue

        # Tranche horaire : on tronque a la minute (sans secondes)
        crea_minute = crea.replace(second=0, microsecond=0)
        label = (f"{jours_fr[crea_minute.weekday()]} "
                 f"{crea_minute.strftime('%d %Y %H:%M')}")
        b = buckets.setdefault(label, AnalyseTrancheItem(tranche_horaire=label))
        b.nb_ticket += 1
        if delai < 3: b.moins_3min += 1
        elif delai < 5: b.entre_3_5min += 1
        elif delai < 7: b.entre_5_7min += 1
        else: b.plus_de_7min += 1

    return sorted(buckets.values(), key=lambda x: x.tranche_horaire)


class PlanningRdvItem(BaseModel):
    """Un rendez-vous a afficher sur le planning visuel."""
    titre: str
    contenu: str
    date_debut: str             # ISO YYYY-MM-DD HH:MM:SS
    date_fin: str
    ressource: str              # 'Crea Ticket' OU nom de l'operateur
    couleur: str                # hex code (#86efac / #fde68a / #fdba74 / #fca5a5)
    delai_label: str            # '< 3 min' / 'Entre 3 et 5 min' / etc.
    delai_min: float
    nb_valide: int              # 0 = icone 'pas valide', >0 = icone OK


# Couleurs (cf legend WinDev : Vert / OrangeClair / OrangeFonce / Rouge)
_COULEURS_DELAI = {
    "< 3 min":           "#86efac",   # vert
    "Entre 3 et 5 min":  "#fde68a",   # jaune
    "Entre 5 et 7 min":  "#fdba74",   # orange
    "> 7 min":           "#fca5a5",   # rouge clair
}


class TicketCallPanierItem(BaseModel):
    id_tk_call_sfr_panier: str
    id_offres_sfr: int = 0
    lib_offre: str = ""
    type: str = ""
    type_vente: int = 0
    num: str = ""
    num_date_saisie: str = ""
    portabilite: bool = False
    num_portabilite: str = ""
    num_prise_rio: str = ""
    num_prise_optique: str = ""
    opt_tv: str = ""
    opt_choisies: str = ""
    test_eligibilite: str = ""
    motif_annulation: str = ""
    statut_prod: int = 0
    a_creer: bool = False             # auto cf code WinDev : NUM<>'' AND statut_prod=1


class TicketCallDetail(BaseModel):
    # TK_Liste
    id_tk_liste: str
    date_crea: str = ""
    id_tk_statut: int = 0
    lib_statut: str = ""
    cloturee: bool = False
    date_cloture: str = ""
    date_report: str = ""
    op_crea: int = 0
    op_crea_nom: str = ""
    # TK_CallSFR (client)
    id_tk_call_sfr: str = ""
    id_salarie: int = 0
    id_salarie_nom: str = ""
    nom_client: str = ""
    prenom_client: str = ""
    nom_marital_client: str = ""
    civilite_client: int = 0
    date_naiss: str = ""
    dep_naiss: str = ""
    adresse1: str = ""
    adresse2: str = ""
    cp: str = ""
    ville: str = ""
    mobile1: str = ""
    adr_mail: str = ""
    type_logement: int = 0
    opt_rappel: bool = False
    opt_partenaire: bool = False
    intervention_vend: bool = False
    info_vente: str = ""
    ref_appel: str = ""
    motif_annulation: str = ""
    code_valid: str = ""
    # Lignes panier
    paniers: list[TicketCallPanierItem] = []


def get_ticket_call_detail(id_tk_liste: int) -> Optional[TicketCallDetail]:
    """Charge le detail complet du ticket : TK_Liste + TK_CallSFR + paniers."""
    db_tk = get_pg_connection("ticket")
    db_bo = get_pg_connection("ticket_bo")
    db_rh = get_pg_connection("rh")

    # TK_Liste
    tl = db_tk.query_one(
        """SELECT tl.id_tk_liste, tl.date_crea, tl.id_tk_statut,
                  tl.cloturee, tl.date_cloture, tl.date_report, tl.op_crea,
                  ts.lib_statut
             FROM ticket.pgt_tk_liste tl
             LEFT JOIN ticket.pgt_tk_statut ts ON ts.id_tk_statut = tl.id_tk_statut
            WHERE tl.id_tk_liste = ? LIMIT 1""",
        (int(id_tk_liste),),
    )
    if not tl:
        return None

    # TK_CallSFR
    tc = db_bo.query_one(
        """SELECT id_tk_call_sfr, id_salarie, nom_client, prenom_client,
                  nom_marital_client, civilite_client, date_naiss, dep_naiss,
                  adresse1, adresse2, cp, ville, mobile1, adr_mail,
                  type_logement, opt_rappel, opt_partenaire,
                  intervention_vend, info_vente, ref_appel,
                  motif_annulation, code_valid
             FROM ticket_bo.pgt_tk_call_sfr
            WHERE id_tk_liste = ?
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            LIMIT 1""",
        (int(id_tk_liste),),
    ) or {}

    # Resolution noms salaries (op_crea + id_salarie du call)
    op_crea = int(tl.get("op_crea") or 0)
    id_sal = int(tc.get("id_salarie") or 0)
    nom_op = ""; nom_sal = ""
    ids_needed = []
    if op_crea: ids_needed.append(op_crea)
    if id_sal and id_sal != op_crea: ids_needed.append(id_sal)
    if ids_needed:
        ids_sql = ",".join(str(i) for i in ids_needed)
        sals = db_rh.query(
            f"SELECT id_salarie, nom, prenom FROM rh.pgt_salarie "
            f"WHERE id_salarie IN ({ids_sql})"
        ) or []
        smap = {int(s["id_salarie"]): s for s in sals}
        if op_crea and op_crea in smap:
            s = smap[op_crea]
            nom_op = f"{(s.get('nom') or '').strip()} {_capitalize((s.get('prenom') or '').strip())}".strip()
        if id_sal and id_sal in smap:
            s = smap[id_sal]
            nom_sal = f"{(s.get('nom') or '').strip()} {_capitalize((s.get('prenom') or '').strip())}".strip()

    # Paniers
    paniers: list[TicketCallPanierItem] = []
    id_tc = int(tc.get("id_tk_call_sfr") or 0)
    if id_tc:
        rows = db_bo.query(
            """SELECT p.id_tk_call_sfr_panier, p.id_offres_sfr, p.type,
                      p.type_vente, p.num, p.num_date_saisie,
                      p.portabilite, p.num_portabilite, p.num_prise_rio,
                      p.num_prise_optique, p.opt_tv, p.opt_choisies,
                      p.test_eligibilite, p.motif_annulation, p.statut_prod,
                      o.lib_offre
                 FROM ticket_bo.pgt_tk_call_sfr_panier p
                 LEFT JOIN adv.pgt_sfr_offres_provad o
                        ON o.id_offres_sfr = p.id_offres_sfr
                WHERE p.id_tk_call_sfr = ?
                  AND (p.modif_elem IS NULL OR p.modif_elem NOT LIKE '%suppr%')
                ORDER BY p.id_tk_call_sfr_panier""",
            (id_tc,),
        ) or []
        for r in rows:
            num = (r.get("num") or "").strip()
            statut_p = int(r.get("statut_prod") or 0)
            paniers.append(TicketCallPanierItem(
                id_tk_call_sfr_panier=str(r["id_tk_call_sfr_panier"]),
                id_offres_sfr=int(r.get("id_offres_sfr") or 0),
                lib_offre=r.get("lib_offre") or "",
                type=r.get("type") or "",
                type_vente=int(r.get("type_vente") or 0),
                num=num,
                num_date_saisie=_date_str(r.get("num_date_saisie")),
                portabilite=bool(r.get("portabilite")),
                num_portabilite=r.get("num_portabilite") or "",
                num_prise_rio=r.get("num_prise_rio") or "",
                num_prise_optique=r.get("num_prise_optique") or "",
                opt_tv=r.get("opt_tv") or "",
                opt_choisies=r.get("opt_choisies") or "",
                test_eligibilite=r.get("test_eligibilite") or "",
                motif_annulation=r.get("motif_annulation") or "",
                statut_prod=statut_p,
                a_creer=bool(num and statut_p == 1),
            ))

    return TicketCallDetail(
        id_tk_liste=str(tl["id_tk_liste"]),
        date_crea=_date_str(tl.get("date_crea")),
        id_tk_statut=int(tl.get("id_tk_statut") or 0),
        lib_statut=tl.get("lib_statut") or "",
        cloturee=bool(tl.get("cloturee")),
        date_cloture=_date_str(tl.get("date_cloture")),
        date_report=_date_str(tl.get("date_report")),
        op_crea=op_crea, op_crea_nom=nom_op,
        id_tk_call_sfr=str(id_tc) if id_tc else "",
        id_salarie=id_sal, id_salarie_nom=nom_sal,
        nom_client=tc.get("nom_client") or "",
        prenom_client=tc.get("prenom_client") or "",
        nom_marital_client=tc.get("nom_marital_client") or "",
        civilite_client=int(tc.get("civilite_client") or 0),
        date_naiss=_date_str(tc.get("date_naiss")),
        dep_naiss=str(tc.get("dep_naiss") or ""),
        adresse1=tc.get("adresse1") or "",
        adresse2=tc.get("adresse2") or "",
        cp=tc.get("cp") or "",
        ville=tc.get("ville") or "",
        mobile1=tc.get("mobile1") or "",
        adr_mail=tc.get("adr_mail") or "",
        type_logement=int(tc.get("type_logement") or 0),
        opt_rappel=bool(tc.get("opt_rappel")),
        opt_partenaire=bool(tc.get("opt_partenaire")),
        intervention_vend=bool(tc.get("intervention_vend")),
        info_vente=tc.get("info_vente") or "",
        ref_appel=tc.get("ref_appel") or "",
        motif_annulation=tc.get("motif_annulation") or "",
        code_valid=tc.get("code_valid") or "",
        paniers=paniers,
    )


def resolve_cin_url(id_call_sfr: int, source: str = "normal") -> str:
    """Test HEAD sur {id_call_sfr}_CIN.jpg, fallback _PieceIdentite.pdf
    cf code WinDev boutons 'Voir la CIN' / 'Voir la CIN SOS'.

    source : 'normal' -> https://groupe-exo.omaya.fr
             'sos'    -> https://sos.groupe-exo.omaya.fr
    Retourne l'URL a ouvrir cote front."""
    import requests
    base = ("https://sos.groupe-exo.omaya.fr" if source == "sos"
            else "https://groupe-exo.omaya.fr")
    url_jpg = f"{base}/DocOmaya/{id_call_sfr}_CIN.jpg"
    try:
        r = requests.head(url_jpg, timeout=5, allow_redirects=True)
        if r.status_code == 404:
            return f"{base}/DocOmaya/{id_call_sfr}_PieceIdentite.pdf"
        return url_jpg
    except Exception:
        # En cas d'erreur reseau, on renvoie le JPG par defaut (le
        # navigateur affichera son propre 404 si vraiment absent).
        return url_jpg


# ====================================================================
# 4. EXTRACTION SFR (Fen_ExtractionSFR)
# ====================================================================


class EtatSfrItem(BaseModel):
    id_etat: int
    lib_etat: str


def list_etats_sfr() -> list[EtatSfrItem]:
    """Combo 'Etat vente SFR' (Fen_ExtractionSFR)."""
    db = get_pg_connection("adv")
    rows = db.query(
        """SELECT id_etat, lib_etat FROM adv.pgt_sfr_etat_contrat
            WHERE (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            ORDER BY lib_etat""",
    ) or []
    return [EtatSfrItem(id_etat=int(r["id_etat"]),
                         lib_etat=r.get("lib_etat") or "") for r in rows]


class ExtractionSfrRow(BaseModel):
    id_contrat: str
    num_bs: str = ""
    lib_produit: str = ""
    type_prod: str = ""           # famille_produit
    type_vente: int = 0
    date_signature: str = ""
    type_etat: str = ""           # lib_type
    etat_contrat: str = ""        # lib_etat
    etat_sfr: int = 0
    couleur_hex: str = ""         # couleur ligne (cf TypeEtatContrat.Couleur_*)
    # Vendeur
    nom_vendeur: str = ""
    agence: str = ""
    equipe: str = ""
    # Client
    client_nom: str = ""
    client_adr: str = ""
    client_cp: str = ""
    client_ville: str = ""
    client_mail: str = ""
    client_mobile: str = ""
    # SFR
    box8: bool = False
    box8_verif: bool = False
    cluster_code: str = ""
    cluster_nom: str = ""
    date_portabilite: str = ""
    date_racc_valid: str = ""
    date_rdv_tech: str = ""
    date_resil: str = ""
    date_validation: str = ""
    internet_garanti: bool = False
    remise: float = 0.0
    self_install: str = ""
    technologie: int = 0
    infos_internes: str = ""
    infos_partagees: str = ""


def search_extraction_sfr(
    du: date, au: date, mode: str = "date_racc",
    id_etat_sfr: int = 0,
) -> list[ExtractionSfrRow]:
    """Recherche contrats SFR pour Fen_ExtractionSFR.
    mode : 'date_racc' / 'rdv_tech' / 'churn'.
    Pour 'churn' : id_etat_sfr ignore (force a tous) + filtre
    id_type_etat=4 (resiliation) + date_resil non vide + nb jours
    (date_racc -> date_resil) <= 30."""
    db = get_pg_connection("adv")

    where = [
        "(c.modif_elem IS NULL OR c.modif_elem NOT LIKE '%suppr%')",
    ]
    params: list = []

    if mode == "rdv_tech":
        where.append("c.date_rdv_tech BETWEEN ? AND ?")
        params.extend([du, au])
    elif mode == "churn":
        where.append("c.date_resil BETWEEN ? AND ?")
        where.append("c.date_resil IS NOT NULL")
        where.append("e.id_type_etat = 4")
        params.extend([du, au])
    else:   # date_racc (default)
        where.append("c.date_racc_activ BETWEEN ? AND ?")
        params.extend([du, au])

    if id_etat_sfr and mode != "churn":
        where.append("c.id_etat_sfr = ?")
        params.append(int(id_etat_sfr))

    rows = db.query(
        f"""SELECT c.id_contrat, c.id_salarie, c.id_client,
                   c.num_bs, c.date_signature, c.box8, c.box8_verif,
                   c.id_sfr_cluster, c.date_portabilite, c.date_racc_activ,
                   c.date_rdv_tech, c.date_resil, c.date_validation,
                   c.id_etat_sfr, c.internet_garanti, c.remise,
                   c.self_install, c.technologie, c.type_vente,
                   c.info_interne, c.info_vente_sfr,
                   p.lib_produit, p.famille,
                   e.lib_etat, e.id_type_etat,
                   t.lib_type, t.couleur_r, t.couleur_v, t.couleur_b
              FROM adv.pgt_sfr_contrat c
              LEFT JOIN adv.pgt_sfr_produit p ON p.id_produit = c.id_produit
              LEFT JOIN adv.pgt_sfr_etat_contrat e
                     ON e.id_etat = c.id_etat_contrat
              LEFT JOIN adv.pgt_type_etat_contrat t
                     ON t.id_type_etat = e.id_type_etat
             WHERE {' AND '.join(where)}
             ORDER BY c.date_signature DESC
             LIMIT 5000""",
        tuple(params),
    ) or []

    # Filtre Churn : nb_jours date_racc -> date_resil <= 30
    if mode == "churn":
        from datetime import date as date_cls
        kept = []
        for r in rows:
            dra = r.get("date_racc_activ"); drs = r.get("date_resil")
            if isinstance(dra, date_cls) and isinstance(drs, date_cls):
                if (drs - dra).days <= 30:
                    kept.append(r)
        rows = kept

    # Resolution clients (batch) et vendeurs + clusters
    id_clients = list({int(r["id_client"]) for r in rows if r.get("id_client")})
    id_sals    = list({int(r["id_salarie"]) for r in rows if r.get("id_salarie")})
    id_clus    = list({int(r["id_sfr_cluster"]) for r in rows if r.get("id_sfr_cluster")})

    clients_map: dict[int, dict] = {}
    if id_clients:
        ids = ",".join(str(i) for i in id_clients)
        cs = db.query(
            f"""SELECT id_client, nom, prenom, adresse1, cp, ville, mail, gsm
                  FROM adv.pgt_client WHERE id_client IN ({ids})""",
        ) or []
        clients_map = {int(c["id_client"]): c for c in cs}

    db_rh = get_pg_connection("rh")
    sals_map: dict[int, dict] = {}
    if id_sals:
        ids = ",".join(str(i) for i in id_sals)
        ss = db_rh.query(
            f"""SELECT id_salarie, nom, prenom FROM rh.pgt_salarie
                 WHERE id_salarie IN ({ids})""",
        ) or []
        sals_map = {int(s["id_salarie"]): s for s in ss}

    clus_map: dict[int, dict] = {}
    if id_clus:
        ids = ",".join(str(i) for i in id_clus)
        cls = db.query(
            f"""SELECT id_sfr_cluster, code_vad, nom_cluster
                  FROM adv.pgt_sfr_cluster
                 WHERE id_sfr_cluster IN ({ids})""",
        ) or []
        clus_map = {int(c["id_sfr_cluster"]): c for c in cls}

    out: list[ExtractionSfrRow] = []
    for r in rows:
        id_sal = int(r.get("id_salarie") or 0)
        sal = sals_map.get(id_sal, {})
        nom_v = (f"{(sal.get('nom') or '').strip()} "
                 f"{_capitalize((sal.get('prenom') or '').strip())}").strip()

        id_clt = int(r.get("id_client") or 0)
        cli = clients_map.get(id_clt, {})

        clu = clus_map.get(int(r.get("id_sfr_cluster") or 0), {})

        # Couleur ligne (hex from R/V/B)
        cr = int(r.get("couleur_r") or 0)
        cv = int(r.get("couleur_v") or 0)
        cb = int(r.get("couleur_b") or 0)
        couleur_hex = ""
        if cr or cv or cb:
            couleur_hex = f"#{cr:02x}{cv:02x}{cb:02x}"

        out.append(ExtractionSfrRow(
            id_contrat=str(r["id_contrat"]),
            num_bs=r.get("num_bs") or "",
            lib_produit=r.get("lib_produit") or "",
            type_prod=r.get("famille") or "",
            type_vente=int(r.get("type_vente") or 0),
            date_signature=_date_str(r.get("date_signature")),
            type_etat=r.get("lib_type") or "",
            etat_contrat=r.get("lib_etat") or "",
            etat_sfr=int(r.get("id_etat_sfr") or 0),
            couleur_hex=couleur_hex,
            nom_vendeur=nom_v,
            client_nom=" ".join(filter(None, [
                (cli.get("nom") or "").strip(),
                _capitalize((cli.get("prenom") or "").strip()),
            ])),
            client_adr=cli.get("adresse1") or "",
            client_cp=cli.get("cp") or "",
            client_ville=cli.get("ville") or "",
            client_mail=cli.get("mail") or "",
            client_mobile=cli.get("gsm") or "",
            box8=bool(r.get("box8")),
            box8_verif=bool(r.get("box8_verif")),
            cluster_code=clu.get("code_vad") or "",
            cluster_nom=clu.get("nom_cluster") or "",
            date_portabilite=_date_str(r.get("date_portabilite")),
            date_racc_valid=_date_str(r.get("date_racc_activ")),
            date_rdv_tech=_date_str(r.get("date_rdv_tech")),
            date_resil=_date_str(r.get("date_resil")),
            date_validation=_date_str(r.get("date_validation")),
            internet_garanti=bool(r.get("internet_garanti")),
            remise=float(r.get("remise") or 0),
            self_install=str(r.get("self_install") or ""),
            technologie=int(r.get("technologie") or 0),
            infos_internes=r.get("info_interne") or "",
            infos_partagees=r.get("info_vente_sfr") or "",
        ))
    return out


# ====================================================================
# 6. CLUSTER SFR (Fen_SFRCluster + Fen_ClusterAjout)
# ====================================================================


class ClusterSfr(BaseModel):
    id_sfr_cluster: str
    region: str = ""
    code_vad: str = ""
    nom_cluster: str = ""
    mail_bo: str = ""


class ClusterPeriode(BaseModel):
    id_sfr_cluster_periode: str
    id_sfr_cluster: str
    du: str = ""
    au: str = ""
    objectif_vv: int = 0


class ClusterPayload(BaseModel):
    region: str = ""
    code_vad: str = ""
    nom_cluster: str = ""
    mail_bo: str = ""


class PeriodePayload(BaseModel):
    id_sfr_cluster: int
    du: date
    au: date
    objectif_vv: int = 0


def list_clusters() -> list[ClusterSfr]:
    db = get_pg_connection("adv")
    rows = db.query(
        """SELECT id_sfr_cluster, region, code_vad, nom_cluster, mail_bo
             FROM adv.pgt_sfr_cluster
            WHERE (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            ORDER BY region, nom_cluster"""
    ) or []
    return [ClusterSfr(
        id_sfr_cluster=str(r["id_sfr_cluster"]),
        region=r.get("region") or "",
        code_vad=r.get("code_vad") or "",
        nom_cluster=r.get("nom_cluster") or "",
        mail_bo=r.get("mail_bo") or "",
    ) for r in rows]


def list_cluster_periodes(id_sfr_cluster: int) -> list[ClusterPeriode]:
    db = get_pg_connection("adv")
    # 989/1206 periodes en BDD sont des "coquilles vides" heritees de WinDev
    # (du=NULL, au=NULL, objectif_vv=0) -> on les masque a l'affichage.
    rows = db.query(
        """SELECT id_sfr_cluster_periode, id_sfr_cluster,
                  du, au, objectif_vv
             FROM adv.pgt_sfr_cluster_periode
            WHERE id_sfr_cluster = ?
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
              AND (du IS NOT NULL OR au IS NOT NULL
                   OR COALESCE(objectif_vv, 0) > 0)
            ORDER BY du DESC NULLS LAST""",
        (int(id_sfr_cluster),),
    ) or []
    return [ClusterPeriode(
        id_sfr_cluster_periode=str(r["id_sfr_cluster_periode"]),
        id_sfr_cluster=str(r["id_sfr_cluster"]),
        du=_date_str(r.get("du")),
        au=_date_str(r.get("au")),
        objectif_vv=int(r.get("objectif_vv") or 0),
    ) for r in rows]


def create_cluster(p: ClusterPayload, op_id: int) -> int:
    db = get_pg_connection("adv")
    id_new = _new_id()
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_sfr_cluster_auto), 0) + 1 AS n FROM adv.pgt_sfr_cluster"
    )
    auto_n = int(auto["n"]) if auto else 1
    db.query(
        """INSERT INTO adv.pgt_sfr_cluster
              (id_sfr_cluster_auto, id_sfr_cluster, region, code_vad,
               nom_cluster, mail_bo,
               modif_date, modif_op, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, NOW(), ?, 'new')""",
        (auto_n, id_new, p.region, p.code_vad, p.nom_cluster, p.mail_bo,
         int(op_id)),
    )
    return id_new


def update_cluster(id_cluster: int, p: ClusterPayload, op_id: int) -> bool:
    db = get_pg_connection("adv")
    db.query(
        """UPDATE adv.pgt_sfr_cluster
              SET region=?, code_vad=?, nom_cluster=?, mail_bo=?,
                  modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_sfr_cluster=?""",
        (p.region, p.code_vad, p.nom_cluster, p.mail_bo,
         int(op_id), int(id_cluster)),
    )
    return True


def delete_cluster(id_cluster: int, op_id: int) -> bool:
    db = get_pg_connection("adv")
    db.query(
        """UPDATE adv.pgt_sfr_cluster
              SET modif_elem='suppr', modif_date=NOW(), modif_op=?
            WHERE id_sfr_cluster=?""",
        (int(op_id), int(id_cluster)),
    )
    return True


def create_cluster_periode(p: PeriodePayload, op_id: int) -> int:
    db = get_pg_connection("adv")
    id_new = _new_id()
    # id_sfr_cluster_duau : cle composite text utilisee par WinDev pour
    # l'indexation. On la construit comme '{id_cluster}_{du}_{au}'.
    duau = f"{int(p.id_sfr_cluster)}_{p.du}_{p.au}"
    db.query(
        """INSERT INTO adv.pgt_sfr_cluster_periode
              (id_sfr_cluster_periode, id_sfr_cluster, du, au,
               objectif_vv, id_sfr_cluster_duau,
               modif_date, modif_op, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, NOW(), ?, 'new')""",
        (id_new, int(p.id_sfr_cluster), p.du, p.au, int(p.objectif_vv),
         duau, int(op_id)),
    )
    return id_new


def update_cluster_periode(
    id_periode: int, p: PeriodePayload, op_id: int,
) -> bool:
    db = get_pg_connection("adv")
    db.query(
        """UPDATE adv.pgt_sfr_cluster_periode
              SET du=?, au=?, objectif_vv=?,
                  modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_sfr_cluster_periode=?""",
        (p.du, p.au, int(p.objectif_vv), int(op_id), int(id_periode)),
    )
    return True


def delete_cluster_periode(id_periode: int, op_id: int) -> bool:
    db = get_pg_connection("adv")
    db.query(
        """UPDATE adv.pgt_sfr_cluster_periode
              SET modif_elem='suppr', modif_date=NOW(), modif_op=?
            WHERE id_sfr_cluster_periode=?""",
        (int(op_id), int(id_periode)),
    )
    return True


# ====================================================================
# 7. OFFRES EZY (Fen_OffresEZY)
# ====================================================================


class OffreEzy(BaseModel):
    id_offres_sfr: str
    type: str = ""
    lib_offre: str = ""
    debit_down: str = ""
    debit_up: str = ""
    prix_offre: float = 0.0
    recurrence: str = ""
    prix_pro_ttc: str = ""
    engagement: str = ""
    en_promo: bool = False
    info_promo: str = ""
    service_inclus: str = ""
    id_produit: int = 0
    lib_produit: str = ""      # jointure sfr_produit
    online: bool = False


class ProduitSfr(BaseModel):
    id_produit: int
    lib_produit: str = ""
    famille: str = ""


class OffreEzyPayload(BaseModel):
    id_produit: int = 0
    online: bool = False


def list_offres_ezy(cat: str, pro: bool = False) -> list[OffreEzy]:
    """Liste les offres SFR Provad par categorie (FIBRE / MOBILE) et
    version Part (defaut) ou Pro. cf code WinDev :
      - FIBRE + Part -> type='FIBRE'
      - FIBRE + Pro  -> type='FIB PRO'
      - MOBILE + Part -> type='MOBILE'
      - MOBILE + Pro  -> type='MOB PRO'
    """
    cat = cat.upper()
    if cat == "FIBRE":
        type_val = "FIB PRO" if pro else "FIBRE"
    elif cat == "MOBILE":
        type_val = "MOB PRO" if pro else "MOBILE"
    else:
        return []

    db = get_pg_connection("adv")
    rows = db.query(
        """SELECT o.id_offres_sfr, o.type, o.lib_offre,
                  o.debit_down, o.debit_up, o.prix_offre, o.recurrence,
                  o.prix_pro_ttc, o.engagement,
                  o.en_promo, o.info_promo, o.service_inclus,
                  o.id_produit, o.online,
                  p.lib_produit
             FROM adv.pgt_sfr_offres_provad o
             LEFT JOIN adv.pgt_sfr_produit p ON p.id_produit = o.id_produit
            WHERE o.type = ?
              AND (o.modif_elem IS NULL OR o.modif_elem NOT LIKE '%suppr%')
            ORDER BY o.lib_offre""",
        (type_val,),
    ) or []
    return [OffreEzy(
        id_offres_sfr=str(r["id_offres_sfr"]),
        type=r.get("type") or "",
        lib_offre=r.get("lib_offre") or "",
        debit_down=r.get("debit_down") or "",
        debit_up=r.get("debit_up") or "",
        prix_offre=float(r.get("prix_offre") or 0),
        recurrence=r.get("recurrence") or "",
        prix_pro_ttc=str(r.get("prix_pro_ttc") or ""),
        engagement=str(r.get("engagement") or ""),
        en_promo=bool(r.get("en_promo")),
        info_promo=str(r.get("info_promo") or ""),
        service_inclus=str(r.get("service_inclus") or ""),
        id_produit=int(r.get("id_produit") or 0),
        lib_produit=r.get("lib_produit") or "",
        online=bool(r.get("online")),
    ) for r in rows]


def list_produits_offres_ezy(famille: str) -> list[ProduitSfr]:
    """Liste des produits SFR pour la combo 'Produit associé'
    (ReqProduitFibre / ReqProduitMobile WinDev)."""
    famille = famille.upper()
    db = get_pg_connection("adv")
    rows = db.query(
        """SELECT id_produit, lib_produit, famille
             FROM adv.pgt_sfr_produit
            WHERE famille = ?
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            ORDER BY lib_produit""",
        (famille,),
    ) or []
    return [ProduitSfr(
        id_produit=int(r["id_produit"]),
        lib_produit=r.get("lib_produit") or "",
        famille=r.get("famille") or "",
    ) for r in rows]


class ImportOffresEzyResult(BaseModel):
    nb_parses: int = 0
    nb_crees: int = 0
    nb_updates: int = 0
    nb_errors: int = 0
    offres: list[dict] = []


def import_offres_ezy(
    html_content: str, source: str, op_id: int,
) -> ImportOffresEzyResult:
    """Parse le HTML Provad puis upsert dans pgt_sfr_offres_provad.
    Cf 5 boutons WinDev de l'onglet Import :
      - source='fibre'      -> Import Offres SFR FIBRE
      - source='mobile'     -> Import Offres SFR Mobile (+ match produit MOBILE)
      - source='secu'       -> Import Offres SFR Maison SECU (+ match produit SECU)
      - source='fibre_pro'  -> Import Offres SFR FIBRE Pro
      - source='mobile_pro' -> Import Offres SFR Mobile Pro (+ match produit MOB PRO)
    """
    from app.intranets.adm.services.offres_ezy_parser import parse_html_import

    offres = parse_html_import(html_content, source)
    res = ImportOffresEzyResult(nb_parses=len(offres), offres=list(offres))
    if not offres:
        return res

    db = get_pg_connection("adv")

    # Famille du produit associe a matcher automatiquement selon la source
    match_famille: str | None = {
        "mobile":     "MOBILE",
        "secu":       "SECU",
        "mobile_pro": "MOB PRO",
    }.get(source.lower())

    for o in offres:
        lib = o["lib_offre"]
        type_val = o["type"]
        try:
            # Match auto du produit associe (uniquement pour mobile/secu/mobpro)
            id_produit = 0
            if match_famille:
                prod = db.query_one(
                    """SELECT id_produit FROM adv.pgt_sfr_produit
                        WHERE famille = ? AND lib_produit = ? LIMIT 1""",
                    (match_famille, lib),
                )
                if prod:
                    id_produit = int(prod["id_produit"])

            # Upsert : match par (lib_offre, type) pour eviter collision Part/Pro
            existing = db.query_one(
                """SELECT id_offres_sfr FROM adv.pgt_sfr_offres_provad
                    WHERE lib_offre = ? AND type = ?
                      AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                    LIMIT 1""",
                (lib, type_val),
            )
            if existing:
                # UPDATE : on ne touche pas id_produit sauf match auto (mobile/secu/mobpro)
                params = [
                    o["debit_down"], o["debit_up"],
                    float(o["prix_offre"]),
                    o["recurrence"],
                    o["prix_pro_ttc"],
                    o["engagement"],
                    bool(o["en_promo"]),
                    o["info_promo"],
                    o["services_inclus"],
                    int(op_id),
                ]
                if match_famille and id_produit:
                    db.query(
                        """UPDATE adv.pgt_sfr_offres_provad SET
                                  debit_down=?, debit_up=?, prix_offre=?,
                                  recurrence=?, prix_pro_ttc=?, engagement=?,
                                  en_promo=?, info_promo=?, service_inclus=?,
                                  id_produit=?, online=TRUE,
                                  modif_date=NOW(), modif_op=?, modif_elem='Modif'
                            WHERE id_offres_sfr=?""",
                        (*params[:9], id_produit, int(op_id),
                         int(existing["id_offres_sfr"])),
                    )
                else:
                    db.query(
                        """UPDATE adv.pgt_sfr_offres_provad SET
                                  debit_down=?, debit_up=?, prix_offre=?,
                                  recurrence=?, prix_pro_ttc=?, engagement=?,
                                  en_promo=?, info_promo=?, service_inclus=?,
                                  online=TRUE,
                                  modif_date=NOW(), modif_op=?, modif_elem='Modif'
                            WHERE id_offres_sfr=?""",
                        (*params, int(existing["id_offres_sfr"])),
                    )
                res.nb_updates += 1
            else:
                id_new = _new_id()
                auto = db.query_one(
                    "SELECT COALESCE(MAX(id_offres_sfr_auto), 0) + 1 AS n FROM adv.pgt_sfr_offres_provad"
                )
                auto_n = int(auto["n"]) if auto else 1
                db.query(
                    """INSERT INTO adv.pgt_sfr_offres_provad
                          (id_offres_sfr_auto, id_offres_sfr, type, lib_offre,
                           debit_down, debit_up, prix_offre, recurrence,
                           prix_pro_ttc, engagement,
                           en_promo, info_promo, service_inclus,
                           id_produit, online,
                           modif_date, modif_op, modif_elem)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, TRUE,
                               NOW(), ?, 'Modif')""",
                    (auto_n, id_new, type_val, lib,
                     o["debit_down"], o["debit_up"], float(o["prix_offre"]),
                     o["recurrence"], o["prix_pro_ttc"], o["engagement"],
                     bool(o["en_promo"]), o["info_promo"], o["services_inclus"],
                     id_produit, int(op_id)),
                )
                res.nb_crees += 1
        except Exception:
            res.nb_errors += 1
    return res


def update_offre_ezy(id_offres_sfr: int, p: OffreEzyPayload, op_id: int) -> bool:
    """Met a jour l'association produit + le flag online sur une offre.
    Cf code WinDev 'Sortie d'une ligne' : seuls IDproduit + Online sont
    modifiables."""
    db = get_pg_connection("adv")
    db.query(
        """UPDATE adv.pgt_sfr_offres_provad
              SET id_produit=?, online=?,
                  modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_offres_sfr=?""",
        (int(p.id_produit), bool(p.online), int(op_id), int(id_offres_sfr)),
    )
    return True


# ====================================================================
# 8. SUIVI RDV TECH (Fen_SuiviRDVTECH)
# ====================================================================


class SuiviRdvTechRow(BaseModel):
    id_tk_liste: str
    id_contrat: str
    id_tk_retour_rdv_tech_fibre: str
    date_crea: str = ""
    date_cloture: str = ""
    cloturee: bool = False
    lib_statut: str = ""
    vendeur: str = ""
    num_bs: str = ""              # depuis TK_RetourRdvTechFIBRE
    num_bs_sfr: str = ""          # depuis SFR_contrat
    date_rdv_tech: str = ""
    periode_rdv_tech: str = ""
    date_signature: str = ""
    id_fibre_statut_rdv: int = 0
    lib_statut_rdv: str = ""
    info_cplt: str = ""


class SfrStatutRdv(BaseModel):
    id_sfr_statut_rdv: int
    lib_statut: str = ""


class RdvTechUpdatePayload(BaseModel):
    id_fibre_statut_rdv: int = 0
    info_cplt: str = ""


def list_sfr_statuts_rdv() -> list[SfrStatutRdv]:
    db = get_pg_connection("adv")
    rows = db.query(
        """SELECT id_sfr_statut_rdv, lib_statut FROM adv.pgt_sfr_statut_rdv
            WHERE (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            ORDER BY lib_statut""",
    ) or []
    return [SfrStatutRdv(
        id_sfr_statut_rdv=int(r["id_sfr_statut_rdv"]),
        lib_statut=r.get("lib_statut") or "",
    ) for r in rows]


def update_rdv_tech(
    id_retour: int, p: RdvTechUpdatePayload, op_id: int,
) -> bool:
    """Met a jour le statut RDV + info complementaire d'un retour
    RDV Tech (pgt_tk_retour_rdv_tech_fibre)."""
    db = get_pg_connection("ticket_bo")
    db.query(
        """UPDATE ticket_bo.pgt_tk_retour_rdv_tech_fibre
              SET id_fibre_statut_rdv=?, info_cplt=?,
                  modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_tk_retour_rdv_tech_fibre=?""",
        (int(p.id_fibre_statut_rdv), p.info_cplt or "",
         int(op_id), int(id_retour)),
    )
    return True


def search_suivi_rdv_tech(du: date, au: date, etat: str = "tous",
                          ) -> list[SuiviRdvTechRow]:
    """cf requete SQL Fen_SuiviRDVTECH : JOIN TK_RetourRdvTechFIBRE
    + TK_Liste + TK_Statut + salarie + SFR_contrat.
    etat : 'ouverts' | 'clotures' | 'tous'.
    """
    db_bo = get_pg_connection("ticket_bo")

    where = [
        "(r.modif_elem IS NULL OR r.modif_elem NOT LIKE '%suppr%')",
        "l.date_crea >= ?",
        "l.date_crea < (?::date + INTERVAL '1 day')",
    ]
    params: list = [du, au]
    if etat == "ouverts":
        where.append("(l.cloturee IS NULL OR l.cloturee = FALSE)")
    elif etat == "clotures":
        where.append("l.cloturee = TRUE")

    rows = db_bo.query(
        f"""SELECT r.id_tk_retour_rdv_tech_fibre, r.id_tk_liste, r.id_contrat,
                   r.num_bs, r.id_fibre_statut_rdv, r.info_cplt,
                   l.date_crea, l.date_cloture, l.cloturee, l.op_crea,
                   l.id_tk_statut
              FROM ticket_bo.pgt_tk_retour_rdv_tech_fibre r
              JOIN ticket.pgt_tk_liste l ON l.id_tk_liste = r.id_tk_liste
             WHERE {' AND '.join(where)}
             ORDER BY l.date_crea DESC
             LIMIT 5000""",
        tuple(params),
    ) or []
    if not rows:
        return []

    # Resolutions batch : SFR_contrat + salarie + tk_statut + fibre_statut_rdv
    id_contrats = list({int(r["id_contrat"]) for r in rows if r.get("id_contrat")})
    id_sals     = list({int(r["op_crea"]) for r in rows if r.get("op_crea")})
    id_statuts  = list({int(r["id_tk_statut"]) for r in rows if r.get("id_tk_statut")})
    id_rdvsts   = list({int(r["id_fibre_statut_rdv"]) for r in rows
                        if r.get("id_fibre_statut_rdv")})

    db_adv = get_pg_connection("adv")
    contrats_map: dict[int, dict] = {}
    if id_contrats:
        ids = ",".join(str(i) for i in id_contrats)
        c = db_adv.query(
            f"""SELECT id_contrat, num_bs, date_rdv_tech, periode_rdv_tech,
                       date_signature
                  FROM adv.pgt_sfr_contrat WHERE id_contrat IN ({ids})""",
        ) or []
        contrats_map = {int(x["id_contrat"]): x for x in c}

    rdvsts_map: dict[int, str] = {}
    if id_rdvsts:
        ids = ",".join(str(i) for i in id_rdvsts)
        s = db_adv.query(
            f"""SELECT id_sfr_statut_rdv, lib_statut FROM adv.pgt_sfr_statut_rdv
                 WHERE id_sfr_statut_rdv IN ({ids})""",
        ) or []
        rdvsts_map = {int(x["id_sfr_statut_rdv"]): x.get("lib_statut") or ""
                       for x in s}

    db_rh = get_pg_connection("rh")
    sals_map: dict[int, str] = {}
    if id_sals:
        ids = ",".join(str(i) for i in id_sals)
        s = db_rh.query(
            f"""SELECT id_salarie, nom, prenom FROM rh.pgt_salarie
                 WHERE id_salarie IN ({ids})""",
        ) or []
        for x in s:
            nom = (x.get("nom") or "").strip().upper()
            prenom = _capitalize((x.get("prenom") or "").strip().lower())
            sals_map[int(x["id_salarie"])] = f"{prenom} {nom}".strip()

    db_tk = get_pg_connection("ticket")
    statuts_map: dict[int, str] = {}
    if id_statuts:
        ids = ",".join(str(i) for i in id_statuts)
        s = db_tk.query(
            f"""SELECT id_tk_statut, lib_statut FROM ticket.pgt_tk_statut
                 WHERE id_tk_statut IN ({ids})""",
        ) or []
        statuts_map = {int(x["id_tk_statut"]): x.get("lib_statut") or ""
                        for x in s}

    out: list[SuiviRdvTechRow] = []
    for r in rows:
        id_c = int(r.get("id_contrat") or 0)
        ctt = contrats_map.get(id_c, {})
        out.append(SuiviRdvTechRow(
            id_tk_liste=str(r["id_tk_liste"]),
            id_contrat=str(r["id_contrat"]),
            id_tk_retour_rdv_tech_fibre=str(r["id_tk_retour_rdv_tech_fibre"]),
            date_crea=_date_str(r.get("date_crea")),
            date_cloture=_date_str(r.get("date_cloture")),
            cloturee=bool(r.get("cloturee")),
            lib_statut=statuts_map.get(int(r.get("id_tk_statut") or 0), ""),
            vendeur=sals_map.get(int(r.get("op_crea") or 0), ""),
            num_bs=r.get("num_bs") or "",
            num_bs_sfr=ctt.get("num_bs") or "",
            date_rdv_tech=_date_str(ctt.get("date_rdv_tech")),
            periode_rdv_tech=ctt.get("periode_rdv_tech") or "",
            date_signature=_date_str(ctt.get("date_signature")),
            id_fibre_statut_rdv=int(r.get("id_fibre_statut_rdv") or 0),
            lib_statut_rdv=rdvsts_map.get(int(r.get("id_fibre_statut_rdv") or 0), ""),
            info_cplt=r.get("info_cplt") or "",
        ))
    return out


# ====================================================================
# 9. EXTRACTION ETP (Fen_ETP)
# ====================================================================


class ExtractionEtpRow(BaseModel):
    code_cluster: str
    libelle_cluster: str
    courtier: str = "EXOSPHERE"
    inf_ou_egal_2: int = 0
    sup_ou_egal_3: int = 0


def search_extraction_etp(du: date, au: date) -> list[ExtractionEtpRow]:
    """Cf code WinDev Fen_ETP :
    Compte pour chaque cluster combien de vendeurs ont <=2 contrats
    et combien ont >=3 contrats SUR LA PERIODE, en FIBRE, hors TK
    et hors HorsCible. Chaque vendeur n'est comptabilise QUE dans
    son cluster majoritaire (celui ou il a le + de contrats grace
    au ORDER BY nbCtt DESC + skip si vendeur deja vu)."""
    db = get_pg_connection("adv")
    rows = db.query(
        """SELECT COUNT(c.id_contrat) AS nb_ctt,
                  cl.code_vad, cl.nom_cluster, c.id_salarie
             FROM adv.pgt_sfr_contrat c
             JOIN adv.pgt_sfr_cluster cl ON cl.id_sfr_cluster = c.id_sfr_cluster
             JOIN adv.pgt_sfr_produit p  ON p.id_produit = c.id_produit
            WHERE (c.modif_elem IS NULL OR c.modif_elem NOT LIKE '%suppr%')
              AND c.date_signature BETWEEN ? AND ?
              AND p.famille = 'FIBRE'
              AND c.id_sfr_cluster <> 0
              AND COALESCE(c.hors_cible, FALSE) = FALSE
              AND c.num_bs NOT ILIKE 'TK%'
            GROUP BY cl.code_vad, cl.nom_cluster, c.id_salarie
           HAVING COUNT(c.id_contrat) > 0
            ORDER BY nb_ctt DESC""",
        (du, au),
    ) or []

    # Dedup vendeur + agrege par cluster
    seen_vendeurs: set[int] = set()
    by_cluster: dict[str, ExtractionEtpRow] = {}
    for r in rows:
        id_sa = int(r.get("id_salarie") or 0)
        if id_sa in seen_vendeurs:
            continue
        seen_vendeurs.add(id_sa)

        code = r.get("code_vad") or ""
        item = by_cluster.get(code)
        if item is None:
            item = ExtractionEtpRow(
                code_cluster=code,
                libelle_cluster=r.get("nom_cluster") or "",
            )
            by_cluster[code] = item

        nb = int(r.get("nb_ctt") or 0)
        if nb <= 2:
            item.inf_ou_egal_2 += 1
        else:
            item.sup_ou_egal_3 += 1

    return list(by_cluster.values())


def export_extraction_etp_xlsx(rows: list[ExtractionEtpRow]) -> bytes:
    """XLSX simple sans couleurs."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    wb = Workbook()
    ws = wb.active; ws.title = "Extraction ETP"
    ws.append(["Code Cluster", "Libellé Cluster", "Courtier",
                "Inf ou égal à 2", "Sup ou égal à 3"])
    header_fill = PatternFill(start_color="FF17494E", end_color="FF17494E",
                               fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill; cell.font = header_font

    total_inf = total_sup = 0
    for r in rows:
        ws.append([r.code_cluster, r.libelle_cluster, r.courtier,
                    r.inf_ou_egal_2, r.sup_ou_egal_3])
        total_inf += r.inf_ou_egal_2
        total_sup += r.sup_ou_egal_3

    # Ligne Somme
    ws.append(["", "Somme", "", total_inf, total_sup])
    total_font = Font(bold=True)
    for cell in ws[ws.max_row]:
        cell.font = total_font

    for i, w in enumerate([14, 32, 12, 14, 14], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ====================================================================
# 5. PARCOURS CHAINES (Fen_ParcoursChaine)
# ====================================================================


# IDTypeDroitAccès = 209 cf code WinDev (droit "tickets Diff SFR")
DROIT_TK_DIFF_SFR = 209


class ParcoursChaineRow(BaseModel):
    id_sa: str
    vendeur: str = ""
    en_activite: bool = False
    agence: str = ""
    equipe: str = ""
    droit_diff: bool = False
    nb_tk_valides: int = 0
    nb_parcours_chaines: int = 0
    nb_tk_chaine_tot: int = 0
    nb_tk_diff: int = 0
    pourcent_global: float = 0.0       # chaines / tk_valides
    pourcent_chaines: float = 0.0      # chaines / tk_chaine_tot
    couleur_hex: str = ""              # selon les 2 taux (cf WinDev)


def _color_parcours(tx1: float, tx2: float) -> str:
    """Cf code WinDev :
      - tx1 >= 0.8 et tx2 >= 0.8 -> VertPastel
      - tx1 >= 0.8 et tx2 <  0.8 -> OrangePastel
      - tx1 <  0.8 et tx2 >= 0.8 -> JaunePastel
      - autre                    -> RougePastel"""
    if tx1 >= 0.8 and tx2 >= 0.8: return "#bbf7d0"   # vert pastel
    if tx1 >= 0.8 and tx2 <  0.8: return "#fed7aa"   # orange pastel
    if tx1 <  0.8 and tx2 >= 0.8: return "#fef3c7"   # jaune pastel
    return "#fecaca"                                  # rouge pastel


def search_parcours_chaines(du: date, au: date) -> list[ParcoursChaineRow]:
    """Calcul par vendeur (OPCrea) :
      - nb_tk_valides : nb tickets distincts avec au moins 1 panier
        ayant un vrai NumBS (non vide, non 'TK%')
      - nb_parcours_chaines : nb tickets distincts ayant FIBRE+MOBILE
        avec vrais NumBS
      - nb_tk_chaine_tot = nb_parcours_chaines (le KO 'AnomalieMobile'
        n'existe pas en PG actuellement)
      - nb_tk_diff = 0 (la colonne ticket_diff/anomalie_mobile n'existe
        pas dans pgt_tk_call_sfr PG)
      - pourcent_global = chaines / tk_valides
      - pourcent_chaines = chaines / tk_chaine_tot
      - droit_diff = pgt_salarie_droit_acces (id_type_droit_acces=209)
    """
    db_bo = get_pg_connection("ticket_bo")
    db_tk = get_pg_connection("ticket")
    db_rh = get_pg_connection("rh")

    # 1. Tk validés : un ticket -> 1 si au moins un panier a un vrai NumBS
    tk_valides = db_bo.query(
        """SELECT DISTINCT tc.id_tk_liste, tl.op_crea
             FROM ticket_bo.pgt_tk_call_sfr tc
             JOIN ticket_bo.pgt_tk_call_sfr_panier p
               ON p.id_tk_call_sfr = tc.id_tk_call_sfr
             JOIN ticket.pgt_tk_liste tl ON tl.id_tk_liste = tc.id_tk_liste
            WHERE (p.modif_elem IS NULL OR p.modif_elem NOT LIKE '%suppr%')
              AND p.num <> ''
              AND p.num NOT ILIKE 'TK%'
              AND tl.date_crea >= ?
              AND tl.date_crea < (?::date + INTERVAL '1 day')""",
        (du, au),
    ) or []

    # 2. Parcours chainés : ticket avec FIBRE+MOBILE non-TK
    chaines = db_bo.query(
        """SELECT DISTINCT tc.id_tk_liste, tl.op_crea
             FROM ticket_bo.pgt_tk_call_sfr tc
             JOIN ticket_bo.pgt_tk_call_sfr_panier fib
               ON fib.id_tk_call_sfr = tc.id_tk_call_sfr
              AND fib.type = 'FIBRE'
              AND fib.num <> ''
              AND fib.num NOT ILIKE 'TK%'
              AND (fib.modif_elem IS NULL OR fib.modif_elem NOT LIKE '%suppr%')
             JOIN ticket_bo.pgt_tk_call_sfr_panier mob
               ON mob.id_tk_call_sfr = tc.id_tk_call_sfr
              AND mob.type = 'MOBILE'
              AND mob.num <> ''
              AND mob.num NOT ILIKE 'TK%'
              AND (mob.modif_elem IS NULL OR mob.modif_elem NOT LIKE '%suppr%')
             JOIN ticket.pgt_tk_liste tl ON tl.id_tk_liste = tc.id_tk_liste
            WHERE tl.date_crea >= ?
              AND tl.date_crea < (?::date + INTERVAL '1 day')""",
        (du, au),
    ) or []

    # Agrege par operateur
    by_op: dict[int, dict] = {}
    for r in tk_valides:
        op = int(r.get("op_crea") or 0)
        if not op: continue
        by_op.setdefault(op, {"valides": set(), "chaines": set()})
        by_op[op]["valides"].add(int(r["id_tk_liste"]))
    for r in chaines:
        op = int(r.get("op_crea") or 0)
        if not op: continue
        by_op.setdefault(op, {"valides": set(), "chaines": set()})
        by_op[op]["chaines"].add(int(r["id_tk_liste"]))

    if not by_op:
        return []

    # 3. Resolution salaries + droit Diff SFR
    op_ids = list(by_op.keys())
    ids_sql = ",".join(str(i) for i in op_ids)
    sals = db_rh.query(
        f"""SELECT id_salarie, nom, prenom, en_activite
              FROM rh.pgt_salarie s
              LEFT JOIN LATERAL (
                SELECT en_activite FROM rh.pgt_salarie_embauche se
                 WHERE se.id_salarie = s.id_salarie LIMIT 1
              ) e ON true
             WHERE id_salarie IN ({ids_sql})""",
    ) or []
    sals_map = {int(s["id_salarie"]): s for s in sals}
    droits = db_rh.query(
        f"""SELECT id_salarie, droit_actif FROM rh.pgt_salarie_droit_acces
             WHERE id_salarie IN ({ids_sql})
               AND id_type_droit_acces = ?
               AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
        (DROIT_TK_DIFF_SFR,),
    ) or []
    droits_map = {int(d["id_salarie"]): bool(d.get("droit_actif")) for d in droits}

    out: list[ParcoursChaineRow] = []
    for op_id, agg in by_op.items():
        s = sals_map.get(op_id, {})
        nom = (s.get("nom") or "").strip()
        prenom = _capitalize((s.get("prenom") or "").strip())
        nb_val = len(agg["valides"])
        nb_chn = len(agg["chaines"])
        tx1 = nb_chn / nb_val if nb_val > 0 else 0.0
        tx2 = nb_chn / nb_chn if nb_chn > 0 else 0.0   # tk_chaine_tot = nb_chaines
        out.append(ParcoursChaineRow(
            id_sa=str(op_id),
            vendeur=f"{nom} {prenom}".strip(),
            en_activite=bool(s.get("en_activite")),
            droit_diff=droits_map.get(op_id, False),
            nb_tk_valides=nb_val,
            nb_parcours_chaines=nb_chn,
            nb_tk_chaine_tot=nb_chn,
            nb_tk_diff=0,
            pourcent_global=round(tx1 * 100, 1),
            pourcent_chaines=round(tx2 * 100, 1),
            couleur_hex=_color_parcours(tx1, tx2),
        ))
    out.sort(key=lambda r: (-r.pourcent_chaines, r.vendeur))
    return out


def set_droit_diff_sfr(ids_salarie: list[int], actif: bool, op_id: int) -> int:
    """Active/desactive le droit IDTypeDroitAcces=209 pour la liste."""
    db = get_pg_connection("rh")
    n = 0
    for id_sal in ids_salarie:
        existing = db.query_one(
            """SELECT id_salarie_droit_acces FROM rh.pgt_salarie_droit_acces
                WHERE id_salarie = ? AND id_type_droit_acces = ?
                LIMIT 1""",
            (int(id_sal), DROIT_TK_DIFF_SFR),
        )
        if existing:
            db.query(
                """UPDATE rh.pgt_salarie_droit_acces
                      SET droit_actif=?, modif_date=NOW(), modif_op=?,
                          modif_elem='modif'
                    WHERE id_salarie_droit_acces=?""",
                (actif, int(op_id), int(existing["id_salarie_droit_acces"])),
            )
        else:
            id_new = _new_id()
            db.query(
                """INSERT INTO rh.pgt_salarie_droit_acces
                      (id_salarie_droit_acces, id_salarie, id_type_droit_acces,
                       droit_actif, modif_date, modif_op, modif_elem)
                   VALUES (?, ?, ?, ?, NOW(), ?, 'new')""",
                (id_new, int(id_sal), DROIT_TK_DIFF_SFR, actif, int(op_id)),
            )
        n += 1
    return n


def export_parcours_chaines_xlsx(rows: list[ParcoursChaineRow]) -> bytes:
    """XLSX avec couleur de fond par ligne (selon les 2 taux)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    columns: list[tuple[str, str]] = [
        ("vendeur", "Vendeur"),
        ("agence", "Agence"),
        ("equipe", "Équipe"),
        ("en_activite", "Actif"),
        ("droit_diff", "Droit Tk Diff"),
        ("nb_tk_valides", "nb PC Validés"),
        ("nb_parcours_chaines", "nb PC potentiels"),
        ("pourcent_global", "% PC validés Prod chaînés potentiel"),
        ("nb_tk_chaine_tot", "nb Tk Validés Prod globale"),
        ("pourcent_chaines", "% PC Validés Prod globale"),
        ("nb_tk_diff", "nb Tk Diff"),
    ]
    wb = Workbook(); ws = wb.active; ws.title = "Parcours chaînés"
    ws.append([lbl for _, lbl in columns])
    header_fill = PatternFill(start_color="FF17494E", end_color="FF17494E",
                               fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([getattr(r, k) for k, _ in columns])
        if r.couleur_hex and r.couleur_hex.startswith("#"):
            hex_code = "FF" + r.couleur_hex[1:].upper()
            fill = PatternFill(start_color=hex_code, end_color=hex_code,
                                fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = fill

    widths = [24, 18, 18, 6, 8, 10, 10, 16, 16, 16, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def export_extraction_sfr_xlsx(rows: list[ExtractionSfrRow]) -> bytes:
    """Genere un XLSX openpyxl avec fond de ligne colore selon le
    couleur_hex de chaque ExtractionSfrRow (= couleur du TypeEtatContrat)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    columns: list[tuple[str, str]] = [
        ("num_bs", "Num BS"),
        ("lib_produit", "Lib Produit"),
        ("type_prod", "Type Prod"),
        ("type_vente", "Type vente"),
        ("date_signature", "Date Signature"),
        ("type_etat", "Type Etat"),
        ("etat_contrat", "Etat contrat"),
        ("nom_vendeur", "Vendeur"),
        ("client_nom", "Client"),
        ("client_adr", "Adresse"),
        ("client_cp", "CP"),
        ("client_ville", "Ville"),
        ("client_mail", "Mail"),
        ("client_mobile", "Mobile"),
        ("cluster_code", "Cluster Code"),
        ("cluster_nom", "Cluster Nom"),
        ("date_portabilite", "Date Portabilité"),
        ("date_racc_valid", "Date Racc"),
        ("date_rdv_tech", "Date RDV Tech"),
        ("date_resil", "Date Résil"),
        ("date_validation", "Date Validation"),
        ("box8", "Box8"),
        ("box8_verif", "Box8 Vérif"),
        ("internet_garanti", "Internet Garanti"),
        ("remise", "Remise"),
        ("self_install", "Self Install"),
        ("technologie", "Techno"),
        ("infos_internes", "Infos Internes"),
        ("infos_partagees", "Infos Partagées"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Extraction SFR"
    ws.append([lbl for _, lbl in columns])
    header_fill = PatternFill(start_color="FF17494E", end_color="FF17494E",
                               fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r in rows:
        values = [getattr(r, k) for k, _ in columns]
        ws.append(values)
        # Applique la couleur de fond sur toute la ligne
        if r.couleur_hex and r.couleur_hex.startswith("#"):
            hex_code = "FF" + r.couleur_hex[1:].upper()
            fill = PatternFill(start_color=hex_code, end_color=hex_code,
                                fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = fill

    # Largeurs colonnes approximatives
    widths = [14, 24, 12, 8, 12, 14, 22, 22, 22, 28, 8, 18, 26, 14,
              12, 18, 14, 12, 12, 12, 14, 6, 8, 12, 8, 12, 8, 30, 30]
    for i, w in enumerate(widths[:len(columns)], start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else
                              chr(64 + (i - 1) // 26) + chr(64 + (i - 1) % 26 + 1)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


def _create_tk_liste(db_tk, id_tk_liste: int, op_id: int,
                      id_type_demande: int) -> None:
    """Cree une entree TK_Liste (cf code WinDev TK_Liste.* dans
    Convertir en Ticket Call RET RDV TECH / Racc)."""
    db_tk.query(
        """INSERT INTO ticket.pgt_tk_liste
              (id_tk_liste, id_tk_liste_auto, date_crea, op_crea, op_dest,
               op_traitement_staff, ordre_traitement_staff,
               service, id_tk_type_demande, id_tk_statut,
               cloturee, modif_date, modif_op, modif_elem)
           VALUES (?, ?, NOW(), ?, ?, 0, 0, 'BO', ?, 1,
                   FALSE, NOW(), ?, 'new')""",
        (int(id_tk_liste), int(id_tk_liste), int(op_id), int(op_id),
         int(id_type_demande), int(op_id)),
    )


def convert_to_ret_rdv_tech(id_contrats: list[int], op_id: int) -> dict:
    """Pour chaque contrat selectionne, cree TK_Liste (IDTK_TypeDemande=26)
    + TK_CallSFR_RetRDVTech."""
    db_tk = get_pg_connection("ticket")
    db_bo = get_pg_connection("ticket_bo")
    nb_ok = 0; nb_ko = 0
    for id_c in id_contrats:
        try:
            id_tk = _new_id()
            _create_tk_liste(db_tk, id_tk, op_id, id_type_demande=26)
            db_bo.query(
                """INSERT INTO ticket_bo.pgt_tk_call_sfr_ret_rdv_tech
                      (id_tk_call_sfr_ret_rdv_tech, id_tk_liste, id_contrat,
                       id_sfr_statut_rdv, ope_traitement,
                       modif_date, modif_op, modif_elem)
                   VALUES (?, ?, ?, 0, 0, NOW(), ?, 'new')""",
                (_new_id(), int(id_tk), int(id_c), int(op_id)),
            )
            nb_ok += 1
        except Exception:
            nb_ko += 1
    return {"nb_ok": nb_ok, "nb_ko": nb_ko}


def convert_to_ret_racc(id_contrats: list[int], op_id: int) -> dict:
    """Idem mais IDTK_TypeDemande=32 + TK_CallSFR_RetRacc."""
    db_tk = get_pg_connection("ticket")
    db_bo = get_pg_connection("ticket_bo")
    nb_ok = 0; nb_ko = 0
    for id_c in id_contrats:
        try:
            id_tk = _new_id()
            _create_tk_liste(db_tk, id_tk, op_id, id_type_demande=32)
            db_bo.query(
                """INSERT INTO ticket_bo.pgt_tk_call_sfr_ret_racc
                      (id_tk_call_sfr_ret_racc, id_tk_liste, id_contrat,
                       id_etat_call_ret, ope_traitement,
                       modif_date, modif_op, modif_elem)
                   VALUES (?, ?, ?, 0, 0, NOW(), ?, 'new')""",
                (_new_id(), int(id_tk), int(id_c), int(op_id)),
            )
            nb_ok += 1
        except Exception:
            nb_ko += 1
    return {"nb_ok": nb_ok, "nb_ko": nb_ko}


# ====================================================================
# Convertir / Cloturer la selection (Fen_TicketCallSFR boutons)
# ====================================================================

# ID vendeur "generique" CALL SFR cf code WinDev
SFR_CALL_GENERIC_SALARIE_ID = 20200715153948361


class ConversionResultItem(BaseModel):
    id_tk_liste: str
    nb_paniers: int = 0
    nb_crees: int = 0
    nb_existants: int = 0
    nb_erreurs: int = 0
    message: str = ""
    cloture_ok: bool = False


def _enregistrer_client(
    db_adv, civilite: int, nom: str, prenom: str,
    date_naiss, adr: str, adr_cplt: str, cp: str, ville: str,
    tel: str, gsm: str, mail: str, op_id: int,
    opt_partenaire: bool = False,
) -> int:
    """Cree (ou retrouve) un client. Retourne id_client (0 si echec).
    cf code WinDev EnregistrerClient + ReqRechercheClient :
    on dedoublonne par (nom, prenom, adresse, cp, ville, tel/gsm).
    """
    tel1 = (tel or "").replace(".", "").replace(" ", "").replace("-", "")
    gsm2 = (gsm or "").replace(".", "").replace(" ", "").replace("-", "")
    # Recherche client existant (nom + prenom + cp + (tel OU gsm))
    existing = db_adv.query_one(
        """SELECT id_client FROM adv.pgt_client
            WHERE UPPER(nom) = UPPER(?) AND UPPER(prenom) = UPPER(?)
              AND COALESCE(cp, '') = ?
              AND (
                    (COALESCE(tel, '') = ? AND ? <> '')
                 OR (COALESCE(gsm, '') = ? AND ? <> '')
              )
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            LIMIT 1""",
        (nom, prenom, cp, gsm2, gsm2, tel1, tel1),
    )
    if existing:
        return int(existing["id_client"])

    id_new = _new_id()
    try:
        auto = db_adv.query_one(
            "SELECT COALESCE(MAX(id_client_auto), 0) + 1 AS n FROM adv.pgt_client"
        )
        auto_n = int(auto["n"]) if auto else 1
        db_adv.query(
            """INSERT INTO adv.pgt_client
                  (id_client_auto, id_client, civilite, nom, prenom,
                   date_naiss, adresse1, adresse2, cp, ville, pays,
                   tel, gsm, mail, opt_partenaire,
                   op_saisie, date_saisie,
                   modif_date, modif_op, modif_elem)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'FRANCE',
                       ?, ?, ?, ?, ?, NOW(), NOW(), ?, 'new')""",
            (auto_n, id_new, int(civilite or 0), nom, prenom,
             date_naiss if date_naiss else None,
             adr, adr_cplt, cp, ville,
             gsm2, tel1, mail, bool(opt_partenaire),
             int(op_id), int(op_id)),
        )
        return id_new
    except Exception:
        return 0


def _modif_fiche_client(
    db_adv, id_client: int, nom: str, prenom: str, date_naiss,
    adr: str, adr_cplt: str, cp: str, ville: str,
    tel: str, gsm: str, mail: str, op_id: int,
) -> None:
    """Met a jour le client UNIQUEMENT si son nom est vide
    (cf code WinDev ModifFicheClient : si client.NOM = '')."""
    if not id_client:
        return
    r = db_adv.query_one(
        "SELECT nom FROM adv.pgt_client WHERE id_client = ?", (int(id_client),))
    if not r or (r.get("nom") or "").strip():
        return
    tel1 = (tel or "").replace(".", "").replace(" ", "").replace("-", "")
    gsm2 = (gsm or "").replace(".", "").replace(" ", "").replace("-", "")
    db_adv.query(
        """UPDATE adv.pgt_client
              SET nom=?, prenom=?, date_naiss=?, adresse1=?, adresse2=?,
                  cp=?, ville=?, pays='FRANCE', tel=?, gsm=?, mail=?,
                  op_saisie=?, date_saisie=NOW(),
                  modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_client=?""",
        (nom, prenom, date_naiss if date_naiss else None,
         adr, adr_cplt, cp, ville, gsm2, tel1, mail,
         int(op_id), int(op_id), int(id_client)),
    )


def _technologie_from_num(num: str) -> int:
    """1 = THD, 2 = CBL, 3 = autre (cf code WinDev)."""
    n = (num or "").upper()
    if n.startswith("THD"): return 1
    if n.startswith("CBL"): return 2
    return 3


def _enregistrer_ctt_sfr(
    db_adv, num_bs: str, id_client: int, id_salarie: int,
    date_signature, statut_panier: int, motif: str,
    id_offres_sfr: int, type_vente: int, opt_tv: str,
    ticket_diff: bool, num_prise_optique: str, parcours_chaines: bool,
    op_id: int,
) -> int:
    """Cree un sfr_contrat. Retourne idStatutTicket :
      - 33 si NumBS deja existant (DOUBLON)
      - 4 si OK
      - 0 si erreur."""
    # Verif doublon
    exists = db_adv.query_one(
        """SELECT id_contrat FROM adv.pgt_sfr_contrat
            WHERE UPPER(num_bs) = UPPER(?)
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            LIMIT 1""",
        (num_bs,),
    )
    if exists:
        return 33

    # Etat selon statut panier (cf selon WinDev)
    if statut_panier == 2:
        id_etat = 29     # Rejet - Vente annulée par le CALL
    elif statut_panier == 4:
        id_etat = 84    # Validé - Différé
    elif "TK" in num_bs.upper():
        id_etat = 30    # Rejet - Panier validé mais non finalise
    else:
        id_etat = 9     # Temporaire non reconnu

    # IdSte du vendeur
    id_ste = None
    if id_salarie:
        try:
            db_rh = get_pg_connection("rh")
            se = db_rh.query_one(
                """SELECT id_ste FROM rh.pgt_salarie_embauche
                    WHERE id_salarie = ? LIMIT 1""",
                (int(id_salarie),),
            )
            if se: id_ste = se.get("id_ste")
        except Exception:
            pass

    # Lookup offre pour id_produit + lib_offre (pour box8)
    offre = db_adv.query_one(
        """SELECT id_produit, lib_offre FROM adv.pgt_sfr_offres_provad
            WHERE id_offres_sfr = ? LIMIT 1""",
        (int(id_offres_sfr),),
    ) or {}
    id_produit = int(offre.get("id_produit") or 0)
    box8 = "8" in (offre.get("lib_offre") or "")

    id_new = _new_id()
    try:
        auto = db_adv.query_one(
            "SELECT COALESCE(MAX(id_contrat_auto), 0) + 1 AS n FROM adv.pgt_sfr_contrat"
        )
        auto_n = int(auto["n"]) if auto else 1
        db_adv.query(
            """INSERT INTO adv.pgt_sfr_contrat
                  (id_contrat_auto, id_contrat, id_client, id_salarie,
                   id_ste, num_bs, date_signature,
                   id_etat_sfr, id_etat_contrat, motif_annulation,
                   technologie, id_produit, type_vente,
                   box8, option_dec,
                   non_call, issu_tk_diff, parcours_chaine,
                   num_prise_vend,
                   op_saisie, date_saisie,
                   hors_cible, notation, portabilite,
                   modif_date, modif_op, modif_elem)
               VALUES (?, ?, ?, ?, ?, ?, ?,
                       ?, ?, ?,
                       ?, ?, ?,
                       ?, ?,
                       TRUE, ?, ?,
                       ?,
                       ?, NOW(),
                       FALSE, 0, FALSE,
                       NOW(), ?, 'new')""",
            (auto_n, id_new, int(id_client), int(id_salarie),
             id_ste, num_bs, date_signature,
             id_etat, id_etat, motif or "",
             _technologie_from_num(num_bs), id_produit, int(type_vente),
             box8, opt_tv or "",
             bool(ticket_diff), bool(parcours_chaines),
             num_prise_optique or "",
             int(op_id),
             int(op_id)),
        )
    except Exception:
        return 0

    # Calcul nbPoints via calcul_point_contrat partage
    try:
        from app.shared.sdtc.bareme import calcul_point_contrat
        prod = db_adv.query_one(
            """SELECT famille, sous_fam FROM adv.pgt_sfr_produit
                WHERE id_produit = ? LIMIT 1""",
            (id_produit,),
        ) or {}
        fam = prod.get("famille") or ""
        sous_fam = prod.get("sous_fam") or ""
        # DonneFamProdSFR(famille, type_vente) : retourne 'FIB CQ' ou
        # 'FIB MIG' selon le contexte. Pour simplifier on prend famille
        # de la table produit.
        nbpt = calcul_point_contrat(fam, sous_fam, 0,
                                     str(date_signature) if date_signature else "",
                                     str(id_new), 0)
        if nbpt:
            db_adv.query(
                """UPDATE adv.pgt_sfr_contrat SET nb_points=?,
                          modif_date=NOW(), modif_op=?
                    WHERE id_contrat=?""",
                (float(nbpt), int(op_id), id_new),
            )
    except Exception:
        pass

    return 4


def convert_selection_to_contracts(
    ids_tk_liste: list[int], op_id: int,
) -> list[ConversionResultItem]:
    """Convertit les tickets selectionnes en contrats SFR puis cloture
    le ticket. Cf code WinDev btn 'Convertir la selection'."""
    db_bo = get_pg_connection("ticket_bo")
    db_tk = get_pg_connection("ticket")
    db_adv = get_pg_connection("adv")
    out: list[ConversionResultItem] = []

    for id_tl in ids_tk_liste:
        item = ConversionResultItem(id_tk_liste=str(id_tl))

        # 1. TK_CallSFR du ticket
        tc = db_bo.query_one(
            """SELECT id_tk_call_sfr, id_salarie, nom_client, prenom_client,
                      nom_marital_client, civilite_client, date_naiss,
                      adresse1, adresse2, cp, ville, mobile1, adr_mail,
                      opt_partenaire
                 FROM ticket_bo.pgt_tk_call_sfr
                WHERE id_tk_liste = ?
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (int(id_tl),),
        )
        if not tc:
            item.message = "TK_CallSFR introuvable"
            out.append(item); continue

        # 2. Lignes panier
        id_tc = int(tc["id_tk_call_sfr"])
        paniers = db_bo.query(
            """SELECT id_tk_call_sfr_panier, num, id_offres_sfr,
                      statut_prod, motif_annulation, type, type_vente,
                      opt_tv, num_prise_optique
                 FROM ticket_bo.pgt_tk_call_sfr_panier
                WHERE id_tk_call_sfr = ?
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
            (id_tc,),
        ) or []
        item.nb_paniers = len(paniers)

        # 3. Recupere date_crea pour signature
        tl = db_tk.query_one(
            "SELECT date_crea FROM ticket.pgt_tk_liste WHERE id_tk_liste=?",
            (int(id_tl),),
        )
        date_crea = tl.get("date_crea") if tl else None
        date_sign = date_crea.date() if date_crea else None

        # NomClt = NomClient (+ ' ep ' + NomMaritalClient si renseigne)
        nom_clt = (tc.get("nom_client") or "").strip()
        if (tc.get("nom_marital_client") or "").strip():
            nom_clt += " ep " + (tc.get("nom_marital_client") or "").strip()
        prenom_clt = (tc.get("prenom_client") or "").strip()

        # Test parcours chaines : FIBRE + MOBILE avec NUM non TK
        test_fib = test_mob = False
        for p in paniers:
            n = (p.get("num") or "").strip().upper()
            if n and not n.startswith("TK"):
                if (p.get("type") or "").upper() == "FIBRE":
                    test_fib = True
                else:
                    test_mob = True
        parcours_chaines = test_fib and test_mob

        id_vendeur = int(tc.get("id_salarie") or 0)

        # 4. Boucle paniers
        for p in paniers:
            id_panier = int(p["id_tk_call_sfr_panier"])
            num_bs = (p.get("num") or "").strip().upper()

            # Si NUM vide -> genere 'TK<id>'
            if not num_bs:
                num_bs = f"TK{id_panier}"
                db_bo.query(
                    """UPDATE ticket_bo.pgt_tk_call_sfr_panier
                          SET num=?, modif_date=NOW(), modif_op=?,
                              modif_elem='modif'
                        WHERE id_tk_call_sfr_panier=?""",
                    (num_bs, int(op_id), id_panier),
                )

            # Verif si contrat existe deja
            existing = db_adv.query_one(
                """SELECT id_contrat, id_salarie, id_client
                     FROM adv.pgt_sfr_contrat
                    WHERE UPPER(num_bs) = UPPER(?)
                      AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                    LIMIT 1""",
                (num_bs,),
            )
            if existing:
                item.nb_existants += 1
                # Reaffectation vendeur si actuel = generique (20200715... ou 0)
                cur_sal = int(existing.get("id_salarie") or 0)
                if cur_sal in (0, SFR_CALL_GENERIC_SALARIE_ID) and id_vendeur > 0:
                    db_adv.query(
                        """UPDATE adv.pgt_sfr_contrat
                              SET id_salarie=?, modif_date=NOW(),
                                  modif_op=?, modif_elem='modif'
                            WHERE id_contrat=?""",
                        (id_vendeur, int(op_id), int(existing["id_contrat"])),
                    )
                # Maj fiche client si vide
                _modif_fiche_client(
                    db_adv, int(existing.get("id_client") or 0),
                    nom_clt, prenom_clt, tc.get("date_naiss"),
                    tc.get("adresse1") or "", tc.get("adresse2") or "",
                    tc.get("cp") or "", tc.get("ville") or "",
                    "", tc.get("mobile1") or "", tc.get("adr_mail") or "",
                    op_id,
                )
            else:
                # Creation client + creation contrat
                id_clt = _enregistrer_client(
                    db_adv, int(tc.get("civilite_client") or 0),
                    nom_clt, prenom_clt, tc.get("date_naiss"),
                    tc.get("adresse1") or "", tc.get("adresse2") or "",
                    tc.get("cp") or "", tc.get("ville") or "",
                    "", tc.get("mobile1") or "", tc.get("adr_mail") or "",
                    op_id, bool(tc.get("opt_partenaire")),
                )
                if not id_clt:
                    item.nb_erreurs += 1
                    continue
                status = _enregistrer_ctt_sfr(
                    db_adv, num_bs, id_clt, id_vendeur, date_sign,
                    int(p.get("statut_prod") or 0),
                    p.get("motif_annulation") or "",
                    int(p.get("id_offres_sfr") or 0),
                    int(p.get("type_vente") or 0),
                    p.get("opt_tv") or "",
                    False,        # TicketDiff : pas dispo en PG, on met False
                    p.get("num_prise_optique") or "",
                    parcours_chaines, op_id,
                )
                if status == 4:
                    item.nb_crees += 1
                else:
                    item.nb_erreurs += 1

        # 5. Cloture le ticket
        try:
            db_tk.query(
                """UPDATE ticket.pgt_tk_liste
                      SET cloturee=TRUE, date_cloture=NOW(),
                          modif_date=NOW(), modif_op=?, modif_elem='modif'
                    WHERE id_tk_liste=?""",
                (int(op_id), int(id_tl)),
            )
            item.cloture_ok = True
        except Exception as e:
            item.message = f"Cloture KO : {e}"

        item.message = (
            f"{item.nb_crees} créé(s), {item.nb_existants} existant(s)"
            + (f", {item.nb_erreurs} erreur(s)" if item.nb_erreurs else "")
        )
        out.append(item)
    return out


def cloture_selection_sans_convertir(
    ids_tk_liste: list[int], op_id: int,
) -> list[ConversionResultItem]:
    """Cloture juste les tickets selectionnes (sans creer de contrat)."""
    db_tk = get_pg_connection("ticket")
    out: list[ConversionResultItem] = []
    for id_tl in ids_tk_liste:
        item = ConversionResultItem(id_tk_liste=str(id_tl))
        try:
            db_tk.query(
                """UPDATE ticket.pgt_tk_liste
                      SET cloturee=TRUE, date_cloture=NOW(),
                          modif_date=NOW(), modif_op=?, modif_elem='modif'
                    WHERE id_tk_liste=?""",
                (int(op_id), int(id_tl)),
            )
            item.cloture_ok = True
            item.message = "Clôturé"
        except Exception as e:
            item.message = f"KO : {e}"
        out.append(item)
    return out


def update_panier_num(
    id_panier: int, new_num: str, id_tk_liste: int, op_id: int,
) -> bool:
    """Met a jour TK_CallSFR_Panier.NUM + bascule TK_Liste.IDTK_Statut=17
    (SFR - Num BS SFR renseigne) cf code WinDev modif colonne NUM."""
    db_bo = get_pg_connection("ticket_bo")
    db_tk = get_pg_connection("ticket")
    norm = (new_num or "").strip().upper()
    db_bo.query(
        """UPDATE ticket_bo.pgt_tk_call_sfr_panier
              SET num=?, modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_tk_call_sfr_panier=?""",
        (norm, int(op_id), int(id_panier)),
    )
    db_tk.query(
        """UPDATE ticket.pgt_tk_liste
              SET id_tk_statut=17, modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_tk_liste=?""",
        (int(op_id), int(id_tk_liste)),
    )
    return True


def planning_appels(
    du: date, au: date, etat: str = "tous",
) -> list[PlanningRdvItem]:
    """Construit la liste des RDV pour le planning visuel (onglet
    'Analyse des Appels').

    Chaque ticket genere 2 RDV :
      1. Un RDV sur la ressource 'Crea Ticket' a date_crea (3 minutes)
      2. Un RDV sur la ressource = nom_operateur entre
         date_deb_prise_en_charge et date_fin_prise_en_charge.

    La couleur = categorie de delai (avant prise en charge).
    """
    rows = _load_ticket_call_sfr(du, au, etat)
    ids = [int(r["id_tk_call_sfr"]) for r in rows]
    paniers_by_id = _load_paniers_by_tickets(ids)

    out: list[PlanningRdvItem] = []
    for r in rows:
        id_tc = int(r["id_tk_call_sfr"])
        crea = r.get("date_crea")
        if not isinstance(crea, datetime):
            continue
        deb = r.get("date_deb_prise_en_charge")
        fin = r.get("date_fin_prise_en_charge")

        # Col_DelaiAvPriseCharge cf code WinDev :
        # DateHeureDifference(DH_crea, DH_Deb) -> delai = deb - crea
        # (et non deb - date_h_appel comme on avait initialement).
        delai = 0.0
        if isinstance(deb, datetime) and isinstance(crea, datetime):
            delai = (deb - crea).total_seconds() / 60.0
            if delai < 0:
                delai = 0.0
        delai_lbl = _delai_label(delai)
        couleur = _COULEURS_DELAI[delai_lbl]

        # Resume panier (multiligne)
        paniers = paniers_by_id.get(id_tc, [])
        nb_valide = sum(1 for p in paniers if int(p.get("statut_prod") or 0) in (1, 3))
        lines = []
        for p in paniers:
            num = p.get("num") or ""
            etat_p = _statut_prod_label(p.get("statut_prod"))
            d = p.get("num_date_saisie")
            line = num
            if d: line += f" ({_date_str(d)})"
            line += f" - {etat_p}"
            lines.append(line)
        contenu = "\n".join(lines)

        titre = (
            (r.get("nom_client") or "").strip() + " "
            + _capitalize((r.get("prenom_client") or "").strip())
        ).strip()

        # On exclut les tickets faits par l'op 6 (cf WinDev op_crea<>6)
        # (deja filtre dans _load_ticket_call_sfr).

        # RDV 1 : creation du ticket (Crea Ticket, 3 min)
        crea_min = crea.replace(microsecond=0, second=0)
        fin_crea = crea_min + timedelta(minutes=3)
        out.append(PlanningRdvItem(
            titre=titre, contenu=contenu,
            date_debut=crea_min.strftime("%Y-%m-%d %H:%M:%S"),
            date_fin=fin_crea.strftime("%Y-%m-%d %H:%M:%S"),
            ressource="Crea Ticket",
            couleur=couleur, delai_label=delai_lbl,
            delai_min=round(delai, 1), nb_valide=nb_valide,
        ))

        # RDV 2 : appel par l'operateur (entre date_deb et date_fin)
        if isinstance(deb, datetime) and isinstance(fin, datetime):
            if deb > fin:
                fin = deb + timedelta(minutes=3)
            sa_nom = (r.get("sa_nom") or "").strip()
            sa_prenom = _capitalize((r.get("sa_prenom") or "").strip())
            nom_ope = f"{sa_nom} {sa_prenom}".strip() or "—"
            out.append(PlanningRdvItem(
                titre=titre, contenu=contenu,
                date_debut=deb.strftime("%Y-%m-%d %H:%M:%S"),
                date_fin=fin.strftime("%Y-%m-%d %H:%M:%S"),
                ressource=nom_ope,
                couleur=couleur, delai_label=delai_lbl,
                delai_min=round(delai, 1), nb_valide=nb_valide,
            ))

    return out


def analyse_ventes_tk_call_sfr(
    du: date, au: date, etat: str = "tous",
) -> AnalyseVentesTotaux:
    """Onglet 'Analyse des ventes' : compteurs validees/annulees/pas_statuees
    + repartition par delai (<3m, 3-5m, 5-7m, >7m)."""
    rows = _load_ticket_call_sfr(du, au, etat)
    ids = [int(r["id_tk_call_sfr"]) for r in rows]
    paniers_by_id = _load_paniers_by_tickets(ids)

    totaux = AnalyseVentesTotaux()
    par_delai: dict[str, AnalyseVentesItem] = {}
    delai_order = ["< 3 min", "Entre 3 et 5 min", "Entre 5 et 7 min", "> 7 min"]
    for d in delai_order:
        par_delai[d] = AnalyseVentesItem(delai=d)

    for r in rows:
        id_tc = int(r["id_tk_call_sfr"])
        deb = r.get("date_deb_prise_en_charge")
        crea = r.get("date_crea")
        # Col_DelaiAvPriseCharge WinDev = date_deb - date_crea
        delai = 0.0
        if isinstance(deb, datetime) and isinstance(crea, datetime):
            delai = (deb - crea).total_seconds() / 60.0
            if delai < 0: delai = 0.0
        delai_lbl = _delai_label(delai)

        for p in paniers_by_id.get(id_tc, []):
            s = int(p.get("statut_prod") or 0)
            if s in (1, 3):
                totaux.ventes_validees += 1
                par_delai[delai_lbl].ventes_valides += 1
            elif s == 2:
                totaux.ventes_annulees += 1
                par_delai[delai_lbl].ventes_annulees += 1
            else:
                totaux.pas_encore_statuees += 1

    totaux.par_delai = [par_delai[d] for d in delai_order]
    return totaux


def send_mails_to_bos(
    ids_contrats: list[int], op_id: int, test_mode: bool = False,
) -> list[SendMailsResult]:
    """Pour chaque contrat selectionne, recupere l'email BO du cluster
    et envoie un mail 'demande RDV technicien'. Met a jour
    sfr_contrat.mail_bo_envoye + mail_bo_date_envoi + info_interne.

    En mode test : envoie tout a a.loudieux@exosphere.fr (sujet prefixe).
    """
    from app.shared.notifications.mail import envoi_mail
    db = get_pg_connection("adv")
    out: list[SendMailsResult] = []

    for id_ctt in ids_contrats:
        r = db.query_one(
            """
            SELECT c.id_contrat, c.num_bs, c.info_interne, c.id_sfr_cluster,
                   clu.mail_bo, clu.nom_cluster
              FROM adv.pgt_sfr_contrat c
              LEFT JOIN adv.pgt_sfr_cluster clu
                ON clu.id_sfr_cluster = c.id_sfr_cluster
             WHERE c.id_contrat = ?
            """,
            (int(id_ctt),),
        )
        if not r:
            out.append(SendMailsResult(
                id_contrat=str(id_ctt), ok=False,
                message="Contrat introuvable",
            ))
            continue

        num_bs = r.get("num_bs") or ""
        mail_bo = (r.get("mail_bo") or "").strip()
        if not mail_bo or "@" not in mail_bo:
            out.append(SendMailsResult(
                id_contrat=str(id_ctt), num_bs=num_bs, ok=False,
                message=f"Mail BO cluster absent/invalide : {mail_bo!r}",
            ))
            continue

        sujet = f"N° {num_bs} – demande rdv technicien"
        html = (
            "Bonjour,<br/><br/>"
            "<p>Ce contrat n'a pas encore de rendez-vous technicien planifié.<br/>"
            "Pouvez-vous le positionner dès que possible et nous en tenir informé.</p>"
            "<p>Je vous remercie par avance.<br/>Cdt,</p>"
            "<p><b>BO Exosphere</b><br/>"
            "<i>Standard : 03.62.27.60.04<br/>"
            "Mail : bo@exosphere.fr</i></p>"
        )

        # Mode test : tout va sur le compte de la dev
        if test_mode:
            to = ["a.loudieux@exosphere.fr"]
            cci = None
            sujet = "TEST - " + sujet
        else:
            to = [mail_bo]
            cci = [
                "bo@exosphere.fr",
                "m.doineau@exosphere.fr",
                "cuneyt.caliskan@sfr.com",
                "intranet@omaya.fr",
            ]

        try:
            sent_ok = envoi_mail(
                sujet=sujet, html=html, destinataires=to,
                cci=cci, expediteur="bo@exosphere.fr",
            )
        except Exception as e:
            out.append(SendMailsResult(
                id_contrat=str(id_ctt), num_bs=num_bs, ok=False,
                message=f"Erreur envoi : {e}",
            ))
            continue

        if not sent_ok:
            out.append(SendMailsResult(
                id_contrat=str(id_ctt), num_bs=num_bs, ok=False,
                message="Echec SMTP",
            ))
            continue

        # Update contrat : marque le mail envoye + log dans info_interne
        old_info = r.get("info_interne") or ""
        new_info = (
            old_info + ("\n" if old_info else "")
            + f"Mail envoyé au BO SFR ({mail_bo}) le "
            + date.today().strftime("%d/%m/%Y")
        )
        try:
            db.query(
                """UPDATE adv.pgt_sfr_contrat
                      SET mail_bo_envoye = TRUE,
                          mail_bo_date_envoi = ?,
                          info_interne = ?,
                          modif_date = NOW(), modif_op = ?,
                          modif_elem = 'modif'
                    WHERE id_contrat = ?""",
                (date.today(), new_info, int(op_id), int(id_ctt)),
            )
        except Exception as e:
            out.append(SendMailsResult(
                id_contrat=str(id_ctt), num_bs=num_bs, ok=True,
                message=f"Mail OK mais erreur UPDATE : {e}",
            ))
            continue

        out.append(SendMailsResult(
            id_contrat=str(id_ctt), num_bs=num_bs, ok=True,
            message=f"Envoyé à {mail_bo}" + (" (TEST)" if test_mode else ""),
        ))

    return out
