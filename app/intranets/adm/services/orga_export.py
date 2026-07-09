"""
Btn 'Exporter la selection' de Fen_Organigramme (WinDev).

Genere un XLSX avec la liste des salaries d'un bloc orga + ses
descendants, avec les mêmes colonnes et styles couleurs que le
WinDev :
  - rouge (204,0,0)   : vendeur non productif
  - violet (153,0,204) : en pause

Colonnes :
  Nom | Prenom | Date Embauche | Agence (orga parent) | Equipe
  (orga courant) | Entite (RS_Interne) | Poste | Resp Equipe |
  Production | NB Ctt W edites | NB Ctt W recus | En pause
"""
from __future__ import annotations

import io
import logging
from datetime import date as _date
from typing import Any

from app.core.database.pg import get_pg_connection


logger = logging.getLogger(__name__)


def _iso_date(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)[:10]
    if s.startswith("1900") or s.startswith("0000"):
        return ""
    return s


def _fmt_date_fr(iso: str) -> str:
    if not iso or len(iso) < 10:
        return ""
    return f"{iso[8:10]}/{iso[5:7]}/{iso[0:4]}"


def _collect_orga_tree(id_orga: int) -> list[dict]:
    """Cf. WinDev ListeOrgaComplet : bloc courant + descendants.
    Retourne [{id, lib, parent_lib}, ...].
    """
    rh = get_pg_connection("rh")
    try:
        rows = rh.query(
            """WITH RECURSIVE tree AS (
                   SELECT idorganigramme AS id, id_parent, lib_orga, 0 AS depth
                     FROM pgt_organigramme
                    WHERE idorganigramme = ?
                      AND (modif_elem IS NULL
                           OR modif_elem NOT LIKE '%suppr%')
                   UNION ALL
                   SELECT o.idorganigramme, o.id_parent, o.lib_orga,
                          t.depth + 1
                     FROM pgt_organigramme o
                     JOIN tree t ON o.id_parent = t.id
                    WHERE (o.modif_elem IS NULL
                           OR o.modif_elem NOT LIKE '%suppr%')
                      AND t.depth < 20
               )
               SELECT id, id_parent, lib_orga FROM tree""",
            (id_orga,),
        ) or []
    except Exception as e:
        logger.exception("_collect_orga_tree id_orga=%s", id_orga)
        # Remonte l'erreur pour que l'endpoint expose un 500 explicite
        raise RuntimeError(f"CTE arbre orga : {e}") from e
    # Map id -> lib pour resoudre le parent_lib
    lib_by_id = {int(r.get("id") or 0): (r.get("lib_orga") or "").strip()
                 for r in rows}
    out: list[dict] = []
    for r in rows:
        oid = int(r.get("id") or 0)
        pid = int(r.get("id_parent") or 0)
        out.append({
            "id": oid,
            "lib": lib_by_id.get(oid, ""),
            "parent_lib": lib_by_id.get(pid, ""),
        })
    return out


def _load_salaries_by_orga(
    orga_ids: list[int], today_iso: str,
) -> list[dict]:
    """Cf. WinDev reqSa : salaries actifs rattaches, tries par
    RespEquipe / RespAdjoint / Nom / Prenom.
    """
    if not orga_ids:
        return []
    rh = get_pg_connection("rh")
    ids_sql = ",".join(str(i) for i in orga_ids)
    try:
        rows = rh.query(
            f"""SELECT DISTINCT
                       so.id_salarie, so.idorganigramme,
                       s.nom, s.prenom,
                       se.id_ste, se.id_type_poste,
                       se.resp_equipe, se.resp_adjoint,
                       se.date_anciennete, se.en_pause, se.id_absence
                  FROM rh.pgt_salarie s
                  JOIN rh.pgt_salarie_embauche se ON se.id_salarie = s.id_salarie
                  JOIN rh.pgt_salarie_organigramme so ON so.id_salarie = s.id_salarie
                 WHERE so.idorganigramme IN ({ids_sql})
                   AND (so.modif_elem IS NULL
                        OR so.modif_elem NOT LIKE '%suppr%')
                   AND so.date_debut IS NOT NULL
                   AND so.date_debut <= ?::date
                   AND (so.date_fin IS NULL
                        OR so.date_fin = '1900-01-01'
                        OR so.date_fin >= ?::date)
                   AND se.en_activite = TRUE
                   AND (s.modif_elem IS NULL
                        OR s.modif_elem NOT LIKE '%suppr%')
                 ORDER BY se.resp_equipe DESC NULLS LAST,
                          se.resp_adjoint DESC NULLS LAST,
                          s.nom ASC, s.prenom ASC""",
            (today_iso, today_iso),
        ) or []
    except Exception:
        logger.exception("_load_salaries_by_orga")
        return []
    return rows


def _resolve_lookups(rows: list[dict]) -> tuple[dict, dict]:
    """Charge societes + type_poste pour les libellés."""
    id_stes = list({int(r.get("id_ste") or 0) for r in rows} - {0})
    id_postes = list({int(r.get("id_type_poste") or 0) for r in rows} - {0})
    rh = get_pg_connection("rh")
    stes: dict[int, str] = {}
    postes: dict[int, str] = {}
    if id_stes:
        try:
            for r in rh.query(
                f"""SELECT id_ste, rs_interne, raison_sociale
                      FROM pgt_societe
                     WHERE id_ste IN ({','.join(str(i) for i in id_stes)})""",
            ) or []:
                stes[int(r.get("id_ste") or 0)] = (
                    (r.get("rs_interne") or r.get("raison_sociale") or "").strip()
                )
        except Exception:
            logger.exception("_resolve_lookups societes")
    if id_postes:
        try:
            for r in rh.query(
                f"""SELECT id_type_poste, lib_poste
                      FROM pgt_type_poste
                     WHERE id_type_poste IN ({','.join(str(i) for i in id_postes)})""",
            ) or []:
                postes[int(r.get("id_type_poste") or 0)] = (
                    (r.get("lib_poste") or "").strip()
                )
        except Exception:
            logger.exception("_resolve_lookups postes")
    return stes, postes


def _load_dernier_ctt(ids_salarie: set[int], today_iso: str) -> dict[int, dict]:
    """Cherche le dernier contrat signe (tous partenaires) par salarie.
    Retourne {id_salarie: {date, partenaire}}.
    """
    if not ids_salarie:
        return {}
    adv = get_pg_connection("adv")
    try:
        parts = adv.query(
            """SELECT prefixe_bdd, lib_partenaire
                 FROM pgt_partenaire
                WHERE is_actif = TRUE AND modif_elem <> 'suppr'""",
        ) or []
    except Exception:
        return {}
    ids_sql = ",".join(f"'{i}'" for i in ids_salarie)
    out: dict[int, dict] = {}
    for part in parts:
        pref = (part.get("prefixe_bdd") or "").strip().lower()
        if not pref:
            continue
        lib_part = (part.get("lib_partenaire") or pref.upper()).strip()
        try:
            rows = adv.query(
                f"""SELECT id_salarie, MAX(date_signature) AS maxdate
                      FROM pgt_{pref}_contrat
                     WHERE id_salarie IN ({ids_sql})
                       AND date_signature IS NOT NULL
                       AND modif_elem NOT LIKE '%suppr%'
                       AND date_signature::date <= ?::date
                     GROUP BY id_salarie""",
                (today_iso,),
            ) or []
        except Exception:
            continue
        for r in rows:
            sid = int(r.get("id_salarie") or 0)
            d = _iso_date(r.get("maxdate") or r.get("max"))
            if not d:
                continue
            cur = out.get(sid)
            if not cur or d > cur["date"]:
                out[sid] = {"date": d, "partenaire": lib_part}
    return out


def _load_docrh_stats(ids_salarie: set[int]) -> dict[int, tuple[int, int]]:
    """Nb Ctt W edites et recus par salarie (cf. reqCttW WinDev)."""
    if not ids_salarie:
        return {}
    ids_sql = ",".join(str(i) for i in ids_salarie)
    rh = get_pg_connection("rh")
    try:
        rows = rh.query(
            f"""SELECT id_salarie,
                       COUNT(id_salarie_doc_rh) AS nb_edit,
                       COALESCE(SUM(CASE WHEN recu THEN 1 ELSE 0 END), 0) AS nb_recu
                  FROM pgt_salarie_doc_rh
                 WHERE id_salarie IN ({ids_sql})
                   AND modif_elem <> 'suppr'
                 GROUP BY id_salarie""",
        ) or []
    except Exception:
        logger.exception("_load_docrh_stats")
        return {}
    return {
        int(r.get("id_salarie") or 0): (
            int(r.get("nb_edit") or 0),
            int(r.get("nb_recu") or 0),
        )
        for r in rows
    }


def _load_absences_debut(ids_absence: set[int]) -> dict[int, str]:
    """Date de debut d'absence (pour la mention 'En pause depuis le XX')."""
    if not ids_absence:
        return {}
    ids_sql = ",".join(str(i) for i in ids_absence)
    rh = get_pg_connection("rh")
    try:
        rows = rh.query(
            f"""SELECT id_absence, date_debut FROM pgt_absence
                 WHERE id_absence IN ({ids_sql})""",
        ) or []
    except Exception:
        return {}
    return {
        int(r.get("id_absence") or 0): _iso_date(r.get("date_debut"))
        for r in rows
    }


# --------------------------------------------------------------------
# Point d'entree : XLSX bytes
# --------------------------------------------------------------------

def export_orga_selection_xlsx(id_orga: str) -> bytes:
    """Genere le fichier Excel de la selection."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        id_o = int(id_orga)
    except (TypeError, ValueError):
        raise ValueError("id_orga invalide")

    logger.info("[export_orga] id=%s : step 1 collect tree", id_o)
    orgas = _collect_orga_tree(id_o)
    if not orgas:
        raise ValueError("Bloc introuvable")
    orga_ids = [o["id"] for o in orgas]
    orga_by_id = {o["id"]: o for o in orgas}
    logger.info(
        "[export_orga] id=%s : tree=%d orgas", id_o, len(orga_ids),
    )

    today = _date.today().isoformat()
    try:
        salaries = _load_salaries_by_orga(orga_ids, today)
    except Exception as e:
        raise RuntimeError(f"Chargement salaries : {e}") from e
    logger.info(
        "[export_orga] id=%s : %d salaries", id_o, len(salaries),
    )
    stes, postes = _resolve_lookups(salaries)
    ids_sal = {int(r.get("id_salarie") or 0) for r in salaries}
    try:
        dernier_ctt = _load_dernier_ctt(ids_sal, today)
    except Exception as e:
        logger.exception("dernier_ctt")
        dernier_ctt = {}
    try:
        docrh = _load_docrh_stats(ids_sal)
    except Exception as e:
        logger.exception("docrh")
        docrh = {}
    ids_abs = {
        int(r.get("id_absence") or 0)
        for r in salaries if r.get("en_pause")
    } - {0}
    try:
        abs_debut = _load_absences_debut(ids_abs)
    except Exception as e:
        logger.exception("abs_debut")
        abs_debut = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Liste des salariés"

    headers = [
        "Nom", "Prénom", "Date Embauche",
        "Agence", "Equipe", "Entité", "Poste",
        "Resp Equipe", "Production",
        "NB Ctt W édités", "NB Ctt W reçus", "En pause",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17494E")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for r in salaries:
        sid = int(r.get("id_salarie") or 0)
        oid = int(r.get("idorganigramme") or 0)
        o = orga_by_id.get(oid, {})
        prod_info = dernier_ctt.get(sid)
        if prod_info:
            prod = (
                f"Dernier ctt signé : {prod_info['partenaire']} le "
                f"{_fmt_date_fr(prod_info['date'])}"
            )
            non_prod = False
        else:
            prod = "Pas encore productif"
            non_prod = True
        en_pause = bool(r.get("en_pause"))
        pause_txt = ""
        if en_pause:
            id_abs = int(r.get("id_absence") or 0)
            db_abs = abs_debut.get(id_abs, "")
            pause_txt = (
                f"En pause depuis le {_fmt_date_fr(db_abs)}"
                if db_abs else "En pause"
            )
        edit_recu = docrh.get(sid, (0, 0))

        values = [
            r.get("nom") or "",
            (r.get("prenom") or "").strip(),
            _fmt_date_fr(_iso_date(r.get("date_anciennete"))),
            o.get("parent_lib", ""),
            o.get("lib", ""),
            stes.get(int(r.get("id_ste") or 0), ""),
            postes.get(int(r.get("id_type_poste") or 0), ""),
            "Oui" if r.get("resp_equipe") else "",
            prod,
            edit_recu[0],
            edit_recu[1],
            pause_txt,
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            # Couleurs conditionnelles (Nom/Prenom/Production)
            if col in (1, 2, 9):
                if en_pause:
                    cell.font = Font(color="9900CC")   # violet
                elif non_prod:
                    cell.font = Font(color="CC0000")   # rouge
        row_idx += 1

    # Largeurs colonnes raisonnables
    widths = [18, 15, 14, 22, 22, 14, 20, 12, 40, 12, 12, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
