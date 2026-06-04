"""
Service Registre RH — transposition de la fenetre WinDev Fen_RegistreRH.

Source de donnees : PostgreSQL (schema rh) via get_pg_connection("rh").

La page WinDev affiche une combo Societe (filtre IDTypeOrga=1) + un tableau
jointure (salarie + coordonnees + embauche + sortie + type_poste) filtre
par IdSte. Tri par DateDebut DESC.
"""

import io
from datetime import date, datetime
from typing import Any

from app.core.database.pg import get_pg_connection


def _str_id(v: Any) -> str:
    """IDs 8 octets exposes en str (cf. feedback_ids_8octets_string)."""
    if v is None:
        return ""
    s = str(v).strip()
    return s if s and s != "0" else ""


def _iso(v: Any) -> str:
    """Date ou timestamp PG -> ISO 'YYYY-MM-DD' (vide si null/zero)."""
    if v is None or v == "":
        return ""
    if isinstance(v, (date, datetime)):
        if v.year < 1900:
            return ""
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s or s.startswith("0000") or s.startswith("1900"):
        return ""
    return s[:10]


def _int(v: Any) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(v)
    except (ValueError, TypeError):
        return 0


def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def list_societes() -> list[dict]:
    """Combo Societe : filtre IDTypeOrga=1 (Internes), tri par RaisonSociale."""
    db = get_pg_connection("rh")
    rows = db.query(
        """SELECT id_ste, rs_interne, raison_sociale
        FROM rh.pgt_societe
        WHERE modif_elem NOT LIKE '%suppr%'
          AND id_type_orga = 1
        ORDER BY raison_sociale ASC NULLS LAST"""
    )
    return [
        {
            "id_ste": _str_id(r.get("id_ste")),
            "rs_interne": _str(r.get("rs_interne")) or _str(r.get("raison_sociale")),
        }
        for r in rows
    ]


def list_refs() -> dict:
    """Combos colonnes : type_ctt, type_horaire, type_sortie."""
    db = get_pg_connection("rh")

    type_ctt = db.query(
        """SELECT id_type_ctt, intitule
        FROM rh.pgt_type_ctt_travail
        WHERE modif_elem NOT LIKE '%suppr%'
        ORDER BY intitule ASC NULLS LAST"""
    )
    type_horaire = db.query(
        """SELECT id_type_horaire, lib_horaire
        FROM rh.pgt_type_horaire_travail
        WHERE modif_elem NOT LIKE '%suppr%'
        ORDER BY lib_horaire ASC NULLS LAST"""
    )
    type_sortie = db.query(
        """SELECT id_type_sortie, lib_sortie
        FROM rh.pgt_type_sortie_salarie
        WHERE modif_elem NOT LIKE '%suppr%'
        ORDER BY lib_sortie ASC NULLS LAST"""
    )

    return {
        "type_ctt": [
            {"id": _int(r.get("id_type_ctt")), "label": _str(r.get("intitule"))}
            for r in type_ctt
        ],
        "type_horaire": [
            {"id": _int(r.get("id_type_horaire")), "label": _str(r.get("lib_horaire"))}
            for r in type_horaire
        ],
        "type_sortie": [
            {"id": _int(r.get("id_type_sortie")), "label": _str(r.get("lib_sortie"))}
            for r in type_sortie
        ],
    }


def list_salaries(id_ste: int) -> list[dict]:
    """
    Liste des salaries d'une societe.

    Transposition de la requete WinDev Table_ReqRegistreRH avec les jointures :
      salarie
      LEFT JOIN salarie_coordonnees ON id_salarie
      LEFT JOIN salarie_embauche    ON id_salarie
      LEFT JOIN type_poste          ON salarie_embauche.id_type_poste
      LEFT JOIN salarie_sortie      ON id_salarie
    WHERE salarie_embauche.id_ste = :id_ste AND salarie.modif_elem NOT LIKE '%suppr%'
    ORDER BY date_debut DESC.
    """
    db = get_pg_connection("rh")
    rows = db.query(
        """SELECT
            s.id_salarie,
            s.civilite, s.nom, s.prenom, s.sexe, s.nationalite,
            s.date_naiss, s.lieu_naiss, s.dep_naiss,
            s.num_ss, s.cpam, s.num_cin, s.travailleur_handi,
            sc.adresse1, sc.adresse2, sc.cp, sc.ville,
            sc.tel_mob, sc.mail, sc.iban,
            sc.urg_nom, sc.urg_lien, sc.urg_tel,
            se.id_ste, se.date_debut, se.date_fin_per_essai,
            se.dpae_num, se.dpae_date,
            se.id_type_poste, tp.lib_poste,
            se.id_type_ctt, se.id_type_horaire,
            se.en_activite, se.coopte, se.coopteur,
            ss.date_sortie_demandee, ss.date_sortie_reelle,
            ss.demandeur_sortie, ss.id_type_sortie
        FROM rh.pgt_salarie s
        LEFT JOIN rh.pgt_salarie_coordonnees sc ON sc.id_salarie = s.id_salarie
        LEFT JOIN rh.pgt_salarie_embauche se    ON se.id_salarie = s.id_salarie
        LEFT JOIN rh.pgt_type_poste tp          ON tp.id_type_poste = se.id_type_poste
        LEFT JOIN rh.pgt_salarie_sortie ss      ON ss.id_salarie = s.id_salarie
        WHERE se.id_ste = ?
          AND s.modif_elem NOT LIKE '%suppr%'
        ORDER BY se.date_debut DESC NULLS LAST""",
        (id_ste,),
    )
    return [
        {
            "id_salarie": _str_id(r.get("id_salarie")),
            "civilite": _int(r.get("civilite")),
            "nom": _str(r.get("nom")),
            "prenom": _str(r.get("prenom")),
            "sexe": _str(r.get("sexe")),
            "nationalite": _str(r.get("nationalite")),
            "date_naiss": _iso(r.get("date_naiss")),
            "lieu_naiss": _str(r.get("lieu_naiss")),
            "dep_naiss": _int(r.get("dep_naiss")),
            "num_ss": _str(r.get("num_ss")),
            "cpam": _str(r.get("cpam")),
            "num_cin": _str(r.get("num_cin")),
            "travailleur_handi": bool(r.get("travailleur_handi")),
            "adresse1": _str(r.get("adresse1")),
            "adresse2": _str(r.get("adresse2")),
            "cp": _str(r.get("cp")),
            "ville": _str(r.get("ville")),
            "tel_mob": _str(r.get("tel_mob")),
            "mail": _str(r.get("mail")),
            "iban": _str(r.get("iban")),
            "urg_nom": _str(r.get("urg_nom")),
            "urg_lien": _str(r.get("urg_lien")),
            "urg_tel": _str(r.get("urg_tel")),
            "id_ste": _str_id(r.get("id_ste")),
            "date_debut": _iso(r.get("date_debut")),
            "date_fin_per_essai": _iso(r.get("date_fin_per_essai")),
            "dpae_num": _str(r.get("dpae_num")),
            "dpae_date": _iso(r.get("dpae_date")),
            "id_type_poste": _int(r.get("id_type_poste")),
            "lib_poste": _str(r.get("lib_poste")),
            "id_type_ctt": _int(r.get("id_type_ctt")),
            "id_type_horaire": _int(r.get("id_type_horaire")),
            "en_activite": bool(r.get("en_activite")),
            "coopte": bool(r.get("coopte")),
            "coopteur": _str_id(r.get("coopteur")),
            "date_sortie_demandee": _iso(r.get("date_sortie_demandee")),
            "date_sortie_reelle": _iso(r.get("date_sortie_reelle")),
            "demandeur_sortie": _str_id(r.get("demandeur_sortie")),
            "id_type_sortie": _int(r.get("id_type_sortie")),
        }
        for r in rows
    ]


# --- Export Excel --------------------------------------------------------

CIVILITE_LIBS = {1: "M.", 2: "Mme", 3: "Mlle"}


def _date_fr(iso: str) -> str:
    """ISO 'YYYY-MM-DD' -> 'DD/MM/YYYY' (vide si vide)."""
    if not iso or len(iso) < 10:
        return ""
    return f"{iso[8:10]}/{iso[5:7]}/{iso[0:4]}"


def export_xlsx(id_ste: int) -> tuple[bytes, str]:
    """Genere un .xlsx du Registre RH pour une societe.

    Retourne (bytes du fichier, nom de fichier suggere).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    db = get_pg_connection("rh")

    # Nom de la societe (pour le titre + nom de fichier)
    soc_row = db.query_one(
        "SELECT rs_interne, raison_sociale FROM rh.pgt_societe WHERE id_ste = ?",
        (id_ste,),
    )
    nom_soc = ""
    if soc_row:
        nom_soc = (soc_row.get("rs_interne") or soc_row.get("raison_sociale") or "").strip()

    # Refs pour traduire les ids
    refs = list_refs()
    ctt_by_id = {r["id"]: r["label"] for r in refs["type_ctt"]}
    horaire_by_id = {r["id"]: r["label"] for r in refs["type_horaire"]}
    sortie_by_id = {r["id"]: r["label"] for r in refs["type_sortie"]}

    rows = list_salaries(id_ste)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registre RH"

    # En-tete : titre + societe + date
    titre_font = Font(bold=True, size=14, color="FFFFFF")
    titre_fill = PatternFill("solid", fgColor="17494E")
    ws.cell(row=1, column=1, value=f"Registre RH — {nom_soc}").font = titre_font
    ws.cell(row=1, column=1).fill = titre_fill
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=24)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.cell(row=2, column=1, value=f"Export du {datetime.now().strftime('%d/%m/%Y a %H:%M')}").font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=24)

    # En-tetes colonnes (ligne 4)
    headers = [
        "Civilité", "Nom", "Prénom", "Nationalité", "Sexe",
        "N° Sécu Soc", "Date Naiss", "Lieu Naiss", "Dép Naiss",
        "Adresse 1", "Adresse 2", "CP", "Ville",
        "N° CIN", "DPAE n°",
        "En activité", "Date début", "Type contrat", "Poste",
        "Date sortie demandée", "Date sortie réelle", "Type sortie",
        "Type horaire", "RQTH",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="4E1D17")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "A5"  # fige le header

    # Lignes
    for i, r in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=CIVILITE_LIBS.get(r["civilite"], ""))
        ws.cell(row=i, column=2, value=r["nom"])
        ws.cell(row=i, column=3, value=r["prenom"])
        ws.cell(row=i, column=4, value=r["nationalite"])
        ws.cell(row=i, column=5, value=r["sexe"])
        ws.cell(row=i, column=6, value=r["num_ss"])
        ws.cell(row=i, column=7, value=_date_fr(r["date_naiss"]))
        ws.cell(row=i, column=8, value=r["lieu_naiss"])
        ws.cell(row=i, column=9, value=r["dep_naiss"] or "")
        ws.cell(row=i, column=10, value=r["adresse1"])
        ws.cell(row=i, column=11, value=r["adresse2"])
        ws.cell(row=i, column=12, value=r["cp"])
        ws.cell(row=i, column=13, value=r["ville"])
        ws.cell(row=i, column=14, value=r["num_cin"])
        ws.cell(row=i, column=15, value=r["dpae_num"])
        ws.cell(row=i, column=16, value="Oui" if r["en_activite"] else "Non")
        ws.cell(row=i, column=17, value=_date_fr(r["date_debut"]))
        ws.cell(row=i, column=18, value=ctt_by_id.get(r["id_type_ctt"], ""))
        ws.cell(row=i, column=19, value=r["lib_poste"])
        ws.cell(row=i, column=20, value=_date_fr(r["date_sortie_demandee"]))
        ws.cell(row=i, column=21, value=_date_fr(r["date_sortie_reelle"]))
        ws.cell(row=i, column=22, value=sortie_by_id.get(r["id_type_sortie"], ""))
        ws.cell(row=i, column=23, value=horaire_by_id.get(r["id_type_horaire"], ""))
        ws.cell(row=i, column=24, value="Oui" if r["travailleur_handi"] else "")

    # Largeurs auto-ish : on calcule la longueur max par colonne
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(h)
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        # Borne raisonnable : entre 8 et 35
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 35)

    # Footer
    footer_row = ws.max_row + 2
    nb_total = len(rows)
    nb_actifs = sum(1 for r in rows if r["en_activite"])
    ws.cell(row=footer_row, column=1, value=f"{nb_total} salarié(s) dont {nb_actifs} en activité").font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(start_row=footer_row, end_row=footer_row, start_column=1, end_column=24)

    # Serialise en bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Nom de fichier (avec sanitization basique)
    safe_soc = "".join(c if c.isalnum() or c in " -_" else "_" for c in nom_soc).strip() or f"ste_{id_ste}"
    filename = f"Registre_RH_{safe_soc}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buf.getvalue(), filename
