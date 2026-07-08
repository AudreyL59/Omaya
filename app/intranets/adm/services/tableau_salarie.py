"""
Service Fen_TableauSalarie - Tableau des salaries par equipe.

Cf. WinDev creaListeVendeur() :
  1. Recupere toutes les orgas descendantes recursivement (ListeOrgaComplet)
  2. Pour chacune : query salariés directement rattaches, chevauchant
     la periode [DateDeb, DateFin]
  3. Pour chaque salarie : lookup activite, sortie, absences, avances
  4. Ajoute ou complete Eq_Terrain (2eme niveau si vendeur en doublon)
  5. Trie par Eq_Terrain, Poste, Nom
"""
from __future__ import annotations

import io as _io
import logging
import re
import unicodedata
from datetime import date, timedelta
from typing import Optional

from app.core.database.pg import get_pg_connection
from app.intranets.adm.schemas.tableau_salarie import (
    ExportXlsxParams, OrgaCombo, RechercherParams, RechercherResult,
    VendeurRow,
)

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------

def _clean_id(v) -> str:
    if v is None:
        return ""
    try:
        n = int(v)
        return str(n) if n else ""
    except (TypeError, ValueError):
        return ""


def _iso_date(v) -> str:
    if v is None:
        return ""
    s = str(v)[:10]
    if s.startswith("1900") or s.startswith("0000"):
        return ""
    return s


def _cap_prenom(p: str) -> str:
    if not p:
        return ""
    return p[0].upper() + p[1:].lower() if len(p) > 1 else p.upper()


def _fr_date_jjmm_aaaa(d) -> str:
    """Formate 'Jjj JJ Mmm AAAA' (cf. WinDev)."""
    if d is None:
        return ""
    if isinstance(d, str):
        try:
            d = date.fromisoformat(d[:10])
        except Exception:
            return d[:10]
    if not hasattr(d, "strftime"):
        return str(d)[:10]
    jours = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    mois = [
        "", "Janv", "Févr", "Mars", "Avr", "Mai", "Juin",
        "Juil", "Août", "Sept", "Oct", "Nov", "Déc",
    ]
    return f"{jours[d.weekday()]} {d.day:02d} {mois[d.month]} {d.year}"


def _premier_jour(mois_paiement: str) -> str:
    """'2026-07' -> '2026-07-01'."""
    return mois_paiement + "-01"


def _dernier_jour(mois_paiement: str) -> str:
    """'2026-07' -> '2026-07-31'."""
    y = int(mois_paiement[:4])
    m = int(mois_paiement[5:7])
    if m == 12:
        d0 = date(y + 1, 1, 1)
    else:
        d0 = date(y, m + 1, 1)
    return (d0 - timedelta(days=1)).isoformat()


# --------------------------------------------------------------------
# Combo Orgas (equipes/agences pour picker)
# --------------------------------------------------------------------

def list_orgas() -> list[OrgaCombo]:
    """Liste toutes les orgas actives niveau 3 (agence) ou 4 (equipe)
    pour picker. Ajoute le libelle du parent pour lisibilite.
    """
    rh = get_pg_connection("rh")
    try:
        rows = rh.query(
            """SELECT o.idorganigramme, o.lib_orga,
                      p.lib_orga AS lib_parent
                 FROM pgt_organigramme o
                 LEFT JOIN pgt_organigramme p
                        ON p.idorganigramme = o.id_parent
                WHERE (o.modif_elem IS NULL
                       OR o.modif_elem NOT LIKE '%suppr%')
                  AND o.id_type_niveau_orga IN (3, 4)
                ORDER BY p.lib_orga ASC NULLS LAST, o.lib_orga ASC""",
        ) or []
    except Exception:
        logger.exception("list_orgas")
        return []
    return [
        OrgaCombo(
            id_orga=_clean_id(r.get("idorganigramme")),
            lib_orga=(r.get("lib_orga") or "").strip(),
            lib_parent=(r.get("lib_parent") or "").strip(),
        )
        for r in rows
    ]


# --------------------------------------------------------------------
# ListeOrgaComplet : recuperation recursive
# --------------------------------------------------------------------

def _liste_orga_complet(id_racine: int) -> list[dict]:
    """Cf. WinDev ListeOrgaComplet : retourne tous les orgas descendants
    (racine incluse) avec libelle du parent.
    """
    rh = get_pg_connection("rh")
    seen: set[int] = {int(id_racine)}
    to_process = [int(id_racine)]
    result_ids: list[int] = [int(id_racine)]
    while to_process:
        parents = tuple(to_process)
        to_process = []
        try:
            placeholders = ",".join(["?"] * len(parents))
            rows = rh.query(
                f"""SELECT idorganigramme FROM pgt_organigramme
                     WHERE id_parent IN ({placeholders})
                       AND (modif_elem IS NULL
                            OR modif_elem NOT LIKE '%suppr%')""",
                parents,
            ) or []
        except Exception:
            break
        for r in rows:
            oid = int(r.get("idorganigramme") or 0)
            if oid and oid not in seen:
                seen.add(oid)
                to_process.append(oid)
                result_ids.append(oid)

    # Charge les infos completes pour chaque orga
    if not result_ids:
        return []
    placeholders = ",".join(["?"] * len(result_ids))
    try:
        rows = rh.query(
            f"""SELECT o.idorganigramme, o.lib_orga,
                      p.lib_orga AS parent_lib
                 FROM pgt_organigramme o
                 LEFT JOIN pgt_organigramme p
                        ON p.idorganigramme = o.id_parent
                WHERE o.idorganigramme IN ({placeholders})""",
            tuple(result_ids),
        ) or []
    except Exception:
        return []
    return rows


# --------------------------------------------------------------------
# creaListeVendeur (proc principale)
# --------------------------------------------------------------------

def rechercher(p: RechercherParams) -> RechercherResult:
    """Cf. WinDev creaListeVendeur() : recupere tous les salaries
    de l'orga selectionnee + ses descendants, sur le mois de paiement.
    """
    if not p.id_orga or p.id_orga == "0":
        return RechercherResult(ok=False, message="Equipe requise")
    if not re.match(r"^\d{4}-\d{2}$", p.mois_paiement or ""):
        return RechercherResult(
            ok=False, message="Mois paiement invalide (YYYY-MM)",
        )

    date_deb = _premier_jour(p.mois_paiement)
    date_fin = _dernier_jour(p.mois_paiement)

    rh = get_pg_connection("rh")
    tk = get_pg_connection("ticket_bo")

    orgas = _liste_orga_complet(int(p.id_orga))
    if not orgas:
        return RechercherResult(
            ok=True, lignes=[],
            message="Aucun sous-organigramme trouve",
        )

    # Accumulateur : { id_salarie : (VendeurRow, index_dans_lignes) }
    vendeurs: dict[int, VendeurRow] = {}
    # Pour respecter l'ordre d'insertion
    ordre: list[int] = []

    for orga in orgas:
        id_orga = int(orga.get("idorganigramme") or 0)
        lib_orga = (orga.get("lib_orga") or "").strip()
        parent_lib = (orga.get("parent_lib") or "").strip()

        # Query salaries rattaches a cette orga sur la periode
        try:
            rows = rh.query(
                """SELECT DISTINCT ON (s.id_salarie)
                        s.id_salarie, s.nom, s.prenom,
                        so.date_debut, so.date_fin,
                        se.en_activite,
                        se.date_debut AS date_embauche,
                        tp.lib_poste,
                        se.resp_equipe
                     FROM pgt_salarie_organigramme so
                     JOIN pgt_salarie s ON s.id_salarie = so.id_salarie
                     LEFT JOIN pgt_salarie_embauche se
                            ON se.id_salarie = s.id_salarie
                           AND (se.modif_elem IS NULL
                                OR se.modif_elem NOT LIKE '%suppr%')
                     LEFT JOIN pgt_type_poste tp
                            ON tp.id_type_poste = se.id_type_poste
                    WHERE so.idorganigramme = ?
                      AND (so.modif_elem IS NULL
                           OR so.modif_elem NOT LIKE '%suppr%')
                      AND so.date_debut <= ?
                      AND (so.date_fin IS NULL
                           OR so.date_fin >= ?
                           OR so.date_fin = '1900-01-01')
                      AND (s.modif_elem IS NULL
                           OR s.modif_elem NOT LIKE '%suppr%')
                    ORDER BY s.id_salarie, se.date_debut DESC NULLS LAST""",
                (id_orga, date_fin, date_deb),
            ) or []
        except Exception:
            logger.exception("rechercher : query orga=%s", id_orga)
            continue

        # Affectation "eq_terrain" pour cette orga : le lib_orga
        affectation = lib_orga

        for r in rows:
            id_sal = int(r.get("id_salarie") or 0)
            if not id_sal:
                continue

            date_embauche_iso = _iso_date(r.get("date_embauche"))
            # Cf. WinDev : si date_embauche > sdateFin -> skip
            if date_embauche_iso and date_embauche_iso > date_fin:
                continue

            en_activite = bool(r.get("en_activite"))

            # Determine le type_sortie
            type_sortie = ""
            date_sortie_iso = ""
            if not en_activite:
                try:
                    sortie = rh.query_one(
                        """SELECT ss.date_sortie_demandee,
                                  ts.lib_sortie, ts.id_type_sortie
                             FROM pgt_salarie_sortie ss
                             LEFT JOIN pgt_type_sortie_salarie ts
                                    ON ts.id_type_sortie = ss.id_type_sortie
                            WHERE ss.id_salarie = ?
                              AND (ss.modif_elem IS NULL
                                   OR ss.modif_elem NOT LIKE '%suppr%')
                            ORDER BY ss.date_sortie_demandee DESC NULLS LAST
                            LIMIT 1""",
                        (id_sal,),
                    )
                except Exception:
                    sortie = None
                if sortie:
                    type_sortie = (sortie.get("lib_sortie") or "").strip()
                    date_sortie_iso = _iso_date(
                        sortie.get("date_sortie_demandee")
                    )
                    # Cf. WinDev : si sortie avant debut periode -> skip
                    if date_sortie_iso and date_sortie_iso < date_deb:
                        continue

            # Deja en table ? Mise a jour affectation si applicable
            if id_sal in vendeurs:
                existing = vendeurs[id_sal]
                if existing.eq_terrain != affectation and en_activite:
                    # Cf. WinDev : si en activite et affectation differente,
                    # on prend l'affectation la plus recente
                    existing.eq_terrain = affectation
                continue

            # Ajoute une nouvelle ligne
            row = VendeurRow(
                id_salarie=str(id_sal),
                nom=(r.get("nom") or "").strip(),
                prenom=_cap_prenom((r.get("prenom") or "").strip()),
                poste=(r.get("lib_poste") or "").strip() or "Non défini",
                is_actif=en_activite,
                is_sortie=not en_activite,
                date_entree=date_embauche_iso,
                type_sortie=type_sortie,
                eq_terrain=affectation,
                is_resp=bool(r.get("resp_equipe")),
                absences="",
                avance=0.0,
            )

            # Charge les absences
            try:
                absences = rh.query(
                    """SELECT a.date_debut, a.date_fin,
                              ta.lib_absence
                         FROM pgt_absence a
                         JOIN pgt_type_absence ta
                              ON ta.id_type_absence = a.id_type_absence
                        WHERE a.id_salarie = ?
                          AND (a.modif_elem IS NULL
                               OR a.modif_elem NOT LIKE '%suppr%')
                          AND (
                              (a.date_debut BETWEEN ? AND ?)
                              OR (a.date_fin BETWEEN ? AND ?)
                          )
                        ORDER BY a.date_debut ASC""",
                    (id_sal, date_deb, date_fin, date_deb, date_fin),
                ) or []
            except Exception:
                absences = []

            lignes_abs: list[str] = []
            for a in absences:
                lib = (a.get("lib_absence") or "").strip()
                d1 = a.get("date_debut")
                d2 = a.get("date_fin")
                d1_str = _iso_date(d1)
                d2_str = _iso_date(d2)
                if d2_str:
                    lignes_abs.append(
                        f"{lib} du {_fr_date_jjmm_aaaa(d1)} "
                        f"au {_fr_date_jjmm_aaaa(d2)}"
                    )
                else:
                    lignes_abs.append(
                        f"{lib} depuis le {_fr_date_jjmm_aaaa(d1)}"
                    )

            # Declarations de presence (absent)
            try:
                decl = rh.query(
                    """SELECT sdp.date_jour AS d,
                              ta.lib_absence
                         FROM pgt_salarie_decl_presence sdp
                         JOIN pgt_type_absence ta
                              ON ta.id_type_absence = sdp.motif_absence
                        WHERE sdp.id_salarie = ?
                          AND sdp.date_jour BETWEEN ? AND ?
                          AND sdp.presence = FALSE
                          AND sdp.motif_absence <> 6
                          AND (sdp.modif_elem IS NULL
                               OR sdp.modif_elem NOT LIKE '%suppr%')
                        ORDER BY sdp.date_jour ASC""",
                    (id_sal, date_deb, date_fin),
                ) or []
            except Exception:
                decl = []

            for d in decl:
                lib = (d.get("lib_absence") or "").strip()
                lignes_abs.append(
                    f"{_fr_date_jjmm_aaaa(d.get('d'))}, "
                    f"Déclaré(e) absent(e) : {lib}"
                )

            row.absences = "\n".join(lignes_abs)

            # Avances TK sur mois paiement
            try:
                avances = tk.query(
                    """SELECT SUM(montant) AS total
                         FROM pgt_tk_demande_avance
                        WHERE (modif_elem IS NULL
                               OR modif_elem NOT LIKE '%suppr%')
                          AND demande_validee = TRUE
                          AND date_paiement = ?
                          AND beneficiaire = ?""",
                    (date_deb, id_sal),
                ) or []
                if avances and avances[0]:
                    row.avance = float(avances[0].get("total") or 0)
            except Exception:
                pass

            vendeurs[id_sal] = row
            ordre.append(id_sal)

    # Tri final : Eq_Terrain, Poste, Nom (cf. WinDev)
    lignes = [vendeurs[i] for i in ordre]
    lignes.sort(key=lambda v: (v.eq_terrain, v.poste, v.nom, v.prenom))
    return RechercherResult(
        ok=True, lignes=lignes,
        message=f"{len(lignes)} salarie(s) trouve(s)",
    )


# --------------------------------------------------------------------
# Export XLSX (grouped by Eq_Terrain)
# --------------------------------------------------------------------

def _sanitize_feuille(s: str) -> str:
    """Nom feuille Excel : max 31 chars + retire caracteres interdits."""
    s = re.sub(r"[/\\?*\[\]:]+", " ", s or "")
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("INDEP", "").strip()
    return s[:31] or "Feuille1"


def generer_xlsx(p: ExportXlsxParams) -> tuple[str, bytes]:
    """Cf. WinDev Btn Export XLS : XLSX avec titre + groupement par Eq_Terrain."""
    try:
        from openpyxl import Workbook  # noqa: PLC0415
        from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415
    except ImportError:
        return ("", b"")

    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_feuille(p.lib_orga)

    font_title = Font(bold=True, size=20)
    font_header = Font(bold=True)
    fill_yellow = PatternFill(
        start_color="FFFACD", end_color="FFFACD", fill_type="solid",
    )
    align_multi = Alignment(vertical="center", wrap_text=True)

    # Titre grand
    ws.cell(row=1, column=1, value=p.lib_orga).font = font_title

    headers = [
        "Nom", "Prénom", "Poste", "Actif", "Sortie",
        "Date Entrée", "Motif Sortie", "Absences", "Avance",
    ]

    li = 4  # cf. WinDev : li += 3 puis += 1 pour equipe
    current_equipe = ""
    for row_v in p.lignes:
        if row_v.eq_terrain != current_equipe:
            # Change d'equipe : nouveau bloc jaune (equipe + entete)
            current_equipe = row_v.eq_terrain
            li += 3
            ws.cell(row=li, column=1, value=current_equipe).font = font_header
            ws.cell(row=li, column=1).fill = fill_yellow
            li += 1
            for i, h in enumerate(headers, 1):
                c = ws.cell(row=li, column=i, value=h)
                c.font = font_header
                c.fill = fill_yellow
                c.alignment = align_multi
                # Fond jaune sur la ligne au-dessus
                ws.cell(row=li - 1, column=i).fill = fill_yellow
            li += 1

        # Ligne de donnee
        ws.cell(row=li, column=1, value=row_v.nom).alignment = align_multi
        ws.cell(row=li, column=2, value=row_v.prenom).alignment = align_multi
        ws.cell(row=li, column=3, value=row_v.poste).alignment = align_multi
        ws.cell(row=li, column=4, value="X" if row_v.is_actif else "").alignment = align_multi
        ws.cell(row=li, column=5, value="X" if row_v.is_sortie else "").alignment = align_multi
        ws.cell(row=li, column=6, value=row_v.date_entree).alignment = align_multi
        ws.cell(row=li, column=7, value=row_v.type_sortie).alignment = align_multi
        ws.cell(row=li, column=8, value=row_v.absences).alignment = align_multi
        ws.cell(row=li, column=9, value=row_v.avance).alignment = align_multi
        li += 1

    # Largeurs colonnes
    widths = [22, 18, 24, 8, 8, 14, 22, 60, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    nom_feuille = _sanitize_feuille(p.lib_orga)
    lib_safe = unicodedata.normalize("NFKD", nom_feuille)
    lib_safe = "".join(c for c in lib_safe if not unicodedata.combining(c))
    lib_safe = re.sub(r"\s+", "_", lib_safe.strip()) or "salaries"
    fic_name = f"Tableau_salaries_{lib_safe}_{p.mois_paiement}.xlsx"
    return (fic_name, buf.getvalue())
