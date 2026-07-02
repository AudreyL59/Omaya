"""Service Fen_ImportSFR (ADM Imports Bases -> SFR/RED).

10 types d'import (cf combo TypeImport WinDev) :
  1. Base Journaliere Fibre   ImportJournalierFibre
  2. Base Journaliere Mobile  ImportJournalierMobile
  3. Base Journaliere CALL    ImportJournalierCALL
  4. Base Hebdo               ImportHebdo
  5. Import Options           ImportOptions
  6. RUN                      ImportRUN
  7. Call RET - KO            ImportCallRET_KO
  8. Call RET - Racc          ImportCallRET_Racc
  9. Call RET - Vente ADD     ImportCallRET_VentesADD
 10. Call RET - RDV Tech      ImportCallRET_RDVTech

Etat actuel : type 1 (Base Journ Fibre) detection only. Les 9 autres
types sont des squelettes a coder au fur et a mesure.

Ligne_départ : 2 pour types 3, 5, 7, 8, 9, 10 ; 3 pour les autres
(les fichiers SFR ont parfois 2 lignes d'entete).
"""

from __future__ import annotations

import base64
import io
from datetime import date, datetime
from typing import Optional

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection
from app.shared.recrutement.services.recherche_cv import _int, _str
from app.intranets.adm.services.import_eni import (
    _new_id, _lookup_or_create_client,
    _col_letter_to_index, _cell, _parse_date_fr, _dernier_jour_mois,
    _parse_int,
)


# Mapping colonnes Base Journaliere Mobile (cf groupe grpJournalier Mobile)
COLS_BJ_MOBILE = {
    "num_bs": "A",              "date_signature": "C",
    "date_activation": "F",     "date_portabilite": "H",
    "date_resil": "BM",         "lib_statut": "I",
    "type_vente": "AC",         "offre": "J",
    "code_vendeur": "X",        "num_mobile": "AT",
    "activ_control": "AP",      "processing_state": "AQ",
    "client_cp": "AJ",          "client_nom": "BQ",
    "client_prenom": "BR",      "client_gsm": "BS",
    "client_mail": "BN",
    "parcours_chaine": "BO",    "parcours_degroupe": "BF",
}


# Mapping colonnes Base Journaliere CALL (vendeur via nom, pas code)
COLS_BJ_CALL = {
    "num_bs": "A",              "vendeur": "K",
    "date_signature": "C",      "date_rdv": "G",
    "lib_statut": "I",          "offre": "J",
    "type_vente": "AC",         "technologie": "AF",
    "client_nom": "BQ",         "client_prenom": "BR",
    "client_adresse": "AI",     "client_cplt": "BG",
    "client_cp": "AJ",          "client_ville": "AL",
    "client_naiss": "BP",       "client_tel_mobile": "BS",
    "comment": "L",
}


# Mapping colonnes Base Hebdo (cf groupe grpHebdo). Pattern Fibre.
COLS_BJ_HEBDO = {
    "num_bs": "A",              "date_signature": "C",
    "date_va": "D",             "date_ra": "E",
    "date_rdv": "G",            "lib_statut": "I",
    "statut_vente": "J",        "motif_annul": "K",
    "type_install": "AM",       "type_vente": "AC",
    "offre": "J",               "technologie": "AF",
    "cluster_region": "AO",     "cluster_code": "AP",
    "cluster_ville": "AQ",
    "client_cp": "AJ",          "client_ville": "AL",
    "client_rue": "AI",         "client_identite": "BG",
    "client_tel": "BS",         "client_mail": "BN",
}


def _lookup_vendeur_by_nom_prenom(vendeur_cell: str) -> int:
    """Lookup salarie par concat nom+prenom (variantes avec %)."""
    if not vendeur_cell or not vendeur_cell.strip():
        return 0
    try:
        db = get_pg_connection("rh")
        s = vendeur_cell.upper()
        for c in ("-", "'", " "):
            s = s.replace(c, "%")
        s = s.replace("%%", "%")
        pattern = f"%{s}%"
        rows = db.query(
            """SELECT id_salarie FROM rh.pgt_salarie
                WHERE UPPER(CONCAT(nom, '%', prenom)) LIKE ?
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 2""",
            (pattern,),
        ) or []
        if len(rows) == 1:
            return int(rows[0]["id_salarie"])
    except Exception:
        pass
    return 0


# Mapping colonnes Base Journaliere Fibre (cf groupe grpJournalier Fibre)
COLS_BJ_FIBRE = {
    "num_bs": "A",                "date_signature": "C",
    "date_va": "D",               "date_ra": "E",
    "date_rdv": "G",              "date_rdv_actu": "H",
    "lib_statut": "I",            "statut_vente": "J",
    "motif_annul": "K",           "comment": "L",
    "instance": "M",              "cluster_ville": "AQ",
    "cluster_code": "AP",         "client_adresse": "AI",
    "client_cp": "AJ",            "client_ville": "AL",
    "type_install": "AM",         "type_vente": "AC",
    "code_offre": "AD",           "technologie": "AF",
    "box8": "AT",                 "internet_garantie": "AU",
    "portabilite": "BA",          "info_tech": "BG",
    "prise_existante": "BJ",      "num_prise": "BL",
    "date_resil": "BM",           "der_motif": "P",
    "parcours_chaine": "BO",      "parcours_degroupe": "BF",
    "prise_saisie": "BT",         "remise": "AR",
    "code_vendeur": "X",          "client_mail": "BN",
    "client_nom": "BQ",           "client_prenom": "BR",
    "client_gsm": "BS",
}


class ImportSfrParams(BaseModel):
    type_import: int                    # 1..10
    simulation: bool = True
    ligne_depart: int = 3               # 2 ou 3 selon type
    periode1_du: str = ""
    periode1_au: str = ""
    periode1_mois_paiement: str = ""
    periode2_du: str = ""
    periode2_au: str = ""
    periode2_mois_paiement: str = ""
    mois_paiement_distrib: str = ""


class ImportSfrResume(BaseModel):
    nb_fichiers: int = 0
    nb_ajoutes: int = 0
    nb_modifies: int = 0
    nb_modif_vend: int = 0
    nb_migrations: int = 0
    nb_non_modifies: int = 0
    nb_erreurs: int = 0
    nb_introuvables: int = 0
    nb_doublons: int = 0
    nb_hors_delai: int = 0


class ImportSfrResult(BaseModel):
    ok: bool
    type_import: int
    type_label: str
    simulation: bool
    resume: ImportSfrResume
    fichiers_traites: list[str] = []
    contrats_ajoutes: list[dict] = []
    contrats_modifies: list[dict] = []
    contrats_non_trouves: list[dict] = []
    contrats_migrations: list[dict] = []
    modif_vendeurs: list[dict] = []
    erreurs: list[dict] = []
    message: str = ""
    xlsx_b64: str = ""
    xlsx_name: str = ""
    mail_envoye: bool = False


TYPE_LABELS = {
    1: "Base Journalière Fibre",
    2: "Base Journalière Mobile",
    3: "Base Journalière CALL",
    4: "Base Hebdo",
    5: "Import Options",
    6: "RUN",
    7: "Call RET - KO",
    8: "Call RET - Racc",
    9: "Call RET - Vente ADD",
    10: "Call RET - RDV Tech",
}

# Helpers metier (transposition WinDev simplifiee)


def _type_vente_fibre(s: str) -> int:
    """typeVenteFibre : 1=Nouvelle, 2=Retention, 3=Migration THD, 4=Mig FTTH..."""
    u = (s or "").upper()
    if "MIG" in u or "MTX" in u:
        return 3
    if "RET" in u:
        return 2
    return 1


def _type_techno_fibre(s: str) -> int:
    """1=FTTH, 2=FTTB, 3=ADSL"""
    u = (s or "").upper()
    if "FTTH" in u:
        return 1
    if "FTTB" in u or "THD" in u:
        return 2
    if "ADSL" in u:
        return 3
    return 0


def _type_offre_fibre(code_offre: str, techno: int,
                      date_sign: Optional[date]) -> int:
    """typeOffreFibre : lookup pgt_sfr_produit by code/lib.
    Fallback : 0 si introuvable."""
    if not code_offre:
        return 0
    try:
        db = get_pg_connection("adv")
        r = db.query_one(
            """SELECT id_produit FROM adv.pgt_sfr_produit
                WHERE LOWER(lib_produit) LIKE LOWER(?)
                LIMIT 1""",
            (f"%{code_offre}%",),
        )
        return _int(r.get("id_produit")) if r else 0
    except Exception:
        return 0


def _id_etat_fibre(lib_statut: str) -> int:
    """donneIdEtatFibre : lookup pgt_sfr_etat_contrat by lib_etat LIKE."""
    if not lib_statut:
        return 0
    try:
        db = get_pg_connection("adv")
        r = db.query_one(
            """SELECT id_etat FROM adv.pgt_sfr_etat_contrat
                WHERE LOWER(lib_etat) LIKE LOWER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (f"{lib_statut}%",),
        )
        return _int(r.get("id_etat")) if r else 0
    except Exception:
        return 0


def _test_cluster_fibre(code_vad: str) -> dict:
    """testClusterFibre : retourne {id_sfr_cluster, hors_cible}."""
    if not code_vad:
        return {"id_sfr_cluster": 0, "hors_cible": True}
    try:
        db = get_pg_connection("adv")
        r = db.query_one(
            """SELECT id_sfr_cluster, hors_cible FROM adv.pgt_sfr_cluster
                WHERE code_vad = ? LIMIT 1""",
            (code_vad,),
        )
        return {"id_sfr_cluster": _int(r.get("id_sfr_cluster")) if r else 0,
                "hors_cible": bool(r.get("hors_cible")) if r else True}
    except Exception:
        return {"id_sfr_cluster": 0, "hors_cible": True}


def _lookup_tk_call_sfr(num_bs: str) -> dict:
    """ReqTkCallSFR_ByNumCtt : lookup pgt_tk_call_sfr by num_bs."""
    if not num_bs:
        return {}
    try:
        db = get_pg_connection("ticket_bo")
        r = db.query_one(
            """SELECT id_salarie, nom_client, nom_marital_client, prenom_client,
                      datenaiss, adresse1, adresse2, cp, ville, mobile1, mobile2,
                      adr_mail, opt_partenaire, num_prise_optique, id_tk_liste
                 FROM ticket_bo.pgt_tk_call_sfr
                WHERE UPPER(num_bs) = UPPER(?) LIMIT 1""",
            (num_bs,),
        )
        return r or {}
    except Exception:
        return {}


def _import_journalier_fibre(
    p: ImportSfrParams, fname: str, content: bytes, op_id: int,
    ajoutes: list, modifies: list, migrations: list,
    modif_vendeurs: list, erreurs: list, resume: ImportSfrResume,
) -> None:
    """Procedure Base Journaliere Fibre (type 1)."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        erreurs.append({"_erreur": f"Lecture {fname} : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_BJ_FIBRE.items()}
    db = get_pg_connection("adv")

    ligne_depart = p.ligne_depart or 3
    for i in range(ligne_depart, (ws.max_row or 0) + 1):
        num_bs = _cell(ws, i, cols["num_bs"]).upper()
        if not num_bs:
            continue

        date_sign = _parse_date_fr(_cell(ws, i, cols["date_signature"]))
        date_va = _parse_date_fr(_cell(ws, i, cols["date_va"]))
        date_ra = _parse_date_fr(_cell(ws, i, cols["date_ra"]))
        date_rdv = _parse_date_fr(_cell(ws, i, cols["date_rdv"]))
        date_rdv_actu = _parse_date_fr(_cell(ws, i, cols["date_rdv_actu"]))
        if date_rdv_actu:
            date_rdv = date_rdv_actu
        type_v_s = _cell(ws, i, cols["type_vente"])
        type_vente = _type_vente_fibre(type_v_s)
        techno = _type_techno_fibre(_cell(ws, i, cols["technologie"]))
        code_offre = _cell(ws, i, cols["code_offre"])
        offre = _type_offre_fibre(code_offre, techno, date_sign)
        lib_statut = _cell(ws, i, cols["lib_statut"])
        instance = _cell(ws, i, cols["instance"]) if "instance" in cols else ""
        instance = instance.replace(" ", "").upper()
        der_motif = _cell(ws, i, cols["der_motif"]) if "der_motif" in cols else ""
        id_etat = _id_etat_fibre(lib_statut)
        # Ajustements motif d'annulation (cf. WinDev) :
        # - id=3 + instance=S1 -> 87 (Client Absent S1)
        # - id in (10, 36) + DerMotif : 14/15/16/86 selon motif
        if id_etat == 3 and instance == "S1":
            id_etat = 87
        if (id_etat == 10 and der_motif != "Nominal") or id_etat == 36:
            dm_up = der_motif.upper()
            if "FIN DE" in dm_up:
                id_etat = 14
            elif "RENON" in dm_up:
                id_etat = 15
            elif "RETRAC" in dm_up:
                id_etat = 16
            elif id_etat != 10:
                id_etat = 86  # Motif incohérent
        cluster_code = _cell(ws, i, cols["cluster_code"]).replace("'", "")
        cluster_ville = _cell(ws, i, cols["cluster_ville"])
        client_cp = _cell(ws, i, cols["client_cp"])
        client_ville = _cell(ws, i, cols["client_ville"])
        client_adr = _cell(ws, i, cols["client_adresse"])
        client_mail = _cell(ws, i, cols["client_mail"])
        client_nom = _cell(ws, i, cols["client_nom"])
        client_prenom = _cell(ws, i, cols["client_prenom"])
        client_gsm = _cell(ws, i, cols["client_gsm"])
        if client_mail == "0":
            client_mail = ""
        if client_gsm and not client_gsm.startswith("0"):
            client_gsm = "0" + client_gsm

        box8 = "oui" in _cell(ws, i, cols["box8"]).lower()
        portabilite = "oui" in _cell(ws, i, cols["portabilite"]).lower()
        prise_existante = "oui" in _cell(ws, i, cols["prise_existante"]).lower()
        prise_saisie = "oui" in _cell(ws, i, cols["prise_saisie"]).lower()
        internet_garantie = "oui" in _cell(ws, i, cols["internet_garantie"]).lower()
        remise = bool(_cell(ws, i, cols["remise"]).strip())
        code_vendeur = _cell(ws, i, cols["code_vendeur"])
        comment = _cell(ws, i, cols["comment"])
        info_tech = _cell(ws, i, cols["info_tech"])
        if comment == "0":
            comment = ""
        if info_tech and info_tech != "0":
            comment += "\n" + info_tech
        motif_annul = _cell(ws, i, cols["motif_annul"])
        if motif_annul == "0":
            motif_annul = ""
        num_prise = _cell(ws, i, cols["num_prise"])
        date_resil = _parse_date_fr(_cell(ws, i, cols["date_resil"]))

        # Detection migration FTTB -> FTTH
        is_mig = ("MTX-THD" in code_offre.upper()
                  and "MIG" in type_v_s.upper())
        if is_mig and type_vente == 3:
            type_vente = 4
            migrations.append({
                "NumBS": num_bs, "DateSign": str(date_sign or ""),
                "ClientCP": client_cp, "ClientVille": client_ville,
                "LibStatut": lib_statut, "Box8": box8,
                "ClusterCode": cluster_code,
            })
            resume.nb_migrations += 1

        cluster = _test_cluster_fibre(cluster_code)

        # Lookup contrat existant
        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, id_client, id_produit,
                      id_etat_contrat, id_etat_sfr, type_vente, date_signature,
                      num_prise_vend, id_sfr_cluster, box8, remise,
                      offre_speciale, internet_garanti, prise_existante,
                      prise_saisie, num_prise_sfr, processing_state,
                      parcours_chaine, parcours_degroupe, date_resil,
                      motif_annulation
                 FROM adv.pgt_sfr_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )

        # Detection offre speciale (SCD dans code_offre)
        offre_speciale = "SCD" in code_offre.upper()

        if not ctt:
            tk = _lookup_tk_call_sfr(num_bs)
            id_salarie = int(tk.get("id_salarie") or 20200715153948361)
            ajoutes.append({
                "NumBS": num_bs,
                "DateSign": str(date_sign or ""),
                "DateVa": str(date_va or ""), "DateRa": str(date_ra or ""),
                "DateRDV": str(date_rdv or ""),
                "ClientCP": client_cp, "ClientVille": client_ville,
                "LibStatut": lib_statut, "IdProduit": offre,
                "TypeVente": type_vente, "Box8": box8,
                "ClusterCode": cluster_code,
                "_payload_create": {
                    "num_bs": num_bs, "id_salarie": id_salarie,
                    "id_produit": offre, "id_etat_contrat": id_etat,
                    "id_etat_sfr": id_etat, "type_vente": type_vente,
                    "technologie": techno, "box8": box8,
                    "portabilite": portabilite, "self_install": False,
                    "id_sfr_cluster": cluster["id_sfr_cluster"],
                    "date_signature": date_sign,
                    "date_validation": date_va,
                    "date_racc_activ": date_ra,
                    "date_rdv_tech": date_rdv,
                    "date_resil": date_resil,
                    "motif_annulation": motif_annul,
                    "info_vente_sfr": comment,
                    "info_interne": code_vendeur,
                    "internet_garanti": internet_garantie,
                    "num_prise_sfr": num_prise,
                    "prise_existante": prise_existante,
                    "prise_saisie": prise_saisie,
                    "remise": remise,
                    "_client": {
                        "nom": client_nom, "prenom": client_prenom,
                        "adresse": client_adr, "cp": client_cp,
                        "ville": client_ville, "mail": client_mail,
                        "gsm": client_gsm,
                    },
                },
            })
            resume.nb_ajoutes += 1
        else:
            # ---- CONTRAT EXISTANT : detection modifs + erreurs ----
            id_contrat = int(ctt["id_contrat"])
            id_sal_db = int(ctt.get("id_salarie") or 0)
            etat_actuel = int(ctt.get("id_etat_contrat") or 0)
            modifs = []
            errs_detected = []  # Erreurs rapportables (feuille 3)
            updates: dict = {}  # Champs a UPDATE en PROD

            # -- Reattribution vendeur si Fibre inconnu --
            if id_sal_db in (0, 20200715153948361):
                tk = _lookup_tk_call_sfr(num_bs)
                tk_sal = int(tk.get("id_salarie") or 0)
                if tk_sal and tk_sal != id_sal_db:
                    modifs.append(f"Vendeur -> {tk_sal}")
                    modif_vendeurs.append({
                        "NumBS": num_bs, "OldIdSalarie": id_sal_db,
                        "NewIdSalarie": tk_sal,
                    })
                    updates["id_salarie"] = tk_sal
                    resume.nb_modif_vend += 1

            # -- Detection erreurs / MAJ champs (cf. WinDev testErr) --
            db_type_vente = int(ctt.get("type_vente") or 0)
            if db_type_vente != type_vente:
                # WinDev : si les 2 sont < 3 on ignore
                if not (db_type_vente < 3 and type_vente < 3):
                    errs_detected.append(("Type Vente", db_type_vente, type_vente))
                    updates["type_vente"] = type_vente
            if offre and int(ctt.get("id_produit") or 0) != offre:
                errs_detected.append(("Offre", ctt.get("id_produit"), offre))
                updates["id_produit"] = offre
            if bool(ctt.get("box8")) != box8:
                errs_detected.append(("Box 8", bool(ctt.get("box8")), box8))
                updates["box8"] = box8
                updates["box8_verif"] = box8
            if bool(ctt.get("remise")) != remise:
                errs_detected.append(("Remise", bool(ctt.get("remise")), remise))
                updates["remise"] = remise
            if bool(ctt.get("offre_speciale")) != offre_speciale:
                errs_detected.append(("Offre spéciale",
                                       bool(ctt.get("offre_speciale")), offre_speciale))
                updates["offre_speciale"] = offre_speciale
            if bool(ctt.get("internet_garanti")) != internet_garantie:
                errs_detected.append(("Internet Garanti",
                                       bool(ctt.get("internet_garanti")), internet_garantie))
                updates["internet_garanti"] = internet_garantie
            # -- Champs silent (pas rapportes dans erreurs) --
            if bool(ctt.get("prise_existante")) != prise_existante:
                updates["prise_existante"] = prise_existante
            if bool(ctt.get("prise_saisie")) != prise_saisie:
                updates["prise_saisie"] = prise_saisie
            if (ctt.get("num_prise_sfr") or "") != num_prise:
                updates["num_prise_sfr"] = num_prise
            if (ctt.get("processing_state") or "") != lib_statut:
                updates["processing_state"] = lib_statut
            if bool(ctt.get("parcours_chaine")) != bool(_cell(ws, i, cols["parcours_chaine"]).strip()):
                updates["parcours_chaine"] = bool(_cell(ws, i, cols["parcours_chaine"]).strip())
            if bool(ctt.get("parcours_degroupe")) != bool(_cell(ws, i, cols["parcours_degroupe"]).strip()):
                updates["parcours_degroupe"] = bool(_cell(ws, i, cols["parcours_degroupe"]).strip())
            if date_resil and ctt.get("date_resil") != date_resil:
                updates["date_resil"] = date_resil
            if motif_annul and motif_annul.upper() not in (ctt.get("motif_annulation") or "").upper():
                new_ma = (ctt.get("motif_annulation") or "") + motif_annul + "\n"
                updates["motif_annulation"] = new_ma

            # -- Traitement changement d'etat --
            etat_change_status = "no_change"  # 'maj', 'no_change', 'non_modifie', 'paye_ko'
            type_etat_old = 0
            if etat_actuel and id_etat and etat_actuel != id_etat:
                etat_info = db.query_one(
                    """SELECT id_type_etat FROM adv.pgt_sfr_etat_contrat
                        WHERE id_etat = ? LIMIT 1""",
                    (etat_actuel,),
                )
                type_etat_old = int((etat_info or {}).get("id_type_etat") or 0)
                # WinDev :
                # - typeOld <> 5 et <> 6 :
                #   - typeOld <=2 ou =7 ou =8 -> MAJ etat + histo
                #   - sinon -> non modifie
                # - sinon (payé/décomm) : non_modifie + rapport special
                #   si typeStatut=Raccordement KO et typeOld=5
                if type_etat_old not in (5, 6):
                    if type_etat_old <= 2 or type_etat_old in (7, 8):
                        etat_change_status = "maj"
                        updates["id_etat_contrat"] = id_etat
                        updates["date_validation"] = date_va
                        updates["date_racc_activ"] = date_ra
                        updates["date_rdv_tech"] = date_rdv
                        updates["date_resil"] = date_resil
                        if comment:
                            new_ii = (ctt.get("info_interne") or "") + "\n" + comment
                            updates["info_interne"] = new_ii
                        updates["motif_annulation"] = motif_annul
                        modifs.append(f"Etat -> {id_etat}")
                        resume.nb_modifies += 1
                    else:
                        etat_change_status = "non_modifie"
                        resume.nb_non_modifies += 1
                else:
                    # Paye/decomm : verifier Raccordement KO
                    if "Raccordement KO" in lib_statut and type_etat_old == 5:
                        errs_detected.append(("Payé passé en KO",
                                               f"etat {etat_actuel}",
                                               "Racc KO"))
                        etat_change_status = "paye_ko"
                    else:
                        etat_change_status = "non_modifie"
                        resume.nb_non_modifies += 1
            elif errs_detected or modifs:
                resume.nb_modifies += 1
            else:
                resume.nb_non_modifies += 1

            # -- Rapport erreurs --
            for lib_err, val_av, val_ap in errs_detected:
                erreurs.append({
                    "NumBS": num_bs,
                    "TypeErreur": lib_err,
                    "AvantImport": str(val_av),
                    "ApresImport": str(val_ap),
                    "DateSign": str(ctt.get("date_signature") or ""),
                    "TypeStatut": lib_statut,
                })
                resume.nb_erreurs += 1

            if modifs:
                modifies.append({
                    "NumBS": num_bs,
                    "DateSign OMAYA": str(ctt.get("date_signature") or ""),
                    "Modifs": " | ".join(modifs),
                    "EtatChangeStatus": etat_change_status,
                })

            # -- Recalcul nb_points si date sign ou racc >= 2022-02-01 --
            if (date_sign and date_sign >= date(2022, 2, 1)) or \
               (date_ra and date_ra >= date(2022, 2, 1)):
                updates["_recalc_nb_points"] = True

            # -- PASSE PROD : UPDATE effectif --
            if not p.simulation and updates:
                do_recalc = updates.pop("_recalc_nb_points", False)
                sets = []
                params = []
                for k, v in updates.items():
                    sets.append(f"{k} = ?")
                    params.append(v)
                if sets:
                    sets.append("modif_date = NOW()")
                    sets.append("modif_op = ?")
                    sets.append("modif_elem = 'modif'")
                    params.append(int(op_id))
                    params.append(id_contrat)
                    try:
                        db.query(
                            f"UPDATE adv.pgt_sfr_contrat SET {', '.join(sets)} "
                            f"WHERE id_contrat = ?",
                            tuple(params),
                        )
                    except Exception as e:
                        erreurs.append({"NumBS": num_bs,
                                          "TypeErreur": "UPDATE",
                                          "AvantImport": str(e)})
                # Historisation si etat a change
                if etat_change_status == "maj":
                    try:
                        _ajoute_histo_sfr_etat(
                            id_contrat, etat_actuel, id_etat, "", op_id,
                        )
                    except NameError:
                        pass  # fonction non definie encore (a implementer)

    wb.close()


def _import_journalier_mobile(
    p: ImportSfrParams, content: bytes, op_id: int,
    ajoutes: list, modifies: list, erreurs: list, resume: ImportSfrResume,
) -> None:
    """Type 2 : Base Journaliere Mobile. Pattern Fibre sans cluster.
    Compare DatePort, DateResil, DateAct, ActivControl, ProcessingState."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        erreurs.append({"_erreur": f"Lecture : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_BJ_MOBILE.items()}
    db = get_pg_connection("adv")
    ligne_depart = p.ligne_depart or 3

    for i in range(ligne_depart, (ws.max_row or 0) + 1):
        num_bs = _cell(ws, i, cols["num_bs"]).upper()
        if not num_bs:
            continue
        date_sign = _parse_date_fr(_cell(ws, i, cols["date_signature"]))
        date_act = _parse_date_fr(_cell(ws, i, cols["date_activation"]))
        date_port = _parse_date_fr(_cell(ws, i, cols["date_portabilite"]))
        date_resil = _parse_date_fr(_cell(ws, i, cols["date_resil"]))
        type_vente = _type_vente_fibre(_cell(ws, i, cols["type_vente"]))
        lib_offre = _cell(ws, i, cols["offre"])
        lib_statut = _cell(ws, i, cols["lib_statut"])
        id_etat = _id_etat_fibre(lib_statut)
        num_mobile = _cell(ws, i, cols["num_mobile"])
        activ_control = _cell(ws, i, cols["activ_control"])
        processing_state = _cell(ws, i, cols["processing_state"])
        client_cp = _cell(ws, i, cols["client_cp"])
        client_nom = _cell(ws, i, cols["client_nom"])
        client_prenom = _cell(ws, i, cols["client_prenom"])
        client_gsm = _cell(ws, i, cols["client_gsm"])
        if client_gsm and not client_gsm.startswith("0"):
            client_gsm = "0" + client_gsm
        client_mail = _cell(ws, i, cols["client_mail"])
        code_vendeur = _cell(ws, i, cols["code_vendeur"])

        # Lookup produit Mobile (sous_fam=MOBILE par defaut)
        offre = _type_offre_fibre(lib_offre, 0, date_sign)

        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, id_client, id_produit,
                      type_vente, date_signature, date_portabilite,
                      date_resil, date_racc_activ, activ_control,
                      processing_state, id_etat_sfr, id_etat_contrat,
                      nb_points, info_interne
                 FROM adv.pgt_sfr_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )
        if not ctt:
            ajoutes.append({
                "NumBS": num_bs, "DateSign": str(date_sign or ""),
                "DateAct": str(date_act or ""),
                "DatePort": str(date_port or ""),
                "DateResil": str(date_resil or ""),
                "ClientCP": client_cp,
                "Client": f"{client_nom} {client_prenom}".strip(),
                "LibOffre": lib_offre, "TypeVente": type_vente,
                "NumMobile": num_mobile, "LibStatut": lib_statut,
            })
            resume.nb_ajoutes += 1
        else:
            id_contrat = int(ctt["id_contrat"])
            id_etat_sfr_actuel = int(ctt.get("id_etat_sfr") or 0)
            etat_ctt_actuel = int(ctt.get("id_etat_contrat") or 0)
            modifs = []
            errs = []
            if int(ctt.get("type_vente") or 0) != type_vente:
                errs.append(("Type Vente", ctt.get("type_vente"), type_vente))
                modifs.append(f"TypeVente -> {type_vente}")
            if int(ctt.get("id_produit") or 0) != offre and offre:
                modifs.append(f"Offre -> {offre}")
            if ctt.get("date_portabilite") != date_port and date_port:
                modifs.append(f"DatePort -> {date_port}")
            if ctt.get("date_resil") != date_resil and date_resil:
                modifs.append(f"DateResil -> {date_resil}")
            if ctt.get("date_racc_activ") != date_act and date_act:
                modifs.append(f"DateAct -> {date_act}")
            if (_str(ctt.get("activ_control")) != activ_control
                    and activ_control):
                modifs.append(f"ActivControl -> {activ_control}")
            if (_str(ctt.get("processing_state")) != processing_state
                    and processing_state):
                modifs.append(f"ProcessingState -> {processing_state}")

            # MAJ GSM client si num_mobile diff (cf. WinDev)
            id_client_db = int(ctt.get("id_client") or 0)
            if id_client_db and num_mobile:
                cl = db.query_one(
                    "SELECT gsm FROM adv.pgt_client WHERE id_client = ? LIMIT 1",
                    (id_client_db,),
                )
                if cl and (cl.get("gsm") or "") != num_mobile:
                    modifs.append(f"Client GSM -> {num_mobile}")
                    if not p.simulation:
                        try:
                            db.query(
                                """UPDATE adv.pgt_client
                                      SET gsm = ?, modif_date = NOW(), modif_op = ?
                                    WHERE id_client = ?""",
                                (num_mobile, int(op_id), id_client_db),
                            )
                        except Exception:
                            pass

            # Changement etat SFR : IDetatSFR <> id_etat and id_etat<>0 and
            # IDetatSFR<>76 -> histo "SFR" (cf. WinDev)
            new_etat_sfr = None
            if (id_etat and id_etat_sfr_actuel != id_etat
                    and id_etat_sfr_actuel != 76):
                new_etat_sfr = id_etat
                modifs.append(f"EtatSFR -> {id_etat}")

            # Changement etat contrat : logique WinDev
            # typeOld <=2 ou 7/8 -> MAJ + histo "Vend"
            # typeOld = 5 -> rapport "Payé passé en KO/Résiliation"
            new_etat_ctt = None
            if id_etat and etat_ctt_actuel != id_etat:
                etat_info = db.query_one(
                    """SELECT id_type_etat, lib_etat FROM adv.pgt_sfr_etat_contrat
                        WHERE id_etat = ? LIMIT 1""",
                    (etat_ctt_actuel,),
                )
                type_etat_old = int((etat_info or {}).get("id_type_etat") or 0)
                if type_etat_old not in (5, 6):
                    if type_etat_old <= 2 or type_etat_old in (7, 8):
                        new_etat_ctt = id_etat
                        modifs.append(f"EtatContrat -> {id_etat}")
                else:
                    # Paye/decomm : rapport special
                    lib_stat_up = lib_statut.upper()
                    if type_etat_old == 5 and (
                            "RESIL" in lib_stat_up or "KO" in lib_stat_up):
                        errs.append((f"Payé passé en {lib_statut}",
                                     etat_ctt_actuel, id_etat))

            # InfoInterne : append code_vendeur si absent
            new_info_interne = None
            if code_vendeur and code_vendeur.upper() not in (
                    ctt.get("info_interne") or "").upper():
                new_info_interne = ((ctt.get("info_interne") or "")
                                    + "\n" + code_vendeur).strip()
                modifs.append("InfoInterne (code vendeur)")

            # Recalcul nb_points fam=MOBILE si date_sign ou date_ra >= 2022-02-01
            recalc = ((date_sign and date_sign >= date(2022, 2, 1))
                      or (date_act and date_act >= date(2022, 2, 1)))
            nbpt_new = None
            if recalc and offre:
                try:
                    from app.intranets.adm.services.sfr_helpers import (
                        _donne_fam_prod_sfr,
                    )
                    from app.shared.sdtc.bareme import calcul_point_contrat
                    prod = db.query_one(
                        "SELECT sous_fam FROM adv.pgt_sfr_produit WHERE id_produit = ? LIMIT 1",
                        (offre,),
                    ) or {}
                    fam = _donne_fam_prod_sfr("MOBILE", type_vente)
                    nbpt = calcul_point_contrat(
                        fam=fam, ss_fam=prod.get("sous_fam") or "",
                        palier=0,
                        date_sign=str(date_sign) if date_sign else "",
                        info_cplt="", palier2=0,
                    )
                    if float(ctt.get("nb_points") or 0) != nbpt:
                        nbpt_new = float(nbpt)
                        modifs.append(f"nb_points -> {nbpt}")
                except Exception:
                    pass

            for lib_err, val_av, val_ap in errs:
                erreurs.append({
                    "NumBS": num_bs,
                    "TypeErreur": lib_err,
                    "InfoAvant": str(val_av),
                    "InfoApres": str(val_ap),
                    "DateSign": str(ctt.get("date_signature") or ""),
                    "NumMobile": num_mobile,
                })
                resume.nb_erreurs += 1

            if modifs:
                modifies.append({
                    "NumBS": num_bs, "Modifs": " | ".join(modifs),
                    "Client": f"{client_nom} {client_prenom}".strip(),
                    "NumMobile": num_mobile,
                    "EtatSFR ap": id_etat_sfr_actuel if new_etat_sfr is None else new_etat_sfr,
                    "EtatContrat ap": etat_ctt_actuel if new_etat_ctt is None else new_etat_ctt,
                })
                resume.nb_modifies += 1
                if not p.simulation:
                    try:
                        sets = ["type_vente = ?", "id_produit = ?",
                                "date_portabilite = COALESCE(?, date_portabilite)",
                                "date_resil = COALESCE(?, date_resil)",
                                "date_racc_activ = COALESCE(?, date_racc_activ)",
                                "activ_control = ?", "processing_state = ?"]
                        params: list = [type_vente, offre, date_port,
                                         date_resil, date_act,
                                         activ_control, processing_state]
                        if new_etat_sfr is not None:
                            sets.append("id_etat_sfr = ?")
                            params.append(new_etat_sfr)
                        if new_etat_ctt is not None:
                            sets.append("id_etat_contrat = ?")
                            params.append(new_etat_ctt)
                        if new_info_interne is not None:
                            sets.append("info_interne = ?")
                            params.append(new_info_interne)
                        if nbpt_new is not None:
                            sets.append("nb_points = ?")
                            params.append(nbpt_new)
                        sets.append("modif_date = NOW()")
                        sets.append("modif_op = ?")
                        sets.append("modif_elem = 'modif'")
                        params.append(int(op_id))
                        params.append(id_contrat)
                        db.query(
                            f"UPDATE adv.pgt_sfr_contrat SET {', '.join(sets)} "
                            f"WHERE id_contrat = ?",
                            tuple(params),
                        )
                        # Historisation
                        if new_etat_sfr is not None:
                            _ajoute_histo_sfr_etat(
                                id_contrat, id_etat_sfr_actuel, new_etat_sfr,
                                "", op_id, categorie="SFR",
                            )
                        if new_etat_ctt is not None:
                            _ajoute_histo_sfr_etat(
                                id_contrat, etat_ctt_actuel, new_etat_ctt,
                                "", op_id, categorie="Vend",
                            )
                    except Exception as e:
                        modifies[-1]["Erreur"] = str(e)
            else:
                resume.nb_non_modifies += 1
    wb.close()


def _import_journalier_call(
    p: ImportSfrParams, content: bytes, op_id: int,
    ajoutes: list, modifies: list, erreurs: list, resume: ImportSfrResume,
) -> None:
    """Type 3 : Base Journaliere CALL (FIBRE). Vendeur via nom complet.
    Pattern : si pas trouve -> erreur 'VENDEUR inconnu'.
    Si contrat existant + vendeur Fibre inconnu -> reattribution.
    Si vendeur DB different -> erreur 'Vendeur différent'."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        erreurs.append({"_erreur": f"Lecture : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_BJ_CALL.items()}
    db = get_pg_connection("adv")
    ligne_depart = p.ligne_depart or 2

    for i in range(ligne_depart, (ws.max_row or 0) + 1):
        num_bs = _cell(ws, i, cols["num_bs"]).upper()
        if not num_bs:
            continue
        date_sign = _parse_date_fr(_cell(ws, i, cols["date_signature"]))
        # Date RDV : prendre les 8 premiers chars JJMMAAAA
        date_rdv_s = _cell(ws, i, cols["date_rdv"])[:8]
        date_rdv = None
        if len(date_rdv_s) == 8 and date_rdv_s.isdigit():
            try:
                date_rdv = date(int(date_rdv_s[4:8]), int(date_rdv_s[2:4]),
                                int(date_rdv_s[0:2]))
            except Exception:
                pass
        vendeur_cell = _cell(ws, i, cols["vendeur"])
        client_nom = _cell(ws, i, cols["client_nom"])
        client_prenom = _cell(ws, i, cols["client_prenom"])
        client_adr = _cell(ws, i, cols["client_adresse"])
        client_cp = _cell(ws, i, cols["client_cp"])
        client_ville = _cell(ws, i, cols["client_ville"])
        client_mobile = _cell(ws, i, cols["client_tel_mobile"])
        nom_offre = _cell(ws, i, cols["offre"])
        type_vente_s = _cell(ws, i, cols["type_vente"])
        comment = _cell(ws, i, cols["comment"])
        type_vente = _type_vente_fibre(type_vente_s)
        techno = _type_techno_fibre(_cell(ws, i, cols["technologie"]))
        offre = _type_offre_fibre(nom_offre, techno, date_sign)
        box8 = "8" in nom_offre

        id_vendeur = _lookup_vendeur_by_nom_prenom(vendeur_cell)
        if id_vendeur == 0:
            erreurs.append({
                "NumBS": num_bs, "Erreur": "VENDEUR inconnu",
                "VendeurCell": vendeur_cell,
                "DateSign": str(date_sign or ""),
                "NomOffre": nom_offre, "TypeVente": type_vente_s,
            })
            resume.nb_erreurs += 1
            # On continue quand meme avec sentinelle FIBRE_INCONNU
            id_vendeur = FIBRE_INCONNU

        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, id_client, id_produit,
                      non_call, type_vente, date_signature
                 FROM adv.pgt_sfr_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )
        if not ctt:
            # ---- BRANCHE AJOUT (cf. WinDev l.274-306) ----
            ajoutes.append({
                "NumBS": num_bs, "Vendeur": vendeur_cell,
                "DateSign": str(date_sign or ""),
                "DateRDV": str(date_rdv or ""),
                "Client": f"{client_nom} {client_prenom}".strip(),
                "ClientCP": client_cp, "ClientVille": client_ville,
                "NomOffre": nom_offre, "TypeVente": type_vente_s,
                "Comment": comment,
            })
            resume.nb_ajoutes += 1
            if not p.simulation:
                try:
                    # 1. Cree/trouve le client via traiter_client
                    from app.intranets.adm.services.import_helpers_common import (
                        traiter_client,
                    )
                    id_client = traiter_client(
                        info_client={
                            "nom": client_nom, "prenom": client_prenom,
                            "adresse1": client_adr, "cp": client_cp,
                            "ville": client_ville, "gsm": client_mobile,
                            "op_saisie": op_id, "modif_op": op_id,
                        }, force_maj=False, op_id=op_id,
                    ) or 0
                    # 2. INSERT contrat via ajout_fiche_contrat_sfr
                    # (cf. WinDev ajoutFicheContrat(monCtt))
                    from app.intranets.adm.services.sfr_helpers import (
                        ajout_fiche_contrat_sfr,
                    )
                    ctt_data = {
                        "id_client": id_client,
                        "id_salarie": id_vendeur,
                        "num_bs": num_bs,
                        "date_signature": date_sign,
                        "date_rdv_tech": date_rdv,
                        "id_etat_sfr": 1, "id_etat_contrat": 1,
                        "id_produit": offre,
                        "technologie": techno,
                        "type_vente": type_vente,
                        "box8": box8, "box8_verif": box8,
                        "hors_cible": False, "option_dec": False,
                        "option_verif": False,
                        "motif_annulation": "",
                        "info_interne": comment,
                        "non_call": True, "remise": False,
                        "issu_tk_diff": 0,
                    }
                    id_ctt_new = ajout_fiche_contrat_sfr(ctt_data, op_id)
                    ajoutes[-1]["IdContratCree"] = id_ctt_new
                    # 3. Historisation etat initial 0 -> 1 (cf. WinDev)
                    try:
                        _ajoute_histo_sfr_etat(
                            id_ctt_new, 0, 1, "", op_id, categorie="Vend",
                        )
                    except Exception:
                        pass
                except Exception as e:
                    ajoutes[-1]["Erreur"] = str(e)
        else:
            # ---- BRANCHE MODIF (cf. WinDev l.308-402) ----
            id_contrat = int(ctt["id_contrat"])
            id_sal_db = int(ctt.get("id_salarie") or 0)
            id_client_db = int(ctt.get("id_client") or 0)
            non_call_db = bool(ctt.get("non_call"))
            modifies.append({
                "NumBS": num_bs, "Vendeur": vendeur_cell,
                "DateSign": str(date_sign or ""),
                "DateRDV": str(date_rdv or ""),
                "Client": f"{client_nom} {client_prenom}".strip(),
                "ClientCP": client_cp, "NomOffre": nom_offre,
                "Comment": comment,
            })
            resume.nb_modifies += 1

            # 1. MAJ client existant (cf. WinDev l.336-349)
            if not p.simulation and id_client_db:
                try:
                    from app.intranets.adm.services.sfr_helpers import (
                        modif_fiche_client_sfr,
                    )
                    modif_fiche_client_sfr(
                        id_client_db,
                        nom=client_nom, prenom=client_prenom,
                        date_naiss=None,
                        adresse1=client_adr, adresse2="",
                        cp=client_cp, ville=client_ville,
                        tel="", gsm=client_mobile, mail="",
                        op_id=op_id,
                    )
                except Exception:
                    pass

            # 2. Repasse NonCALL a FALSE si etait TRUE (cf. WinDev l.351-360)
            if non_call_db and not p.simulation:
                try:
                    db.query(
                        """UPDATE adv.pgt_sfr_contrat
                              SET non_call = FALSE,
                                  modif_date = NOW(), modif_op = ?,
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (int(op_id), id_contrat),
                    )
                    modifies[-1]["Note"] = "NonCALL -> False"
                except Exception:
                    pass

            # 3. Reattribution si fibre inconnu et nouveau vendeur trouve
            if id_sal_db == FIBRE_INCONNU and id_vendeur != FIBRE_INCONNU:
                modifies[-1]["Note"] = "Contrat Fibre Inconnu réattribué"
                if not p.simulation:
                    try:
                        db.query(
                            """UPDATE adv.pgt_sfr_contrat
                                  SET id_salarie = ?, non_call = FALSE,
                                      modif_date = NOW(), modif_op = ?,
                                      modif_elem = 'modif'
                                WHERE id_contrat = ?""",
                            (id_vendeur, int(op_id), id_contrat),
                        )
                    except Exception as e:
                        modifies[-1]["ErreurReattrib"] = str(e)
            elif id_sal_db != id_vendeur and id_vendeur != FIBRE_INCONNU:
                # Vendeur different -> pas de reattribution, juste rapport
                erreurs.append({
                    "NumBS": num_bs, "Erreur": "Vendeur différent",
                    "VendeurOmaya": id_sal_db, "VendeurImport": vendeur_cell,
                    "DateSign": str(date_sign or ""),
                })
    wb.close()


def _import_journalier_hebdo(
    p: ImportSfrParams, content: bytes, op_id: int,
    ajoutes: list, modifies: list, migrations: list,
    erreurs: list, resume: ImportSfrResume,
) -> None:
    """Type 4 : Base Hebdo. Pattern Fibre (T1) adapte :
    - Dates lues avec [:10] (formats variables)
    - Detection 'FTTB VERS FTTH' pour migrations
    - client_identite splittee par ',' (nom, prenom)

    Note WinDev : la variable dateref = '20201001' est declaree ligne
    197 mais JAMAIS utilisee dans le code effectif (une seule
    occurrence, a la declaration). Le commentaire WinDev dit "on
    n'importe pas les contrats anterieures au 01/10/2020" mais le
    filtrage n'est jamais applique.
    Ancienne implementation Python appliquait ce filtre a tort ->
    supprime pour rester fidele WinDev.
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        erreurs.append({"_erreur": f"Lecture : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_BJ_HEBDO.items()}
    db = get_pg_connection("adv")
    ligne_depart = p.ligne_depart or 3

    for i in range(ligne_depart, (ws.max_row or 0) + 1):
        num_bs = _cell(ws, i, cols["num_bs"]).upper()
        if not num_bs:
            continue
        # Dates : prendre [:10] avant parse
        date_sign = _parse_date_fr(_cell(ws, i, cols["date_signature"])[:10])
        date_va = _parse_date_fr(_cell(ws, i, cols["date_va"])[:10])
        date_ra = _parse_date_fr(_cell(ws, i, cols["date_ra"])[:10])
        date_rdv = _parse_date_fr(_cell(ws, i, cols["date_rdv"])[:10])
        type_v_s = _cell(ws, i, cols["type_vente"])
        type_vente = _type_vente_fibre(type_v_s)
        techno = _type_techno_fibre(_cell(ws, i, cols["technologie"]))
        lib_offre = _cell(ws, i, cols["offre"])
        offre = _type_offre_fibre(lib_offre, techno, date_sign)
        type_install = _cell(ws, i, cols["type_install"])
        lib_statut = _cell(ws, i, cols["lib_statut"])
        id_etat = _id_etat_fibre(lib_statut)
        cluster_code = _cell(ws, i, cols["cluster_code"])
        cluster_ville = _cell(ws, i, cols["cluster_ville"])
        client_cp = _cell(ws, i, cols["client_cp"])
        client_ville = _cell(ws, i, cols["client_ville"])
        client_rue = _cell(ws, i, cols["client_rue"])
        client_identite = _cell(ws, i, cols["client_identite"])
        client_tel = _cell(ws, i, cols["client_tel"])
        client_mail = _cell(ws, i, cols["client_mail"])
        motif_annul = _cell(ws, i, cols["motif_annul"])

        # Split identite "NOM, PRENOM"
        parts = client_identite.split(",", 1)
        client_nom = parts[0].strip() if parts else ""
        client_prenom = parts[1].strip() if len(parts) > 1 else ""

        # Detection migration FTTB VERS FTTH
        is_mig = "FTTB VERS FTTH" in lib_offre.upper()
        if is_mig and type_vente == 3:
            type_vente = 4
            migrations.append({
                "NumBS": num_bs, "DateSign": str(date_sign or ""),
                "ClientCP": client_cp, "ClientVille": client_ville,
                "LibStatut": lib_statut, "ClusterCode": cluster_code,
            })
            resume.nb_migrations += 1

        cluster = _test_cluster_fibre(cluster_code)
        box8 = "8" in lib_offre

        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, id_client, id_produit,
                      id_etat_contrat, type_vente, date_signature,
                      id_sfr_cluster
                 FROM adv.pgt_sfr_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )
        if not ctt:
            # ---- BRANCHE AJOUT (cf. WinDev l.275-300) ----
            ajoutes.append({
                "NumBS": num_bs, "DateSign": str(date_sign or ""),
                "DateVa": str(date_va or ""), "DateRa": str(date_ra or ""),
                "DateRDV": str(date_rdv or ""), "ClientCP": client_cp,
                "ClientVille": client_ville, "LibStatut": lib_statut,
                "TypeVente": type_vente, "ClusterCode": cluster_code,
            })
            resume.nb_ajoutes += 1
            if not p.simulation:
                try:
                    from app.intranets.adm.services.import_helpers_common import (
                        traiter_client,
                    )
                    from app.intranets.adm.services.sfr_helpers import (
                        ajout_fiche_contrat_sfr,
                    )
                    id_client = traiter_client(
                        info_client={
                            "nom": client_nom, "prenom": client_prenom,
                            "adresse1": client_rue, "cp": client_cp,
                            "ville": client_ville, "tel": client_tel,
                            "mail": client_mail,
                            "op_saisie": op_id, "modif_op": op_id,
                        }, force_maj=False, op_id=op_id,
                    ) or 0
                    ctt_data = {
                        "id_client": id_client,
                        "id_salarie": FIBRE_INCONNU,
                        "num_bs": num_bs,
                        "date_signature": date_sign,
                        "date_validation": date_va,
                        "date_racc_activ": date_ra,
                        "date_rdv_tech": date_rdv,
                        "id_etat_sfr": id_etat, "id_etat_contrat": id_etat,
                        "id_sfr_cluster": int((cluster or {}).get("id_sfr_cluster") or 0),
                        "id_produit": offre,
                        "technologie": techno,
                        "self_install": "SELF" in type_install.upper(),
                        "type_vente": type_vente,
                        "box8": box8, "box8_verif": box8,
                        "motif_annulation": motif_annul,
                        "non_call": True,
                        "hors_cible": bool((cluster or {}).get("hors_cible")),
                        "issu_tk_diff": 0,
                    }
                    id_ctt_new = ajout_fiche_contrat_sfr(ctt_data, op_id)
                    ajoutes[-1]["IdContratCree"] = id_ctt_new
                    try:
                        _ajoute_histo_sfr_etat(
                            id_ctt_new, 0, id_etat, "", op_id,
                            categorie="Vend",
                        )
                    except Exception:
                        pass
                except Exception as e:
                    ajoutes[-1]["Erreur"] = str(e)
        else:
            # ---- BRANCHE MODIF (cf. WinDev l.301-500) ----
            id_contrat = int(ctt["id_contrat"])
            id_client_db = int(ctt.get("id_client") or 0)
            etat_ctt_actuel = int(ctt.get("id_etat_contrat") or 0)
            tv_db = int(ctt.get("type_vente") or 0)
            id_prod_db = int(ctt.get("id_produit") or 0)
            modifs = []
            errs = []

            # 1. Detection erreurs Type Vente + Offre (rapport uniquement)
            # cf. WinDev l.333-355 : si les 2 < 3, on ignore. Sinon rapport.
            if tv_db != type_vente and not (tv_db < 3 and type_vente < 3):
                errs.append(("Type Vente", tv_db, type_vente))
            if offre and id_prod_db != offre:
                errs.append(("Offre", id_prod_db, offre))

            # 2. Determine type_etat_old pour la logique WinDev
            etat_info = db.query_one(
                """SELECT id_type_etat FROM adv.pgt_sfr_etat_contrat
                    WHERE id_etat = ? LIMIT 1""",
                (etat_ctt_actuel,),
            )
            type_etat_old = int((etat_info or {}).get("id_type_etat") or 0)

            # 3. MAJ etat contrat WinDev :
            # - type_etat_old NOT IN (5, 6) :
            #   * type_etat_old <=2 ou 7/8 -> UPDATE + histo "Vend"
            # - type_etat_old IN (5, 6) :
            #   * "Raccordement KO" et type_etat_old=5 -> rapport
            new_etat_ctt = None
            if id_etat and etat_ctt_actuel != id_etat:
                if type_etat_old not in (5, 6):
                    if type_etat_old <= 2 or type_etat_old in (7, 8):
                        new_etat_ctt = id_etat
                        modifs.append(f"Etat -> {id_etat}")
                else:
                    if type_etat_old == 5 and (
                            "Raccordement KO" in lib_statut
                            or "RESIL" in lib_statut.upper()):
                        errs.append(("Payé passé en KO",
                                     etat_ctt_actuel, id_etat))

            # 4. Cluster auto si etat "Temporaire" et cluster inconnu avec "-"
            # cf. WinDev l.381-397
            new_cluster_id = None
            if (etat_ctt_actuel == 1
                    and not (cluster or {}).get("id_sfr_cluster")
                    and "-" in cluster_code
                    and not p.simulation):
                try:
                    new_cluster_id = _new_id_sfr()
                    db.query(
                        """INSERT INTO adv.pgt_sfr_cluster
                              (id_sfr_cluster_auto, id_sfr_cluster,
                               region, code_vad, nom_cluster,
                               modif_date, modif_op, modif_elem)
                           VALUES (?, ?, ?, ?, ?, NOW(), ?, 'new')""",
                        (new_cluster_id, new_cluster_id, cluster_ville,
                         cluster_code, cluster_ville, int(op_id)),
                    )
                    modifs.append(f"Cluster auto -> {cluster_code}")
                except Exception:
                    new_cluster_id = None

            # 5. Rapports erreurs
            for lib_err, val_av, val_ap in errs:
                erreurs.append({
                    "NumBS": num_bs, "TypeErreur": lib_err,
                    "AvantImport": str(val_av), "ApresImport": str(val_ap),
                    "DateSign": str(ctt.get("date_signature") or ""),
                    "TypeStatut": type_install, "LibStatut": lib_statut,
                })
                resume.nb_erreurs += 1

            if modifs or errs:
                modifies.append({
                    "NumBS": num_bs, "Modifs": " | ".join(modifs) or "-",
                    "DateSign Omaya": str(ctt.get("date_signature") or ""),
                    "ClientCP": client_cp,
                    "Erreurs": [e[0] for e in errs] if errs else None,
                })
                resume.nb_modifies += 1
            else:
                resume.nb_non_modifies += 1

            # 6. PROD : UPDATE client + UPDATE contrat + histo si etat change
            if not p.simulation:
                # 6a. MAJ client (cf. WinDev l.307-320)
                if id_client_db:
                    try:
                        from app.intranets.adm.services.sfr_helpers import (
                            modif_fiche_client_sfr,
                        )
                        modif_fiche_client_sfr(
                            id_client_db,
                            nom=client_nom, prenom=client_prenom,
                            date_naiss=None,
                            adresse1=client_rue, adresse2="",
                            cp=client_cp, ville=client_ville,
                            tel=client_tel, gsm="", mail=client_mail,
                            op_id=op_id,
                        )
                    except Exception:
                        pass

                # 6b. UPDATE contrat (dates + self_install + techno + cluster)
                try:
                    sets = ["date_validation = COALESCE(?, date_validation)",
                            "date_racc_activ = COALESCE(?, date_racc_activ)",
                            "date_rdv_tech = COALESCE(?, date_rdv_tech)",
                            "self_install = ?"]
                    params: list = [date_va, date_ra, date_rdv,
                                     "SELF" in type_install.upper()]
                    if new_etat_ctt is not None:
                        sets.append("id_etat_contrat = ?")
                        params.append(new_etat_ctt)
                        sets.append("motif_annulation = ?")
                        params.append(motif_annul)
                    if new_cluster_id is not None:
                        sets.append("id_sfr_cluster = ?")
                        params.append(new_cluster_id)
                        sets.append("technologie = ?")
                        params.append(techno)
                    sets.append("modif_date = NOW()")
                    sets.append("modif_op = ?")
                    sets.append("modif_elem = 'modif'")
                    params.append(int(op_id))
                    params.append(id_contrat)
                    db.query(
                        f"UPDATE adv.pgt_sfr_contrat "
                        f"SET {', '.join(sets)} WHERE id_contrat = ?",
                        tuple(params),
                    )
                except Exception as e:
                    if modifs and modifies:
                        modifies[-1]["Erreur"] = str(e)

                # 6c. Historisation etat contrat (cf. WinDev l.447)
                if new_etat_ctt is not None:
                    try:
                        _ajoute_histo_sfr_etat(
                            id_contrat, etat_ctt_actuel, new_etat_ctt,
                            "", op_id, categorie="Vend",
                        )
                    except Exception:
                        pass
    wb.close()


def _import_placeholder(
    p: ImportSfrParams, fname: str, type_lbl: str,
    erreurs: list, resume: ImportSfrResume,
) -> None:
    """Placeholder pour les types non encore codes."""
    erreurs.append({
        "Fichier": fname,
        "Erreur": f"Type '{type_lbl}' non encore implementé (squelette).",
        "Note": "Logique métier WinDev à transposer.",
    })
    resume.nb_erreurs += 1


# ---------------------------------------------------------------------------
# Type 5 : ImportOptions (MAJ Box8Verif / OptionVerif)
# ---------------------------------------------------------------------------


def _import_options(
    p: ImportSfrParams, content: bytes, op_id: int,
    modifies: list, erreurs: list, resume: ImportSfrResume,
) -> None:
    """Type 5 : pour chaque BS, MAJ box8_verif ou option_verif=TRUE si
    StatOpt=VERSEE, selon Lib_Option (Box 8 vs autre)."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        erreurs.append({"_erreur": f"Lecture : {e}"})
        return
    ws = wb.active
    cols_map = {"num_bs": "A", "lib_option": "B", "statut_option": "C"}
    cols = {k: _col_letter_to_index(v) for k, v in cols_map.items()}
    db = get_pg_connection("adv")
    ligne_dep = p.ligne_depart or 2

    for i in range(ligne_dep, (ws.max_row or 0) + 1):
        num_bs = _cell(ws, i, cols["num_bs"]).upper().strip()
        if not num_bs:
            continue
        lib_opt = _cell(ws, i, cols["lib_option"])
        stat_opt = _cell(ws, i, cols["statut_option"]).upper()

        ctt = db.query_one(
            """SELECT id_contrat, num_bs, box8, box8_verif, option_dec,
                      option_verif, id_salarie, date_signature, id_etat_contrat
                 FROM adv.pgt_sfr_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )
        if not ctt:
            erreurs.append({"NumBS": num_bs, "Erreur": "BS introuvable",
                            "LibOption": lib_opt, "Statut": stat_opt})
            resume.nb_introuvables += 1
            continue

        id_contrat = int(ctt["id_contrat"])
        b8_avant = bool(ctt.get("box8_verif"))
        opt_avant = bool(ctt.get("option_verif"))

        is_box8 = "BOX 8" in lib_opt.upper()
        is_versee = stat_opt == "VERSEE"

        if is_box8 and is_versee:
            new_b8_verif = True
            new_opt_verif = opt_avant
        elif (not is_box8) and is_versee:
            new_b8_verif = b8_avant
            new_opt_verif = True
        else:
            # Pas de changement
            new_b8_verif = b8_avant
            new_opt_verif = opt_avant

        if new_b8_verif != b8_avant or new_opt_verif != opt_avant:
            modifies.append({
                "NumBS": num_bs, "LibOption": lib_opt, "Statut": stat_opt,
                "Box8Verif Avant": b8_avant, "Box8Verif Apres": new_b8_verif,
                "OptionVerif Avant": opt_avant, "OptionVerif Apres": new_opt_verif,
            })
            resume.nb_modifies += 1
            if not p.simulation:
                try:
                    db.query(
                        """UPDATE adv.pgt_sfr_contrat
                              SET box8_verif = ?, option_verif = ?,
                                  modif_date = NOW(), modif_op = ?,
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (new_b8_verif, new_opt_verif, int(op_id), id_contrat),
                    )
                except Exception as e:
                    modifies[-1]["Erreur"] = str(e)
        else:
            resume.nb_non_modifies += 1
    wb.close()


# ---------------------------------------------------------------------------
# Type 6 : ImportRUN (squelette multi-feuilles - logique metier partielle)
# ---------------------------------------------------------------------------


def _import_run(
    p: ImportSfrParams, content: bytes, op_id: int,
    ajoutes: list, modifies: list, erreurs: list,
    resume: ImportSfrResume,
) -> None:
    """Type 6 : ImportRUN. Lit 3 feuilles (Offre/Booster, Mobile, Option/
    Volumique) avec colonnes differentes par feuille. Pour chaque ligne :
    lookup BS, detection periode, creation/MAJ SFR_Contrat_Remun, verif
    montant officiel, hors delai.

    NOTE : implementation simplifiee, traitementOngletRun WinDev est tres
    riche (gestion VV/Racc, regul negative, geste co, statut Operateur,
    montant officiel par produit/typeVente/dateSign). A enrichir au fur
    et a mesure.
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        erreurs.append({"_erreur": f"Lecture : {e}"})
        return
    db = get_pg_connection("adv")

    p1_du = _parse_date_fr(p.periode1_du) or date(1900, 1, 1)
    p1_au = _parse_date_fr(p.periode1_au) or date(2100, 12, 31)
    p2_du = _parse_date_fr(p.periode2_du) or date(1900, 1, 1)
    p2_au = _parse_date_fr(p.periode2_au) or date(2100, 12, 31)
    mp1 = _dernier_jour_mois(p.periode1_mois_paiement)
    mp2 = _dernier_jour_mois(p.periode2_mois_paiement)
    mp_distrib = _dernier_jour_mois(p.mois_paiement_distrib)

    # Colonnes par feuille (cf grpValide WinDev Run1=Offres, Run2=Mobile/Option,
    # Run3=Volumique). Defaults raisonnables a ajuster selon vrai fichier.
    cols_per_sheet = {
        0: {"num_bs": "A", "techno": "B", "offre": "C", "type_vente": "D",
            "statut_ra": "E", "motif_annul": "F", "date_ra": "G",
            "hors_zone": "H", "type_rem": "I", "montant_rem": "J",
            "date_sign": "K", "client_nom": "L", "client_prenom": "M",
            "statut": "N", "motif_dero": "O", "geste_co": "P"},
        1: {"num_bs": "A", "techno": "B", "offre": "C", "type_vente": "D",
            "statut_ra": "E", "motif_annul": "F", "date_ra": "G",
            "hors_zone": "H", "type_opt": "I", "lib_opt": "J",
            "type_rem": "K", "montant_rem": "L", "date_sign": "M",
            "client_nom": "N", "client_prenom": "O", "statut": "P",
            "motif_dero": "Q"},
        2: {"num_bs": "A", "techno": "B", "offre": "C", "type_vente": "D",
            "statut_ra": "E", "motif_annul": "F", "date_ra": "G",
            "hors_zone": "H", "type_rem": "I", "montant_rem": "J",
            "periode": "K", "date_sign": "L", "client_nom": "M",
            "client_prenom": "N", "statut": "O"},
    }

    nb_introu = 0; nb_paye = 0; nb_deja_p = 0; nb_va_non_p = 0
    nb_mont0 = 0; nb_err_rem = 0; nb_err_hd = 0

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        if sheet_idx >= 3:
            break  # On ne traite que les 3 premieres feuilles
        ws = wb[sheet_name]
        cols_map = cols_per_sheet.get(sheet_idx, cols_per_sheet[0])
        cols = {k: _col_letter_to_index(v) for k, v in cols_map.items()}

        for i in range(2, (ws.max_row or 0) + 1):
            num_bs = _cell(ws, i, cols.get("num_bs", 1)).upper().strip()
            if not num_bs:
                continue
            client_nom = _cell(ws, i, cols.get("client_nom", 1))
            client_prenom = _cell(ws, i, cols.get("client_prenom", 1))
            techno = _cell(ws, i, cols.get("techno", 1))
            offre = _cell(ws, i, cols.get("offre", 1))
            type_vente = _cell(ws, i, cols.get("type_vente", 1))
            statut_ra = _cell(ws, i, cols.get("statut_ra", 1))
            motif_annul = _cell(ws, i, cols.get("motif_annul", 1))
            type_rem = _cell(ws, i, cols.get("type_rem", 1))
            hors_zone = _cell(ws, i, cols.get("hors_zone", 1))

            # Montant : test numerique + detection montant negatif
            mt_s = _cell(ws, i, cols.get("montant_rem", 1)).replace(",", ".")
            test_montant_rem = False
            montant_rem = 0.0
            try:
                if mt_s:
                    montant_rem = float(mt_s)
                    test_montant_rem = True
            except ValueError:
                test_montant_rem = False
            test_montant_neg = montant_rem < 0

            date_ra = _parse_date_fr(_cell(ws, i, cols.get("date_ra", 1)))
            date_sign = _parse_date_fr(_cell(ws, i, cols.get("date_sign", 1)))
            # Format MM-YYYY (cf. WinDev DateVersChaine "MM-AAAA")
            mois_sign = date_sign.strftime("%m-%Y") if date_sign else ""
            mois_racc = date_ra.strftime("%m-%Y") if date_ra else ""

            statut = _cell(ws, i, cols.get("statut", 1))
            # WinDev : Motif_Dero vide si feuille "Volumique"
            motif_dero = ""
            if "Volumique" not in sheet_name:
                motif_dero = _cell(ws, i, cols.get("motif_dero", 1))

            # GesteCo : boolean seulement si feuille "Offre"
            geste_co = False
            if "Offre" in sheet_name:
                geste_co = "GEST" in _cell(ws, i, cols.get("geste_co", 1)).upper()

            lib_opt = _cell(ws, i, cols.get("lib_opt", 1)) if "lib_opt" in cols else ""
            type_opt = _cell(ws, i, cols.get("type_opt", 1)) if "type_opt" in cols else ""
            periode = _cell(ws, i, cols.get("periode", 1)) if "periode" in cols else ""

            lib_rem = offre
            if sheet_name == "Option":  # exact match WinDev
                lib_rem = f"{type_opt} : {lib_opt}"
            elif "Volumique" in sheet_name:
                lib_rem = periode

            ctt = db.query_one(
                """SELECT id_contrat, id_salarie, date_signature, type_vente,
                          id_produit, remise, id_etat_sfr, info_vente_sfr
                     FROM adv.pgt_sfr_contrat
                    WHERE UPPER(num_bs) = UPPER(?)
                      AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                    LIMIT 1""",
                (num_bs,),
            )
            if not ctt:
                nb_introu += 1
                erreurs.append({
                    "Onglet": sheet_name, "NumBS": num_bs,
                    "DateSign": str(date_sign or ""),
                    "Client": f"{client_nom} {client_prenom}".strip(),
                    "Offre": offre, "TypeRem": type_rem,
                    "Montant": montant_rem, "Statut": statut,
                    "Erreur": "Introuvable",
                })
                resume.nb_introuvables += 1
                continue

            id_contrat = int(ctt["id_contrat"])
            id_sal = int(ctt.get("id_salarie") or 0)
            date_sign_db = ctt.get("date_signature")
            type_vente_db = int(ctt.get("type_vente") or 0)

            # Detection periode
            agence, equipe, is_distrib = _affectation_sfr_min(id_sal)
            mois_p, periode_lbl = _detect_periode_sfr(
                date_sign_db, is_distrib,
                p1_du, p1_au, mp1, p2_du, p2_au, mp2, mp_distrib,
            )
            mois_p_str = mois_p.strftime("%m-%Y") if mois_p else ""

            if periode_lbl == "HORS_DELAI":
                nb_err_hd += 1
                erreurs.append({
                    "Onglet": sheet_name, "NumBS": num_bs,
                    "Erreur": "Hors Délai",
                    "DateSign": str(date_sign_db or ""),
                    "Agence": agence, "Equipe": equipe,
                })
                resume.nb_hors_delai += 1

            # ---- BLOC 2 : MAJ statut contrat -> 76 (Rejet Operateur) ----
            # cf. WinDev traitementOngletRun :
            # si Contient(feuille,"Offre") ET Contient(TypeRem,"Offre") ET
            #    (Statut="NON_VERSABLE" OU (Statut="SUSPENDUE" ET StatutRa_Run="OK"))
            # Attention precedence WinDev : ET a priorite sur OU, donc :
            #    (Statut=NON_VERSABLE) OR (Statut=SUSPENDUE AND StatutRa_Run=OK)
            id_etat_sfr_db = int(ctt.get("id_etat_sfr") or 0)
            statut_up = statut.upper()
            statut_ra_up = statut_ra.upper()
            should_mark_76 = (
                "OFFRE" in sheet_name.upper()
                and "OFFRE" in type_rem.upper()
                and (statut_up == "NON_VERSABLE"
                     or (statut_up == "SUSPENDUE" and statut_ra_up == "OK"))
                and id_etat_sfr_db != 76
            )
            if should_mark_76:
                if not p.simulation:
                    try:
                        _ajoute_histo_sfr_etat(
                            id_contrat, id_etat_sfr_db, 76, "",
                            op_id, categorie="SFR",
                        )
                        new_info = (ctt.get("info_vente_sfr") or "")
                        if motif_dero:
                            new_info = f"{new_info}\n{motif_dero}"
                        db.query(
                            """UPDATE adv.pgt_sfr_contrat
                                  SET id_etat_sfr = 76,
                                      info_vente_sfr = ?,
                                      modif_date = NOW(), modif_op = ?,
                                      modif_elem = 'modif'
                                WHERE id_contrat = ?""",
                            (new_info, int(op_id), id_contrat),
                        )
                    except Exception as e:
                        erreurs.append({
                            "Onglet": sheet_name, "NumBS": num_bs,
                            "Erreur": f"UPDATE etat 76 : {e}",
                        })
                modifies.append({
                    "Onglet": sheet_name, "NumBS": num_bs,
                    "NouvelEtatSFR": 76,
                    "AncienEtatSFR": id_etat_sfr_db,
                    "Info": "Statut SFR mis en Rejet Opérateur - non payable",
                })

            if montant_rem == 0:
                # ---- BLOC 3 : Montant a 0 (feuille 4 WinDev) ----
                nb_mont0 += 1
                nom_vend = ""
                try:
                    v = get_pg_connection("rh").query_one(
                        "SELECT nom, prenom FROM rh.pgt_salarie WHERE id_salarie = ? LIMIT 1",
                        (id_sal,),
                    ) or {}
                    nom_vend = (
                        f"{v.get('nom') or ''} "
                        f"{(v.get('prenom') or '').title()}"
                    ).strip()
                except Exception:
                    pass
                modifies.append({
                    "Onglet": sheet_name,
                    "NumBS": num_bs,
                    "DateSign": str(date_sign or ""),
                    "DateRa": str(date_ra or ""),
                    "Client": f"{client_nom} {(client_prenom or '').title()}".strip(),
                    "Offre": offre,
                    "StatutRa": statut_ra,
                    "MotifAnnul": motif_annul,
                    "TypeRem": type_rem,
                    "LibRem": lib_rem,
                    "Periode": periode,
                    "MontantRem": montant_rem,
                    "Vendeur": nom_vend,
                    "Agence": agence,
                    "Equipe": equipe,
                    "MoisPaiement": mois_p_str,
                    "Statut": statut,
                    "MotifDero": motif_dero,
                    "Note": "Montant à 0",
                })
                continue

            lib_type_rem = type_rem.split(" (")[0]
            if test_montant_neg:
                lib_type_rem += " - Régul"

            # Rapport commun (18 champs feuille WinDev)
            def _report_row(feuille_note: str, extra: dict | None = None) -> dict:
                nom_vend = ""
                try:
                    v = get_pg_connection("rh").query_one(
                        "SELECT nom, prenom FROM rh.pgt_salarie WHERE id_salarie = ? LIMIT 1",
                        (id_sal,),
                    ) or {}
                    nom_vend = (
                        f"{v.get('nom') or ''} "
                        f"{(v.get('prenom') or '').title()}"
                    ).strip()
                except Exception:
                    pass
                r = {
                    "Onglet": sheet_name, "NumBS": num_bs,
                    "DateSign": str(date_sign or ""),
                    "DateRa": str(date_ra or ""),
                    "Client": f"{client_nom} {(client_prenom or '').title()}".strip(),
                    "Offre": offre, "StatutRa": statut_ra,
                    "MotifAnnul": motif_annul, "TypeRem": type_rem,
                    "LibRem": lib_rem, "Periode": periode,
                    "MontantRem": montant_rem, "Vendeur": nom_vend,
                    "Agence": agence, "Equipe": equipe,
                    "MoisPaiement": mois_p_str,
                    "Statut": statut, "MotifDero": motif_dero,
                    "Note": feuille_note,
                }
                if extra: r.update(extra)
                return r

            # Verifie rem deja enregistree (cf. WinDev ReqHistoRemSFR_contrat)
            existing_rem = None
            try:
                existing_rem = db.query_one(
                    """SELECT id_sfr_contrat_remun, ra_montant, ra_mois_p,
                              raccordement, validation
                         FROM adv.pgt_sfr_contrat_remun
                        WHERE id_contrat = ? AND type_rem = ? AND lib_option = ?
                        LIMIT 1""",
                    (id_contrat, lib_type_rem, lib_rem),
                )
            except Exception:
                existing_rem = None

            # -------- BLOC 4 : Techno "ABO" (Mobile) --------
            if techno.upper() == "ABO":
                if not existing_rem:
                    # INSERT + verif REM officielle
                    nb_paye += 1
                    ajoutes.append(_report_row("Payé"))
                    if not p.simulation:
                        try:
                            new_id = _new_id_sfr()
                            db.query(
                                """INSERT INTO adv.pgt_sfr_contrat_remun
                                      (id_sfr_contrat_remun_auto,
                                       id_sfr_contrat_remun,
                                       id_contrat, num, type_rem, lib_option,
                                       validation, va_mois_p, va_montant,
                                       va_statut, va_motif,
                                       raccordement, ra_mois_p, ra_montant,
                                       ra_statut, ra_motif,
                                       modif_date, modif_op, modif_elem)
                                   VALUES (?, ?, ?, ?, ?, ?,
                                           FALSE, '', 0, '', '',
                                           TRUE, ?, ?, ?, ?,
                                           NOW(), ?, 'new')""",
                                (new_id, new_id, id_contrat, num_bs,
                                 lib_type_rem, lib_rem,
                                 mois_p_str, montant_rem, statut, motif_dero,
                                 int(op_id)),
                            )
                        except Exception as e:
                            ajoutes[-1]["Erreur"] = str(e)
                    # Verif REM officielle
                    mnt_off = _verif_rem_officielle(offre, type_vente_db, date_sign_db)
                    if mnt_off > 0:
                        nb_err_rem += 1
                        if mnt_off > montant_rem:
                            note_err = "Rem insuffisante"
                        elif mnt_off < montant_rem:
                            note_err = "Versement en trop"
                        else:
                            note_err = "OK"
                        erreurs.append(_report_row(
                            note_err,
                            {"MontantOfficiel": mnt_off},
                        ))
                elif existing_rem.get("raccordement"):
                    # Deja paye
                    nb_deja_p += 1
                    modifies.append(_report_row("Déjà payé", {
                        "MontantOmaya": float(existing_rem.get("ra_montant") or 0),
                        "MoisPOmaya": str(existing_rem.get("ra_mois_p") or ""),
                        "MontantImport": montant_rem,
                        "MoisPImport": mois_p_str,
                    }))
                else:
                    # Existe mais raccordement=False -> UPDATE + Payé
                    nb_paye += 1
                    ajoutes.append(_report_row("Payé (MAJ)"))
                    if not p.simulation:
                        try:
                            db.query(
                                """UPDATE adv.pgt_sfr_contrat_remun
                                      SET raccordement = TRUE, ra_mois_p = ?,
                                          ra_montant = ?, ra_statut = ?,
                                          ra_motif = ?,
                                          modif_date = NOW(), modif_op = ?,
                                          modif_elem = 'modif'
                                    WHERE id_sfr_contrat_remun = ?""",
                                (mois_p_str, montant_rem, statut, motif_dero,
                                 int(op_id),
                                 int(existing_rem["id_sfr_contrat_remun"])),
                            )
                        except Exception as e:
                            ajoutes[-1]["Erreur"] = str(e)
                    # Verif REM officielle (idem que la branche INSERT)
                    mnt_off = _verif_rem_officielle(offre, type_vente_db, date_sign_db)
                    if mnt_off > 0:
                        nb_err_rem += 1
                        if mnt_off > montant_rem: note_err = "Rem insuffisante"
                        elif mnt_off < montant_rem: note_err = "Versement en trop"
                        else: note_err = "OK"
                        erreurs.append(_report_row(
                            note_err, {"MontantOfficiel": mnt_off},
                        ))
                continue

            # -------- BLOC 5 : Techno autre (Internet) --------
            is_vv = "(VV)" in type_rem
            id_produit_db = int(ctt.get("id_produit") or 0)
            remise_db = bool(ctt.get("remise"))

            def _do_verif_rem_internet() -> None:
                """Rapport verif REM Internet (cf. WinDev)."""
                nonlocal nb_err_rem
                if test_montant_neg:
                    # Regul REM (feuille 5)
                    nb_err_rem += 1
                    erreurs.append(_report_row("Regul REM", {
                        "MontantOfficiel": 0.0,
                        "Remise": remise_db, "GesteCo": geste_co,
                    }))
                    return
                mnt_off = _verif_rem_officielle_internet(
                    id_produit_db, type_vente_db, date_sign_db,
                    sheet_name, type_rem, remise_db,
                    mois_racc, mois_sign,
                )
                if mnt_off <= 0:
                    return
                nb_err_rem += 1
                if remise_db != geste_co:
                    note_err = "Erreur Geste Co"
                elif mnt_off > montant_rem:
                    note_err = "Rem insuffisante"
                elif mnt_off < montant_rem:
                    note_err = "Versement en trop"
                else:
                    note_err = "OK"
                erreurs.append(_report_row(note_err, {
                    "MontantOfficiel": mnt_off,
                    "Remise": remise_db, "GesteCo": geste_co,
                }))
                # Feuille 'Option' avec mnt_off > 0 -> MAJ OptionVerif
                if "option" in sheet_name.lower() and not p.simulation:
                    try:
                        db.query(
                            """UPDATE adv.pgt_sfr_contrat
                                  SET option_verif = TRUE,
                                      mois_p_option = ?,
                                      modif_date = NOW(), modif_op = ?,
                                      modif_elem = 'modif'
                                WHERE id_contrat = ?""",
                            (mois_p_str, int(op_id), id_contrat),
                        )
                    except Exception:
                        pass

            if not existing_rem:
                # ---- INSERT ----
                nb_paye += 1
                ajoutes.append(_report_row("Payé"))
                resume.nb_modifies += 1
                # cf. WinDev :
                # - Si (VV) : Validation=True + Va_*, Raccordement=False
                # - Sinon (Ra) : Raccordement=True + Ra_*
                #   Si MoisRacc=MoisSign : Validation=True + Va_MoisP + Va_Montant=0
                validation = is_vv or (not is_vv and mois_racc == mois_sign)
                va_mois_p = ""; va_montant = 0.0; va_statut = ""; va_motif = ""
                ra_mois_p = ""; ra_montant = 0.0; ra_statut = ""; ra_motif = ""
                if is_vv:
                    va_mois_p, va_montant = mois_p_str, montant_rem
                    va_statut, va_motif = statut, motif_dero
                else:
                    if mois_racc == mois_sign:
                        va_mois_p, va_montant = mois_p_str, 0.0
                        va_statut, va_motif = statut, motif_dero
                    ra_mois_p, ra_montant = mois_p_str, montant_rem
                    ra_statut, ra_motif = statut, motif_dero
                raccordement = not is_vv
                if not p.simulation:
                    try:
                        new_id = _new_id_sfr()
                        db.query(
                            """INSERT INTO adv.pgt_sfr_contrat_remun
                                  (id_sfr_contrat_remun_auto,
                                   id_sfr_contrat_remun,
                                   id_contrat, num, type_rem, lib_option,
                                   validation, va_mois_p, va_montant,
                                   va_statut, va_motif,
                                   raccordement, ra_mois_p, ra_montant,
                                   ra_statut, ra_motif,
                                   modif_date, modif_op, modif_elem)
                               VALUES (?, ?, ?, ?, ?, ?,
                                       ?, ?, ?, ?, ?,
                                       ?, ?, ?, ?, ?,
                                       NOW(), ?, 'new')""",
                            (new_id, new_id, id_contrat, num_bs,
                             lib_type_rem, lib_rem,
                             validation, va_mois_p, va_montant,
                             va_statut, va_motif,
                             raccordement, ra_mois_p, ra_montant,
                             ra_statut, ra_motif,
                             int(op_id)),
                        )
                    except Exception as e:
                        ajoutes[-1]["Erreur"] = str(e)
                _do_verif_rem_internet()
            elif is_vv:
                # ---- EXISTE + (VV) ----
                if existing_rem.get("validation"):
                    # Déjà payé Va
                    nb_deja_p += 1
                    modifies.append(_report_row("Déjà payé (Va)", {
                        "MontantOmaya": float(existing_rem.get("ra_montant") or 0),
                        "MontantImport": montant_rem,
                        "MoisPImport": mois_p_str,
                    }))
                else:
                    # UPDATE validation + Va_*
                    nb_paye += 1
                    ajoutes.append(_report_row("Payé (Va MAJ)"))
                    if not p.simulation:
                        try:
                            db.query(
                                """UPDATE adv.pgt_sfr_contrat_remun
                                      SET validation = TRUE, va_mois_p = ?,
                                          va_montant = ?, va_statut = ?,
                                          va_motif = ?,
                                          modif_date = NOW(), modif_op = ?,
                                          modif_elem = 'modif'
                                    WHERE id_sfr_contrat_remun = ?""",
                                (mois_p_str, montant_rem, statut, motif_dero,
                                 int(op_id),
                                 int(existing_rem["id_sfr_contrat_remun"])),
                            )
                        except Exception as e:
                            ajoutes[-1]["Erreur"] = str(e)
                    _do_verif_rem_internet()
            else:
                # ---- EXISTE + Ra ----
                # Cas 1 : Validation=False ET MoisSign≠MoisRacc -> nb_va_non_p
                if (not existing_rem.get("validation")
                        and mois_sign != mois_racc):
                    nb_va_non_p += 1
                    modifies.append(_report_row("Va non payé", {
                        "MontantExistantRa": float(existing_rem.get("ra_montant") or 0),
                        "MoisPExistantRa": str(existing_rem.get("ra_mois_p") or ""),
                    }))
                # Cas 2 : Raccordement=True -> Déjà payé Ra
                if existing_rem.get("raccordement"):
                    nb_deja_p += 1
                    modifies.append(_report_row("Déjà payé (Ra)", {
                        "MontantOmaya": float(existing_rem.get("ra_montant") or 0),
                        "MoisPOmaya": str(existing_rem.get("ra_mois_p") or ""),
                        "MontantImport": montant_rem,
                        "MoisPImport": mois_p_str,
                    }))
                else:
                    # Cas 3 : UPDATE Ra + éventuellement Va si MoisRacc=MoisSign
                    nb_paye += 1
                    ajoutes.append(_report_row("Payé (Ra MAJ)"))
                    if not p.simulation:
                        try:
                            sets = ["raccordement = TRUE",
                                    "ra_mois_p = ?", "ra_montant = ?",
                                    "ra_statut = ?", "ra_motif = ?"]
                            params: list = [mois_p_str, montant_rem,
                                            statut, motif_dero]
                            if mois_racc == mois_sign:
                                sets = ["validation = TRUE",
                                        "va_mois_p = ?", "va_montant = 0",
                                        "va_statut = ?", "va_motif = ?"] + sets
                                params = [mois_p_str, statut, motif_dero] + params
                            sets.append("modif_date = NOW()")
                            sets.append("modif_op = ?")
                            sets.append("modif_elem = 'modif'")
                            params.append(int(op_id))
                            params.append(int(existing_rem["id_sfr_contrat_remun"]))
                            db.query(
                                f"UPDATE adv.pgt_sfr_contrat_remun "
                                f"SET {', '.join(sets)} "
                                f"WHERE id_sfr_contrat_remun = ?",
                                tuple(params),
                            )
                        except Exception as e:
                            ajoutes[-1]["Erreur"] = str(e)
                    _do_verif_rem_internet()

    wb.close()
    # Reporting global
    erreurs.append({
        "Note": "Resume",
        "NB Paye": nb_paye, "NB Deja Paye": nb_deja_p,
        "NB VA Non Paye": nb_va_non_p, "NB Introuvable": nb_introu,
        "NB Montant 0": nb_mont0, "NB Err REM": nb_err_rem,
        "NB Hors Delai": nb_err_hd,
    })


def _affectation_sfr_min(id_salarie: int) -> tuple[str, str, bool]:
    """Wrapper minimal - reutilise _affectation_oen-like."""
    if not id_salarie:
        return ("", "", False)
    try:
        db = get_pg_connection("rh")
        rows = db.query(
            """SELECT o.lib_orga, o.id_type_niveau_orga, o.id_type_orga
                 FROM rh.pgt_salarie_organigramme so
                 JOIN rh.pgt_organigramme o ON o.idorganigramme = so.idorganigramme
                WHERE so.id_salarie = ?
                  AND (so.modif_elem IS NULL OR so.modif_elem NOT LIKE '%suppr%')""",
            (int(id_salarie),),
        ) or []
        agence = ""; equipe = ""; is_distrib = False
        for r in rows:
            lvl = r.get("id_type_niveau_orga")
            lib = r.get("lib_orga") or ""
            if lvl == 3 and not agence:
                agence = lib
                if int(r.get("id_type_orga") or 0) == 3:
                    is_distrib = True
            elif lvl == 4 and not equipe:
                equipe = lib
        return (agence, equipe, is_distrib)
    except Exception:
        return ("", "", False)


def _detect_periode_sfr(
    date_sign, is_distrib: bool,
    p1_du: date, p1_au: date, mp1, p2_du: date, p2_au: date, mp2,
    mp_distrib,
):
    if not date_sign:
        return (None, "")
    if is_distrib:
        return (mp_distrib, "Distrib")
    if p1_du <= date_sign <= p1_au:
        return (mp1, "Période 1")
    if p2_du <= date_sign <= p2_au:
        return (mp2, "Période 2")
    p1_du_m1 = (p1_du.replace(month=p1_du.month - 1) if p1_du.month > 1
                else p1_du.replace(year=p1_du.year - 1, month=12))
    p1_au_m1 = (p1_au.replace(month=p1_au.month - 1) if p1_au.month > 1
                else p1_au.replace(year=p1_au.year - 1, month=12))
    if p1_du_m1 <= date_sign <= p1_au_m1:
        return (mp1, "Période -1 mois")
    return (None, "HORS_DELAI")


# ---------------------------------------------------------------------------
# CALL RET (types 7, 8, 9, 10) - vendeurs sentinelles + helpers
# ---------------------------------------------------------------------------

# Salaries sentinelles (cf code WinDev)
ABASSI_AHD = 20230920112006538          # vendeur cible Call RET
FIBRE_INCONNU = 20200715153948361       # vendeur par defaut si TK pas trouve
ID_AFFECT_DISTRIB_CALL_RETENTION = 20230601160425831


def _donne_statut_call(libelle: str) -> tuple[int, str]:
    """donneStatutCall : lookup pgt_sfr_etat_call_ret by lib_etat LIKE.
    Retourne (id_etat_call_ret, lib_etat)."""
    if not libelle:
        return (0, "")
    try:
        db = get_pg_connection("adv")
        r = db.query_one(
            """SELECT id_etat_call_ret, lib_etat FROM adv.pgt_sfr_etat_call_ret
                WHERE LOWER(lib_etat) LIKE LOWER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (f"%{libelle}%",),
        )
        return ((_int(r.get("id_etat_call_ret")), _str(r.get("lib_etat")))
                if r else (0, ""))
    except Exception:
        return (0, "")


def _iter_sheets(content: bytes):
    """Itere toutes les feuilles d'un classeur (nb_sheets de WinDev)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        yield wb[sheet], sheet
    wb.close()


def _verif_rem_officielle_internet(
    id_produit: int, type_vente: int, date_sign: Optional[date],
    sheet_name: str, type_rem: str, remise: bool,
    mois_racc: str, mois_sign: str,
) -> float:
    """cf. WinDev bloc 5 (Internet) verif REM officielle.

    Regles :
    - Feuille contient 'Offre' ET TypeRem contient 'Offre' :
      * TypeRem contient '(VV)' : montant_va_remise ou montant_va selon remise
      * sinon : montant_ra(_remise) + montant_va(_remise) si MoisRacc=MoisSign
    - Feuille contient 'Option' : abonnement_tv
    - Feuille contient 'Volumique' : prime_volumique

    TypeVente=2 (Migration) -> traite comme 1 (Vente).
    """
    if not id_produit or not date_sign:
        return 0.0
    tv = 1 if type_vente == 2 else type_vente
    try:
        db = get_pg_connection("adv")
        r = db.query_one(
            """SELECT montant_va, montant_va_remise, montant_ra,
                      montant_ra_remise, abonnement_tv, prime_volumique
                 FROM adv.pgt_sfr_remun
                WHERE id_produit = ? AND type_vente = ?
                  AND ? BETWEEN date_debut AND date_fin
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (id_produit, tv, date_sign),
        )
        if not r: return 0.0
        sn = sheet_name.lower()
        tr = type_rem.lower()
        mnt = 0.0
        if "offre" in sn and "offre" in tr:
            if "(vv)" in tr:
                mnt = float((r.get("montant_va_remise")
                              if remise else r.get("montant_va")) or 0)
            else:
                if remise:
                    mnt = float(r.get("montant_ra_remise") or 0)
                    if mois_racc == mois_sign:
                        mnt += float(r.get("montant_va_remise") or 0)
                else:
                    mnt = float(r.get("montant_ra") or 0)
                    if mois_racc == mois_sign:
                        mnt += float(r.get("montant_va") or 0)
        elif "option" in sn:
            mnt = float(r.get("abonnement_tv") or 0)
        elif "volumique" in sn:
            mnt = float(r.get("prime_volumique") or 0)
        return mnt
    except Exception:
        return 0.0


def _verif_rem_officielle(offre_run: str, type_vente: int,
                           date_sign: Optional[date]) -> float:
    """cf. WinDev ReqTrouveRemCttSFR(Offre_Run, typeV, DateSi).

    Cherche la ligne de la grille pgt_sfr_remun matching :
    - lib_produit LIKE '<offre_run>%' (ou 'Internet Partout%' si commence par)
    - type_vente = typeV (si TypeVente=2 -> 1)
    - date_sign entre date_debut et date_fin (grille valide)

    Retourne montant_ra ou 0.0 si pas de correspondance.
    """
    if not offre_run or not date_sign:
        return 0.0
    # cf. WinDev : si Offre_Run commence par "Internet Partout" -> "Internet Partout"
    lib_search = offre_run
    if offre_run.upper().startswith("INTERNET PARTOUT"):
        lib_search = "Internet Partout"
    # TypeVente = 2 (Migration) -> traite comme 1 (Vente)
    tv = 1 if type_vente == 2 else type_vente
    try:
        db = get_pg_connection("adv")
        r = db.query_one(
            """SELECT g.montant_ra
                 FROM adv.pgt_sfr_remun g
                 JOIN adv.pgt_sfr_produit p ON p.id_produit = g.id_produit
                WHERE p.lib_produit LIKE ?
                  AND g.type_vente = ?
                  AND ? BETWEEN g.date_debut AND g.date_fin
                  AND (g.modif_elem IS NULL OR g.modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (f"{lib_search}%", tv, date_sign),
        )
        return float(r.get("montant_ra") or 0) if r else 0.0
    except Exception:
        return 0.0


def _ajoute_histo_sfr_etat(id_contrat: int, old_etat: int, new_etat: int,
                            date_paiement: str, op_id: int,
                            categorie: str = "Vend") -> None:
    """Historise un changement d'etat SFR (cf. WinDev ajouteHistoContrat).

    2 tables selon categorie :
    - 'Vend' (defaut) -> pgt_sfr_histo_etat_ctt
    - 'SFR'           -> pgt_sfr_histo_etat_ctt_sfr
    """
    if not id_contrat:
        return
    db = get_pg_connection("adv")
    table = ("pgt_sfr_histo_etat_ctt_sfr" if categorie.upper() == "SFR"
             else "pgt_sfr_histo_etat_ctt")
    auto = db.query_one(
        f"SELECT COALESCE(MAX(id_histo_auto), 0) + 1 AS n FROM adv.{table}"
    )
    db.query(
        f"""INSERT INTO adv.{table}
              (id_histo_auto, id_histo, id_contrat, op_saisie, date,
               old_etat, new_etat, date_paiement,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, NOW(), ?, ?, ?, ?, NOW(), 'new')""",
        (int(auto["n"]) if auto else 1, _new_id_sfr(),
         int(id_contrat), int(op_id),
         int(old_etat) if old_etat else 0,
         int(new_etat) if new_etat else 0,
         date_paiement or "", int(op_id)),
    )


def _new_id_sfr() -> int:
    """Id 8 octets construit depuis la date/heure (equivalent
    idEntierDateHeureSys WinDev)."""
    from datetime import datetime as _dt
    n = _dt.now()
    return int(n.strftime("%Y%m%d%H%M%S")) * 1000 + n.microsecond // 1000


def _create_sfr_contrat_callret(orig: dict, new_bs: str,
                                date_sign: Optional[date], op_id: int,
                                info_motif: str) -> int:
    """Cree un nouveau SFR_contrat (rattrape par Call RET) en cascadant les
    valeurs du contrat origine.

    Delegue a ajout_fiche_contrat_sfr (sfr_helpers) qui gere ~50 colonnes
    et le recalcul nb_points fam FIB CQ + bonus. Ajoute ensuite
    l'historisation de l'etat initial (cf. WinDev ajouteHistoContrat).
    """
    from app.intranets.adm.services.sfr_helpers import (
        ajout_fiche_contrat_sfr,
    )

    # Construction dict compatible ST_CONTRAT_SFR
    ctt = {
        "id_client": int(orig.get("id_client") or 0),
        "id_salarie": int(ABASSI_AHD),  # cible Call RET
        "id_ste": 0,
        "num_bs": new_bs,
        "date_signature": date_sign,
        "id_etat_sfr": 1,
        "id_etat_contrat": 1,
        "id_sfr_cluster": int(orig.get("id_sfr_cluster") or 0),
        "id_produit": int(orig.get("id_produit") or 0),
        "technologie": int(orig.get("technologie") or 0),
        "self_install": bool(orig.get("self_install")),
        "type_vente": int(orig.get("type_vente") or 0),
        "box8": bool(orig.get("box8")),
        "box8_verif": bool(orig.get("box8")),
        "option_dec": bool(orig.get("option_dec")),
        "option_verif": bool(orig.get("option_verif")),
        "motif_annulation": "",
        "info_interne": info_motif,
        "non_call": True,
        "remise": bool(orig.get("remise")),
        "hors_cible": bool(orig.get("hors_cible")),
        "issu_tk_diff": 0,
    }
    id_contrat = ajout_fiche_contrat_sfr(ctt, op_id)

    # Historisation etat initial (cf. WinDev ajouteHistoContrat)
    try:
        _ajoute_histo_sfr_etat(id_contrat, 0, 1, "", op_id, categorie="Vend")
    except Exception:
        pass

    return id_contrat


def _import_callret_ko(
    p: ImportSfrParams, content: bytes, op_id: int,
    ajoutes: list, modifies: list, erreurs: list, resume: ImportSfrResume,
) -> None:
    """Type 7 : ImportCallRET_KO. Lit toutes les feuilles, pour chaque ligne
    lookup BS, ajoute new BS si fourni + MAJ id_etat_call_ret + obs."""
    cols_map = {"num_bs": "A", "statut": "I", "comment": "L", "new_bs": "AT"}
    cols = {k: _col_letter_to_index(v) for k, v in cols_map.items()}
    db = get_pg_connection("adv")
    ligne_dep = p.ligne_depart or 2

    for ws, _sheet in _iter_sheets(content):
        for i in range(ligne_dep, (ws.max_row or 0) + 1):
            num_bs = _cell(ws, i, cols["num_bs"]).upper().replace(" ", "")
            if not num_bs:
                continue
            statut = _cell(ws, i, cols["statut"])
            comment = _cell(ws, i, cols["comment"])
            new_bs = _cell(ws, i, cols["new_bs"]).upper().replace(" ", "")
            id_statut, lib_etat_call = _donne_statut_call(statut)
            if statut and id_statut == 0:
                erreurs.append({
                    "NumBS": num_bs, "Erreur": "Statut inconnu",
                    "Statut": statut,
                })
                resume.nb_erreurs += 1

            orig = db.query_one(
                """SELECT id_contrat, id_client, id_salarie, id_produit,
                          date_signature, id_sfr_cluster, technologie,
                          self_install, type_vente, box8, option_dec,
                          option_verif, remise, hors_cible, id_etat_call_ret,
                          obs_call_ret, id_contrat_ret
                     FROM adv.pgt_sfr_contrat
                    WHERE UPPER(num_bs) = UPPER(?)
                      AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                    LIMIT 1""",
                (num_bs,),
            )
            if not orig:
                erreurs.append({"NumBS": num_bs, "Erreur": "BS introuvable",
                                "Statut": statut})
                resume.nb_introuvables += 1
                continue

            id_contrat_orig = int(orig["id_contrat"])
            modifs_done = []

            # Ajout d'un nouveau BS si fourni
            if new_bs:
                # Verifier si le new_bs existe deja
                exist = db.query_one(
                    """SELECT id_contrat FROM adv.pgt_sfr_contrat
                        WHERE UPPER(num_bs) = UPPER(?) LIMIT 1""",
                    (new_bs,),
                )
                row_a = {
                    "NumBS_Origine": num_bs, "NewBS": new_bs,
                    "LibEtatCall": lib_etat_call,
                    "Existant": bool(exist),
                }
                if not exist:
                    if not p.simulation:
                        try:
                            new_id = _create_sfr_contrat_callret(
                                orig, new_bs, orig.get("date_signature"),
                                op_id,
                                f"Contrat KO Rattrape par le Call RET, "
                                f"Num Origine : {num_bs}",
                            )
                            row_a["NewIdContrat"] = new_id
                            # MAJ id_contrat_ret sur l'origine
                            db.query(
                                """UPDATE adv.pgt_sfr_contrat
                                      SET id_contrat_ret = ?, modif_date = NOW(),
                                          modif_op = ?, modif_elem = 'modif'
                                    WHERE id_contrat = ?""",
                                (new_id, int(op_id), id_contrat_orig),
                            )
                        except Exception as e:
                            row_a["Erreur"] = str(e)
                ajoutes.append(row_a)
                resume.nb_ajoutes += 1

            # MAJ id_etat_call_ret + obs si different
            if id_statut and int(orig.get("id_etat_call_ret") or 0) != id_statut:
                modifs_done.append(f"EtatCallRet -> {id_statut} ({lib_etat_call})")
                if not p.simulation:
                    new_obs = (f"{datetime.now().strftime('%d/%m/%Y %H:%M , ')}"
                               f"{lib_etat_call} : {comment}\n")
                    db.query(
                        """UPDATE adv.pgt_sfr_contrat
                              SET id_etat_call_ret = ?, obs_call_ret = ?,
                                  modif_date = NOW(), modif_op = ?,
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (id_statut, new_obs, int(op_id), id_contrat_orig),
                    )

            if modifs_done:
                modifies.append({
                    "NumBS": num_bs, "LibEtatCall": lib_etat_call,
                    "Modifs": " | ".join(modifs_done), "Comment": comment,
                })
                resume.nb_modifies += 1


def _import_callret_racc(
    p: ImportSfrParams, content: bytes, op_id: int,
    ajoutes: list, modifies: list, erreurs: list, resume: ImportSfrResume,
) -> None:
    """Type 8 : ImportCallRET_Racc. 1 BS origine + jusqu'a 4 nouveaux BS."""
    cols_map = {"num_bs": "A", "comment": "L",
                "new_bs1": "AT", "new_bs2": "AU", "new_bs3": "AV", "new_bs4": "AW"}
    cols = {k: _col_letter_to_index(v) for k, v in cols_map.items()}
    db = get_pg_connection("adv")
    ligne_dep = p.ligne_depart or 2

    for ws, _sheet in _iter_sheets(content):
        for i in range(ligne_dep, (ws.max_row or 0) + 1):
            num_bs = _cell(ws, i, cols["num_bs"]).upper().replace(" ", "")
            if not num_bs:
                continue
            comment = _cell(ws, i, cols["comment"])
            new_bss = [_cell(ws, i, cols[k]).upper().replace(" ", "")
                       for k in ("new_bs1", "new_bs2", "new_bs3", "new_bs4")]
            new_bss = [b for b in new_bss if b]

            orig = db.query_one(
                """SELECT id_contrat, id_client, id_salarie, id_produit,
                          date_signature, id_sfr_cluster, technologie,
                          self_install, type_vente, box8, option_dec,
                          option_verif, remise, hors_cible, info_interne
                     FROM adv.pgt_sfr_contrat
                    WHERE UPPER(num_bs) = UPPER(?)
                      AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                    LIMIT 1""",
                (num_bs,),
            )
            if not orig:
                erreurs.append({"NumBS": num_bs, "Erreur": "BS introuvable"})
                resume.nb_introuvables += 1
                continue

            id_orig = int(orig["id_contrat"])
            id_clt = int(orig.get("id_client") or 0)
            liste_bs_add = []

            for new_bs in new_bss:
                exist = db.query_one(
                    """SELECT id_contrat, id_client, id_salarie
                         FROM adv.pgt_sfr_contrat
                        WHERE UPPER(num_bs) = UPPER(?) LIMIT 1""",
                    (new_bs,),
                )
                row_a = {"NumBS_Origine": num_bs, "NewBS": new_bs,
                         "Existant": bool(exist)}
                if exist:
                    # MAJ client / vendeur si different
                    if int(exist.get("id_client") or 0) != id_clt:
                        if not p.simulation:
                            db.query(
                                """UPDATE adv.pgt_sfr_contrat
                                      SET id_client = ?, modif_date = NOW(),
                                          modif_op = ?, modif_elem = 'modif'
                                    WHERE id_contrat = ?""",
                                (id_clt, int(op_id), int(exist["id_contrat"])),
                            )
                        row_a["Note"] = "ID Client mis a jour"
                    sal_db = int(exist.get("id_salarie") or 0)
                    if sal_db in (0, FIBRE_INCONNU):
                        if not p.simulation:
                            db.query(
                                """UPDATE adv.pgt_sfr_contrat
                                      SET id_salarie = ?, modif_date = NOW(),
                                          modif_op = ?, modif_elem = 'modif'
                                    WHERE id_contrat = ?""",
                                (int(ABASSI_AHD), int(op_id),
                                 int(exist["id_contrat"])),
                            )
                        row_a["Note"] = (row_a.get("Note", "")
                                         + " | ID Vendeur mis a jour").strip(" |")
                else:
                    if not p.simulation:
                        try:
                            new_id = _create_sfr_contrat_callret(
                                orig, new_bs, orig.get("date_signature"),
                                op_id,
                                f"Vente RACC par le Call RET, "
                                f"Num Origine : {num_bs}",
                            )
                            row_a["NewIdContrat"] = new_id
                        except Exception as e:
                            row_a["Erreur"] = str(e)
                ajoutes.append(row_a)
                resume.nb_ajoutes += 1
                liste_bs_add.append(new_bs)

            # MAJ info_interne origine
            if liste_bs_add and not p.simulation:
                info_add = (f"\n{datetime.now().strftime('%d/%m/%Y %H:%M , ')}"
                            f"Ventes ADD par Call RET\n    - "
                            + "\n    - ".join(liste_bs_add))
                obs_add = (f"{datetime.now().strftime('%d/%m/%Y %H:%M , ')}"
                           f"{comment}\n" if comment else "")
                try:
                    db.query(
                        """UPDATE adv.pgt_sfr_contrat
                              SET info_interne = COALESCE(info_interne, '') || ?,
                                  obs_call_ret = CASE WHEN ? <> ''
                                                      THEN ? ELSE obs_call_ret END,
                                  modif_date = NOW(), modif_op = ?,
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (info_add, obs_add, obs_add, int(op_id), id_orig),
                    )
                except Exception as e:
                    modifies.append({"NumBS": num_bs, "Erreur": str(e)})


def _import_callret_rdvtech(
    p: ImportSfrParams, content: bytes, op_id: int,
    ajoutes: list, modifies: list, erreurs: list, resume: ImportSfrResume,
) -> None:
    """Type 10 : ImportCallRET_RDVTech. MAJ date_rdv_tech + id_sfr_statut_rdv."""
    cols_map = {"num_bs": "A", "statut": "I", "comment": "L", "date_rdv": "BB"}
    cols = {k: _col_letter_to_index(v) for k, v in cols_map.items()}
    db = get_pg_connection("adv")
    ligne_dep = p.ligne_depart or 2

    for ws, _sheet in _iter_sheets(content):
        for i in range(ligne_dep, (ws.max_row or 0) + 1):
            num_bs = _cell(ws, i, cols["num_bs"]).upper().replace(" ", "")
            if not num_bs:
                continue
            statut = _cell(ws, i, cols["statut"])
            comment = _cell(ws, i, cols["comment"])
            date_rdv = _parse_date_fr(_cell(ws, i, cols["date_rdv"]))

            id_statut, lib_etat = _donne_statut_call(statut) if statut else (0, "")
            id_statut_rdv = 0
            lib_statut_rdv = ""
            if id_statut:
                # Lookup id_etat_rdv_tech sur etat_call_ret
                r = db.query_one(
                    """SELECT id_etat_rdv_tech FROM adv.pgt_sfr_etat_call_ret
                        WHERE id_etat_call_ret = ? LIMIT 1""",
                    (id_statut,),
                )
                if r and r.get("id_etat_rdv_tech"):
                    r2 = db.query_one(
                        """SELECT id_sfr_statut_rdv, lib_statut
                             FROM adv.pgt_sfr_statut_rdv
                            WHERE id_sfr_statut_rdv = ? LIMIT 1""",
                        (int(r["id_etat_rdv_tech"]),),
                    )
                    if r2:
                        id_statut_rdv = int(r2.get("id_sfr_statut_rdv") or 0)
                        lib_statut_rdv = _str(r2.get("lib_statut"))

            orig = db.query_one(
                """SELECT id_contrat, date_rdv_tech, id_sfr_statut_rdv,
                          info_interne
                     FROM adv.pgt_sfr_contrat
                    WHERE UPPER(num_bs) = UPPER(?)
                      AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                    LIMIT 1""",
                (num_bs,),
            )
            if not orig:
                erreurs.append({"NumBS": num_bs, "Erreur": "Contrat introuvable",
                                "Statut": statut, "Comment": comment})
                resume.nb_introuvables += 1
                continue

            id_orig = int(orig["id_contrat"])
            modifs_done = []
            new_info = orig.get("info_interne") or ""
            new_rdv = orig.get("date_rdv_tech")
            new_statut_rdv = int(orig.get("id_sfr_statut_rdv") or 0)

            if (comment and "Import Call Ret RDV Tech :"
                    not in (orig.get("info_interne") or "")):
                modifs_done.append("Comment ajoute")
                new_info += (f"\n{datetime.now().strftime('%d/%m/%Y %H:%M , ')}"
                             f"Import Call Ret RDV Tech : {comment}")
            if id_statut_rdv and new_statut_rdv != id_statut_rdv:
                modifs_done.append(f"StatutRDV -> {id_statut_rdv} ({lib_statut_rdv})")
                new_statut_rdv = id_statut_rdv
            if date_rdv and (not new_rdv or new_rdv < date_rdv):
                old_str = (new_rdv.strftime("%d/%m/%Y") if new_rdv else "")
                modifs_done.append(f"DateRDV {old_str} -> {date_rdv}")
                new_info += (f"\nModif RDV du {old_str} au "
                             f"{date_rdv.strftime('%d/%m/%Y')}")
                new_rdv = date_rdv

            row_snap = {
                "NumBS": num_bs, "StatutRDV": lib_statut_rdv,
                "DateRDV": str(date_rdv or ""),
                "Modifs": " | ".join(modifs_done) if modifs_done else "(aucune)",
            }
            if modifs_done:
                ajoutes.append(row_snap)
                resume.nb_modifies += 1
                if not p.simulation:
                    try:
                        db.query(
                            """UPDATE adv.pgt_sfr_contrat
                                  SET date_rdv_tech = ?, id_sfr_statut_rdv = ?,
                                      info_interne = ?,
                                      modif_date = NOW(), modif_op = ?,
                                      modif_elem = 'modif'
                                WHERE id_contrat = ?""",
                            (new_rdv, new_statut_rdv, new_info,
                             int(op_id), id_orig),
                        )
                    except Exception as e:
                        row_snap["Erreur"] = str(e)


def _import_callret_ventesadd(
    p: ImportSfrParams, content: bytes, op_id: int,
    ajoutes: list, modifies: list, erreurs: list, resume: ImportSfrResume,
) -> None:
    """Type 9 : ImportCallRET_VentesADD. Reattribution vendeur a ABASSI si
    fibre inconnu/anonymous. Sinon : si vendeur deja en Distrib Call Retention,
    on signale en erreur."""
    cols_map = {"num_bs": "A", "statut": "I", "comment": "L",
                "date_sign": "C", "categorie": "M", "offre": "J",
                "date_racc": "E"}
    cols = {k: _col_letter_to_index(v) for k, v in cols_map.items()}
    db = get_pg_connection("adv")
    ligne_dep = p.ligne_depart or 2

    for ws, _sheet in _iter_sheets(content):
        for i in range(ligne_dep, (ws.max_row or 0) + 1):
            num_bs = _cell(ws, i, cols["num_bs"]).upper().replace(" ", "")
            if not num_bs:
                continue
            statut = _cell(ws, i, cols["statut"])
            comment = _cell(ws, i, cols["comment"])
            offre = _cell(ws, i, cols["offre"])
            id_statut, lib_etat_call = _donne_statut_call(statut)

            orig = db.query_one(
                """SELECT id_contrat, id_salarie
                     FROM adv.pgt_sfr_contrat
                    WHERE UPPER(num_bs) = UPPER(?)
                      AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                    LIMIT 1""",
                (num_bs,),
            )
            if not orig:
                erreurs.append({"NumBS": num_bs, "Erreur": "Contrat introuvable",
                                "Statut": statut, "Comment": comment})
                resume.nb_introuvables += 1
                continue

            id_orig = int(orig["id_contrat"])
            id_sal_db = int(orig.get("id_salarie") or 0)

            if id_sal_db in (0, FIBRE_INCONNU):
                # Reattribution a ABASSI
                ajoutes.append({
                    "NumBS": num_bs, "Offre": offre,
                    "LibEtatCall": lib_etat_call,
                    "OldIdSalarie": id_sal_db,
                    "NewIdSalarie": ABASSI_AHD,
                })
                resume.nb_modif_vend += 1
                if not p.simulation:
                    try:
                        db.query(
                            """UPDATE adv.pgt_sfr_contrat
                                  SET id_salarie = ?, modif_date = NOW(),
                                      modif_op = ?, modif_elem = 'modif'
                                WHERE id_contrat = ?""",
                            (int(ABASSI_AHD), int(op_id), id_orig),
                        )
                    except Exception as e:
                        ajoutes[-1]["Erreur"] = str(e)
            else:
                # cf. WinDev l.174-183 : verifier si le vendeur est deja
                # dans l'organigramme Distrib Call Retention (via
                # id_parent = ID_AFFECT_DISTRIB_CALL_RETENTION).
                # - Si oui   -> silence (deja OK)
                # - Si non   -> erreur 'attribue a un vendeur externe'
                #   (contrat attribue a un autre vendeur, ne pas reattribuer)
                is_deja_call_ret = False
                try:
                    rows = db.query(
                        """SELECT o.id_parent
                             FROM rh.pgt_salarie_organigramme so
                             JOIN rh.pgt_organigramme o
                                  ON o.idorganigramme = so.idorganigramme
                            WHERE so.id_salarie = ?
                              AND (so.modif_elem IS NULL
                                   OR so.modif_elem NOT LIKE '%suppr%')""",
                        (id_sal_db,),
                    ) or []
                    is_deja_call_ret = any(
                        int(r.get("id_parent") or 0)
                            == ID_AFFECT_DISTRIB_CALL_RETENTION
                        for r in rows
                    )
                except Exception:
                    is_deja_call_ret = False

                if not is_deja_call_ret:
                    erreurs.append({
                        "NumBS": num_bs,
                        "Erreur": "Contrat attribué à un vendeur",
                        "Statut": statut, "IdSalarie": id_sal_db,
                    })
                    resume.nb_erreurs += 1


def run_import_sfr(
    p: ImportSfrParams, files: list[tuple[str, bytes]], op_id: int,
) -> ImportSfrResult:
    """Dispatcher principal."""
    label = TYPE_LABELS.get(p.type_import, "?")
    if not files:
        return ImportSfrResult(
            ok=False, type_import=p.type_import, type_label=label,
            simulation=p.simulation, resume=ImportSfrResume(),
            message="Aucun fichier fourni.",
        )

    resume = ImportSfrResume(nb_fichiers=len(files))
    ajoutes: list[dict] = []
    modifies: list[dict] = []
    non_trouves: list[dict] = []
    migrations: list[dict] = []
    modif_vendeurs: list[dict] = []
    erreurs: list[dict] = []
    fichiers_traites: list[str] = []

    for fname, content in files:
        fichiers_traites.append(fname)
        try:
            if p.type_import == 1:
                _import_journalier_fibre(
                    p, fname, content, op_id,
                    ajoutes, modifies, migrations, modif_vendeurs,
                    erreurs, resume,
                )
            elif p.type_import == 2:
                _import_journalier_mobile(p, content, op_id,
                                          ajoutes, modifies, erreurs, resume)
            elif p.type_import == 3:
                _import_journalier_call(p, content, op_id,
                                        ajoutes, modifies, erreurs, resume)
            elif p.type_import == 4:
                _import_journalier_hebdo(p, content, op_id,
                                         ajoutes, modifies, migrations,
                                         erreurs, resume)
            elif p.type_import == 5:
                _import_options(p, content, op_id,
                                modifies, erreurs, resume)
            elif p.type_import == 6:
                _import_run(p, content, op_id,
                            ajoutes, modifies, erreurs, resume)
            elif p.type_import == 7:
                _import_callret_ko(p, content, op_id,
                                   ajoutes, modifies, erreurs, resume)
            elif p.type_import == 8:
                _import_callret_racc(p, content, op_id,
                                     ajoutes, modifies, erreurs, resume)
            elif p.type_import == 9:
                _import_callret_ventesadd(p, content, op_id,
                                          ajoutes, modifies, erreurs, resume)
            elif p.type_import == 10:
                _import_callret_rdvtech(p, content, op_id,
                                        ajoutes, modifies, erreurs, resume)
            else:
                _import_placeholder(p, fname, label, erreurs, resume)
        except Exception as e:
            erreurs.append({"Fichier": fname, "Erreur": str(e)})
            resume.nb_erreurs += 1

    # Cleanup payloads
    for row in ajoutes:
        row.pop("_payload_create", None)

    res = ImportSfrResult(
        ok=True, type_import=p.type_import, type_label=label,
        simulation=p.simulation, resume=resume,
        fichiers_traites=fichiers_traites,
        contrats_ajoutes=ajoutes,
        contrats_modifies=modifies,
        contrats_non_trouves=non_trouves,
        contrats_migrations=migrations,
        modif_vendeurs=modif_vendeurs,
        erreurs=erreurs,
        message=(
            f"{len(files)} fichier(s) | "
            f"Ajoutés {resume.nb_ajoutes} | Modifiés {resume.nb_modifies} | "
            f"Migr. FTTB->FTTH {resume.nb_migrations} | "
            f"Modif vend. {resume.nb_modif_vend} | "
            f"Non modifiés {resume.nb_non_modifies} | "
            f"Erreurs {resume.nb_erreurs}. "
            + ("(SIMULATION)" if p.simulation else "(PRODUCTION)")
        ),
    )
    _attach_xlsx_and_mail_sfr(res, op_id)
    return res


def _build_xlsx_sfr(res: ImportSfrResult) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook(); ws = wb.active; ws.title = "Résumé"
    header_fill = PatternFill("solid", fgColor="17494E")
    header_font = Font(bold=True, color="FFFFFF")
    items = [
        ("NB Fichiers", res.resume.nb_fichiers),
        ("NB Ajoutés", res.resume.nb_ajoutes),
        ("NB Modifiés", res.resume.nb_modifies),
        ("NB Modif vendeurs", res.resume.nb_modif_vend),
        ("NB Migrations FTTB→FTTH", res.resume.nb_migrations),
        ("NB Non modifiés", res.resume.nb_non_modifies),
        ("NB Erreurs", res.resume.nb_erreurs),
        ("NB Introuvables", res.resume.nb_introuvables),
        ("NB Doublons", res.resume.nb_doublons),
        ("NB Hors délai", res.resume.nb_hors_delai),
    ]
    ws.append(["Indicateur", "Nombre"])
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for lbl, n in items:
        ws.append([lbl, n])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    for title, rows in [
        ("Ajoutés", res.contrats_ajoutes),
        ("Modifiés", res.contrats_modifies),
        ("Migrations FTTB-FTTH", res.contrats_migrations),
        ("Modif Vendeurs", res.modif_vendeurs),
        ("Erreurs", res.erreurs),
    ]:
        if not rows:
            continue
        sh = wb.create_sheet(title=title[:31])
        keys = list(rows[0].keys())
        sh.append(keys)
        for c in sh[1]:
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for r in rows:
            sh.append([str(r.get(k, "")) if r.get(k) is not None else "" for k in keys])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def _attach_xlsx_and_mail_sfr(res: ImportSfrResult, op_id: int) -> None:
    from app.shared.notifications.mail import envoi_mail
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    suffix = "_SIMU" if res.simulation else ""
    prefix_map = {
        1: "ImportSFRFibre", 2: "ImportSFRMobile", 3: "ImportSFRCall",
        4: "ImportSFRHebdo", 5: "ImportSFROptions", 6: "ImportSFRRun",
        7: "ImportSFRCallRetKO", 8: "ImportSFRCallRetRacc",
        9: "ImportSFRCallRetVADD", 10: "ImportSFRCallRetRDVTech",
    }
    xlsx_name = f"{prefix_map.get(res.type_import, 'ImportSFR')}_{ts}{suffix}.xlsx"
    try:
        xlsx_bytes = _build_xlsx_sfr(res)
    except Exception:
        return
    res.xlsx_name = xlsx_name
    res.xlsx_b64 = base64.b64encode(xlsx_bytes).decode("ascii")

    try:
        db = get_pg_connection("rh")
        r = db.query_one(
            "SELECT mail FROM rh.pgt_salarie_coordonnees WHERE id_salarie = ? LIMIT 1",
            (int(op_id),),
        )
        op_mail = (r.get("mail") if r else "") or ""
    except Exception:
        op_mail = ""
    destinataires = [op_mail] if op_mail else ["intranet@omaya.fr"]
    cc = ["intranet@omaya.fr"] if op_mail and op_mail != "intranet@omaya.fr" else []

    sujet_pref = "SIMULATION : " if res.simulation else ""
    sujet = (f"{sujet_pref}Importation {res.type_label} SFR "
             f"du {date.today().strftime('%d/%m/%Y')}")
    html = (
        "<p>Bonjour,</p>"
        f"<p>Fin importation le : {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>"
        f"<p><strong>{res.message}</strong></p>"
        "<p>Service Importation EXOSPHERE</p>"
    )
    try:
        res.mail_envoye = envoi_mail(
            sujet=sujet, html=html,
            destinataires=destinataires, cc=cc,
            expediteur="intranet@omaya.fr",
            attachments=[(xlsx_name, xlsx_bytes)],
        )
    except Exception:
        res.mail_envoye = False
