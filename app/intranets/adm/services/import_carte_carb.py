"""
Service Fen_ImportFournisseurCarte (ADM Ulease -> Import des releves
fournisseur depuis un fichier Excel).

Pour l'instant : import Total Energies. Autres fournisseurs a venir.

Pipeline ImportTotalEnergies (cf. WinDev) :
  Pour chaque ligne du fichier Excel :
    1. Lit les colonnes (CodeCarte, NumCarte, Date, Heure, Lieu, LibType,
       MontantHT, MontantTTC, IdFacturation, CompteClient).
    2. Match la carte (pgt_cartecarburant) par CodeCarte + NumCarte LIKE.
    3. Match (ou cree si pas simulation) le type de releve par Lib_Type.
    4. Verifie qu'un releve n'existe pas deja (meme fournisseur + date +
       heure) -> sinon INSERT dans pgt_cartecarbrelevefournisseur.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from app.core.database.pg import get_pg_connection


def _str(v: Any) -> str:
    return "" if v is None else str(v)


def _int(v: Any) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def _float(v: Any) -> float:
    """cf. audit : tolerant sur le format FR (virgule). Sans replace ','
    -> '.', une cellule Excel Texte '12,50' renvoyait 0.0."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except (TypeError, ValueError):
        return 0.0


# Compteur monotone pour eviter les collisions d'id sur import de masse
# (le format YYYYMMDDHHMMSSmmm ne garantit pas l'unicite si 2 appels
# tombent dans la meme ms - avere en boucle serree).
_new_id_counter = 0


def _new_id() -> int:
    """Id unique 17 chiffres : timestamp ms + compteur mod 1000.
    cf. audit : anti-collision sur import serre (INDEX UNIQUE
    id_type_releve_fournisseur, id_releve_fournisseur, ...).
    """
    global _new_id_counter
    _new_id_counter = (_new_id_counter + 1) % 1000
    base = int(datetime.now().strftime("%Y%m%d%H%M%S%f")[:14])
    # Concatene compteur 3 chiffres apres le timestamp 14 chars
    return base * 1000 + _new_id_counter


def _next_auto(db, schema: str, table: str, col: str) -> int:
    """Tables HFSQL migrees sans sequence PG : MAX(_auto)+1."""
    r = db.query_one(
        f"SELECT COALESCE(MAX({col}),0)+1 AS n FROM {schema}.{table}",
    )
    return _int(r.get("n")) if r else 1


def _col_letter_to_idx(letter: str) -> int:
    """'A'->0, 'B'->1, ..., 'Z'->25, 'AA'->26."""
    s = (letter or "").strip().upper()
    if not s:
        return -1
    n = 0
    for c in s:
        if not ("A" <= c <= "Z"):
            return -1
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n - 1


def _cell_str(row: tuple, idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float)):
        # nombre lu : on retourne sans decimales si entier
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    return str(v)


def _cell_date(row: tuple, idx: int):
    """Lit une cellule comme date. Tolerant sur les formats :
    JJ/MM/AAAA, J/M/AAAA, JJ-MM-AAAA, AAAA-MM-JJ, avec ou sans zero
    leading. Renvoie None si non parsable.
    """
    if idx < 0 or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        return v.date() if hasattr(v, "hour") else v
    s = str(v).strip()
    if not s:
        return None
    # Tente plusieurs formats connus (avec/sans zero leading)
    for fmt in (
        "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y",
        "%Y-%m-%d", "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s[:10] if fmt.endswith("Y") else s,
                                     fmt).date()
        except Exception:
            pass
    # Fallback : parsing "2/1/26" style via split
    for sep in ("/", "-"):
        if sep in s:
            parts = s.split(sep, 2)
            if len(parts) == 3:
                try:
                    a, b, c = (int(p.strip()) for p in parts)
                    # Heuristique : si le 3e est < 100, c'est yy (annee 20xx)
                    if c < 100:
                        c = 2000 + c
                    return datetime(c, b, a).date()
                except Exception:
                    pass
    return None


def _cell_time(row: tuple, idx: int):
    """Lit une cellule comme heure HH:MM ou H:MM (sans zero leading).
    Renvoie une chaine HH:MM:SS ou None.
    """
    if idx < 0 or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    if hasattr(v, "hour") and hasattr(v, "minute"):
        return v.time() if hasattr(v, "year") else v
    s = str(v).strip()
    if not s:
        return None
    # Cherche le premier ':' - permet H:MM ou HH:MM
    if ":" in s:
        try:
            hm = s.split(":", 2)
            if len(hm) >= 2:
                h = int(hm[0])
                m = int(hm[1][:2])
                if 0 <= h < 24 and 0 <= m < 60:
                    return f"{h:02d}:{m:02d}:00"
        except Exception:
            return None
    return None


def _cell_float(row: tuple, idx: int) -> float:
    """Lit une cellule comme float. Tolerant sur le format FR (virgule).
    Cf. audit : sans replace(',', '.'), une cellule '12,50' renvoyait
    0.0 silencieusement (float() plante puis except renvoie 0).
    """
    if idx < 0 or idx >= len(row):
        return 0.0
    v = row[idx]
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _format_norm(s: str) -> str:
    """ChaineFormate(..., ccSansEspace + ccMajuscule)."""
    return s.replace(" ", "").upper().strip()


def _format_no_accent_upper(s: str) -> str:
    """ChaineFormate(..., ccSansAccent + ccMajuscule)."""
    import unicodedata
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).upper().strip()


def _rollback_manual(
    db, releve_ids: list[int], type_ids: list[int],
) -> None:
    """Rollback manuel en cas d'erreur mi-import.

    L'API PGConnection.query() est autocommit -> chaque INSERT est deja
    commit avant qu'on puisse detecter une erreur ligne suivante. On
    supprime donc explicitement les enregistrements deja crees pour
    rendre l'operation atomique (best-effort : swallow exceptions).
    """
    if releve_ids:
        try:
            ids_sql = ",".join(str(int(i)) for i in releve_ids)
            db.query(
                f"""DELETE FROM ulease.pgt_cartecarbrelevefournisseur
                     WHERE id_carte_carb_releve_fournisseur IN ({ids_sql})"""
            )
        except Exception:
            pass
    if type_ids:
        try:
            ids_sql = ",".join(str(int(i)) for i in type_ids)
            db.query(
                f"""DELETE FROM ulease.pgt_typerelevefournisseur
                     WHERE id_type_releve_fournisseur IN ({ids_sql})"""
            )
        except Exception:
            pass


def import_total_energies(
    file_bytes: bytes,
    id_carte_fournisseur: int,
    ligne_depart: int,
    cols: dict,
    simulation: bool,
    op_id: int,
) -> dict:
    """Import Excel Total Energies. cols est un dict des lettres de colonne :
      { 'id_facture': 'A', 'compte_client': 'B', 'num_carte': 'C',
        'code_carte': 'D', 'date': 'E', 'heure': 'F', 'lieu': 'G',
        'lib_type': 'H', 'montant_ht': 'I', 'montant_ttc': 'J' }
    """
    if not id_carte_fournisseur:
        return {"ok": False, "error": "Fournisseur manquant"}
    if not file_bytes:
        return {"ok": False, "error": "Fichier vide"}

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        return {"ok": False, "error": f"Fichier Excel illisible : {e}"}

    idx = {k: _col_letter_to_idx(v) for k, v in cols.items()}
    needed = ("id_facture", "compte_client", "num_carte", "code_carte",
              "date", "heure", "lieu", "lib_type", "montant_ht", "montant_ttc")
    for k in needed:
        if k not in idx:
            return {"ok": False, "error": f"Colonne manquante : {k}"}

    db = get_pg_connection("ulease")
    resume: list[str] = []
    nb_ajout = 0
    nb_lus = 0

    # cf. audit : l'API query() est autocommit -> pas de transaction
    # native. On collecte les IDs des relevees inseres pour permettre
    # un rollback manuel en cas d'erreur mi-import (DELETE bulk sur
    # les IDs concernes).
    inserted_releve_ids: list[int] = []
    inserted_type_ids: list[int] = []

    # Cache des types_releve dans la session
    types_known = {
        _format_no_accent_upper(r["lib_type"]): _int(r["id_type_releve_fournisseur"])
        for r in db.query(
            """SELECT id_type_releve_fournisseur, lib_type
                 FROM ulease.pgt_typerelevefournisseur
                WHERE (modif_elem IS NULL OR modif_elem <> 'suppr')""",
        ) or []
    }

    # cf. audit : pre-calcul des _auto counters pour eviter N+1 SELECT
    # MAX() par INSERT (perf sur 5000 lignes). Incrementation en memoire.
    _cur_type_auto = _next_auto(
        db, "ulease", "pgt_typerelevefournisseur",
        "id_type_releve_fournisseur_auto",
    ) - 1
    _cur_releve_auto = _next_auto(
        db, "ulease", "pgt_cartecarbrelevefournisseur",
        "id_carte_carb_releve_fournisseur_auto",
    ) - 1

    def _incr_type_auto() -> int:
        nonlocal _cur_type_auto
        _cur_type_auto += 1
        return _cur_type_auto

    def _incr_releve_auto() -> int:
        nonlocal _cur_releve_auto
        _cur_releve_auto += 1
        return _cur_releve_auto

    start = max(2, int(ligne_depart))  # cf. audit : garde-fou >= 2
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < start:
            continue
        nb_lus += 1
        # Valeurs lues
        # cf. audit : id_facturation en varchar(50) - troncature defensive
        id_facture = _format_norm(_cell_str(row, idx["id_facture"]))[:50]
        cpte_client = _format_norm(_cell_str(row, idx["compte_client"]))
        num_carte = _format_norm(_cell_str(row, idx["num_carte"]))
        code_carte = _format_norm(_cell_str(row, idx["code_carte"]))
        date_v = _cell_date(row, idx["date"])
        heure_v = _cell_time(row, idx["heure"])
        lieu = _cell_str(row, idx["lieu"])
        lib_type_raw = _cell_str(row, idx["lib_type"])
        lib_type = _format_no_accent_upper(lib_type_raw)
        # cf. audit : _cell_float supporte format FR (virgule)
        montant_ht = _cell_float(row, idx["montant_ht"])
        montant_ttc = _cell_float(row, idx["montant_ttc"])

        # Ligne vide -> skip silencieux
        if not num_carte and not code_carte and not date_v:
            continue

        # 1. Match carte (CodeCarte + NumCarte LIKE)
        carte = db.query_one(
            """SELECT id_carte_carburant FROM ulease.pgt_cartecarburant
                WHERE code_carte = ?
                  AND num_carte LIKE ?
                  AND (modif_elem IS NULL OR modif_elem <> 'suppr')
                LIMIT 1""",
            (code_carte, num_carte + "%"),
        )
        if not carte:
            resume.append(f"Ligne {r_idx} : carte {cpte_client} {num_carte} inconnue")
            continue
        id_carte = _int(carte.get("id_carte_carburant"))

        # 2. Match (ou cree) type releve
        id_type = types_known.get(lib_type, 0)
        if id_type == 0:
            if simulation:
                resume.append(f"Ligne {r_idx} : type '{lib_type_raw}' inconnu")
            else:
                id_type = _new_id()
                try:
                    db.query(
                        """INSERT INTO ulease.pgt_typerelevefournisseur
                             (id_type_releve_fournisseur_auto,
                              id_type_releve_fournisseur, lib_type, categorie,
                              modif_op, modif_date, modif_elem)
                           VALUES (?, ?, ?, 'Non déterminée', ?, NOW(), 'new')""",
                        (_incr_type_auto(), id_type, lib_type, int(op_id)),
                    )
                    inserted_type_ids.append(id_type)
                    types_known[lib_type] = id_type
                    resume.append(f"Type '{lib_type_raw}' ajouté")
                except Exception as e:
                    resume.append(
                        f"Ligne {r_idx} : ERREUR INSERT type '{lib_type_raw}' "
                        f": {e}"
                    )
                    _rollback_manual(db, inserted_releve_ids, inserted_type_ids)
                    return {
                        "ok": False,
                        "error": f"Erreur INSERT type ligne {r_idx} : {e}",
                        "nb_lus": nb_lus, "nb_ajout": nb_ajout,
                        "resume": resume,
                        "rollback": "auto (types + releves inseres supprimes)",
                    }

        # 3. Verifie existence releve : cf. audit - dedup par (id_carte,
        # date, heure) au lieu de (id_carte_fournisseur, date, heure).
        # Evite le faux positif de 2 pleins simultanes sur 2 cartes du
        # meme fournisseur.
        if date_v and heure_v:
            exists = db.query_one(
                """SELECT id_carte_carb_releve_fournisseur
                     FROM ulease.pgt_cartecarbrelevefournisseur
                    WHERE (modif_elem IS NULL OR modif_elem <> 'suppr')
                      AND id_carte_fournisseur = ?
                      AND id_carte_carburant = ?
                      AND date = ?
                      AND heure = ?
                    LIMIT 1""",
                (int(id_carte_fournisseur), id_carte, date_v, heure_v),
            )
        else:
            exists = None
        if exists:
            continue

        # 4. INSERT releve (si pas simulation)
        if not simulation:
            id_releve = _new_id()
            try:
                db.query(
                    """INSERT INTO ulease.pgt_cartecarbrelevefournisseur
                         (id_carte_carb_releve_fournisseur_auto,
                          id_carte_carb_releve_fournisseur, id_facturation,
                          id_carte_fournisseur, id_carte_carburant,
                          id_type_releve_fournisseur, date, heure, lieu,
                          montant_ht, montant_ttc,
                          modif_op, modif_date, modif_elem)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NOW(), 'new')""",
                    (
                        _incr_releve_auto(), id_releve, id_facture,
                        int(id_carte_fournisseur),
                        id_carte, id_type, date_v, heure_v, lieu,
                        montant_ht, montant_ttc, int(op_id),
                    ),
                )
                inserted_releve_ids.append(id_releve)
            except Exception as e:
                resume.append(
                    f"Ligne {r_idx} : ERREUR INSERT releve : {e}"
                )
                _rollback_manual(db, inserted_releve_ids, inserted_type_ids)
                return {
                    "ok": False,
                    "error": f"Erreur INSERT releve ligne {r_idx} : {e}",
                    "nb_lus": nb_lus, "nb_ajout": nb_ajout,
                    "resume": resume,
                    "rollback": "auto (types + releves inseres supprimes)",
                }
            nb_ajout += 1

    resume.append(f"nb Ajout : {nb_ajout}")
    resume.append("Import Terminé")
    return {
        "ok": True,
        "simulation": simulation,
        "nb_lus": nb_lus,
        "nb_ajout": nb_ajout,
        "resume": resume,
    }
