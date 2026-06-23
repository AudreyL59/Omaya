"""
Service Fen_importFormIAG (ADM Salaries -> Suivi des formations IAG).

Import d'un fichier Excel listant les formations IAG effectuees :
  - Code IAG (colonne B) : matche pgt_salarie_partenaire.code WHERE
    id_partenaire = id du partenaire 'IAG'.
  - Date Formation (C), Nom (D), Prenom (E), Score (G).

Pipeline (cf. WinDev) :
  1. Pour chaque ligne >= 2 (header en 1), parse les valeurs.
  2. Si Date Formation == 'Aucune participation' : skip (V1).
  3. Si Date < (today - limit_jours) : skip.
  4. Cherche le salarie par CodeIAG.
  5. Si pas trouve : cherche par Nom + Prenom (LIKE avec % sur '-' et ' ').
  6. Si trouve : compare score + date avec valeurs courantes.
     - Si different : ajoute au resume + UPDATE (si pas simulation).
  7. Si pas trouve : ajoute a la liste des erreurs.

Liste 'a former' : salaries actifs avec FormationIAG=0 ou score<13
filtres sur poste (VRP/MANAGER/AGENCE/REGION/DA) + societe > 7 sauf 11
+ exclusions ID hardcodees (cf. WinDev).
"""

from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import load_workbook

from app.core.database.pg import get_pg_connection


def _str(v: Any) -> str:
    return "" if v is None else str(v)


def _int(v: Any) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def _col_letter_to_idx(letter: str) -> int:
    s = (letter or "").strip().upper()
    if not s:
        return -1
    n = 0
    for c in s:
        if not ("A" <= c <= "Z"):
            return -1
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n - 1


def _cell(row: tuple, idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _cell_str(row: tuple, idx: int) -> str:
    v = _cell(row, idx)
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _parse_date(v: Any) -> date | None:
    """Format Excel : datetime, 'AAAA.MM.JJ' ou 'JJ/MM/AAAA'."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    # AAAA.MM.JJ
    if len(s) >= 10 and s[4] == "." and s[7] == ".":
        try:
            return date(int(s[:4]), int(s[5:7]), int(s[8:10]))
        except Exception:
            return None
    # JJ/MM/AAAA
    if len(s) >= 10 and s[2] == "/" and s[5] == "/":
        try:
            return date(int(s[6:10]), int(s[3:5]), int(s[:2]))
        except Exception:
            return None
    # AAAA-MM-JJ
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        try:
            return date(int(s[:4]), int(s[5:7]), int(s[8:10]))
        except Exception:
            return None
    return None


def _id_partenaire_iag() -> int:
    db = get_pg_connection("adv")
    r = db.query_one(
        """SELECT id_partenaire FROM adv.pgt_partenaire
            WHERE UPPER(prefixe_bdd) = 'IAG'
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            LIMIT 1""",
    )
    return _int(r.get("id_partenaire")) if r else 0


# ---------------------------------------------------------------------------
# Liste 'a former' (tableau du bas dans WinDev)
# ---------------------------------------------------------------------------


def list_a_former() -> list[dict]:
    """ReqlisteFormationIAG : salaries actifs concernes (VRP/MANAGER/
    AGENCE/REGION/DA) avec FormationIAG=0 OU score<13."""
    db = get_pg_connection("rh")
    rows = db.query(
        """SELECT s.id_salarie, s.nom, s.prenom,
                  e.formation_iag, e.formation_iag_score, e.formation_iag_date,
                  e.id_ste, e.en_pause, e.date_debut,
                  tp.lib_poste
             FROM rh.pgt_salarie s
       INNER JOIN rh.pgt_salarie_embauche e
               ON e.id_salarie = s.id_salarie
       INNER JOIN rh.pgt_type_poste tp
               ON tp.id_type_poste = e.id_type_poste
            WHERE (s.modif_elem IS NULL OR s.modif_elem NOT LIKE '%suppr%')
              AND COALESCE(e.en_activite, FALSE) = TRUE
              AND (
                    COALESCE(e.formation_iag, FALSE) = FALSE
                 OR (COALESCE(e.formation_iag, FALSE) = TRUE
                     AND COALESCE(e.formation_iag_score, 0) < 13)
                  )
              AND (
                    tp.lib_poste ILIKE '%VRP%'
                 OR tp.lib_poste ILIKE '%MANAGER%'
                 OR tp.lib_poste ILIKE '%AGENCE%'
                 OR tp.lib_poste ILIKE '%REGION%'
                 OR tp.lib_poste ILIKE '%DA%'
                  )
              AND e.id_ste > 7
              AND e.id_ste <> 11
              AND s.id_salarie NOT IN (20190827110730867, 20200715153948361)
         ORDER BY s.nom ASC, s.prenom ASC""",
    ) or []
    out = []
    for r in rows:
        prenom = _str(r.get("prenom")).strip()
        if prenom:
            prenom = prenom[:1].upper() + prenom[1:].lower()
        out.append({
            "id_salarie": str(_int(r.get("id_salarie"))),
            "nom": _str(r.get("nom")),
            "prenom": prenom,
            "formation_iag": bool(r.get("formation_iag")),
            "formation_iag_score": _int(r.get("formation_iag_score")),
            "formation_iag_date": _str(r.get("formation_iag_date"))[:10],
            "id_ste": _int(r.get("id_ste")),
            "en_pause": bool(r.get("en_pause")),
            "lib_poste": _str(r.get("lib_poste")),
            "date_debut": _str(r.get("date_debut"))[:10],
        })
    return out


# ---------------------------------------------------------------------------
# Recherche salarie (Code IAG -> Nom/Prenom)
# ---------------------------------------------------------------------------


def _find_by_code_iag(code: str, id_part_iag: int) -> list[int]:
    """Retourne les id_salarie associes au code IAG dans
    pgt_salarie_partenaire."""
    if not code or not id_part_iag:
        return []
    db = get_pg_connection("rh")
    rows = db.query(
        """SELECT id_salarie FROM rh.pgt_salarie_partenaire
            WHERE id_partenaire = ?
              AND code = ?
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
        (int(id_part_iag), code),
    ) or []
    return [_int(r.get("id_salarie")) for r in rows if _int(r.get("id_salarie"))]


def _find_by_nom_prenom(nom_pat: str, prenom_pat: str) -> list[int]:
    """Cherche par nom + prenom avec patterns LIKE (% remplace - et ' ')."""
    if not nom_pat or not prenom_pat:
        return []
    db = get_pg_connection("rh")
    rows = db.query(
        """SELECT id_salarie FROM rh.pgt_salarie
            WHERE UPPER(nom) LIKE ?
              AND UPPER(prenom) LIKE ?
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
        (nom_pat, prenom_pat),
    ) or []
    return [_int(r.get("id_salarie")) for r in rows if _int(r.get("id_salarie"))]


def _info_embauche(id_salarie: int) -> dict:
    """Retourne {nom, prenom, formation_iag, formation_iag_date,
    formation_iag_score} pour comparaison + UPDATE."""
    db = get_pg_connection("rh")
    r = db.query_one(
        """SELECT s.nom, s.prenom,
                  e.formation_iag, e.formation_iag_date, e.formation_iag_score
             FROM rh.pgt_salarie s
        LEFT JOIN rh.pgt_salarie_embauche e
               ON e.id_salarie = s.id_salarie
              AND (e.modif_elem IS NULL OR e.modif_elem NOT LIKE '%suppr%')
              AND COALESCE(e.en_activite, FALSE) = TRUE
            WHERE s.id_salarie = ? LIMIT 1""",
        (int(id_salarie),),
    )
    return r or {}


# ---------------------------------------------------------------------------
# Import principal
# ---------------------------------------------------------------------------


def import_formations(
    file_bytes: bytes,
    cols: dict,
    limit_jours: int,
    simulation: bool,
    op_id: int,
) -> dict:
    """Import Excel formations IAG. cols = {code_iag, date, nom, prenom,
    score} chaque valeur etant une lettre de colonne ('B', 'C', ...).

    Retourne {ok, simulation, nb_lus, nb_maj, resume:[lignes],
    vendeurs_erreur:[{nom, prenom, date_formation, score, nb_fiche}]}."""
    if not file_bytes:
        return {"ok": False, "error": "Fichier vide"}

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        return {"ok": False, "error": f"Fichier Excel illisible : {e}"}

    idx = {k: _col_letter_to_idx(v) for k, v in cols.items()}
    for k in ("code_iag", "date", "nom", "prenom", "score"):
        if k not in idx or idx[k] < 0:
            return {"ok": False, "error": f"Colonne manquante : {k}"}

    id_part_iag = _id_partenaire_iag()
    if not id_part_iag:
        return {"ok": False, "error": "Partenaire IAG introuvable"}

    today = date.today()
    limit_date = today - timedelta(days=int(limit_jours)) if limit_jours > 0 else None
    db_rh = get_pg_connection("rh")
    resume: list[str] = []
    nb_lus = 0
    nb_maj = 0
    erreurs: list[dict] = []

    resume.append(
        f"Début importation : {today.strftime('%d/%m/%Y')} à "
        f"{datetime.now().strftime('%H:%M')}"
    )

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < 2:
            continue  # header
        nb_lus += 1
        code = _cell_str(row, idx["code_iag"])
        date_raw = _cell(row, idx["date"])
        nom_raw = _cell_str(row, idx["nom"]).upper()
        prenom_raw = _cell_str(row, idx["prenom"]).upper()
        score_raw = _cell_str(row, idx["score"])
        try:
            score = int(float(score_raw)) if score_raw else 0
        except (TypeError, ValueError):
            score = 0

        # Cas 'Aucune participation' -> skip V1
        if isinstance(date_raw, str) and "aucune" in date_raw.lower():
            continue
        d = _parse_date(date_raw)
        if not d:
            continue
        # Filtre par 'limiter_aux N derniers jours'
        if limit_date and d < limit_date:
            continue

        # 1. Recherche par CodeIAG
        ids = _find_by_code_iag(code, id_part_iag)
        # 2. Si pas trouve : par Nom + Prenom
        if len(ids) != 1:
            # Nettoyage des noms : MAJUSCULE + - et ' ' -> %
            nom_pat = nom_raw.replace("-", "%").replace(" ", "%") + "%"
            prenom_pat = prenom_raw.replace("-", "%").replace(" ", "%") + "%"
            ids = _find_by_nom_prenom(nom_pat, prenom_pat)

        if len(ids) != 1:
            erreurs.append({
                "nom": _cell_str(row, idx["nom"]),
                "prenom": _cell_str(row, idx["prenom"]),
                "date_formation": d.strftime("%Y-%m-%d"),
                "score": score,
                "nb_fiche": len(ids),
                "code_iag": code,
            })
            continue

        id_s = ids[0]
        info = _info_embauche(id_s)
        cur_date = info.get("formation_iag_date")
        if hasattr(cur_date, "strftime"):
            cur_date_d: date | None = (
                cur_date if isinstance(cur_date, date) and not isinstance(cur_date, datetime)
                else cur_date.date() if isinstance(cur_date, datetime)
                else None
            )
        else:
            cur_date_d = None
        cur_score = _int(info.get("formation_iag_score"))

        if cur_score != score or cur_date_d != d:
            nb_maj += 1
            resume.append(
                f"{_str(info.get('nom'))} {_str(info.get('prenom'))} : "
                f"{score} pt(s) le {d.strftime('%d/%m/%Y')}"
            )
            if not simulation:
                db_rh.query(
                    """UPDATE rh.pgt_salarie_embauche
                          SET formation_iag = TRUE,
                              formation_iag_date = ?,
                              formation_iag_score = ?,
                              modif_date = NOW(),
                              modif_op = ?,
                              modif_elem = 'modif'
                        WHERE id_salarie = ?
                          AND (modif_elem IS NULL
                               OR modif_elem NOT LIKE '%suppr%')
                          AND COALESCE(en_activite, FALSE) = TRUE""",
                    (d, score, int(op_id), id_s),
                )

    resume.append("Fin d'importation")
    return {
        "ok": True,
        "simulation": simulation,
        "nb_lus": nb_lus,
        "nb_maj": nb_maj,
        "resume": resume,
        "vendeurs_erreur": erreurs,
    }
