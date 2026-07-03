"""Service Fen_AjoutColonneImport (ADM Imports Bases -> Ajout colonne import).

Permet d'enrichir un fichier Excel avec des colonnes calculees depuis la BDD :
  - Vendeur / Agence / Equipe   <- IMPLEMENTE
  - Date de Signature
  - Lib Produit
  - Etat Omaya
  - Info Client
  - CAR (ENI uniquement)
  - Infos RUN et REM SFR

Le fichier d'entree doit avoir une colonne 'num_contrat' (defaut B).
Le mode partenaire est soit :
  - 'liste'  : tous les contrats sont du meme partenaire (combo)
  - 'colonne': le partenaire est dans une colonne du fichier
"""

from __future__ import annotations

import base64
import io
from datetime import date, datetime
from typing import Optional

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection
from app.shared.recrutement.services.recherche_cv import _int, _str
from app.intranets.adm.services.import_eni import _col_letter_to_index, _cell


PARTENAIRES_SUPPORTES = ["eni", "iag", "oen", "pro", "sfr", "str", "val"]


class AjoutColonneParams(BaseModel):
    col_num_contrat: str = "B"
    mode_partenaire: str = "liste"      # 'liste' ou 'colonne'
    partenaire: str = ""                # si mode='liste' : prefixe BDD
    col_partenaire: str = ""            # si mode='colonne' : colonne du fichier
    action: str = "vendeur_agence_equipe"


class AjoutColonneResult(BaseModel):
    ok: bool
    nb_lignes_traitees: int = 0
    nb_lignes_enrichies: int = 0
    xlsx_b64: str = ""
    xlsx_name: str = ""
    message: str = ""


def _lookup_contrat_basic(partenaire: str, num_bs: str) -> Optional[dict]:
    """Lookup id_salarie + date_signature + id_etat + id_produit basics."""
    p = partenaire.lower()
    if p not in PARTENAIRES_SUPPORTES:
        return None
    try:
        db = get_pg_connection("adv")
        return db.query_one(
            f"""SELECT id_contrat, id_salarie, id_client, id_produit,
                       id_etat_contrat, date_signature
                  FROM adv.pgt_{p}_contrat
                 WHERE UPPER(num_bs) = UPPER(?)
                   AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                 LIMIT 1""",
            (num_bs,),
        )
    except Exception:
        return None


def _info_salarie(id_salarie: int) -> dict:
    if not id_salarie:
        return {}
    try:
        db = get_pg_connection("rh")
        r = db.query_one(
            "SELECT nom, prenom FROM rh.pgt_salarie WHERE id_salarie = ? LIMIT 1",
            (int(id_salarie),),
        )
        return r or {}
    except Exception:
        return {}


def _affectation(id_salarie: int, date_ref: str = "") -> tuple[str, str]:
    """(agence, equipe) a la date `date_ref` (YYYY-MM-DD ou vide).

    cf. WinDev affectationVendeurByDate(idVend, DateSig) : retourne
    l'affectation en vigueur a la date de signature (pas l'actuelle).
    Un vendeur mute apres signature avait son ancienne agence/equipe
    sur les contrats historiques - critique pour l'export.

    Si date_ref est vide, prend l'affectation actuelle (fallback WinDev).
    """
    if not id_salarie:
        return ("", "")
    try:
        db = get_pg_connection("rh")
        # Filtre les affectations valides a la date_ref si fournie
        # (date_debut <= date_ref <= date_fin OU date_fin NULL)
        if date_ref:
            rows = db.query(
                """SELECT o.lib_orga, o.id_type_niveau_orga
                     FROM rh.pgt_salarie_organigramme so
                     JOIN rh.pgt_organigramme o
                          ON o.idorganigramme = so.idorganigramme
                    WHERE so.id_salarie = ?
                      AND (so.modif_elem IS NULL
                           OR so.modif_elem NOT LIKE '%suppr%')
                      AND (so.date_debut IS NULL OR so.date_debut <= ?)
                      AND (so.date_fin IS NULL OR so.date_fin >= ?)""",
                (int(id_salarie), date_ref, date_ref),
            ) or []
        else:
            rows = db.query(
                """SELECT o.lib_orga, o.id_type_niveau_orga
                     FROM rh.pgt_salarie_organigramme so
                     JOIN rh.pgt_organigramme o
                          ON o.idorganigramme = so.idorganigramme
                    WHERE so.id_salarie = ?
                      AND (so.modif_elem IS NULL
                           OR so.modif_elem NOT LIKE '%suppr%')""",
                (int(id_salarie),),
            ) or []
        agence = ""; equipe = ""
        for r in rows:
            lvl = r.get("id_type_niveau_orga")
            lib = r.get("lib_orga") or ""
            if lvl == 3 and not agence:
                agence = lib
            elif lvl == 4 and not equipe:
                equipe = lib
        return (agence, equipe)
    except Exception:
        return ("", "")


def _read_and_copy_workbook(content: bytes) -> tuple:
    """Charge le XLSX d'entree et retourne (in_wb, out_wb, in_ws, out_ws,
    nb_cols)."""
    from openpyxl import load_workbook, Workbook
    in_wb = load_workbook(io.BytesIO(content), data_only=True)
    in_ws = in_wb.active
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = in_ws.title or "Sheet1"
    nb_cols = in_ws.max_column or 0
    # Copie header + toutes les lignes
    for row in in_ws.iter_rows(values_only=True):
        out_ws.append(list(row))
    return (in_wb, out_wb, in_ws, out_ws, nb_cols)


def _save_workbook_b64(wb, name_prefix: str = "AjoutCol") -> tuple[str, str]:
    """Sauve un Workbook en bytes b64 + retourne (b64, name)."""
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    xlsx_name = f"Modif_{name_prefix}_{ts}.xlsx"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return (base64.b64encode(buf.read()).decode("ascii"), xlsx_name)


def add_vendeur_agence_equipe(
    p: AjoutColonneParams, content: bytes,
) -> AjoutColonneResult:
    """Bouton 1 : ajoute 4 colonnes Nom Vendeur / Prenom Vendeur / Agence /
    Equipe a la fin du fichier."""
    try:
        in_wb, out_wb, in_ws, out_ws, nb_cols = _read_and_copy_workbook(content)
    except Exception as e:
        return AjoutColonneResult(ok=False, message=f"Lecture XLSX : {e}")

    col_num = _col_letter_to_index(p.col_num_contrat)
    col_part = _col_letter_to_index(p.col_partenaire) if p.col_partenaire else 0

    # Headers ajoutes
    out_ws.cell(row=1, column=nb_cols + 1, value="Nom Vendeur")
    out_ws.cell(row=1, column=nb_cols + 2, value="Prénom Vendeur")
    out_ws.cell(row=1, column=nb_cols + 3, value="Agence")
    out_ws.cell(row=1, column=nb_cols + 4, value="Equipe")

    nb_traitees = 0; nb_enrichies = 0
    for i in range(2, (in_ws.max_row or 0) + 1):
        num_ctt = _cell(in_ws, i, col_num).strip()
        if not num_ctt:
            continue
        nb_traitees += 1
        if p.mode_partenaire == "colonne" and col_part:
            partenaire = _cell(in_ws, i, col_part).strip().lower()
        else:
            partenaire = p.partenaire.lower()
        if partenaire not in PARTENAIRES_SUPPORTES:
            continue
        ctt = _lookup_contrat_basic(partenaire, num_ctt)
        if not ctt:
            continue
        id_sal = int(ctt.get("id_salarie") or 0)
        sal = _info_salarie(id_sal)
        # cf. WinDev affectationVendeurByDate(idVend, DateSig) : passe
        # la date de signature pour retrouver l'agence/equipe historique
        # (evite d'ecrire l'affectation actuelle si le vendeur a mute).
        date_sig = ctt.get("date_signature")
        date_ref = str(date_sig)[:10] if date_sig else ""
        agence, equipe = _affectation(id_sal, date_ref)
        if sal:
            out_ws.cell(row=i, column=nb_cols + 1, value=sal.get("nom") or "")
            out_ws.cell(row=i, column=nb_cols + 2, value=sal.get("prenom") or "")
            out_ws.cell(row=i, column=nb_cols + 3, value=agence)
            out_ws.cell(row=i, column=nb_cols + 4, value=equipe)
            nb_enrichies += 1

    in_wb.close()
    b64, name = _save_workbook_b64(out_wb, "Vendeur")
    return AjoutColonneResult(
        ok=True, nb_lignes_traitees=nb_traitees,
        nb_lignes_enrichies=nb_enrichies, xlsx_b64=b64, xlsx_name=name,
        message=(f"{nb_traitees} ligne(s) traitée(s), "
                 f"{nb_enrichies} enrichie(s) avec vendeur/agence/equipe."),
    )


def add_date_signature(
    p: AjoutColonneParams, content: bytes,
) -> AjoutColonneResult:
    """Bouton 2 : ajoute 1 colonne Date Signature."""
    try:
        in_wb, out_wb, in_ws, out_ws, nb_cols = _read_and_copy_workbook(content)
    except Exception as e:
        return AjoutColonneResult(ok=False, message=f"Lecture XLSX : {e}")
    col_num = _col_letter_to_index(p.col_num_contrat)
    col_part = _col_letter_to_index(p.col_partenaire) if p.col_partenaire else 0
    out_ws.cell(row=1, column=nb_cols + 1, value="Date Signature")
    nb_t = 0; nb_e = 0
    for i in range(2, (in_ws.max_row or 0) + 1):
        num_ctt = _cell(in_ws, i, col_num).strip()
        if not num_ctt:
            continue
        nb_t += 1
        partenaire = (_cell(in_ws, i, col_part).strip().lower()
                      if p.mode_partenaire == "colonne" and col_part
                      else p.partenaire.lower())
        if partenaire not in PARTENAIRES_SUPPORTES:
            continue
        ctt = _lookup_contrat_basic(partenaire, num_ctt)
        if ctt and ctt.get("date_signature"):
            # cf. WinDev DateVersChaine(..., 'JJ/MM/AAAA') : format DD/MM/YYYY
            # (Python renvoyait YYYY-MM-DD ISO -> livre visuellement different).
            date_v = ctt["date_signature"]
            try:
                iso = str(date_v)[:10]
                if len(iso) == 10 and iso[4] == "-" and iso[7] == "-":
                    fr = f"{iso[8:10]}/{iso[5:7]}/{iso[0:4]}"
                else:
                    fr = str(date_v)
            except Exception:
                fr = str(date_v)
            cell = out_ws.cell(row=i, column=nb_cols + 1, value=fr)
            cell.number_format = "DD/MM/YYYY"
            nb_e += 1
    in_wb.close()
    b64, name = _save_workbook_b64(out_wb, "DateSign")
    return AjoutColonneResult(ok=True, nb_lignes_traitees=nb_t,
                              nb_lignes_enrichies=nb_e, xlsx_b64=b64,
                              xlsx_name=name,
                              message=f"{nb_t} lignes, {nb_e} dates ajoutées.")


def add_lib_produit(
    p: AjoutColonneParams, content: bytes,
) -> AjoutColonneResult:
    """Bouton 3 : ajoute 1 colonne Lib Produit."""
    try:
        in_wb, out_wb, in_ws, out_ws, nb_cols = _read_and_copy_workbook(content)
    except Exception as e:
        return AjoutColonneResult(ok=False, message=f"Lecture XLSX : {e}")
    col_num = _col_letter_to_index(p.col_num_contrat)
    col_part = _col_letter_to_index(p.col_partenaire) if p.col_partenaire else 0
    out_ws.cell(row=1, column=nb_cols + 1, value="Lib Produit")
    db = get_pg_connection("adv")
    nb_t = 0; nb_e = 0
    for i in range(2, (in_ws.max_row or 0) + 1):
        num_ctt = _cell(in_ws, i, col_num).strip()
        if not num_ctt:
            continue
        nb_t += 1
        partenaire = (_cell(in_ws, i, col_part).strip().lower()
                      if p.mode_partenaire == "colonne" and col_part
                      else p.partenaire.lower())
        if partenaire not in PARTENAIRES_SUPPORTES:
            continue
        ctt = _lookup_contrat_basic(partenaire, num_ctt)
        if not ctt:
            continue
        id_prod = int(ctt.get("id_produit") or 0)
        if id_prod:
            r = db.query_one(
                f"SELECT lib_produit FROM adv.pgt_{partenaire}_produit "
                f"WHERE id_produit = ? LIMIT 1", (id_prod,),
            )
            if r:
                out_ws.cell(row=i, column=nb_cols + 1,
                            value=r.get("lib_produit") or "")
                nb_e += 1
    in_wb.close()
    b64, name = _save_workbook_b64(out_wb, "LibProduit")
    return AjoutColonneResult(ok=True, nb_lignes_traitees=nb_t,
                              nb_lignes_enrichies=nb_e, xlsx_b64=b64,
                              xlsx_name=name,
                              message=f"{nb_t} lignes, {nb_e} produits ajoutés.")


def add_etat_omaya(
    p: AjoutColonneParams, content: bytes,
) -> AjoutColonneResult:
    """Bouton 4 : ajoute 2 colonnes 'Type Etat' + 'Lib Etat' Omaya."""
    try:
        in_wb, out_wb, in_ws, out_ws, nb_cols = _read_and_copy_workbook(content)
    except Exception as e:
        return AjoutColonneResult(ok=False, message=f"Lecture XLSX : {e}")
    col_num = _col_letter_to_index(p.col_num_contrat)
    col_part = _col_letter_to_index(p.col_partenaire) if p.col_partenaire else 0
    out_ws.cell(row=1, column=nb_cols + 1, value="Type Etat")
    out_ws.cell(row=1, column=nb_cols + 2, value="Lib Etat")
    db = get_pg_connection("adv")
    nb_t = 0; nb_e = 0
    for i in range(2, (in_ws.max_row or 0) + 1):
        num_ctt = _cell(in_ws, i, col_num).strip()
        if not num_ctt:
            continue
        nb_t += 1
        partenaire = (_cell(in_ws, i, col_part).strip().lower()
                      if p.mode_partenaire == "colonne" and col_part
                      else p.partenaire.lower())
        if partenaire not in PARTENAIRES_SUPPORTES:
            continue
        ctt = _lookup_contrat_basic(partenaire, num_ctt)
        if not ctt:
            continue
        id_etat = int(ctt.get("id_etat_contrat") or 0)
        if id_etat:
            r = db.query_one(
                f"""SELECT e.lib_etat, t.lib_type
                      FROM adv.pgt_{partenaire}_etat_contrat e
                      LEFT JOIN adv.pgt_type_etat_contrat t
                             ON t.id_type_etat = e.id_type_etat
                     WHERE e.id_etat = ? LIMIT 1""", (id_etat,),
            )
            if r:
                out_ws.cell(row=i, column=nb_cols + 1,
                            value=r.get("lib_type") or "")
                out_ws.cell(row=i, column=nb_cols + 2,
                            value=r.get("lib_etat") or "")
                nb_e += 1
    in_wb.close()
    b64, name = _save_workbook_b64(out_wb, "EtatOmaya")
    return AjoutColonneResult(ok=True, nb_lignes_traitees=nb_t,
                              nb_lignes_enrichies=nb_e, xlsx_b64=b64,
                              xlsx_name=name,
                              message=f"{nb_t} lignes, {nb_e} états ajoutés.")


def add_info_client(
    p: AjoutColonneParams, content: bytes,
) -> AjoutColonneResult:
    """Bouton 5 : ajoute Nom/Prenom/CP/Ville/Mail/GSM client."""
    try:
        in_wb, out_wb, in_ws, out_ws, nb_cols = _read_and_copy_workbook(content)
    except Exception as e:
        return AjoutColonneResult(ok=False, message=f"Lecture XLSX : {e}")
    col_num = _col_letter_to_index(p.col_num_contrat)
    col_part = _col_letter_to_index(p.col_partenaire) if p.col_partenaire else 0
    # Cf WinDev : Nom Client (concat Nom+Prenom), Mail, GSM, CP, Ville
    headers = ["Nom Client", "Mail", "GSM", "CP", "Ville"]
    for j, h in enumerate(headers, start=1):
        out_ws.cell(row=1, column=nb_cols + j, value=h)
    db = get_pg_connection("adv")
    nb_t = 0; nb_e = 0
    for i in range(2, (in_ws.max_row or 0) + 1):
        num_ctt = _cell(in_ws, i, col_num).strip()
        if not num_ctt:
            continue
        nb_t += 1
        partenaire = (_cell(in_ws, i, col_part).strip().lower()
                      if p.mode_partenaire == "colonne" and col_part
                      else p.partenaire.lower())
        if partenaire not in PARTENAIRES_SUPPORTES:
            continue
        ctt = _lookup_contrat_basic(partenaire, num_ctt)
        if not ctt:
            continue
        id_clt = int(ctt.get("id_client") or 0)
        if id_clt:
            cli = db.query_one(
                """SELECT nom, prenom, cp, ville, mail, gsm
                     FROM adv.pgt_client WHERE id_client = ? LIMIT 1""",
                (id_clt,),
            )
            if cli:
                nom_complet = f"{cli.get('nom') or ''} {(cli.get('prenom') or '').title()}".strip()
                out_ws.cell(row=i, column=nb_cols + 1, value=nom_complet)
                out_ws.cell(row=i, column=nb_cols + 2, value=cli.get("mail") or "")
                out_ws.cell(row=i, column=nb_cols + 3, value=cli.get("gsm") or "")
                out_ws.cell(row=i, column=nb_cols + 4, value=cli.get("cp") or "")
                out_ws.cell(row=i, column=nb_cols + 5, value=cli.get("ville") or "")
                nb_e += 1
    in_wb.close()
    b64, name = _save_workbook_b64(out_wb, "InfoClient")
    return AjoutColonneResult(ok=True, nb_lignes_traitees=nb_t,
                              nb_lignes_enrichies=nb_e, xlsx_b64=b64,
                              xlsx_name=name,
                              message=f"{nb_t} lignes, {nb_e} clients ajoutés.")


def add_car_eni(
    p: AjoutColonneParams, content: bytes,
) -> AjoutColonneResult:
    """Bouton 6 : ajoute 'CAR Relevée' (ENI uniquement, sur GazCarRelevee)."""
    try:
        in_wb, out_wb, in_ws, out_ws, nb_cols = _read_and_copy_workbook(content)
    except Exception as e:
        return AjoutColonneResult(ok=False, message=f"Lecture XLSX : {e}")
    col_num = _col_letter_to_index(p.col_num_contrat)
    # Cf WinDev : "CAR OMAYA Relevée" + "CAR OMAYA Déclarée"
    out_ws.cell(row=1, column=nb_cols + 1, value="CAR OMAYA Relevée")
    out_ws.cell(row=1, column=nb_cols + 2, value="CAR OMAYA Déclarée")
    db = get_pg_connection("adv")
    col_part = _col_letter_to_index(p.col_partenaire) if p.col_partenaire else 0
    nb_t = 0; nb_e = 0
    for i in range(2, (in_ws.max_row or 0) + 1):
        num_ctt = _cell(in_ws, i, col_num).strip()
        if not num_ctt:
            continue
        nb_t += 1
        # cf. WinDev L44 : si CelPart <> 'ENI' alors on skip.
        partenaire = (_cell(in_ws, i, col_part).strip().lower()
                      if p.mode_partenaire == "colonne" and col_part
                      else p.partenaire.lower())
        if partenaire != "eni":
            continue
        try:
            r = db.query_one(
                """SELECT gaz_car_relevee, gaz_car_declaree
                     FROM adv.pgt_eni_contrat
                    WHERE UPPER(num_bs) = UPPER(?)
                      AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                    LIMIT 1""", (num_ctt,),
            )
        except Exception:
            r = None
        if r:
            # cf. audit : ecrire la valeur brute (peut contenir decimales)
            # au lieu de _int() qui tronque.
            out_ws.cell(row=i, column=nb_cols + 1,
                        value=r.get("gaz_car_relevee"))
            out_ws.cell(row=i, column=nb_cols + 2,
                        value=r.get("gaz_car_declaree"))
            nb_e += 1
    in_wb.close()
    b64, name = _save_workbook_b64(out_wb, "CAR_ENI")
    return AjoutColonneResult(ok=True, nb_lignes_traitees=nb_t,
                              nb_lignes_enrichies=nb_e, xlsx_b64=b64,
                              xlsx_name=name,
                              message=f"{nb_t} lignes, {nb_e} CAR ENI ajoutées.")


def add_infos_run_rem_sfr(
    p: AjoutColonneParams, content: bytes,
) -> AjoutColonneResult:
    """Bouton 7 : SFR uniquement. Itere les remunerations dans
    pgt_sfr_contrat_remun. Ajoute Va RUN / Mois P Va RUN / Ra RUN /
    Mois P Ra RUN (cf WinDev)."""
    try:
        in_wb, out_wb, in_ws, out_ws, nb_cols = _read_and_copy_workbook(content)
    except Exception as e:
        return AjoutColonneResult(ok=False, message=f"Lecture XLSX : {e}")
    col_num = _col_letter_to_index(p.col_num_contrat)
    col_part = _col_letter_to_index(p.col_partenaire) if p.col_partenaire else 0
    out_ws.cell(row=1, column=nb_cols + 1, value="Va RUN")
    out_ws.cell(row=1, column=nb_cols + 2, value="Mois P Va RUN")
    out_ws.cell(row=1, column=nb_cols + 3, value="Ra RUN")
    out_ws.cell(row=1, column=nb_cols + 4, value="Mois P Ra RUN")
    db = get_pg_connection("adv")
    nb_t = 0; nb_e = 0
    for i in range(2, (in_ws.max_row or 0) + 1):
        num_ctt = _cell(in_ws, i, col_num).strip()
        if not num_ctt:
            continue
        nb_t += 1
        # cf. WinDev : Run + REM SFR uniquement pour partenaire SFR.
        partenaire = (_cell(in_ws, i, col_part).strip().lower()
                      if p.mode_partenaire == "colonne" and col_part
                      else p.partenaire.lower())
        if partenaire != "sfr":
            continue
        try:
            rows = db.query(
                """SELECT validation, va_mois_p, raccordement, ra_mois_p
                     FROM adv.pgt_sfr_contrat_remun
                    WHERE UPPER(num) = UPPER(?)
                      AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
                (num_ctt,),
            ) or []
        except Exception:
            rows = []
        if not rows:
            continue
        for r in rows:
            if r.get("validation"):
                out_ws.cell(row=i, column=nb_cols + 1, value=True)
                out_ws.cell(row=i, column=nb_cols + 2,
                            value=str(r.get("va_mois_p") or ""))
            if r.get("raccordement"):
                out_ws.cell(row=i, column=nb_cols + 3, value=True)
                out_ws.cell(row=i, column=nb_cols + 4,
                            value=str(r.get("ra_mois_p") or ""))
        nb_e += 1
    in_wb.close()
    b64, name = _save_workbook_b64(out_wb, "RunRemSFR")
    return AjoutColonneResult(ok=True, nb_lignes_traitees=nb_t,
                              nb_lignes_enrichies=nb_e, xlsx_b64=b64,
                              xlsx_name=name,
                              message=f"{nb_t} lignes, {nb_e} run/rem SFR ajoutés.")


def dispatch(p: AjoutColonneParams, content: bytes) -> AjoutColonneResult:
    """Dispatcher selon p.action."""
    actions = {
        "vendeur_agence_equipe": add_vendeur_agence_equipe,
        "date_signature": add_date_signature,
        "lib_produit": add_lib_produit,
        "etat_omaya": add_etat_omaya,
        "info_client": add_info_client,
        "car_eni": add_car_eni,
        "infos_run_rem_sfr": add_infos_run_rem_sfr,
    }
    fn = actions.get(p.action)
    if not fn:
        return AjoutColonneResult(ok=False,
                                  message=f"Action inconnue : {p.action}")
    return fn(p, content)
