"""Service Fen_ImportENI (ADM Imports Bases -> Import partenaire ENI/PLENITUDE).

5 types d'import :
 1. Base Journaliere CALL  -> ImportJournalier()
 2. RUN Valides             -> importRUNValides()
 3. RUN Resils              -> importRUNResil()
 4. Salesforce              -> ImportSalesforce()
 5. Base journaliere ENI    -> ImportJournalierENI()

Pour chaque type, lecture d'un Excel + analyse + 6 tables resultats :
 - Resume importation (compteurs)
 - Contrats modifies
 - Contrats non trouves
 - Contrats RUN
 - Probleme Vendeur
 - Contrat Import Journ ENI

Les procedures metier seront codees au fur et a mesure (placeholder
pour l'instant : retourne un resume vide + erreur 'pas encore code').
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from typing import Optional

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection

logger = logging.getLogger(__name__)


class ImportEniParams(BaseModel):
    type_import: int                  # 1..5
    simulation: bool = True
    periode1_du: str = ""
    periode1_au: str = ""
    periode1_mois_paiement: str = ""  # 'MM-YYYY'
    periode2_du: str = ""
    periode2_au: str = ""
    periode2_mois_paiement: str = ""
    mois_paiement_distrib: str = ""
    maj_produit_contrat_stand: bool = True
    maj_etats_contrats_existants: bool = False


# Mapping colonnes par defaut pour 'Base Journaliere ENI'
# (cf. ecran onglet 'Base Journ. ENI').
COLS_BJ_ENI = {
    "num_contrat": "A",
    "vendeur_nom": "D",
    "date_signature": "E",
    "type_offre": "F",
    "statut": "G",
    "offre": "I",
    "puissance": "J",
    "type_comptage": "K",
    "car": "L",
    "client_cp": "S",
    "client_gsm": "T",
    "date_heure_crea": "W",
    "addon": "AF",
    "opt_mail": "AH",
    "mandat_rib": "AI",
    "note_globale": "AJ",
    "info_notation": "AK",
}


# Mapping colonnes par defaut pour 'RUN Resils' (cf. ecran onglet
# 'RUN Resils').
COLS_RUN_R = {
    "num_contrat": "A",
    "vendeur_nom": "G",
    "type_offre": "H",
    "date_signature": "L",
}


# Mapping colonnes par defaut pour 'RUN Valides' (cf. ecran Fen_ImportENI
# onglet 'RUN Valide'). Les lettres correspondent aux colonnes Excel.
COLS_RUN_V = {
    "num_contrat": "A",
    "vendeur_nom": "G",
    "type_offre": "H",
    "offre_reclassifiee": "I",
    "statut_elec": "J",
    "statut_gaz": "K",
    "date_signature": "M",
    "car": "P",
    "puissance": "Q",
    "type_comptage": "R",
    "mail_fourni": "AB",
    # Cf. WinDev importRUNValides (TXT 2026-07 : ligne 53-54) : ce qui
    # etait 'Protection' est devenu 'HomeServe' (map vers opt_entretien
    # en DB), + nouvelle colonne 'Mandat' (map vers pgt_eni_contrat.opt_mandat).
    # NB : la colonne opt_mandat doit exister sur pgt_eni_contrat (migration
    # SQL fournie separement), sinon comparaison desactivee via fallback.
    "home_serve": "AC",
    "mandat": "AD",
}


# Mapping colonnes Excel par defaut pour 'Base Journaliere CALL'
# (cf. ecran Fen_ImportENI). Les lettres correspondent aux colonnes Excel
# (A=1, B=2, ...). Sera surchargee via parametres BDD au fil de l'eau.
COLS_BJ_CALL = {
    "num_contrat": "A",
    "date_signature": "B",
    "vendeur_nom": "C",
    "client_nom": "D",
    "client_prenom": "E",
    "client_adresse": "F",
    "client_cplt": "G",
    "client_cp": "H",
    "client_ville": "I",
    "client_tel": "J",
    "client_gsm": "K",
    "client_mail": "L",
    "car": "M",
    "offre": "M",  # cf WinDev (doublon CAR/Offre M)
    "delai_retract": "N",
    "lib_statut": "S",
    "comment": "R",
    "heure": "U",
    "indice": "V",
    "date_naissance": "W",
    "serv_entretien": "X",
    "code_enr": "Y",
    "puissance": "Z",
}


def _col_letter_to_index(letter: str) -> int:
    """A->1, B->2, ..., Z->26, AA->27, etc."""
    letter = (letter or "").strip().upper()
    if not letter:
        return 0
    n = 0
    for c in letter:
        n = n * 26 + (ord(c) - 64)
    return n


class ImportEniResume(BaseModel):
    nb_valides: int = 0
    nb_deja_statues: int = 0
    nb_doublons: int = 0
    nb_introuvables: int = 0
    nb_decommissions: int = 0
    nb_resilies: int = 0
    nb_modifications: int = 0
    nb_erreurs_mails: int = 0
    nb_erreurs_entretien: int = 0
    nb_contrats_hors_delai: int = 0
    nb_erreurs_offres: int = 0
    nb_erreurs_type_comptage: int = 0
    nb_erreurs_energie_verte: int = 0
    nb_erreurs_car: int = 0
    nb_erreurs_puiss: int = 0
    nb_erreurs_reforest: int = 0
    nb_erreurs_protection: int = 0
    nb_erreurs_mandat: int = 0


class ImportEniResult(BaseModel):
    ok: bool
    type_import: int
    type_label: str
    simulation: bool
    resume: ImportEniResume
    contrats_modifies: list[dict] = []
    contrats_non_trouves: list[dict] = []
    contrats_run: list[dict] = []
    pb_vendeur: list[dict] = []
    contrat_import_journ_eni: list[dict] = []
    message: str = ""
    xlsx_b64: str = ""           # rapport XLSX en base64 (pour DL frontend)
    xlsx_name: str = ""          # nom du fichier
    mail_envoye: bool = False    # True si l'envoi a reussi


TYPE_LABELS = {
    1: "Base Journalière CALL",
    2: "RUN Valides",
    3: "RUN Résils",
    4: "Base journalière ENI",
}


def run_import(p: ImportEniParams, file_bytes: bytes, op_id: int) -> ImportEniResult:
    """Dispatcher principal."""
    label = TYPE_LABELS.get(p.type_import, "?")
    if not file_bytes:
        return ImportEniResult(
            ok=False, type_import=p.type_import, type_label=label,
            simulation=p.simulation, resume=ImportEniResume(),
            message="Aucun fichier fourni.",
        )

    if p.type_import == 1:
        return _import_journalier_call(p, file_bytes, op_id)
    if p.type_import == 2:
        return _import_run_valides(p, file_bytes, op_id)
    if p.type_import == 3:
        return _import_run_resil(p, file_bytes, op_id)
    if p.type_import == 4:
        return _import_journalier_eni(p, file_bytes, op_id)

    # TODO : autres procedures
    return ImportEniResult(
        ok=True, type_import=p.type_import, type_label=label,
        simulation=p.simulation, resume=ImportEniResume(),
        message=(f"Import {label} : procedure non encore codée. "
                 f"Fichier reçu ({len(file_bytes)} octets), mode "
                 f"{'simulation' if p.simulation else 'PRODUCTION'}."),
    )


# ---------------------------------------------------------------------------
# Type 1 : Base Journaliere CALL (ImportJournalier)
# ---------------------------------------------------------------------------


def _cell(ws, row: int, col_idx: int):
    """Lit une cellule Excel et renvoie une str trim."""
    if col_idx <= 0:
        return ""
    v = ws.cell(row=row, column=col_idx).value
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def _parse_date_fr(s: str) -> Optional[date]:
    """JJ/MM/AAAA ou variantes ISO -> date."""
    if not s:
        return None
    s = s.strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_int(s: str) -> int:
    if not s:
        return 0
    m = re.search(r"-?\d+", s.replace(" ", ""))
    return int(m.group(0)) if m else 0


def _normaliser_nom(s: str) -> str:
    """Uppercase + sans accent + sans doubles espaces + sans tirets."""
    if not s:
        return ""
    s = s.upper().strip()
    # Strip accents (approximatif)
    for a, b in [("À", "A"), ("Â", "A"), ("Ä", "A"), ("É", "E"), ("È", "E"),
                 ("Ê", "E"), ("Ë", "E"), ("Î", "I"), ("Ï", "I"), ("Ô", "O"),
                 ("Ö", "O"), ("Ù", "U"), ("Û", "U"), ("Ü", "U"), ("Ç", "C")]:
        s = s.replace(a, b)
    s = s.replace("-", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s.strip()


def _affectation_vendeur(id_salarie: int, _date_sign: Optional[date]) -> tuple[str, str]:
    """Retourne (lib_agence, lib_equipe) du vendeur a une date donnee.
    Pour l'instant : derniere affectation connue (simplification, pas
    de versioning historique date)."""
    if not id_salarie:
        return ("", "")
    db = get_pg_connection("rh")
    rows = db.query(
        """SELECT o.lib_orga, o.id_type_niveau_orga
             FROM rh.pgt_salarie_organigramme so
             JOIN rh.pgt_organigramme o ON o.idorganigramme = so.idorganigramme
            WHERE so.id_salarie = ?
              AND (so.modif_elem IS NULL OR so.modif_elem NOT LIKE '%suppr%')""",
        (int(id_salarie),),
    ) or []
    agence = ""
    equipe = ""
    for r in rows:
        lvl = r.get("id_type_niveau_orga")
        lib = r.get("lib_orga") or ""
        if lvl == 3 and not agence:
            agence = lib
        elif lvl == 4 and not equipe:
            equipe = lib
    return (agence, equipe)


def _lookup_tk_call_eni(num_bs: str) -> dict:
    """Cf. WinDev ReqTkCall_ByNumCtt(NumCtt, 'ENI') : recupere le
    ticket call associe au numero de contrat pour peupler les
    coordonnees client + eventuellement l'id_salarie.

    Renvoie dict compatible traiter_client + 'id_salarie' + hint
    non_call (implicite : si tk trouve, non_call=FALSE).
    Dict vide si absent.
    """
    if not num_bs:
        return {}
    try:
        db_bo = get_pg_connection("ticket_bo")
        tk = db_bo.query_one(
            """SELECT id_salarie, nom_client, nom_marital_client,
                      prenom_client, adresse1, adresse2, cp, ville,
                      mobile1, date_naiss, adr_mail
                 FROM ticket_bo.pgt_tk_call
                WHERE UPPER(num_bs) = UPPER(?)
                  AND UPPER(partenaire) = 'ENI'
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )
        if not tk:
            return {}
        nom = tk.get("nom_client") or ""
        if tk.get("nom_marital_client"):
            nom += f" ep {tk['nom_marital_client']}"
        return {
            "id_salarie": int(tk.get("id_salarie") or 0),
            "nom": nom, "prenom": tk.get("prenom_client") or "",
            "adresse": tk.get("adresse1") or "",
            "cplt": tk.get("adresse2") or "",
            "cp": tk.get("cp") or "", "ville": tk.get("ville") or "",
            "gsm": tk.get("mobile1") or "",
            "mail": tk.get("adr_mail") or "",
            "date_naiss": tk.get("date_naiss"),
        }
    except Exception:
        return {}


def _is_distrib_eni(id_salarie: int) -> bool:
    """Cf. WinDev IsDistrib : le vendeur appartient-il a une orga type
    distributeur (id_type_orga=3) au niveau equipe/agence ?

    Utilise pour ImportRUN Valides et Resils : si distributeur ->
    mois_paiement = Mois_paiementDistrib (au lieu des periodes 1/2/-1mois).
    """
    if not id_salarie:
        return False
    try:
        db = get_pg_connection("rh")
        rows = db.query(
            """SELECT o.id_type_orga
                 FROM rh.pgt_salarie_organigramme so
                 JOIN rh.pgt_organigramme o
                   ON o.idorganigramme = so.idorganigramme
                WHERE so.id_salarie = ?
                  AND (so.modif_elem IS NULL
                       OR so.modif_elem NOT LIKE '%suppr%')""",
            (int(id_salarie),),
        ) or []
        return any(int(r.get("id_type_orga") or 0) == 3 for r in rows)
    except Exception:
        return False


def _verif_activite_distrib(id_salarie: int, _date_sign: Optional[date]) -> str:
    """Cf. WinDev vérifActivitéDistrib(monOpé, DateSignature) : verifie
    que le salarie est actif dans salarie_embauche. Le code source WinDev
    n'a pas ete transpose entierement (pas de MAJ de statut, uniquement
    remontee d'info).

    Renvoie un message d'alerte non vide si le salarie n'est pas actif
    (a remonter dans pb_vendeur). Renvoie chaine vide si tout OK."""
    if not id_salarie:
        return ""
    db = get_pg_connection("rh")
    r = db.query_one(
        """SELECT en_activite FROM rh.pgt_salarie_embauche
            WHERE id_salarie = ? LIMIT 1""",
        (int(id_salarie),),
    )
    if r is not None and r.get("en_activite") is False:
        return "Salarié inactif à la date de signature"
    return ""


def _info_salarie(id_salarie: int) -> dict:
    if not id_salarie:
        return {}
    db = get_pg_connection("rh")
    r = db.query_one(
        """SELECT id_salarie, nom, prenom
             FROM rh.pgt_salarie WHERE id_salarie = ?""",
        (int(id_salarie),),
    )
    return r or {}


def _import_journalier_call(
    p: ImportEniParams, file_bytes: bytes, op_id: int,
) -> ImportEniResult:
    """Import 'Base Journaliere CALL' ENI : pour chaque ligne, cherche
    le contrat dans adv.pgt_eni_contrat par num_bs et :
      - Si trouve : ajoute a contrats_modifies + check vendeur (-> pb_vendeur si mismatch)
        + applique etat_contrat (15=Refus CALL si 'refus' dans statut, 51->37)
        + si pas simulation : UPDATE code_enr + non_call=false
      - Si pas trouve : ajoute a contrats_non_trouves
    """
    from openpyxl import load_workbook

    label = TYPE_LABELS.get(1, "Base Journalière CALL")
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return ImportEniResult(
            ok=False, type_import=1, type_label=label,
            simulation=p.simulation, resume=ImportEniResume(),
            message=f"Lecture Excel KO : {e}",
        )
    ws = wb.active

    # Mapping colonnes (defaut ; sera parametre BDD plus tard)
    cols = {k: _col_letter_to_index(v) for k, v in COLS_BJ_CALL.items()}

    resume = ImportEniResume()
    modifs: list[dict] = []
    non_trouves: list[dict] = []
    pb_vendeur: list[dict] = []

    db = get_pg_connection("adv")
    n_rows = ws.max_row or 0

    for i in range(2, n_rows + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"]).upper()
        if not num_contrat:
            continue
        date_sign_s = _cell(ws, i, cols["date_signature"])
        date_sign = _parse_date_fr(date_sign_s)
        vendeur_cell = _cell(ws, i, cols["vendeur_nom"])
        car = _cell(ws, i, cols["car"])
        offre = _cell(ws, i, cols["offre"])
        lib_statut = _cell(ws, i, cols["lib_statut"])

        client = {
            "nom": _cell(ws, i, cols["client_nom"]),
            "prenom": _cell(ws, i, cols["client_prenom"]),
            "adresse": _cell(ws, i, cols["client_adresse"]),
            "cplt": _cell(ws, i, cols["client_cplt"]),
            "cp": _cell(ws, i, cols["client_cp"]),
            "ville": _cell(ws, i, cols["client_ville"]),
            "tel": _cell(ws, i, cols["client_tel"]),
            "gsm": _cell(ws, i, cols["client_gsm"]),
            "mail": _cell(ws, i, cols["client_mail"]),
            "date_naiss": _cell(ws, i, cols.get("date_naissance") or 0),
        }

        code_enr = _cell(ws, i, cols["code_enr"])
        heure = _cell(ws, i, cols["heure"])
        indice = _cell(ws, i, cols["indice"])
        if not code_enr:
            heure_norm = (heure or "0000").replace(":", "")[:4]
            code_enr = f"AGT_0000_IND_{indice}_TEL_123456789_TIME_{heure_norm}00"

        puiss = _parse_int(_cell(ws, i, cols["puissance"]))

        # Lookup contrat par num_bs (= num_contrat)
        r = db.query_one(
            """SELECT id_contrat, id_salarie, id_client, id_ste, num_bs,
                      id_produit, id_etat_contrat, date_signature,
                      gaz_car_relevee, elec_puissance, gaz_actif
                 FROM adv.pgt_eni_contrat
                WHERE UPPER(num_bs) = ?
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_contrat,),
        )

        if not r:
            non_trouves.append({
                "NumCtt": num_contrat,
                "DateSign": date_sign_s,
                "Vendeur": vendeur_cell,
                "Produit": offre,
                "CAR": car,
                "Puiss": puiss,
                "LibStatut": lib_statut,
                "Nom Cli": client["nom"],
                "Prenom Cli": client["prenom"],
                "CP": client["cp"],
                "Ville": client["ville"],
                "GSM": client["gsm"],
                "CodeEnr": code_enr,
            })
            resume.nb_introuvables += 1
            continue

        # Trouve : on note la modif
        id_contrat = int(r["id_contrat"])
        id_salarie_db = int(r.get("id_salarie") or 0)
        etat_actuel = int(r.get("id_etat_contrat") or 0)
        nouvel_etat = etat_actuel
        if etat_actuel == 51:
            nouvel_etat = 37
        if "refus" in (lib_statut or "").lower():
            nouvel_etat = 15

        # Check vendeur cell vs vendeur en BDD
        sal_info = _info_salarie(id_salarie_db)
        nom_db = _normaliser_nom(
            f"{sal_info.get('nom', '')} {sal_info.get('prenom', '')}".strip()
        )
        nom_cell = _normaliser_nom(vendeur_cell)
        agence, equipe = _affectation_vendeur(id_salarie_db, date_sign)

        row_common = {
            "NumCtt": r.get("num_bs"),
            "DateSign": str(r.get("date_signature") or ""),
            "IdSalarie": id_salarie_db,
            "Societe": int(r.get("id_ste") or 0),
            "Agence": agence,
            "Equipe": equipe,
            "Produit": int(r.get("id_produit") or 0),
            "CAR": int(r.get("gaz_car_relevee") or 0),
            "Puiss Elec": int(r.get("elec_puissance") or 0),
            "Etat actuel": etat_actuel,
            "Nouvel etat": nouvel_etat,
            "Gaz actif": bool(r.get("gaz_actif")),
            "CodeEnr": code_enr,
            "Cli Nom": client["nom"],
            "Cli Prenom": client["prenom"],
            "Cli GSM": client["gsm"],
            "Cli Mail": client["mail"],
        }

        if nom_cell and nom_db and nom_cell != nom_db:
            pb_vendeur.append({
                **row_common,
                "Vendeur Import": vendeur_cell,
                "Vendeur OMAYA": nom_db,
            })

        # vérifActivitéDistrib : remonte si salarie inactif a la date sign
        msg_act = _verif_activite_distrib(id_salarie_db, date_sign)
        if msg_act:
            pb_vendeur.append({**row_common, "Vendeur Import": vendeur_cell,
                                "Alerte": msg_act})

        modifs.append(row_common)
        resume.nb_modifications += 1

        # MAJ reelle (hors simulation)
        if not p.simulation:
            db.query(
                """UPDATE adv.pgt_eni_contrat
                      SET code_enr = ?,
                          non_call = FALSE,
                          id_etat_contrat = ?,
                          modif_date = NOW(),
                          modif_op = ?,
                          modif_elem = 'modif'
                    WHERE id_contrat = ?""",
                (code_enr, nouvel_etat, int(op_id), id_contrat),
            )

            # Traitement client : lookup / update / create
            ci = {
                "nom": client["nom"], "prenom": client["prenom"],
                "adresse": client["adresse"], "cplt": client["cplt"],
                "cp": client["cp"], "ville": client["ville"],
                "tel": client["tel"], "gsm": client["gsm"],
                "mail": client["mail"], "date_naiss": client["date_naiss"],
            }
            id_client_new = _lookup_or_create_client(ci, op_id)
            id_client_old = int(r.get("id_client") or 0)
            if id_client_new and id_client_new != id_client_old:
                # Re-link contrat -> nouveau client + soft-delete ancien
                db.query(
                    """UPDATE adv.pgt_eni_contrat
                          SET id_client = ?, modif_date = NOW()
                        WHERE id_contrat = ?""",
                    (id_client_new, id_contrat),
                )
                if id_client_old:
                    db.query(
                        """UPDATE adv.pgt_client
                              SET modif_elem = 'suppr', modif_date = NOW(),
                                  modif_op = ?
                            WHERE id_client = ?""",
                        (int(op_id), id_client_old),
                    )

            # Cree option si absente
            opt_exists = db.query_one(
                """SELECT 1 FROM adv.pgt_eni_contrat_option
                    WHERE id_contrat = ? LIMIT 1""",
                (id_contrat,),
            )
            if not opt_exists:
                auto = db.query_one(
                    "SELECT COALESCE(MAX(id_contrat_option_auto), 0) + 1 AS n "
                    "FROM adv.pgt_eni_contrat_option"
                )
                db.query(
                    """INSERT INTO adv.pgt_eni_contrat_option
                          (id_contrat_option_auto, id_contrat, num_bs,
                           modif_op, modif_date, modif_elem)
                       VALUES (?, ?, ?, ?, NOW(), 'new')""",
                    (int(auto["n"]) if auto else 1,
                     id_contrat, r.get("num_bs") or "", int(op_id)),
                )

    wb.close()

    res = ImportEniResult(
        ok=True, type_import=1, type_label=label,
        simulation=p.simulation, resume=resume,
        contrats_modifies=modifs,
        contrats_non_trouves=non_trouves,
        pb_vendeur=pb_vendeur,
        message=(
            f"{n_rows - 1} ligne(s) lue(s) | "
            f"{resume.nb_modifications} contrat(s) trouvé(s) | "
            f"{resume.nb_introuvables} introuvable(s) | "
            f"{len(pb_vendeur)} pb vendeur. "
            f"{'(SIMULATION : aucune écriture)' if p.simulation else '(PRODUCTION : MAJ appliquées)'}"
        ),
    )
    _attach_xlsx_and_mail(res, op_id, "ImportCallENI")
    return res


# ---------------------------------------------------------------------------
# Type 2 : RUN Valides (importRUNValides)
# ---------------------------------------------------------------------------


def _conso_gaz_type(car: int) -> int:
    if 1000 <= car <= 6000: return 1
    if 6001 <= car <= 13000: return 2
    if car >= 13001: return 3
    return 0


def _conso_elec_type(puiss: int) -> int:
    if puiss == 3: return 1
    if puiss == 6: return 2
    if puiss == 9: return 3
    if puiss >= 12: return 4
    return 0


def _dernier_jour_mois(s: str) -> Optional[date]:
    """'06-2025' ou 'MM-YYYY' -> dernier jour du mois."""
    if not s:
        return None
    s = s.strip()
    try:
        if "-" in s:
            mm, yy = s.split("-")
            mois = int(mm); annee = int(yy)
        else:
            return None
        # Calcul dernier jour
        if mois == 12:
            return date(annee + 1, 1, 1) - __import__("datetime").timedelta(days=1)
        return date(annee, mois + 1, 1) - __import__("datetime").timedelta(days=1)
    except (ValueError, TypeError):
        return None


def _import_run_valides(
    p: ImportEniParams, file_bytes: bytes, op_id: int,
) -> ImportEniResult:
    """Import 'RUN Valides' ENI : analyse contrats valides et calcule
    nouvel etat (19/46/47/...) + MoisP selon la periode.

    Phase 1 : detection seulement (pas de MAJ), pas de XLSX/mail."""
    from openpyxl import load_workbook

    label = TYPE_LABELS.get(2, "RUN Valides")
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return ImportEniResult(
            ok=False, type_import=2, type_label=label,
            simulation=p.simulation, resume=ImportEniResume(),
            message=f"Lecture Excel KO : {e}",
        )
    ws = wb.active

    cols = {k: _col_letter_to_index(v) for k, v in COLS_RUN_V.items()}

    # Dates des periodes
    d1_du = _parse_date_fr(p.periode1_du) or date(1900, 1, 1)
    d1_au = _parse_date_fr(p.periode1_au) or date(2100, 12, 31)
    d2_du = _parse_date_fr(p.periode2_du) or date(1900, 1, 1)
    d2_au = _parse_date_fr(p.periode2_au) or date(2100, 12, 31)
    mp1 = _dernier_jour_mois(p.periode1_mois_paiement)
    mp2 = _dernier_jour_mois(p.periode2_mois_paiement)
    mp_distrib = _dernier_jour_mois(p.mois_paiement_distrib)
    import datetime as _dt
    d1_du_m1 = d1_du.replace(month=d1_du.month - 1) if d1_du.month > 1 else d1_du.replace(year=d1_du.year - 1, month=12)
    d1_au_m1 = d1_au.replace(month=d1_au.month - 1) if d1_au.month > 1 else d1_au.replace(year=d1_au.year - 1, month=12)

    resume = ImportEniResume()
    contrats_run: list[dict] = []
    hors_delai: list[dict] = []
    doublons: list[dict] = []
    non_trouves: list[dict] = []

    db = get_pg_connection("adv")
    n_rows = ws.max_row or 0

    for i in range(2, n_rows + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"])
        if not num_contrat:
            continue
        type_offre = _cell(ws, i, cols["type_offre"]).strip()
        vend_full = _cell(ws, i, cols["vendeur_nom"])
        vendeur_nom = vend_full.split(" / ")[0] if vend_full else ""
        offre_reclass = _cell(ws, i, cols["offre_reclassifiee"]).strip()
        statut_gaz = _cell(ws, i, cols["statut_gaz"]).upper()
        statut_elec = _cell(ws, i, cols["statut_elec"]).upper()
        b_gaz_actif = (statut_gaz == "ON_SUPPLY")
        b_elec_actif = (statut_elec == "ON_SUPPLY")
        date_sign = _parse_date_fr(_cell(ws, i, cols["date_signature"]))
        car = _parse_int(_cell(ws, i, cols["car"]))
        puiss_elec = _parse_int(_cell(ws, i, cols["puissance"]))
        type_comptage = _cell(ws, i, cols["type_comptage"])
        opt_hphc = "HPHC" in type_comptage.upper()
        montant_mails = _parse_int(_cell(ws, i, cols["mail_fourni"]))
        montant_hs = _parse_int(_cell(ws, i, cols["home_serve"]))
        montant_mandat = _parse_int(_cell(ws, i, cols["mandat"]))
        opt_mails = montant_mails > 0
        opt_home_serve = montant_hs > 0
        opt_mandat_xls = montant_mandat > 0
        # cf. WinDev importRUNValides : Opt_Protection est mis a 0
        opt_protect = False

        # Lookup contrat(s)
        rows_ctt = db.query(
            """SELECT id_contrat, id_salarie, id_client, id_ste, num_bs,
                      id_produit, id_etat_contrat, date_signature,
                      gaz_car_relevee, gaz_car_declaree, elec_puissance,
                      gaz_actif, elec_actif, code_enr, non_call,
                      mois_p, notation
                 FROM adv.pgt_eni_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                """,
            (num_contrat,),
        ) or []

        if len(rows_ctt) == 0:
            non_trouves.append({
                "NumCtt": num_contrat,
                "DateSign": _cell(ws, i, cols["date_signature"]),
                "Vendeur": vendeur_nom,
                "CAR": car,
                "Puiss": puiss_elec,
                "GazActif": b_gaz_actif,
                "ElecActif": b_elec_actif,
                "TypeOffre": type_offre,
            })
            resume.nb_introuvables += 1
            continue
        if len(rows_ctt) > 1:
            resume.nb_doublons += 1
            for r in rows_ctt:
                doublons.append({
                    "NumCtt": num_contrat,
                    "id_contrat_OMAYA": r.get("id_contrat"),
                    "DateSign_OMAYA": str(r.get("date_signature") or ""),
                    "TypeOffre": type_offre,
                    "CAR": car,
                    "Puiss": puiss_elec,
                    "GazActif_OMAYA": bool(r.get("gaz_actif")),
                    "ElecActif_OMAYA": bool(r.get("elec_actif")),
                })
            continue

        r = rows_ctt[0]
        id_contrat = int(r["id_contrat"])
        date_sign_omaya = r.get("date_signature")
        etat_actuel = int(r.get("id_etat_contrat") or 0)
        car_omaya = int(r.get("gaz_car_relevee") or 0)
        puiss_omaya = int(r.get("elec_puissance") or 0)
        id_produit = int(r.get("id_produit") or 0)
        id_salarie_contrat = int(r.get("id_salarie") or 0)

        # Lookup produit pour SousFAM (cf. WinDev : flagOffre matching
        # TypeOffre vs ENI_produit.SousFAM)
        prod_info = db.query_one(
            """SELECT sous_fam, famille, lib_produit
                 FROM adv.pgt_eni_produit
                WHERE id_produit = ? LIMIT 1""",
            (id_produit,),
        )
        sous_fam = (prod_info or {}).get("sous_fam") or ""

        # Periode
        # cf. WinDev l.248-250 : IsDistrib -> Mois_paiementDistrib
        # (branche prioritaire avant les periodes 1/2/-1mois).
        # Sans ce check, les contrats des distributeurs sont classes
        # "Hors Délai" a tort et jamais payes.
        mois_paiement = None
        periode_lbl = ""
        if date_sign_omaya:
            if _is_distrib_eni(id_salarie_contrat):
                mois_paiement = mp_distrib
                periode_lbl = "Distrib"
            elif d1_du <= date_sign_omaya <= d1_au:
                mois_paiement = mp1; periode_lbl = "Période 1"
            elif d2_du <= date_sign_omaya <= d2_au:
                mois_paiement = mp2; periode_lbl = "Période 2"
            elif d1_du_m1 <= date_sign_omaya <= d1_au_m1:
                mois_paiement = mp1; periode_lbl = "Période -1 mois"
            else:
                resume.nb_contrats_hors_delai += 1
                hors_delai.append({
                    "NumCtt": num_contrat,
                    "DateSign": str(date_sign_omaya),
                    "TypeOffre": type_offre,
                    "CAR": car,
                    "Puiss": puiss_elec,
                })

        # Etat actuel : check IDTypeEtat
        etat_info = db.query_one(
            """SELECT id_type_etat, lib_etat FROM adv.pgt_eni_etat_contrat
                WHERE id_etat = ? LIMIT 1""",
            (etat_actuel,),
        )
        id_type_etat = int(etat_info.get("id_type_etat") or 0) if etat_info else 0

        # flagOffre : matching TypeOffre vs sous_fam produit (cf. WinDev)
        flag_offre = True
        if type_offre == "Dual" and sous_fam != "GAZ-ELEC":
            flag_offre = False
        elif type_offre == "Mono Gaz" and sous_fam != "GAZ":
            flag_offre = False
        elif type_offre == "Mono Elec" and sous_fam != "ELEC":
            flag_offre = False
        if not flag_offre:
            resume.nb_erreurs_offres += 1

        # testDualResilPartielle : Dual + IDTypeEtat=4 + au moins un
        # cote actif -> le contrat peut etre repasse en Valide partiel
        # (cf. WinDev : si ENI_contrat.IDetatContrat<>59 et TypeOffre=Dual
        # et IDTypeEtat=4 et (GazValide=1 ou ElecValide=1)).
        test_dual_resil_partielle = False
        if (etat_actuel != 59
                and type_offre == "Dual"
                and id_type_etat == 4
                and (b_gaz_actif or b_elec_actif)):
            test_dual_resil_partielle = True

        # Determination du nouvel etat (cf WinDev)
        nouvel_etat = etat_actuel
        gaz_valide_nouv = b_gaz_actif
        elec_valide_nouv = b_elec_actif
        maj_etat_flag = False

        # IDTypeEtat 1, 2 ou 8, ou IDetat 54, ou resil partielle detectee
        # -> contrat eligible MAJ (cf WinDev importRUNValides.txt l.423,
        # maj 2026-07-06 : ajout du type 8).
        eligible = ((id_type_etat in (1, 2, 8))
                    or etat_actuel == 54
                    or test_dual_resil_partielle)
        if eligible and flag_offre:
            if not offre_reclass:
                nouvel_etat = 19  # Valide - Paye par l'opérateur
                if type_offre == "Dual":
                    gaz_valide_nouv = True; elec_valide_nouv = True
                elif "Gaz" in type_offre or "gaz" in type_offre:
                    gaz_valide_nouv = True; elec_valide_nouv = False
                elif "Elec" in type_offre or "elec" in type_offre:
                    gaz_valide_nouv = False; elec_valide_nouv = True
            else:
                if offre_reclass == "Solo Elec":
                    nouvel_etat = 46  # Partiel Elec
                    gaz_valide_nouv = False
                else:
                    nouvel_etat = 47  # Partiel Gaz
                    elec_valide_nouv = False
            maj_etat_flag = True
            resume.nb_valides += 1
        elif eligible and not flag_offre:
            # Eligible mais mauvaise offre -> pas de MAJ etat mais
            # comptage en 'erreurs offre' (deja fait ci-dessus)
            pass
        else:
            resume.nb_deja_statues += 1

        # Erreurs CAR / Puissance a la baisse
        maj_car_flag = False
        if car < car_omaya:
            resume.nb_erreurs_car += 1
            if p.maj_produit_contrat_stand:
                maj_car_flag = True
        if puiss_elec < puiss_omaya:
            resume.nb_erreurs_puiss += 1
            if p.maj_produit_contrat_stand:
                maj_car_flag = True

        # Options : look up
        opt = db.query_one(
            """SELECT opt_mail, opt_e_facture, opt_e_communication,
                      opt_optin_commercial, opt_hp_hc, opt_protection,
                      opt_entretien, opt_energie_verte_gaz,
                      opt_energie_verte_elec, opt_reforestation
                 FROM adv.pgt_eni_contrat_option WHERE id_contrat = ? LIMIT 1""",
            (id_contrat,),
        )
        # opt_mandat sur pgt_eni_contrat (colonne peut ne pas exister
        # si migration SQL pas encore faite - fallback False)
        db_opt_mandat = False
        try:
            r_mandat = db.query_one(
                "SELECT opt_mandat FROM adv.pgt_eni_contrat WHERE id_contrat = ?",
                (id_contrat,),
            )
            db_opt_mandat = bool(r_mandat and r_mandat.get("opt_mandat"))
        except Exception:
            # colonne opt_mandat pas encore en base
            db_opt_mandat = False

        if opt:
            if (not opt_mails and opt.get("opt_e_facture")
                    and opt.get("opt_e_communication")
                    and opt.get("opt_optin_commercial")):
                resume.nb_erreurs_mails += 1
            if not opt_hphc and opt.get("opt_hp_hc"):
                resume.nb_erreurs_type_comptage += 1
            # cf. WinDev TXT 2026-07 ligne 537 : HomeServe XLSX vs opt_entretien DB
            if not opt_home_serve and opt.get("opt_entretien"):
                resume.nb_erreurs_entretien += 1
        # cf. WinDev TXT 2026-07 ligne 513 : Mandat XLSX vs pgt_eni_contrat.opt_mandat
        if not opt_mandat_xls and db_opt_mandat:
            resume.nb_erreurs_mandat += 1

        contrats_run.append({
            "NumCtt": r.get("num_bs"),
            "DateSign": str(date_sign_omaya or ""),
            "Periode": periode_lbl,
            "MoisP": str(mois_paiement) if mois_paiement else "",
            "TypeOffre": type_offre,
            "OffreReclass": offre_reclass,
            "CAR Excel": car,
            "CAR OMAYA": car_omaya,
            "Puiss Excel": puiss_elec,
            "Puiss OMAYA": puiss_omaya,
            "GazActif Excel": b_gaz_actif,
            "ElecActif Excel": b_elec_actif,
            "ConsoGaz": _conso_gaz_type(car_omaya),
            "ConsoElec": _conso_elec_type(puiss_omaya),
            "Etat actuel": etat_actuel,
            "TypeEtat": id_type_etat,
            "Nouvel etat": nouvel_etat,
            "MAJ etat": maj_etat_flag,
            "MAJ car": maj_car_flag,
            "OptMail Excel": opt_mails,
            "OptHPHC Excel": opt_hphc,
            "OptHomeServe Excel": opt_home_serve,
            "OptMandat Excel": opt_mandat_xls,
            # Payload pour MAJ prod
            "_maj_data": {
                "id_contrat": id_contrat,
                "id_produit": int(r.get("id_produit") or 0),
                "date_signature": date_sign_omaya,
                "maj_etat": maj_etat_flag,
                "maj_car": maj_car_flag,
                "etat_actuel": etat_actuel,
                "nouvel_etat": nouvel_etat,
                "gaz_valide": gaz_valide_nouv,
                "elec_valide": elec_valide_nouv,
                "mois_p": mois_paiement,
                "mois_p_old": str(r.get("mois_p") or ""),
                "car_excel": car,
                "puiss_excel": puiss_elec,
                "opt_mail": opt_mails,
                "opt_home_serve": opt_home_serve,
                "opt_mandat": opt_mandat_xls,
                # cf. WinDev : Opt_Protection est mis a 0 dans importRUNValides
                "opt_protect": False,
                "notation": r.get("notation"),
            },
        })

    wb.close()

    # -- PASSE PROD : MAJ contrat + option + histo --
    nb_majs = 0
    if not p.simulation or p.maj_produit_contrat_stand:
        for row in contrats_run:
            md = row.pop("_maj_data", {})
            if not md:
                continue
            try:
                done = _apply_maj_run_valides(md, op_id, p.simulation)
                if done:
                    nb_majs += 1
            except Exception as e:
                row["Erreur"] = str(e)

    res = ImportEniResult(
        ok=True, type_import=2, type_label=label,
        simulation=p.simulation, resume=resume,
        contrats_run=contrats_run + hors_delai,
        contrats_non_trouves=non_trouves,
        contrats_modifies=doublons,  # on reutilise cet onglet pour les doublons
        message=(
            f"{n_rows - 1} ligne(s) lue(s) | "
            f"Validés {resume.nb_valides} | "
            f"Déjà statués {resume.nb_deja_statues} | "
            f"Hors délai {resume.nb_contrats_hors_delai} | "
            f"Doublons {resume.nb_doublons} | "
            f"Introuvables {resume.nb_introuvables}. "
            + (f"{nb_majs} MAJ appliquées." if not p.simulation
               else "(SIMULATION)")
        ),
    )
    _attach_xlsx_and_mail(res, op_id, "ImportRunENIValide")
    return res


def _apply_maj_run_valides(md: dict, op_id: int, simulation: bool) -> bool:
    """MAJ eni_contrat + eni_contrat_option pour RUN Valides.
    Si simulation=True : ne fait que la MAJ CAR/Puissance (cf WinDev
    InterrupteurMAJCAR). Sinon : MAJ complete + histo.
    Retourne True si une MAJ a ete appliquee."""
    db = get_pg_connection("adv")
    id_contrat = int(md["id_contrat"])
    has_change = False

    # MAJ CAR / Puissance (MAJ produit STAND : peut tourner meme en simu)
    if md.get("maj_car"):
        # Recalcul nb_points
        prod = db.query_one(
            """SELECT famille, sous_fam FROM adv.pgt_eni_produit
                WHERE id_produit = ? LIMIT 1""",
            (int(md.get("id_produit") or 0),),
        )
        nb_pts = 0.0
        if prod:
            # cf. WinDev importRUNValides (MAJ 2026-07-06) :
            # la chaine nbOption depend de la date de signature.
            # - Post-2026-05-01 : "NOTE:<notation>//" + "RIB//" (opt_mandat)
            #   + "MAIL//" (opt_mail) + "MAINT//" (opt_entretien)
            # - Pre-2026-05-01  : compte des 4 options historiques
            #   (opt_mail + opt_reforestation + opt_energie_verte_elec/gaz)
            opts = db.query_one(
                """SELECT opt_mail, opt_entretien,
                          opt_reforestation, opt_energie_verte_elec,
                          opt_energie_verte_gaz
                     FROM adv.pgt_eni_contrat_option
                    WHERE id_contrat = ?
                      AND (modif_elem IS NULL
                           OR modif_elem NOT LIKE '%suppr%')
                    LIMIT 1""",
                (id_contrat,),
            ) or {}
            date_sign = md.get("date_signature")
            date_sign_iso = str(date_sign)[:10] if date_sign else ""
            is_post_20260501 = date_sign_iso >= "2026-05-01"
            if is_post_20260501:
                # NB : opt_mandat n'est pas encore en base PG (colonne
                # absente sur pgt_eni_contrat). Traite comme False tant
                # que la migration n'est pas faite.
                opt_mandat = bool(md.get("opt_mandat") or False)
                parts = [f"NOTE:{md.get('notation') or 0}//"]
                if opt_mandat:
                    parts.append("RIB//")
                if opts.get("opt_mail"):
                    parts.append("MAIL//")
                if opts.get("opt_entretien"):
                    parts.append("MAINT//")
                info_cplt = "".join(parts)
            else:
                nb_opts = (
                    int(bool(opts.get("opt_mail")))
                    + int(bool(opts.get("opt_reforestation")))
                    + int(bool(opts.get("opt_energie_verte_elec")))
                    + int(bool(opts.get("opt_energie_verte_gaz")))
                )
                info_cplt = str(nb_opts)
            nb_pts = _calcul_points_eni(
                prod.get("famille") or "", prod.get("sous_fam") or "",
                int(md.get("car_excel") or 0),
                date_sign,
                info_cplt,
                int(md.get("puiss_excel") or 0),
            )
        db.query(
            """UPDATE adv.pgt_eni_contrat
                  SET gaz_car_relevee = ?, elec_puissance = ?,
                      nb_points = ?,
                      modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                WHERE id_contrat = ?""",
            (int(md.get("car_excel") or 0), int(md.get("puiss_excel") or 0),
             nb_pts, int(op_id), id_contrat),
        )
        has_change = True

    if simulation:
        return has_change

    # MAJ etat + option + histo (PRODUCTION uniquement)
    if md.get("maj_etat"):
        db.query(
            """UPDATE adv.pgt_eni_contrat
                  SET id_etat_contrat = ?, gaz_actif = ?, elec_actif = ?,
                      mois_p = ?,
                      modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                WHERE id_contrat = ?""",
            (int(md["nouvel_etat"]), bool(md.get("gaz_valide")),
             bool(md.get("elec_valide")), md.get("mois_p"),
             int(op_id), id_contrat),
        )
        _ajoute_histo_eni_etat(
            id_contrat, int(md["etat_actuel"]), int(md["nouvel_etat"]),
            md.get("mois_p_old", ""), op_id,
        )
        has_change = True

    # MAJ options (Mail + HomeServe/Entretien + Protection)
    # cf. WinDev TXT 2026-07 :
    #   XLSX 'HomeServe' -> pgt_eni_contrat_option.opt_entretien
    #   XLSX 'Mandat' -> pgt_eni_contrat.opt_mandat
    opt_home_serve = bool(md.get("opt_home_serve"))
    if md.get("opt_mail"):
        db.query(
            """UPDATE adv.pgt_eni_contrat_option
                  SET opt_mail = TRUE, opt_e_facture = TRUE,
                      opt_e_communication = TRUE, opt_optin_commercial = TRUE,
                      opt_entretien = ?,
                      opt_protection = ?,
                      modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                WHERE id_contrat = ?""",
            (opt_home_serve, bool(md.get("opt_protect")),
             int(op_id), id_contrat),
        )
        has_change = True
    else:
        db.query(
            """UPDATE adv.pgt_eni_contrat_option
                  SET opt_mail = FALSE, opt_entretien = ?, opt_protection = ?,
                      modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                WHERE id_contrat = ?""",
            (opt_home_serve, bool(md.get("opt_protect")),
             int(op_id), id_contrat),
        )
        has_change = True

    # MAJ opt_mandat sur pgt_eni_contrat (colonne peut ne pas exister
    # si migration SQL pas encore faite - skip silencieux)
    try:
        db.query(
            """UPDATE adv.pgt_eni_contrat
                  SET opt_mandat = ?,
                      modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                WHERE id_contrat = ?""",
            (bool(md.get("opt_mandat")), int(op_id), id_contrat),
        )
        has_change = True
    except Exception:
        logger.warning(
            "opt_mandat MAJ KO (migration pgt_eni_contrat.opt_mandat requise)",
        )

    return has_change


# ---------------------------------------------------------------------------
# Type 3 : RUN Resils (importRUNResil)
# ---------------------------------------------------------------------------


def _import_run_resil(
    p: ImportEniParams, file_bytes: bytes, op_id: int,
) -> ImportEniResult:
    """Import 'RUN Resils' ENI : pour chaque ligne, determine si le
    contrat doit etre :
      - Resilie (etat 16) si type_etat in (1,2) ou etat=54 ou
        (type_etat=5 ET meme mois paiement)
      - Decommissionne (20=total, 49=partiel ELEC, 50=partiel GAZ)
        si type_etat=5 et mois paiement different
      - Deja statue sinon
    Phase 1 : detection seulement."""
    from openpyxl import load_workbook

    label = TYPE_LABELS.get(3, "RUN Résils")
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return ImportEniResult(
            ok=False, type_import=3, type_label=label,
            simulation=p.simulation, resume=ImportEniResume(),
            message=f"Lecture Excel KO : {e}",
        )
    ws = wb.active

    cols = {k: _col_letter_to_index(v) for k, v in COLS_RUN_R.items()}

    d1_du = _parse_date_fr(p.periode1_du) or date(1900, 1, 1)
    d1_au = _parse_date_fr(p.periode1_au) or date(2100, 12, 31)
    d2_du = _parse_date_fr(p.periode2_du) or date(1900, 1, 1)
    d2_au = _parse_date_fr(p.periode2_au) or date(2100, 12, 31)
    mp1 = _dernier_jour_mois(p.periode1_mois_paiement)
    mp2 = _dernier_jour_mois(p.periode2_mois_paiement)
    mp_distrib = _dernier_jour_mois(p.mois_paiement_distrib)
    d1_du_m1 = (d1_du.replace(month=d1_du.month - 1) if d1_du.month > 1
                else d1_du.replace(year=d1_du.year - 1, month=12))
    d1_au_m1 = (d1_au.replace(month=d1_au.month - 1) if d1_au.month > 1
                else d1_au.replace(year=d1_au.year - 1, month=12))

    resume = ImportEniResume()
    contrats_run: list[dict] = []
    doublons: list[dict] = []
    non_trouves: list[dict] = []
    hors_delai: list[dict] = []

    db = get_pg_connection("adv")
    n_rows = ws.max_row or 0

    for i in range(2, n_rows + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"])
        if not num_contrat:
            continue
        type_offre = _cell(ws, i, cols["type_offre"]).strip()
        vend_full = _cell(ws, i, cols["vendeur_nom"])
        vendeur_nom = vend_full.split(" / ")[0] if vend_full else ""
        date_sign_s = _cell(ws, i, cols["date_signature"])

        rows_ctt = db.query(
            """SELECT id_contrat, id_salarie, id_client, num_bs,
                      id_etat_contrat, date_signature, mois_p
                 FROM adv.pgt_eni_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
            (num_contrat,),
        ) or []

        if len(rows_ctt) == 0:
            non_trouves.append({
                "NumCtt": num_contrat,
                "DateSign": date_sign_s,
                "Vendeur": vendeur_nom,
                "TypeOffre": type_offre,
                "Onglet": "RESIL",
            })
            resume.nb_introuvables += 1
            continue
        if len(rows_ctt) > 1:
            resume.nb_doublons += 1
            for r in rows_ctt:
                doublons.append({
                    "NumCtt": num_contrat,
                    "id_contrat_OMAYA": r.get("id_contrat"),
                    "DateSign_OMAYA": str(r.get("date_signature") or ""),
                    "TypeOffre": type_offre,
                    "Onglet": "RESIL",
                })
            continue

        r = rows_ctt[0]
        id_contrat = int(r["id_contrat"])
        date_sign_omaya = r.get("date_signature")
        etat_actuel = int(r.get("id_etat_contrat") or 0)
        mois_p_omaya = r.get("mois_p")
        id_salarie_contrat = int(r.get("id_salarie") or 0)

        # Periode + MoisP
        # cf. WinDev l.196-198 : IsDistrib -> Mois_paiementDistrib
        # (branche prioritaire avant les periodes standards)
        mois_paiement = None
        periode_lbl = ""
        if date_sign_omaya:
            if _is_distrib_eni(id_salarie_contrat):
                mois_paiement = mp_distrib
                periode_lbl = "Distrib"
            elif d1_du <= date_sign_omaya <= d1_au:
                mois_paiement = mp1; periode_lbl = "Période 1"
            elif d2_du <= date_sign_omaya <= d2_au:
                mois_paiement = mp2; periode_lbl = "Période 2"
            elif d1_du_m1 <= date_sign_omaya <= d1_au_m1:
                mois_paiement = mp1; periode_lbl = "Période -1 mois"
            else:
                resume.nb_contrats_hors_delai += 1
                hors_delai.append({
                    "NumCtt": num_contrat,
                    "DateSign": str(date_sign_omaya),
                    "TypeOffre": type_offre,
                    "Onglet": "RESIL HORS DELAI",
                })

        # Lookup type_etat
        etat_info = db.query_one(
            """SELECT id_type_etat, lib_etat FROM adv.pgt_eni_etat_contrat
                WHERE id_etat = ? LIMIT 1""",
            (etat_actuel,),
        )
        id_type_etat = int(etat_info.get("id_type_etat") or 0) if etat_info else 0
        lib_etat = (etat_info.get("lib_etat") or "") if etat_info else ""

        # Determination du nouvel etat (cf WinDev)
        nouvel_etat = etat_actuel
        nouveau_mois_p = ""
        majetat_flag = False
        traitement = "deja_statue"

        same_month = (
            mois_paiement and mois_p_omaya
            and mois_paiement.month == mois_p_omaya.month
            and mois_paiement.year == mois_p_omaya.year
        )

        is_resiliable = (
            id_type_etat in (1, 2) or etat_actuel == 54
            or (id_type_etat == 5 and same_month)
        )

        if is_resiliable:
            nouvel_etat = 16  # Resilie par operateur
            nouveau_mois_p = ""
            majetat_flag = True
            traitement = "resilie"
            resume.nb_resilies += 1
        elif id_type_etat == 5:
            # Decommissionne (mois_p_omaya different)
            majetat_flag = True
            if "partielle" in lib_etat.lower():
                if "elec" in lib_etat.lower():
                    nouvel_etat = 49  # Decommission Partielle ELEC
                    traitement = "decomm_part_elec"
                else:
                    nouvel_etat = 50  # Decommission Partielle GAZ
                    traitement = "decomm_part_gaz"
            else:
                nouvel_etat = 20  # Decommissionne total
                traitement = "decomm_total"
            nouveau_mois_p = str(mois_paiement) if mois_paiement else ""
            resume.nb_decommissions += 1
        else:
            # Deja statue (autres cas)
            resume.nb_deja_statues += 1

        contrats_run.append({
            "NumCtt": r.get("num_bs"),
            "DateSign": str(date_sign_omaya or ""),
            "Periode": periode_lbl,
            "TypeOffre": type_offre,
            "Etat actuel": etat_actuel,
            "TypeEtat": id_type_etat,
            "Lib Etat": lib_etat,
            "MoisP OMAYA": str(mois_p_omaya) if mois_p_omaya else "",
            "Nouvel etat": nouvel_etat,
            "Nouveau MoisP": nouveau_mois_p,
            "Traitement": traitement,
            "MAJ etat": majetat_flag,
            "_maj_data": {
                "id_contrat": id_contrat,
                "maj_etat": majetat_flag,
                "etat_actuel": etat_actuel,
                "nouvel_etat": nouvel_etat,
                "mois_p": mois_paiement if traitement != "resilie" else None,
                "mois_p_old": str(mois_p_omaya) if mois_p_omaya else "",
            },
        })

    wb.close()

    # -- PASSE PROD : MAJ etat + option + histo (si pas simulation) --
    nb_majs = 0
    if not p.simulation:
        for row in contrats_run:
            md = row.pop("_maj_data", {})
            if not md or not md.get("maj_etat"):
                continue
            try:
                id_ctt = int(md["id_contrat"])
                db.query(
                    """UPDATE adv.pgt_eni_contrat
                          SET id_etat_contrat = ?, mois_p = ?,
                              modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (int(md["nouvel_etat"]), md.get("mois_p"),
                     int(op_id), id_ctt),
                )
                # MAJ contrat_option : cf. WinDev L544-552 (importRUNResil)
                # Le WinDev copie explicitement OPT_EnergieVerteElec,
                # OPT_EnergieVerteGaz, OPT_Reforestation, OPT_Protection
                # dans la MAJ meme si les valeurs sont identiques
                # (declenche la replication + trace modif date/op +
                # tolerance schema si trigger sur ces colonnes).
                opt_actuelles = db.query_one(
                    """SELECT opt_energie_verte_elec, opt_energie_verte_gaz,
                              opt_reforestation, opt_protection
                         FROM adv.pgt_eni_contrat_option
                        WHERE id_contrat = ? LIMIT 1""",
                    (id_ctt,),
                ) or {}
                db.query(
                    """UPDATE adv.pgt_eni_contrat_option
                          SET opt_energie_verte_elec = ?,
                              opt_energie_verte_gaz = ?,
                              opt_reforestation = ?,
                              opt_protection = ?,
                              modif_date = NOW(), modif_op = ?,
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (bool(opt_actuelles.get("opt_energie_verte_elec")),
                     bool(opt_actuelles.get("opt_energie_verte_gaz")),
                     bool(opt_actuelles.get("opt_reforestation")),
                     bool(opt_actuelles.get("opt_protection")),
                     int(op_id), id_ctt),
                )
                _ajoute_histo_eni_etat(
                    id_ctt,
                    int(md["etat_actuel"]), int(md["nouvel_etat"]),
                    md.get("mois_p_old", ""), op_id,
                )
                nb_majs += 1
            except Exception as e:
                row["Erreur"] = str(e)

    res = ImportEniResult(
        ok=True, type_import=3, type_label=label,
        simulation=p.simulation, resume=resume,
        contrats_run=contrats_run + hors_delai,
        contrats_non_trouves=non_trouves,
        contrats_modifies=doublons,
        message=(
            f"{n_rows - 1} ligne(s) lue(s) | "
            f"Résiliés {resume.nb_resilies} | "
            f"Décommissions {resume.nb_decommissions} | "
            f"Déjà statués {resume.nb_deja_statues} | "
            f"Hors délai {resume.nb_contrats_hors_delai} | "
            f"Doublons {resume.nb_doublons} | "
            f"Introuvables {resume.nb_introuvables}. "
            + (f"{nb_majs} MAJ appliquées." if not p.simulation
               else "(SIMULATION)")
        ),
    )
    _attach_xlsx_and_mail(res, op_id, "ImportRunENIResilDecom")
    return res


# ---------------------------------------------------------------------------
# Type 4 : Base journaliere ENI (ImportJournalierENI)
# ---------------------------------------------------------------------------


def _parse_datetime_fr(s: str) -> Optional[datetime]:
    """Parse 'AAAA-MM-JJ HH:mm:SS.CCC' ou 'AAAA-JJ-MM HH:mm:SS.CCC'.

    Le code WinDev tente d'abord AAAA-JJ-MM puis bascule sur AAAA-MM-JJ
    si invalide. On essaie les 2 ordres ici aussi."""
    if not s:
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S",
                "%Y-%d-%m %H:%M:%S.%f", "%Y-%d-%m %H:%M:%S",
                "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _id_produit_from_type_offre(type_offre: str) -> int:
    """Dual=67, Elec=66, Gaz=65 (cf WinDev)."""
    if type_offre == "Dual":
        return 67
    if type_offre == "Elec":
        return 66
    return 65  # default = Gaz


def _etat_contrat_journ_eni(lib_statut: str, id_prod: int, offre: str,
                             non_call: bool,
                             etat_actuel: int = 0) -> tuple[int, bool, bool]:
    """Determine (id_etat, gaz_actif, elec_actif) selon WinDev :
      - CANCELLED :
        - Dual : 43 (Resil partielle GAZ) ou 42 (partielle ELEC) selon offre
        - Sinon : 57 (Retractation Client) + gaz/elec off
      - Sinon :
        - Pas NonCALL :
          * Si etat_actuel = 72 -> on ne touche pas (cf. WinDev
            ImportJournalierENI.txt l.290, maj 2026-07-06).
          * Sinon : 37 (En cours traitement) ou 70 (En cours valid.)
        - NonCALL : 51 (Temporaire SF)
    """
    statut = (lib_statut or "").upper()
    if "CANCELLED" in statut:
        if id_prod == 67:  # Dual
            if "GAZ" in (offre or "").upper():
                return (43, False, True)  # Resil partielle GAZ
            else:
                return (42, True, False)  # Resil partielle ELEC
        return (57, False, False)  # Retractation Client
    # Pas CANCELLED
    if not non_call:
        # cf. WinDev l.290 (maj 2026-07-06) : etat 72 preserve
        if etat_actuel == 72:
            return (72, id_prod != 65, id_prod != 66)
        if "COMPLETED" in statut:
            return (70, id_prod != 65, id_prod != 66)  # BS CALL En cours valid.
        return (37, id_prod != 65, id_prod != 66)  # En cours traitement
    return (51, id_prod != 65, id_prod != 66)  # Temporaire SF


# ---------------------------------------------------------------------------
# Helpers phase 2 (mode PRODUCTION)
# ---------------------------------------------------------------------------


def _new_id() -> int:
    """Id 8 octets construit depuis la date/heure (cf idEntierDateHeureSys
    WinDev)."""
    n = datetime.now()
    return int(n.strftime("%Y%m%d%H%M%S")) * 1000 + n.microsecond // 1000


def _calcul_points_eni(famille: str, sous_fam: str, car: int,
                        date_sign: Optional[date], options_str: str,
                        puissance: int) -> float:
    """Calcule nb_points en deleguant a calcul_point_contrat (transposition
    fidele WinDev calculPointContrat).

    Mapping WinDev :
      - Palier  = valeur metier (CAR pour Gaz, puissance pour Elec)
      - palier2 = autre valeur (puissance si GAZ, CAR si ELEC) - utilise
        pour le bonus ELEC sur GAZ-ELEC depuis 2023-02-07
      - InfoCplt = options_str (RIB//, MAIL//, MAINT//, NOTE: depuis 2026-05-01)
    """
    from app.shared.sdtc.bareme import calcul_point_contrat
    sous_fam_up = (sous_fam or "").upper()
    if "GAZ" in sous_fam_up:
        palier = car or 0
        palier2 = puissance or 0
    else:
        palier = puissance or 0
        palier2 = car or 0
    return calcul_point_contrat(
        fam=famille or "ENI",
        ss_fam=sous_fam or "",
        palier=palier,
        date_sign=str(date_sign) if date_sign else "",
        info_cplt=options_str or "",
        palier2=palier2,
    )


def _ajoute_histo_eni_etat(id_contrat: int, old_etat: int, new_etat: int,
                            date_paiement: str, op_id: int) -> None:
    """Ajoute une ligne dans adv.pgt_eni_histo_etat_ctt (transition etat)."""
    if not id_contrat:
        return
    db = get_pg_connection("adv")
    new_id = _new_id()
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_histo_auto), 0) + 1 AS n FROM adv.pgt_eni_histo_etat_ctt"
    )
    auto_n = int(auto["n"]) if auto else 1
    db.query(
        """INSERT INTO adv.pgt_eni_histo_etat_ctt
              (id_histo_auto, id_histo, id_contrat, op_saisie, date,
               old_etat, new_etat, date_paiement,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, NOW(), ?, ?, ?, ?, NOW(), 'new')""",
        (auto_n, new_id, int(id_contrat), int(op_id),
         int(old_etat) if old_etat else 0,
         int(new_etat) if new_etat else 0,
         date_paiement or "", int(op_id)),
    )


def _lookup_or_create_client(info: dict, op_id: int) -> int:
    """Wrapper autour de traiter_client (cf. WinDev traiterClient).

    Historiquement ce helper faisait une dedup gsm+nom ad hoc. Il
    delegue maintenant a la procedure commune traiter_client qui
    fait :
    - normalisation (mail lower, gsm formate, sans accent...)
    - geocodage adresse (best-effort api-adresse.data.gouv.fr)
    - dedup par mail (WinDev : matche si HNbEnr=1)
    - INSERT si absent (avec force id fourni ou genere)

    Le mapping des cles preserve la compatibilite avec les appelants
    existants ('adresse' -> 'adresse1', 'cplt' -> 'adresse2').
    """
    try:
        from app.intranets.adm.services.import_helpers_common import (
            traiter_client,
        )
        return traiter_client(
            info_client={
                "nom": info.get("nom") or "",
                "prenom": info.get("prenom") or "",
                "adresse1": info.get("adresse1") or info.get("adresse") or "",
                "adresse2": info.get("adresse2") or info.get("cplt") or "",
                "cp": info.get("cp") or "",
                "ville": info.get("ville") or "",
                "gsm": info.get("gsm") or "",
                "mail": info.get("mail") or "",
                "tel": info.get("tel") or "",
                "date_naiss": info.get("date_naiss"),
                "op_saisie": op_id, "modif_op": op_id,
            },
            force_maj=False, op_id=op_id,
        ) or 0
    except Exception:
        return 0


def _create_eni_contrat(td: dict, op_id: int) -> int:
    """Cree un eni_contrat depuis un dict 'à créer' (cf TableImportCttJournENI).
    Retourne id_contrat cree."""
    db = get_pg_connection("adv")
    id_contrat = _new_id()
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_contrat_auto), 0) + 1 AS n FROM adv.pgt_eni_contrat"
    )
    auto_n = int(auto["n"]) if auto else 1

    date_sign = td.get("date_signature")
    if isinstance(date_sign, str):
        date_sign = _parse_date_fr(date_sign)
    date_crea = td.get("date_crea")
    if isinstance(date_crea, str):
        date_crea = _parse_datetime_fr(date_crea)

    # Calcul nb_points a la creation (cf. WinDev L513 :
    # ENI_contrat.nbPoints = TableImportCttJournENI[ind].NBPts)
    nb_pts = 0.0
    id_produit = int(td.get("id_produit") or 0)
    if id_produit:
        prod = db.query_one(
            """SELECT famille, sous_fam FROM adv.pgt_eni_produit
                WHERE id_produit = ? LIMIT 1""",
            (id_produit,),
        )
        if prod:
            note = float(td.get("note") or 0)
            options_parts = [f"NOTE:{note}"]
            if td.get("mandat_rib"): options_parts.append("RIB")
            if td.get("opt_mail"): options_parts.append("MAIL")
            if td.get("opt_maint"): options_parts.append("MAINT")
            options_str = "//".join(options_parts) + "//"
            try:
                nb_pts = float(_calcul_points_eni(
                    prod.get("famille") or "", prod.get("sous_fam") or "",
                    int(td.get("car") or 0),
                    date_sign, options_str,
                    int(td.get("puissance") or 0),
                ) or 0)
            except Exception:
                nb_pts = 0.0

    db.query(
        """INSERT INTO adv.pgt_eni_contrat
              (id_contrat_auto, id_contrat, id_client, id_salarie, id_ste,
               num_bs, id_produit, id_etat_contrat,
               date_signature, gaz_car_relevee, gaz_car_declaree,
               elec_puissance, gaz_actif, elec_actif,
               op_saisie, date_saisie, non_call, code_enr,
               notation, notation_info, nb_points, opt_mandat,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, ?,
                   ?, ?, ?,
                   ?, ?, ?,
                   ?, ?, ?,
                   ?, ?, ?, '',
                   ?, ?, ?, ?,
                   ?, NOW(), 'new')""",
        (auto_n, id_contrat,
         int(td.get("id_client") or 0),
         int(td.get("id_salarie") or 0),
         int(td.get("id_ste") or 0),
         td.get("num_bs") or "",
         int(td.get("id_produit") or 0),
         int(td.get("etat_contrat") or 0),
         date_sign, int(td.get("car") or 0), int(td.get("car") or 0),
         int(td.get("puissance") or 0),
         bool(td.get("gaz_actif")), bool(td.get("elec_actif")),
         int(op_id), date_crea or datetime.now(),
         bool(td.get("non_call")),
         float(td.get("note") or 0), td.get("info_note") or "",
         nb_pts, bool(td.get("mandat_rib")),
         int(op_id)),
    )

    # Cree la ligne option vide (defaults)
    auto_o = db.query_one(
        "SELECT COALESCE(MAX(id_contrat_option_auto), 0) + 1 AS n FROM adv.pgt_eni_contrat_option"
    )
    auto_o_n = int(auto_o["n"]) if auto_o else 1
    db.query(
        """INSERT INTO adv.pgt_eni_contrat_option
              (id_contrat_option_auto, id_contrat, num_bs,
               opt_mail, opt_entretien, opt_hp_hc,
               opt_index_gaz, opt_index_elec, opt_delai_retra,
               opt_energie_verte_gaz, opt_energie_verte_elec,
               opt_deja_client_eni, opt_reforestation,
               opt_optin_commercial, opt_e_facture, opt_e_communication,
               opt_pdc, opt_accept_com_parte, opt_consent_consult_distri,
               opt_protection,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, FALSE, FALSE, FALSE,
                   FALSE, FALSE, FALSE, FALSE,
                   FALSE, FALSE, FALSE,
                   FALSE, FALSE, FALSE,
                   FALSE,
                   ?, NOW(), 'new')""",
        (auto_o_n, id_contrat, td.get("num_bs") or "",
         bool(td.get("opt_mail")), bool(td.get("opt_maint")),
         bool(td.get("opt_hphc")), int(op_id)),
    )
    return id_contrat


def _lookup_vendeur_by_nom_prenom(nom_complet: str) -> tuple[int, int]:
    """Cherche un salarie par 'NOM Prenom' ou variantes.
    Reproduit les 3 strategies WinDev (ReqVendeurByConcatPrenomNom
    avec retry '-'->' ', puis ReqVendeurByDebPrenomFinNom).

    Retourne (id_salarie, id_ste). 0,0 si introuvable."""
    if not nom_complet or not nom_complet.strip():
        return (0, 0)

    db = get_pg_connection("rh")

    def _query_concat(pattern: str) -> list[dict]:
        # Match sur CONCAT(nom, prenom) ou CONCAT(prenom, nom) avec pattern LIKE
        return db.query(
            """SELECT s.id_salarie, e.id_ste
                 FROM rh.pgt_salarie s
                 LEFT JOIN rh.pgt_salarie_embauche e ON e.id_salarie = s.id_salarie
                WHERE (UPPER(CONCAT(s.nom, '%', s.prenom)) LIKE ?
                    OR UPPER(CONCAT(s.prenom, '%', s.nom)) LIKE ?)
                  AND (s.modif_elem IS NULL OR s.modif_elem NOT LIKE '%suppr%')
                LIMIT 2""",
            (pattern, pattern),
        ) or []

    def _normalise_pattern(cell: str) -> str:
        """Cf. WinDev : Majuscule + Remplace('-','%') + '(' -> '%' etc."""
        s = cell.upper()
        for c in ("-", "'", " "):
            s = s.replace(c, "%")
        while "%%" in s:
            s = s.replace("%%", "%")
        return f"%{s}%"

    # Strategie 1 : concat nom%prenom avec substitutions
    rows = _query_concat(_normalise_pattern(nom_complet))
    if len(rows) == 1:
        return (int(rows[0]["id_salarie"]),
                int(rows[0].get("id_ste") or 0))

    # Strategie 2 : retry avec remplacement tirets par espaces (cf. WinDev)
    if "-" in nom_complet:
        alt = nom_complet.replace("-", " ")
        rows = _query_concat(_normalise_pattern(alt))
        if len(rows) == 1:
            return (int(rows[0]["id_salarie"]),
                    int(rows[0].get("id_ste") or 0))

    # Strategie 3 : ReqVendeurByDebPrenomFinNom : prenom au debut,
    # nom en fin (ExtraitChaine DepuisDebut / DepuisFin)
    parts = nom_complet.strip().split()
    if len(parts) >= 2:
        prenom = parts[0].upper()
        nom = parts[-1].upper()
        rows = db.query(
            """SELECT s.id_salarie, e.id_ste
                 FROM rh.pgt_salarie s
                 LEFT JOIN rh.pgt_salarie_embauche e ON e.id_salarie = s.id_salarie
                WHERE UPPER(s.nom) LIKE ? AND UPPER(s.prenom) LIKE ?
                  AND (s.modif_elem IS NULL OR s.modif_elem NOT LIKE '%suppr%')
                LIMIT 2""",
            (f"{nom}%", f"{prenom}%"),
        ) or []
        if len(rows) == 1:
            return (int(rows[0]["id_salarie"]),
                    int(rows[0].get("id_ste") or 0))
    return (0, 0)


def _import_journalier_eni(
    p: ImportEniParams, file_bytes: bytes, op_id: int,
) -> ImportEniResult:
    """Import 'Base Journaliere ENI' : pour chaque ligne du fichier,
    cree le contrat s'il n'existe pas ou met a jour si necessaire.

    Phase 1 : detection only :
      - Si contrat n'existe pas : flag 'a creer' + recherche vendeur
        + determination idProduit + etat initial
      - Si contrat existe : compare valeurs Excel vs BDD, si modif et
        etat eligible (type_etat<=2 ou etat=70) : flag 'a modifier'
        + nouvel etat
    Pas de creation ni MAJ effective en phase 1."""
    from openpyxl import load_workbook

    label = TYPE_LABELS.get(4, "Base journalière ENI")
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return ImportEniResult(
            ok=False, type_import=4, type_label=label,
            simulation=p.simulation, resume=ImportEniResume(),
            message=f"Lecture Excel KO : {e}",
        )
    ws = wb.active

    cols = {k: _col_letter_to_index(v) for k, v in COLS_BJ_ENI.items()}

    resume = ImportEniResume()
    journ_eni: list[dict] = []

    db = get_pg_connection("adv")
    n_rows = ws.max_row or 0
    nb_creer = 0
    nb_modif = 0
    nb_inchanges = 0

    for i in range(2, n_rows + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"]).upper()
        if not num_contrat:
            continue

        date_sign_s = _cell(ws, i, cols["date_signature"])
        date_sign = _parse_date_fr(date_sign_s)
        date_crea = _parse_datetime_fr(_cell(ws, i, cols["date_heure_crea"]))
        lib_statut = _cell(ws, i, cols["statut"])
        vend_raw = _cell(ws, i, cols["vendeur_nom"])
        vendeur_nom = vend_raw.split("/")[0].strip() if vend_raw else ""

        offre = _cell(ws, i, cols["offre"])
        type_offre = _cell(ws, i, cols["type_offre"]).strip()
        car = _parse_int(_cell(ws, i, cols["car"]))
        puiss_s = _cell(ws, i, cols["puissance"])
        puiss = _parse_int(puiss_s.split(".")[0]) if puiss_s else 0
        type_comptage = _cell(ws, i, cols["type_comptage"])
        opt_hphc = "HPHC" in type_comptage.upper()
        _addon_up = _cell(ws, i, cols["addon"]).upper()
        opt_maint = ("PCK_HOME_P" in _addon_up) or ("HSV_PACK_PM" in _addon_up)
        opt_mail = "MAIL" in _cell(ws, i, cols["opt_mail"]).upper()
        mandat_rib = "OUI" in _cell(ws, i, cols["mandat_rib"]).upper()
        note_g = _parse_int(_cell(ws, i, cols["note_globale"]))
        note_s5 = note_g / 2 if note_g else 0
        info_note = _cell(ws, i, cols["info_notation"])
        client_cp = _cell(ws, i, cols["client_cp"])
        client_gsm = _cell(ws, i, cols["client_gsm"])

        id_produit = _id_produit_from_type_offre(type_offre)

        # Lookup contrat (sans opt_mandat : colonne absente du schema PG)
        r = db.query_one(
            """SELECT id_contrat, id_salarie, id_etat_contrat,
                      date_signature, date_saisie,
                      gaz_car_relevee, elec_puissance,
                      gaz_actif, elec_actif, non_call,
                      notation, notation_info
                 FROM adv.pgt_eni_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_contrat,),
        )

        if not r:
            # ---- CONTRAT N'EXISTE PAS : a CREER ----
            # Lookup TkCall pour peupler les coordonnees client + eventuel
            # override id_vendeur + non_call = FALSE (cf. WinDev
            # ReqTkCall_ByNumCtt).
            tk_client = _lookup_tk_call_eni(num_contrat)
            non_call = True
            if tk_client:
                non_call = False

            # Recherche vendeur
            id_vendeur, id_ste = _lookup_vendeur_by_nom_prenom(vendeur_nom)
            # Si XLS ne donne pas de vendeur mais TkCall a un id_salarie,
            # utilise celui du ticket (cf. WinDev)
            if id_vendeur == 0 and tk_client.get("id_salarie"):
                id_vendeur = int(tk_client["id_salarie"])

            etat, gaz_actif, elec_actif = _etat_contrat_journ_eni(
                lib_statut, id_produit, offre, non_call=non_call,
            )

            # Peuple _client_info depuis TkCall (avant : nom/prenom
            # etaient toujours vides -> clients cree avec nom=null)
            client_info = {
                "nom": tk_client.get("nom") or "",
                "prenom": tk_client.get("prenom") or "",
                "adresse1": tk_client.get("adresse") or "",
                "adresse2": tk_client.get("cplt") or "",
                "cp": tk_client.get("cp") or client_cp,
                "ville": tk_client.get("ville") or "",
                "gsm": tk_client.get("gsm") or client_gsm,
                "mail": tk_client.get("mail") or "",
                "date_naiss": tk_client.get("date_naiss"),
            }

            journ_eni.append({
                "NumCtt": num_contrat,
                "Statut": "à créer",
                "DateSigne": str(date_sign or ""),
                "DateCrea": str(date_crea or ""),
                "NomVend": vendeur_nom,
                "IdSalarie": id_vendeur,
                "Vendeur trouvé": id_vendeur > 0,
                "Société": id_ste,
                "LibProduit (id)": id_produit,
                "TypeOffre": type_offre,
                "CAR": car,
                "Puissance": puiss,
                "OptHpHc": opt_hphc,
                "OptMail": opt_mail,
                "OptMaint": opt_maint,
                "OptMandatRib": mandat_rib,
                "Note": note_s5,
                "InfoNote": info_note,
                "EtatContrat (initial)": etat,
                "GazActif": gaz_actif,
                "ElecActif": elec_actif,
                "NonCALL": non_call,
                "CltCP": client_cp,
                "CltGSM": client_gsm,
                # ---- payload pour creation en mode prod ----
                "_client_info": client_info,
                "_contrat_data": {
                    "num_bs": num_contrat,
                    "id_salarie": id_vendeur,
                    "id_ste": id_ste,
                    "id_produit": id_produit,
                    "etat_contrat": etat,
                    "date_signature": date_sign,
                    "date_crea": date_crea,
                    "car": car,
                    "puissance": puiss,
                    "gaz_actif": gaz_actif,
                    "elec_actif": elec_actif,
                    "non_call": non_call,
                    "note": note_s5,
                    "info_note": info_note,
                    "opt_mail": opt_mail,
                    "opt_maint": opt_maint,
                    "opt_hphc": opt_hphc,
                    "mandat_rib": mandat_rib,
                },
            })
            nb_creer += 1
            continue

        # ---- CONTRAT EXISTE : compare ----
        id_contrat = int(r["id_contrat"])
        non_call_bdd = bool(r.get("non_call"))
        etat_actuel = int(r.get("id_etat_contrat") or 0)
        car_bdd = int(r.get("gaz_car_relevee") or 0)
        puiss_bdd = int(r.get("elec_puissance") or 0)

        # Options actuelles
        opt = db.query_one(
            """SELECT opt_entretien, opt_mail
                 FROM adv.pgt_eni_contrat_option WHERE id_contrat = ? LIMIT 1""",
            (id_contrat,),
        ) or {}

        # Differences detectees
        diffs: list[str] = []
        if date_crea and r.get("date_saisie") != date_crea:
            diffs.append(f"date_saisie {r.get('date_saisie')} -> {date_crea}")
        if puiss > 0 and puiss != puiss_bdd:
            diffs.append(f"puissance {puiss_bdd} -> {puiss}")
        if car > 0 and car != car_bdd:
            diffs.append(f"car_gaz {car_bdd} -> {car}")
        if bool(opt.get("opt_entretien")) != opt_maint:
            diffs.append(f"opt_entretien -> {opt_maint}")
        if bool(opt.get("opt_mail")) != opt_mail:
            diffs.append(f"opt_mail -> {opt_mail}")

        # Eligibilite MAJ etat
        etat_info = db.query_one(
            """SELECT id_type_etat FROM adv.pgt_eni_etat_contrat
                WHERE id_etat = ? LIMIT 1""",
            (etat_actuel,),
        )
        id_type_etat = int(etat_info.get("id_type_etat") or 0) if etat_info else 0
        eligible_etat = (id_type_etat <= 2) or etat_actuel == 70

        nouvel_etat = etat_actuel
        if eligible_etat:
            nouvel_etat, _, _ = _etat_contrat_journ_eni(
                lib_statut, id_produit, offre, non_call=non_call_bdd,
                etat_actuel=etat_actuel,
            )
            if nouvel_etat != etat_actuel:
                diffs.append(f"etat {etat_actuel} -> {nouvel_etat}")

        if diffs:
            journ_eni.append({
                "NumCtt": num_contrat,
                "Statut": "à modifier",
                "DateSigne": str(r.get("date_signature") or ""),
                "EtatActuel": etat_actuel,
                "TypeEtat": id_type_etat,
                "NouvelEtat": nouvel_etat,
                "Diffs": " | ".join(diffs),
                "LibProduit (id)": id_produit,
                "TypeOffre": type_offre,
                "OptMail Excel": opt_mail,
                "OptMaint Excel": opt_maint,
                # Payload pour MAJ en prod
                "_maj_data": {
                    "id_contrat": id_contrat,
                    "id_produit": id_produit,
                    "date_saisie": date_crea,
                    "date_signature": r.get("date_signature"),
                    "puissance": puiss,
                    "car": car,
                    "opt_maint": opt_maint,
                    "opt_mail": opt_mail,
                    "mandat_rib": mandat_rib,
                    "note": note_s5,
                    "etat_actuel": etat_actuel,
                    "nouvel_etat": nouvel_etat if eligible_etat else etat_actuel,
                    "etat_change": eligible_etat and (nouvel_etat != etat_actuel),
                },
            })
            nb_modif += 1
        else:
            nb_inchanges += 1

    wb.close()

    # -- PASSE PROD : creation + MAJ si pas simulation --
    nb_crees = 0
    nb_majs = 0
    if not p.simulation:
        for row in journ_eni:
            try:
                if row.get("Statut") == "à créer":
                    ci = row.pop("_client_info", {})
                    cd = row.pop("_contrat_data", {})
                    if ci:
                        cd["id_client"] = _lookup_or_create_client(ci, op_id)
                    new_id = _create_eni_contrat(cd, op_id)
                    row["IdContratCree"] = new_id
                    nb_crees += 1
                elif row.get("Statut") == "à modifier":
                    md = row.pop("_maj_data", {})
                    if md and md.get("id_contrat"):
                        _apply_maj_eni_journ(md, op_id)
                        nb_majs += 1
            except Exception as e:
                row["Erreur"] = str(e)

    resume.nb_modifications = nb_modif
    resume.nb_valides = nb_creer

    res = ImportEniResult(
        ok=True, type_import=4, type_label=label,
        simulation=p.simulation, resume=resume,
        contrat_import_journ_eni=journ_eni,
        message=(
            f"{n_rows - 1} ligne(s) lue(s) | "
            f"À créer : {nb_creer} | À modifier : {nb_modif} | "
            f"Inchangés : {nb_inchanges}. "
            + (f"PRODUCTION : {nb_crees} créés, {nb_majs} MAJ."
               if not p.simulation else "(SIMULATION)")
        ),
    )
    _attach_xlsx_and_mail(res, op_id, "ImportJournPLENITUDE")
    return res


def _apply_maj_eni_journ(md: dict, op_id: int) -> None:
    """MAJ eni_contrat + eni_contrat_option pour un contrat existant
    (ImportJournalierENI). Historise si l'etat change + recalcul nb_points."""
    db = get_pg_connection("adv")
    id_contrat = int(md["id_contrat"])

    # Recalcul nb_points (cf. WinDev : ENI_contrat.nbPoints =
    # calculPointContrat(...))
    nb_pts = None
    id_produit = int(md.get("id_produit") or 0)
    if id_produit:
        prod = db.query_one(
            """SELECT famille, sous_fam FROM adv.pgt_eni_produit
                WHERE id_produit = ? LIMIT 1""",
            (id_produit,),
        )
        if prod:
            # Options string (cf. WinDev : "NOTE:x//RIB//MAIL//MAINT//")
            note = float(md.get("note") or 0)
            options_str_parts = [f"NOTE:{note}"]
            if md.get("mandat_rib"): options_str_parts.append("RIB")
            if md.get("opt_mail"): options_str_parts.append("MAIL")
            if md.get("opt_maint"): options_str_parts.append("MAINT")
            options_str = "//".join(options_str_parts) + "//"
            nb_pts = _calcul_points_eni(
                prod.get("famille") or "", prod.get("sous_fam") or "",
                int(md.get("car") or 0),
                md.get("date_signature"),
                options_str,
                int(md.get("puissance") or 0),
            )

    sets = []
    params: list = []
    if md.get("date_saisie"):
        sets.append("date_saisie = ?")
        params.append(md["date_saisie"])
    if md.get("puissance"):
        sets.append("elec_puissance = ?")
        params.append(int(md["puissance"]))
    if md.get("car"):
        sets.append("gaz_car_relevee = ?")
        params.append(int(md["car"]))
    if md.get("etat_change"):
        sets.append("id_etat_contrat = ?")
        params.append(int(md["nouvel_etat"]))
    if md.get("mandat_rib") is not None:
        # cf. WinDev : ENI_contrat.Opt_Mandat = MandatRib
        sets.append("opt_mandat = ?")
        params.append(bool(md["mandat_rib"]))
    if nb_pts is not None:
        sets.append("nb_points = ?")
        params.append(float(nb_pts))

    if sets:
        sets.append("modif_date = NOW()")
        sets.append("modif_op = ?")
        sets.append("modif_elem = 'modif'")
        params.append(int(op_id))
        params.append(id_contrat)
        db.query(
            f"""UPDATE adv.pgt_eni_contrat
                  SET {', '.join(sets)}
                WHERE id_contrat = ?""",
            tuple(params),
        )

    # MAJ option
    db.query(
        """UPDATE adv.pgt_eni_contrat_option
              SET opt_entretien = ?, opt_mail = ?,
                  modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
            WHERE id_contrat = ?""",
        (bool(md.get("opt_maint")), bool(md.get("opt_mail")),
         int(op_id), id_contrat),
    )

    if md.get("etat_change"):
        _ajoute_histo_eni_etat(
            id_contrat, int(md["etat_actuel"]), int(md["nouvel_etat"]),
            "", op_id,
        )


# ---------------------------------------------------------------------------
# Export XLSX + envoi mail BO (commun aux 4 types)
# ---------------------------------------------------------------------------


def _resume_to_lines(resume: ImportEniResume) -> list[tuple[str, int]]:
    return [
        ("NB Validés", resume.nb_valides),
        ("NB Déjà Statués", resume.nb_deja_statues),
        ("NB Doublons", resume.nb_doublons),
        ("NB Introuvables", resume.nb_introuvables),
        ("NB Décommissions", resume.nb_decommissions),
        ("NB Résiliés", resume.nb_resilies),
        ("NB Modifications", resume.nb_modifications),
        ("NB Erreurs Mails", resume.nb_erreurs_mails),
        ("NB Erreurs Entretien", resume.nb_erreurs_entretien),
        ("NB Contrats Hors Délai", resume.nb_contrats_hors_delai),
        ("NB Erreurs Offres", resume.nb_erreurs_offres),
        ("NB Erreurs Type comptage", resume.nb_erreurs_type_comptage),
        ("NB Erreurs Énergie Verte", resume.nb_erreurs_energie_verte),
        ("NB Erreurs CAR", resume.nb_erreurs_car),
        ("NB Erreurs PUISS", resume.nb_erreurs_puiss),
        ("NB Erreurs Reforestation", resume.nb_erreurs_reforest),
        ("NB Erreurs Protection", resume.nb_erreurs_protection),
        ("NB Erreurs Mandat", resume.nb_erreurs_mandat),
    ]


def _build_xlsx_import_eni(res: ImportEniResult) -> bytes:
    """Genere un XLSX recapitulatif avec :
      - Feuille 1 : Resume (compteurs)
      - Feuille N : une feuille par tableau non vide (cf. WinDev 8 feuilles
        pour RUN Valides, 5 pour RUN Resils, 3 pour Journ CALL, 1 pour
        Journ ENI). On dynamise selon ce qui est rempli."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    # Feuille 1 : Resume
    ws = wb.active
    ws.title = "Résumé"
    ws.append(["Indicateur", "Nombre"])
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="17494E")
    header_font = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for lbl, n in _resume_to_lines(res.resume):
        ws.append([lbl, n])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    sections: list[tuple[str, list[dict]]] = [
        ("Contrats modifiés", res.contrats_modifies),
        ("Contrats non trouvés", res.contrats_non_trouves),
        ("Contrats RUN", res.contrats_run),
        ("Problème Vendeur", res.pb_vendeur),
        ("Contrat Import Journ ENI", res.contrat_import_journ_eni),
    ]
    for title, rows in sections:
        if not rows:
            continue
        sheet = wb.create_sheet(title=title[:31])
        keys = list(rows[0].keys())
        sheet.append(keys)
        for c in sheet[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for r in rows:
            sheet.append([
                str(r.get(k, "")) if r.get(k) is not None else ""
                for k in keys
            ])
        # Largeur auto basique
        for i, k in enumerate(keys, start=1):
            col = chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"
            sheet.column_dimensions[col].width = max(12, min(40, len(k) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _send_mail_import_eni(
    res: ImportEniResult, xlsx_bytes: bytes, xlsx_name: str, op_id: int,
) -> bool:
    """Envoie le mail BO recap a l'op connecte + intranet@omaya.fr.
    Mail BO (pas RH) : expediteur affiche 'intranet@omaya.fr'."""
    from app.shared.notifications.mail import envoi_mail
    from app.core.database.pg import get_pg_connection

    # Mail de l'op connecte
    db = get_pg_connection("rh")
    r = db.query_one(
        """SELECT mail FROM rh.pgt_salarie_coordonnees
            WHERE id_salarie = ? LIMIT 1""",
        (int(op_id),),
    )
    op_mail = (r.get("mail") if r else "") or ""

    destinataires = []
    if op_mail:
        destinataires.append(op_mail)
    # Toujours envoyer une copie a intranet (mail BO support)
    intranet_mail = "intranet@omaya.fr"
    cc = []
    if intranet_mail not in destinataires:
        cc.append(intranet_mail)

    if not destinataires and not cc:
        return False

    if not destinataires:
        destinataires = cc
        cc = []

    sujet_prefix = "SIMULATION : " if res.simulation else ""
    sujet = f"{sujet_prefix}Importation {res.type_label} du {date.today().strftime('%d/%m/%Y')}"

    resume_html = "".join(
        f"<li><strong>{lbl} :</strong> {n}</li>"
        for lbl, n in _resume_to_lines(res.resume) if n > 0
    )
    html = (
        "<p>Bonjour,</p>"
        f"<p>Fin importation le : {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>"
        f"<p><strong>{res.message}</strong></p>"
        f"<ul>{resume_html}</ul>"
        "<p>Service Importation EXOSPHERE</p>"
    )

    try:
        return envoi_mail(
            sujet=sujet, html=html,
            destinataires=destinataires,
            cc=cc,
            expediteur="intranet@omaya.fr",
            attachments=[(xlsx_name, xlsx_bytes)],
        )
    except Exception:
        return False


def _attach_xlsx_and_mail(
    res: ImportEniResult, op_id: int, file_prefix: str = "ImportENI",
) -> None:
    """Genere le XLSX, l'attache en base64 dans res.xlsx_b64, envoie le
    mail. Modifie res in-place."""
    import base64
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    suffix = "_SIMU" if res.simulation else ""
    xlsx_name = f"{file_prefix}_{ts}{suffix}.xlsx"
    try:
        xlsx_bytes = _build_xlsx_import_eni(res)
    except Exception:
        return
    res.xlsx_name = xlsx_name
    res.xlsx_b64 = base64.b64encode(xlsx_bytes).decode("ascii")
    res.mail_envoye = _send_mail_import_eni(
        res, xlsx_bytes, xlsx_name, op_id,
    )
