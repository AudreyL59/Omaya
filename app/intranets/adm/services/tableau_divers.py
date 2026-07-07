"""
Service Fen_TableauDivers - Generation de tableaux divers.

3 fonctionnalites :
- lister_demandes : liste les salaries embauches entre Du et Au + calcul
  du poste et de l'affectation (agence/equipe)
- generer_valandre : XLSX 6 colonnes pour accreditation Valandre
- generer_comptable : XLSX 20 colonnes pour le comptable
"""
import io as _io
import logging
import re
import unicodedata
from datetime import date, datetime
from typing import Optional

from app.core.database.pg import get_pg_connection
from app.intranets.adm.schemas.tableau_divers import (
    DemandeRow, GenererComptableParams, GenererValandreParams,
    ListerDemandesParams, ListerDemandesResult,
)

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------

def _clean_id(v) -> str:
    if v is None:
        return ""
    try:
        n = int(v)
        return str(n) if n else ""
    except (TypeError, ValueError):
        return ""


def _dmY(v) -> str:
    """Date ISO / datetime -> 'DD/MM/YYYY' (WinDev maskDateSysteme)."""
    if v is None:
        return ""
    s = str(v)[:10]
    if s.startswith("1900") or s.startswith("0000") or not s:
        return ""
    if len(s) == 10 and s[4] == "-":
        return f"{s[8:10]}/{s[5:7]}/{s[:4]}"
    return s


def _extraire_num_voie(adresse1: str) -> tuple[str, str]:
    """Cf. WinDev : si le 1er token de l'adresse est numerique, split.
    Retourne (numero, voie_sans_num). Sinon ('', adresse_complete)."""
    if not adresse1:
        return ("", "")
    parts = adresse1.strip().split(" ", 1)
    if len(parts) >= 2 and parts[0].isdigit():
        return (parts[0], parts[1].strip())
    return ("", adresse1.strip())


# --------------------------------------------------------------------
# Affectation (reutilise le pattern de paies_bs)
# --------------------------------------------------------------------

def _load_agence_equipe(id_salarie: int) -> tuple[str, str]:
    """Retourne (agence, equipe) courante d'un salarie (segment actuel)."""
    rh = get_pg_connection("rh")
    try:
        rows = rh.query(
            """SELECT o.lib_orga, o.id_type_niveau_orga, o.id_parent
                 FROM pgt_salarie_organigramme so
                 JOIN pgt_organigramme o
                      ON o.idorganigramme = so.idorganigramme
                WHERE so.id_salarie = ?
                  AND (so.modif_elem IS NULL
                       OR so.modif_elem NOT LIKE '%suppr%')
                  AND (so.date_fin IS NULL OR so.date_fin >= CURRENT_DATE)
                  AND (so.date_debut IS NULL OR so.date_debut <= CURRENT_DATE)
                ORDER BY so.date_debut DESC NULLS LAST""",
            (int(id_salarie),),
        ) or []
    except Exception:
        return ("", "")
    agence = ""
    equipe = ""
    id_agence = 0
    for s in rows:
        lvl = s.get("id_type_niveau_orga")
        lib = (s.get("lib_orga") or "").strip()
        if lvl == 3 and not agence:
            agence = lib
        elif lvl == 4 and not equipe:
            equipe = lib
            id_agence = int(s.get("id_parent") or 0)
    if equipe and not agence and id_agence:
        rh = get_pg_connection("rh")
        try:
            r = rh.query_one(
                "SELECT lib_orga FROM pgt_organigramme WHERE idorganigramme = ?",
                (id_agence,),
            )
            if r:
                agence = (r.get("lib_orga") or "").strip()
        except Exception:
            pass
    return (agence, equipe)


def _norm_lib(s: str) -> str:
    """Retire accents + uppercase pour tests 'contains' insensibles."""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    ascii_s = "".join(c for c in nfkd if not unicodedata.combining(c))
    return ascii_s.upper()


def _poste_formate(
    lib_poste: str,
    resp_equipe: bool,
    agence_brute: str,
    equipe_brute: str,
) -> tuple[str, str, str]:
    """Cf. WinDev ListerDemandes : mappe lib_poste -> (poste_lib, agence, equipe).

    - MANAGER      -> 'Chef d equipe'
    - AGENCE       -> 'Responsable agence' (agence = 2eme niveau)
    - REGION       -> 'Dir. Co. Partenaire' + fix NORD->LILLE, OUEST->RENNES
    - ADM ou TECH  -> 'Resp. BO Eni' si resp_equipe sinon 'Agent BO Eni' + agence='LILLE'
    - defaut       -> 'Vendeur'
    """
    up = _norm_lib(lib_poste)
    # Agence brute : retire prefixe 'Agence '
    ag = re.sub(r"^Agence\s+", "", agence_brute or "", flags=re.IGNORECASE)
    # Equipe : garde jusqu'au premier '/' + retire prefixe 'Equipe '
    eq = (equipe_brute or "").split("/", 1)[0]
    eq = re.sub(r"^Equipe\s+", "", eq, flags=re.IGNORECASE).strip()
    ag = ag.strip()

    if "MANAGER" in up:
        return ("Chef d'équipe", ag, eq)
    if "AGENCE" in up:
        return ("Responsable agence", ag, "")
    if "REGION" in up:
        # Agence prend le 2eme niveau de l'affectation (region -> ville)
        # On applique le fix NORD/OUEST
        ag_up = _norm_lib(ag)
        if "NORD" in ag_up:
            ag = "LILLE"
        elif "OUEST" in ag_up:
            ag = "RENNES"
        return ("Dir. Co. Partenaire", ag, "")
    if "ADM" in up or "TECH" in up:
        poste = "Resp. BO Eni" if resp_equipe else "Agent BO Eni"
        return (poste, "LILLE", "")
    return ("Vendeur", ag, eq)


# --------------------------------------------------------------------
# Lister les demandes
# --------------------------------------------------------------------

def lister_demandes(p: ListerDemandesParams) -> ListerDemandesResult:
    """Cf. WinDev ListerDemandes() : liste les salaries embauches
    entre Du et Au (en_activite=TRUE) + calcul poste et affectation.
    """
    if not p.du or not p.au or len(p.du) < 10 or len(p.au) < 10:
        return ListerDemandesResult(
            ok=False, message="Dates invalides (YYYY-MM-DD)",
        )
    rh = get_pg_connection("rh")
    try:
        rows = rh.query(
            """SELECT s.id_salarie, s.nom, s.prenom,
                      e.date_debut, e.id_ste, e.id_type_poste,
                      e.resp_equipe,
                      c.mail,
                      tp.lib_poste
                 FROM pgt_salarie s
                 JOIN pgt_salarie_embauche e ON e.id_salarie = s.id_salarie
                 LEFT JOIN pgt_salarie_coordonnees c ON c.id_salarie = s.id_salarie
                 LEFT JOIN pgt_type_poste tp ON tp.id_type_poste = e.id_type_poste
                WHERE e.date_debut BETWEEN ? AND ?
                  AND e.en_activite = TRUE
                  AND (s.modif_elem IS NULL OR s.modif_elem NOT LIKE '%suppr%')
                  AND (e.modif_elem IS NULL OR e.modif_elem NOT LIKE '%suppr%')
                ORDER BY s.nom, s.prenom""",
            (p.du[:10], p.au[:10]),
        ) or []
    except Exception as e:
        logger.exception("lister_demandes")
        return ListerDemandesResult(ok=False, message=f"Erreur SQL : {e}")

    lignes: list[DemandeRow] = []
    for r in rows:
        id_sal = _clean_id(r.get("id_salarie"))
        try:
            ag_brute, eq_brute = _load_agence_equipe(int(id_sal))
        except Exception:
            ag_brute, eq_brute = ("", "")
        poste, agence, equipe = _poste_formate(
            (r.get("lib_poste") or "").strip(),
            bool(r.get("resp_equipe")),
            ag_brute, eq_brute,
        )
        lignes.append(DemandeRow(
            id_salarie=id_sal,
            choix=True,
            nom=(r.get("nom") or "").strip(),
            prenom=(r.get("prenom") or "").strip(),
            poste=poste,
            email=(r.get("mail") or "").strip(),
            agence=agence,
            equipe=equipe,
            type_demande="Première demande",
        ))
    return ListerDemandesResult(
        ok=True, lignes=lignes,
        message=f"{len(lignes)} salarie(s) trouve(s)",
    )


# --------------------------------------------------------------------
# Generer XLSX Valandre EXO
# --------------------------------------------------------------------

def generer_valandre_xlsx(p: GenererValandreParams) -> tuple[str, bytes]:
    """XLSX 6 colonnes pour accreditation Valandre :
    Prenom / NOM / Mail / Societe='EXOSPHERE' / Identifiant / Nouveau mot de passe.

    Les lignes cochees (choix=True) apparaissent avec un fond vert
    (RVB 146,208,80 = #92D050).
    """
    try:
        from openpyxl import Workbook  # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill  # noqa: PLC0415
    except ImportError:
        return ("", b"")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sample"

    font = Font(name="Calibri", size=10)
    fill_vert = PatternFill(
        start_color="92D050", end_color="92D050", fill_type="solid",
    )
    # En-tetes
    headers = [
        "Prénom", "NOM", "Mail", "Société",
        "Identifiant", "Nouveau mot de passe",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = font
    # Largeurs (cf. WinDev - proportions ramenees en unites xlsx)
    widths = [15, 12, 25, 12, 16, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Lignes cochees uniquement
    r = 2
    for row in p.lignes:
        if not row.choix:
            continue
        ws.cell(row=r, column=1, value=row.prenom).font = font
        ws.cell(row=r, column=2, value=row.nom).font = font
        ws.cell(row=r, column=3, value=row.email).font = font
        ws.cell(row=r, column=4, value="EXOSPHERE").font = font
        # Fond vert sur colonnes 1-4 (cf. WinDev)
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill_vert
        r += 1

    buf = _io.BytesIO()
    wb.save(buf)
    today = datetime.now().strftime("%Y-%m-%d")
    fic_name = f"{today}_ACCREDITATION_VALANDRE.xlsx"
    return (fic_name, buf.getvalue())


# --------------------------------------------------------------------
# Generer XLSX Comptable (fichier nouveaux salaries)
# --------------------------------------------------------------------

def generer_comptable_xlsx(
    p: GenererComptableParams,
) -> tuple[str, bytes]:
    """XLSX 20 colonnes pour le comptable :
    MATRICULE / NOM / Prenom / Emploi / DATE ENTREE / N° / VOIE /
    COMPLEMENT / CP / VILLE / CODE INSEE COMMUNE / Code pays Nationalite /
    NUMERO INSEE / DATE NAISS / DEPT NAISS / LIEU / Code pays Naissance /
    DEBUT de CONTRAT / Entite / DATE SORTIE REELLE.

    Filtre : embauche entre Du et Au, id_ste <> 4 (cf. WinDev - exclusion
    d'une entite specifique).
    """
    try:
        from openpyxl import Workbook  # noqa: PLC0415
        from openpyxl.styles import Font  # noqa: PLC0415
    except ImportError:
        return ("", b"")

    if not p.du or not p.au or len(p.du) < 10 or len(p.au) < 10:
        raise ValueError("Dates invalides (YYYY-MM-DD)")

    rh = get_pg_connection("rh")
    rows = rh.query(
        """SELECT s.id_salarie, s.nom, s.prenom, s.date_naiss,
                  s.dep_naiss, s.lieu_naiss, s.nationalite, s.num_ss,
                  c.adresse1, c.adresse2, c.cp, c.ville,
                  e.date_debut,
                  ste.rs_interne,
                  tp.lib_poste,
                  ss.date_sortie_reelle
             FROM pgt_salarie s
             JOIN pgt_salarie_embauche e ON e.id_salarie = s.id_salarie
             LEFT JOIN pgt_salarie_coordonnees c ON c.id_salarie = s.id_salarie
             LEFT JOIN pgt_societe ste ON ste.id_ste = e.id_ste
             LEFT JOIN pgt_type_poste tp ON tp.id_type_poste = e.id_type_poste
             LEFT JOIN pgt_salarie_sortie ss ON ss.id_salarie = s.id_salarie
                    AND (ss.modif_elem IS NULL OR ss.modif_elem NOT LIKE '%suppr%')
            WHERE e.date_debut BETWEEN ? AND ?
              AND e.id_ste <> 4
              AND (s.modif_elem IS NULL OR s.modif_elem NOT LIKE '%suppr%')
              AND (e.modif_elem IS NULL OR e.modif_elem NOT LIKE '%suppr%')
            ORDER BY e.date_debut, s.nom, s.prenom""",
        (p.du[:10], p.au[:10]),
    ) or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Sample"
    font = Font(name="Calibri", size=10)

    headers = [
        "MATRICULE", "NOM", "Prenom", "Emploi", "DATE D'ENTREE",
        "N°", "VOIE", "COMPLEMENT", "CP", "VILLE",
        "CODE INSEE COMMUNE", "Code pays Nationalité", "NUMERO INSEE",
        "DATE NAISS", "DEPT NAISS", "LIEU", "Code pays Naissance",
        "DEBUT de CONTRAT", "Entité", "DATE SORTIE REELLE",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h).font = font

    r = 2
    for x in rows:
        lib_poste = (x.get("lib_poste") or "").strip()
        # cf. WinDev : 'Non defini' -> 'VRP multicartes'
        if lib_poste.lower() == "non défini" or lib_poste.lower() == "non defini":
            lib_poste = "VRP multicartes"

        num_voie, voie = _extraire_num_voie(x.get("adresse1") or "")

        date_deb = _dmY(x.get("date_debut"))
        vals = [
            "",                                          # MATRICULE
            (x.get("nom") or "").strip(),                # NOM
            (x.get("prenom") or "").strip(),             # Prenom
            lib_poste,                                    # Emploi
            date_deb,                                    # DATE D'ENTREE
            num_voie,                                    # N°
            voie,                                        # VOIE
            (x.get("adresse2") or "").strip(),           # COMPLEMENT
            (x.get("cp") or "").strip(),                 # CP
            (x.get("ville") or "").strip(),              # VILLE
            "",                                          # CODE INSEE COMMUNE
            (x.get("nationalite") or "").strip(),        # Code pays Nationalite
            (x.get("num_ss") or "").strip(),             # NUMERO INSEE
            _dmY(x.get("date_naiss")),                   # DATE NAISS
            str(x.get("dep_naiss") or "").strip(),       # DEPT NAISS (smallint)
            (x.get("lieu_naiss") or "").strip(),         # LIEU
            "",                                          # Code pays Naissance
            date_deb,                                    # DEBUT de CONTRAT
            (x.get("rs_interne") or "").strip(),         # Entite
            _dmY(x.get("date_sortie_reelle")),           # DATE SORTIE REELLE
        ]
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value=v).font = font
        r += 1

    buf = _io.BytesIO()
    wb.save(buf)
    today = datetime.now().strftime("%Y-%m-%d")
    fic_name = f"{today}_FICHIER_NOUVEAUX_SALARIES.xlsx"
    return (fic_name, buf.getvalue())
