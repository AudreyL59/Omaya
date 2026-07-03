"""Service Fen_ImportNotation (ADM Imports Bases -> Import notations).

Importe les notations (Notation + NotationInfo) sur les contrats depuis
un fichier Excel. Le mapping colonnes + le bareme (notes_sur) varient
selon le partenaire :
  - ENI : Num H, Note F, Info G, sur 5
  - SFR : Num A, Note B, Info C, sur 10
  - autre : pas de mapping defini (notes_sur = 1)

La note est NORMALISEE sur 5 avant ecriture en BDD :
    note_db = (note_xlsx / notes_sur) * 5
Si note = 0, on met 0.01 (pour distinguer 'note zero' de 'pas de note').
"""

from __future__ import annotations

import base64
import io
from datetime import datetime
from typing import Optional

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection
from app.shared.recrutement.services.recherche_cv import _int, _str
from app.intranets.adm.services.import_eni import _col_letter_to_index, _cell


PARTENAIRES_SUPPORTES = ["eni", "iag", "oen", "pro", "sfr", "str", "val"]


# Mapping par defaut selon partenaire (cf. WinDev combo selon MoiMeme).
# ENI et SFR ont des mappings XLS specifiques ; pour les autres partenaires
# on utilise le layout standard 'export Provad' (Num A, Note B, Info C)
# sur 10 par convention. Le user peut override via les params col_*.
_LAYOUT_STANDARD = {
    "col_num": "A", "col_note": "B", "col_info": "C", "notes_sur": 10,
}
MAPPING_PAR_PARTENAIRE = {
    "eni": {"col_num": "H", "col_note": "F", "col_info": "G", "notes_sur": 5},
    "sfr": {"col_num": "A", "col_note": "B", "col_info": "C", "notes_sur": 10},
    "iag": dict(_LAYOUT_STANDARD),
    "pro": dict(_LAYOUT_STANDARD),
    "oen": dict(_LAYOUT_STANDARD),
    "str": dict(_LAYOUT_STANDARD),
    "val": dict(_LAYOUT_STANDARD),
    "tlc": dict(_LAYOUT_STANDARD),
}


class ImportNotationParams(BaseModel):
    partenaire: str
    col_num_contrat: str = ""
    col_note: str = ""
    col_info: str = ""
    notes_sur: float = 5.0
    simulation: bool = True


class NotationLigne(BaseModel):
    num_bs: str
    vendeur: str = ""
    agence: str = ""
    equipe: str = ""
    date_signature: str = ""
    note_normalisee: float = 0.0
    info_notes: str = ""
    statut: str = ""


class ImportNotationResult(BaseModel):
    ok: bool
    partenaire: str
    nb_lignes: int = 0
    nb_importees: int = 0
    nb_introuvables: int = 0
    nb_erreurs: int = 0
    lignes: list[NotationLigne] = []
    erreurs: list[dict] = []
    xlsx_b64: str = ""
    xlsx_name: str = ""
    message: str = ""


def get_mapping_default(partenaire: str) -> dict:
    """Retourne le mapping par defaut (col_num, col_note, col_info, notes_sur)
    selon partenaire, ou un mapping vide si non defini."""
    p = (partenaire or "").lower()
    return MAPPING_PAR_PARTENAIRE.get(
        p, {"col_num": "", "col_note": "", "col_info": "", "notes_sur": 1}
    )


def _info_salarie(id_salarie: int) -> dict:
    if not id_salarie:
        return {}
    try:
        db = get_pg_connection("rh")
        r = db.query_one(
            "SELECT nom, prenom FROM rh.pgt_salarie WHERE id_salarie = ? LIMIT 1",
            (int(id_salarie),),
        )
        return r or {}
    except Exception:
        return {}


def _affectation(id_salarie: int) -> tuple[str, str]:
    if not id_salarie:
        return ("", "")
    try:
        db = get_pg_connection("rh")
        rows = db.query(
            """SELECT o.lib_orga, o.id_type_niveau_orga
                 FROM rh.pgt_salarie_organigramme so
                 JOIN rh.pgt_organigramme o ON o.idorganigramme = so.idorganigramme
                WHERE so.id_salarie = ?
                  AND (so.modif_elem IS NULL OR so.modif_elem NOT LIKE '%suppr%')""",
            (int(id_salarie),),
        ) or []
        agence = ""; equipe = ""
        for r in rows:
            lvl = r.get("id_type_niveau_orga")
            lib = r.get("lib_orga") or ""
            if lvl == 3 and not agence:
                agence = lib
            elif lvl == 4 and not equipe:
                equipe = lib
        return (agence, equipe)
    except Exception:
        return ("", "")


def run_import_notation(
    p: ImportNotationParams, content: bytes, op_id: int,
) -> ImportNotationResult:
    """Importe les notations d'un fichier XLSX vers les contrats du
    partenaire choisi.

    Pour chaque ligne (a partir de la 2e) :
    - normalise la note : (note/notes_sur)*5 (sur 5)
    - si note=0 -> 0.01 (pour distinguer)
    - lookup contrat par num_bs
    - si trouve : UPDATE notation + notation_info + modif_date
    - sinon : ajout aux erreurs

    Retourne un XLSX recapitulatif (2 feuilles : notations + erreurs).
    """
    from openpyxl import load_workbook, Workbook
    partenaire = (p.partenaire or "").lower()
    if partenaire not in PARTENAIRES_SUPPORTES:
        return ImportNotationResult(
            ok=False, partenaire=partenaire,
            message=f"Partenaire inconnu : {p.partenaire}",
        )

    # Mapping : prend ce que le user a saisi, sinon default partenaire
    default = get_mapping_default(partenaire)
    col_num = p.col_num_contrat or default["col_num"]
    col_note = p.col_note or default["col_note"]
    col_info = p.col_info or default["col_info"]
    notes_sur = p.notes_sur if p.notes_sur > 0 else default["notes_sur"]
    if not col_num or not col_note:
        return ImportNotationResult(
            ok=False, partenaire=partenaire,
            message="Mapping colonnes incomplet (col_num + col_note requis)",
        )

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return ImportNotationResult(
            ok=False, partenaire=partenaire, message=f"Lecture XLSX : {e}",
        )
    ws = wb.active
    idx_num = _col_letter_to_index(col_num)
    idx_note = _col_letter_to_index(col_note)
    idx_info = _col_letter_to_index(col_info) if col_info else 0
    db = get_pg_connection("adv")

    lignes: list[NotationLigne] = []
    erreurs: list[dict] = []
    nb_imp = 0

    for i in range(2, (ws.max_row or 0) + 1):
        num_bs = _cell(ws, i, idx_num).upper().strip()
        if not num_bs:
            continue
        note_raw = _cell(ws, i, idx_note).replace(",", ".").strip()
        try:
            note_val = float(note_raw) if note_raw else 0.0
        except ValueError:
            note_val = 0.0
        info_notes = _cell(ws, i, idx_info) if idx_info else ""

        # Normalisation sur 5
        note_norm = (note_val / notes_sur) * 5 if notes_sur else 0.0
        if note_norm == 0:
            note_norm = 0.01

        ctt = db.query_one(
            f"""SELECT id_contrat, id_salarie, date_signature, num_bs
                 FROM adv.pgt_{partenaire}_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )
        if not ctt:
            erreurs.append({"num_bs": num_bs, "erreur": "Contrat introuvable"})
            continue

        id_contrat = int(ctt["id_contrat"])
        id_sal = int(ctt.get("id_salarie") or 0)
        sal = _info_salarie(id_sal)
        ag, eq = _affectation(id_sal)
        date_sig = ctt.get("date_signature")
        vendeur = (f"{_str(sal.get('nom'))} "
                   f"{_str(sal.get('prenom')).title()}".strip()) if sal else ""

        statut = "Mode Simu"
        if not p.simulation:
            try:
                db.query(
                    f"""UPDATE adv.pgt_{partenaire}_contrat
                          SET notation = ?, notation_info = ?,
                              modif_date = NOW(), modif_op = ?,
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (note_norm, info_notes, int(op_id), id_contrat),
                )
                statut = "Import OK"
                nb_imp += 1
            except Exception as e:
                statut = f"Erreur : {e}"
        else:
            nb_imp += 1

        lignes.append(NotationLigne(
            num_bs=num_bs, vendeur=vendeur, agence=ag, equipe=eq,
            date_signature=str(date_sig or ""),
            note_normalisee=round(note_norm, 2),
            info_notes=info_notes, statut=statut,
        ))
    wb.close()

    # Build XLSX recapitulatif
    out = Workbook()
    sh1 = out.active
    sh1.title = "Liste des notations"
    sh1.append(["Num BS", "Vendeur", "Agence", "Equipe", "Date Signature",
                "Note (/5)", "Info Notes", "Statut"])
    for l in lignes:
        sh1.append([l.num_bs, l.vendeur, l.agence, l.equipe,
                    l.date_signature, l.note_normalisee, l.info_notes,
                    l.statut])
    if erreurs:
        sh2 = out.create_sheet(title="Erreurs")
        sh2.append(["Num BS", "Type Erreur"])
        for e in erreurs:
            sh2.append([e["num_bs"], e["erreur"]])

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    xlsx_name = f"ImportNotes_{partenaire.upper()}_{ts}{'_SIMU' if p.simulation else ''}.xlsx"
    buf = io.BytesIO(); out.save(buf); buf.seek(0)
    xlsx_b64 = base64.b64encode(buf.read()).decode("ascii")

    return ImportNotationResult(
        ok=True, partenaire=partenaire,
        nb_lignes=len(lignes) + len(erreurs),
        nb_importees=nb_imp, nb_introuvables=len(erreurs),
        nb_erreurs=sum(1 for l in lignes if l.statut.startswith("Erreur")),
        lignes=lignes, erreurs=erreurs,
        xlsx_b64=xlsx_b64, xlsx_name=xlsx_name,
        message=(f"{len(lignes) + len(erreurs)} ligne(s) | "
                 f"Importées {nb_imp} | Introuvables {len(erreurs)}. "
                 + ("(SIMULATION)" if p.simulation else "(PRODUCTION)")),
    )
