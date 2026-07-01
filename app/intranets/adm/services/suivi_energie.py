"""Services pour Fen_SuiviEnergie (module ADM > Suivi Énergie).

Sous-fenêtres :
  - Fen_ExtractionEnergie : extraction tickets Call Energie OEN par
    periode + toggle Validé/Annulé + export XLSX
  - Fen_TicketCall (à venir)
"""

from datetime import date, datetime, timedelta
from typing import Any

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection


# ID type demande = 22 pour les tickets Call Energie (cf code WinDev)
TK_TYPE_DEMANDE_ENERGIE = 22
# IDTK_Statut = 28 = Cloturé annulé (exclu de l'extraction)
TK_STATUT_EXCLU = 28


def _date_str(v: Any) -> str:
    if v is None: return ""
    if isinstance(v, (date, datetime)): return v.strftime("%Y-%m-%d")
    return str(v)


def _capitalize(s: str) -> str:
    return s.capitalize() if s else ""


# ====================================================================
# 1. EXTRACTION ENERGIE (Fen_ExtractionEnergie)
# ====================================================================


class ExtractionEnergieRow(BaseModel):
    id_tk_liste: str
    date_souscription: str = ""
    numero_cm: str = ""
    nom: str = ""
    prenom: str = ""
    telephone: str = ""
    adresse_mail: str = ""
    date_activation: str = ""
    type_contrat: str = ""      # ELEC / GAZ / DUAL
    commercial: str = ""


def search_extraction_energie(
    du: date, au: date, statut: str = "valide",
) -> list[ExtractionEnergieRow]:
    """cf Fen_ExtractionEnergie : liste les tickets Call energie OEN
    sur la periode Du/Au (date_crea).

    statut :
      - 'valide'  : StatutProd IN (1, 3) (cf BoutonSegmente1 = 1)
      - 'annule'  : StatutProd = 2

    Agrege par ticket : si 2 paniers -> TYPE = 'DUAL',
    sinon TYPE = OEN_produit.SousFAM.
    """
    if statut == "annule":
        statut_prod = (2,)
    else:
        statut_prod = (1, 3)

    db_bo = get_pg_connection("ticket_bo")
    db_tk = get_pg_connection("ticket")
    db_adv = get_pg_connection("adv")
    db_rh = get_pg_connection("rh")

    placeholder = ",".join("?" * len(statut_prod))
    rows = db_bo.query(
        f"""SELECT c.id_tk_liste, c.id_salarie, c.nom_client, c.prenom_client,
                   c.mobile1, c.adr_mail, c.ref_appel,
                   l.date_crea, l.op_crea, l.id_tk_statut,
                   p.id_produit, p.partenaire, p.statut_prod,
                   p.observations, p.date_entree
              FROM ticket_bo.pgt_tk_call c
              JOIN ticket.pgt_tk_liste l ON l.id_tk_liste = c.id_tk_liste
              JOIN ticket_bo.pgt_tk_call_panier p
                ON p.id_tk_liste = l.id_tk_liste
             WHERE (l.modif_elem IS NULL OR l.modif_elem NOT LIKE '%suppr%')
               AND l.id_tk_type_demande = ?
               AND l.id_tk_statut <> ?
               AND l.date_crea >= ?
               AND l.date_crea < (?::date + INTERVAL '1 day')
               AND p.partenaire = 'OEN'
               AND p.statut_prod IN ({placeholder})
             ORDER BY l.date_crea ASC""",
        (TK_TYPE_DEMANDE_ENERGIE, TK_STATUT_EXCLU, du, au, *statut_prod),
    ) or []
    if not rows:
        return []

    # Resolutions batch : produits OEN + vendeurs
    id_prods = list({int(r["id_produit"]) for r in rows if r.get("id_produit")})
    id_sals  = list({int(r["op_crea"]) for r in rows if r.get("op_crea")})

    prods_map: dict[int, str] = {}
    if id_prods:
        ids = ",".join(str(i) for i in id_prods)
        p = db_adv.query(
            f"""SELECT id_produit, sous_fam FROM adv.pgt_oen_produit
                 WHERE id_produit IN ({ids})""",
        ) or []
        prods_map = {int(x["id_produit"]): (x.get("sous_fam") or "") for x in p}

    sals_map: dict[int, str] = {}
    if id_sals:
        ids = ",".join(str(i) for i in id_sals)
        s = db_rh.query(
            f"""SELECT id_salarie, nom, prenom FROM rh.pgt_salarie
                 WHERE id_salarie IN ({ids})""",
        ) or []
        for x in s:
            prenom = _capitalize((x.get("prenom") or "").lower())
            nom = (x.get("nom") or "").upper()
            sals_map[int(x["id_salarie"])] = f"{prenom} {nom}".strip()

    # Agrege par id_tk_liste : si 2 paniers -> DUAL
    by_tk: dict[str, ExtractionEnergieRow] = {}
    for r in rows:
        id_tl = str(r["id_tk_liste"])
        item = by_tk.get(id_tl)
        if item is None:
            id_p = int(r.get("id_produit") or 0)
            item = ExtractionEnergieRow(
                id_tk_liste=id_tl,
                date_souscription=_date_str(r.get("date_crea")),
                numero_cm=r.get("observations") or "",
                nom=r.get("nom_client") or "",
                prenom=r.get("prenom_client") or "",
                telephone=r.get("mobile1") or "",
                adresse_mail=r.get("adr_mail") or "",
                date_activation=_date_str(r.get("date_entree")),
                type_contrat=prods_map.get(id_p, ""),
                commercial=sals_map.get(int(r.get("op_crea") or 0), ""),
            )
            by_tk[id_tl] = item
        else:
            # 2e panier sur le meme ticket -> DUAL
            item.type_contrat = "DUAL"
            # Mise a jour date d'activation si plus recente
            de = _date_str(r.get("date_entree"))
            if de:
                item.date_activation = de
    return list(by_tk.values())


# ====================================================================
# 2. TICKET CALL ENERGIE (Fen_TicketCall)
# ====================================================================


class TicketCallItem(BaseModel):
    id_tk_liste: str
    id_tk_call: str
    id_salarie: int
    date_crea: str = ""
    date_h_appel: str = ""
    date_deb_prise_en_charge: str = ""
    date_fin_prise_en_charge: str = ""
    lib_statut: str = ""
    nom_client: str = ""
    prenom_client: str = ""
    cp: str = ""
    ville: str = ""
    adr_mail: str = ""
    mobile1: str = ""
    ref_appel: str = ""
    ope_appel: int = 0
    nom_operateur: str = ""
    nb_ctt: int = 0
    liste_num_ctt: str = ""     # Col_ListeNumCtt : Lib_produit + NumBS
    num_cm: str = ""            # Col_NumCM (Observations quand partenaire OEN)
    delai_prise_charge_min: float = 0.0
    duree_appel_sec: float = 0.0
    partenaires: list[str] = []
    row_color_alert: bool = False  # >1h entre crea et num_date_saisie


class AnalyseTrancheEnergie(BaseModel):
    tranche_horaire: str
    nb_ticket: int = 0
    moins_3_min: int = 0
    entre_3_et_5: int = 0
    entre_5_et_7: int = 0
    plus_de_7: int = 0


class AnalyseVentesItem(BaseModel):
    delai: str
    ventes_validees: int = 0
    ventes_annulees: int = 0


class AnalyseVentesTotaux(BaseModel):
    tranches: list[AnalyseVentesItem] = []
    nb_ventes_validees: int = 0
    nb_ventes_annulees: int = 0
    nb_ventes_pas_statuees: int = 0


class PlanningRdvItem(BaseModel):
    id_tk_call: str
    id_tk_liste: str
    id_salarie: int
    titre: str = ""       # nom + prenom client
    contenu: str = ""     # panier(s)
    ressource: str = ""   # 'Crea Ticket' ou nom operateur
    date_debut: str = ""
    date_fin: str = ""
    couleur_hex: str = ""
    nb_valide: int = 0


_COULEURS_DELAI = {
    "moins_3": "#bbf7d0",     # vert pastel (<3min)
    "entre_3_5": "#fef08a",   # jaune pastel
    "entre_5_7": "#fed7aa",   # orange pastel
    "plus_7": "#fecaca",      # rouge pastel
}


def _delai_label(en_min: float) -> tuple[str, str]:
    """Retourne (libelle, key_couleur) selon le delai en minutes."""
    if en_min < 3: return ("< 3 min", "moins_3")
    if en_min < 5: return ("3 à 5 min", "entre_3_5")
    if en_min < 7: return ("5 à 7 min", "entre_5_7")
    return ("> 7 min", "plus_7")


def _search_tickets_call_raw(
    du: date, au: date, etat: str,
) -> list[dict]:
    """Requete brute : cf ReqTkCall WinDev."""
    db_bo = get_pg_connection("ticket_bo")

    where = [
        "(c.modif_elem IS NULL OR c.modif_elem NOT LIKE '%suppr%')",
        "l.date_crea >= ?",
        "l.date_crea < (?::date + INTERVAL '1 day')",
        # Optionnel mais safe : ne prend que les tickets Energie
        "l.id_tk_type_demande = 22",
    ]
    params: list = [du, au]
    if etat == "ouverts":
        where.append("(l.cloturee IS NULL OR l.cloturee = FALSE)")
    elif etat == "clotures":
        where.append("l.cloturee = TRUE")

    rows = db_bo.query(
        f"""SELECT c.id_tk_call, c.id_tk_liste, c.id_salarie,
                   c.civilite_client, c.nom_client, c.nom_marital_client,
                   c.prenom_client, c.date_naiss, c.dep_naiss,
                   c.adresse1, c.adresse2, c.cp, c.ville, c.adr_mail,
                   c.mobile1, c.type_logement, c.appel_en_cours,
                   c.date_h_appel, c.ope_appel, c.ref_appel,
                   c.date_deb_prise_en_charge, c.date_fin_prise_en_charge,
                   l.date_crea, l.id_tk_statut, l.date_cloture, l.cloturee
              FROM ticket_bo.pgt_tk_call c
              JOIN ticket.pgt_tk_liste l ON l.id_tk_liste = c.id_tk_liste
             WHERE {' AND '.join(where)}
             ORDER BY l.date_crea DESC
             LIMIT 5000""",
        tuple(params),
    ) or []
    return rows


def list_ticket_call_energie(
    du: date, au: date, etat: str = "tous",
) -> list[TicketCallItem]:
    """cf Fen_TicketCall Energie : liste des tickets Call + resolution
    panier + operateur + statut."""
    rows = _search_tickets_call_raw(du, au, etat)
    if not rows:
        return []

    db_bo = get_pg_connection("ticket_bo")
    db_tk = get_pg_connection("ticket")
    db_rh = get_pg_connection("rh")
    db_adv = get_pg_connection("adv")

    # Panier par id_tk_call
    id_calls = [int(r["id_tk_call"]) for r in rows]
    ids_sql = ",".join(str(i) for i in id_calls)
    paniers = db_bo.query(
        f"""SELECT p.id_tk_call, p.id_produit, p.partenaire, p.num_bs,
                   p.num_date_saisie, p.observations, p.statut_prod
              FROM ticket_bo.pgt_tk_call_panier p
             WHERE p.id_tk_call IN ({ids_sql})
               AND (p.modif_elem IS NULL OR p.modif_elem NOT LIKE '%suppr%')""",
    ) or []
    paniers_by_call: dict[int, list[dict]] = {}
    for p in paniers:
        paniers_by_call.setdefault(int(p["id_tk_call"]), []).append(p)

    # Lookup produits par partenaire (schema commun adv)
    all_prods: dict[tuple[str, int], str] = {}
    for p in paniers:
        part = (p.get("partenaire") or "").upper()
        pid = int(p.get("id_produit") or 0)
        if not part or not pid: continue
        table = f"pgt_{part.lower()}_produit"
        if (part, pid) in all_prods: continue
        try:
            r = db_adv.query_one(
                f"SELECT lib_produit FROM adv.{table} WHERE id_produit = ? LIMIT 1",
                (pid,),
            )
            all_prods[(part, pid)] = (r or {}).get("lib_produit") or ""
        except Exception:
            all_prods[(part, pid)] = ""

    # Statuts + operateurs + salaries
    id_statuts = list({int(r.get("id_tk_statut") or 0) for r in rows if r.get("id_tk_statut")})
    id_ope     = list({int(r.get("ope_appel") or 0) for r in rows if r.get("ope_appel")})

    statuts_map: dict[int, str] = {}
    if id_statuts:
        ids = ",".join(str(i) for i in id_statuts)
        s = db_tk.query(
            f"SELECT id_tk_statut, lib_statut FROM ticket.pgt_tk_statut WHERE id_tk_statut IN ({ids})",
        ) or []
        statuts_map = {int(x["id_tk_statut"]): x.get("lib_statut") or "" for x in s}

    ope_map: dict[int, str] = {}
    if id_ope:
        ids = ",".join(str(i) for i in id_ope)
        s = db_rh.query(
            f"SELECT id_salarie, nom, prenom FROM rh.pgt_salarie WHERE id_salarie IN ({ids})",
        ) or []
        for x in s:
            prenom = _capitalize((x.get("prenom") or "").lower())
            ope_map[int(x["id_salarie"])] = f"{prenom} {(x.get('nom') or '').upper()}".strip()

    out: list[TicketCallItem] = []
    for r in rows:
        id_c = int(r["id_tk_call"])
        pans = paniers_by_call.get(id_c, [])

        # Construit Col_ListeNumCtt + partenaires + row_color_alert
        libs: list[str] = []
        parts_set: set[str] = set()
        num_cm = ""
        row_alert = False
        for p in pans:
            part = (p.get("partenaire") or "").upper()
            parts_set.add(part)
            pid = int(p.get("id_produit") or 0)
            lib = all_prods.get((part, pid)) or f"Prod {part} inconnu"
            num_bs = p.get("num_bs") or ""
            line = f"{lib} {num_bs}"
            nds = p.get("num_date_saisie")
            if isinstance(nds, datetime):
                line += f" ({nds.strftime('%d/%m/%Y %H:%M')})"
                # Diff crea vs num_date_saisie : >1h -> rouge
                dc = r.get("date_crea")
                if isinstance(dc, datetime):
                    diff_h = (nds - dc).total_seconds() / 3600.0
                    if diff_h >= 1:
                        row_alert = True
            libs.append(line)
            if part == "OEN":
                num_cm = p.get("observations") or ""

        # Delai + duree
        delai_min = 0.0
        duree_sec = 0.0
        dc = r.get("date_crea")
        deb = r.get("date_deb_prise_en_charge")
        fin = r.get("date_fin_prise_en_charge")
        if isinstance(dc, datetime) and isinstance(deb, datetime):
            delai_min = max(0.0, (deb - dc).total_seconds() / 60.0)
        if isinstance(deb, datetime) and isinstance(fin, datetime):
            duree_sec = max(0.0, (fin - deb).total_seconds())

        out.append(TicketCallItem(
            id_tk_liste=str(r["id_tk_liste"]),
            id_tk_call=str(r["id_tk_call"]),
            id_salarie=int(r.get("id_salarie") or 0),
            date_crea=_date_str(dc) if not isinstance(dc, datetime) else dc.strftime("%Y-%m-%d %H:%M:%S"),
            date_h_appel=(r.get("date_h_appel").strftime("%Y-%m-%d %H:%M:%S")
                          if isinstance(r.get("date_h_appel"), datetime) else ""),
            date_deb_prise_en_charge=(deb.strftime("%Y-%m-%d %H:%M:%S")
                                       if isinstance(deb, datetime) else ""),
            date_fin_prise_en_charge=(fin.strftime("%Y-%m-%d %H:%M:%S")
                                       if isinstance(fin, datetime) else ""),
            lib_statut=statuts_map.get(int(r.get("id_tk_statut") or 0), ""),
            nom_client=r.get("nom_client") or "",
            prenom_client=r.get("prenom_client") or "",
            cp=r.get("cp") or "",
            ville=r.get("ville") or "",
            adr_mail=r.get("adr_mail") or "",
            mobile1=r.get("mobile1") or "",
            ref_appel=r.get("ref_appel") or "",
            ope_appel=int(r.get("ope_appel") or 0),
            nom_operateur=ope_map.get(int(r.get("ope_appel") or 0), ""),
            nb_ctt=len(pans),
            liste_num_ctt="\n".join(libs),
            num_cm=num_cm,
            delai_prise_charge_min=round(delai_min, 2),
            duree_appel_sec=round(duree_sec, 1),
            partenaires=sorted(parts_set),
            row_color_alert=row_alert,
        ))
    return out


def analyse_ventes_tk_call_energie(
    du: date, au: date, etat: str = "tous",
) -> AnalyseVentesTotaux:
    """Analyse des ventes par tranche de delai (cf onglet Analyse ventes
    Fen_TicketCall) : compte les paniers Validés (statut_prod IN 1,3)
    vs Annulés (statut_prod=2) par tranche de delai (<3 / 3-5 / 5-7 / >7)."""
    rows = _search_tickets_call_raw(du, au, etat)
    if not rows:
        return AnalyseVentesTotaux()

    db_bo = get_pg_connection("ticket_bo")
    id_calls = [int(r["id_tk_call"]) for r in rows]
    ids_sql = ",".join(str(i) for i in id_calls)
    paniers = db_bo.query(
        f"""SELECT id_tk_call, statut_prod FROM ticket_bo.pgt_tk_call_panier
             WHERE id_tk_call IN ({ids_sql})
               AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
    ) or []
    paniers_by_call: dict[int, list[dict]] = {}
    for p in paniers:
        paniers_by_call.setdefault(int(p["id_tk_call"]), []).append(p)

    # Agrege par tranche de delai
    tranches_map: dict[str, AnalyseVentesItem] = {}
    tot_val = tot_ann = tot_pas = 0
    for r in rows:
        deb = r.get("date_deb_prise_en_charge")
        crea = r.get("date_crea")
        delai = 0.0
        if isinstance(deb, datetime) and isinstance(crea, datetime):
            delai = max(0.0, (deb - crea).total_seconds() / 60.0)
        lbl, _ = _delai_label(delai)

        item = tranches_map.get(lbl)
        if item is None:
            item = AnalyseVentesItem(delai=lbl)
            tranches_map[lbl] = item

        for p in paniers_by_call.get(int(r["id_tk_call"]), []):
            sp = int(p.get("statut_prod") or 0)
            if sp in (1, 3):
                item.ventes_validees += 1
                tot_val += 1
            elif sp == 2:
                item.ventes_annulees += 1
                tot_ann += 1
            else:
                tot_pas += 1

    order = {"< 3 min": 0, "3 à 5 min": 1, "5 à 7 min": 2, "> 7 min": 3}
    tranches = sorted(tranches_map.values(), key=lambda x: order.get(x.delai, 99))
    return AnalyseVentesTotaux(
        tranches=tranches,
        nb_ventes_validees=tot_val,
        nb_ventes_annulees=tot_ann,
        nb_ventes_pas_statuees=tot_pas,
    )


def planning_appels_energie(
    du: date, au: date, etat: str = "tous",
) -> list[PlanningRdvItem]:
    """cf onglet Planning : genere les RDV pour le calendrier
    (colonne 'Crea Ticket' + colonnes operateurs). Chaque ticket
    genere 2 RDV : un a la date de crea, un a la prise en charge."""
    rows = _search_tickets_call_raw(du, au, etat)
    if not rows:
        return []

    db_bo = get_pg_connection("ticket_bo")
    db_rh = get_pg_connection("rh")
    id_calls = [int(r["id_tk_call"]) for r in rows]
    ids_sql = ",".join(str(i) for i in id_calls)
    paniers = db_bo.query(
        f"""SELECT id_tk_call, num_bs, num_date_saisie, statut_prod, partenaire
              FROM ticket_bo.pgt_tk_call_panier
             WHERE id_tk_call IN ({ids_sql})
               AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
    ) or []
    paniers_by_call: dict[int, list[dict]] = {}
    for p in paniers:
        paniers_by_call.setdefault(int(p["id_tk_call"]), []).append(p)

    id_ope = list({int(r.get("ope_appel") or 0) for r in rows if r.get("ope_appel")})
    ope_map: dict[int, str] = {}
    if id_ope:
        ids = ",".join(str(i) for i in id_ope)
        s = db_rh.query(
            f"SELECT id_salarie, nom, prenom FROM rh.pgt_salarie WHERE id_salarie IN ({ids})",
        ) or []
        for x in s:
            prenom = _capitalize((x.get("prenom") or "").lower())
            ope_map[int(x["id_salarie"])] = f"{prenom} {(x.get('nom') or '').upper()}".strip()

    out: list[PlanningRdvItem] = []
    for r in rows:
        # Skip vendeur ID=6 cf code WinDev
        id_sa = int(r.get("id_salarie") or 0)
        if id_sa == 6: continue

        dc = r.get("date_crea")
        if not isinstance(dc, datetime): continue

        # Delai + couleur
        deb = r.get("date_deb_prise_en_charge")
        delai_min = 0.0
        if isinstance(deb, datetime):
            delai_min = max(0.0, (deb - dc).total_seconds() / 60.0)
        _, key = _delai_label(delai_min)
        couleur = _COULEURS_DELAI[key]

        # Titre + contenu
        titre = f"{r.get('nom_client') or ''} {_capitalize((r.get('prenom_client') or '').lower())}".strip()
        pans = paniers_by_call.get(int(r["id_tk_call"]), [])
        contenu_parts = []
        nb_val = 0
        for p in pans:
            sp = int(p.get("statut_prod") or 0)
            etat_prod = "Pas statué"
            if sp == 1: etat_prod = "Validé"; nb_val += 1
            elif sp == 2: etat_prod = "Annulé"
            elif sp == 3: etat_prod = "Num BS ajouté"; nb_val += 1
            line = p.get("num_bs") or ""
            nds = p.get("num_date_saisie")
            if isinstance(nds, datetime):
                line += f" ({nds.strftime('%d/%m/%Y %H:%M')})"
            line += f" - {etat_prod}"
            contenu_parts.append(line)
        contenu = "\n".join(contenu_parts)

        # RDV 1 : Crea Ticket (colonne 'Crea Ticket')
        deb_crea = dc.replace(second=0, microsecond=0)
        fin_crea = deb_crea + timedelta(minutes=3)
        out.append(PlanningRdvItem(
            id_tk_call=str(r["id_tk_call"]),
            id_tk_liste=str(r["id_tk_liste"]),
            id_salarie=id_sa,
            titre=titre, contenu=contenu,
            ressource="Crea Ticket",
            date_debut=deb_crea.strftime("%Y-%m-%d %H:%M:%S"),
            date_fin=fin_crea.strftime("%Y-%m-%d %H:%M:%S"),
            couleur_hex=couleur, nb_valide=nb_val,
        ))

        # RDV 2 : Colonne operateur (Prise en charge)
        ope_id = int(r.get("ope_appel") or 0)
        nom_ope = ope_map.get(ope_id, "")
        if nom_ope and isinstance(deb, datetime):
            fin = r.get("date_fin_prise_en_charge")
            if isinstance(fin, datetime) and deb <= fin:
                d_deb, d_fin = deb, fin
            else:
                d_deb, d_fin = deb, deb + timedelta(minutes=3)
            out.append(PlanningRdvItem(
                id_tk_call=str(r["id_tk_call"]),
                id_tk_liste=str(r["id_tk_liste"]),
                id_salarie=id_sa,
                titre=titre, contenu=contenu,
                ressource=nom_ope,
                date_debut=d_deb.strftime("%Y-%m-%d %H:%M:%S"),
                date_fin=d_fin.strftime("%Y-%m-%d %H:%M:%S"),
                couleur_hex=couleur, nb_valide=nb_val,
            ))
    return out


class TicketCallPanierItem(BaseModel):
    id_tk_call_panier: str
    id_produit: int
    partenaire: str = ""
    lib_offre: str = ""
    num_bs: str = ""
    num_date_saisie: str = ""
    statut_prod: int = 0
    motif_annulation: str = ""
    date_entree: str = ""             # DateActiv
    observations: str = ""            # RefClient
    opt_mail: bool = False
    opt_e_facture: bool = False
    opt_e_communication: bool = False
    opt_optin_commercial: bool = False
    opt_consent_consult_distri: bool = False
    opt_accept_com_parte: bool = False
    opt_mandat: bool = False
    opt_energie_verte_gaz: bool = False
    opt_reforestation: bool = False
    format_numerique: bool = False
    a_creer: bool = False


class TicketCallDetail(BaseModel):
    id_tk_liste: str
    id_tk_call: str
    id_client: str
    # infos ticket
    date_crea: str = ""
    lib_statut: str = ""
    date_report: str = ""
    cloturee: bool = False
    date_cloture: str = ""
    op_dest: int = 0
    nom_dest: str = ""
    # infos client
    civilite: int = 0
    nom_client: str = ""
    prenom_client: str = ""
    nom_marital_client: str = ""
    date_naiss: str = ""
    dep_naiss: str = ""
    adresse1: str = ""
    adresse2: str = ""
    cp: str = ""
    ville: str = ""
    type_logement: str = ""
    mobile1: str = ""
    adr_mail: str = ""
    date_sign: str = ""              # date_crea..PartieDate
    ref_appel: str = ""
    info_vente: str = ""
    opt_partenaire: bool = False     # Consent rappel partenaires
    intervention_vend: bool = False
    # salarie (auteur)
    id_salarie: int = 0
    nom_operateur: str = ""
    # paniers
    paniers: list[TicketCallPanierItem] = []


def get_ticket_call_detail(id_tk_liste: int) -> TicketCallDetail | None:
    """Charge le detail complet du ticket Call Energie pour le modal
    Fen_ContenuTicketCall."""
    db_bo = get_pg_connection("ticket_bo")
    db_tk = get_pg_connection("ticket")
    db_rh = get_pg_connection("rh")
    db_adv = get_pg_connection("adv")

    # TK_Liste
    tl = db_tk.query_one(
        """SELECT id_tk_liste, date_crea, op_crea, op_dest, id_tk_statut,
                   date_report, cloturee, date_cloture
             FROM ticket.pgt_tk_liste WHERE id_tk_liste = ? LIMIT 1""",
        (int(id_tk_liste),),
    )
    if not tl:
        return None

    # TK_Call
    tc = db_bo.query_one(
        """SELECT id_tk_call, id_salarie, id_client,
                   civilite_client, nom_client, prenom_client,
                   nom_marital_client, date_naiss, dep_naiss,
                   adresse1, adresse2, cp, ville, type_logement,
                   mobile1, adr_mail, ref_appel, info_vente,
                   intervention_vend, opt_partenaire
              FROM ticket_bo.pgt_tk_call
             WHERE id_tk_liste = ?
               AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
             LIMIT 1""",
        (int(id_tk_liste),),
    ) or {}

    # Statut ticket
    lib_statut = ""
    if tl.get("id_tk_statut"):
        s = db_tk.query_one(
            "SELECT lib_statut FROM ticket.pgt_tk_statut WHERE id_tk_statut = ? LIMIT 1",
            (int(tl["id_tk_statut"]),),
        )
        lib_statut = (s or {}).get("lib_statut") or ""

    # Operateur (id_salarie de TK_Call sinon op_crea)
    id_sal = int(tc.get("id_salarie") or 0) or int(tl.get("op_crea") or 0)
    nom_ope = ""
    id_dest = int(tl.get("op_dest") or 0)
    nom_dest = ""
    if id_sal or id_dest:
        ids = ",".join(str(i) for i in [id_sal, id_dest] if i)
        if ids:
            sals = db_rh.query(
                f"SELECT id_salarie, nom, prenom FROM rh.pgt_salarie WHERE id_salarie IN ({ids})",
            ) or []
            for x in sals:
                prenom = _capitalize((x.get("prenom") or "").lower())
                lbl = f"{prenom} {(x.get('nom') or '').upper()}".strip()
                if int(x["id_salarie"]) == id_sal: nom_ope = lbl
                if int(x["id_salarie"]) == id_dest: nom_dest = lbl

    # Paniers
    paniers: list[TicketCallPanierItem] = []
    if tc.get("id_tk_call"):
        pans = db_bo.query(
            """SELECT id_tk_call_panier, id_produit, partenaire, num_bs,
                       num_date_saisie, statut_prod, motif_annulation,
                       date_entree, observations,
                       opt_mail, opt_e_facture, opt_e_communication,
                       opt_optin_commercial, opt_consent_consult_distri,
                       opt_accept_com_parte, opt_mandat,
                       opt_energie_verte_gaz, opt_reforestation,
                       format_numerique
                  FROM ticket_bo.pgt_tk_call_panier
                 WHERE id_tk_call = ?
                   AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
            (int(tc["id_tk_call"]),),
        ) or []
        # Lookup produits par partenaire
        prods_cache: dict[tuple[str, int], str] = {}
        for p in pans:
            part = (p.get("partenaire") or "").upper()
            pid = int(p.get("id_produit") or 0)
            lib = ""
            if part and pid:
                if (part, pid) not in prods_cache:
                    try:
                        table = f"pgt_{part.lower()}_produit"
                        r = db_adv.query_one(
                            f"SELECT lib_produit FROM adv.{table} WHERE id_produit = ? LIMIT 1",
                            (pid,),
                        )
                        prods_cache[(part, pid)] = (r or {}).get("lib_produit") or ""
                    except Exception:
                        prods_cache[(part, pid)] = ""
                lib = prods_cache[(part, pid)]
            num_bs = p.get("num_bs") or ""
            statut_prod = int(p.get("statut_prod") or 0)
            paniers.append(TicketCallPanierItem(
                id_tk_call_panier=str(p["id_tk_call_panier"]),
                id_produit=pid,
                partenaire=part,
                lib_offre=lib,
                num_bs=num_bs,
                num_date_saisie=_date_str(p.get("num_date_saisie")),
                statut_prod=statut_prod,
                motif_annulation=p.get("motif_annulation") or "",
                date_entree=_date_str(p.get("date_entree")),
                observations=p.get("observations") or "",
                opt_mail=bool(p.get("opt_mail")),
                opt_e_facture=bool(p.get("opt_e_facture")),
                opt_e_communication=bool(p.get("opt_e_communication")),
                opt_optin_commercial=bool(p.get("opt_optin_commercial")),
                opt_consent_consult_distri=bool(p.get("opt_consent_consult_distri")),
                opt_accept_com_parte=bool(p.get("opt_accept_com_parte")),
                opt_mandat=bool(p.get("opt_mandat")),
                opt_energie_verte_gaz=bool(p.get("opt_energie_verte_gaz")),
                opt_reforestation=bool(p.get("opt_reforestation")),
                format_numerique=bool(p.get("format_numerique")),
                a_creer=(num_bs != "" and statut_prod == 1),
            ))

    dc = tl.get("date_crea")
    return TicketCallDetail(
        id_tk_liste=str(tl["id_tk_liste"]),
        id_tk_call=str(tc.get("id_tk_call") or ""),
        id_client=str(tc.get("id_client") or ""),
        date_crea=(dc.strftime("%Y-%m-%d %H:%M:%S")
                   if isinstance(dc, datetime) else _date_str(dc)),
        lib_statut=lib_statut,
        date_report=_date_str(tl.get("date_report")),
        cloturee=bool(tl.get("cloturee")),
        date_cloture=_date_str(tl.get("date_cloture")),
        op_dest=id_dest,
        nom_dest=nom_dest,
        civilite=int(tc.get("civilite_client") or 0),
        nom_client=tc.get("nom_client") or "",
        prenom_client=tc.get("prenom_client") or "",
        nom_marital_client=tc.get("nom_marital_client") or "",
        date_naiss=_date_str(tc.get("date_naiss")),
        dep_naiss=str(tc.get("dep_naiss") or ""),
        adresse1=tc.get("adresse1") or "",
        adresse2=tc.get("adresse2") or "",
        cp=tc.get("cp") or "",
        ville=tc.get("ville") or "",
        type_logement=str(tc.get("type_logement") or ""),
        mobile1=tc.get("mobile1") or "",
        adr_mail=tc.get("adr_mail") or "",
        date_sign=_date_str(dc.date() if isinstance(dc, datetime) else dc),
        ref_appel=tc.get("ref_appel") or "",
        info_vente=tc.get("info_vente") or "",
        opt_partenaire=bool(tc.get("opt_partenaire")),
        intervention_vend=bool(tc.get("intervention_vend")),
        id_salarie=id_sal,
        nom_operateur=nom_ope,
        paniers=paniers,
    )


def update_panier_call_energie(
    id_panier: int, num: str, statut_prod: int, op_id: int,
) -> bool:
    """Update NUM + StatutProd d'un panier."""
    db = get_pg_connection("ticket_bo")
    db.query(
        """UPDATE ticket_bo.pgt_tk_call_panier
              SET num_bs=?, num_date_saisie=NOW(), statut_prod=?,
                  modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_tk_call_panier=?""",
        (num.upper(), int(statut_prod), int(op_id), int(id_panier)),
    )
    return True


# ====================================================================
# 2c. CONVERTIR LA SELECTION EN CONTRAT (multi-partenaire OEN/ENI/etc.)
# ====================================================================


# ID vendeur "generique" cf code WinDev
CALL_GENERIC_SALARIE_ID = 20200715153948361


class ConvertPanierResult(BaseModel):
    id_panier: str
    partenaire: str = ""
    num_bs: str = ""
    id_statut: int = 0     # cf codes WinDev
    action: str = ""       # 'created' | 'updated' | 'skipped' | 'error' | 'doublon'
    message: str = ""


class ConvertSelectionResult(BaseModel):
    nb_crees: int = 0
    nb_updates: int = 0
    nb_erreurs: int = 0
    nb_skipped: int = 0
    results: list[ConvertPanierResult] = []


def _new_id() -> int:
    """ID entier 8 octets = timestamp au format yyyyMMddHHmmssSSS."""
    return int(datetime.now().strftime("%Y%m%d%H%M%S%f")[:17])


def _enregistrer_client_energie(
    db, tc_row: dict, op_id: int,
) -> int:
    """Cf procedure EnregistrerClient WinDev : dedup par (nom, prenom,
    adresse, cp, ville, tel, gsm) puis crea si nouveau."""
    tel1 = (tc_row.get("mobile1") or "").replace(".", "").replace(" ", "").replace("-", "")
    gsm2 = tel1   # WinDev met gsm et tel identiques (ClientTel = "" en appel)
    nom = tc_row.get("nom_client") or ""
    prenom = tc_row.get("prenom_client") or ""
    cp = tc_row.get("cp") or ""
    ville = tc_row.get("ville") or ""

    existing = db.query_one(
        """SELECT id_client FROM adv.pgt_client
            WHERE UPPER(nom) = UPPER(?) AND UPPER(prenom) = UPPER(?)
              AND COALESCE(cp, '') = ?
              AND COALESCE(ville, '') = ?
              AND (COALESCE(gsm, '') = ? OR COALESCE(tel, '') = ?)
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            LIMIT 1""",
        (nom, prenom, cp, ville, gsm2, tel1),
    )
    if existing:
        return int(existing["id_client"])

    id_new = _new_id()
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_client_auto), 0) + 1 AS n FROM adv.pgt_client"
    )
    auto_n = int(auto["n"]) if auto else 1
    db.query(
        """INSERT INTO adv.pgt_client
              (id_client_auto, id_client, civilite, nom, prenom,
               date_naiss, adresse1, adresse2, cp, ville, pays,
               tel, gsm, mail, opt_partenaire,
               op_saisie, date_saisie, modif_date, modif_op, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'FRANCE',
                   ?, ?, ?, ?, ?, NOW(), NOW(), ?, 'new')""",
        (auto_n, id_new, int(tc_row.get("civilite_client") or 0),
         nom, prenom,
         tc_row.get("date_naiss") if tc_row.get("date_naiss") else None,
         tc_row.get("adresse1") or "", tc_row.get("adresse2") or "",
         cp, ville, gsm2, tel1, tc_row.get("adr_mail") or "",
         bool(tc_row.get("opt_partenaire")),
         int(op_id), int(op_id)),
    )
    return id_new


def _modif_fiche_client(db, id_client: int, tc_row: dict, op_id: int) -> None:
    """cf ModifFicheClient WinDev : maj UNIQUEMENT si client.NOM = ''."""
    if not id_client: return
    r = db.query_one(
        "SELECT nom FROM adv.pgt_client WHERE id_client = ?", (int(id_client),))
    if not r or (r.get("nom") or "").strip():
        return
    tel = (tc_row.get("mobile1") or "").replace(".", "").replace(" ", "")
    db.query(
        """UPDATE adv.pgt_client
              SET nom=?, prenom=?, date_naiss=?, adresse1=?, adresse2=?,
                  cp=?, ville=?, pays='FRANCE', tel=?, gsm=?, mail=?,
                  opt_partenaire=?, op_saisie=?, date_saisie=NOW(),
                  modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_client=?""",
        (tc_row.get("nom_client") or "",
         tc_row.get("prenom_client") or "",
         tc_row.get("date_naiss") if tc_row.get("date_naiss") else None,
         tc_row.get("adresse1") or "", tc_row.get("adresse2") or "",
         tc_row.get("cp") or "", tc_row.get("ville") or "",
         tel, tel, tc_row.get("adr_mail") or "",
         bool(tc_row.get("opt_partenaire")),
         int(op_id), int(op_id), int(id_client)),
    )


def _get_etat_contrat_oen(statut_prod: int, num_bs: str) -> int:
    """cf WinDev partenaire OEN :
      0 (non defini) -> 42 (Rejet BO - NON CALL)
      2 (annule)     -> 38
      TK dans NUM    -> 39
      autre          -> 40"""
    if statut_prod == 0: return 42
    if statut_prod == 2: return 38
    if "TK" in (num_bs or "").upper(): return 39
    return 40


def _get_etat_contrat_eni(statut_prod: int, num_bs: str) -> int:
    """cf WinDev partenaire ENI :
      2 (annule)     -> 66
      TK dans NUM    -> 67
      autre          -> 37"""
    if statut_prod == 2: return 66
    if "TK" in (num_bs or "").upper(): return 67
    return 37


def _get_etat_contrat_pro(statut_prod: int, num_bs: str) -> int:
    """cf WinDev partenaire PRO :
      2 (annule)     -> 6
      TK dans NUM    -> 7
      autre          -> 5"""
    if statut_prod == 2: return 6
    if "TK" in (num_bs or "").upper(): return 7
    return 5


def _get_etat_contrat_oen_v2(statut_prod: int, num_bs: str) -> int:
    """Version bouton 'Convertir la selection' au niveau liste :
      0 (non defini) + TK -> 51 (Ticket non finalise et non call)
      0 (non defini) + autre -> 42 (Rejet BO NON CALL)
      2 (annule)     -> 38
      TK dans NUM    -> 39
      autre          -> 40"""
    has_tk = "TK" in (num_bs or "").upper()
    if statut_prod == 0:
        return 51 if has_tk else 42
    if statut_prod == 2: return 38
    return 39 if has_tk else 40


def _get_etat_contrat_default(statut_prod: int, num_bs: str) -> int:
    """Autres partenaires (VAL, STR...) : meme mapping que ENI."""
    return _get_etat_contrat_eni(statut_prod, num_bs)


def _insert_contrat(
    db, partenaire: str, panier: dict, tc_row: dict,
    id_client: int, id_salarie: int, id_ste: int | None,
    id_etat_contrat: int, num_bs: str, date_signature,
    lib_produit: str, famille: str, sous_fam: str, op_id: int,
) -> tuple[int, float]:
    """Insere le contrat dans adv.pgt_{partenaire}_contrat.
    Retourne (id_contrat, nb_points)."""
    from app.shared.sdtc.bareme import calcul_point_contrat

    part = partenaire.lower()
    table = f"pgt_{part}_contrat"
    id_new = _new_id()

    # Calcul nbPoints
    nbpt = 0.0
    try:
        nbpt = calcul_point_contrat(
            famille or "", sous_fam or "", "",
            date_signature.strftime("%Y-%m-%d") if isinstance(date_signature, (date, datetime)) else str(date_signature or ""),
            num_bs, 0,
        ) or 0.0
    except Exception:
        pass

    # Auto n
    auto = db.query_one(
        f"SELECT COALESCE(MAX(id_contrat_auto), 0) + 1 AS n FROM adv.{table}"
    )
    auto_n = int(auto["n"]) if auto else 1

    # Colonnes communes
    if part == "oen":
        db.query(
            f"""INSERT INTO adv.{table}
                  (id_contrat_auto, id_contrat, id_client, id_salarie, id_ste,
                   num_bs, ref_client, id_produit, id_etat_contrat, date_signature,
                   date_activation, is_dual,
                   op_saisie, date_saisie, non_call, nb_points,
                   modif_op, modif_date, modif_elem)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NOW(), FALSE, ?, ?, NOW(), 'new')""",
            (auto_n, id_new, int(id_client), int(id_salarie), id_ste,
             num_bs, panier.get("observations") or "",
             int(panier.get("id_produit") or 0), id_etat_contrat,
             date_signature,
             panier.get("date_entree") if panier.get("date_entree") else None,
             (panier.get("motif_annulation") or "").upper().find("DUAL") >= 0,
             int(op_id), float(nbpt), int(op_id)),
        )
    elif part == "eni":
        db.query(
            f"""INSERT INTO adv.{table}
                  (id_contrat_auto, id_contrat, id_client, id_salarie, id_ste,
                   num_bs, id_produit, id_etat_contrat, date_signature,
                   op_saisie, date_saisie, non_call, nb_points,
                   modif_op, modif_date, modif_elem)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NOW(), FALSE, ?, ?, NOW(), 'new')""",
            (auto_n, id_new, int(id_client), int(id_salarie), id_ste,
             num_bs, int(panier.get("id_produit") or 0), id_etat_contrat,
             date_signature, int(op_id), float(nbpt), int(op_id)),
        )
    else:
        # VAL, STR, PRO... : colonnes minimales communes
        cols = ["id_contrat_auto", "id_contrat", "id_client", "id_salarie",
                "id_ste", "num_bs", "id_produit", "id_etat_contrat",
                "date_signature", "op_saisie", "date_saisie",
                "non_call", "nb_points", "modif_op", "modif_date", "modif_elem"]
        vals = [auto_n, id_new, int(id_client), int(id_salarie), id_ste,
                num_bs, int(panier.get("id_produit") or 0), id_etat_contrat,
                date_signature, int(op_id), datetime.now(),
                False, float(nbpt), int(op_id), datetime.now(), "new"]
        # Colonnes optionnelles specifiques
        if part == "val" and panier.get("format_numerique"):
            cols.append("format_numerique"); vals.append(True)
        if part == "str" and panier.get("opt_mandat"):
            cols.append("opt_mandat"); vals.append(True)
        col_sql = ", ".join(cols)
        ph = ", ".join(["?"] * len(vals))
        db.query(
            f"INSERT INTO adv.{table} ({col_sql}) VALUES ({ph})",
            tuple(vals),
        )

    return id_new, float(nbpt)


def _insert_contrat_option(
    db, partenaire: str, id_contrat: int, num_bs: str, panier: dict,
    vte_add_part: str, op_id: int,
) -> None:
    """Insere l'option pour OEN ou ENI (les autres n'ont pas d'option)."""
    part = partenaire.lower()
    if part not in ("oen", "eni"): return
    table = f"pgt_{part}_contrat_option"
    auto = db.query_one(
        f"SELECT COALESCE(MAX(id_contrat_option_auto), 0) + 1 AS n FROM adv.{table}"
    )
    auto_n = int(auto["n"]) if auto else 1

    common = {
        "opt_mail": bool(panier.get("opt_mail")),
        "opt_energie_verte_gaz": bool(panier.get("opt_energie_verte_gaz")),
        "opt_energie_verte_elec": False,
        "opt_reforestation": bool(panier.get("opt_reforestation")),
        "opt_optin_commercial": bool(panier.get("opt_optin_commercial")),
        "opt_e_facture": bool(panier.get("opt_e_facture")),
        "opt_e_communication": bool(panier.get("opt_e_communication")),
        "opt_accept_com_parte": bool(panier.get("opt_accept_com_parte")),
        "opt_consent_consult_distri": bool(panier.get("opt_consent_consult_distri")),
    }
    if part == "oen":
        db.query(
            f"""INSERT INTO adv.{table}
                  (id_contrat_option_auto, id_contrat, num_bs,
                   opt_mail, opt_energie_verte_gaz, opt_energie_verte_elec,
                   opt_reforestation, opt_optin_commercial,
                   opt_e_facture, opt_e_communication,
                   opt_accept_com_parte, opt_consent_consult_distri,
                   opt_vte_add_part,
                   modif_op, modif_date, modif_elem)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NOW(), 'new')""",
            (auto_n, id_contrat, num_bs,
             common["opt_mail"], common["opt_energie_verte_gaz"],
             common["opt_energie_verte_elec"],
             common["opt_reforestation"], common["opt_optin_commercial"],
             common["opt_e_facture"], common["opt_e_communication"],
             common["opt_accept_com_parte"], common["opt_consent_consult_distri"],
             vte_add_part or "", int(op_id)),
        )
    else:  # eni
        db.query(
            f"""INSERT INTO adv.{table}
                  (id_contrat_option_auto, id_contrat, num_bs,
                   opt_mail, opt_energie_verte_gaz, opt_energie_verte_elec,
                   opt_reforestation, opt_optin_commercial,
                   opt_e_facture, opt_e_communication,
                   opt_accept_com_parte, opt_consent_consult_distri,
                   modif_op, modif_date, modif_elem)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NOW(), 'new')""",
            (auto_n, id_contrat, num_bs,
             common["opt_mail"], common["opt_energie_verte_gaz"],
             common["opt_energie_verte_elec"],
             common["opt_reforestation"], common["opt_optin_commercial"],
             common["opt_e_facture"], common["opt_e_communication"],
             common["opt_accept_com_parte"], common["opt_consent_consult_distri"],
             int(op_id)),
        )


def convert_selection_energie(
    id_tk_liste: int, ids_paniers: list[int], op_id: int,
) -> ConvertSelectionResult:
    """Convertit les paniers selectionnes en contrats.
    Cf code WinDev Fen_ContenuTicketCall 'Convertir la selection en contrat'.

    Regle 'testVteAddOen' : si un panier OEN est present dans le ticket,
    SEULS les paniers OEN sont convertis (les autres sont skippes)."""
    db_bo = get_pg_connection("ticket_bo")
    db_adv = get_pg_connection("adv")
    db_rh = get_pg_connection("rh")
    result = ConvertSelectionResult()

    # Charge TK_Call pour info client + vendeur
    tc = db_bo.query_one(
        """SELECT id_tk_call, id_salarie, id_client, civilite_client,
                   nom_client, prenom_client, nom_marital_client,
                   date_naiss, adresse1, adresse2, cp, ville,
                   mobile1, adr_mail, opt_partenaire
              FROM ticket_bo.pgt_tk_call
             WHERE id_tk_liste = ? LIMIT 1""",
        (int(id_tk_liste),),
    )
    if not tc:
        result.nb_erreurs = 1
        return result

    id_salarie = int(tc.get("id_salarie") or 0)
    if id_salarie == 0:
        # Fallback op_crea du TK_Liste
        db_tk = get_pg_connection("ticket")
        tl = db_tk.query_one(
            "SELECT op_crea, date_crea FROM ticket.pgt_tk_liste WHERE id_tk_liste = ?",
            (int(id_tk_liste),),
        )
        if tl: id_salarie = int(tl.get("op_crea") or 0)
    else:
        db_tk = get_pg_connection("ticket")
        tl = db_tk.query_one(
            "SELECT date_crea FROM ticket.pgt_tk_liste WHERE id_tk_liste = ?",
            (int(id_tk_liste),),
        )

    date_signature = None
    if tl and tl.get("date_crea"):
        d = tl["date_crea"]
        date_signature = d.date() if isinstance(d, datetime) else d

    # Id ste du vendeur
    id_ste = None
    if id_salarie:
        se = db_rh.query_one(
            "SELECT id_ste FROM rh.pgt_salarie_embauche WHERE id_salarie = ? LIMIT 1",
            (int(id_salarie),),
        )
        if se: id_ste = se.get("id_ste")

    # Charge TOUS les paniers du ticket (pour testVteAddOen)
    all_pans = db_bo.query(
        """SELECT id_tk_call_panier, id_produit, partenaire, num_bs,
                  statut_prod, motif_annulation, date_entree, observations,
                  opt_mail, opt_e_facture, opt_e_communication,
                  opt_optin_commercial, opt_consent_consult_distri,
                  opt_accept_com_parte, opt_mandat,
                  opt_energie_verte_gaz, opt_reforestation, format_numerique
             FROM ticket_bo.pgt_tk_call_panier
            WHERE id_tk_call = ?
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
        (int(tc["id_tk_call"]),),
    ) or []
    test_vte_add_oen = any((p.get("partenaire") or "").upper() == "OEN"
                           for p in all_pans)
    # Construit VteAddPart = liste concatenee
    vte_add_part = ";".join(
        f"{(p.get('partenaire') or '').upper()}-{int(p.get('id_produit') or 0)}"
        for p in all_pans
    )

    # Filtre sur les paniers demandes
    ids_pans_set = {int(x) for x in ids_paniers}
    paniers_to_process = [p for p in all_pans
                          if int(p["id_tk_call_panier"]) in ids_pans_set]

    for panier in paniers_to_process:
        id_p = int(panier["id_tk_call_panier"])
        partenaire = (panier.get("partenaire") or "").upper()
        num_bs = (panier.get("num_bs") or "").upper()
        statut_prod = int(panier.get("statut_prod") or 0)
        res = ConvertPanierResult(id_panier=str(id_p), partenaire=partenaire,
                                   num_bs=num_bs)

        try:
            # Regle testVteAddOen : skip si OEN present et panier != OEN
            if test_vte_add_oen and partenaire != "OEN":
                res.action = "skipped"
                res.message = "OEN présent dans le ticket → panier non-OEN skippé"
                result.nb_skipped += 1
                result.results.append(res)
                continue

            # Genere NUM si vide (TK + id_panier)
            if not num_bs:
                num_bs = f"TK{id_p}"
                db_bo.query(
                    """UPDATE ticket_bo.pgt_tk_call_panier
                          SET num_bs=?, num_date_saisie=NOW(),
                              modif_date=NOW(), modif_op=?, modif_elem='modif'
                        WHERE id_tk_call_panier=?""",
                    (num_bs, int(op_id), id_p),
                )
                res.num_bs = num_bs

            # Table cible
            part_lower = partenaire.lower()
            table_ctt = f"pgt_{part_lower}_contrat"

            # Verif doublon
            try:
                existing = db_adv.query_one(
                    f"""SELECT id_contrat, id_salarie, id_client, id_etat_contrat
                          FROM adv.{table_ctt}
                         WHERE UPPER(num_bs) = UPPER(?)
                           AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                         LIMIT 1""",
                    (num_bs,),
                )
            except Exception:
                existing = None    # Table inconnue -> nouveau partenaire

            if existing:
                # DOUBLON - reaffecte vendeur si generic
                cur_sal = int(existing.get("id_salarie") or 0)
                if cur_sal in (0, CALL_GENERIC_SALARIE_ID) and id_salarie > 0:
                    db_adv.query(
                        f"""UPDATE adv.{table_ctt}
                              SET id_salarie=?, modif_date=NOW(), modif_op=?,
                                  modif_elem='modif'
                            WHERE id_contrat=?""",
                        (id_salarie, int(op_id), int(existing["id_contrat"])),
                    )
                    res.action = "updated"
                    res.id_statut = 4
                    res.message = "Contrat existant : vendeur réaffecté"
                    result.nb_updates += 1
                else:
                    res.action = "doublon"
                    res.id_statut = 33
                    res.message = "Contrat DOUBLON (déjà existant, vendeur ≠ générique)"
                    result.nb_skipped += 1

                # Update fiche client
                _modif_fiche_client(db_adv, int(existing.get("id_client") or 0),
                                     tc, op_id)

                # Update etat_contrat selon regles WinDev
                cur_etat = int(existing.get("id_etat_contrat") or 0)
                new_etat = None
                if partenaire == "OEN" and cur_etat == 8:
                    new_etat = _get_etat_contrat_oen(statut_prod, num_bs)
                elif partenaire == "ENI" and cur_etat == 51:
                    new_etat = _get_etat_contrat_eni(statut_prod, num_bs)
                if new_etat is not None:
                    db_adv.query(
                        f"""UPDATE adv.{table_ctt}
                              SET id_etat_contrat=?, modif_date=NOW(),
                                  modif_op=?, modif_elem='modif'
                            WHERE id_contrat=?""",
                        (new_etat, int(op_id), int(existing["id_contrat"])),
                    )
                # Update VteAddPart pour OEN
                if partenaire == "OEN":
                    db_adv.query(
                        """UPDATE adv.pgt_oen_contrat_option
                              SET opt_vte_add_part=?, modif_date=NOW(),
                                  modif_op=?, modif_elem='modif'
                            WHERE id_contrat=?""",
                        (vte_add_part, int(op_id), int(existing["id_contrat"])),
                    )
            else:
                # Nouveau contrat
                # 1. Client
                id_client = _enregistrer_client_energie(db_adv, tc, op_id)
                if not id_client:
                    res.action = "error"
                    res.message = "Échec création client"
                    result.nb_erreurs += 1
                    result.results.append(res)
                    continue

                # 2. Etat contrat selon partenaire
                if partenaire == "OEN":
                    id_etat = _get_etat_contrat_oen(statut_prod, num_bs)
                elif partenaire == "ENI":
                    id_etat = _get_etat_contrat_eni(statut_prod, num_bs)
                else:
                    id_etat = _get_etat_contrat_default(statut_prod, num_bs)

                # 3. Lookup lib_produit + famille + sous_fam
                lib = ""; fam = ""; ss_fam = ""
                pid = int(panier.get("id_produit") or 0)
                if pid:
                    table_prod = f"pgt_{part_lower}_produit"
                    try:
                        pr = db_adv.query_one(
                            f"SELECT lib_produit, famille, sous_fam FROM adv.{table_prod} WHERE id_produit = ? LIMIT 1",
                            (pid,),
                        )
                        if pr:
                            lib = pr.get("lib_produit") or ""
                            fam = pr.get("famille") or ""
                            ss_fam = pr.get("sous_fam") or ""
                    except Exception:
                        pass

                # 4. INSERT contrat + option
                id_ctt, nbpt = _insert_contrat(
                    db_adv, partenaire, panier, tc, id_client, id_salarie,
                    id_ste, id_etat, num_bs, date_signature,
                    lib, fam, ss_fam, op_id,
                )
                _insert_contrat_option(
                    db_adv, partenaire, id_ctt, num_bs, panier,
                    vte_add_part, op_id,
                )

                res.action = "created"
                res.id_statut = 4
                res.message = f"Contrat créé (nbPts={nbpt:.2f})"
                result.nb_crees += 1
        except Exception as e:
            res.action = "error"
            res.message = f"Erreur : {e}"
            result.nb_erreurs += 1
        result.results.append(res)

    return result


class ConvertTicketResult(BaseModel):
    id_tk_liste: str
    nb_paniers: int = 0
    nb_updates: int = 0
    nb_skipped: int = 0
    nb_erreurs: int = 0
    cloture_ok: bool = False
    message: str = ""


def convert_tickets_selection_energie(
    ids_tk_liste: list[int], op_id: int,
) -> list[ConvertTicketResult]:
    """Bouton 'Convertir la selection' au niveau liste Fen_TicketCall.

    ATTENTION : cf code WinDev, cette version NE CREE PAS de nouveaux
    contrats (le bloc EnregistrerCtt est commente). Elle fait uniquement
    des updates sur les contrats existants + cloture le ticket a la fin.

    Regles par partenaire pour l'update d'etat :
      - PRO : si etat_courant=1  -> nouvel etat selon PRO (5/6/7)
      - ENI : si etat_courant=51 -> nouvel etat selon ENI (37/66/67)
      - OEN : si type_etat <= 2  -> nouvel etat selon OEN v2
              (40/38/39/42/51 selon statut_prod + TK) +
              update option OPT_VteAdd_Part
    """
    db_bo = get_pg_connection("ticket_bo")
    db_tk = get_pg_connection("ticket")
    db_adv = get_pg_connection("adv")
    out: list[ConvertTicketResult] = []

    for id_tl in ids_tk_liste:
        res = ConvertTicketResult(id_tk_liste=str(id_tl))

        tc = db_bo.query_one(
            """SELECT id_tk_call, id_salarie, id_client, civilite_client,
                      nom_client, prenom_client, nom_marital_client,
                      date_naiss, adresse1, adresse2, cp, ville,
                      mobile1, adr_mail, opt_partenaire
                 FROM ticket_bo.pgt_tk_call
                WHERE id_tk_liste = ? LIMIT 1""",
            (int(id_tl),),
        )
        if not tc:
            res.message = "TK_Call introuvable"
            out.append(res); continue

        # Vendeur du ticket (via TK_Call ou op_crea)
        id_salarie_ticket = int(tc.get("id_salarie") or 0)
        if id_salarie_ticket == 0:
            tl = db_tk.query_one(
                "SELECT op_crea FROM ticket.pgt_tk_liste WHERE id_tk_liste = ?",
                (int(id_tl),),
            )
            if tl: id_salarie_ticket = int(tl.get("op_crea") or 0)

        # Tous les paniers
        pans = db_bo.query(
            """SELECT id_tk_call_panier, id_produit, partenaire, num_bs,
                      statut_prod, motif_annulation, opt_mail,
                      opt_e_facture, opt_e_communication,
                      opt_optin_commercial, opt_consent_consult_distri,
                      opt_accept_com_parte, opt_mandat,
                      opt_energie_verte_gaz, opt_reforestation,
                      format_numerique
                 FROM ticket_bo.pgt_tk_call_panier
                WHERE id_tk_call = ?
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
            (int(tc["id_tk_call"]),),
        ) or []
        res.nb_paniers = len(pans)

        test_vte_add_oen = any((p.get("partenaire") or "").upper() == "OEN"
                                for p in pans)
        vte_add_part = ";".join(
            f"{(p.get('partenaire') or '').upper()}-{int(p.get('id_produit') or 0)}"
            for p in pans
        )

        for panier in pans:
            id_p = int(panier["id_tk_call_panier"])
            partenaire = (panier.get("partenaire") or "").upper()
            num_bs = (panier.get("num_bs") or "").upper()
            statut_prod = int(panier.get("statut_prod") or 0)

            try:
                # Skip si OEN present et panier != OEN
                if test_vte_add_oen and partenaire != "OEN":
                    res.nb_skipped += 1
                    continue

                # Genere NUM si vide (TK + id_panier)
                if not num_bs:
                    num_bs = f"TK{id_p}"
                    db_bo.query(
                        """UPDATE ticket_bo.pgt_tk_call_panier
                              SET num_bs=?, num_date_saisie=NOW(),
                                  modif_date=NOW(), modif_op=?, modif_elem='modif'
                            WHERE id_tk_call_panier=?""",
                        (num_bs, int(op_id), id_p),
                    )

                # Cherche doublon dans adv.pgt_{part}_contrat
                part_lower = partenaire.lower()
                table_ctt = f"pgt_{part_lower}_contrat"
                try:
                    existing = db_adv.query_one(
                        f"""SELECT id_contrat, id_salarie, id_client,
                                   id_etat_contrat
                              FROM adv.{table_ctt}
                             WHERE UPPER(num_bs) = UPPER(?)
                               AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                             LIMIT 1""",
                        (num_bs,),
                    )
                except Exception:
                    existing = None

                if not existing:
                    # Bouton liste : PAS de creation nouveau contrat (WinDev commente)
                    res.nb_skipped += 1
                    continue

                # DOUBLON : update selon partenaire
                # Reaffecte vendeur si generic (WinDev : if IDSalarie=0)
                cur_sal = int(existing.get("id_salarie") or 0)
                if cur_sal == 0 and id_salarie_ticket > 0:
                    db_adv.query(
                        f"""UPDATE adv.{table_ctt}
                              SET id_salarie=?, modif_date=NOW(), modif_op=?,
                                  modif_elem='modif'
                            WHERE id_contrat=?""",
                        (id_salarie_ticket, int(op_id),
                         int(existing["id_contrat"])),
                    )

                # Update fiche client si vide
                _modif_fiche_client(db_adv,
                                     int(existing.get("id_client") or 0),
                                     tc, op_id)

                # Update etat_contrat selon partenaire
                cur_etat = int(existing.get("id_etat_contrat") or 0)
                new_etat = None
                if partenaire == "PRO" and cur_etat == 1:
                    new_etat = _get_etat_contrat_pro(statut_prod, num_bs)
                elif partenaire == "ENI" and cur_etat == 51:
                    new_etat = _get_etat_contrat_eni(statut_prod, num_bs)
                elif partenaire == "OEN":
                    # Regle OEN : verifie IDTypeEtat <= 2 via pgt_oen_etat_contrat
                    try:
                        oen_etat = db_adv.query_one(
                            """SELECT id_type_etat FROM adv.pgt_oen_etat_contrat
                                WHERE id_etat = ? LIMIT 1""",
                            (cur_etat,),
                        )
                        if oen_etat and int(oen_etat.get("id_type_etat") or 0) <= 2:
                            new_etat = _get_etat_contrat_oen_v2(statut_prod, num_bs)
                    except Exception:
                        pass

                if new_etat is not None:
                    db_adv.query(
                        f"""UPDATE adv.{table_ctt}
                              SET id_etat_contrat=?, modif_date=NOW(),
                                  modif_op=?, modif_elem='modif'
                            WHERE id_contrat=?""",
                        (new_etat, int(op_id), int(existing["id_contrat"])),
                    )

                # OEN : update ou insert opt_vte_add_part
                if partenaire == "OEN":
                    opt_exists = db_adv.query_one(
                        """SELECT id_contrat_option_auto
                             FROM adv.pgt_oen_contrat_option
                            WHERE id_contrat = ? LIMIT 1""",
                        (int(existing["id_contrat"]),),
                    )
                    if opt_exists:
                        db_adv.query(
                            """UPDATE adv.pgt_oen_contrat_option
                                  SET opt_vte_add_part=?, modif_date=NOW(),
                                      modif_op=?, modif_elem='modif'
                                WHERE id_contrat=?""",
                            (vte_add_part, int(op_id),
                             int(existing["id_contrat"])),
                        )
                    else:
                        auto = db_adv.query_one(
                            "SELECT COALESCE(MAX(id_contrat_option_auto), 0) + 1 AS n FROM adv.pgt_oen_contrat_option"
                        )
                        auto_n = int(auto["n"]) if auto else 1
                        db_adv.query(
                            """INSERT INTO adv.pgt_oen_contrat_option
                                  (id_contrat_option_auto, id_contrat, num_bs,
                                   opt_vte_add_part, opt_energie_verte_elec,
                                   modif_op, modif_date, modif_elem)
                               VALUES (?, ?, ?, ?, FALSE, ?, NOW(), 'new')""",
                            (auto_n, int(existing["id_contrat"]), num_bs,
                             vte_add_part, int(op_id)),
                        )

                res.nb_updates += 1
            except Exception as e:
                res.nb_erreurs += 1
                if not res.message:
                    res.message = f"Erreur panier {id_p} : {e}"

        # Cloture le ticket (equivalent au reqUdateTicket final WinDev)
        try:
            db_tk.query(
                """UPDATE ticket.pgt_tk_liste
                      SET cloturee=TRUE, date_cloture=NOW(),
                          modif_date=NOW(), modif_op=?, modif_elem='modif'
                    WHERE id_tk_liste=?""",
                (int(op_id), int(id_tl)),
            )
            res.cloture_ok = True
        except Exception as e:
            res.message = (res.message + " | " if res.message else "") + f"Cloture KO : {e}"

        if not res.message:
            res.message = f"{res.nb_updates} maj, {res.nb_skipped} skip"
        out.append(res)
    return out


def cloture_selection_tickets_energie(
    ids_tk_liste: list[int], op_id: int,
) -> list[ConvertTicketResult]:
    """Bouton 'Cloturer sans convertir' : UPDATE cloturee=TRUE sur
    tous les tickets selectionnes."""
    db_tk = get_pg_connection("ticket")
    out: list[ConvertTicketResult] = []
    for id_tl in ids_tk_liste:
        res = ConvertTicketResult(id_tk_liste=str(id_tl))
        try:
            db_tk.query(
                """UPDATE ticket.pgt_tk_liste
                      SET cloturee=TRUE, date_cloture=NOW(),
                          modif_date=NOW(), modif_op=?, modif_elem='modif'
                    WHERE id_tk_liste=?""",
                (int(op_id), int(id_tl)),
            )
            res.cloture_ok = True
            res.message = "Clôturé"
        except Exception as e:
            res.message = f"KO : {e}"
        out.append(res)
    return out


def cloturer_ticket_call_energie(id_tk_liste: int, op_id: int) -> bool:
    """Clôture le ticket call (cf 2e OuiNon 'Souhaitez-vous cloturé le ticket ?')."""
    db = get_pg_connection("ticket")
    db.query(
        """UPDATE ticket.pgt_tk_liste
              SET cloturee=TRUE, date_cloture=NOW(),
                  modif_date=NOW(), modif_op=?, modif_elem='modif'
            WHERE id_tk_liste=?""",
        (int(op_id), int(id_tk_liste)),
    )
    return True


def resolve_call_justif_url(
    id_tk_call: int, id_panier: int, partenaire: str, source: str = "normal",
) -> str:
    """Voir le justif (Fen_ContenuTicketCall) :
    - Si partenaire OEN : {IDTK_Call_Panier}_Clarification.pdf
    - Sinon : {idCall}_Justif.jpg
    source : 'normal' -> groupe-exo, 'sos' -> sos.groupe-exo.

    Pas de HEAD ici : URL unique (pas de fallback jpg->pdf comme pour
    les CIN SFR), donc pas besoin de la lib 'requests'."""
    base = ("https://sos.groupe-exo.omaya.fr" if source == "sos"
            else "https://groupe-exo.omaya.fr")
    if (partenaire or "").upper() == "OEN":
        return f"{base}/DocOmaya/{id_panier}_Clarification.pdf"
    return f"{base}/DocOmaya/{id_tk_call}_Justif.jpg"


def export_extraction_energie_xlsx(
    rows: list[ExtractionEnergieRow],
) -> bytes:
    """XLSX simple (pas de couleurs necessaires ici)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    columns: list[tuple[str, str]] = [
        ("date_souscription", "Date de souscription"),
        ("numero_cm", "Numéro de CM"),
        ("nom", "Nom"),
        ("prenom", "Prénom"),
        ("telephone", "Téléphone"),
        ("adresse_mail", "Adresse mail"),
        ("date_activation", "Date d'activation"),
        ("type_contrat", "Type de contrat"),
        ("commercial", "Commercial"),
    ]
    wb = Workbook(); ws = wb.active; ws.title = "Extraction Énergie"
    ws.append([lbl for _, lbl in columns])
    header_fill = PatternFill(start_color="FF17494E", end_color="FF17494E",
                               fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill; cell.font = header_font

    for r in rows:
        ws.append([getattr(r, k) for k, _ in columns])

    for i, w in enumerate([16, 20, 20, 20, 14, 28, 16, 12, 22], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()
