"""
Service Fen_CalculPointsBS - Recalcul des points contrats.

Cf. WinDev Recalcul_Point(part) : pour chaque partenaire, recalcule
les points de chaque contrat entre Du et Au. Compare avec la valeur
en base et retourne la liste des contrats dont les points changent.
En mode non simulation : UPDATE {partenaire}_contrat.nb_points.

Reutilise :
  - app.shared.sdtc.bareme.calcul_point_contrat (238 lignes deja
    transposees depuis calculPointContrat WinDev)
  - app.shared.sdtc.helpers.donne_fam_prod_sfr (mapping SFR)
"""
from __future__ import annotations

import io as _io
import logging
from datetime import datetime
from typing import Optional

from app.core.database.pg import get_pg_connection
from app.intranets.adm.schemas.calcul_points_bs import (
    ContratModifieRow, ExportXlsxParams, PartenaireCombo,
    RecalculParams, RecalculResult,
)

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------

def _clean_id(v) -> str:
    if v is None:
        return ""
    try:
        n = int(v)
        return str(n) if n else ""
    except (TypeError, ValueError):
        return ""


def _iso_date(v) -> str:
    if v is None:
        return ""
    s = str(v)[:10]
    if s.startswith("1900") or s.startswith("0000"):
        return ""
    return s


# --------------------------------------------------------------------
# Combo Partenaires actifs
# --------------------------------------------------------------------

def list_partenaires() -> list[PartenaireCombo]:
    """Cf. Combo Partenaire : liste des partenaires actifs.
    Filtre : is_actif=TRUE + modif_elem NOT LIKE 'suppr%'.
    """
    db = get_pg_connection("rh")
    try:
        rows = db.query(
            """SELECT lib_partenaire, prefixe_bdd
                 FROM adv.pgt_partenaire
                WHERE (is_actif IS NULL OR is_actif = TRUE)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                ORDER BY lib_partenaire ASC""",
        ) or []
    except Exception:
        logger.exception("list_partenaires")
        return []
    return [
        PartenaireCombo(
            prefixe_bdd=(r.get("prefixe_bdd") or "").strip(),
            lib_partenaire=(r.get("lib_partenaire") or "").strip(),
        )
        for r in rows
        if (r.get("prefixe_bdd") or "").strip()
    ]


# --------------------------------------------------------------------
# Recalcul_Point : proc principale
# --------------------------------------------------------------------

_KNOWN_PREFIXES = {"ENI", "IAG", "SFR", "STR", "VAL", "TLC", "OEN", "PRO"}


def _load_contrats(
    prefixe: str, du: str, au: str,
) -> list[dict]:
    """Cf. WinDev reqProd : charge les contrats du partenaire entre
    Du et Au avec produit + etat_contrat.
    """
    if prefixe not in _KNOWN_PREFIXES:
        return []
    p = prefixe.lower()
    db = get_pg_connection("rh")

    # Colonnes de base
    base_cols = (
        "ct.id_contrat, ct.id_salarie, ct.num_bs, "
        "ct.info_interne, ct.id_produit, "
        "ct.id_etat_contrat, ct.date_signature, "
        "ct.nb_points, ct.mois_p AS mois_p, "
        "prod.lib_produit, prod.famille, prod.sous_fam, "
        "et.lib_etat, et.id_type_etat"
    )

    # Colonnes speciales ENI
    if prefixe == "ENI":
        base_cols += (
            ", ct.gaz_car_declaree, ct.gaz_car_relevee, "
            "ct.elec_puissance, ct.gaz_actif, ct.elec_actif, "
            "ct.notation, ct.opt_mandat"
        )

    # Colonnes speciales SFR (type_vente)
    if prefixe == "SFR":
        base_cols += ", ct.type_vente"

    sql = f"""
        SELECT {base_cols}
          FROM adv.pgt_{p}_contrat ct
          JOIN adv.pgt_{p}_produit prod ON prod.id_produit = ct.id_produit
          JOIN adv.pgt_{p}_etat_contrat et ON et.id_etat = ct.id_etat_contrat
         WHERE ct.date_signature BETWEEN ? AND ?
           AND (ct.modif_elem IS NULL
                OR ct.modif_elem NOT LIKE '%suppr%')
        ORDER BY ct.date_signature ASC"""
    try:
        return db.query(sql, (du, au)) or []
    except Exception:
        logger.exception("_load_contrats prefixe=%s", prefixe)
        return []


def _eni_options_string(id_contrat: int) -> str:
    """Cf. WinDev : construit 'NOTE:{notation}//RIB//MAIL//MAINT//'
    depuis pgt_eni_contrat + pgt_eni_contrat_option.
    """
    db = get_pg_connection("rh")
    try:
        r = db.query_one(
            """SELECT c.notation, c.opt_mandat,
                      opt.opt_mail, opt.opt_entretien
                 FROM adv.pgt_eni_contrat c
                 LEFT JOIN adv.pgt_eni_contrat_option opt
                        ON opt.id_contrat = c.id_contrat
                WHERE c.id_contrat = ?
                LIMIT 1""",
            (id_contrat,),
        )
    except Exception:
        return ""
    if not r:
        return ""
    parts: list[str] = []
    notation = r.get("notation")
    if notation is not None:
        parts.append(f"NOTE:{int(notation) if notation == int(notation) else notation}//")
    if bool(r.get("opt_mandat")):
        parts.append("RIB//")
    if bool(r.get("opt_mail")):
        parts.append("MAIL//")
    if bool(r.get("opt_entretien")):
        parts.append("MAINT//")
    return "".join(parts)


def recalcul_points(p: RecalculParams) -> RecalculResult:
    """Cf. WinDev Recalcul_Point(part) - transposition Python.

    Boucle sur chaque contrat du partenaire entre Du et Au :
    - Cas SFR : mapping DonneFamProdSFR + branche FIBRE vs MOBILE
    - Cas ENI : Car = GazCarRelevee sinon GazCarDeclaree, string options
      composee (NOTE/RIB/MAIL/MAINT), palier2 = PuissanceElec
    - Autres : appel simple avec NumBS en info_cplt
    Puis compare avec nb_points en base : si different, ajoute a la
    liste + eventuel UPDATE.
    """
    if p.prefixe not in _KNOWN_PREFIXES:
        return RecalculResult(
            ok=False, message=f"Partenaire inconnu : {p.prefixe}",
        )
    if not p.du or not p.au or len(p.du) < 10 or len(p.au) < 10:
        return RecalculResult(
            ok=False, message="Dates invalides (YYYY-MM-DD)",
        )
    if p.du > p.au:
        return RecalculResult(
            ok=False, message="Erreur sur le choix des dates !",
        )

    # Imports lazy (evite cycles si dispatch au boot)
    from app.shared.sdtc.bareme import calcul_point_contrat
    from app.shared.sdtc.helpers import donne_fam_prod_sfr

    contrats = _load_contrats(p.prefixe, p.du, p.au)
    lignes: list[ContratModifieRow] = []
    nb_maj = 0

    db_maj = get_pg_connection("rh")
    prefixe_lower = p.prefixe.lower()

    for r in contrats:
        id_ctt = int(r.get("id_contrat") or 0)
        famille = (r.get("famille") or "").strip()
        sous_fam = (r.get("sous_fam") or "").strip()
        date_sign_raw = r.get("date_signature")
        date_sign_str = ""
        if date_sign_raw:
            # Format WinDev "YYYYMMDD" pour les comparaisons
            date_sign_str = str(date_sign_raw)[:10].replace("-", "")

        nb_points_actuel = float(r.get("nb_points") or 0)
        nb_pt_new: float = 0.0

        # Champs remplis dans TableContratModif
        car_col = 0
        kva_col = 0
        nb_opt_str = ""

        if p.prefixe == "SFR":
            type_vente = int(r.get("type_vente") or 0)
            fam_sfr = donne_fam_prod_sfr(famille, type_vente)
            if famille == "FIBRE":
                # calculPointContrat(fam, ss_fam, 0, dateref, id_contrat)
                nb_pt_new = calcul_point_contrat(
                    fam_sfr, sous_fam, 0, date_sign_str, str(id_ctt),
                )
            else:
                # calculPointContrat(fam, ss_fam, "", dateref, lib_produit)
                nb_pt_new = calcul_point_contrat(
                    fam_sfr, sous_fam, "", date_sign_str,
                    (r.get("lib_produit") or "").strip(),
                )
        elif p.prefixe == "ENI":
            car_col = int(r.get("gaz_car_declaree") or 0)
            car_relevee = int(r.get("gaz_car_relevee") or 0)
            if car_relevee > 0:
                car_col = car_relevee
            kva_col = int(r.get("elec_puissance") or 0)
            nb_opt_str = _eni_options_string(id_ctt)
            nb_pt_new = calcul_point_contrat(
                famille, sous_fam, car_col,
                date_sign_str, nb_opt_str, kva_col,
            )
        else:
            # Autres partenaires : info_cplt = NumBS
            nb_pt_new = calcul_point_contrat(
                famille, sous_fam, "",
                date_sign_str, (r.get("num_bs") or "").strip(),
            )

        # Arrondi 2 decimales (fidele WinDev)
        nb_pt_new_r = round(float(nb_pt_new), 2)
        if abs(nb_points_actuel - nb_pt_new_r) < 0.001:
            continue

        lignes.append(ContratModifieRow(
            id_contrat=_clean_id(id_ctt),
            part=p.prefixe,
            num_bs=(r.get("num_bs") or "").strip(),
            date_signature=_iso_date(r.get("date_signature")),
            famille=famille,
            ss_fam=sous_fam,
            car=car_col,
            kva=kva_col,
            nb_opt=nb_opt_str,
            lib_etat=(r.get("lib_etat") or "").strip(),
            id_type_etat=int(r.get("id_type_etat") or 0),
            nb_point_av=nb_points_actuel,
            nb_point_ap=nb_pt_new_r,
        ))

        # UPDATE en base si non simulation
        if not p.simulation:
            try:
                db_maj.execute(
                    f"""UPDATE adv.pgt_{prefixe_lower}_contrat
                          SET nb_points = ?, modif_date = NOW(),
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (nb_pt_new_r, id_ctt),
                )
                nb_maj += 1
            except Exception:
                logger.exception(
                    "UPDATE nb_points KO id=%s", id_ctt,
                )

    msg_base = (
        f"{len(contrats)} contrat(s) lu(s), "
        f"{len(lignes)} avec point(s) modifie(s)"
    )
    if p.simulation:
        message = msg_base + " (SIMULATION)"
    else:
        message = msg_base + f" - {nb_maj} MAJ base"

    return RecalculResult(
        ok=True,
        nb_ctts_lus=len(contrats),
        nb_modifies=len(lignes),
        lignes=lignes,
        message=message,
    )


# --------------------------------------------------------------------
# Export XLSX
# --------------------------------------------------------------------

def generer_xlsx(p: ExportXlsxParams) -> tuple[str, bytes]:
    """Cf. WinDev VersExcel : XLSX de TableContratModif avec mise en forme."""
    try:
        from openpyxl import Workbook  # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill  # noqa: PLC0415
    except ImportError:
        return ("", b"")

    wb = Workbook()
    ws = wb.active
    ws.title = "Points"
    font_h = Font(bold=True)
    fill_h = PatternFill(
        start_color="17494E", end_color="17494E", fill_type="solid",
    )
    font_h_white = Font(bold=True, color="FFFFFF")

    headers = [
        "Part", "NumBS", "Date Signature",
        "Famille", "Ss Fam", "Car", "KVA",
        "Options", "Lib Etat", "Pts Avant", "Pts Apres", "Diff",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = font_h_white
        c.fill = fill_h

    for r, line in enumerate(p.lignes, 2):
        ws.cell(row=r, column=1, value=line.part)
        ws.cell(row=r, column=2, value=line.num_bs)
        ws.cell(row=r, column=3, value=line.date_signature)
        ws.cell(row=r, column=4, value=line.famille)
        ws.cell(row=r, column=5, value=line.ss_fam)
        ws.cell(row=r, column=6, value=line.car)
        ws.cell(row=r, column=7, value=line.kva)
        ws.cell(row=r, column=8, value=line.nb_opt)
        ws.cell(row=r, column=9, value=line.lib_etat)
        ws.cell(row=r, column=10, value=line.nb_point_av)
        ws.cell(row=r, column=11, value=line.nb_point_ap)
        ws.cell(row=r, column=12,
                value=round(line.nb_point_ap - line.nb_point_av, 2))

    widths = [8, 15, 14, 15, 12, 8, 8, 30, 20, 10, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    fic_name = (
        f"Calcul_points_{p.prefixe}_du_{p.du}_au_{p.au}_"
        f"{datetime.now().strftime('%H%M')}.xlsx"
    )
    return (fic_name, buf.getvalue())
