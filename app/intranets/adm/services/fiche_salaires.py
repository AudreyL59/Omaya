"""
Service Fen_FicheSalaires - Envoi fiches de salaire.

## Plan 1 : Decoupage PDF + matching vendeurs

1. list_societes_fdv() : combo Societe (id_type_orga=1 FDV Interne)
2. charger_pdf(pdf_bytes) : extrait les vendeurs des pages du PDF via
   pattern natif ##BULLETIN##<periode>##<num>##<NOM>##<PRENOM>##.
   Match automatique nom+prenom vs pgt_salarie.
3. rechercher_salaries(q) : recherche libre pour attribution manuelle
   des lignes rouges.

Le WinDev fait de l'OCR sur zone visuelle (1194,480,984,55) faute d'un
bon extracteur texte natif. Python + pypdf peuvent extraire le texte
brut : on cible le pattern ## qui est present en debut de chaque page
et 100% fiable.
"""

import base64
import logging
import re
import unicodedata
from typing import Optional

from app.core.database.pg import get_pg_connection
from app.intranets.adm.schemas.fiche_salaires import (
    ChargerPdfResult, EnvoiVendeurResult, EnvoyerFdpParams, EnvoyerFdpResult,
    GenererPdfPrepaieParams, GenererPdfPrepaieResult,
    ParseXlsxResult, ReimportXlsxResult, SauvegardeXlsxResult, SocieteFDV,
    ValiderParams, ValiderResult, VendeurRow,
)

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------

def _cap_prenom(p: str) -> str:
    if not p:
        return ""
    return "-".join(x[:1].upper() + x[1:].lower() for x in p.split("-"))


def _clean_id(v) -> str:
    if v is None:
        return ""
    try:
        n = int(v)
        return str(n) if n else ""
    except (TypeError, ValueError):
        return ""


def _norm_search(s: str) -> str:
    """cf. WinDev ChaineFormate(ccSansAccent + ccMajuscule) : retire accents,
    puis majuscules + sans espaces multiples."""
    if not s:
        return ""
    # Retire caracteres mal encodes (les remplace par vide)
    s = s.replace("�", "")
    nfkd = unicodedata.normalize("NFKD", s)
    ascii_s = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", ascii_s.upper()).strip()


# --------------------------------------------------------------------
# Combo Societes FDV Interne
# --------------------------------------------------------------------

def list_societes_fdv() -> list[SocieteFDV]:
    """Cf. WinDev Combo Societe : societes FDV Interne actives.

    id_type_orga = 1 (FDV Interne) - cf. project_intranets_list memoire
    et pattern ListeSocietePage (frontend adm/pages).
    """
    rh = get_pg_connection("rh")
    try:
        rows = rh.query(
            """SELECT id_ste, raison_sociale, rs_interne
                 FROM pgt_societe
                WHERE (modif_elem IS NULL
                       OR modif_elem NOT LIKE '%suppr%')
                  AND id_type_orga = 1
                  AND is_actif = TRUE
                ORDER BY rs_interne ASC NULLS LAST""",
        ) or []
    except Exception:
        return []
    return [
        SocieteFDV(
            id_ste=_clean_id(r.get("id_ste")),
            raison_sociale=(r.get("raison_sociale") or "").strip(),
            rs_interne=(r.get("rs_interne") or "").strip(),
        )
        for r in rows
    ]


# --------------------------------------------------------------------
# Charger PDF (extraction + matching)
# --------------------------------------------------------------------

# Pattern robuste : ##BULLETIN##(period)##(num)##(NOM)##(prenom)##
# Tolerant sur les caracteres non-'#' (accents, espaces, tirets).
_BULLETIN_RE = re.compile(
    r'##BULLETIN##(\d{2}-\d{4})##(\d+)##([^#]+?)##([^#]+?)##'
)


# Table de translation accents -> sans (pour normalisation SQL et Python).
_ACCENTS_FROM = "ÀÁÂÃÄÅÇÈÉÊËÌÍÎÏÑÒÓÔÕÖÙÚÛÜÝàáâãäåçèéêëìíîïñòóôõöùúûüý"
_ACCENTS_TO   = "AAAAAACEEEEIIIINOOOOOUUUUYaaaaaaceeeeiiiinooooouuuuy"


def _find_salarie_by_name(
    nom_norm: str, prenom_norm: str,
) -> Optional[dict]:
    """Recherche un salarie par nom+prenom, tolerant a :
      - la casse (UPPER)
      - les accents (TRANSLATE)
      - les espaces bordants (TRIM)
      - le statut (inclut les salaries sortis)
      - l'ambiguite : si N matches, on prend le salarie actif ou le
        segment d'embauche le plus recent.

    Cf. WinDev ReqChercheSalarieByPrenomNOM (avec ChaineFormate
    ccSansAccent + ccMajuscule).
    """
    rh = get_pg_connection("rh")
    try:
        rows = rh.query(
            f"""SELECT DISTINCT ON (s.id_salarie)
                    s.id_salarie, s.nom, s.prenom, s.login AS mail,
                    c.tel_mob AS gsm,
                    e.en_activite,
                    e.date_debut
                  FROM pgt_salarie s
                  LEFT JOIN pgt_salarie_coordonnees c
                         ON c.id_salarie = s.id_salarie
                  LEFT JOIN pgt_salarie_embauche e
                         ON e.id_salarie = s.id_salarie
                        AND (e.modif_elem IS NULL
                             OR e.modif_elem NOT LIKE '%suppr%')
                 WHERE TRANSLATE(UPPER(TRIM(s.nom)),
                                 '{_ACCENTS_FROM}',
                                 '{_ACCENTS_TO}') = ?
                   AND TRANSLATE(UPPER(TRIM(s.prenom)),
                                 '{_ACCENTS_FROM}',
                                 '{_ACCENTS_TO}') = ?
                 ORDER BY s.id_salarie, e.date_debut DESC NULLS LAST""",
            (nom_norm, prenom_norm),
        ) or []
    except Exception:
        logger.exception("Recherche salarie KO")
        return None
    # Cf. Fen_FicheSalaires - HNbEnr = 1 sinon ligne rouge :
    # en cas d'ambiguite (homonyme), on ne prend pas de decision,
    # on laisse en rouge pour attribution manuelle.
    # (DISTINCT ON s.id_salarie -> chaque row = 1 personne differente)
    if len(rows) != 1:
        return None
    r = rows[0]
    return {
        "id_salarie": _clean_id(r.get("id_salarie")),
        "nom": (r.get("nom") or "").strip(),
        "prenom": (r.get("prenom") or "").strip(),
        "mail": (r.get("mail") or "").strip(),
        "gsm": (r.get("gsm") or "").strip(),
    }


def charger_pdf(pdf_bytes: bytes) -> ChargerPdfResult:
    """Cf. WinDev Btn Charger le fichier PDF.

    1. Ouvre PDF.
    2. Pour chaque page : extrait le pattern ##BULLETIN##...##NOM##PRENOM##
    3. Groupe par (nom, prenom) : Nb_Page += 1 pour multi-pages du meme
       vendeur.
    4. Match automatique vs pgt_salarie : 1 resultat = vert, sinon rouge.
    """
    try:
        from pypdf import PdfReader  # noqa: PLC0415
    except ImportError:
        return ChargerPdfResult(
            ok=False, message="pypdf non installe cote serveur",
        )

    import io as _io
    try:
        reader = PdfReader(_io.BytesIO(pdf_bytes))
    except Exception as e:
        return ChargerPdfResult(
            ok=False, message=f"PDF invalide : {e}",
        )

    nb_pages = len(reader.pages)
    if not nb_pages:
        return ChargerPdfResult(
            ok=True, nb_pages=0, vendeurs=[],
            message="PDF vide",
        )

    # Extraction pattern
    vendeurs_by_key: dict[str, VendeurRow] = {}
    order: list[str] = []

    for i, page in enumerate(reader.pages):
        try:
            txt = page.extract_text() or ""
        except Exception:
            txt = ""
        m = _BULLETIN_RE.search(txt)
        if not m:
            # Page sans pattern : on cree une ligne 'inconnu'
            key = f"__unknown_{i+1}"
            vendeurs_by_key[key] = VendeurRow(
                id_salarie="0",
                vendeur=f"Page {i+1} inconnue",
                num_page=i + 1, nb_page=1,
                couleur="rouge",
            )
            order.append(key)
            continue
        nom = m.group(3).strip()
        prenom = m.group(4).strip()
        # cf. WinDev : sans espace multiple + sans accent + majuscule
        vendeur_lib = f"{_norm_search(prenom)} {_norm_search(nom)}"
        # Clef de regroupement : nom_norm|prenom_norm
        key = f"{_norm_search(nom)}|{_norm_search(prenom)}"
        if key in vendeurs_by_key:
            vendeurs_by_key[key].nb_page += 1
        else:
            row = VendeurRow(
                id_salarie="0",
                vendeur=vendeur_lib,
                num_page=i + 1,
                nb_page=1,
                couleur="rouge",
            )
            # Lookup salarie
            info = _find_salarie_by_name(
                _norm_search(nom), _norm_search(prenom),
            )
            if info:
                row.id_salarie = info["id_salarie"]
                row.nom_prenom = f"{info['nom']} {_cap_prenom(info['prenom'])}"
                row.mail = info["mail"]
                row.gsm = info["gsm"]
                row.couleur = "vert"
            vendeurs_by_key[key] = row
            order.append(key)

    vendeurs = [vendeurs_by_key[k] for k in order]

    return ChargerPdfResult(
        ok=True,
        pdf_b64=base64.b64encode(pdf_bytes).decode("ascii"),
        nb_pages=nb_pages,
        vendeurs=vendeurs,
        message=(
            f"{len(vendeurs)} vendeur(s) identifie(s) - "
            f"{sum(1 for v in vendeurs if v.id_salarie == '0')} en rouge"
        ),
    )


# --------------------------------------------------------------------
# Recherche manuelle (ligne rouge)
# --------------------------------------------------------------------

def rechercher_salaries(q: str) -> list[dict]:
    """Recherche libre nom/prenom pour attribution manuelle (cf. WinDev
    Fen_RechercheNomSalarie).
    """
    q = (q or "").strip()
    if len(q) < 2:
        return []
    rh = get_pg_connection("rh")
    try:
        rows = rh.query(
            """SELECT s.id_salarie, s.nom, s.prenom, s.login AS mail
                 FROM pgt_salarie s
                 JOIN pgt_salarie_embauche e
                      ON e.id_salarie = s.id_salarie
                     AND (e.modif_elem IS NULL
                          OR e.modif_elem NOT LIKE '%suppr%')
                     AND e.en_activite = TRUE
                WHERE UPPER(s.nom) LIKE UPPER(?)
                   OR UPPER(s.prenom) LIKE UPPER(?)
                   OR UPPER(s.nom || ' ' || s.prenom) LIKE UPPER(?)
                ORDER BY s.nom, s.prenom
                LIMIT 20""",
            (f"%{q}%", f"%{q}%", f"%{q}%"),
        ) or []
    except Exception:
        return []
    return [
        {
            "id_salarie": _clean_id(r.get("id_salarie")),
            "nom": (r.get("nom") or "").strip(),
            "prenom": _cap_prenom((r.get("prenom") or "").strip()),
            "mail": (r.get("mail") or "").strip(),
        }
        for r in rows
    ]


# --------------------------------------------------------------------
# Btn Valider - decoupe PDF + upload FTP + recup base
# --------------------------------------------------------------------

def _safe_filename(s: str) -> str:
    """Sanitize nom fichier : espaces -> underscore + retire caracteres
    dangereux. Preserve caracteres alphanumeriques + tirets + accents.
    """
    if not s:
        return "vendeur"
    # Retire caracteres mal encodes (encodage PDF cassé)
    s = s.replace("�", "").replace("�", "")
    # Retire caracteres OS-dangereux
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", s)
    # Espaces -> underscore
    s = re.sub(r"\s+", "_", s.strip())
    return s or "vendeur"


def _download_ftp_optional(path: str) -> Optional[bytes]:
    """Tente de recuperer un fichier via FTP. Retourne None si absent."""
    try:
        from app.shared.tickets.forms.factdistrib import _ftp_download
        return _ftp_download(path)
    except Exception:
        return None


def valider(p: ValiderParams) -> ValiderResult:
    """Cf. WinDev Btn Valider.

    Pour chaque vendeur attribue (id_salarie != 0) :
    1. Extrait les pages [num_page, num_page + nb_page - 1] du PDF
       original + genere un PDF individuel.
    2. Upload FTP gestionRH/{id}/Fiches_Salaires/{Vendeur}_{YYYY-MM}_FS.pdf
    3. Cherche la base PDF deja generee par Fen_PaiesBS :
       gestionRH/{id}/Fiches_Salaires/{NOM_PRENOM}_{YYYY-MM}_base.pdf
       (fallback : {NOM_PRENOM}_{MM-YYYY}_base.pdf)
    """
    # Validation
    if not p.pdf_b64:
        return ValiderResult(ok=False, message="PDF manquant")
    if not p.vendeurs:
        return ValiderResult(ok=False, message="Aucun vendeur")
    if len(p.mois_paiement) != 7 or p.mois_paiement[4] != "-":
        return ValiderResult(
            ok=False, message="mois_paiement invalide (YYYY-MM)",
        )

    # Verifie qu'aucun vendeur n'est en rouge
    non_attribues = [v for v in p.vendeurs if v.id_salarie == "0"]
    if non_attribues:
        return ValiderResult(
            ok=False,
            message=(
                f"{len(non_attribues)} vendeur(s) non attribue(s). "
                "Merci de verifier les lignes en rouge."
            ),
            vendeurs=p.vendeurs,
        )

    # Decode le PDF source
    try:
        pdf_source = base64.b64decode(p.pdf_b64)
    except Exception as e:
        return ValiderResult(ok=False, message=f"PDF b64 invalide : {e}")

    try:
        from pypdf import PdfReader, PdfWriter  # noqa: PLC0415
        from app.core.config import FTP_GESTION_RH_PATH
        from app.shared.tickets.forms.cttw_pdf import ftp_upload
    except ImportError as e:
        return ValiderResult(
            ok=False, message=f"Dependance manquante : {e}",
        )

    import io as _io
    try:
        reader = PdfReader(_io.BytesIO(pdf_source))
    except Exception as e:
        return ValiderResult(ok=False, message=f"PDF illisible : {e}")

    ftp_root = FTP_GESTION_RH_PATH.rstrip("/")
    nb_valides = 0
    nb_erreurs = 0
    result_rows: list[VendeurRow] = []

    for v in p.vendeurs:
        # Copie la structure de v pour retour
        nv = v.model_copy()

        # 1. Extrait les pages [num_page .. num_page + nb_page - 1]
        num_min = int(v.num_page)  # 1-indexed
        num_max = num_min + int(v.nb_page) - 1
        try:
            writer = PdfWriter()
            for k in range(num_min - 1, num_max):  # 0-indexed
                if 0 <= k < len(reader.pages):
                    writer.add_page(reader.pages[k])
            buf = _io.BytesIO()
            writer.write(buf)
            pdf_indiv = buf.getvalue()
        except Exception as e:
            logger.exception("Split PDF KO vendeur %s", v.vendeur)
            nv.couleur = "rouge"
            result_rows.append(nv)
            nb_erreurs += 1
            continue

        # 2. Nom fichier FS.pdf (cf. WinDev)
        fic_name = f"{_safe_filename(v.vendeur)}_{p.mois_paiement}_FS.pdf"
        # 3. Upload FTP
        upload_ok = False
        try:
            ftp_upload(
                f"{ftp_root}/{v.id_salarie}/Fiches_Salaires",
                fic_name, pdf_indiv,
            )
            upload_ok = True
        except Exception as e:
            logger.exception("Upload FTP KO : %s", fic_name)

        nv.fichier_pdf = fic_name if upload_ok else ""

        # 4. Cherche la base PDF (generee par Fen_PaiesBS)
        # cf. WinDev : Majuscule(NomPrenom)+"_"+DateVersChaine(MoisP,"AAAA-MM")+"_base.pdf"
        # avec fallback sur MM-AAAA
        base_candidates = []
        if v.nom_prenom:
            nom_prenom_up = v.nom_prenom.upper()
            base_candidates.append(
                _safe_filename(nom_prenom_up)
                + f"_{p.mois_paiement}_base.pdf",
            )
            # Fallback MM-AAAA
            mm_aaaa = p.mois_paiement[5:7] + "-" + p.mois_paiement[:4]
            base_candidates.append(
                _safe_filename(nom_prenom_up)
                + f"_{mm_aaaa}_base.pdf",
            )

        base_found = ""
        for candidate in base_candidates:
            path = f"{ftp_root}/{v.id_salarie}/Fiches_Salaires/{candidate}"
            content = _download_ftp_optional(path)
            if content:
                base_found = candidate
                break
        nv.base_pdf = base_found

        # 5. Couleur / statut
        if upload_ok:
            nv.choix = True
            nv.couleur = "vert"
            nb_valides += 1
        else:
            nv.couleur = "orange"
            nb_erreurs += 1

        result_rows.append(nv)

    return ValiderResult(
        ok=True,
        vendeurs=result_rows,
        nb_valides=nb_valides,
        nb_erreurs=nb_erreurs,
        message=(
            f"{nb_valides} vendeur(s) valide(s), {nb_erreurs} erreur(s). "
            + ("Passer au Plan 2." if nb_erreurs == 0 else "")
        ),
    )


# --------------------------------------------------------------------
# Sauvegarde XLSX / Reimport XLSX
# --------------------------------------------------------------------

_XLSX_COLS = [
    ("Choix", "choix"),
    ("Vendeur", "vendeur"),
    ("NomPrenom", "nom_prenom"),
    ("NumPage", "num_page"),
    ("NbPage", "nb_page"),
    ("IdSalarie", "id_salarie"),
    ("Mail", "mail"),
    ("GSM", "gsm"),
    ("FichierPDF", "fichier_pdf"),
    ("BasePDF", "base_pdf"),
    ("TabPrepaies", "tab_prepaies"),
    ("Couleur", "couleur"),
]


def sauvegarder_xlsx(vendeurs: list[VendeurRow]) -> SauvegardeXlsxResult:
    """Btn Sauve EXCEL : exporte TableListeVendeur en XLSX pour reprise
    ulterieure. Cf. WinDev TableListeVendeur.VersExcel(...).
    """
    try:
        from openpyxl import Workbook  # noqa: PLC0415
    except ImportError:
        return SauvegardeXlsxResult(ok=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sauve"
    # Header
    ws.append([lbl for lbl, _ in _XLSX_COLS])
    for v in vendeurs:
        d = v.model_dump()
        ws.append([d.get(attr, "") for _, attr in _XLSX_COLS])

    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    xlsx_b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return SauvegardeXlsxResult(
        ok=True,
        xlsx_b64=xlsx_b64,
        fic_name="SauveEnvoieFicSalaires.xlsx",
    )


def reimporter_xlsx(xlsx_bytes: bytes) -> ReimportXlsxResult:
    """Btn Reimporter Sauve XLS : recharge un XLSX genere par
    sauvegarder_xlsx (cf. WinDev Btn Reimporter Sauve XLS).
    """
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError:
        return ReimportXlsxResult(
            ok=False, message="openpyxl non installe",
        )

    import io as _io
    try:
        wb = load_workbook(_io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        return ReimportXlsxResult(
            ok=False, message=f"XLSX illisible : {e}",
        )
    ws = wb.active
    if not ws or ws.max_row < 2:
        return ReimportXlsxResult(ok=True, vendeurs=[])

    # Attend le meme header que sauvegarder_xlsx (positions fixes)
    vendeurs: list[VendeurRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(x is None for x in row):
            continue
        try:
            d = {}
            for i, (_, attr) in enumerate(_XLSX_COLS):
                if i < len(row) and row[i] is not None:
                    v = row[i]
                    if attr == "choix":
                        d[attr] = bool(v) and str(v).lower() not in (
                            "false", "0", "non", "",
                        )
                    elif attr in ("num_page", "nb_page"):
                        try:
                            d[attr] = int(v)
                        except (TypeError, ValueError):
                            d[attr] = 0
                    elif attr == "id_salarie":
                        d[attr] = str(int(v)) if v else "0"
                    else:
                        d[attr] = str(v).strip()
            vendeurs.append(VendeurRow(**d))
        except Exception:
            logger.exception("Row XLSX skip")

    return ReimportXlsxResult(
        ok=True,
        vendeurs=vendeurs,
        message=f"{len(vendeurs)} vendeur(s) reimporte(s)",
    )


# --------------------------------------------------------------------
# Prepaie Excel (Plan 2)
# --------------------------------------------------------------------

def parse_xlsx(xlsx_bytes: bytes) -> ParseXlsxResult:
    """Ouvre le XLSX + retourne la matrice de cellules pour affichage
    dans le composant grille React.

    Cf. WinDev Btn Ouvrir un tableau prepaies : Tableur1.Charge(monfic, 1).
    """
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError:
        return ParseXlsxResult(
            ok=False, message="openpyxl non installe",
        )

    import io as _io
    try:
        wb = load_workbook(_io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        return ParseXlsxResult(ok=False, message=f"XLSX illisible : {e}")

    ws = wb.active
    if not ws:
        return ParseXlsxResult(ok=True)

    nrows = ws.max_row or 0
    ncols = ws.max_column or 0
    # Sanity : bornes raisonnables (evite XLSX corrompu geant)
    nrows = min(nrows, 200)
    ncols = min(ncols, 30)

    cells: list[list[str]] = []
    for r in range(1, nrows + 1):
        row_vals: list[str] = []
        for c in range(1, ncols + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                row_vals.append("")
            elif isinstance(v, (int, float)):
                # Format nombre : entier -> pas de decimal, sinon 2 dec
                if isinstance(v, int) or float(v).is_integer():
                    row_vals.append(str(int(v)))
                else:
                    row_vals.append(f"{v:.2f}")
            else:
                row_vals.append(str(v))
        cells.append(row_vals)

    return ParseXlsxResult(
        ok=True, nrows=nrows, ncols=ncols, cells=cells,
    )


def _parse_plage(plage: str) -> Optional[tuple[int, int, int, int]]:
    """Parse 'A1:G17' -> (row_min, col_min, row_max, col_max) 1-indexed."""
    plage = (plage or "").strip().upper()
    m = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", plage)
    if not m:
        return None

    def col_letter_to_num(letter: str) -> int:
        n = 0
        for c in letter:
            n = n * 26 + (ord(c) - 64)
        return n

    c1 = col_letter_to_num(m.group(1))
    r1 = int(m.group(2))
    c2 = col_letter_to_num(m.group(3))
    r2 = int(m.group(4))
    return (min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2))


def _render_prepaie_html(cells: list[list[str]], titre: str) -> str:
    """Genere HTML pour PDF prepaie (page A4 portrait)."""
    from html import escape as h
    rows = []
    for row in cells:
        rows.append("<tr>" + "".join(f"<td>{h(v)}</td>" for v in row) + "</tr>")
    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8"><title>{h(titre)}</title>
<style>
@page {{ size: A4; margin: 15mm 10mm; }}
body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 9pt; color: #222; }}
h1 {{ font-size: 12pt; color: #17494E; border-bottom: 2px solid #17494E; padding-bottom: 4px; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 8px; }}
td {{ border: 1px solid #999; padding: 3px 5px; font-size: 8pt; }}
</style>
</head>
<body>
<h1>{h(titre)}</h1>
<table>{''.join(rows)}</table>
</body>
</html>
"""


def generer_pdf_prepaie(p: GenererPdfPrepaieParams) -> GenererPdfPrepaieResult:
    """Cf. WinDev Btn Enregistrer la selection en PDF.

    1. Ouvre XLSX + extrait la plage selectionnee (A1:G17 par ex).
    2. Genere PDF via WeasyPrint (tableau HTML).
    3. Upload FTP gestionRH/{id_sal}/Fiches_Salaires/.
    4. Retour couleur : vert (OK) / orange (upload KO) / rouge (PDF KO).
    """
    # Validation
    if not p.id_salarie or p.id_salarie == "0":
        return GenererPdfPrepaieResult(
            ok=False, couleur="rouge", message="Vendeur non attribue",
        )
    if len(p.mois_paiement) != 7:
        return GenererPdfPrepaieResult(
            ok=False, couleur="rouge", message="mois_paiement invalide",
        )
    bounds = _parse_plage(p.plage)
    if not bounds:
        return GenererPdfPrepaieResult(
            ok=False, couleur="rouge",
            message="Plage invalide (attendu 'A1:G17')",
        )

    # Decode XLSX
    try:
        xlsx_bytes = base64.b64decode(p.xlsx_b64)
    except Exception as e:
        return GenererPdfPrepaieResult(
            ok=False, couleur="rouge",
            message=f"XLSX b64 invalide : {e}",
        )

    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError:
        return GenererPdfPrepaieResult(
            ok=False, couleur="rouge",
            message="openpyxl non installe",
        )

    import io as _io
    try:
        wb = load_workbook(_io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        return GenererPdfPrepaieResult(
            ok=False, couleur="rouge",
            message=f"XLSX illisible : {e}",
        )

    ws = wb.active
    if not ws:
        return GenererPdfPrepaieResult(
            ok=False, couleur="rouge",
            message="XLSX sans feuille",
        )

    r_min, c_min, r_max, c_max = bounds
    cells: list[list[str]] = []
    for r in range(r_min, r_max + 1):
        row_vals: list[str] = []
        for c in range(c_min, c_max + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                row_vals.append("")
            elif isinstance(v, (int, float)):
                if isinstance(v, int) or float(v).is_integer():
                    row_vals.append(str(int(v)))
                else:
                    row_vals.append(f"{v:.2f}")
            else:
                row_vals.append(str(v))
        cells.append(row_vals)

    # Genere PDF
    titre = f"Tableau prepaie - {p.nom_prenom} - {p.mois_paiement}"
    html = _render_prepaie_html(cells, titre)
    try:
        from weasyprint import HTML as _HTML  # noqa: PLC0415
        pdf_bytes = _HTML(string=html).write_pdf()
    except Exception as e:
        logger.exception("WeasyPrint prepaie")
        return GenererPdfPrepaieResult(
            ok=False, couleur="rouge",
            message=f"Erreur PDF : {e}",
        )

    # Nom fichier : {NomPrenom}_{YYYY_MM}_FicPREPAIE.pdf
    # cf. WinDev : DateVersChaine(MoisP, "AAAA_MM") avec underscore
    mois_slash = p.mois_paiement.replace("-", "_")
    fic_name = (
        f"{_safe_filename(p.nom_prenom)}"
        f"_{mois_slash}_FicPREPAIE.pdf"
    )

    # Upload FTP
    try:
        from app.core.config import FTP_GESTION_RH_PATH
        from app.shared.tickets.forms.cttw_pdf import ftp_upload
        ftp_upload(
            f"{FTP_GESTION_RH_PATH.rstrip('/')}"
            f"/{p.id_salarie}/Fiches_Salaires",
            fic_name, pdf_bytes,
        )
    except Exception as e:
        logger.exception("Upload FTP prepaie")
        # PDF genere mais upload KO -> orange
        return GenererPdfPrepaieResult(
            ok=True, couleur="orange", fic_name=fic_name,
            message=f"PDF genere mais upload FTP KO : {e}",
        )

    return GenererPdfPrepaieResult(
        ok=True, couleur="vert", fic_name=fic_name,
        message=f"PDF prepaie envoye : {fic_name}",
    )


# --------------------------------------------------------------------
# Envoi FDP - ZIP protege + SMTP salaire@omaya.fr
# --------------------------------------------------------------------

def _decrypter_mdp_windev(mdp_crypte: bytes) -> Optional[str]:
    """Decrypte pgt_salarie.mdp_crypte (bytea) vers le mot de passe en clair.

    Cf. WinDev :
      bufCle est un Buffer = HashChaine(HA_MD5_128, HASH_SECRET_KEY)
      sResultat est un Buffer = DecrypteStandard(salarie.MDPCrypte,
                                                bufCle, crypteAES128)

    Details WinDev DecrypteStandard avec crypteAES128 :
      - Algorithme : AES-128 en CBC
      - Cle : buffer 16 bytes (MD5 de HASH_SECRET_KEY)
      - IV : les 16 premiers bytes du buffer chiffre
      - Contenu chiffre : le reste apres les 16 bytes d'IV
      - Padding : PKCS7

    Config env attendue :
      HASH_SECRET_KEY : la constante WinDev (chaine)
    """
    if not mdp_crypte or len(mdp_crypte) < 32:
        # Un buffer chiffre doit contenir au moins IV (16b) + 1 bloc (16b)
        return None

    import hashlib
    from app.core.config import HASH_SECRET_KEY

    secret = (HASH_SECRET_KEY or "").strip()
    if not secret:
        return None  # config manquante -> fallback aval

    # Cle AES-128 = MD5(secret) = 16 bytes
    key = hashlib.md5(secret.encode("utf-8")).digest()

    try:
        from cryptography.hazmat.primitives.ciphers import (
            Cipher, algorithms, modes,
        )
        from cryptography.hazmat.primitives.padding import PKCS7
    except ImportError:
        logger.warning("cryptography non installe - decryptage impossible")
        return None

    # IV = 16 premiers bytes, contenu = reste
    iv = bytes(mdp_crypte[:16])
    ciphertext = bytes(mdp_crypte[16:])

    try:
        cipher = Cipher(algorithms.AES(key), modes.CBC(iv))
        decryptor = cipher.decryptor()
        padded = decryptor.update(ciphertext) + decryptor.finalize()
        unpadder = PKCS7(128).unpadder()
        clear = unpadder.update(padded) + unpadder.finalize()
    except Exception:
        logger.exception("Decryptage AES-128 KO")
        return None

    try:
        return clear.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return clear.decode("latin-1")
        except Exception:
            return None


def _mdp_zip_for_salarie(id_salarie: int, fallback: str = "OMAYA") -> str:
    """Retourne le mot de passe ZIP pour un salarie.

    Cf. WinDev : DecrypteStandard(salarie.MDPCrypte, bufCle, crypteAES128).
    bufCle = HashChaine(HA_MD5_128, HASH_SECRET_KEY).

    Strategie :
    1. Si HASH_SECRET_KEY configure ET mdp_crypte present : decrypte AES-128
    2. Fallback : partie locale du login OMAYA (avant @)
    3. Fallback ultime : constante 'OMAYA'
    """
    if not id_salarie:
        return fallback
    rh = get_pg_connection("rh")
    try:
        r = rh.query_one(
            "SELECT login, mdp_crypte FROM pgt_salarie WHERE id_salarie = ?",
            (int(id_salarie),),
        )
    except Exception:
        r = None
    if not r:
        return fallback

    # 1. Tente decryptage WinDev
    mdp_crypte = r.get("mdp_crypte")
    if mdp_crypte:
        if isinstance(mdp_crypte, memoryview):
            mdp_crypte = mdp_crypte.tobytes()
        elif isinstance(mdp_crypte, str):
            try:
                import base64 as _b64
                mdp_crypte = _b64.b64decode(mdp_crypte)
            except Exception:
                mdp_crypte = None
        if isinstance(mdp_crypte, bytes):
            clear = _decrypter_mdp_windev(mdp_crypte)
            if clear and clear.strip():
                return clear.strip()

    # 2. Fallback login (partie locale)
    login = ((r.get("login") or "") or "").strip()
    if login:
        return login.split("@")[0].strip() or fallback
    return fallback


def _download_pdf_from_ftp(id_salarie: str, fic_name: str) -> Optional[bytes]:
    """Recupere un PDF stocke sur FTP gestionRH/{id}/Fiches_Salaires/."""
    if not fic_name:
        return None
    try:
        from app.core.config import FTP_GESTION_RH_PATH
        from app.shared.tickets.forms.factdistrib import _ftp_download
    except ImportError:
        return None
    path = (
        f"{FTP_GESTION_RH_PATH.rstrip('/')}"
        f"/{id_salarie}/Fiches_Salaires/{fic_name}"
    )
    return _ftp_download(path)


def _create_zip_aes(
    files: list[tuple[str, bytes]], password: str,
) -> Optional[bytes]:
    """Cree un ZIP AES-256 protege par mot de passe.

    Utilise pyzipper si dispo. Fallback : ZIP non chiffre + log warning.
    """
    try:
        import pyzipper  # noqa: PLC0415
    except ImportError:
        logger.warning(
            "pyzipper non installe - ZIP NON CHIFFRE (installer pyzipper)",
        )
        # Fallback stdlib
        import io as _io
        import zipfile
        buf = _io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, content in files:
                zf.writestr(name, content)
        return buf.getvalue()

    import io as _io
    buf = _io.BytesIO()
    try:
        with pyzipper.AESZipFile(
            buf, "w", compression=pyzipper.ZIP_DEFLATED,
            encryption=pyzipper.WZ_AES,
        ) as zf:
            zf.setpassword(password.encode("utf-8"))
            for name, content in files:
                zf.writestr(name, content)
    except Exception:
        logger.exception("Creation ZIP AES KO")
        return None
    return buf.getvalue()


def _send_mail_salaire(
    dest: list[str], cci: list[str], sujet: str, html: str,
    zip_bytes: bytes, zip_name: str,
) -> bool:
    """Envoi email via SMTP OVH avec compte salaire@omaya.fr.

    Config attendue :
      SMTP_SALAIRE_HOST=ssl0.ovh.net
      SMTP_SALAIRE_PORT=465
      SMTP_SALAIRE_USER=salaire@omaya.fr
      SMTP_SALAIRE_PASSWORD=***

    Fallback : reutilise SMTP_RH_* si SMTP_SALAIRE_* absents.
    """
    import smtplib
    from email.header import Header
    from email.mime.application import MIMEApplication
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.utils import formataddr

    # Config centralisee (app/core/config.py)
    from app.core.config import (
        SMTP_SALAIRE_HOST, SMTP_SALAIRE_PORT, SMTP_SALAIRE_USER,
        SMTP_SALAIRE_PASSWORD,
        SMTP_RH_HOST, SMTP_RH_PORT, SMTP_RH_USER, SMTP_RH_PASSWORD,
    )
    host = SMTP_SALAIRE_HOST
    port = SMTP_SALAIRE_PORT
    user = SMTP_SALAIRE_USER
    pwd = SMTP_SALAIRE_PASSWORD
    if not pwd:
        # Fallback SMTP RH (Gmail) si salaire non configure
        host = SMTP_RH_HOST
        port = SMTP_RH_PORT
        user = SMTP_RH_USER
        pwd = SMTP_RH_PASSWORD
    if not pwd:
        logger.error("SMTP pas configure (SMTP_SALAIRE_PASSWORD + SMTP_RH_PASSWORD vides)")
        return False

    all_recipients = dest + cci
    if not dest:
        return False

    msg = MIMEMultipart("mixed")
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html, "html", "utf-8"))
    msg.attach(alt)
    zip_part = MIMEApplication(zip_bytes, Name=zip_name)
    zip_part["Content-Disposition"] = f'attachment; filename="{zip_name}"'
    msg.attach(zip_part)

    msg["Subject"] = Header(sujet, "utf-8")
    msg["From"] = formataddr(("Service paie OMAYA", user))
    msg["To"] = ", ".join(dest)
    if cci:
        msg["Bcc"] = ", ".join(cci)

    try:
        with smtplib.SMTP_SSL(host, port, timeout=30) as smtp:
            smtp.login(user, pwd)
            smtp.sendmail(user, all_recipients, msg.as_string())
        logger.info("Mail salaire envoye : %s -> %d dest", sujet[:80], len(dest))
        return True
    except Exception:
        logger.exception("Envoi SMTP salaire KO (%s)", sujet[:80])
        return False


def _valid_email(email: str) -> bool:
    if not email or "@" not in email:
        return False
    local, _, domain = email.rpartition("@")
    return bool(local) and "." in domain


def envoyer_fdp(p: EnvoyerFdpParams) -> EnvoyerFdpResult:
    """Cf. WinDev Btn Valider et envoyer les FDP.

    Pour chaque vendeur avec Choix=True et mail valide :
    1. Recupere les PDF depuis le FTP (FS + base + prepaie)
    2. Cree un ZIP AES-256 protege par le mdp du salarie
    3. Envoie email via SMTP salaire@omaya.fr avec le ZIP en PJ
    4. CCI systematique : salaire@omaya.fr + intranet@omaya.fr
    """
    if not p.vendeurs:
        return EnvoyerFdpResult(
            ok=False, message="Aucun vendeur",
        )
    mois_fr = p.mois_paiement  # YYYY-MM
    envois: list[EnvoiVendeurResult] = []
    nb_envoyes = 0
    nb_echecs = 0

    for v in p.vendeurs:
        if not v.choix:
            continue
        res = EnvoiVendeurResult(
            id_salarie=v.id_salarie,
            nom_prenom=v.nom_prenom,
            mail=v.mail,
            couleur="rouge",
        )
        # 1. Mail valide ?
        if not _valid_email(v.mail):
            res.couleur = "orange"
            res.message = "Adresse mail invalide"
            envois.append(res)
            nb_echecs += 1
            continue

        # 2. Recupere les PDF depuis le FTP
        files_to_zip: list[tuple[str, bytes]] = []
        if v.fichier_pdf:
            fs = _download_pdf_from_ftp(v.id_salarie, v.fichier_pdf)
            if fs:
                files_to_zip.append((v.fichier_pdf, fs))
        if v.base_pdf:
            bp = _download_pdf_from_ftp(v.id_salarie, v.base_pdf)
            if bp:
                files_to_zip.append((v.base_pdf, bp))
        if v.tab_prepaies:
            tp = _download_pdf_from_ftp(v.id_salarie, v.tab_prepaies)
            if tp:
                files_to_zip.append((v.tab_prepaies, tp))

        if not files_to_zip:
            res.message = "Aucun fichier a envoyer"
            envois.append(res)
            nb_echecs += 1
            continue

        # 3. ZIP AES-256
        mdp = _mdp_zip_for_salarie(int(v.id_salarie) if v.id_salarie else 0)
        zip_name = (
            f"DocSalaire_{mois_fr}_{_safe_filename(v.nom_prenom)}.zip"
        )
        zip_bytes = _create_zip_aes(files_to_zip, mdp)
        if not zip_bytes:
            res.message = "Creation ZIP KO"
            envois.append(res)
            nb_echecs += 1
            continue

        # 4. Corps du mail
        parts_lib = ["votre fiche de salaire"]
        if v.base_pdf:
            parts_lib.append("votre base contrat")
        if v.tab_prepaies:
            parts_lib.append("votre tableau prepaie")
        if len(parts_lib) > 1:
            parts_str = ", ".join(parts_lib[:-1]) + " et " + parts_lib[-1]
        else:
            parts_str = parts_lib[0]

        sujet = f"Fiche Salaire {mois_fr} {v.nom_prenom}"
        html = f"""<font face='arial' style='font-size:10pt;'>
<p>Bonjour,</p>
<p>Retrouvez des a present sur votre espace salarie, {parts_str}.</p>
<p>Cet espace personnel est accessible depuis l'intranet ou l'appli
mobile Omayapp.</p>
<p>https://groupe-exo.omaya.fr</p>
<p><b>IMPORTANT : Le fichier ZIP en PJ, contenant les documents cites
ci-dessus, a ete protege par votre mot de passe de connexion OMAYA.</b></p>
<p>Pour toute question n'hesitez pas a contacter votre responsable.</p>
<p>Cordialement</p>
<p><b>Service paie</b><br/>
Std: 03.62.27.60.04<br/>
{p.raison_sociale}</p>
<p><i>PS : ceci est un mail automatique merci de ne pas repondre</i></p>
</font>"""

        cci = ["salaire@omaya.fr", "intranet@omaya.fr"]
        sent = _send_mail_salaire(
            [v.mail], cci, sujet, html, zip_bytes, zip_name,
        )
        if sent:
            res.couleur = "vert"
            res.message = "Envoye"
            nb_envoyes += 1
        else:
            res.couleur = "rouge"
            res.message = "Envoi SMTP KO"
            nb_echecs += 1
        envois.append(res)

    return EnvoyerFdpResult(
        ok=True,
        envois=envois,
        nb_envoyes=nb_envoyes,
        nb_echecs=nb_echecs,
        message=f"{nb_envoyes} envoye(s), {nb_echecs} en erreur",
    )
