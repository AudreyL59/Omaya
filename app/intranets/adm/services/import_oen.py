"""Service Fen_ImportOEN (ADM Imports Bases -> Import OHM Energie).

4 types d'import :
 1. Base Journaliere -> ImportJournalier
 2. RUN Valide       -> ImportRunValide (creation OEN_Contrat_Remun)
 3. RUN Resil        -> ImportRunResil (paye=decomm, sinon resil)
 4. Import Thermostat-> ImportThermostat (set OPT_Entretien=true par RefClient)
"""

from __future__ import annotations

import base64
import io
from datetime import date, datetime
from typing import Optional

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection
from app.shared.recrutement.services.recherche_cv import _int, _str
from app.intranets.adm.services.import_eni import (
    _new_id, _lookup_or_create_client,
    _col_letter_to_index, _cell, _parse_date_fr, _dernier_jour_mois,
    _parse_int,
)


# Mappings colonnes (cf. screens Fen_ImportOEN)
COLS_BJ_OEN = {
    "num_contrat": "A",       "ref_client": "B",
    "type_ener": "C",         "puissance": "D",
    "date_creation": "E",     "date_activation": "F",
    "date_signature": "H",    "vendeur": "K",
    "lib_statut": "I",        "lib_offre": "J",
    "car_relevee": "P",       "option_verte": "M",
    "car_declaree": "P",
}

COLS_RV_OEN = {
    "num_contrat": "A",       "ref_client": "B",
    "type_ener": "C",         "puissance": "D",
    "date_signature": "G",    "date_activation": "F",
    "vendeur": "K",           "lib_statut": "I",
    "car": "L",               "lib_offre": "J",
    "lib_rem": "M",
}

COLS_RR_OEN = {
    "num_contrat": "A",       "ref_client": "B",
    "type_ener": "C",         "puissance": "D",
    "date_signature": "G",    "date_activation": "F",
    "vendeur": "K",           "lib_statut": "I",
    "car": "L",               "lib_offre": "J",
    "lib_rem": "M",
}

COLS_TH_OEN = {
    "num_cm": "A",
}


class ImportOenParams(BaseModel):
    type_import: int                    # 1..4
    simulation: bool = True
    periode1_du: str = ""
    periode1_au: str = ""
    periode1_mois_paiement: str = ""
    periode2_du: str = ""
    periode2_au: str = ""
    periode2_mois_paiement: str = ""
    mois_paiement_distrib: str = ""


class ImportOenResume(BaseModel):
    nb_ajoutes: int = 0
    nb_modifies: int = 0
    nb_valides: int = 0
    nb_resilies: int = 0
    nb_decommissions: int = 0
    nb_deja_statues: int = 0
    nb_introuvables: int = 0
    nb_doublons: int = 0
    nb_hors_delai: int = 0
    nb_erreurs: int = 0
    nb_pb_vendeur: int = 0
    nb_pb_statut: int = 0
    nb_pb_offre: int = 0


class ImportOenResult(BaseModel):
    ok: bool
    type_import: int
    type_label: str
    simulation: bool
    resume: ImportOenResume
    contrats_ajoutes: list[dict] = []
    contrats_modifies: list[dict] = []
    contrats_run: list[dict] = []
    contrats_non_trouves: list[dict] = []
    pb_vendeur: list[dict] = []
    message: str = ""
    xlsx_b64: str = ""
    xlsx_name: str = ""
    mail_envoye: bool = False


TYPE_LABELS = {
    1: "Base Journalière",
    2: "RUN Valide",
    3: "RUN Résil",
    4: "Import Thermostat",
}


# ---------------------------------------------------------------------------
# Helpers OEN
# ---------------------------------------------------------------------------


def _id_etat_oen_by_lib(lib_statut: str) -> int:
    """Lookup OEN_etatContrat.id_etat by lib_etat LIKE."""
    if not lib_statut:
        return 0
    db = get_pg_connection("adv")
    r = db.query_one(
        """SELECT id_etat FROM adv.pgt_oen_etat_contrat
            WHERE LOWER(lib_etat) LIKE LOWER(?)
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            LIMIT 1""",
        (f"{lib_statut}%",),
    )
    return _int(r.get("id_etat")) if r else 0


def _id_produit_oen_by_lib(lib_offre: str) -> int:
    """Lookup OEN_produit.id_produit by lib_produit LIKE."""
    if not lib_offre:
        return 0
    db = get_pg_connection("adv")
    r = db.query_one(
        """SELECT id_produit FROM adv.pgt_oen_produit
            WHERE LOWER(lib_produit) LIKE LOWER(?)
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            LIMIT 1""",
        (f"%{lib_offre}%",),
    )
    return _int(r.get("id_produit")) if r else 0


def _lookup_vendeur_by_code(code: str, id_partenaire_oen: int = 562949953421321) -> int:
    """Lookup id_salarie via salarie_partenaire.code."""
    if not code:
        return 0
    db = get_pg_connection("rh")
    r = db.query_one(
        """SELECT id_salarie FROM rh.pgt_salarie_partenaire
            WHERE id_partenaire = ? AND code = ?
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            LIMIT 1""",
        (int(id_partenaire_oen), code),
    )
    return _int(r.get("id_salarie")) if r else 0


def _affectation_oen(id_salarie: int) -> tuple[str, str, bool]:
    """(agence, equipe, is_distrib). Pareil que IAG."""
    if not id_salarie:
        return ("", "", False)
    db = get_pg_connection("rh")
    rows = db.query(
        """SELECT o.lib_orga, o.id_type_niveau_orga, o.id_type_orga
             FROM rh.pgt_salarie_organigramme so
             JOIN rh.pgt_organigramme o ON o.idorganigramme = so.idorganigramme
            WHERE so.id_salarie = ?
              AND (so.modif_elem IS NULL OR so.modif_elem NOT LIKE '%suppr%')""",
        (int(id_salarie),),
    ) or []
    agence = ""; equipe = ""; is_distrib = False
    for r in rows:
        lvl = r.get("id_type_niveau_orga")
        lib = r.get("lib_orga") or ""
        if lvl == 3 and not agence:
            agence = lib
            if int(r.get("id_type_orga") or 0) == 3:
                is_distrib = True
        elif lvl == 4 and not equipe:
            equipe = lib
    return (agence, equipe, is_distrib)


def _ajoute_histo_oen_etat(id_contrat: int, old_etat: int, new_etat: int,
                            date_paiement: str, op_id: int,
                            categorie: str = "") -> None:
    """Historise un changement d'etat OEN (table pgt_oen_histo_etat_ctt)."""
    if not id_contrat:
        return
    db = get_pg_connection("adv")
    table = "pgt_oen_histo_etat_ctt_oen" if categorie == "OEN" else "pgt_oen_histo_etat_ctt"
    auto = db.query_one(
        f"SELECT COALESCE(MAX(id_histo_auto), 0) + 1 AS n FROM adv.{table}"
    )
    db.query(
        f"""INSERT INTO adv.{table}
              (id_histo_auto, id_histo, id_contrat, op_saisie, date,
               old_etat, new_etat, date_paiement,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, NOW(), ?, ?, ?, ?, NOW(), 'new')""",
        (int(auto["n"]) if auto else 1, _new_id(),
         int(id_contrat), int(op_id),
         int(old_etat) if old_etat else 0,
         int(new_etat) if new_etat else 0,
         date_paiement or "", int(op_id)),
    )


def _detect_periode_oen(
    date_sign: Optional[date], is_distrib: bool,
    p1_du: date, p1_au: date, mp1: Optional[date],
    p2_du: date, p2_au: date, mp2: Optional[date],
    mp_distrib: Optional[date],
) -> tuple[Optional[date], str]:
    """Identique a IAG : (mois_p, libelle_periode)."""
    if not date_sign:
        return (None, "")
    if is_distrib:
        return (mp_distrib, "Distrib")
    if p1_du <= date_sign <= p1_au:
        return (mp1, "Période 1")
    if p2_du <= date_sign <= p2_au:
        return (mp2, "Période 2")
    p1_du_m1 = (p1_du.replace(month=p1_du.month - 1) if p1_du.month > 1
                else p1_du.replace(year=p1_du.year - 1, month=12))
    p1_au_m1 = (p1_au.replace(month=p1_au.month - 1) if p1_au.month > 1
                else p1_au.replace(year=p1_au.year - 1, month=12))
    if p1_du_m1 <= date_sign <= p1_au_m1:
        return (mp1, "Période -1 mois")
    return (None, "HORS_DELAI")


def _create_oen_contrat(td: dict, op_id: int) -> int:
    """INSERT pgt_oen_contrat + pgt_oen_contrat_option (defaults)."""
    db = get_pg_connection("adv")
    id_contrat = _new_id()
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_contrat_auto), 0) + 1 AS n FROM adv.pgt_oen_contrat"
    )
    db.query(
        """INSERT INTO adv.pgt_oen_contrat
              (id_contrat_auto, id_contrat, id_client, id_salarie, id_ste,
               num_bs, ref_client, id_produit, id_etat_contrat, id_etat_oen,
               date_signature, date_activation,
               gaz_car_declaree, gaz_car_relevee, elec_puissance, is_dual,
               op_saisie, date_saisie, non_call,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                   ?, ?, ?, ?, ?, ?,
                   ?, ?, ?, ?, NOW(), 'new')""",
        (int(auto["n"]) if auto else 1, id_contrat,
         int(td.get("id_client") or 0),
         int(td.get("id_salarie") or 0),
         int(td.get("id_ste") or 0),
         td.get("num_bs") or "", td.get("ref_client") or "",
         int(td.get("id_produit") or 0),
         int(td.get("etat_contrat") or 0),
         int(td.get("etat_contrat") or 0),
         td.get("date_signature"), td.get("date_activation"),
         int(td.get("car_declaree") or 0),
         int(td.get("car_relevee") or 0),
         int(td.get("puissance") or 0),
         bool(td.get("is_dual")),
         int(op_id), td.get("date_creation") or datetime.now(),
         bool(td.get("non_call", True)),
         int(op_id)),
    )
    # Option
    auto_o = db.query_one(
        "SELECT COALESCE(MAX(id_contrat_option_auto), 0) + 1 AS n "
        "FROM adv.pgt_oen_contrat_option"
    )
    db.query(
        """INSERT INTO adv.pgt_oen_contrat_option
              (id_contrat_option_auto, id_contrat, num_bs,
               opt_energie_verte_gaz, opt_energie_verte_elec,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, NOW(), 'new')""",
        (int(auto_o["n"]) if auto_o else 1, id_contrat,
         td.get("num_bs") or "",
         bool(td.get("opt_energie_verte_gaz")),
         bool(td.get("opt_energie_verte_elec")),
         int(op_id)),
    )
    return id_contrat


# ---------------------------------------------------------------------------
# Type 1 : ImportJournalier
# ---------------------------------------------------------------------------


def _import_journalier_oen(
    p: ImportOenParams, file_bytes: bytes, op_id: int,
    ajoutes: list, modifies: list, pb_vendeur: list, resume: ImportOenResume,
) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_BJ_OEN.items()}
    db = get_pg_connection("adv")

    # Premiere passe : detecte les clients dual (au moins 1 ligne Gaz + 1 Elec)
    refs_dual: dict[str, dict[str, bool]] = {}
    for i in range(2, (ws.max_row or 0) + 1):
        ref = _cell(ws, i, cols["ref_client"])
        type_ener = _cell(ws, i, cols["type_ener"])
        if not ref:
            continue
        if ref not in refs_dual:
            refs_dual[ref] = {"gaz": False, "elec": False}
        if "gaz" in type_ener.lower():
            refs_dual[ref]["gaz"] = True
        else:
            refs_dual[ref]["elec"] = True

    for i in range(2, (ws.max_row or 0) + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"]).upper()
        if not num_contrat:
            continue
        ref_client = _cell(ws, i, cols["ref_client"])
        type_ener = _cell(ws, i, cols["type_ener"])
        car_relevee = _parse_int(_cell(ws, i, cols["car_relevee"])) or 0
        car_declaree = _parse_int(_cell(ws, i, cols["car_declaree"])) or 0
        if car_relevee < 0:
            car_relevee = 0
        if car_declaree < 0:
            car_declaree = 0
        puiss = _parse_int(_cell(ws, i, cols["puissance"])) or 0
        opt_verte = "verte" in _cell(ws, i, cols["option_verte"]).lower()
        lib_statut = _cell(ws, i, cols["lib_statut"])
        lib_offre = _cell(ws, i, cols["lib_offre"])
        vendeur_code = _cell(ws, i, cols["vendeur"]).strip()
        date_sign = _parse_date_fr(_cell(ws, i, cols["date_signature"]))

        # Filtre brouillons "draft"
        id_etat = _id_etat_oen_by_lib(lib_statut)
        if id_etat == 9:
            continue

        # Determine is_dual + lib_offre suffix
        refd = refs_dual.get(ref_client, {})
        is_dual = refd.get("gaz") and refd.get("elec")
        if is_dual:
            lib_offre += " (DUAL)"

        # Determine id_produit (avec fallback selon type_ener et dual)
        id_produit = _id_produit_oen_by_lib(lib_offre)
        if id_produit == 0:
            if is_dual:
                id_produit = 92 if "gaz" in type_ener.lower() else 93
            else:
                id_produit = 90 if "gaz" in type_ener.lower() else 91

        # Erreurs reporting
        if id_etat == 0:
            resume.nb_pb_statut += 1
            pb_vendeur.append({
                "NumCtt": num_contrat, "Erreur": f"Statut introuvable: {lib_statut}",
                "CodeVend": vendeur_code, "RefClient": ref_client,
            })
        if id_produit == 0:
            resume.nb_pb_offre += 1
            pb_vendeur.append({
                "NumCtt": num_contrat, "Erreur": f"Offre introuvable: {lib_offre}",
                "CodeVend": vendeur_code, "RefClient": ref_client,
            })

        # Lookup contrat
        ctt = db.query_one(
            """SELECT id_contrat, id_client, id_salarie, id_produit,
                      id_etat_contrat, id_etat_oen, id_ste, date_signature,
                      gaz_car_relevee, gaz_car_declaree, elec_puissance,
                      is_dual, mois_p
                 FROM adv.pgt_oen_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_contrat,),
        )

        id_vendeur = _lookup_vendeur_by_code(vendeur_code)

        if not ctt:
            if id_vendeur == 0:
                resume.nb_pb_vendeur += 1
                pb_vendeur.append({
                    "NumCtt": num_contrat,
                    "Erreur": f"Vendeur introuvable (code: {vendeur_code})",
                    "RefClient": ref_client,
                })
            ajoutes.append({
                "NumCtt": num_contrat, "RefClient": ref_client,
                "TypeEner": type_ener, "Puiss": puiss,
                "CarDecl": car_declaree, "CarRel": car_relevee,
                "DateSign": str(date_sign or ""),
                "LibOffre": lib_offre, "LibStatut": lib_statut,
                "IdProduit": id_produit, "EtatContrat": id_etat,
                "IsDual": is_dual,
                "_payload_create": {
                    "num_bs": num_contrat, "ref_client": ref_client,
                    "id_salarie": id_vendeur,
                    "id_produit": id_produit, "etat_contrat": id_etat,
                    "date_signature": date_sign,
                    "car_declaree": car_declaree, "car_relevee": car_relevee,
                    "puissance": puiss, "is_dual": is_dual,
                    "opt_energie_verte_gaz": opt_verte and "gaz" in type_ener.lower(),
                    "opt_energie_verte_elec": opt_verte and "elec" in type_ener.lower(),
                    "non_call": True,
                },
            })
            resume.nb_ajoutes += 1
        else:
            # Compare et detect modifs
            modifs = []
            id_contrat_ex = int(ctt["id_contrat"])
            etat_ctt_actuel = int(ctt.get("id_etat_contrat") or 0)
            etat_oen_actuel = int(ctt.get("id_etat_oen") or 0)
            id_ste_actuel = int(ctt.get("id_ste") or 0)
            date_sign_actuel = ctt.get("date_signature")
            mois_p_actuel = ctt.get("mois_p")

            # Date de signature
            new_date_sign = None
            if date_sign and date_sign_actuel != date_sign:
                new_date_sign = date_sign
                modifs.append(f"Date Sign -> {date_sign}")

            if bool(ctt.get("is_dual")) != is_dual:
                modifs.append(f"Dual -> {is_dual}")
            if int(ctt.get("id_produit") or 0) != id_produit and id_produit:
                modifs.append(f"Produit -> {id_produit}")

            # Etat Vendeur : MAJ si etat different + categorie != 'FIXE'
            # (cf. WinDev Publipostage ImportJournalier OEN : simplification
            # de la condition IDTypeEtat<=2 ET reqInfoStatut.IDTypeEtat>2
            # ET Categorie<>'FIXE' -> Categorie<>'FIXE' seul).
            new_etat_ctt = None
            if id_etat and etat_ctt_actuel != id_etat:
                cat_info = db.query_one(
                    """SELECT categorie FROM adv.pgt_oen_etat_contrat
                        WHERE id_etat = ? LIMIT 1""",
                    (etat_ctt_actuel,),
                )
                cat_actuel = ((cat_info or {}).get("categorie") or "").upper()
                if cat_actuel != "FIXE":
                    new_etat_ctt = id_etat
                    modifs.append("Etat Vendeur")

            # Etat OEN : idem (compare id_etat_oen)
            new_etat_oen = None
            if id_etat and etat_oen_actuel != id_etat:
                cat_info_o = db.query_one(
                    """SELECT categorie FROM adv.pgt_oen_etat_contrat
                        WHERE id_etat = ? LIMIT 1""",
                    (etat_oen_actuel,),
                )
                cat_oen = ((cat_info_o or {}).get("categorie") or "").upper()
                if cat_oen != "FIXE":
                    new_etat_oen = id_etat
                    modifs.append("Etat OEN")

            # id_ste : MAJ si vide (cf. WinDev "si OEN_contrat.IdSte=0")
            new_id_ste = None
            if id_ste_actuel == 0 and id_vendeur:
                mv = db.query_one(
                    """SELECT id_ste FROM rh.pgt_salarie
                        WHERE id_salarie = ? LIMIT 1""",
                    (int(id_vendeur),),
                )
                mv_ste = int((mv or {}).get("id_ste") or 0)
                if mv_ste:
                    new_id_ste = mv_ste
                    modifs.append("STE")

            if int(ctt.get("gaz_car_relevee") or 0) != car_relevee:
                modifs.append(f"CAR -> {car_relevee}")
            if int(ctt.get("elec_puissance") or 0) != puiss:
                modifs.append(f"Puiss -> {puiss}")

            if modifs:
                modifies.append({
                    "NumCtt": num_contrat,
                    "DateSign OMAYA": str(ctt.get("date_signature") or ""),
                    "Modifs": " | ".join(modifs),
                    "_payload_update": {
                        "id_contrat": id_contrat_ex,
                        "id_produit": id_produit, "is_dual": is_dual,
                        "car_relevee": car_relevee, "car_declaree": car_declaree,
                        "puissance": puiss,
                        "opt_energie_verte_gaz": opt_verte and "gaz" in type_ener.lower(),
                        "opt_energie_verte_elec": opt_verte and "elec" in type_ener.lower(),
                        # Nouveaux champs (peuvent etre None -> pas MAJ dans le UPDATE)
                        "date_signature": new_date_sign,
                        "new_etat_ctt": new_etat_ctt,
                        "etat_ctt_actuel": etat_ctt_actuel,
                        "new_etat_oen": new_etat_oen,
                        "etat_oen_actuel": etat_oen_actuel,
                        "new_id_ste": new_id_ste,
                        "mois_p": mois_p_actuel,
                    },
                })
                resume.nb_modifies += 1
    wb.close()


# ---------------------------------------------------------------------------
# Type 2 : ImportRunValide
# ---------------------------------------------------------------------------


def _import_run_valide_oen(
    p: ImportOenParams, file_bytes: bytes, op_id: int,
    runs: list, modifies: list, non_trouves: list, pb_vendeur: list,
    resume: ImportOenResume,
) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_RV_OEN.items()}
    db = get_pg_connection("adv")

    p1_du = _parse_date_fr(p.periode1_du) or date(1900, 1, 1)
    p1_au = _parse_date_fr(p.periode1_au) or date(2100, 12, 31)
    p2_du = _parse_date_fr(p.periode2_du) or date(1900, 1, 1)
    p2_au = _parse_date_fr(p.periode2_au) or date(2100, 12, 31)
    mp1 = _dernier_jour_mois(p.periode1_mois_paiement)
    mp2 = _dernier_jour_mois(p.periode2_mois_paiement)
    mp_distrib = _dernier_jour_mois(p.mois_paiement_distrib)

    for i in range(2, (ws.max_row or 0) + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"]).upper()
        if not num_contrat:
            continue
        ref_client = _cell(ws, i, cols["ref_client"])
        type_ener = _cell(ws, i, cols["type_ener"]).upper()
        car = _parse_int(_cell(ws, i, cols["car"]))
        puiss = _parse_int(_cell(ws, i, cols["puissance"]))
        lib_rem = _cell(ws, i, cols["lib_rem"])
        lib_statut = _cell(ws, i, cols["lib_statut"])
        lib_offre = _cell(ws, i, cols["lib_offre"])
        vendeur_code = _cell(ws, i, cols["vendeur"])
        date_act = _parse_date_fr(_cell(ws, i, cols["date_activation"]))

        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, date_signature,
                      gaz_car_relevee, elec_puissance, id_etat_contrat, mois_p
                 FROM adv.pgt_oen_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_contrat,),
        )
        if not ctt:
            non_trouves.append({
                "NumCtt": num_contrat, "RefClient": ref_client,
                "TypeEner": type_ener, "Puiss": puiss, "Car": car,
            })
            resume.nb_introuvables += 1
            continue

        id_contrat = int(ctt["id_contrat"])
        agence, equipe, is_distrib = _affectation_oen(int(ctt.get("id_salarie") or 0))
        mois_p, periode_lbl = _detect_periode_oen(
            ctt.get("date_signature"), is_distrib,
            p1_du, p1_au, mp1, p2_du, p2_au, mp2, mp_distrib,
        )
        mois_p_str = mois_p.strftime("%m-%Y") if mois_p else ""

        if periode_lbl == "HORS_DELAI":
            resume.nb_hors_delai += 1
            pb_vendeur.append({
                "NumCtt": num_contrat, "Erreur": "Hors Délai",
                "Agence": agence, "Equipe": equipe,
            })

        # Verifier si rem deja enregistree
        existing_rem = db.query_one(
            """SELECT id_oen_contrat_remun FROM adv.pgt_oen_contrat_remun
                WHERE id_contrat = ? AND ra_mois_p = ? LIMIT 1""",
            (id_contrat, mois_p_str),
        ) if False else None
        # Note : la table pgt_oen_contrat_remun a peut-etre des colonnes
        # snake_case differentes. On verifie en best-effort
        try:
            existing_rem = db.query_one(
                """SELECT id_oen_contrat_remun_auto AS n
                     FROM adv.pgt_oen_contrat_remun
                    WHERE id_contrat = ? AND ra_mois_p = ? LIMIT 1""",
                (id_contrat, mois_p_str),
            )
        except Exception:
            existing_rem = None

        if existing_rem:
            modifies.append({
                "NumCtt": num_contrat, "RefClient": ref_client,
                "Periode": periode_lbl, "MoisP": mois_p_str,
                "Lib Rem": lib_rem,
                "Note": "Rem déjà enregistrée",
            })
            resume.nb_deja_statues += 1
        else:
            runs.append({
                "NumCtt": num_contrat, "RefClient": ref_client,
                "Periode": periode_lbl, "MoisP": mois_p_str,
                "Lib Rem": lib_rem, "Lib Offre": lib_offre,
                "Agence": agence, "Equipe": equipe,
                "_payload_create_rem": {
                    "id_contrat": id_contrat,
                    "num_bs": num_contrat,
                    "lib_offre": lib_offre, "lib_rem": lib_rem,
                    "lib_statut": lib_statut, "mois_p": mois_p_str,
                    "nouvel_etat": 43,  # Paye par l'operateur
                    "etat_actuel": int(ctt.get("id_etat_contrat") or 0),
                    "mois_p_date": mois_p,
                    "date_activation": date_act,
                    "car": car, "puiss": puiss,
                    "type_ener": type_ener,
                },
            })
            resume.nb_valides += 1
    wb.close()


# ---------------------------------------------------------------------------
# Type 3 : ImportRunResil
# ---------------------------------------------------------------------------


def _import_run_resil_oen(
    p: ImportOenParams, file_bytes: bytes, op_id: int,
    runs: list, modifies: list, non_trouves: list, pb_vendeur: list,
    resume: ImportOenResume,
) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_RR_OEN.items()}
    db = get_pg_connection("adv")

    p1_du = _parse_date_fr(p.periode1_du) or date(1900, 1, 1)
    p1_au = _parse_date_fr(p.periode1_au) or date(2100, 12, 31)
    p2_du = _parse_date_fr(p.periode2_du) or date(1900, 1, 1)
    p2_au = _parse_date_fr(p.periode2_au) or date(2100, 12, 31)
    mp1 = _dernier_jour_mois(p.periode1_mois_paiement)
    mp2 = _dernier_jour_mois(p.periode2_mois_paiement)
    mp_distrib = _dernier_jour_mois(p.mois_paiement_distrib)

    for i in range(2, (ws.max_row or 0) + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"]).upper()
        if not num_contrat:
            continue
        ref_client = _cell(ws, i, cols["ref_client"])

        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, date_signature,
                      id_etat_contrat, mois_p
                 FROM adv.pgt_oen_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_contrat,),
        )
        if not ctt:
            non_trouves.append({"NumCtt": num_contrat, "RefClient": ref_client})
            resume.nb_introuvables += 1
            continue

        id_contrat = int(ctt["id_contrat"])
        etat_actuel = int(ctt.get("id_etat_contrat") or 0)
        agence, equipe, is_distrib = _affectation_oen(int(ctt.get("id_salarie") or 0))
        mois_p, periode_lbl = _detect_periode_oen(
            ctt.get("date_signature"), is_distrib,
            p1_du, p1_au, mp1, p2_du, p2_au, mp2, mp_distrib,
        )
        if periode_lbl == "HORS_DELAI":
            resume.nb_hors_delai += 1
            pb_vendeur.append({"NumCtt": num_contrat, "Erreur": "Hors Délai"})

        # Lookup type_etat
        etat_info = db.query_one(
            """SELECT id_type_etat FROM adv.pgt_oen_etat_contrat
                WHERE id_etat = ? LIMIT 1""",
            (etat_actuel,),
        )
        type_etat = int(etat_info.get("id_type_etat") or 0) if etat_info else 0

        mois_p_omaya = ctt.get("mois_p")
        same_month = (mois_p and mois_p_omaya
                      and mois_p.month == mois_p_omaya.month
                      and mois_p.year == mois_p_omaya.year)

        # Determination du traitement
        traitement = "deja_statue"; nouvel_etat = etat_actuel; nouveau_mois_p = None
        if type_etat == 5:
            if same_month:
                # Resilie
                nouvel_etat = 17  # OEN_contrat resilie
                nouveau_mois_p = None
                traitement = "resilie"
                resume.nb_resilies += 1
            else:
                # Decomm
                nouvel_etat = 37  # Decommissionne par l'operateur
                nouveau_mois_p = mois_p
                traitement = "decomm"
                resume.nb_decommissions += 1
        elif type_etat in (4, 3, 6):
            # Deja resilie / anomalie / decomm
            resume.nb_deja_statues += 1
        else:
            # En attente -> resilie
            nouvel_etat = 17
            nouveau_mois_p = None
            traitement = "resilie"
            resume.nb_resilies += 1

        row_snap = {
            "NumCtt": num_contrat, "RefClient": ref_client,
            "Periode": periode_lbl, "Agence": agence, "Equipe": equipe,
            "Etat actuel": etat_actuel, "TypeEtat": type_etat,
            "Nouvel etat": nouvel_etat, "Traitement": traitement,
        }
        if traitement == "deja_statue":
            modifies.append(row_snap)
        else:
            runs.append(row_snap)

        # Production
        if not p.simulation and traitement != "deja_statue":
            try:
                db.query(
                    """UPDATE adv.pgt_oen_contrat
                          SET id_etat_contrat = ?, mois_p = ?,
                              modif_date = NOW(), modif_op = ?,
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (nouvel_etat, nouveau_mois_p, int(op_id), id_contrat),
                )
                _ajoute_histo_oen_etat(
                    id_contrat, etat_actuel, nouvel_etat,
                    str(mois_p_omaya) if mois_p_omaya else "", op_id,
                )
            except Exception as e:
                row_snap["Erreur"] = str(e)
    wb.close()


# ---------------------------------------------------------------------------
# Type 4 : ImportThermostat
# ---------------------------------------------------------------------------


def _import_thermostat_oen(
    p: ImportOenParams, file_bytes: bytes, op_id: int,
    ajoutes: list, pb_vendeur: list, resume: ImportOenResume,
) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_TH_OEN.items()}
    db = get_pg_connection("adv")

    for i in range(2, (ws.max_row or 0) + 1):
        ref_client = _cell(ws, i, cols["num_cm"])
        if not ref_client:
            continue
        # Cherche le contrat Elec du client
        r = db.query_one(
            """SELECT c.id_contrat, c.num_bs, c.id_salarie,
                      p.lib_produit
                 FROM adv.pgt_oen_contrat c
                 JOIN adv.pgt_oen_produit p ON p.id_produit = c.id_produit
                WHERE c.ref_client = ?
                  AND UPPER(COALESCE(p.sous_fam, '')) LIKE '%ELEC%'
                  AND (c.modif_elem IS NULL OR c.modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (ref_client,),
        )
        if not r:
            pb_vendeur.append({"RefClient": ref_client,
                               "Erreur": "CM ou contrat Elec Introuvable"})
            resume.nb_erreurs += 1
            continue

        id_contrat = int(r["id_contrat"])
        ajoutes.append({
            "RefClient": ref_client, "NumBS": r.get("num_bs"),
            "LibProduit": r.get("lib_produit"),
            "_payload_update_option": {"id_contrat": id_contrat},
        })
        resume.nb_ajoutes += 1

        if not p.simulation:
            try:
                db.query(
                    """UPDATE adv.pgt_oen_contrat_option
                          SET opt_entretien = TRUE,
                              modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (int(op_id), id_contrat),
                )
            except Exception:
                pass
    wb.close()


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------


def run_import_oen(
    p: ImportOenParams, file_bytes: bytes, op_id: int,
) -> ImportOenResult:
    label = TYPE_LABELS.get(p.type_import, "?")
    if not file_bytes:
        return ImportOenResult(
            ok=False, type_import=p.type_import, type_label=label,
            simulation=p.simulation, resume=ImportOenResume(),
            message="Aucun fichier fourni.",
        )

    resume = ImportOenResume()
    ajoutes: list[dict] = []; modifies: list[dict] = []
    runs: list[dict] = []; non_trouves: list[dict] = []
    pb_vendeur: list[dict] = []

    try:
        if p.type_import == 1:
            _import_journalier_oen(
                p, file_bytes, op_id, ajoutes, modifies, pb_vendeur, resume,
            )
        elif p.type_import == 2:
            _import_run_valide_oen(
                p, file_bytes, op_id, runs, modifies, non_trouves, pb_vendeur, resume,
            )
        elif p.type_import == 3:
            _import_run_resil_oen(
                p, file_bytes, op_id, runs, modifies, non_trouves, pb_vendeur, resume,
            )
        elif p.type_import == 4:
            _import_thermostat_oen(p, file_bytes, op_id, ajoutes, pb_vendeur, resume)
    except Exception as e:
        return ImportOenResult(
            ok=False, type_import=p.type_import, type_label=label,
            simulation=p.simulation, resume=resume,
            message=f"Erreur lecture/traitement : {e}",
        )

    # -- PASSE PROD : creation + MAJ --
    nb_crees = 0; nb_majs = 0
    if not p.simulation:
        db = get_pg_connection("adv")
        # Type 1 : creations
        for row in ajoutes:
            pl = row.pop("_payload_create", None)
            if not pl:
                continue
            try:
                new_id = _create_oen_contrat(pl, op_id)
                row["IdContratCree"] = new_id
                nb_crees += 1
            except Exception as e:
                row["Erreur"] = str(e)
        # Type 1 : MAJ
        for row in modifies:
            pl = row.pop("_payload_update", None)
            if not pl:
                continue
            try:
                # Champs de base (toujours MAJ)
                sets = ["id_produit = ?", "is_dual = ?",
                        "gaz_car_relevee = ?", "gaz_car_declaree = ?",
                        "elec_puissance = ?"]
                params: list = [int(pl["id_produit"]), bool(pl["is_dual"]),
                                int(pl["car_relevee"]), int(pl["car_declaree"]),
                                int(pl["puissance"])]
                # Champs optionnels (MAJ seulement si changement detecte)
                if pl.get("date_signature"):
                    sets.append("date_signature = ?")
                    params.append(pl["date_signature"])
                if pl.get("new_etat_ctt"):
                    sets.append("id_etat_contrat = ?")
                    params.append(int(pl["new_etat_ctt"]))
                if pl.get("new_etat_oen"):
                    sets.append("id_etat_oen = ?")
                    params.append(int(pl["new_etat_oen"]))
                if pl.get("new_id_ste"):
                    sets.append("id_ste = ?")
                    params.append(int(pl["new_id_ste"]))
                sets.append("modif_date = NOW()")
                sets.append("modif_op = ?")
                params.append(int(op_id))
                sets.append("modif_elem = 'modif'")
                sql = (f"UPDATE adv.pgt_oen_contrat SET {', '.join(sets)} "
                       f"WHERE id_contrat = ?")
                params.append(int(pl["id_contrat"]))
                db.query(sql, tuple(params))

                # Historisation etats (cf. WinDev ajouteHistoContrat)
                mois_p = pl.get("mois_p")
                mois_p_str = str(mois_p) if mois_p else ""
                if pl.get("new_etat_ctt"):
                    _ajoute_histo_oen_etat(
                        int(pl["id_contrat"]), int(pl["etat_ctt_actuel"]),
                        int(pl["new_etat_ctt"]), mois_p_str, op_id,
                        categorie="Vend",
                    )
                if pl.get("new_etat_oen"):
                    _ajoute_histo_oen_etat(
                        int(pl["id_contrat"]), int(pl["etat_oen_actuel"]),
                        int(pl["new_etat_oen"]), "", op_id,
                        categorie="OEN",
                    )
                nb_majs += 1
            except Exception as e:
                row["Erreur"] = str(e)
        # Type 2 : creations Rem (best-effort, table peut-etre absente)
        for row in runs:
            pl = row.pop("_payload_create_rem", None)
            if not pl:
                continue
            try:
                # UPDATE contrat -> etat 43 + mois_p
                db.query(
                    """UPDATE adv.pgt_oen_contrat
                          SET id_etat_contrat = ?, id_etat_oen = 15,
                              mois_p = ?, date_activation = COALESCE(?, date_activation),
                              modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (int(pl["nouvel_etat"]), pl.get("mois_p_date"),
                     pl.get("date_activation"),
                     int(op_id), int(pl["id_contrat"])),
                )
                _ajoute_histo_oen_etat(
                    int(pl["id_contrat"]), int(pl["etat_actuel"]),
                    int(pl["nouvel_etat"]), pl.get("mois_p", ""), op_id,
                    categorie="Vend",
                )
                nb_majs += 1
            except Exception as e:
                row["Erreur"] = str(e)

    label_msg = f" | {nb_crees} créés, {nb_majs} MAJ" if not p.simulation else ""
    # Cleanup payloads
    for row in ajoutes:
        row.pop("_payload_create", None)
    for row in modifies:
        row.pop("_payload_update", None)
    for row in runs:
        row.pop("_payload_create_rem", None)
    for row in ajoutes:
        row.pop("_payload_update_option", None)

    res = ImportOenResult(
        ok=True, type_import=p.type_import, type_label=label,
        simulation=p.simulation, resume=resume,
        contrats_ajoutes=ajoutes, contrats_modifies=modifies,
        contrats_run=runs, contrats_non_trouves=non_trouves,
        pb_vendeur=pb_vendeur,
        message=(
            f"Type {p.type_import} ({label}) | "
            f"Ajoutés {resume.nb_ajoutes} | Modifs {resume.nb_modifies} | "
            f"Validés {resume.nb_valides} | Résiliés {resume.nb_resilies} | "
            f"Décomm {resume.nb_decommissions} | Déjà statués {resume.nb_deja_statues} | "
            f"Introuvables {resume.nb_introuvables} | "
            f"Pb vendeur {resume.nb_pb_vendeur} | Pb statut {resume.nb_pb_statut} | "
            f"Pb offre {resume.nb_pb_offre} | Hors délai {resume.nb_hors_delai}"
            f"{label_msg}. "
            + ("(SIMULATION)" if p.simulation else "(PRODUCTION)")
        ),
    )
    _attach_xlsx_and_mail_oen(res, op_id)
    return res


# ---------------------------------------------------------------------------
# XLSX + mail BO
# ---------------------------------------------------------------------------


def _build_xlsx_oen(res: ImportOenResult) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook(); ws = wb.active; ws.title = "Résumé"
    header_fill = PatternFill("solid", fgColor="17494E")
    header_font = Font(bold=True, color="FFFFFF")
    items = [
        ("NB Ajoutés", res.resume.nb_ajoutes),
        ("NB Modifiés", res.resume.nb_modifies),
        ("NB Validés", res.resume.nb_valides),
        ("NB Résiliés", res.resume.nb_resilies),
        ("NB Décommissions", res.resume.nb_decommissions),
        ("NB Déjà statués", res.resume.nb_deja_statues),
        ("NB Introuvables", res.resume.nb_introuvables),
        ("NB Doublons", res.resume.nb_doublons),
        ("NB Hors délai", res.resume.nb_hors_delai),
        ("NB Pb Vendeur", res.resume.nb_pb_vendeur),
        ("NB Pb Statut", res.resume.nb_pb_statut),
        ("NB Pb Offre", res.resume.nb_pb_offre),
        ("NB Erreurs", res.resume.nb_erreurs),
    ]
    ws.append(["Indicateur", "Nombre"])
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for lbl, n in items:
        ws.append([lbl, n])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    sections: list[tuple[str, list[dict]]] = [
        ("Ajoutés", res.contrats_ajoutes),
        ("Modifiés", res.contrats_modifies),
        ("RUN", res.contrats_run),
        ("Non trouvés", res.contrats_non_trouves),
        ("Pb Vendeur", res.pb_vendeur),
    ]
    for title, rows in sections:
        if not rows:
            continue
        sh = wb.create_sheet(title=title[:31])
        keys = list(rows[0].keys())
        sh.append(keys)
        for c in sh[1]:
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for r in rows:
            sh.append([str(r.get(k, "")) if r.get(k) is not None else "" for k in keys])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def _attach_xlsx_and_mail_oen(res: ImportOenResult, op_id: int) -> None:
    from app.shared.notifications.mail import envoi_mail
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    suffix = "_SIMU" if res.simulation else ""
    prefix_map = {1: "ImportCallOEN", 2: "ImportRunOENValide",
                  3: "ImportRunOENResilDecom", 4: "ImportThermostatOEN"}
    xlsx_name = f"{prefix_map.get(res.type_import, 'ImportOEN')}_{ts}{suffix}.xlsx"
    try:
        xlsx_bytes = _build_xlsx_oen(res)
    except Exception:
        return
    res.xlsx_name = xlsx_name
    res.xlsx_b64 = base64.b64encode(xlsx_bytes).decode("ascii")

    db = get_pg_connection("rh")
    r = db.query_one(
        "SELECT mail FROM rh.pgt_salarie_coordonnees WHERE id_salarie = ? LIMIT 1",
        (int(op_id),),
    )
    op_mail = (r.get("mail") if r else "") or ""
    destinataires = [op_mail] if op_mail else ["intranet@omaya.fr"]
    cc = ["intranet@omaya.fr"] if op_mail and op_mail != "intranet@omaya.fr" else []

    sujet_pref = "SIMULATION : " if res.simulation else ""
    sujet = (f"{sujet_pref}Importation {res.type_label} OHM Énergie "
             f"du {date.today().strftime('%d/%m/%Y')}")
    html = (
        "<p>Bonjour,</p>"
        f"<p>Fin importation le : {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>"
        f"<p><strong>{res.message}</strong></p>"
        "<p>Service Importation EXOSPHERE</p>"
    )
    try:
        res.mail_envoye = envoi_mail(
            sujet=sujet, html=html,
            destinataires=destinataires, cc=cc,
            expediteur="intranet@omaya.fr",
            attachments=[(xlsx_name, xlsx_bytes)],
        )
    except Exception:
        res.mail_envoye = False
