"""Service Fen_ImportPRO (ADM Imports Bases -> PROTECTED).

Pattern identique a IAG : 2 types (Base Journaliere + RUN avec
dispatch valide/resil selon 'ANNUL' dans le nom).

PRO = assurance / protection des appareils, structure de contrat
simple (pas de gaz/elec/CAR/puissance/is_dual).
"""

from __future__ import annotations

import base64
import io
from datetime import date, datetime
from typing import Optional

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection
from app.shared.recrutement.services.recherche_cv import _int, _str
from app.intranets.adm.services.import_eni import (
    _new_id, _lookup_or_create_client,
    _col_letter_to_index, _cell, _parse_date_fr, _dernier_jour_mois,
)


COLS_BJ_PRO = {
    "num_contrat": "A",        "date_signature": "C",
    "vendeur_prenom": "E",     "vendeur_nom": "F",
    "signe": "G",              "pack": "J",
}

COLS_RUN_PRO = {
    "num_contrat": "A",        "date_signature": "B",
    "client_nom": "E",         "client_prenom": "F",
    "vendeur": "K",            "commentaire": "N",
}


class ImportProParams(BaseModel):
    type_import: int                   # 1 = Base Journ, 2 = RUN
    simulation: bool = True
    format_vendeur: str = "prenom_nom"
    periode1_du: str = ""
    periode1_au: str = ""
    periode1_mois_paiement: str = ""
    periode2_du: str = ""
    periode2_au: str = ""
    periode2_mois_paiement: str = ""
    mois_paiement_distrib: str = ""


class ImportProResume(BaseModel):
    nb_fichiers: int = 0
    nb_ajoutes: int = 0
    nb_valides: int = 0
    nb_resilies: int = 0
    nb_deja_saisis: int = 0
    nb_deja_statues: int = 0
    nb_introuvables: int = 0
    nb_doublons: int = 0
    nb_pb_vendeur: int = 0
    nb_erreurs: int = 0


class ImportProResult(BaseModel):
    ok: bool
    type_import: int
    type_label: str
    simulation: bool
    resume: ImportProResume
    fichiers_traites: list[str] = []
    contrats_ajoutes: list[dict] = []
    contrats_modifies: list[dict] = []
    contrats_non_trouves: list[dict] = []
    contrats_run: list[dict] = []
    pb_vendeur: list[dict] = []
    message: str = ""
    xlsx_b64: str = ""
    xlsx_name: str = ""
    mail_envoye: bool = False


TYPE_LABELS = {1: "Base Journalière", 2: "RUN"}


def _id_produit_pro_by_lib(lib_offre: str) -> tuple[int, str]:
    """Determine (id_produit, lib_produit) selon le libelle Offre PRO.

    Mapping fidele WinDev (selon Vrai dans importJournalier) :
      - Basic -> 100
      - Safe  -> 102
      - autre -> 101
    """
    s = (lib_offre or "").upper()
    if "BASIC" in s:
        id_p = 100
    elif "SAFE" in s:
        id_p = 102
    else:
        id_p = 101
    db = get_pg_connection("adv")
    r = db.query_one(
        "SELECT lib_produit FROM adv.pgt_pro_produit WHERE id_produit = ? LIMIT 1",
        (id_p,),
    )
    return (id_p, _str(r.get("lib_produit")) if r else lib_offre)


def _etat_contrat_pro(signe: str, lib_statut: str = "") -> int:
    """37 par defaut (En cours). 66 si refus dans statut, 38 si doublon."""
    s = (lib_statut or "").lower()
    if "doublon" in s:
        return 38
    if "refus" in s:
        return 66
    return 37


def _lookup_vendeur_pro(nom_complet: str, format_v: str) -> int:
    """Cherche un salarie par nom/prenom selon format. Retourne id ou 0."""
    if not nom_complet or not nom_complet.strip():
        return 0
    db = get_pg_connection("rh")
    s = nom_complet.upper()
    for c in ("-", "'", " "):
        s = s.replace(c, "%")
    s = s.replace("%%", "%")
    pattern = f"%{s}%"

    if format_v == "nom_prenom":
        sql = """SELECT id_salarie FROM rh.pgt_salarie
                  WHERE UPPER(CONCAT(nom, '%', prenom)) LIKE ?
                    AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                  LIMIT 2"""
    else:
        sql = """SELECT id_salarie FROM rh.pgt_salarie
                  WHERE UPPER(CONCAT(prenom, '%', nom)) LIKE ?
                    AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                  LIMIT 2"""
    rows = db.query(sql, (pattern,)) or []
    if len(rows) == 1:
        return int(rows[0]["id_salarie"])
    return 0


def _info_salarie_pro(id_sal: int) -> dict:
    if not id_sal:
        return {}
    db = get_pg_connection("rh")
    r = db.query_one(
        """SELECT s.id_salarie, s.nom, s.prenom, e.id_ste
             FROM rh.pgt_salarie s
             LEFT JOIN rh.pgt_salarie_embauche e ON e.id_salarie = s.id_salarie
            WHERE s.id_salarie = ? LIMIT 1""",
        (int(id_sal),),
    )
    return r or {}


def _ajoute_histo_pro_etat(id_contrat: int, old_etat: int, new_etat: int,
                            date_paiement: str, op_id: int) -> None:
    if not id_contrat:
        return
    db = get_pg_connection("adv")
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_histo_auto), 0) + 1 AS n FROM adv.pgt_pro_histo_etat_ctt"
    )
    db.query(
        """INSERT INTO adv.pgt_pro_histo_etat_ctt
              (id_histo_auto, id_histo, id_contrat, op_saisie, date,
               old_etat, new_etat, date_paiement,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, NOW(), ?, ?, ?, ?, NOW(), 'new')""",
        (int(auto["n"]) if auto else 1, _new_id(),
         int(id_contrat), int(op_id),
         int(old_etat) if old_etat else 0,
         int(new_etat) if new_etat else 0,
         date_paiement or "", int(op_id)),
    )


def _detect_periode_pro(
    date_sign: Optional[date], is_distrib: bool,
    p1_du: date, p1_au: date, mp1: Optional[date],
    p2_du: date, p2_au: date, mp2: Optional[date],
    mp_distrib: Optional[date],
) -> tuple[Optional[date], str]:
    if not date_sign:
        return (None, "")
    if is_distrib:
        return (mp_distrib, "Distrib")
    if p1_du <= date_sign <= p1_au:
        return (mp1, "Période 1")
    if p2_du <= date_sign <= p2_au:
        return (mp2, "Période 2")
    p1_du_m1 = (p1_du.replace(month=p1_du.month - 1) if p1_du.month > 1
                else p1_du.replace(year=p1_du.year - 1, month=12))
    p1_au_m1 = (p1_au.replace(month=p1_au.month - 1) if p1_au.month > 1
                else p1_au.replace(year=p1_au.year - 1, month=12))
    if p1_du_m1 <= date_sign <= p1_au_m1:
        return (mp1, "Période -1 mois")
    return (None, "HORS_DELAI")


def _affectation_pro(id_salarie: int) -> tuple[str, str, bool]:
    if not id_salarie:
        return ("", "", False)
    db = get_pg_connection("rh")
    rows = db.query(
        """SELECT o.lib_orga, o.id_type_niveau_orga, o.id_type_orga
             FROM rh.pgt_salarie_organigramme so
             JOIN rh.pgt_organigramme o ON o.idorganigramme = so.idorganigramme
            WHERE so.id_salarie = ?
              AND (so.modif_elem IS NULL OR so.modif_elem NOT LIKE '%suppr%')""",
        (int(id_salarie),),
    ) or []
    agence = ""; equipe = ""; is_distrib = False
    for r in rows:
        lvl = r.get("id_type_niveau_orga")
        lib = r.get("lib_orga") or ""
        if lvl == 3 and not agence:
            agence = lib
            if int(r.get("id_type_orga") or 0) == 3:
                is_distrib = True
        elif lvl == 4 and not equipe:
            equipe = lib
    return (agence, equipe, is_distrib)


def _create_pro_contrat(td: dict, op_id: int) -> int:
    db = get_pg_connection("adv")
    id_contrat = _new_id()
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_contrat_auto), 0) + 1 AS n FROM adv.pgt_pro_contrat"
    )
    date_sign = td.get("date_signature")
    if isinstance(date_sign, str):
        date_sign = _parse_date_fr(date_sign)
    db.query(
        """INSERT INTO adv.pgt_pro_contrat
              (id_contrat_auto, id_contrat, id_client, id_salarie, id_ste,
               num_bs, id_produit, id_etat_contrat,
               date_signature, op_saisie, date_saisie, non_call,
               info_interne, code_enr,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?,
                   ?, ?, NOW(), ?, ?, '',
                   ?, NOW(), 'new')""",
        (int(auto["n"]) if auto else 1, id_contrat,
         int(td.get("id_client") or 0),
         int(td.get("id_salarie") or 0),
         int(td.get("id_ste") or 0),
         td.get("num_bs") or "",
         int(td.get("id_produit") or 0),
         int(td.get("etat_contrat") or 37),
         date_sign, int(op_id),
         bool(td.get("non_call", True)),
         td.get("commentaire") or "",
         int(op_id)),
    )
    return id_contrat


def _import_journalier_pro(
    p: ImportProParams, fname: str, content: bytes, op_id: int,
    ajoutes: list, modifies: list, pb_vendeur: list, resume: ImportProResume,
) -> None:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        modifies.append({"_erreur": f"Lecture {fname} : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_BJ_PRO.items()}
    db = get_pg_connection("adv")

    for i in range(2, (ws.max_row or 0) + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"]).upper()
        if not num_contrat:
            continue
        date_sign = _parse_date_fr(_cell(ws, i, cols["date_signature"]))
        vendeur_nom = _cell(ws, i, cols["vendeur_nom"])
        vendeur_prenom = _cell(ws, i, cols["vendeur_prenom"])
        signe = _cell(ws, i, cols["signe"])
        pack = _cell(ws, i, cols["pack"])

        vendeur_complet = f"{vendeur_prenom} {vendeur_nom}".strip()
        id_vendeur = _lookup_vendeur_pro(vendeur_complet, p.format_vendeur)
        id_produit, lib_prod = _id_produit_pro_by_lib(pack)
        etat_initial = _etat_contrat_pro(signe)

        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, num_bs, date_signature, id_produit
                 FROM adv.pgt_pro_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_contrat,),
        )
        info_sal = _info_salarie_pro(id_vendeur) if id_vendeur else {}
        id_ste = int(info_sal.get("id_ste") or 0)

        if not ctt:
            ajoutes.append({
                "NumCtt": num_contrat,
                "DateSigne": str(date_sign or ""),
                "Vendeur": vendeur_complet,
                "IdSalarie": id_vendeur,
                "Société": id_ste,
                "LibProduit": id_produit,
                "Pack": pack,
                "Signé": signe,
                "EtatContrat": etat_initial,
                "_payload_create": {
                    "num_bs": num_contrat,
                    "id_salarie": id_vendeur,
                    "id_ste": id_ste,
                    "id_produit": id_produit,
                    "etat_contrat": etat_initial,
                    "date_signature": date_sign,
                    "non_call": True,
                    "commentaire": signe,
                },
            })
            if id_vendeur == 0:
                pb_vendeur.append({
                    "NumCtt": num_contrat,
                    "DateSigne": str(date_sign or ""),
                    "Vendeur Import": vendeur_complet,
                    "Erreur": "Vendeur inconnu",
                })
                resume.nb_pb_vendeur += 1
            resume.nb_ajoutes += 1
        else:
            id_contrat = int(ctt["id_contrat"])
            id_sal_db = int(ctt.get("id_salarie") or 0)
            modifies.append({
                "NumCtt": ctt.get("num_bs"),
                "DateSigne": str(ctt.get("date_signature") or ""),
                "IdSalarie DB": id_sal_db,
                "LibProduit DB": int(ctt.get("id_produit") or 0),
                "Vendeur Import": vendeur_complet,
            })
            resume.nb_deja_saisis += 1
            if id_sal_db != id_vendeur and id_vendeur != 0:
                pb_vendeur.append({
                    "NumCtt": ctt.get("num_bs"),
                    "DateSigne": str(ctt.get("date_signature") or ""),
                    "Vendeur Import": vendeur_complet,
                    "Erreur": "vendeur réattribué",
                    "OldIdSalarie": id_sal_db,
                    "NewIdSalarie": id_vendeur,
                    "_id_contrat": id_contrat,
                })
    wb.close()


def _import_run_pro(
    p: ImportProParams, fname: str, content: bytes, op_id: int, mode: str,
    runs: list, modifies: list, non_trouves: list, pb_vendeur: list,
    resume: ImportProResume,
) -> None:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        modifies.append({"_erreur": f"Lecture {fname} : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_RUN_PRO.items()}
    db = get_pg_connection("adv")

    p1_du = _parse_date_fr(p.periode1_du) or date(1900, 1, 1)
    p1_au = _parse_date_fr(p.periode1_au) or date(2100, 12, 31)
    p2_du = _parse_date_fr(p.periode2_du) or date(1900, 1, 1)
    p2_au = _parse_date_fr(p.periode2_au) or date(2100, 12, 31)
    mp1 = _dernier_jour_mois(p.periode1_mois_paiement)
    mp2 = _dernier_jour_mois(p.periode2_mois_paiement)
    mp_distrib = _dernier_jour_mois(p.mois_paiement_distrib)

    for i in range(2, (ws.max_row or 0) + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"]).upper()
        if not num_contrat:
            continue
        date_sign_s = _cell(ws, i, cols["date_signature"])
        vendeur_cell = _cell(ws, i, cols["vendeur"])
        comment = _cell(ws, i, cols["commentaire"])

        rows_ctt = db.query(
            """SELECT id_contrat, id_salarie, num_bs, id_etat_contrat,
                      date_signature, mois_p, id_produit
                 FROM adv.pgt_pro_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
            (num_contrat,),
        ) or []
        if len(rows_ctt) == 0:
            non_trouves.append({
                "NumCtt": num_contrat, "DateSign": date_sign_s,
                "Vendeur": vendeur_cell,
                "Statut": "VALID" if mode == "valide" else "RESIL",
            })
            resume.nb_introuvables += 1
            continue
        if len(rows_ctt) > 1:
            resume.nb_doublons += 1
            for r in rows_ctt:
                pb_vendeur.append({
                    "NumCtt": num_contrat,
                    "DateSigne": str(r.get("date_signature") or ""),
                    "Erreur": f"DOUBLON - {'Valid' if mode == 'valide' else 'Resil'}",
                })
            continue

        r = rows_ctt[0]
        id_contrat = int(r["id_contrat"])
        id_sal_db = int(r.get("id_salarie") or 0)
        etat_actuel = int(r.get("id_etat_contrat") or 0)
        mois_p_omaya = r.get("mois_p")

        agence, equipe, is_distrib = _affectation_pro(id_sal_db)
        mois_p, periode_lbl = _detect_periode_pro(
            r.get("date_signature"), is_distrib,
            p1_du, p1_au, mp1, p2_du, p2_au, mp2, mp_distrib,
        )
        if periode_lbl == "HORS_DELAI":
            pb_vendeur.append({
                "NumCtt": num_contrat,
                "Erreur": "Hors Délai - " + ("Valid" if mode == "valide" else "Résil"),
                "Agence": agence, "Equipe": equipe,
            })

        etat_info = db.query_one(
            """SELECT id_type_etat, lib_etat FROM adv.pgt_pro_etat_contrat
                WHERE id_etat = ? LIMIT 1""",
            (etat_actuel,),
        )
        id_type_etat = int(etat_info.get("id_type_etat") or 0) if etat_info else 0
        lib_etat = (etat_info.get("lib_etat") or "") if etat_info else ""

        info_sal = _info_salarie_pro(id_sal_db)
        nom_vend = (f"{_str(info_sal.get('nom'))} "
                    f"{_str(info_sal.get('prenom')).title()}").strip()

        nouvel_etat = etat_actuel
        nouveau_mois_p: Optional[date] = None
        traitement = "deja_statue"

        if mode == "valide":
            # Eligible si type_etat in (1,2) ou etats specifiques
            if id_type_etat in (1, 2) or etat_actuel in (29, 30):
                nouvel_etat = 19  # Valide
                nouveau_mois_p = mois_p_omaya if etat_actuel in (29, 30) else mois_p
                traitement = "valide"
                resume.nb_valides += 1
            else:
                resume.nb_deja_statues += 1
        else:  # resil
            if id_type_etat == 5:  # Paye -> decomm
                nouvel_etat = 20
                nouveau_mois_p = mois_p
                traitement = "decomm"
                resume.nb_resilies += 1
            elif id_type_etat in (1, 2):  # En attente -> resil
                nouvel_etat = 16
                nouveau_mois_p = None
                traitement = "resilie"
                resume.nb_resilies += 1
            else:
                resume.nb_deja_statues += 1

        row_snap = {
            "NumCtt": num_contrat,
            "DateSigne": str(r.get("date_signature") or ""),
            "Vendeur": nom_vend,
            "Periode": periode_lbl,
            "Agence": agence, "Equipe": equipe,
            "Etat actuel": etat_actuel, "TypeEtat": id_type_etat,
            "Lib Etat": lib_etat,
            "Nouvel etat": nouvel_etat,
            "Nouveau MoisP": str(nouveau_mois_p) if nouveau_mois_p else "",
            "Traitement": traitement,
        }
        if traitement == "deja_statue":
            modifies.append(row_snap)
        else:
            runs.append(row_snap)

        if not p.simulation and traitement != "deja_statue":
            try:
                if mode == "valide":
                    db.query(
                        """UPDATE adv.pgt_pro_contrat
                              SET id_etat_contrat = ?, mois_p = ?,
                                  modif_date = NOW(), modif_op = ?,
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (nouvel_etat, nouveau_mois_p, int(op_id), id_contrat),
                    )
                elif mode == "resil":
                    if traitement == "decomm":
                        db.query(
                            """UPDATE adv.pgt_pro_contrat
                                  SET id_etat_contrat = ?, mois_p = ?,
                                      info_interne = COALESCE(info_interne, '') || '\n' || ?,
                                      modif_date = NOW(), modif_op = ?,
                                      modif_elem = 'modif'
                                WHERE id_contrat = ?""",
                            (nouvel_etat, nouveau_mois_p, comment,
                             int(op_id), id_contrat),
                        )
                    else:
                        db.query(
                            """UPDATE adv.pgt_pro_contrat
                                  SET id_etat_contrat = ?, mois_p = NULL,
                                      info_interne = COALESCE(info_interne, '') ||
                                                     '\nMotif Résil : ' || ?,
                                      modif_date = NOW(), modif_op = ?,
                                      modif_elem = 'modif'
                                WHERE id_contrat = ?""",
                            (nouvel_etat, comment, int(op_id), id_contrat),
                        )
                _ajoute_histo_pro_etat(
                    id_contrat, etat_actuel, nouvel_etat,
                    str(mois_p_omaya) if mois_p_omaya else "", op_id,
                )
            except Exception as e:
                row_snap["Erreur"] = str(e)
    wb.close()


def run_import_pro(
    p: ImportProParams, files: list[tuple[str, bytes]], op_id: int,
) -> ImportProResult:
    label = TYPE_LABELS.get(p.type_import, "?")
    if not files:
        return ImportProResult(
            ok=False, type_import=p.type_import, type_label=label,
            simulation=p.simulation, resume=ImportProResume(),
            message="Aucun fichier fourni.",
        )

    resume = ImportProResume(nb_fichiers=len(files))
    ajoutes: list[dict] = []; modifies: list[dict] = []
    non_trouves: list[dict] = []; runs: list[dict] = []
    pb_vendeur: list[dict] = []
    fichiers_traites: list[str] = []

    for fname, content in files:
        fichiers_traites.append(fname)
        if p.type_import == 1:
            _import_journalier_pro(
                p, fname, content, op_id,
                ajoutes, modifies, pb_vendeur, resume,
            )
        elif p.type_import == 2:
            mode = "resil" if "ANNUL" in fname.upper() else "valide"
            _import_run_pro(
                p, fname, content, op_id, mode,
                runs, modifies, non_trouves, pb_vendeur, resume,
            )

    # PASSE PROD
    nb_crees = 0; nb_reattrib = 0
    if not p.simulation:
        db = get_pg_connection("adv")
        for row in ajoutes:
            pl = row.pop("_payload_create", None)
            if not pl or not pl.get("id_salarie"):
                continue
            try:
                new_id = _create_pro_contrat(pl, op_id)
                row["IdContratCree"] = new_id
                nb_crees += 1
            except Exception as e:
                row["Erreur"] = str(e)
        for row in pb_vendeur:
            id_ct = row.pop("_id_contrat", None)
            if not id_ct or row.get("Erreur") != "vendeur réattribué":
                continue
            try:
                db.query(
                    """UPDATE adv.pgt_pro_contrat
                          SET id_salarie = ?,
                              modif_op = ?, modif_date = NOW(),
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (int(row["NewIdSalarie"]), int(op_id), int(id_ct)),
                )
                nb_reattrib += 1
            except Exception as e:
                row["ErreurMaj"] = str(e)

    for row in ajoutes:
        row.pop("_payload_create", None)
    for row in pb_vendeur:
        row.pop("_id_contrat", None)

    res = ImportProResult(
        ok=True, type_import=p.type_import, type_label=label,
        simulation=p.simulation, resume=resume,
        fichiers_traites=fichiers_traites,
        contrats_ajoutes=ajoutes, contrats_modifies=modifies,
        contrats_non_trouves=non_trouves, contrats_run=runs,
        pb_vendeur=pb_vendeur,
        message=(
            f"{len(files)} fichier(s) | Ajoutés {resume.nb_ajoutes} | "
            f"Déjà saisis {resume.nb_deja_saisis} | Validés {resume.nb_valides} | "
            f"Résiliés {resume.nb_resilies} | Pb vendeur {resume.nb_pb_vendeur}. "
            + (f"PRODUCTION : {nb_crees} créés, {nb_reattrib} reattributions."
               if not p.simulation else "(SIMULATION)")
        ),
    )
    _attach_xlsx_and_mail_pro(res, op_id)
    return res


def _build_xlsx_pro(res: ImportProResult) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook(); ws = wb.active; ws.title = "Résumé"
    header_fill = PatternFill("solid", fgColor="17494E")
    header_font = Font(bold=True, color="FFFFFF")
    items = [
        ("NB Fichiers", res.resume.nb_fichiers),
        ("NB Ajoutés", res.resume.nb_ajoutes),
        ("NB Validés", res.resume.nb_valides),
        ("NB Résiliés", res.resume.nb_resilies),
        ("NB Déjà saisis", res.resume.nb_deja_saisis),
        ("NB Déjà statués", res.resume.nb_deja_statues),
        ("NB Introuvables", res.resume.nb_introuvables),
        ("NB Doublons", res.resume.nb_doublons),
        ("NB Pb Vendeur", res.resume.nb_pb_vendeur),
        ("NB Erreurs", res.resume.nb_erreurs),
    ]
    ws.append(["Indicateur", "Nombre"])
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for lbl, n in items:
        ws.append([lbl, n])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    for title, rows in [
        ("Ajoutés", res.contrats_ajoutes),
        ("Déjà saisis", res.contrats_modifies),
        ("Non trouvés", res.contrats_non_trouves),
        ("RUN", res.contrats_run),
        ("Erreurs Vendeurs", res.pb_vendeur),
    ]:
        if not rows:
            continue
        sh = wb.create_sheet(title=title[:31])
        keys = list(rows[0].keys())
        sh.append(keys)
        for c in sh[1]:
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for r in rows:
            sh.append([str(r.get(k, "")) if r.get(k) is not None else "" for k in keys])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def _attach_xlsx_and_mail_pro(res: ImportProResult, op_id: int) -> None:
    from app.shared.notifications.mail import envoi_mail
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    suffix = "_SIMU" if res.simulation else ""
    prefix = "ImportJournalierPRO" if res.type_import == 1 else "ImportRunPRO"
    xlsx_name = f"{prefix}_{ts}{suffix}.xlsx"
    try:
        xlsx_bytes = _build_xlsx_pro(res)
    except Exception:
        return
    res.xlsx_name = xlsx_name
    res.xlsx_b64 = base64.b64encode(xlsx_bytes).decode("ascii")

    db = get_pg_connection("rh")
    r = db.query_one(
        "SELECT mail FROM rh.pgt_salarie_coordonnees WHERE id_salarie = ? LIMIT 1",
        (int(op_id),),
    )
    op_mail = (r.get("mail") if r else "") or ""
    destinataires = [op_mail] if op_mail else ["intranet@omaya.fr"]
    cc = ["intranet@omaya.fr"] if op_mail and op_mail != "intranet@omaya.fr" else []

    sujet_pref = "SIMULATION : " if res.simulation else ""
    sujet = (f"{sujet_pref}Importation {res.type_label} PROTECTED "
             f"du {date.today().strftime('%d/%m/%Y')}")
    html = (
        "<p>Bonjour,</p>"
        f"<p>Fin importation le : {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>"
        f"<p><strong>{res.message}</strong></p>"
        "<p>Service Importation EXOSPHERE</p>"
    )
    try:
        res.mail_envoye = envoi_mail(
            sujet=sujet, html=html,
            destinataires=destinataires, cc=cc,
            expediteur="intranet@omaya.fr",
            attachments=[(xlsx_name, xlsx_bytes)],
        )
    except Exception:
        res.mail_envoye = False
