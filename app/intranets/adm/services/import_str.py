"""Service Fen_ImportSTR (ADM Imports Bases -> STRATO).

3 types d'import (cf combo TypeImport WinDev) :
  1. Base Journaliere   -> ImportJournalier
  2. RUN                -> ImportRUN
  3. Import Resil Hebdo -> ImportResilHebdo

Pattern multi-fichier (sFichier = noms separes par RC).
"""

from __future__ import annotations

import base64
import io
from datetime import date, datetime
from typing import Optional

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection
from app.shared.recrutement.services.recherche_cv import _int, _str
from app.intranets.adm.services.import_eni import (
    _new_id, _lookup_or_create_client,
    _col_letter_to_index, _cell, _parse_date_fr, _dernier_jour_mois,
)


# Mapping colonnes Base Journaliere STR (cf groupe grAlexJournalier)
# Visible sur l'ecran : Date Signature E, Vendeur nom C, prenom D, Num BS N,
# Statut vente Q, Remarques R, Distributeur B, Client (Civilite I/Nom J/Prenom K),
# Client CP L, VILLE M
COLS_BJ_STR = {
    "distributeur": "B",        "vendeur_nom": "C",
    "vendeur_prenom": "D",      "date_signature": "E",
    "client_civilite": "I",     "client_nom": "J",
    "client_prenom": "K",       "client_cp": "L",
    "client_ville": "M",        "num_bs": "N",
    "statut_vente": "Q",        "remarques": "R",
}


# Mapping colonnes Import Resil Hebdo STR (cf groupe _AA1 dans grAlexJournalier)
# D'apres l'ecran : Date Sign E, Vendeur nom C/prenom D, Num BS S, Statut V,
# Remarques W, Client Civilite I, Nom J, Prenom K, adr1 L, adr2 M, CP N,
# VILLE O, Tel P, Mobile Q, DNAISS R
COLS_RH_STR = {
    "vendeur_nom": "C",         "vendeur_prenom": "D",
    "date_signature": "E",      "client_civilite": "I",
    "client_nom": "J",          "client_prenom": "K",
    "client_adr1": "L",         "client_adr2": "M",
    "client_cp": "N",           "client_ville": "O",
    "client_tel": "P",          "client_mobile": "Q",
    "client_naiss": "R",        "num_bs": "S",
    "statut": "V",              "remarques": "W",
}


# Mapping colonnes RUN STR (cf groupe _AA, meme layout que _AA1 du Resil Hebdo
# mais sans suffixe 1). Pareil que COLS_RH_STR : C/D vendeur, E dateSign,
# S NumBS, V Statut, W Remarques, I-R Client.
COLS_RUN_STR = {
    "vendeur_nom": "C",         "vendeur_prenom": "D",
    "date_signature": "E",      "client_civilite": "I",
    "client_nom": "J",          "client_prenom": "K",
    "client_adr1": "L",         "client_adr2": "M",
    "client_cp": "N",           "client_ville": "O",
    "client_tel": "P",          "client_mobile": "Q",
    "client_naiss": "R",        "num_bs": "S",
    "statut": "V",              "remarques": "W",
}


class ImportStrParams(BaseModel):
    type_import: int                    # 1, 2, 3
    simulation: bool = True
    periode1_du: str = ""
    periode1_au: str = ""
    periode1_mois_paiement: str = ""
    periode2_du: str = ""
    periode2_au: str = ""
    periode2_mois_paiement: str = ""
    mois_paiement_distrib: str = ""


class ImportStrResume(BaseModel):
    nb_fichiers: int = 0
    nb_ajoutes: int = 0
    nb_valides: int = 0
    nb_resilies: int = 0
    nb_decommissions: int = 0
    nb_deja_saisis: int = 0
    nb_deja_statues: int = 0
    nb_introuvables: int = 0
    nb_doublons: int = 0
    nb_hors_delai: int = 0
    nb_pb_vendeur: int = 0
    nb_erreurs: int = 0


class ImportStrResult(BaseModel):
    ok: bool
    type_import: int
    type_label: str
    simulation: bool
    resume: ImportStrResume
    fichiers_traites: list[str] = []
    contrats_ajoutes: list[dict] = []
    contrats_modifies: list[dict] = []
    contrats_non_trouves: list[dict] = []
    contrats_run: list[dict] = []
    pb_vendeur: list[dict] = []
    message: str = ""
    xlsx_b64: str = ""
    xlsx_name: str = ""
    mail_envoye: bool = False


TYPE_LABELS = {
    1: "Base Journalière",
    2: "RUN",
    3: "Import Résil Hebdo",
}


# ---------------------------------------------------------------------------
# Helpers STR
# ---------------------------------------------------------------------------


def _id_produit_str_by_lib(lib_produit: str) -> tuple[int, str]:
    """Lookup pgt_str_produit by lib_produit LIKE."""
    if not lib_produit:
        return (0, "")
    db = get_pg_connection("adv")
    r = db.query_one(
        """SELECT id_produit, lib_produit FROM adv.pgt_str_produit
            WHERE LOWER(lib_produit) LIKE LOWER(?)
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            LIMIT 1""",
        (f"%{lib_produit}%",),
    )
    return ((_int(r.get("id_produit")), _str(r.get("lib_produit")))
            if r else (0, ""))


def _id_etat_str_by_lib(lib_statut: str) -> int:
    """Lookup pgt_str_etat_contrat by lib_etat LIKE."""
    if not lib_statut:
        return 0
    db = get_pg_connection("adv")
    r = db.query_one(
        """SELECT id_etat FROM adv.pgt_str_etat_contrat
            WHERE LOWER(lib_etat) LIKE LOWER(?)
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
            LIMIT 1""",
        (f"{lib_statut}%",),
    )
    return _int(r.get("id_etat")) if r else 0


def _lookup_vendeur_str_nom_prenom(nom: str, prenom: str) -> int:
    """Lookup salarie via 5 variantes (cf WinDev ReqRechercheSalarieByNomPrenom) :
    1. nom + prenom exacts
    2. nom + capitalise(prenom) (LIKE)
    3. nom + premiere lettre prenom (%)
    4. echange : prenom + nom
    5. prenom + premiere lettre nom (%)
    Retourne 0 si introuvable ou ambigu (>1 match)."""
    if not nom and not prenom:
        return 0
    nom = (nom or "").strip()
    prenom = (prenom or "").strip()
    if not nom and not prenom:
        return 0
    db = get_pg_connection("rh")

    def _q(n: str, p: str) -> int:
        rows = db.query(
            """SELECT id_salarie FROM rh.pgt_salarie
                WHERE UPPER(nom) LIKE UPPER(?) AND UPPER(prenom) LIKE UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 2""",
            (n, p),
        ) or []
        return int(rows[0]["id_salarie"]) if len(rows) == 1 else 0

    if nom and prenom:
        for n, p in [
            (nom, prenom),
            (nom, prenom.title()),
            (nom, prenom[:1] + "%" if prenom else ""),
            (prenom, nom),
            (prenom, nom[:1] + "%" if nom else ""),
        ]:
            r = _q(n, p)
            if r:
                return r
    return 0


def _id_etat_str_journ(statut: str, remarques: str) -> int:
    """Determine id_etat_contrat selon statut + remarques (cf WinDev) :
    - default = 37 (En cours)
    - 'RESILIE' -> 16
    - 'NON VALIDE' -> 14 (ou 22 si signatures/IBAN/SEPA)
    - 'REJET' ou 'KO' -> 14 (ou 22 si signatures/IBAN/SEPA,
      12 si doublon, 63 si absence mail, 13 si telephone/adresse)
    - default + 'EN ATTENTE' dans remarques -> 69 (Temporaire)
    """
    s = (statut or "").upper()
    r = (remarques or "").upper()
    id_etat = 37

    if "RESILIE" in s:
        id_etat = 16
    elif "NON VALIDE" in s:
        id_etat = 14
        if ("SIGNATURES NON CONFORMES" in r or "IBAN" in r or "SEPA" in r):
            id_etat = 22
    elif "REJET" in s or "KO" in s:
        id_etat = 14
        if ("SIGNATURES NON CONFORMES" in r or "IBAN" in r or "SEPA" in r):
            id_etat = 22
        if "DOUBLON" in r:
            id_etat = 12
        if "ABSENCE MAIL" in r or "MAIL DOCUSIGN DIFFERENT" in r:
            id_etat = 63
        if "TÉLÉPHONE ERRONÉ" in r or "TELEPHONE ERRONE" in r or \
           "ABSENCE ADRESSE POSTALE" in r:
            id_etat = 13

    if id_etat == 37 and "EN ATTENTE" in r:
        id_etat = 69
    return id_etat


def _lookup_tk_call_str(num_bs: str) -> dict:
    """ReqTkCall_ByNumCtt pour STR : recupere id_salarie + info client."""
    if not num_bs:
        return {}
    try:
        db = get_pg_connection("ticket_bo")
        r = db.query_one(
            """SELECT t.id_salarie, t.nom_client, t.nom_marital_client,
                      t.prenom_client, t.datenaiss, t.adresse1, t.adresse2,
                      t.cp, t.ville, t.mobile1, t.adr_mail
                 FROM ticket_bo.pgt_tk_call t
                 JOIN ticket_bo.pgt_tk_call_panier p ON p.id_tk_liste = t.id_tk_liste
                WHERE UPPER(p.num_bs) = UPPER(?)
                  AND (t.modif_elem IS NULL OR t.modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )
        return r or {}
    except Exception:
        return {}


def _info_salarie_str(id_sal: int) -> dict:
    if not id_sal:
        return {}
    db = get_pg_connection("rh")
    r = db.query_one(
        """SELECT s.id_salarie, s.nom, s.prenom, e.id_ste
             FROM rh.pgt_salarie s
             LEFT JOIN rh.pgt_salarie_embauche e ON e.id_salarie = s.id_salarie
            WHERE s.id_salarie = ? LIMIT 1""",
        (int(id_sal),),
    )
    return r or {}


def _ajoute_histo_str_etat(id_contrat: int, old_etat: int, new_etat: int,
                            date_paiement: str, op_id: int) -> None:
    if not id_contrat:
        return
    db = get_pg_connection("adv")
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_histo_auto), 0) + 1 AS n FROM adv.pgt_str_histo_etat_ctt"
    )
    db.query(
        """INSERT INTO adv.pgt_str_histo_etat_ctt
              (id_histo_auto, id_histo, id_contrat, op_saisie, date,
               old_etat, new_etat, date_paiement,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, NOW(), ?, ?, ?, ?, NOW(), 'new')""",
        (int(auto["n"]) if auto else 1, _new_id(),
         int(id_contrat), int(op_id),
         int(old_etat) if old_etat else 0,
         int(new_etat) if new_etat else 0,
         date_paiement or "", int(op_id)),
    )


def _create_str_contrat(td: dict, op_id: int) -> int:
    """Cf. WinDev STR_ajoutFicheContrat + traiterClient.

    Cree le client (via traiter_client si _client fourni) + calcule
    nb_points via calcul_point_contrat + INSERT contrat + historisation
    etat initial.
    """
    db = get_pg_connection("adv")
    id_contrat = _new_id()
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_contrat_auto), 0) + 1 AS n FROM adv.pgt_str_contrat"
    )
    date_sign = td.get("date_signature")
    if isinstance(date_sign, str):
        date_sign = _parse_date_fr(date_sign)

    # 1. Cree/enrichit le client via traiter_client si _client fourni
    id_client = int(td.get("id_client") or 0)
    tk_client = td.get("_client") or {}
    if not id_client and tk_client and (
            tk_client.get("nom") or tk_client.get("mail")):
        try:
            from app.intranets.adm.services.import_helpers_common import (
                traiter_client,
            )
            id_client = traiter_client(
                info_client={
                    "nom": tk_client.get("nom") or "",
                    "prenom": tk_client.get("prenom") or "",
                    "adresse1": tk_client.get("adresse") or "",
                    "adresse2": tk_client.get("cplt") or "",
                    "cp": tk_client.get("cp") or "",
                    "ville": tk_client.get("ville") or "",
                    "gsm": tk_client.get("gsm") or "",
                    "mail": tk_client.get("mail") or "",
                    "date_naiss": tk_client.get("date_naiss"),
                    "op_saisie": op_id, "modif_op": op_id,
                }, force_maj=False, op_id=op_id,
            ) or 0
        except Exception:
            id_client = 0

    # 2. Calcul nb_points (cf. WinDev L33 ajoutFicheContrat :
    # calculPointContrat('ASSU','ST',0,DateSignature,''))
    nb_pts = 0.0
    id_prod = int(td.get("id_produit") or 0)
    if date_sign:
        try:
            from app.shared.sdtc.bareme import calcul_point_contrat
            # Lookup famille/sous_fam si dispo, sinon defaut ASSU/ST
            fam, ss_fam = "ASSU", "ST"
            if id_prod:
                prod = db.query_one(
                    """SELECT famille, sous_fam FROM adv.pgt_str_produit
                        WHERE id_produit = ? LIMIT 1""",
                    (id_prod,),
                )
                if prod:
                    fam = prod.get("famille") or fam
                    ss_fam = prod.get("sous_fam") or ss_fam
            nb_pts = float(calcul_point_contrat(
                fam=fam, ss_fam=ss_fam, palier=0,
                date_sign=str(date_sign),
                info_cplt="", palier2=0,
            ) or 0)
        except Exception:
            nb_pts = 0.0

    etat_initial = int(td.get("etat_contrat") or 1)

    # 3. INSERT contrat complet (cf. WinDev ajoutFicheContrat)
    db.query(
        """INSERT INTO adv.pgt_str_contrat
              (id_contrat_auto, id_contrat, id_client, id_salarie, id_ste,
               num_bs, id_produit, id_etat_contrat,
               date_signature, info_partagee, info_interne, mois_p, nb_points,
               op_saisie, date_saisie, non_call, opt_mandat, code_enr,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?,
                   ?, ?, ?, '', ?,
                   ?, NOW(), ?, ?, '',
                   ?, NOW(), 'new')""",
        (int(auto["n"]) if auto else 1, id_contrat,
         id_client,
         int(td.get("id_salarie") or 0),
         int(td.get("id_ste") or 0),
         td.get("num_bs") or "",
         id_prod,
         etat_initial,
         date_sign,
         td.get("info_partagee") or "",
         td.get("info_interne") or "",
         nb_pts,
         int(op_id),
         # cf. WinDev NonCALL = Faux par defaut
         bool(td.get("non_call", False)),
         bool(td.get("opt_mandat")),
         int(op_id)),
    )

    # 4. Historisation etat initial
    try:
        _ajoute_histo_str_etat(id_contrat, 0, etat_initial, "", op_id)
    except Exception:
        pass

    return id_contrat


# ---------------------------------------------------------------------------
# Type 1 : ImportJournalier (squelette en attente du code WinDev)
# ---------------------------------------------------------------------------


def _import_journalier_str(
    p: ImportStrParams, fname: str, content: bytes, op_id: int,
    ajoutes: list, modifies: list, pb_vendeur: list, resume: ImportStrResume,
) -> None:
    """Type 1 : Base Journaliere STR (cf WinDev ImportJournalier) :
    - idProd = 78 par defaut (STR)
    - Determine idEtat selon statut + remarques (cf _id_etat_str_journ)
    - Si BS existe -> 'Contrats deja saisis' (modifies)
      * Si idEtat different + type_etat actuel in (1,2) -> MAJ etat + histo
    - Si BS n'existe pas :
      * Lookup vendeur multi-variantes (5 fallbacks)
      * Si pas trouve -> sheet 'Vendeur introuvable' (pb_vendeur)
      * Lookup TK_Call STR -> remplace info client si trouve
      * Ajout (sheet 1, creation contrat en prod via _create_str_contrat)
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        modifies.append({"_erreur": f"Lecture {fname} : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_BJ_STR.items()}
    db = get_pg_connection("adv")

    ID_PROD_STR = 78
    # Lookup type prod (cache)
    prod_info = db.query_one(
        "SELECT lib_produit FROM adv.pgt_str_produit WHERE id_produit = ? LIMIT 1",
        (ID_PROD_STR,),
    )
    type_prod = (prod_info.get("lib_produit") if prod_info else "") or ""

    for i in range(2, (ws.max_row or 0) + 1):
        num_bs = _cell(ws, i, cols["num_bs"]).upper()
        if not num_bs:
            continue
        distributeur = _cell(ws, i, cols["distributeur"])
        vendeur_nom = _cell(ws, i, cols["vendeur_nom"])
        vendeur_prenom = _cell(ws, i, cols["vendeur_prenom"])
        date_sign = _parse_date_fr(_cell(ws, i, cols["date_signature"]))
        client_civilite = _cell(ws, i, cols["client_civilite"])
        client_nom = _cell(ws, i, cols["client_nom"])
        client_prenom = _cell(ws, i, cols["client_prenom"])
        client_cp = _cell(ws, i, cols["client_cp"])
        client_ville = _cell(ws, i, cols["client_ville"])
        statut_vente = _cell(ws, i, cols["statut_vente"])
        remarques = _cell(ws, i, cols["remarques"])

        # Determine id_etat
        id_etat = _id_etat_str_journ(statut_vente, remarques)
        etat_info = db.query_one(
            "SELECT lib_etat FROM adv.pgt_str_etat_contrat WHERE id_etat = ? LIMIT 1",
            (id_etat,),
        )
        lib_etat = (etat_info.get("lib_etat") if etat_info else "") or ""

        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, id_client, num_bs,
                      date_signature, id_etat_contrat, opt_mandat
                 FROM adv.pgt_str_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )

        if ctt:
            # Contrat existant
            id_contrat = int(ctt["id_contrat"])
            id_sal_db = int(ctt.get("id_salarie") or 0)
            etat_db = int(ctt.get("id_etat_contrat") or 0)
            etat_db_info = db.query_one(
                """SELECT id_type_etat FROM adv.pgt_str_etat_contrat
                    WHERE id_etat = ? LIMIT 1""",
                (etat_db,),
            )
            type_etat_db = (int(etat_db_info.get("id_type_etat") or 0)
                            if etat_db_info else 0)

            info_sal_db = _info_salarie_str(id_sal_db)
            agence, equipe, _ = _affectation_str(id_sal_db)
            changement_etat = "non"

            if id_etat != etat_db and type_etat_db in (1, 2):
                changement_etat = "oui"
                if not p.simulation:
                    try:
                        db.query(
                            """UPDATE adv.pgt_str_contrat
                                  SET id_etat_contrat = ?,
                                      modif_date = NOW(), modif_op = ?,
                                      modif_elem = 'modif'
                                WHERE id_contrat = ?""",
                            (id_etat, int(op_id), id_contrat),
                        )
                        _ajoute_histo_str_etat(id_contrat, etat_db, id_etat,
                                               "", op_id)
                    except Exception as e:
                        changement_etat = f"oui (err: {e})"

            modifies.append({
                "NumBS": num_bs,
                "ClientNom": client_nom, "ClientPrenom": client_prenom,
                "ClientCP": client_cp, "ClientVille": client_ville,
                "Vendeur": (f"{_str(info_sal_db.get('nom'))} "
                            f"{capitalise_str(_str(info_sal_db.get('prenom')))}".strip()),
                "DateSign": str(ctt.get("date_signature") or ""),
                "TypeProd": type_prod, "Statut": lib_etat,
                "Agence": agence, "Equipe": equipe,
                "SEPADemat": bool(ctt.get("opt_mandat")),
                "Remarques": remarques, "MAJ Etat": changement_etat,
            })
            resume.nb_deja_saisis += 1
            continue

        # Contrat nouveau : lookup vendeur
        id_vendeur = _lookup_vendeur_str_nom_prenom(vendeur_nom, vendeur_prenom)
        nom_vend_final = f"{vendeur_nom} {vendeur_prenom}".strip()
        agence = ""; equipe = ""
        id_ste = 0

        if id_vendeur == 0:
            # Vendeur introuvable -> sheet 3
            pb_vendeur.append({
                "NumBS": num_bs, "ClientNom": client_nom,
                "ClientPrenom": client_prenom, "ClientCP": client_cp,
                "ClientVille": client_ville,
                "Vendeur": nom_vend_final,
                "DateSign": str(date_sign or ""),
                "TypeProd": type_prod, "Statut": lib_etat,
                "Remarques": remarques,
                "Erreur": "Vendeur introuvable",
            })
            resume.nb_pb_vendeur += 1
        else:
            info_sal = _info_salarie_str(id_vendeur)
            id_ste = int(info_sal.get("id_ste") or 0)
            agence, equipe, _ = _affectation_str(id_vendeur)
            nom_vend_final = (f"{_str(info_sal.get('nom'))} "
                              f"{capitalise_str(_str(info_sal.get('prenom')))}".strip())

        # Lookup TK_Call STR pour remplacer info client si trouve
        tk = _lookup_tk_call_str(num_bs)
        if tk:
            if tk.get("id_salarie"):
                id_vendeur = int(tk["id_salarie"])
                info_sal = _info_salarie_str(id_vendeur)
                id_ste = int(info_sal.get("id_ste") or 0)
                agence, equipe, _ = _affectation_str(id_vendeur)
                nom_vend_final = (f"{_str(info_sal.get('nom'))} "
                                  f"{capitalise_str(_str(info_sal.get('prenom')))}".strip())
            client_nom = tk.get("nom_client") or client_nom
            if tk.get("nom_marital_client"):
                client_nom += f" ep {tk['nom_marital_client']}"
            client_prenom = tk.get("prenom_client") or client_prenom
            client_cp = tk.get("cp") or client_cp
            client_ville = tk.get("ville") or client_ville

        # Ajout sheet 1
        ajoutes.append({
            "NumBS": num_bs, "ClientNom": client_nom,
            "ClientPrenom": client_prenom, "ClientCP": client_cp,
            "ClientVille": client_ville,
            "ClientTel": _str(tk.get("mobile1")) if tk else "",
            "ClientMail": _str(tk.get("adr_mail")) if tk else "",
            "Vendeur": nom_vend_final, "DateSign": str(date_sign or ""),
            "TypeProd": type_prod, "Statut": lib_etat,
            "Agence": agence, "Equipe": equipe, "SEPADemat": False,
            "Distributeur": distributeur, "Remarques": remarques,
            "_payload_create": {
                "num_bs": num_bs, "id_salarie": id_vendeur,
                "id_ste": id_ste, "id_produit": ID_PROD_STR,
                "etat_contrat": id_etat,
                "date_signature": date_sign,
                "non_call": True, "info_interne": remarques,
                "opt_mandat": False,
                "_client": {
                    "nom": client_nom, "prenom": client_prenom,
                    "cp": client_cp, "ville": client_ville,
                    "adresse": tk.get("adresse1") if tk else "",
                    "adresse2": tk.get("adresse2") if tk else "",
                    "tel": tk.get("mobile1") if tk else "",
                    "mail": tk.get("adr_mail") if tk else "",
                    "datenaiss": tk.get("datenaiss") if tk else None,
                },
            },
        })
        resume.nb_ajoutes += 1

    wb.close()


def _affectation_str(id_salarie: int) -> tuple[str, str, bool]:
    """(agence, equipe, is_distrib) via JOIN organigramme + niveau 3/4."""
    if not id_salarie:
        return ("", "", False)
    try:
        db = get_pg_connection("rh")
        rows = db.query(
            """SELECT o.lib_orga, o.id_type_niveau_orga, o.id_type_orga
                 FROM rh.pgt_salarie_organigramme so
                 JOIN rh.pgt_organigramme o ON o.idorganigramme = so.idorganigramme
                WHERE so.id_salarie = ?
                  AND (so.modif_elem IS NULL OR so.modif_elem NOT LIKE '%suppr%')""",
            (int(id_salarie),),
        ) or []
        agence = ""; equipe = ""; is_distrib = False
        for r in rows:
            lvl = r.get("id_type_niveau_orga")
            lib = r.get("lib_orga") or ""
            if lvl == 3 and not agence:
                agence = lib
                if int(r.get("id_type_orga") or 0) == 3:
                    is_distrib = True
            elif lvl == 4 and not equipe:
                equipe = lib
        return (agence, equipe, is_distrib)
    except Exception:
        return ("", "", False)


def _detect_periode_str(
    date_sign: Optional[date], is_distrib: bool,
    p1_du: date, p1_au: date, mp1: Optional[date],
    p2_du: date, p2_au: date, mp2: Optional[date],
    mp_distrib: Optional[date],
) -> tuple[Optional[date], str]:
    """Identique aux autres partenaires : (mois_p, periode_lbl)."""
    if not date_sign:
        return (None, "")
    if is_distrib:
        return (mp_distrib, "Période Distrib")
    if p1_du <= date_sign <= p1_au:
        return (mp1, "Période 1")
    if p2_du <= date_sign <= p2_au:
        return (mp2, "Période 2")
    p1_du_m1 = (p1_du.replace(month=p1_du.month - 1) if p1_du.month > 1
                else p1_du.replace(year=p1_du.year - 1, month=12))
    p1_au_m1 = (p1_au.replace(month=p1_au.month - 1) if p1_au.month > 1
                else p1_au.replace(year=p1_au.year - 1, month=12))
    if p1_du_m1 <= date_sign <= p1_au_m1:
        return (mp1, "Période -1 mois")
    return (None, "HORS_DELAI")


def _import_run_str(
    p: ImportStrParams, fname: str, content: bytes, op_id: int,
    runs: list, modifies: list, non_trouves: list, pb_vendeur: list,
    resume: ImportStrResume,
) -> None:
    """Type 2 : RUN STR. Pour chaque ligne :
    - Lookup BS, si pas trouve -> introuvable
    - Detection periode (P1/P2/-1mois/Distrib/HORS_DELAI)
    - Si HORS_DELAI -> sheet 4 (hors_delai)
    - Si statut contient 'VALIDE' :
      * Si type_etat in (1,2) ou etat in (29,30) -> MAJ etat 19
        (Valide-Paye operateur) + mois_p (garde ancien si etat 29/30)
        + histo (couleur verte)
      * Sinon -> deja statue
    - Sinon (resiliation) :
      * Si type_etat in (1,2) -> MAJ etat 16 (Resilie op) ou 57
        (Retractation si statut='RESILIE') + mois_p=null + concat
        Motif Resil (couleur rouge)
      * Sinon -> deja statue
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        modifies.append({"_erreur": f"Lecture {fname} : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_RUN_STR.items()}
    db = get_pg_connection("adv")

    p1_du = _parse_date_fr(p.periode1_du) or date(1900, 1, 1)
    p1_au = _parse_date_fr(p.periode1_au) or date(2100, 12, 31)
    p2_du = _parse_date_fr(p.periode2_du) or date(1900, 1, 1)
    p2_au = _parse_date_fr(p.periode2_au) or date(2100, 12, 31)
    mp1 = _dernier_jour_mois(p.periode1_mois_paiement)
    mp2 = _dernier_jour_mois(p.periode2_mois_paiement)
    mp_distrib = _dernier_jour_mois(p.mois_paiement_distrib)

    for i in range(2, (ws.max_row or 0) + 1):
        num_bs = _cell(ws, i, cols["num_bs"]).upper()
        if not num_bs:
            continue
        statut = _cell(ws, i, cols["statut"])
        remarques = _cell(ws, i, cols["remarques"])
        vendeur_nom = _cell(ws, i, cols["vendeur_nom"])
        vendeur_prenom = _cell(ws, i, cols["vendeur_prenom"])
        client_nom = _cell(ws, i, cols["client_nom"])
        client_prenom = _cell(ws, i, cols["client_prenom"])
        client_adr1 = _cell(ws, i, cols["client_adr1"])
        client_cp = _cell(ws, i, cols["client_cp"])
        client_ville = _cell(ws, i, cols["client_ville"])

        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, num_bs, date_signature,
                      id_etat_contrat, info_partagee, mois_p
                 FROM adv.pgt_str_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )
        if not ctt:
            non_trouves.append({
                "NumBS": num_bs, "ClientNom": client_nom,
                "ClientPrenom": client_prenom, "ClientCP": client_cp,
                "ClientVille": client_ville,
                "Vendeur": f"{vendeur_nom} {capitalise_str(vendeur_prenom)}".strip(),
                "Statut": statut, "Remarques": remarques,
            })
            resume.nb_introuvables += 1
            continue

        id_contrat = int(ctt["id_contrat"])
        id_etat_actuel = int(ctt.get("id_etat_contrat") or 0)
        date_sign_db = ctt.get("date_signature")
        mois_p_omaya = ctt.get("mois_p")
        info_existing = ctt.get("info_partagee") or ""

        # Lookup type_etat actuel
        etat_info = db.query_one(
            """SELECT id_type_etat, lib_etat FROM adv.pgt_str_etat_contrat
                WHERE id_etat = ? LIMIT 1""",
            (id_etat_actuel,),
        )
        id_type_etat = int(etat_info.get("id_type_etat") or 0) if etat_info else 0
        lib_etat_actuel = (etat_info.get("lib_etat") or "") if etat_info else ""

        info_sal = _info_salarie_str(int(ctt.get("id_salarie") or 0))
        nom_vend_db = f"{_str(info_sal.get('nom'))} {capitalise_str(_str(info_sal.get('prenom')))}".strip()

        # Detection periode
        agence, equipe, is_distrib = _affectation_str(int(ctt.get("id_salarie") or 0))
        mois_p, periode_lbl = _detect_periode_str(
            date_sign_db, is_distrib,
            p1_du, p1_au, mp1, p2_du, p2_au, mp2, mp_distrib,
        )
        mois_p_str = mois_p.strftime("%m-%Y") if mois_p else ""

        if periode_lbl == "HORS_DELAI":
            resume.nb_hors_delai += 1
            pb_vendeur.append({
                "NumBS": num_bs, "ClientNom": client_nom,
                "ClientPrenom": client_prenom,
                "DateSign": str(date_sign_db or ""),
                "Vendeur": nom_vend_db,
                "TypeEtat actuel": str(id_type_etat),
                "Lib Etat actuel": lib_etat_actuel,
                "Agence": agence, "Equipe": equipe,
                "Statut Import": statut, "Erreur": "Hors Délai",
            })
            continue

        is_valide = "VALIDE" in statut.upper()

        if is_valide:
            # Cas VALIDE : eligible si type_etat in (1,2) ou etat in (29,30)
            if id_type_etat in (1, 2) or id_etat_actuel in (29, 30):
                # Garde mois_p ancien si etat etait 29/30
                new_mois_p = mois_p_omaya if id_etat_actuel in (29, 30) else mois_p
                runs.append({
                    "NumBS": num_bs, "ClientNom": client_nom,
                    "ClientPrenom": client_prenom, "ClientAdr": client_adr1,
                    "ClientCP": client_cp, "ClientVille": client_ville,
                    "DateSign": str(date_sign_db or ""),
                    "Vendeur": nom_vend_db, "Agence": agence, "Equipe": equipe,
                    "Periode": periode_lbl, "MoisP": mois_p_str,
                    "AncienEtat": lib_etat_actuel,
                    "NouvelEtat": "Valide - Payé par l'opérateur",
                    "_id_contrat": id_contrat,
                    "_old_etat": id_etat_actuel, "_new_etat": 19,
                    "_new_mois_p": new_mois_p, "_traitement": "valide",
                })
                resume.nb_valides += 1
            else:
                modifies.append({
                    "NumBS": num_bs, "ClientNom": client_nom,
                    "ClientPrenom": client_prenom, "ClientCP": client_cp,
                    "ClientVille": client_ville,
                    "DateSign": str(date_sign_db or ""),
                    "Vendeur": nom_vend_db, "Agence": agence, "Equipe": equipe,
                    "Statut Import": statut, "EtatEnregistre": lib_etat_actuel,
                    "Periode": periode_lbl,
                })
                resume.nb_deja_statues += 1
        else:
            # Cas RESILIATION : eligible si type_etat in (1,2)
            if id_type_etat in (1, 2):
                new_etat = 57 if statut.upper() == "RESILIE" else 16
                new_etat_info = db.query_one(
                    """SELECT lib_etat FROM adv.pgt_str_etat_contrat
                        WHERE id_etat = ? LIMIT 1""",
                    (new_etat,),
                )
                new_etat_lib = ((new_etat_info.get("lib_etat") or "")
                                if new_etat_info else "")
                runs.append({
                    "NumBS": num_bs, "ClientNom": client_nom,
                    "ClientPrenom": client_prenom, "ClientAdr": client_adr1,
                    "ClientCP": client_cp, "ClientVille": client_ville,
                    "DateSign": str(date_sign_db or ""),
                    "Vendeur": nom_vend_db, "Agence": agence, "Equipe": equipe,
                    "Periode": periode_lbl, "MoisP": mois_p_str,
                    "AncienEtat": lib_etat_actuel,
                    "NouvelEtat": new_etat_lib,
                    "_id_contrat": id_contrat,
                    "_old_etat": id_etat_actuel, "_new_etat": new_etat,
                    "_remarques": remarques,
                    "_info_existing": info_existing,
                    "_traitement": "resil",
                })
                resume.nb_resilies += 1
            else:
                modifies.append({
                    "NumBS": num_bs, "ClientNom": client_nom,
                    "ClientPrenom": client_prenom, "ClientCP": client_cp,
                    "ClientVille": client_ville,
                    "DateSign": str(date_sign_db or ""),
                    "Vendeur": nom_vend_db, "Agence": agence, "Equipe": equipe,
                    "Statut Import": statut, "EtatEnregistre": lib_etat_actuel,
                    "Periode": periode_lbl,
                })
                resume.nb_deja_statues += 1

    # PASSE PROD : MAJ etat sur les runs
    if not p.simulation:
        for row in runs:
            id_ct = row.pop("_id_contrat", None)
            old_etat = row.pop("_old_etat", 0)
            new_etat = row.pop("_new_etat", 0)
            traitement = row.pop("_traitement", "")
            new_mois_p = row.pop("_new_mois_p", None)
            remarques = row.pop("_remarques", "")
            info_existing = row.pop("_info_existing", "")
            if not id_ct or not new_etat:
                continue
            try:
                if traitement == "valide":
                    db.query(
                        """UPDATE adv.pgt_str_contrat
                              SET id_etat_contrat = ?, mois_p = ?,
                                  modif_op = ?, modif_date = NOW(),
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (int(new_etat), new_mois_p, int(op_id), int(id_ct)),
                    )
                else:  # resil
                    new_info = (info_existing or "") + f"\nMotif Résil : {remarques}"
                    db.query(
                        """UPDATE adv.pgt_str_contrat
                              SET id_etat_contrat = ?, info_partagee = ?,
                                  mois_p = NULL,
                                  modif_op = ?, modif_date = NOW(),
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (int(new_etat), new_info, int(op_id), int(id_ct)),
                    )
                _ajoute_histo_str_etat(int(id_ct), int(old_etat), int(new_etat),
                                       str(new_mois_p or ""), op_id)
            except Exception as e:
                row["Erreur"] = str(e)
    else:
        for row in runs:
            for k in ("_id_contrat", "_old_etat", "_new_etat", "_traitement",
                      "_new_mois_p", "_remarques", "_info_existing"):
                row.pop(k, None)

    wb.close()


def _import_resil_hebdo_str(
    p: ImportStrParams, fname: str, content: bytes, op_id: int,
    runs: list, modifies: list, non_trouves: list, pb_vendeur: list,
    resume: ImportStrResume,
) -> None:
    """Type 3 : Import Resil Hebdo STR. Pour chaque ligne :
    - Lookup contrat par num_bs
    - Si trouve + statut != 'VALIDE' :
      * Si type_etat in (1, 2) (en attente) -> passe a 16 (Resilie par
        operateur) ou 57 (Retractation client) si statut='RESILIE'.
        + concat 'Motif Resil' a info_partagee + reset mois_p + histo.
      * Sinon -> deja statue (sheet 3).
    - Si statut contient 'VALIDE' -> on ne fait rien (cf WinDev).
    - Si pas trouve -> contrat introuvable (sheet 2).
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        modifies.append({"_erreur": f"Lecture {fname} : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_RH_STR.items()}
    db = get_pg_connection("adv")

    for i in range(2, (ws.max_row or 0) + 1):
        num_bs = _cell(ws, i, cols["num_bs"]).upper()
        if not num_bs:
            continue
        statut = _cell(ws, i, cols["statut"])
        remarques = _cell(ws, i, cols["remarques"])
        vendeur_nom = _cell(ws, i, cols["vendeur_nom"])
        vendeur_prenom = _cell(ws, i, cols["vendeur_prenom"])
        client_nom = _cell(ws, i, cols["client_nom"])
        client_prenom = _cell(ws, i, cols["client_prenom"])
        client_adr1 = _cell(ws, i, cols["client_adr1"])
        client_cp = _cell(ws, i, cols["client_cp"])
        client_ville = _cell(ws, i, cols["client_ville"])

        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, num_bs, date_signature,
                      id_etat_contrat, info_partagee
                 FROM adv.pgt_str_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_bs,),
        )
        if not ctt:
            non_trouves.append({
                "NumBS": num_bs, "ClientNom": client_nom,
                "ClientPrenom": client_prenom, "ClientCP": client_cp,
                "ClientVille": client_ville,
                "Vendeur": f"{vendeur_nom} {capitalise_str(vendeur_prenom)}".strip(),
                "Statut": statut, "Remarques": remarques,
            })
            resume.nb_introuvables += 1
            continue

        # Si statut contient VALIDE -> rien a faire
        if "VALIDE" in statut.upper():
            continue

        id_contrat = int(ctt["id_contrat"])
        id_etat_actuel = int(ctt.get("id_etat_contrat") or 0)
        date_sign_db = ctt.get("date_signature")

        # Lookup type_etat actuel
        etat_info = db.query_one(
            """SELECT id_type_etat, lib_etat FROM adv.pgt_str_etat_contrat
                WHERE id_etat = ? LIMIT 1""",
            (id_etat_actuel,),
        )
        id_type_etat = int(etat_info.get("id_type_etat") or 0) if etat_info else 0
        lib_etat_actuel = (etat_info.get("lib_etat") or "") if etat_info else ""

        info_sal = _info_salarie_str(int(ctt.get("id_salarie") or 0))
        nom_vend_db = f"{_str(info_sal.get('nom'))} {_str(info_sal.get('prenom'))}".strip()

        if id_type_etat in (1, 2):
            # En attente -> on resilie
            new_etat = 57 if statut.upper() == "RESILIE" else 16
            new_etat_info = db.query_one(
                """SELECT lib_etat FROM adv.pgt_str_etat_contrat
                    WHERE id_etat = ? LIMIT 1""",
                (new_etat,),
            )
            new_etat_lib = (new_etat_info.get("lib_etat") or "") if new_etat_info else ""

            modifies.append({
                "NumBS": num_bs, "ClientNom": client_nom,
                "ClientPrenom": client_prenom,
                "ClientAdr": client_adr1, "ClientCP": client_cp,
                "ClientVille": client_ville,
                "DateSign": str(date_sign_db or ""),
                "Vendeur": nom_vend_db,
                "AncienEtat": lib_etat_actuel,
                "NouvelEtat": new_etat_lib,
                "Remarques": remarques,
                "_id_contrat": id_contrat, "_old_etat": id_etat_actuel,
                "_new_etat": new_etat, "_info_existing": ctt.get("info_partagee") or "",
            })
            resume.nb_resilies += 1
        else:
            # Deja statue
            runs.append({
                "NumBS": num_bs, "ClientNom": client_nom,
                "ClientPrenom": client_prenom, "ClientCP": client_cp,
                "ClientVille": client_ville,
                "DateSign": str(date_sign_db or ""),
                "Vendeur": nom_vend_db,
                "StatutImport": statut, "EtatEnregistre": lib_etat_actuel,
            })
            resume.nb_deja_statues += 1

    # PASSE PROD : reset etat sur les modifies
    if not p.simulation:
        for row in modifies:
            id_ct = row.pop("_id_contrat", None)
            old_etat = row.pop("_old_etat", 0)
            new_etat = row.pop("_new_etat", 0)
            info_existing = row.pop("_info_existing", "")
            if not id_ct or not new_etat:
                continue
            try:
                new_info = (info_existing or "") + f"\nMotif Résil : {row.get('Remarques', '')}"
                db.query(
                    """UPDATE adv.pgt_str_contrat
                          SET id_etat_contrat = ?, info_partagee = ?,
                              mois_p = NULL,
                              modif_op = ?, modif_date = NOW(),
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (int(new_etat), new_info, int(op_id), int(id_ct)),
                )
                _ajoute_histo_str_etat(int(id_ct), int(old_etat), int(new_etat),
                                       "", op_id)
            except Exception as e:
                row["Erreur"] = str(e)
    else:
        # Nettoyage internal keys quand meme
        for row in modifies:
            row.pop("_id_contrat", None)
            row.pop("_old_etat", None)
            row.pop("_new_etat", None)
            row.pop("_info_existing", None)

    wb.close()


def capitalise_str(s: str) -> str:
    """Capitalise la 1ere lettre (equivalent WinDev capitalise)."""
    return s[:1].upper() + s[1:].lower() if s else ""


def run_import_str(
    p: ImportStrParams, files: list[tuple[str, bytes]], op_id: int,
) -> ImportStrResult:
    """Dispatcher principal."""
    label = TYPE_LABELS.get(p.type_import, "?")
    if not files:
        return ImportStrResult(
            ok=False, type_import=p.type_import, type_label=label,
            simulation=p.simulation, resume=ImportStrResume(),
            message="Aucun fichier fourni.",
        )

    resume = ImportStrResume(nb_fichiers=len(files))
    ajoutes: list[dict] = []; modifies: list[dict] = []
    non_trouves: list[dict] = []; runs: list[dict] = []
    pb_vendeur: list[dict] = []
    fichiers_traites: list[str] = []

    for fname, content in files:
        fichiers_traites.append(fname)
        if p.type_import == 1:
            _import_journalier_str(
                p, fname, content, op_id,
                ajoutes, modifies, pb_vendeur, resume,
            )
        elif p.type_import == 2:
            _import_run_str(
                p, fname, content, op_id,
                runs, modifies, non_trouves, pb_vendeur, resume,
            )
        elif p.type_import == 3:
            _import_resil_hebdo_str(
                p, fname, content, op_id,
                runs, modifies, non_trouves, pb_vendeur, resume,
            )

    # PASSE PROD
    nb_crees = 0; nb_reattrib = 0
    if not p.simulation:
        db = get_pg_connection("adv")
        for row in ajoutes:
            pl = row.pop("_payload_create", None)
            if not pl or not pl.get("id_salarie"):
                continue
            try:
                new_id = _create_str_contrat(pl, op_id)
                row["IdContratCree"] = new_id
                nb_crees += 1
            except Exception as e:
                row["Erreur"] = str(e)
        for row in pb_vendeur:
            id_ct = row.pop("_id_contrat", None)
            if not id_ct or row.get("Erreur") != "vendeur réattribué":
                continue
            try:
                db.query(
                    """UPDATE adv.pgt_str_contrat
                          SET id_salarie = ?,
                              modif_op = ?, modif_date = NOW(),
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (int(row["NewIdSalarie"]), int(op_id), int(id_ct)),
                )
                nb_reattrib += 1
            except Exception as e:
                row["ErreurMaj"] = str(e)

    for row in ajoutes:
        row.pop("_payload_create", None)
    for row in pb_vendeur:
        row.pop("_id_contrat", None)

    res = ImportStrResult(
        ok=True, type_import=p.type_import, type_label=label,
        simulation=p.simulation, resume=resume,
        fichiers_traites=fichiers_traites,
        contrats_ajoutes=ajoutes, contrats_modifies=modifies,
        contrats_non_trouves=non_trouves, contrats_run=runs,
        pb_vendeur=pb_vendeur,
        message=(
            f"{len(files)} fichier(s) | Ajoutés {resume.nb_ajoutes} | "
            f"Déjà saisis {resume.nb_deja_saisis} | Validés {resume.nb_valides} | "
            f"Résiliés {resume.nb_resilies} | Introuvables {resume.nb_introuvables} | "
            f"Pb vendeur {resume.nb_pb_vendeur}. "
            + (f"PRODUCTION : {nb_crees} créés, {nb_reattrib} reattributions."
               if not p.simulation else "(SIMULATION)")
        ),
    )
    _attach_xlsx_and_mail_str(res, op_id)
    return res


def _build_xlsx_str(res: ImportStrResult) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook(); ws = wb.active; ws.title = "Résumé"
    header_fill = PatternFill("solid", fgColor="17494E")
    header_font = Font(bold=True, color="FFFFFF")
    items = [
        ("NB Fichiers", res.resume.nb_fichiers),
        ("NB Ajoutés", res.resume.nb_ajoutes),
        ("NB Validés", res.resume.nb_valides),
        ("NB Résiliés", res.resume.nb_resilies),
        ("NB Déjà saisis", res.resume.nb_deja_saisis),
        ("NB Déjà statués", res.resume.nb_deja_statues),
        ("NB Introuvables", res.resume.nb_introuvables),
        ("NB Pb Vendeur", res.resume.nb_pb_vendeur),
        ("NB Erreurs", res.resume.nb_erreurs),
    ]
    ws.append(["Indicateur", "Nombre"])
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for lbl, n in items:
        ws.append([lbl, n])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    for title, rows in [
        ("Ajoutés", res.contrats_ajoutes),
        ("Déjà saisis", res.contrats_modifies),
        ("Non trouvés", res.contrats_non_trouves),
        ("RUN", res.contrats_run),
        ("Pb Vendeur", res.pb_vendeur),
    ]:
        if not rows:
            continue
        sh = wb.create_sheet(title=title[:31])
        keys = list(rows[0].keys())
        sh.append(keys)
        for c in sh[1]:
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for r in rows:
            sh.append([str(r.get(k, "")) if r.get(k) is not None else "" for k in keys])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def _attach_xlsx_and_mail_str(res: ImportStrResult, op_id: int) -> None:
    from app.shared.notifications.mail import envoi_mail
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    suffix = "_SIMU" if res.simulation else ""
    prefix_map = {1: "ImportJournalierSTR", 2: "ImportRunSTR",
                  3: "ImportResilHebdoSTR"}
    xlsx_name = f"{prefix_map.get(res.type_import, 'ImportSTR')}_{ts}{suffix}.xlsx"
    try:
        xlsx_bytes = _build_xlsx_str(res)
    except Exception:
        return
    res.xlsx_name = xlsx_name
    res.xlsx_b64 = base64.b64encode(xlsx_bytes).decode("ascii")

    try:
        db = get_pg_connection("rh")
        r = db.query_one(
            "SELECT mail FROM rh.pgt_salarie_coordonnees WHERE id_salarie = ? LIMIT 1",
            (int(op_id),),
        )
        op_mail = (r.get("mail") if r else "") or ""
    except Exception:
        op_mail = ""
    destinataires = [op_mail] if op_mail else ["intranet@omaya.fr"]
    cc = ["intranet@omaya.fr"] if op_mail and op_mail != "intranet@omaya.fr" else []

    sujet_pref = "SIMULATION : " if res.simulation else ""
    sujet = (f"{sujet_pref}Importation {res.type_label} STRATO "
             f"du {date.today().strftime('%d/%m/%Y')}")
    html = (
        "<p>Bonjour,</p>"
        f"<p>Fin importation le : {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>"
        f"<p><strong>{res.message}</strong></p>"
        "<p>Service Importation EXOSPHERE</p>"
    )
    try:
        res.mail_envoye = envoi_mail(
            sujet=sujet, html=html,
            destinataires=destinataires, cc=cc,
            expediteur="intranet@omaya.fr",
            attachments=[(xlsx_name, xlsx_bytes)],
        )
    except Exception:
        res.mail_envoye = False
