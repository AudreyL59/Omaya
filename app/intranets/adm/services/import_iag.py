"""Service Fen_ImportIAG (ADM Imports Bases -> Import partenaire IAG).

2 types d'import (vs 4 pour ENI) :
 1. Base Journalière -> importJournalier()
 2. RUN              -> importValide() ou importResil() selon le nom
    du fichier ('ANNUL' dedans -> resil, sinon valide).

Multi-fichier : sFichier WinDev contient plusieurs noms separes par RC,
le frontend peut en uploader plusieurs d'un coup.

Specificite IAG : c'est une assurance, donc pas de gaz/elec/CAR/puissance
(simplifie vs ENI). Juste num_bs, id_produit, id_etat_contrat,
date_signature, id_salarie, id_client.

Les procedures metier seront codees au fur et a mesure (placeholder
pour l'instant).
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Optional

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection
from app.shared.recrutement.services.recherche_cv import _int, _str
# Reutilise helpers ENI (id, client, histo, points)
from app.intranets.adm.services.import_eni import (
    _new_id, _calcul_points_eni, _lookup_or_create_client,
    _col_letter_to_index, _cell, _parse_date_fr, _normaliser_nom,
    _dernier_jour_mois,
)


# Mapping colonnes par defaut pour 'Base Journaliere IAG'
# (cf. ecran onglet 'Import Journalier' Fen_ImportIAG)
COLS_BJ_IAG = {
    "num_contrat": "A",
    "date_signature": "B",
    "produit_assu": "D",
    "client_nom": "E",
    "client_prenom": "F",
    "client_adresse": "G",
    "client_cp": "H",
    "client_ville": "I",
    "vendeur": "K",
    "lib_statut": "M",
    "commentaire": "N",
}


# Mapping colonnes par defaut pour 'RUN IAG' (cf. onglet 'Import RUN')
COLS_RUN_IAG = {
    "num_contrat": "A",
    "date_signature": "B",
    "client_nom": "E",
    "client_prenom": "F",
    "vendeur": "K",
    "commentaire": "N",
}


class ImportIagParams(BaseModel):
    type_import: int                # 1 = Base Journ, 2 = RUN
    simulation: bool = True
    format_vendeur: str = "prenom_nom"   # 'prenom_nom' ou 'nom_prenom'
    periode1_du: str = ""
    periode1_au: str = ""
    periode1_mois_paiement: str = ""
    periode2_du: str = ""
    periode2_au: str = ""
    periode2_mois_paiement: str = ""
    mois_paiement_distrib: str = ""


class ImportIagResume(BaseModel):
    nb_fichiers: int = 0
    nb_ajoutes: int = 0
    nb_valides: int = 0
    nb_resilies: int = 0
    nb_decommissions: int = 0
    nb_deja_saisis: int = 0           # type 1 : "deja saisi"
    nb_deja_statues: int = 0          # type 2 : "deja statue"
    nb_introuvables: int = 0
    nb_doublons: int = 0
    nb_pb_vendeur: int = 0
    nb_erreurs: int = 0


class ImportIagResult(BaseModel):
    ok: bool
    type_import: int
    type_label: str
    simulation: bool
    resume: ImportIagResume
    fichiers_traites: list[str] = []
    contrats_ajoutes: list[dict] = []
    contrats_modifies: list[dict] = []         # 'deja saisis' / 'deja statues'
    contrats_non_trouves: list[dict] = []
    contrats_run: list[dict] = []
    pb_vendeur: list[dict] = []                # = onglet 'Erreurs' si type 2
    message: str = ""
    xlsx_b64: str = ""
    xlsx_name: str = ""
    mail_envoye: bool = False


TYPE_LABELS = {
    1: "Base Journalière",
    2: "RUN",
}


def _id_produit_iag(produit_str: str, num_contrat: str) -> tuple[int, str]:
    """Determine (id_produit, lib_produit) selon le libelle ASSU
    (cf code WinDev importJournalier).

    - HOME PROTEKT -> 58
    - SAFE         -> 64 (KSM), 68 si num_contrat commence par >=30 (TABLETTE)
    - KSM          -> 63
    - PROTECT ELEC / GO PROTEKT -> 77 (MAX ELEC)
    - ELECTRO      -> 113
    - default      -> 76 (MAX ASSURANCE)
    """
    p = (produit_str or "").upper()
    if "HOME PROTEKT" in p:
        return (58, "HOME PROTEKT")
    if "SAFE" in p:
        # Si num_contrat commence par >= 30 -> TABLETTE
        try:
            if int((num_contrat or "")[:2] or 0) >= 30:
                return (68, "SAFE ASSURANCE KSM - TABLETTE")
        except ValueError:
            pass
        return (64, "SAFE ASSURANCE KSM")
    if "KSM" in p:
        return (63, "KSM Coup de main")
    if "PROTECT ELEC" in p or "GO PROTEKT" in p:
        return (77, "MAX ELEC")
    if "ELECTRO" in p:
        return (113, "Electro Assist")
    return (76, "MAX ASSURANCE")


def _etat_contrat_iag(lib_statut: str) -> int:
    """37 (En cours) / 38 (Doublon) / 66 (Refus)."""
    s = (lib_statut or "").lower()
    if "doublon" in s:
        return 38
    if "refus" in s:
        return 66
    return 37


def _lookup_vendeur_iag(nom_complet: str, format_v: str) -> int:
    """Cherche un vendeur par concat selon le format :
      - 'prenom_nom' : NOM%PRENOM (concat) OU prenom_v + nom_v split
      - 'nom_prenom' : NOM%PRENOM en sens inverse
    Retourne id_salarie ou 0 si introuvable / ambigu."""
    if not nom_complet or not nom_complet.strip():
        return 0
    db = get_pg_connection("rh")

    s = nom_complet.upper()
    for c in ("-", "'", " "):
        s = s.replace(c, "%")
    s = s.replace("%%", "%")
    pattern = f"%{s}%"

    if format_v == "nom_prenom":
        sql = """SELECT id_salarie FROM rh.pgt_salarie
                  WHERE UPPER(CONCAT(nom, '%', prenom)) LIKE ?
                    AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                  LIMIT 2"""
    else:  # prenom_nom
        sql = """SELECT id_salarie FROM rh.pgt_salarie
                  WHERE UPPER(CONCAT(prenom, '%', nom)) LIKE ?
                    AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                  LIMIT 2"""

    rows = db.query(sql, (pattern,)) or []
    if len(rows) == 1:
        return int(rows[0]["id_salarie"])

    # Fallback split
    parts = nom_complet.strip().split()
    if len(parts) >= 2:
        a = parts[0].upper(); b = parts[-1].upper()
        if format_v == "nom_prenom":
            nom, prenom = a, b
        else:
            prenom, nom = a, b
        rows = db.query(
            """SELECT id_salarie FROM rh.pgt_salarie
                WHERE UPPER(nom) LIKE ? AND UPPER(prenom) LIKE ?
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 2""",
            (f"{nom}%", f"{prenom}%"),
        ) or []
        if len(rows) == 1:
            return int(rows[0]["id_salarie"])
    return 0


def _info_salarie_iag(id_sal: int) -> dict:
    if not id_sal:
        return {}
    db = get_pg_connection("rh")
    r = db.query_one(
        """SELECT s.id_salarie, s.nom, s.prenom, e.id_ste
             FROM rh.pgt_salarie s
             LEFT JOIN rh.pgt_salarie_embauche e ON e.id_salarie = s.id_salarie
            WHERE s.id_salarie = ? LIMIT 1""",
        (int(id_sal),),
    )
    return r or {}


def _calcul_points_iag(id_produit: int, date_sign: Optional[date]) -> float:
    """Calcule nb_points IAG en deleguant a calcul_point_contrat (transposition
    fidele WinDev calculPointContrat).

    Lookup famille/sous_fam dans pgt_iag_produit puis appelle le bareme
    central pgt_bareme_point via calcul_point_contrat.
    """
    if not id_produit:
        return 0.0
    db = get_pg_connection("adv")
    prod = db.query_one(
        """SELECT famille, sous_fam FROM adv.pgt_iag_produit
            WHERE id_produit = ? LIMIT 1""",
        (int(id_produit),),
    )
    if not prod:
        return 0.0
    from app.shared.sdtc.bareme import calcul_point_contrat
    return calcul_point_contrat(
        fam=prod.get("famille") or "",
        ss_fam=prod.get("sous_fam") or "",
        palier=0,
        date_sign=str(date_sign) if date_sign else "",
    )


def _create_iag_contrat(td: dict, op_id: int) -> int:
    """Cree un iag_contrat. Retourne id_contrat cree."""
    db = get_pg_connection("adv")
    id_contrat = _new_id()
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_contrat_auto), 0) + 1 AS n FROM adv.pgt_iag_contrat"
    )
    auto_n = int(auto["n"]) if auto else 1

    date_sign = td.get("date_signature")
    if isinstance(date_sign, str):
        date_sign = _parse_date_fr(date_sign)

    db.query(
        """INSERT INTO adv.pgt_iag_contrat
              (id_contrat_auto, id_contrat, id_client, id_salarie, id_ste,
               num_bs, id_produit, id_etat_contrat,
               date_signature, op_saisie, date_saisie, non_call,
               nb_points, code_enr, info_interne,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?,
                   ?, ?, NOW(), ?, ?, '', ?,
                   ?, NOW(), 'new')""",
        (auto_n, id_contrat,
         int(td.get("id_client") or 0),
         int(td.get("id_salarie") or 0),
         int(td.get("id_ste") or 0),
         td.get("num_bs") or "",
         int(td.get("id_produit") or 0),
         int(td.get("etat_contrat") or 37),
         date_sign, int(op_id),
         bool(td.get("non_call", True)),
         float(td.get("nb_points") or 0),
         td.get("commentaire") or "",
         int(op_id)),
    )
    return id_contrat


def _import_journalier_iag(
    p: ImportIagParams, fname: str, content: bytes, op_id: int,
    ajoutes: list[dict], modifies: list[dict], pb_vendeur: list[dict],
    resume: ImportIagResume,
) -> None:
    """Procedure pour un fichier Base Journaliere IAG."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        modifies.append({"_erreur": f"Lecture {fname} : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_BJ_IAG.items()}
    db = get_pg_connection("adv")

    for i in range(2, (ws.max_row or 0) + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"]).upper()
        if not num_contrat:
            continue
        date_sign_s = _cell(ws, i, cols["date_signature"])
        date_sign = _parse_date_fr(date_sign_s)
        vendeur_cell = _cell(ws, i, cols["vendeur"])
        comment = _cell(ws, i, cols["commentaire"])
        lib_statut = _cell(ws, i, cols["lib_statut"])
        produit_assu = _cell(ws, i, cols["produit_assu"])
        client = {
            "nom": _cell(ws, i, cols["client_nom"]),
            "prenom": _cell(ws, i, cols["client_prenom"]),
            "adresse": _cell(ws, i, cols["client_adresse"]),
            "cp": _cell(ws, i, cols["client_cp"]),
            "ville": _cell(ws, i, cols["client_ville"]),
        }

        id_prod_assu, lib_prod = _id_produit_iag(produit_assu, num_contrat)
        etat_initial = _etat_contrat_iag(lib_statut)

        # Lookup vendeur : d'abord TK_Call par num_contrat (cf. WinDev
        # ReqTkCall_ByNumCtt(NumContrat, "IAG")), sinon par nom.
        id_vendeur = 0
        non_call = True
        tk_client: dict = {}
        try:
            db_bo = get_pg_connection("ticket_bo")
            tk = db_bo.query_one(
                """SELECT id_salarie, nom_client, nom_marital_client,
                          prenom_client, adresse1, adresse2, cp, ville,
                          mobile1, date_naiss, adr_mail
                     FROM ticket_bo.pgt_tk_call
                    WHERE UPPER(num_bs) = UPPER(?)
                      AND UPPER(partenaire) = 'IAG'
                      AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                    LIMIT 1""",
                (num_contrat,),
            )
            if tk:
                id_vendeur = int(tk.get("id_salarie") or 0)
                if id_vendeur:
                    non_call = False
                nom_cli = tk.get("nom_client") or ""
                if tk.get("nom_marital_client"):
                    nom_cli += f" ep {tk['nom_marital_client']}"
                tk_client = {
                    "nom": nom_cli,
                    "prenom": tk.get("prenom_client") or "",
                    "adresse": tk.get("adresse1") or "",
                    "cplt": tk.get("adresse2") or "",
                    "cp": tk.get("cp") or "",
                    "ville": tk.get("ville") or "",
                    "gsm": tk.get("mobile1") or "",
                    "date_naiss": tk.get("date_naiss"),
                    "mail": tk.get("adr_mail") or "",
                }
        except Exception:
            pass  # tolere si base BO indispo / table absente

        # Enrichit le client avec les infos TkCall si trouvees
        if tk_client:
            for k, v in tk_client.items():
                if v and k in client:
                    client[k] = v
                elif v:
                    client[k] = v

        if id_vendeur == 0:
            id_vendeur = _lookup_vendeur_iag(vendeur_cell, p.format_vendeur)

        # Lookup contrat
        ctt = db.query_one(
            """SELECT id_contrat, id_salarie, id_client, num_bs,
                      id_etat_contrat, date_signature, id_produit
                 FROM adv.pgt_iag_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""",
            (num_contrat,),
        )

        info_sal = _info_salarie_iag(id_vendeur) if id_vendeur else {}
        vendeur_libelle = (
            f"{_str(info_sal.get('nom')).upper()} "
            f"{_str(info_sal.get('prenom')).title()}".strip()
            if info_sal else (vendeur_cell or "-Vide-")
        )
        id_ste = int(info_sal.get("id_ste") or 0)

        if not ctt:
            # ---- A AJOUTER ----
            ajoutes.append({
                "NumCtt": num_contrat,
                "DateSigne": date_sign_s,
                "Vendeur": vendeur_libelle,
                "IdSalarie": id_vendeur,
                "Société": id_ste,
                "LibProduit": id_prod_assu,
                "LibProduitName": lib_prod,
                "EtatContrat": etat_initial,
                "CltNom": client["nom"],
                "CltPrenom": client["prenom"],
                "CltCP": client["cp"],
                "CltVille": client["ville"],
                "_payload_create": {
                    "num_bs": num_contrat,
                    "id_salarie": id_vendeur,
                    "id_ste": id_ste,
                    "id_produit": id_prod_assu,
                    "etat_contrat": etat_initial,
                    "date_signature": date_sign,
                    "non_call": non_call,
                    "commentaire": comment,
                    "_client": {
                        "nom": client["nom"], "prenom": client["prenom"],
                        "adresse": client["adresse"], "cp": client["cp"],
                        "ville": client["ville"],
                    },
                },
            })
            if id_vendeur == 0:
                pb_vendeur.append({
                    "NumCtt": num_contrat,
                    "DateSigne": date_sign_s,
                    "Vendeur Import": vendeur_cell,
                    "Erreur": "Vendeur inconnu",
                    "LibProduit": id_prod_assu,
                    "EtatContrat": etat_initial,
                })
                resume.nb_pb_vendeur += 1
            resume.nb_ajoutes += 1
        else:
            # ---- DEJA EXISTE ----
            id_contrat = int(ctt["id_contrat"])
            id_sal_db = int(ctt.get("id_salarie") or 0)
            modifies.append({
                "NumCtt": ctt.get("num_bs"),
                "DateSigne": str(ctt.get("date_signature") or ""),
                "IdSalarie DB": id_sal_db,
                "LibProduit": int(ctt.get("id_produit") or 0),
                "EtatContrat": int(ctt.get("id_etat_contrat") or 0),
                "Vendeur Import": vendeur_cell,
                "CltNom": client["nom"],
                "CltPrenom": client["prenom"],
            })
            resume.nb_deja_saisis += 1
            # Vendeur different + un seul match -> reattribution
            # (cf. WinDev : UPDATE IAG_contrat SET IDSalarie = ...)
            if id_sal_db != id_vendeur and id_vendeur != 0:
                pb_vendeur.append({
                    "NumCtt": ctt.get("num_bs"),
                    "DateSigne": str(ctt.get("date_signature") or ""),
                    "Vendeur Import": vendeur_cell,
                    "Erreur": "vendeur réattribué",
                    "OldIdSalarie": id_sal_db,
                    "NewIdSalarie": id_vendeur,
                    "_id_contrat": id_contrat,
                })
                if not p.simulation:
                    db.query(
                        """UPDATE adv.pgt_iag_contrat
                              SET id_salarie = ?, modif_op = ?,
                                  modif_date = NOW(), modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (id_vendeur, int(op_id), id_contrat),
                    )

    wb.close()


def _ajoute_histo_iag_etat(id_contrat: int, old_etat: int, new_etat: int,
                            date_paiement: str, op_id: int) -> None:
    """Historise un changement d'etat IAG."""
    if not id_contrat:
        return
    db = get_pg_connection("adv")
    auto = db.query_one(
        "SELECT COALESCE(MAX(id_histo_auto), 0) + 1 AS n FROM adv.pgt_iag_histo_etat_ctt"
    )
    db.query(
        """INSERT INTO adv.pgt_iag_histo_etat_ctt
              (id_histo_auto, id_histo, id_contrat, op_saisie, date,
               old_etat, new_etat, date_paiement,
               modif_op, modif_date, modif_elem)
           VALUES (?, ?, ?, ?, NOW(), ?, ?, ?, ?, NOW(), 'new')""",
        (int(auto["n"]) if auto else 1, _new_id(),
         int(id_contrat), int(op_id),
         int(old_etat) if old_etat else 0,
         int(new_etat) if new_etat else 0,
         date_paiement or "", int(op_id)),
    )


def _detect_periode_iag(
    date_sign: Optional[date], is_distrib: bool,
    p1_du: date, p1_au: date, mp1: Optional[date],
    p2_du: date, p2_au: date, mp2: Optional[date],
    mp_distrib: Optional[date],
) -> tuple[Optional[date], str]:
    """Determine (mois_paiement, libelle_periode) selon les dates."""
    if not date_sign:
        return (None, "")
    if is_distrib:
        return (mp_distrib, "Distrib")
    if p1_du <= date_sign <= p1_au:
        return (mp1, "Période 1")
    if p2_du <= date_sign <= p2_au:
        return (mp2, "Période 2")
    # Periode -1 mois (fallback)
    import datetime as _dt
    p1_du_m1 = (p1_du.replace(month=p1_du.month - 1) if p1_du.month > 1
                else p1_du.replace(year=p1_du.year - 1, month=12))
    p1_au_m1 = (p1_au.replace(month=p1_au.month - 1) if p1_au.month > 1
                else p1_au.replace(year=p1_au.year - 1, month=12))
    if p1_du_m1 <= date_sign <= p1_au_m1:
        return (mp1, "Période -1 mois")
    return (None, "HORS_DELAI")


def _affectation_iag(id_salarie: int) -> tuple[str, str, bool]:
    """Retourne (agence, equipe, is_distrib). Simplification : derniere
    affectation connue (pas de versioning historique date)."""
    if not id_salarie:
        return ("", "", False)
    db = get_pg_connection("rh")
    rows = db.query(
        """SELECT o.lib_orga, o.id_type_niveau_orga, o.id_type_orga
             FROM rh.pgt_salarie_organigramme so
             JOIN rh.pgt_organigramme o ON o.idorganigramme = so.idorganigramme
            WHERE so.id_salarie = ?
              AND (so.modif_elem IS NULL OR so.modif_elem NOT LIKE '%suppr%')""",
        (int(id_salarie),),
    ) or []
    agence = ""; equipe = ""; is_distrib = False
    for r in rows:
        lvl = r.get("id_type_niveau_orga")
        lib = r.get("lib_orga") or ""
        if lvl == 3 and not agence:
            agence = lib
            if int(r.get("id_type_orga") or 0) == 3:
                is_distrib = True
        elif lvl == 4 and not equipe:
            equipe = lib
    return (agence, equipe, is_distrib)


def _import_run_iag(
    p: ImportIagParams, fname: str, content: bytes, op_id: int,
    mode: str,  # 'valide' ou 'resil'
    runs: list[dict], modifies: list[dict], non_trouves: list[dict],
    pb_vendeur: list[dict], resume: ImportIagResume,
) -> None:
    """Logique commune importValide + importResil IAG."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        resume.nb_erreurs += 1
        modifies.append({"_erreur": f"Lecture {fname} : {e}"})
        return
    ws = wb.active
    cols = {k: _col_letter_to_index(v) for k, v in COLS_RUN_IAG.items()}

    p1_du = _parse_date_fr(p.periode1_du) or date(1900, 1, 1)
    p1_au = _parse_date_fr(p.periode1_au) or date(2100, 12, 31)
    p2_du = _parse_date_fr(p.periode2_du) or date(1900, 1, 1)
    p2_au = _parse_date_fr(p.periode2_au) or date(2100, 12, 31)
    mp1 = _dernier_jour_mois(p.periode1_mois_paiement)
    mp2 = _dernier_jour_mois(p.periode2_mois_paiement)
    mp_distrib = _dernier_jour_mois(p.mois_paiement_distrib)

    db = get_pg_connection("adv")

    for i in range(2, (ws.max_row or 0) + 1):
        num_contrat = _cell(ws, i, cols["num_contrat"]).upper()
        if not num_contrat:
            continue
        vendeur_cell = _cell(ws, i, cols["vendeur"])
        date_sign_s = _cell(ws, i, cols["date_signature"])
        comment = _cell(ws, i, cols["commentaire"])
        client = {
            "nom": _cell(ws, i, cols["client_nom"]),
            "prenom": _cell(ws, i, cols["client_prenom"]),
        }

        # Lookup contrat(s)
        rows_ctt = db.query(
            """SELECT id_contrat, id_salarie, id_client, num_bs,
                      id_etat_contrat, date_signature, mois_p, id_produit
                 FROM adv.pgt_iag_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
            (num_contrat,),
        ) or []

        if len(rows_ctt) == 0:
            non_trouves.append({
                "NumCtt": num_contrat,
                "DateSign": date_sign_s,
                "Vendeur": vendeur_cell,
                "CltNom": client["nom"],
                "CltPrenom": client["prenom"],
                "Statut": "VALID" if mode == "valide" else "RESIL",
            })
            resume.nb_introuvables += 1
            continue
        if len(rows_ctt) > 1:
            resume.nb_doublons += 1
            for r in rows_ctt:
                pb_vendeur.append({
                    "NumCtt": num_contrat,
                    "id_contrat": int(r.get("id_contrat") or 0),
                    "DateSigne": str(r.get("date_signature") or ""),
                    "Erreur": "DOUBLON - " + ("Valid" if mode == "valide" else "Résil"),
                })
            continue

        r = rows_ctt[0]
        id_contrat = int(r["id_contrat"])
        id_sal_db = int(r.get("id_salarie") or 0)
        etat_actuel = int(r.get("id_etat_contrat") or 0)
        mois_p_omaya = r.get("mois_p")
        date_sign_omaya = r.get("date_signature")

        # Affectation + periode
        agence, equipe, is_distrib = _affectation_iag(id_sal_db)
        mois_p, periode_lbl = _detect_periode_iag(
            date_sign_omaya, is_distrib,
            p1_du, p1_au, mp1, p2_du, p2_au, mp2, mp_distrib,
        )
        if periode_lbl == "HORS_DELAI":
            resume.nb_pb_vendeur += 1
            pb_vendeur.append({
                "NumCtt": num_contrat,
                "DateSigne": str(date_sign_omaya or ""),
                "Erreur": "Hors Délai - " + ("Valid" if mode == "valide" else "Résil"),
                "Agence": agence, "Equipe": equipe,
                "_id_contrat": id_contrat,
            })

        # Lookup type_etat
        etat_info = db.query_one(
            """SELECT id_type_etat, lib_etat FROM adv.pgt_iag_etat_contrat
                WHERE id_etat = ? LIMIT 1""",
            (etat_actuel,),
        )
        id_type_etat = int(etat_info.get("id_type_etat") or 0) if etat_info else 0
        lib_etat_actuel = (etat_info.get("lib_etat") or "") if etat_info else ""

        # ---- Determination du traitement (cf WinDev) ----
        info_sal = _info_salarie_iag(id_sal_db)
        nom_vend = (f"{_str(info_sal.get('nom'))} "
                    f"{_str(info_sal.get('prenom')).title()}").strip()

        nouvel_etat = etat_actuel
        nouveau_mois_p: Optional[date] = None
        traitement = "deja_statue"

        if mode == "valide":
            # Eligible si type_etat in (1,2) OU etat in (29,30)
            if id_type_etat in (1, 2) or etat_actuel in (29, 30):
                nouvel_etat = 19  # Valide - Paye par l'operateur
                # Si etat=29 ou 30 -> garde mois_p ancien
                if etat_actuel in (29, 30):
                    nouveau_mois_p = mois_p_omaya
                else:
                    nouveau_mois_p = mois_p
                traitement = "valide"
                resume.nb_valides += 1
            else:
                resume.nb_deja_statues += 1
        else:  # mode == 'resil'
            if id_type_etat == 5:  # Paye -> decommissionne
                nouvel_etat = 20
                nouveau_mois_p = mois_p
                traitement = "decomm"
                resume.nb_decommissions += 1
            elif id_type_etat in (1, 2):  # En attente -> resilie
                nouvel_etat = 16
                nouveau_mois_p = None  # mois_p = ""
                traitement = "resilie"
                resume.nb_resilies += 1
            else:
                resume.nb_deja_statues += 1

        # Snapshot a afficher
        row_snap = {
            "NumCtt": num_contrat,
            "DateSigne": str(date_sign_omaya or ""),
            "Vendeur": nom_vend,
            "Société": int(r.get("id_produit") or 0),
            "Agence": agence,
            "Equipe": equipe,
            "Periode": periode_lbl,
            "Etat actuel": etat_actuel,
            "Lib Etat": lib_etat_actuel,
            "Nouvel etat": nouvel_etat,
            "Nouveau MoisP": str(nouveau_mois_p) if nouveau_mois_p else "",
            "Traitement": traitement,
        }

        if traitement == "deja_statue":
            modifies.append(row_snap)
        else:
            runs.append(row_snap)

        # -- Action prod : UPDATE + histo --
        if not p.simulation and traitement != "deja_statue":
            old_mois_p_str = str(mois_p_omaya) if mois_p_omaya else ""
            try:
                if mode == "valide":
                    db.query(
                        """UPDATE adv.pgt_iag_contrat
                              SET id_etat_contrat = ?, mois_p = ?,
                                  modif_date = NOW(), modif_op = ?,
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (nouvel_etat, nouveau_mois_p, int(op_id), id_contrat),
                    )
                elif mode == "resil":
                    if traitement == "decomm":
                        db.query(
                            """UPDATE adv.pgt_iag_contrat
                                  SET id_etat_contrat = ?, mois_p = ?,
                                      info_interne = COALESCE(info_interne, '') || '\n' || ?,
                                      modif_date = NOW(), modif_op = ?,
                                      modif_elem = 'modif'
                                WHERE id_contrat = ?""",
                            (nouvel_etat, nouveau_mois_p, comment,
                             int(op_id), id_contrat),
                        )
                    else:  # resilie
                        db.query(
                            """UPDATE adv.pgt_iag_contrat
                                  SET id_etat_contrat = ?, mois_p = NULL,
                                      info_interne = COALESCE(info_interne, '') ||
                                                     '\nMotif Résil : ' || ?,
                                      modif_date = NOW(), modif_op = ?,
                                      modif_elem = 'modif'
                                WHERE id_contrat = ?""",
                            (nouvel_etat, comment, int(op_id), id_contrat),
                        )
                _ajoute_histo_iag_etat(
                    id_contrat, etat_actuel, nouvel_etat,
                    old_mois_p_str, op_id,
                )
            except Exception as e:
                row_snap["Erreur"] = str(e)

    wb.close()


def run_import_iag(
    p: ImportIagParams, files: list[tuple[str, bytes]], op_id: int,
) -> ImportIagResult:
    """Dispatcher principal."""
    label = TYPE_LABELS.get(p.type_import, "?")
    if not files:
        return ImportIagResult(
            ok=False, type_import=p.type_import, type_label=label,
            simulation=p.simulation, resume=ImportIagResume(),
            message="Aucun fichier fourni.",
        )

    resume = ImportIagResume(nb_fichiers=len(files))
    ajoutes: list[dict] = []
    modifies: list[dict] = []
    non_trouves: list[dict] = []
    runs: list[dict] = []
    pb_vendeur: list[dict] = []
    fichiers_traites: list[str] = []

    for fname, content in files:
        fichiers_traites.append(fname)
        if p.type_import == 1:
            _import_journalier_iag(
                p, fname, content, op_id,
                ajoutes, modifies, pb_vendeur, resume,
            )
        elif p.type_import == 2:
            # Dispatch valide/resil selon 'ANNUL' dans le nom de fichier
            mode = "resil" if "ANNUL" in fname.upper() else "valide"
            _import_run_iag(
                p, fname, content, op_id, mode,
                runs, modifies, non_trouves, pb_vendeur, resume,
            )

    # -- PASSE PROD : creation contrat + reattribution vendeur (si pas simu) --
    nb_crees = 0
    nb_reattrib = 0
    if not p.simulation:
        db = get_pg_connection("adv")
        # Creation des nouveaux contrats
        for row in ajoutes:
            pl = row.pop("_payload_create", None)
            if not pl or not pl.get("id_salarie"):
                continue  # pas de vendeur trouve -> pas de creation auto
            try:
                cli = pl.pop("_client", {})
                pl["id_client"] = _lookup_or_create_client(cli, op_id)
                pl["nb_points"] = _calcul_points_iag(
                    pl["id_produit"], pl.get("date_signature"),
                )
                new_id = _create_iag_contrat(pl, op_id)
                row["IdContratCree"] = new_id
                nb_crees += 1
            except Exception as e:
                row["Erreur"] = str(e)
        # Reattribution vendeur (lignes pb_vendeur 'vendeur reattribue')
        for row in pb_vendeur:
            id_ct = row.pop("_id_contrat", None)
            if not id_ct or row.get("Erreur") != "vendeur réattribué":
                continue
            try:
                db.query(
                    """UPDATE adv.pgt_iag_contrat
                          SET id_salarie = ?,
                              modif_op = ?, modif_date = NOW(),
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (int(row["NewIdSalarie"]), int(op_id), int(id_ct)),
                )
                nb_reattrib += 1
            except Exception as e:
                row["ErreurMaj"] = str(e)

    # Nettoyer les payloads internes restants (ne pas exposer en JSON)
    for row in ajoutes:
        row.pop("_payload_create", None)
    for row in pb_vendeur:
        row.pop("_id_contrat", None)

    res = ImportIagResult(
        ok=True, type_import=p.type_import, type_label=label,
        simulation=p.simulation, resume=resume,
        fichiers_traites=fichiers_traites,
        contrats_ajoutes=ajoutes,
        contrats_modifies=modifies,
        contrats_non_trouves=non_trouves,
        contrats_run=runs,
        pb_vendeur=pb_vendeur,
        message=(
            f"{len(files)} fichier(s) traité(s) | "
            f"À ajouter : {resume.nb_ajoutes} | Déjà saisis : {resume.nb_deja_saisis} | "
            f"Pb vendeur : {resume.nb_pb_vendeur}. "
            + (f"PRODUCTION : {nb_crees} créés, {nb_reattrib} vendeurs réattribués."
               if not p.simulation else "(SIMULATION)")
        ),
    )
    _attach_xlsx_and_mail_iag(res, op_id)
    return res


# ---------------------------------------------------------------------------
# XLSX + mail BO (reuse pattern ENI)
# ---------------------------------------------------------------------------


def _build_xlsx_iag(res: ImportIagResult) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Résumé"
    header_fill = PatternFill("solid", fgColor="17494E")
    header_font = Font(bold=True, color="FFFFFF")

    items = [
        ("NB Fichiers", res.resume.nb_fichiers),
        ("NB Ajoutés", res.resume.nb_ajoutes),
        ("NB Validés", res.resume.nb_valides),
        ("NB Résiliés", res.resume.nb_resilies),
        ("NB Déjà saisis", res.resume.nb_deja_saisis),
        ("NB Déjà statués", res.resume.nb_deja_statues),
        ("NB Introuvables", res.resume.nb_introuvables),
        ("NB Pb Vendeur", res.resume.nb_pb_vendeur),
        ("NB Erreurs", res.resume.nb_erreurs),
    ]
    ws.append(["Indicateur", "Nombre"])
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for lbl, n in items:
        ws.append([lbl, n])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    sections: list[tuple[str, list[dict]]] = [
        ("Contrats ajoutés", res.contrats_ajoutes),
        ("Contrats déjà saisis", res.contrats_modifies),
        ("Contrats non trouvés", res.contrats_non_trouves),
        ("Contrats RUN", res.contrats_run),
        ("Erreurs Vendeurs", res.pb_vendeur),
    ]
    for title, rows in sections:
        if not rows:
            continue
        sh = wb.create_sheet(title=title[:31])
        keys = list(rows[0].keys())
        sh.append(keys)
        for c in sh[1]:
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for r in rows:
            sh.append([str(r.get(k, "")) if r.get(k) is not None else "" for k in keys])
        for i, k in enumerate(keys, start=1):
            col = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
            sh.column_dimensions[col].width = max(12, min(40, len(k) + 4))

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


def _attach_xlsx_and_mail_iag(res: ImportIagResult, op_id: int) -> None:
    import base64
    from app.shared.notifications.mail import envoi_mail

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    suffix = "_SIMU" if res.simulation else ""
    prefix = "ImportJournalierASSU" if res.type_import == 1 else "ImportRunIAG"
    xlsx_name = f"{prefix}_{ts}{suffix}.xlsx"
    try:
        xlsx_bytes = _build_xlsx_iag(res)
    except Exception:
        return
    res.xlsx_name = xlsx_name
    res.xlsx_b64 = base64.b64encode(xlsx_bytes).decode("ascii")

    # Mail BO
    db = get_pg_connection("rh")
    r = db.query_one(
        "SELECT mail FROM rh.pgt_salarie_coordonnees WHERE id_salarie = ? LIMIT 1",
        (int(op_id),),
    )
    op_mail = (r.get("mail") if r else "") or ""
    destinataires = [op_mail] if op_mail else ["intranet@omaya.fr"]
    cc = ["intranet@omaya.fr"] if op_mail and op_mail != "intranet@omaya.fr" else []

    sujet_pref = "SIMULATION : " if res.simulation else ""
    sujet = f"{sujet_pref}Importation {res.type_label} IAG du {date.today().strftime('%d/%m/%Y')}"
    html = (
        "<p>Bonjour,</p>"
        f"<p>Fin importation le : {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>"
        f"<p><strong>{res.message}</strong></p>"
        f"<ul>"
        f"<li>NB Ajout(s) : {res.resume.nb_ajoutes}</li>"
        f"<li>NB Déjà Saisi(s) : {res.resume.nb_deja_saisis}</li>"
        f"<li>NB Erreur(s) Vendeur : {res.resume.nb_pb_vendeur}</li>"
        f"</ul>"
        "<p>Service Importation EXOSPHERE</p>"
    )
    try:
        res.mail_envoye = envoi_mail(
            sujet=sujet, html=html,
            destinataires=destinataires, cc=cc,
            expediteur="intranet@omaya.fr",
            attachments=[(xlsx_name, xlsx_bytes)],
        )
    except Exception:
        res.mail_envoye = False
