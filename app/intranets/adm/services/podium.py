"""
Service Fen_GestionPodium - Gestion des Podiums.

Ce fichier contient :
  - Combos (types podium actifs, distributeurs)
  - CRUD PodiumType et PodiumTypePart (onglet Parametres)
  - Valider annee (onglet Annee Podium)
  - Recherche podium vendeurs + score visible + telecharger (onglet 1)
  - Calcul podium (proc Podium_Calcul dans podium_calcul.py)
"""
from __future__ import annotations

import io as _io
import logging
from datetime import date, datetime, timedelta
from typing import Optional

from app.core.database.pg import get_pg_connection
from app.intranets.adm.schemas.podium import (
    CalculPodiumParams, CalculPodiumResult, ComboItem, PodiumType,
    PodiumTypePart, PodiumTypePartPayload, PodiumTypePayload,
    RechercherPodiumParams, RechercherPodiumResult, SauveScoreVisibleParams,
    TelechargerParams, ValiderAnneeParams, ValiderAnneeResult,
    VendeurPodiumRow,
)

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------
# Helpers communs
# --------------------------------------------------------------------

def _clean_id(v) -> str:
    if v is None:
        return ""
    try:
        n = int(v)
        return str(n) if n else ""
    except (TypeError, ValueError):
        return ""


def _iso_date(v) -> str:
    if v is None:
        return ""
    s = str(v)[:10]
    if s.startswith("1900") or s.startswith("0000"):
        return ""
    return s


def _cap_prenom(p: str) -> str:
    if not p:
        return ""
    return p[0].upper() + p[1:].lower() if len(p) > 1 else p.upper()


# --------------------------------------------------------------------
# Combos (Types podium, Distributeurs)
# --------------------------------------------------------------------

def list_types_podium_actifs() -> list[ComboItem]:
    """Combo 'Type Podium' onglet 1.
    Cf. WinDev : PodiumType actifs, tri par ordre_affichage.
    """
    db = get_pg_connection("rh")
    try:
        rows = db.query(
            """SELECT id_podium_type, lib_podium_type
                 FROM divers.pgt_podium_type
                WHERE is_actif = TRUE
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                ORDER BY ordre_affichage ASC NULLS LAST,
                         lib_podium_type ASC""",
        ) or []
    except Exception:
        logger.exception("list_types_podium_actifs")
        return []
    return [
        ComboItem(
            id=_clean_id(r.get("id_podium_type")),
            lib=(r.get("lib_podium_type") or "").strip(),
        )
        for r in rows
    ]


def list_distributeurs() -> list[ComboItem]:
    """Combo 'Distrib' onglet 1. Cf. WinDev ReqListeDitrib :
    organigramme dont le parent est un enfant de racine (id_ste=4),
    en excluant l'orga 20160729152638792.
    """
    rh = get_pg_connection("rh")
    try:
        rows = rh.query(
            """SELECT o.idorganigramme, o.lib_orga
                 FROM pgt_organigramme o
                 JOIN pgt_organigramme p
                      ON p.idorganigramme = o.id_parent
                WHERE p.id_parent = 0
                  AND p.idorganigramme <> 20160729152638792
                  AND p.id_ste = 4
                  AND (o.modif_elem IS NULL
                       OR o.modif_elem NOT LIKE '%suppr%')
                ORDER BY o.lib_orga ASC""",
        ) or []
    except Exception:
        logger.exception("list_distributeurs")
        return []
    return [
        ComboItem(
            id=_clean_id(r.get("idorganigramme")),
            lib=(r.get("lib_orga") or "").strip(),
        )
        for r in rows
    ]


# --------------------------------------------------------------------
# Onglet 2 - CRUD PodiumType (gauche)
# --------------------------------------------------------------------

def list_podium_types() -> list[PodiumType]:
    """Liste tous les PodiumType non supprimes (tri par ordre_affichage)."""
    db = get_pg_connection("rh")
    try:
        rows = db.query(
            """SELECT id_podium_type, lib_podium_type, lib_court,
                      prod_groupe, qualite, espoir, is_actif,
                      ordre_affichage
                 FROM divers.pgt_podium_type
                WHERE (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                ORDER BY ordre_affichage ASC NULLS LAST,
                         lib_podium_type ASC""",
        ) or []
    except Exception:
        logger.exception("list_podium_types")
        return []
    return [
        PodiumType(
            id_podium_type=_clean_id(r.get("id_podium_type")),
            lib_podium_type=(r.get("lib_podium_type") or "").strip(),
            lib_court=(r.get("lib_court") or "").strip(),
            prod_groupe=bool(r.get("prod_groupe")),
            qualite=bool(r.get("qualite")),
            espoir=bool(r.get("espoir")),
            is_actif=bool(r.get("is_actif")),
            ordre_affichage=int(r.get("ordre_affichage") or 0),
        )
        for r in rows
    ]


def create_podium_type(p: PodiumTypePayload, op_id: int) -> str:
    """Cree un PodiumType. Retourne l'id cree en string."""
    db = get_pg_connection("rh")
    r = db.query_one(
        """INSERT INTO divers.pgt_podium_type
              (lib_podium_type, lib_court, prod_groupe, qualite, espoir,
               is_actif, ordre_affichage,
               modif_date, modif_op, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, ?, NOW(), ?, 'new')
           RETURNING id_podium_type""",
        (
            p.lib_podium_type.strip(), p.lib_court.strip(),
            p.prod_groupe, p.qualite, p.espoir,
            p.is_actif, p.ordre_affichage, int(op_id),
        ),
    )
    return _clean_id(r.get("id_podium_type")) if r else ""


def update_podium_type(id_pt: str, p: PodiumTypePayload, op_id: int) -> bool:
    if not id_pt or id_pt == "0":
        return False
    db = get_pg_connection("rh")
    db.execute(
        """UPDATE divers.pgt_podium_type
              SET lib_podium_type = ?, lib_court = ?, prod_groupe = ?,
                  qualite = ?, espoir = ?, is_actif = ?,
                  ordre_affichage = ?,
                  modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
            WHERE id_podium_type = ?""",
        (
            p.lib_podium_type.strip(), p.lib_court.strip(),
            p.prod_groupe, p.qualite, p.espoir, p.is_actif,
            p.ordre_affichage, int(op_id), int(id_pt),
        ),
    )
    return True


def delete_podium_type(id_pt: str, op_id: int) -> bool:
    """Soft delete : marque modif_elem='suppr' (cf. WinDev)."""
    if not id_pt or id_pt == "0":
        return False
    db = get_pg_connection("rh")
    db.execute(
        """UPDATE divers.pgt_podium_type
              SET modif_date = NOW(), modif_op = ?, modif_elem = 'suppr'
            WHERE id_podium_type = ?""",
        (int(op_id), int(id_pt)),
    )
    return True


# --------------------------------------------------------------------
# Onglet 2 - CRUD PodiumTypePart (droite)
# --------------------------------------------------------------------

def list_podium_type_parts(id_podium_type: str) -> list[PodiumTypePart]:
    """Liste les Parts d'un PodiumType donne."""
    if not id_podium_type or id_podium_type == "0":
        return []
    db = get_pg_connection("rh")
    try:
        rows = db.query(
            """SELECT id_podium_type_part, id_podium_type,
                      famille, sous_fam, prefixe_bdd, type_prod,
                      option_vente, jour_cial_deb, jour_cial_fin
                 FROM divers.pgt_podium_type_part
                WHERE id_podium_type = ?
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                ORDER BY prefixe_bdd ASC, famille ASC""",
            (int(id_podium_type),),
        ) or []
    except Exception:
        logger.exception("list_podium_type_parts")
        return []
    return [
        PodiumTypePart(
            id_podium_type_part=_clean_id(r.get("id_podium_type_part")),
            id_podium_type=_clean_id(r.get("id_podium_type")),
            famille=(r.get("famille") or "Tous").strip(),
            sous_fam=(r.get("sous_fam") or "Tous").strip(),
            prefixe_bdd=(r.get("prefixe_bdd") or "").strip(),
            type_prod=(r.get("type_prod") or "").strip(),
            option_vente=(r.get("option_vente") or "").strip(),
            jour_cial_deb=int(r.get("jour_cial_deb") or 1),
            jour_cial_fin=int(r.get("jour_cial_fin") or 31),
        )
        for r in rows
    ]


def create_podium_type_part(p: PodiumTypePartPayload, op_id: int) -> str:
    db = get_pg_connection("rh")
    r = db.query_one(
        """INSERT INTO divers.pgt_podium_type_part
              (id_podium_type, famille, sous_fam, prefixe_bdd, type_prod,
               option_vente, jour_cial_deb, jour_cial_fin,
               modif_date, modif_op, modif_elem)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, NOW(), ?, 'new')
           RETURNING id_podium_type_part""",
        (
            int(p.id_podium_type),
            p.famille.strip() or "Tous",
            p.sous_fam.strip() or "Tous",
            p.prefixe_bdd.strip(),
            p.type_prod.strip(),
            p.option_vente.strip(),
            int(p.jour_cial_deb),
            int(p.jour_cial_fin),
            int(op_id),
        ),
    )
    return _clean_id(r.get("id_podium_type_part")) if r else ""


def update_podium_type_part(
    id_ptp: str, p: PodiumTypePartPayload, op_id: int,
) -> bool:
    if not id_ptp or id_ptp == "0":
        return False
    db = get_pg_connection("rh")
    db.execute(
        """UPDATE divers.pgt_podium_type_part
              SET id_podium_type = ?, famille = ?, sous_fam = ?,
                  prefixe_bdd = ?, type_prod = ?, option_vente = ?,
                  jour_cial_deb = ?, jour_cial_fin = ?,
                  modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
            WHERE id_podium_type_part = ?""",
        (
            int(p.id_podium_type),
            p.famille.strip() or "Tous",
            p.sous_fam.strip() or "Tous",
            p.prefixe_bdd.strip(), p.type_prod.strip(),
            p.option_vente.strip(),
            int(p.jour_cial_deb), int(p.jour_cial_fin),
            int(op_id), int(id_ptp),
        ),
    )
    return True


def delete_podium_type_part(id_ptp: str, op_id: int) -> bool:
    if not id_ptp or id_ptp == "0":
        return False
    db = get_pg_connection("rh")
    db.execute(
        """UPDATE divers.pgt_podium_type_part
              SET modif_date = NOW(), modif_op = ?, modif_elem = 'suppr'
            WHERE id_podium_type_part = ?""",
        (int(op_id), int(id_ptp)),
    )
    return True


# --------------------------------------------------------------------
# Onglet 3 - Valider annee
# --------------------------------------------------------------------

def valider_annee(p: ValiderAnneeParams, op_id: int) -> ValiderAnneeResult:
    """Cf. WinDev 'Valider l'annee' :
    Pour chaque PodiumType non supprime, cree 12 lignes PodiumMois si
    elles n'existent pas encore (score_visible = TRUE par defaut).
    """
    if p.annee < 2020 or p.annee > 2100:
        return ValiderAnneeResult(
            ok=False, message="Annee invalide",
        )
    db = get_pg_connection("rh")
    try:
        types = db.query(
            """SELECT id_podium_type FROM divers.pgt_podium_type
                WHERE (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
        ) or []
    except Exception as e:
        return ValiderAnneeResult(ok=False, message=f"Erreur SQL : {e}")

    nb_crees = 0
    for t in types:
        id_pt = int(t.get("id_podium_type") or 0)
        if not id_pt:
            continue
        for mois in range(1, 13):
            try:
                existing = db.query_one(
                    """SELECT id_podium_mois FROM divers.pgt_podium_mois
                        WHERE id_podium_type = ? AND mois = ? AND annee = ?
                        LIMIT 1""",
                    (id_pt, mois, str(p.annee)),
                )
                if existing:
                    continue
                db.execute(
                    """INSERT INTO divers.pgt_podium_mois
                          (mois, annee, id_podium_type, score_visible,
                           modif_date, modif_op, modif_elem)
                       VALUES (?, ?, ?, TRUE, NOW(), ?, 'new')""",
                    (mois, str(p.annee), id_pt, int(op_id)),
                )
                nb_crees += 1
            except Exception:
                logger.exception(
                    "valider_annee : insert id_pt=%s mois=%s",
                    id_pt, mois,
                )
    return ValiderAnneeResult(
        ok=True,
        nb_crees=nb_crees,
        message=(
            f"{nb_crees} PodiumMois cree(s) pour l'annee {p.annee}"
            if nb_crees else f"Annee {p.annee} deja complete"
        ),
    )


# --------------------------------------------------------------------
# Onglet 1 - Recherche Podium Vendeurs
# --------------------------------------------------------------------

def _last_day_of_month(y: int, m: int) -> date:
    """Retourne le dernier jour du mois y/m."""
    if m == 12:
        return date(y, 12, 31)
    return date(y, m + 1, 1) - timedelta(days=1)


def _sub_months(d: date, n: int) -> date:
    """Retire n mois d'une date, en clampant le jour au dernier possible."""
    y = d.year
    m = d.month - n
    while m <= 0:
        m += 12
        y -= 1
    while m > 12:
        m -= 12
        y += 1
    last_day = _last_day_of_month(y, m).day
    return date(y, m, min(d.day, last_day))


def _find_lib_equipe(id_equipe: int) -> str:
    """Retourne le lib_orga d'une equipe."""
    if not id_equipe:
        return ""
    rh = get_pg_connection("rh")
    try:
        r = rh.query_one(
            "SELECT lib_orga FROM pgt_organigramme WHERE idorganigramme = ?",
            (id_equipe,),
        )
    except Exception:
        return ""
    return (r.get("lib_orga") or "").strip() if r else ""


def _resp_equipe_at(id_orga: int, date_ref: date) -> Optional[dict]:
    """Retourne le salarie responsable d'une orga (id_type_niveau_orga=4)
    a une date donnee. Utilise pgt_salarie_organigramme.
    """
    rh = get_pg_connection("rh")
    try:
        r = rh.query_one(
            """SELECT so.id_salarie, s.nom, s.prenom
                 FROM pgt_salarie_organigramme so
                 JOIN pgt_salarie s ON s.id_salarie = so.id_salarie
                 JOIN pgt_salarie_embauche e
                      ON e.id_salarie = s.id_salarie
                     AND (e.modif_elem IS NULL
                          OR e.modif_elem NOT LIKE '%suppr%')
                WHERE so.idorganigramme = ?
                  AND (so.modif_elem IS NULL
                       OR so.modif_elem NOT LIKE '%suppr%')
                  AND (so.date_debut IS NULL OR so.date_debut <= ?)
                  AND (so.date_fin IS NULL OR so.date_fin >= ?)
                  AND e.resp_equipe = TRUE
                LIMIT 1""",
            (id_orga, date_ref.isoformat(), date_ref.isoformat()),
        )
    except Exception:
        return None
    return r


def _equipe_terrain(id_salarie: int, date_ref: date) -> int:
    """Retourne id_orga terrain (niveau 4 = equipe) d'un salarie a une date."""
    rh = get_pg_connection("rh")
    try:
        r = rh.query_one(
            """SELECT o.idorganigramme
                 FROM pgt_salarie_organigramme so
                 JOIN pgt_organigramme o
                      ON o.idorganigramme = so.idorganigramme
                WHERE so.id_salarie = ?
                  AND o.id_type_niveau_orga = 4
                  AND (so.modif_elem IS NULL
                       OR so.modif_elem NOT LIKE '%suppr%')
                  AND (so.date_debut IS NULL OR so.date_debut <= ?)
                  AND (so.date_fin IS NULL OR so.date_fin >= ?)
                ORDER BY so.date_debut DESC NULLS LAST
                LIMIT 1""",
            (id_salarie, date_ref.isoformat(), date_ref.isoformat()),
        )
    except Exception:
        return 0
    return int((r or {}).get("idorganigramme") or 0)


def _liste_orga_complet(id_racine: int, date_ref: date) -> set[int]:
    """Recupere recursivement tous les orgas descendants d'un noeud
    racine (utilise pour filtrer par distributeur cf. WinDev
    ListeOrgaComplet).
    """
    rh = get_pg_connection("rh")
    result: set[int] = {int(id_racine)}
    to_process = [int(id_racine)]
    while to_process:
        parents = tuple(to_process)
        to_process = []
        try:
            rows = rh.query(
                f"""SELECT idorganigramme FROM pgt_organigramme
                     WHERE id_parent IN ({','.join(['?'] * len(parents))})
                       AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                     """,
                parents,
            ) or []
        except Exception:
            break
        for r in rows:
            oid = int(r.get("idorganigramme") or 0)
            if oid and oid not in result:
                result.add(oid)
                to_process.append(oid)
    return result


def rechercher_podium(
    p: RechercherPodiumParams,
) -> RechercherPodiumResult:
    """Cf. WinDev 'Podiums Vendeurs - Btn Rechercher'.

    1. Lit le PodiumType courant (flags qualite, prod_groupe, espoir)
    2. Verifie l'existence du PodiumMois (score_visible + id_podium_mois)
    3. Pour chaque PodiumTypePart :
       - Calcule dateDeb / dateFin selon jour_cial_deb/fin
       - Si qualite : dateDeb -= 2 mois (espoir salarie)
       - Query pgt_podium_vendeur + pgt_podium_vendeur_part
       - Filtre par distributeur (organigramme) si is_distrib
       - Aggrege par salarie ou par equipe (si prod_groupe)
    4. Applique le calcul du taux si qualite
    5. Trie -Taux, -Valeur
    """
    if not p.id_podium_type or p.id_podium_type == "0":
        return RechercherPodiumResult(
            ok=False, message="PodiumType requis",
        )
    if p.mois < 1 or p.mois > 12:
        return RechercherPodiumResult(ok=False, message="Mois invalide")
    if p.annee < 2020 or p.annee > 2100:
        return RechercherPodiumResult(ok=False, message="Annee invalide")

    db = get_pg_connection("rh")

    # 1. Lit le PodiumType courant
    pt = db.query_one(
        """SELECT prod_groupe, qualite, espoir
             FROM divers.pgt_podium_type
            WHERE id_podium_type = ?""",
        (int(p.id_podium_type),),
    )
    if not pt:
        return RechercherPodiumResult(
            ok=False, message="PodiumType introuvable",
        )
    is_prod_groupe = bool(pt.get("prod_groupe"))
    is_qualite = bool(pt.get("qualite"))
    is_espoir = bool(pt.get("espoir"))

    # 2. Verifie le PodiumMois
    pm = db.query_one(
        """SELECT id_podium_mois, score_visible
             FROM divers.pgt_podium_mois
            WHERE id_podium_type = ? AND mois = ? AND annee = ?
            LIMIT 1""",
        (int(p.id_podium_type), p.mois, str(p.annee)),
    )
    id_podium_mois = _clean_id((pm or {}).get("id_podium_mois"))
    score_visible = bool((pm or {}).get("score_visible"))

    # 3. Liste les PodiumTypePart
    parts = db.query(
        """SELECT id_podium_type_part, type_prod, jour_cial_deb, jour_cial_fin
             FROM divers.pgt_podium_type_part
            WHERE id_podium_type = ?
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
        (int(p.id_podium_type),),
    ) or []

    # Filtre distributeur : precalcule l'ensemble des orgas
    orga_distrib: set[int] = set()
    if p.is_distrib and p.id_distrib and p.id_distrib != "0":
        orga_distrib = _liste_orga_complet(int(p.id_distrib), date.today())

    # Agrege les lignes par salarie ou equipe
    agg: dict[str, VendeurPodiumRow] = {}

    for part in parts:
        id_ptp = int(part.get("id_podium_type_part") or 0)
        type_prod = (part.get("type_prod") or "").strip()
        j_deb = int(part.get("jour_cial_deb") or 1)
        j_fin = int(part.get("jour_cial_fin") or 31)

        # Calcul dateDeb / dateFin
        if j_deb == 1:
            date_deb = date(p.annee, p.mois, 1)
            date_fin = _last_day_of_month(p.annee, p.mois)
        else:
            date_deb = _sub_months(
                date(p.annee, p.mois, min(j_deb, 28)), 1,
            )
            date_fin = date(p.annee, p.mois, min(j_fin, 28))

        # Requete podium vendeur
        # Selection differente selon prod_groupe
        if is_prod_groupe:
            sql = """
                SELECT pv.id_equipe,
                       SUM(pvp.qte_brut) AS qte_brut,
                       SUM(pvp.qte_hors_rejet) AS qte_hors_rejet,
                       SUM(pvp.qte_paye) AS qte_paye
                  FROM divers.pgt_podium_vendeur pv
                  JOIN divers.pgt_podium_vendeur_part pvp
                       ON pvp.id_podium_vendeur = pv.id_podium_vendeur
                 WHERE (pv.modif_elem IS NULL
                        OR pv.modif_elem NOT LIKE '%suppr%')
                   AND (pvp.modif_elem IS NULL
                        OR pvp.modif_elem NOT LIKE '%suppr%')
                   AND pv.date_jour BETWEEN ? AND ?
                   AND pv.id_podium_type = ?
                   AND pv.distributeur = ?
                   AND pvp.id_podium_type_part = ?
                 GROUP BY pv.id_equipe"""
            params = (
                date_deb.isoformat(), date_fin.isoformat(),
                int(p.id_podium_type), p.is_distrib, id_ptp,
            )
        else:
            sql = """
                SELECT pv.id_salarie,
                       s.nom, s.prenom,
                       e.en_activite, e.resp_equipe, e.date_anciennete,
                       e.en_pause,
                       SUM(pvp.qte_brut) AS qte_brut,
                       SUM(pvp.qte_hors_rejet) AS qte_hors_rejet,
                       SUM(pvp.qte_paye) AS qte_paye
                  FROM divers.pgt_podium_vendeur pv
                  JOIN divers.pgt_podium_vendeur_part pvp
                       ON pvp.id_podium_vendeur = pv.id_podium_vendeur
                  JOIN pgt_salarie s ON s.id_salarie = pv.id_salarie
                  JOIN pgt_salarie_embauche e ON e.id_salarie = pv.id_salarie
                 WHERE (pv.modif_elem IS NULL
                        OR pv.modif_elem NOT LIKE '%suppr%')
                   AND (pvp.modif_elem IS NULL
                        OR pvp.modif_elem NOT LIKE '%suppr%')
                   AND (e.modif_elem IS NULL
                        OR e.modif_elem NOT LIKE '%suppr%')
                   AND pv.date_jour BETWEEN ? AND ?
                   AND pv.id_podium_type = ?
                   AND pv.distributeur = ?
                   AND pvp.id_podium_type_part = ?
                 GROUP BY pv.id_salarie, s.nom, s.prenom,
                          e.en_activite, e.resp_equipe, e.date_anciennete,
                          e.en_pause"""
            params = (
                date_deb.isoformat(), date_fin.isoformat(),
                int(p.id_podium_type), p.is_distrib, id_ptp,
            )
            # Filtre espoir (WinDev {ESPOIR})
            if is_espoir:
                date_espoir = _sub_months(date_deb, 2)
                sql += " HAVING MAX(e.date_anciennete) >= ?"
                params = params + (date_espoir.isoformat(),)

        try:
            rows = db.query(sql, params) or []
        except Exception:
            logger.exception("rechercher_podium : query part=%s", id_ptp)
            continue

        for r in rows:
            qte_brut = float(r.get("qte_brut") or 0)
            qte_hr = float(r.get("qte_hors_rejet") or 0)
            qte_paye = float(r.get("qte_paye") or 0)

            if is_prod_groupe:
                # Aggregation par equipe
                id_eq = int(r.get("id_equipe") or 0)
                key = f"eq:{id_eq}"
                if key not in agg:
                    # Recupere le responsable a la date_deb->date_fin
                    resp = _resp_equipe_at(id_eq, date_deb) or {}
                    lib_eq = _find_lib_equipe(id_eq)
                    agg[key] = VendeurPodiumRow(
                        id_salarie=_clean_id(resp.get("id_salarie")),
                        nom=(
                            f"{(resp.get('nom') or '').strip()} "
                            f"{_cap_prenom((resp.get('prenom') or '').strip())}"
                        ).strip(),
                        id_equipe=str(id_eq) if id_eq else "",
                        equipe_lib=lib_eq,
                        visible=True,
                    )
                row = agg[key]
                # Aggregation valeur selon TypeProd
                if is_qualite:
                    row.brut += qte_brut
                    row.paye += qte_paye
                elif type_prod.lower() == "brut":
                    row.valeur += qte_brut
                elif type_prod.lower() == "horsrejet":
                    row.valeur += qte_hr
                else:  # 'Paye' ou autre
                    row.valeur += qte_paye
            else:
                # Aggregation par salarie
                id_sal = int(r.get("id_salarie") or 0)
                key = f"sal:{id_sal}"
                if key not in agg:
                    id_eq = _equipe_terrain(id_sal, date_fin)
                    lib_eq = _find_lib_equipe(id_eq)
                    en_pause = bool(r.get("en_pause"))
                    en_activite = bool(r.get("en_activite"))
                    visible = en_activite and not en_pause
                    # Filtre distributeur (cf. WinDev)
                    if p.is_distrib and orga_distrib:
                        if id_eq not in orga_distrib:
                            continue
                    agg[key] = VendeurPodiumRow(
                        id_salarie=str(id_sal),
                        nom=(
                            f"{(r.get('nom') or '').strip()} "
                            f"{_cap_prenom((r.get('prenom') or '').strip())}"
                        ).strip(),
                        date_anciennete=_iso_date(r.get("date_anciennete")),
                        id_equipe=str(id_eq) if id_eq else "",
                        equipe_lib=lib_eq,
                        visible=visible,
                    )
                row = agg[key]
                if is_qualite:
                    row.brut += qte_brut
                    row.paye += qte_paye
                elif type_prod.lower() == "brut":
                    row.valeur += qte_brut
                elif type_prod.lower() == "horsrejet":
                    row.valeur += qte_hr
                else:
                    row.valeur += qte_paye

    # 4. Calcul du taux si qualite
    if is_qualite:
        for row in agg.values():
            row.valeur = row.brut
            if row.brut > 20:
                row.taux = row.paye / row.brut

    # 5. Trie -Taux, -Valeur
    lignes = list(agg.values())
    lignes.sort(key=lambda x: (-x.taux, -x.valeur))

    return RechercherPodiumResult(
        ok=True,
        id_podium_mois=id_podium_mois,
        score_visible=score_visible,
        is_qualite=is_qualite,
        is_prod_groupe=is_prod_groupe,
        lignes=lignes,
        message=f"{len(lignes)} ligne(s)",
    )


# --------------------------------------------------------------------
# Btn Disquette (sauvegarde score visible)
# --------------------------------------------------------------------

def sauver_score_visible(p: SauveScoreVisibleParams, op_id: int) -> bool:
    """Cf. WinDev 'Btn Disquette' : PodiumMois.ScoreVisible = ..."""
    if not p.id_podium_mois or p.id_podium_mois == "0":
        return False
    db = get_pg_connection("rh")
    db.execute(
        """UPDATE divers.pgt_podium_mois
              SET score_visible = ?,
                  modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
            WHERE id_podium_mois = ?""",
        (p.score_visible, int(op_id), int(p.id_podium_mois)),
    )
    return True


# --------------------------------------------------------------------
# Btn Telecharger XLSX
# --------------------------------------------------------------------

def generer_xlsx_podium(p: TelechargerParams) -> tuple[str, bytes]:
    """Genere le XLSX du podium tel qu'affiche a l'ecran."""
    try:
        from openpyxl import Workbook  # noqa: PLC0415
        from openpyxl.styles import Font  # noqa: PLC0415
    except ImportError:
        return ("", b"")

    wb = Workbook()
    ws = wb.active
    ws.title = "Podium"
    font_h = Font(bold=True)

    ws.cell(row=1, column=1, value="Nom").font = font_h
    ws.cell(row=1, column=2, value="Date anciennete").font = font_h
    ws.cell(row=1, column=3, value="Equipe").font = font_h
    ws.cell(row=1, column=4, value="Valeur").font = font_h
    ws.cell(row=1, column=5, value="Taux").font = font_h

    r = 2
    for line in p.lignes:
        ws.cell(row=r, column=1, value=line.nom)
        ws.cell(row=r, column=2, value=line.date_anciennete)
        ws.cell(row=r, column=3, value=line.equipe_lib or line.id_equipe)
        ws.cell(row=r, column=4, value=line.valeur)
        ws.cell(row=r, column=5, value=line.taux)
        r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10

    buf = _io.BytesIO()
    wb.save(buf)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M")
    lib = (p.lib_podium or "podium").strip()
    lib = "".join(c for c in lib if c.isalnum() or c in " -_") or "podium"
    lib = lib.replace(" ", "_")
    fic_name = f"{lib}_{p.annee}-{p.mois:02d}_{ts}.xlsx"
    return (fic_name, buf.getvalue())
