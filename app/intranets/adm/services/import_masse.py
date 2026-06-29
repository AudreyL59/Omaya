"""Service Fen_ImportMasse (ADM Imports Bases -> Import en masse).

Permet d'appliquer en MASSE une modification depuis une liste de num_bs
dans un fichier Excel (col A = num_bs ligne par ligne).

5 onglets (cf WinDev) :
  1. Modif Etat (vendeur ou operateur) -- IMPLEMENTE
  2. Modif Produit
  3. Modif Options
  4. Ajout Infos Internes
  5. Modif Vendeur

Multi-partenaire : ENI, IAG, OEN, PRO, SFR, STR, VAL (selon Combo
Partenaire). Le service genere dynamiquement les requetes SQL selon
le partenaire (table pgt_xxx_contrat / pgt_xxx_etat_contrat /
pgt_xxx_produit).
"""

from __future__ import annotations

import base64
import io
from datetime import date, datetime
from typing import Optional

from pydantic import BaseModel

from app.core.database.pg import get_pg_connection
from app.shared.recrutement.services.recherche_cv import _int, _str
from app.intranets.adm.services.import_eni import (
    _new_id, _col_letter_to_index, _cell, _parse_date_fr, _dernier_jour_mois,
)


# Partenaires supportes (prefixes BDD en minuscules pour les tables PG)
PARTENAIRES_SUPPORTES = ["eni", "iag", "oen", "pro", "sfr", "str", "val"]


class MasseEtatParams(BaseModel):
    partenaire: str                     # 'eni', 'sfr', 'oen', etc.
    id_etat_new: int
    mois_paiement: str = ""             # 'MM-AAAA'
    col_num: str = "A"                  # colonne du fichier Excel
    simulation: bool = True
    modif_deja_statues: bool = False
    modif_uniquement_attente: bool = True   # 'En attente / temporaire'
    recoche_energies: bool = False           # recalcul nbPoints ENI
    mode: str = "vendeur"               # 'vendeur' ou 'operateur'


class MasseResume(BaseModel):
    nb_lignes: int = 0
    nb_modifies: int = 0
    nb_deja_statues: int = 0
    nb_non_modifies: int = 0
    nb_introuvables: int = 0
    nb_erreurs: int = 0


class MasseLigneResult(BaseModel):
    num_ctt: str = ""
    id_contrat: int = 0
    produit: str = ""
    ancien_etat: str = ""
    nouvel_etat: str = ""
    mois_paiement: str = ""
    statut: str = ""


class MasseResult(BaseModel):
    ok: bool
    partenaire: str
    mode: str
    id_etat_new: int
    lib_etat_new: str = ""
    simulation: bool
    resume: MasseResume
    lignes: list[MasseLigneResult] = []
    message: str = ""


def list_partenaires() -> list[dict]:
    """Liste les partenaires actifs (cf table pgt_partenaire)."""
    db = get_pg_connection("adv")
    in_list = ",".join(f"'{p}'" for p in PARTENAIRES_SUPPORTES)
    rows = db.query(
        f"""SELECT id_partenaire, prefixe_bdd, lib_partenaire, is_actif
             FROM adv.pgt_partenaire
            WHERE is_actif = TRUE
              AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
              AND LOWER(prefixe_bdd) IN ({in_list})
            ORDER BY lib_partenaire"""
    ) or []
    return [{"id_partenaire": _str(r.get("id_partenaire")),
             "prefixe_bdd": r.get("prefixe_bdd"),
             "lib_partenaire": r.get("lib_partenaire"),
             "is_actif": bool(r.get("is_actif"))} for r in rows]


def list_etats(partenaire: str) -> list[dict]:
    """Liste les etats du partenaire (avec lib_type_etat prefixe)."""
    p = (partenaire or "").lower()
    if p not in PARTENAIRES_SUPPORTES:
        return []
    db = get_pg_connection("adv")
    rows = db.query(
        f"""SELECT e.id_etat, e.lib_etat, e.id_type_etat,
                   t.lib_type
              FROM adv.pgt_{p}_etat_contrat e
              LEFT JOIN adv.pgt_type_etat_contrat t
                     ON t.id_type_etat = e.id_type_etat
             WHERE (e.modif_elem IS NULL OR e.modif_elem NOT LIKE '%suppr%')
             ORDER BY t.lib_type, e.lib_etat"""
    ) or []
    return [{"id_etat": int(r["id_etat"]),
             "lib_etat": r.get("lib_etat") or "",
             "lib_type_etat": r.get("lib_type") or "",
             "id_type_etat": int(r.get("id_type_etat") or 0),
             "lib_complet": f"{r.get('lib_type') or ''} - {r.get('lib_etat') or ''}"}
            for r in rows]


def list_produits(partenaire: str) -> list[dict]:
    """Liste les produits du partenaire (avec famille/sous_fam pre-formatte)."""
    p = (partenaire or "").lower()
    if p not in PARTENAIRES_SUPPORTES:
        return []
    db = get_pg_connection("adv")
    rows = db.query(
        f"""SELECT id_produit, lib_produit, famille, sous_fam, pro_actif
              FROM adv.pgt_{p}_produit
             WHERE (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
             ORDER BY famille, sous_fam, lib_produit"""
    ) or []
    out = []
    for r in rows:
        lib = r.get("lib_produit") or ""
        sf = r.get("sous_fam") or ""
        if p == "sfr":
            lib = f"{r.get('famille') or ''} - {lib}"
        elif p == "eni":
            if sf == "GAZ-ELEC":
                lib = f"Dual - {lib}"
            elif sf:
                lib = f"{sf} - {lib}"
        actif = " (Prod actif)" if r.get("pro_actif") else " (Prod désactivé)"
        out.append({"id_produit": int(r["id_produit"]),
                    "lib_produit": lib + actif,
                    "is_actif": bool(r.get("pro_actif"))})
    out.sort(key=lambda x: x["lib_produit"])
    return out


def _lookup_contrat_partenaire(partenaire: str, num_bs: str) -> Optional[dict]:
    """Lookup contrat + produit + etat pour un partenaire donne (mode vendeur).
    Retourne id_contrat, id_etat_contrat, num_bs, sous_fam, famille,
    id_type_etat, lib_produit, lib_etat ou None."""
    p = partenaire.lower()
    if p not in PARTENAIRES_SUPPORTES:
        return None
    db = get_pg_connection("adv")
    return db.query_one(
        f"""SELECT c.id_contrat, c.id_etat_contrat, c.num_bs,
                   p.sous_fam, p.famille, e.id_type_etat,
                   p.lib_produit, e.lib_etat
              FROM adv.pgt_{p}_contrat c
              JOIN adv.pgt_{p}_produit p ON p.id_produit = c.id_produit
              JOIN adv.pgt_{p}_etat_contrat e
                ON e.id_etat = c.id_etat_contrat
             WHERE UPPER(c.num_bs) = UPPER(?)
               AND (c.modif_elem IS NULL OR c.modif_elem NOT LIKE '%suppr%')
             LIMIT 1""",
        (num_bs,),
    )


def _lookup_contrat_operateur(partenaire: str, num_bs: str) -> Optional[dict]:
    """Lookup contrat avec id_etat OPERATEUR (SFR=id_etat_sfr, OEN=id_etat_oen)."""
    p = partenaire.lower()
    if p == "sfr":
        col_etat = "id_etat_sfr"
    elif p == "oen":
        col_etat = "id_etat_oen"
    else:
        return None
    db = get_pg_connection("adv")
    return db.query_one(
        f"""SELECT c.id_contrat, c.{col_etat} AS id_etat_ope, c.num_bs,
                   p.sous_fam, p.famille, e.id_type_etat,
                   p.lib_produit, e.lib_etat
              FROM adv.pgt_{p}_contrat c
              JOIN adv.pgt_{p}_produit p ON p.id_produit = c.id_produit
              JOIN adv.pgt_{p}_etat_contrat e ON e.id_etat = c.{col_etat}
             WHERE UPPER(c.num_bs) = UPPER(?)
               AND (c.modif_elem IS NULL OR c.modif_elem NOT LIKE '%suppr%')
             LIMIT 1""",
        (num_bs,),
    )


def _lookup_etat(partenaire: str, id_etat: int) -> Optional[dict]:
    p = partenaire.lower()
    if p not in PARTENAIRES_SUPPORTES:
        return None
    db = get_pg_connection("adv")
    return db.query_one(
        f"""SELECT id_etat, lib_etat, id_type_etat
              FROM adv.pgt_{p}_etat_contrat
             WHERE id_etat = ? LIMIT 1""",
        (int(id_etat),),
    )


def _ajoute_histo(partenaire: str, id_contrat: int, old_etat: int,
                   new_etat: int, mois_p: str, op_id: int,
                   categorie: str = "Vend") -> None:
    """Historise la transition d'etat dans pgt_xxx_histo_etat_ctt."""
    p = partenaire.lower()
    if p not in PARTENAIRES_SUPPORTES or not id_contrat:
        return
    try:
        db = get_pg_connection("adv")
        auto = db.query_one(
            f"SELECT COALESCE(MAX(id_histo_auto), 0) + 1 AS n "
            f"FROM adv.pgt_{p}_histo_etat_ctt"
        )
        db.query(
            f"""INSERT INTO adv.pgt_{p}_histo_etat_ctt
                  (id_histo_auto, id_histo, id_contrat, op_saisie, date,
                   old_etat, new_etat, date_paiement,
                   modif_op, modif_date, modif_elem)
               VALUES (?, ?, ?, ?, NOW(), ?, ?, ?, ?, NOW(), 'new')""",
            (int(auto["n"]) if auto else 1, _new_id(),
             int(id_contrat), int(op_id),
             int(old_etat) if old_etat else 0,
             int(new_etat) if new_etat else 0,
             mois_p or "", int(op_id)),
        )
    except Exception:
        pass


def run_modif_etat(
    p: MasseEtatParams, content: bytes, op_id: int,
) -> MasseResult:
    """Onglet 1 : Modif Etat (vendeur ou operateur) en masse.

    - Lit le fichier Excel (col `col_num`, defaut A) = liste de num_bs
    - Pour chaque num : lookup contrat
    - Si trouve et etat = etat cible -> 'Deja statue'
    - Si trouve et options autorisent -> UPDATE (avec logique speciale
      SFR : DISTRIB cochee, MoisP_Va/Ra/RaDistri/VaDistri)
    - Historisation
    """
    from openpyxl import load_workbook

    partenaire = (p.partenaire or "").lower()
    if partenaire not in PARTENAIRES_SUPPORTES:
        return MasseResult(
            ok=False, partenaire=partenaire, mode=p.mode,
            id_etat_new=p.id_etat_new, simulation=p.simulation,
            resume=MasseResume(),
            message=f"Partenaire inconnu : {p.partenaire}",
        )

    etat_new = _lookup_etat(partenaire, p.id_etat_new)
    if not etat_new:
        return MasseResult(
            ok=False, partenaire=partenaire, mode=p.mode,
            id_etat_new=p.id_etat_new, simulation=p.simulation,
            resume=MasseResume(),
            message=f"État cible {p.id_etat_new} introuvable pour {partenaire}",
        )
    lib_etat_new = _str(etat_new.get("lib_etat"))
    type_etat_new = int(etat_new.get("id_type_etat") or 0)

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return MasseResult(
            ok=False, partenaire=partenaire, mode=p.mode,
            id_etat_new=p.id_etat_new, simulation=p.simulation,
            resume=MasseResume(), message=f"Lecture XLSX : {e}",
        )
    ws = wb.active
    col_idx = _col_letter_to_index(p.col_num)
    db = get_pg_connection("adv")
    resume = MasseResume()
    lignes: list[MasseLigneResult] = []

    for i in range(2, (ws.max_row or 0) + 1):
        num_ctt = _cell(ws, i, col_idx).upper().strip()
        if not num_ctt:
            continue
        resume.nb_lignes += 1

        if p.mode == "operateur":
            ctt = _lookup_contrat_operateur(partenaire, num_ctt)
            etat_actuel_col = "id_etat_ope"
        else:
            ctt = _lookup_contrat_partenaire(partenaire, num_ctt)
            etat_actuel_col = "id_etat_contrat"

        if not ctt:
            lignes.append(MasseLigneResult(
                num_ctt=num_ctt, statut="Contrat non trouvé",
            ))
            resume.nb_introuvables += 1
            continue

        id_contrat = int(ctt["id_contrat"])
        id_etat_actuel = int(ctt.get(etat_actuel_col) or 0)
        type_etat_actuel = int(ctt.get("id_type_etat") or 0)
        lib_etat_actuel = _str(ctt.get("lib_etat"))
        lib_produit = _str(ctt.get("lib_produit"))
        sous_fam = _str(ctt.get("sous_fam"))

        # Deja statue ?
        if id_etat_actuel == p.id_etat_new and not p.modif_deja_statues:
            lignes.append(MasseLigneResult(
                num_ctt=num_ctt, id_contrat=id_contrat, produit=lib_produit,
                ancien_etat=lib_etat_actuel, nouvel_etat=lib_etat_new,
                mois_paiement=p.mois_paiement, statut="Déjà statué",
            ))
            resume.nb_deja_statues += 1
            continue

        # Filtre 'UNIQUEMENT ATTENTE/TEMPORAIRE' (type 1, 2, 8)
        # ou si type_etat=5 et new=6 (decom apres valide)
        autorise = True
        if p.mode == "vendeur" and p.modif_uniquement_attente:
            if not (type_etat_actuel in (1, 2, 8)
                    or (type_etat_actuel == 5 and type_etat_new == 6)):
                autorise = False

        if not autorise:
            lignes.append(MasseLigneResult(
                num_ctt=num_ctt, id_contrat=id_contrat, produit=lib_produit,
                ancien_etat=lib_etat_actuel, nouvel_etat=lib_etat_new,
                statut="Contrat non modifié (filtre attente)",
            ))
            resume.nb_non_modifies += 1
            continue

        # MAJ
        mois_p_db = _dernier_jour_mois(p.mois_paiement)
        mois_p_str = p.mois_paiement
        if not p.simulation:
            try:
                if p.mode == "vendeur":
                    _update_etat_vendeur(
                        partenaire, id_contrat, p.id_etat_new,
                        lib_etat_new, type_etat_new, mois_p_db, op_id,
                    )
                    if (p.recoche_energies and partenaire == "eni"
                            and type_etat_new == 5):
                        _recalc_points_eni(id_contrat, sous_fam, op_id)
                else:
                    _update_etat_operateur(
                        partenaire, id_contrat, p.id_etat_new,
                        mois_p_db, op_id,
                    )
                _ajoute_histo(partenaire, id_contrat, id_etat_actuel,
                              p.id_etat_new, mois_p_str, op_id)
                lignes.append(MasseLigneResult(
                    num_ctt=num_ctt, id_contrat=id_contrat,
                    produit=lib_produit, ancien_etat=lib_etat_actuel,
                    nouvel_etat=lib_etat_new, mois_paiement=mois_p_str,
                    statut="Contrat modifié",
                ))
                resume.nb_modifies += 1
            except Exception as e:
                lignes.append(MasseLigneResult(
                    num_ctt=num_ctt, id_contrat=id_contrat,
                    produit=lib_produit, ancien_etat=lib_etat_actuel,
                    nouvel_etat=lib_etat_new,
                    statut=f"Erreur : {e}",
                ))
                resume.nb_erreurs += 1
        else:
            lignes.append(MasseLigneResult(
                num_ctt=num_ctt, id_contrat=id_contrat,
                produit=lib_produit, ancien_etat=lib_etat_actuel,
                nouvel_etat=lib_etat_new, mois_paiement=mois_p_str,
                statut="Mode Simu",
            ))
            resume.nb_modifies += 1
    wb.close()

    return MasseResult(
        ok=True, partenaire=partenaire, mode=p.mode,
        id_etat_new=p.id_etat_new, lib_etat_new=lib_etat_new,
        simulation=p.simulation, resume=resume, lignes=lignes,
        message=(f"{resume.nb_lignes} ligne(s) | Modifs {resume.nb_modifies} | "
                 f"Déjà statués {resume.nb_deja_statues} | "
                 f"Non modifiés {resume.nb_non_modifies} | "
                 f"Introuvables {resume.nb_introuvables} | "
                 f"Erreurs {resume.nb_erreurs}. "
                 + ("(SIMULATION)" if p.simulation else "(PRODUCTION)")),
    )


def _update_etat_vendeur(
    partenaire: str, id_contrat: int, id_etat_new: int,
    lib_etat_new: str, type_etat_new: int,
    mois_p_db: Optional[date], op_id: int,
) -> None:
    """UPDATE id_etat_contrat + mois_p selon les regles WinDev :
    - SFR + 'DISTRIB' dans lib + type=5 (VALIDE/PAYE) : MoisP_Va/Ra/Distri,
      pas de changement etat vendeur
    - SFR autre + type=5 : MoisP_Va ou MoisP_Ra selon 'Raccordement/Activation'
    - SFR + type=6 (DECOMM) : MoisP_Ra
    - Autres partenaires + type=5/6 : mois_p
    - Sinon : reset mois_p
    """
    db = get_pg_connection("adv")
    p = partenaire.lower()
    lib_up = lib_etat_new.upper()
    is_distrib = "DISTRIB" in lib_up
    is_racc_activ = ("RACCORDEMENT" in lib_up or "ACTIVATION" in lib_up)

    if p == "sfr":
        if type_etat_new == 5:
            if is_distrib:
                # On ne change pas l'etat vendeur, juste Distri
                if is_racc_activ:
                    db.query(
                        """UPDATE adv.pgt_sfr_contrat
                              SET mois_p_ra_distri = ?, paye_ra_distri = TRUE,
                                  modif_date = NOW(), modif_op = ?,
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (mois_p_db, int(op_id), int(id_contrat)),
                    )
                else:
                    db.query(
                        """UPDATE adv.pgt_sfr_contrat
                              SET mois_p_va_distri = ?, paye_va_distri = TRUE,
                                  modif_date = NOW(), modif_op = ?,
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (mois_p_db, int(op_id), int(id_contrat)),
                    )
            else:
                col_mois = "mois_p_ra" if is_racc_activ else "mois_p_va"
                db.query(
                    f"""UPDATE adv.pgt_sfr_contrat
                          SET id_etat_contrat = ?, {col_mois} = ?,
                              modif_date = NOW(), modif_op = ?,
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (id_etat_new, mois_p_db, int(op_id), int(id_contrat)),
                )
        elif type_etat_new == 6:
            db.query(
                """UPDATE adv.pgt_sfr_contrat
                      SET id_etat_contrat = ?, mois_p_ra = ?,
                          modif_date = NOW(), modif_op = ?,
                          modif_elem = 'modif'
                    WHERE id_contrat = ?""",
                (id_etat_new, mois_p_db, int(op_id), int(id_contrat)),
            )
        else:
            db.query(
                """UPDATE adv.pgt_sfr_contrat
                      SET id_etat_contrat = ?, mois_p_ra = NULL,
                          modif_date = NOW(), modif_op = ?,
                          modif_elem = 'modif'
                    WHERE id_contrat = ?""",
                (id_etat_new, int(op_id), int(id_contrat)),
            )
    else:
        # Autres partenaires : mois_p general
        if type_etat_new in (5, 6):
            db.query(
                f"""UPDATE adv.pgt_{p}_contrat
                      SET id_etat_contrat = ?, mois_p = ?,
                          modif_date = NOW(), modif_op = ?,
                          modif_elem = 'modif'
                    WHERE id_contrat = ?""",
                (id_etat_new, mois_p_db, int(op_id), int(id_contrat)),
            )
        else:
            db.query(
                f"""UPDATE adv.pgt_{p}_contrat
                      SET id_etat_contrat = ?, mois_p = NULL,
                          modif_date = NOW(), modif_op = ?,
                          modif_elem = 'modif'
                    WHERE id_contrat = ?""",
                (id_etat_new, int(op_id), int(id_contrat)),
            )


def _update_etat_operateur(
    partenaire: str, id_contrat: int, id_etat_new: int,
    mois_p_db: Optional[date], op_id: int,
) -> None:
    """UPDATE id_etat operateur (SFR=id_etat_sfr, OEN=id_etat_oen)."""
    db = get_pg_connection("adv")
    p = partenaire.lower()
    if p == "sfr":
        db.query(
            """UPDATE adv.pgt_sfr_contrat
                  SET id_etat_sfr = ?,
                      modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                WHERE id_contrat = ?""",
            (id_etat_new, int(op_id), int(id_contrat)),
        )
    elif p == "oen":
        db.query(
            """UPDATE adv.pgt_oen_contrat
                  SET id_etat_oen = ?, mois_p = ?,
                      modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                WHERE id_contrat = ?""",
            (id_etat_new, mois_p_db, int(op_id), int(id_contrat)),
        )


# ---------------------------------------------------------------------------
# Onglet 2 : Modif Produit
# ---------------------------------------------------------------------------


class MasseProduitParams(BaseModel):
    partenaire: str
    id_produit_new: int
    col_num: str = "A"
    simulation: bool = True


def run_modif_produit(
    p: MasseProduitParams, content: bytes, op_id: int,
) -> MasseResult:
    """Onglet 2 : Modif Produit en masse.
    Particularite : si id_produit_new=72, on suffixe le NumBS par '-TLCDUO'
    (apres extraction de la 1ere partie avant '-')."""
    from openpyxl import load_workbook
    partenaire = (p.partenaire or "").lower()
    if partenaire not in PARTENAIRES_SUPPORTES:
        return MasseResult(
            ok=False, partenaire=partenaire, mode="produit",
            id_etat_new=p.id_produit_new, simulation=p.simulation,
            resume=MasseResume(), message=f"Partenaire inconnu : {p.partenaire}",
        )
    db = get_pg_connection("adv")
    prod = db.query_one(
        f"SELECT lib_produit FROM adv.pgt_{partenaire}_produit "
        f"WHERE id_produit = ? LIMIT 1", (p.id_produit_new,),
    )
    lib_prod_new = (prod.get("lib_produit") if prod else "") or ""

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return MasseResult(
            ok=False, partenaire=partenaire, mode="produit",
            id_etat_new=p.id_produit_new, simulation=p.simulation,
            resume=MasseResume(), message=f"Lecture XLSX : {e}",
        )
    ws = wb.active
    col_idx = _col_letter_to_index(p.col_num)
    resume = MasseResume()
    lignes: list[MasseLigneResult] = []

    for i in range(2, (ws.max_row or 0) + 1):
        num_ctt = _cell(ws, i, col_idx).upper().strip()
        if not num_ctt:
            continue
        resume.nb_lignes += 1
        ctt = db.query_one(
            f"""SELECT c.id_contrat, c.num_bs, c.id_produit, p.lib_produit
                 FROM adv.pgt_{partenaire}_contrat c
                 LEFT JOIN adv.pgt_{partenaire}_produit p
                        ON p.id_produit = c.id_produit
                WHERE UPPER(c.num_bs) = UPPER(?)
                  AND (c.modif_elem IS NULL OR c.modif_elem NOT LIKE '%suppr%')
                LIMIT 1""", (num_ctt,),
        )
        if not ctt:
            lignes.append(MasseLigneResult(num_ctt=num_ctt,
                                            statut="Contrat non trouvé"))
            resume.nb_introuvables += 1
            continue
        id_contrat = int(ctt["id_contrat"])
        old_id_prod = int(ctt.get("id_produit") or 0)
        if old_id_prod == p.id_produit_new:
            lignes.append(MasseLigneResult(
                num_ctt=num_ctt, id_contrat=id_contrat,
                produit=_str(ctt.get("lib_produit")),
                ancien_etat=_str(ctt.get("lib_produit")),
                nouvel_etat=lib_prod_new, statut="Produit non modifié",
            ))
            resume.nb_non_modifies += 1
            continue
        if not p.simulation:
            try:
                # Cas special id_produit_new=72 -> suffixe NumBS '-TLCDUO'
                if p.id_produit_new == 72:
                    new_num_bs = ctt["num_bs"].split("-", 1)[0] + "-TLCDUO"
                    db.query(
                        f"""UPDATE adv.pgt_{partenaire}_contrat
                              SET id_produit = ?, num_bs = ?,
                                  modif_date = NOW(), modif_op = ?,
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (p.id_produit_new, new_num_bs, int(op_id), id_contrat),
                    )
                else:
                    db.query(
                        f"""UPDATE adv.pgt_{partenaire}_contrat
                              SET id_produit = ?,
                                  modif_date = NOW(), modif_op = ?,
                                  modif_elem = 'modif'
                            WHERE id_contrat = ?""",
                        (p.id_produit_new, int(op_id), id_contrat),
                    )
                statut = "Produit modifié"
                resume.nb_modifies += 1
            except Exception as e:
                statut = f"Erreur : {e}"
                resume.nb_erreurs += 1
        else:
            statut = "Mode Simu"
            resume.nb_modifies += 1
        lignes.append(MasseLigneResult(
            num_ctt=num_ctt, id_contrat=id_contrat,
            produit=_str(ctt.get("lib_produit")),
            ancien_etat=_str(ctt.get("lib_produit")),
            nouvel_etat=lib_prod_new, statut=statut,
        ))
    wb.close()
    return MasseResult(
        ok=True, partenaire=partenaire, mode="produit",
        id_etat_new=p.id_produit_new, lib_etat_new=lib_prod_new,
        simulation=p.simulation, resume=resume, lignes=lignes,
        message=(f"{resume.nb_lignes} ligne(s) | Modifs {resume.nb_modifies} | "
                 f"Non modifiés {resume.nb_non_modifies} | "
                 f"Introuvables {resume.nb_introuvables}. "
                 + ("(SIMULATION)" if p.simulation else "(PRODUCTION)")),
    )


# ---------------------------------------------------------------------------
# Onglet 3 : Modif Options SFR (HorsCible/Cluster)
# ---------------------------------------------------------------------------


class MasseOptionSfrParams(BaseModel):
    hors_cluster: bool                  # nouvelle valeur HorsCible
    col_num: str = "A"
    simulation: bool = True


def run_modif_option_sfr(
    p: MasseOptionSfrParams, content: bytes, op_id: int,
) -> MasseResult:
    """Onglet 3 : MAJ HorsCible (hors_cluster) sur les contrats SFR."""
    from openpyxl import load_workbook
    db = get_pg_connection("adv")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return MasseResult(
            ok=False, partenaire="sfr", mode="option_sfr",
            id_etat_new=0, simulation=p.simulation,
            resume=MasseResume(), message=f"Lecture XLSX : {e}",
        )
    ws = wb.active
    col_idx = _col_letter_to_index(p.col_num)
    resume = MasseResume()
    lignes: list[MasseLigneResult] = []
    target = "Hors cluster" if p.hors_cluster else "Dans cluster"

    for i in range(2, (ws.max_row or 0) + 1):
        num_ctt = _cell(ws, i, col_idx).upper().strip()
        if not num_ctt:
            continue
        resume.nb_lignes += 1
        ctt = db.query_one(
            """SELECT c.id_contrat, c.num_bs, c.hors_cible,
                      p.lib_produit, p.sous_fam, p.famille
                 FROM adv.pgt_sfr_contrat c
                 JOIN adv.pgt_sfr_produit p ON p.id_produit = c.id_produit
                WHERE UPPER(c.num_bs) = UPPER(?)
                  AND (c.modif_elem IS NULL OR c.modif_elem NOT LIKE '%suppr%')
                LIMIT 1""", (num_ctt,),
        )
        if not ctt:
            lignes.append(MasseLigneResult(num_ctt=num_ctt,
                                            statut="Contrat non trouvé"))
            resume.nb_introuvables += 1
            continue
        id_contrat = int(ctt["id_contrat"])
        actuel = bool(ctt.get("hors_cible"))
        ancien_lbl = "Hors cluster" if actuel else "Dans cluster"
        if actuel == p.hors_cluster:
            lignes.append(MasseLigneResult(
                num_ctt=num_ctt, id_contrat=id_contrat,
                produit=_str(ctt.get("lib_produit")),
                ancien_etat=ancien_lbl, nouvel_etat=target,
                statut="Option déjà à jour",
            ))
            resume.nb_non_modifies += 1
            continue
        if not p.simulation:
            try:
                db.query(
                    """UPDATE adv.pgt_sfr_contrat
                          SET hors_cible = ?,
                              modif_date = NOW(), modif_op = ?,
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (p.hors_cluster, int(op_id), id_contrat),
                )
                statut = "Option modifiée"; resume.nb_modifies += 1
            except Exception as e:
                statut = f"Erreur : {e}"; resume.nb_erreurs += 1
        else:
            statut = "Mode Simu"; resume.nb_modifies += 1
        lignes.append(MasseLigneResult(
            num_ctt=num_ctt, id_contrat=id_contrat,
            produit=_str(ctt.get("lib_produit")),
            ancien_etat=ancien_lbl, nouvel_etat=target, statut=statut,
        ))
    wb.close()
    return MasseResult(
        ok=True, partenaire="sfr", mode="option_sfr", id_etat_new=0,
        lib_etat_new=target, simulation=p.simulation, resume=resume,
        lignes=lignes,
        message=(f"{resume.nb_lignes} ligne(s) | Modifs {resume.nb_modifies} | "
                 f"Deja a jour {resume.nb_non_modifies} | "
                 f"Introuvables {resume.nb_introuvables}. "
                 + ("(SIMULATION)" if p.simulation else "(PRODUCTION)")),
    )


# ---------------------------------------------------------------------------
# Onglet 4 : Ajout Infos Internes (concat datetime + commentaire)
# ---------------------------------------------------------------------------


class MasseInfoParams(BaseModel):
    partenaire: str
    col_num: str = "A"
    col_comment: str = "B"
    simulation: bool = True


def run_ajout_info_interne(
    p: MasseInfoParams, content: bytes, op_id: int,
) -> MasseResult:
    """Onglet 4 : pour chaque ligne (num_bs, commentaire), concat
    'datetime + commentaire' a la fin de info_interne."""
    from openpyxl import load_workbook
    partenaire = (p.partenaire or "").lower()
    if partenaire not in PARTENAIRES_SUPPORTES:
        return MasseResult(
            ok=False, partenaire=partenaire, mode="info_interne",
            id_etat_new=0, simulation=p.simulation,
            resume=MasseResume(), message=f"Partenaire inconnu : {p.partenaire}",
        )
    db = get_pg_connection("adv")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return MasseResult(
            ok=False, partenaire=partenaire, mode="info_interne",
            id_etat_new=0, simulation=p.simulation,
            resume=MasseResume(), message=f"Lecture XLSX : {e}",
        )
    ws = wb.active
    col_id = _col_letter_to_index(p.col_num)
    col_comment = _col_letter_to_index(p.col_comment)
    resume = MasseResume()
    lignes: list[MasseLigneResult] = []
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    for i in range(2, (ws.max_row or 0) + 1):
        num_ctt = _cell(ws, i, col_id).upper().strip()
        comment = _cell(ws, i, col_comment).strip()
        if not num_ctt or not comment:
            continue
        resume.nb_lignes += 1
        ctt = db.query_one(
            f"""SELECT id_contrat, num_bs, info_interne
                 FROM adv.pgt_{partenaire}_contrat
                WHERE UPPER(num_bs) = UPPER(?)
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                LIMIT 1""", (num_ctt,),
        )
        if not ctt:
            lignes.append(MasseLigneResult(num_ctt=num_ctt,
                                            statut="Contrat non trouvé"))
            resume.nb_introuvables += 1
            continue
        id_contrat = int(ctt["id_contrat"])
        old_info = _str(ctt.get("info_interne"))
        new_info = (old_info + "\n" if old_info else "") + f"{now_str} : {comment}"
        if not p.simulation:
            try:
                db.query(
                    f"""UPDATE adv.pgt_{partenaire}_contrat
                          SET info_interne = ?,
                              modif_date = NOW(), modif_op = ?,
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (new_info, int(op_id), id_contrat),
                )
                statut = "Commentaire ajouté"; resume.nb_modifies += 1
            except Exception as e:
                statut = f"Erreur : {e}"; resume.nb_erreurs += 1
        else:
            statut = "Mode Simu"; resume.nb_modifies += 1
        lignes.append(MasseLigneResult(
            num_ctt=num_ctt, id_contrat=id_contrat, produit=partenaire.upper(),
            ancien_etat=old_info[:60] + ("..." if len(old_info) > 60 else ""),
            nouvel_etat=new_info[-80:], statut=statut,
        ))
    wb.close()
    return MasseResult(
        ok=True, partenaire=partenaire, mode="info_interne",
        id_etat_new=0, simulation=p.simulation, resume=resume, lignes=lignes,
        message=(f"{resume.nb_lignes} ligne(s) | Modifs {resume.nb_modifies} | "
                 f"Introuvables {resume.nb_introuvables}. "
                 + ("(SIMULATION)" if p.simulation else "(PRODUCTION)")),
    )


# ---------------------------------------------------------------------------
# Onglet 5 : Modif Vendeur (reattribution en masse vers un coopteur)
# ---------------------------------------------------------------------------


class MasseVendeurParams(BaseModel):
    partenaire: str
    id_salarie_new: int                 # id du coopteur cible
    col_num: str = "A"
    simulation: bool = True


def run_modif_vendeur(
    p: MasseVendeurParams, content: bytes, op_id: int,
) -> MasseResult:
    """Onglet 5 : reattribue tous les contrats du fichier au coopteur cible."""
    from openpyxl import load_workbook
    partenaire = (p.partenaire or "").lower()
    if partenaire not in PARTENAIRES_SUPPORTES:
        return MasseResult(
            ok=False, partenaire=partenaire, mode="vendeur_reattrib",
            id_etat_new=0, simulation=p.simulation,
            resume=MasseResume(), message=f"Partenaire inconnu : {p.partenaire}",
        )
    db = get_pg_connection("adv")
    db_rh = get_pg_connection("rh")
    sal = db_rh.query_one(
        "SELECT nom, prenom FROM rh.pgt_salarie WHERE id_salarie = ? LIMIT 1",
        (int(p.id_salarie_new),),
    )
    nom_coopteur = (f"{_str(sal.get('nom'))} {_str(sal.get('prenom')).title()}"
                    if sal else f"ID {p.id_salarie_new}")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return MasseResult(
            ok=False, partenaire=partenaire, mode="vendeur_reattrib",
            id_etat_new=0, simulation=p.simulation,
            resume=MasseResume(), message=f"Lecture XLSX : {e}",
        )
    ws = wb.active
    col_idx = _col_letter_to_index(p.col_num)
    resume = MasseResume()
    lignes: list[MasseLigneResult] = []

    for i in range(2, (ws.max_row or 0) + 1):
        num_ctt = _cell(ws, i, col_idx).upper().strip()
        if not num_ctt:
            continue
        resume.nb_lignes += 1
        ctt = db.query_one(
            f"""SELECT c.id_contrat, c.id_salarie, c.num_bs, p.lib_produit
                 FROM adv.pgt_{partenaire}_contrat c
                 LEFT JOIN adv.pgt_{partenaire}_produit p
                        ON p.id_produit = c.id_produit
                WHERE UPPER(c.num_bs) = UPPER(?)
                  AND (c.modif_elem IS NULL OR c.modif_elem NOT LIKE '%suppr%')
                LIMIT 1""", (num_ctt,),
        )
        if not ctt:
            lignes.append(MasseLigneResult(num_ctt=num_ctt,
                                            statut="Contrat non trouvé"))
            resume.nb_introuvables += 1
            continue
        id_contrat = int(ctt["id_contrat"])
        id_sal_db = int(ctt.get("id_salarie") or 0)
        old_sal = db_rh.query_one(
            "SELECT nom, prenom FROM rh.pgt_salarie WHERE id_salarie = ? LIMIT 1",
            (id_sal_db,),
        ) if id_sal_db else None
        old_lbl = (f"{_str(old_sal.get('nom'))} "
                   f"{_str(old_sal.get('prenom')).title()}".strip()
                   if old_sal else f"ID {id_sal_db}")
        if id_sal_db == p.id_salarie_new:
            lignes.append(MasseLigneResult(
                num_ctt=num_ctt, id_contrat=id_contrat,
                produit=_str(ctt.get("lib_produit")),
                ancien_etat=old_lbl, nouvel_etat=nom_coopteur,
                statut="Déjà réattribué",
            ))
            resume.nb_non_modifies += 1
            continue
        if not p.simulation:
            try:
                db.query(
                    f"""UPDATE adv.pgt_{partenaire}_contrat
                          SET id_salarie = ?,
                              modif_date = NOW(), modif_op = ?,
                              modif_elem = 'modif'
                        WHERE id_contrat = ?""",
                    (int(p.id_salarie_new), int(op_id), id_contrat),
                )
                statut = "Contrat modifié"; resume.nb_modifies += 1
            except Exception as e:
                statut = f"Erreur : {e}"; resume.nb_erreurs += 1
        else:
            statut = "Mode Simu"; resume.nb_modifies += 1
        lignes.append(MasseLigneResult(
            num_ctt=num_ctt, id_contrat=id_contrat,
            produit=_str(ctt.get("lib_produit")),
            ancien_etat=old_lbl, nouvel_etat=nom_coopteur, statut=statut,
        ))
    wb.close()
    return MasseResult(
        ok=True, partenaire=partenaire, mode="vendeur_reattrib",
        id_etat_new=p.id_salarie_new, lib_etat_new=nom_coopteur,
        simulation=p.simulation, resume=resume, lignes=lignes,
        message=(f"{resume.nb_lignes} ligne(s) | Modifs {resume.nb_modifies} | "
                 f"Deja reattribues {resume.nb_non_modifies} | "
                 f"Introuvables {resume.nb_introuvables}. "
                 + ("(SIMULATION)" if p.simulation else "(PRODUCTION)")),
    )


def _recalc_points_eni(id_contrat: int, sous_fam: str, op_id: int) -> None:
    """Recalcul nbPoints ENI apres MAJ etat = 5 (Valide-Paye).
    Recalcule via calcul_point_contrat (table pgt_bareme_point) +
    met a jour GazActif/ElecActif selon sous_fam.
    """
    db = get_pg_connection("adv")

    # Recupere les infos contrat necessaires au recalcul
    ctt = db.query_one(
        """SELECT date_signature, gaz_car_relevee, gaz_car_declaree,
                  elec_puissance, opt_rib, opt_mail, opt_maint, notation
             FROM adv.pgt_eni_contrat WHERE id_contrat = ?""",
        (int(id_contrat),),
    )

    gaz_actif = elec_actif = False
    sf = (sous_fam or "").upper()
    if sf == "GAZ":
        gaz_actif = True
    elif sf == "ELEC":
        elec_actif = True
    else:
        gaz_actif = True; elec_actif = True

    # Recalcul nb_points via le bareme central
    nb_points = 0.0
    if ctt and ctt.get("date_signature"):
        from app.shared.sdtc.bareme import calcul_point_contrat
        car = int(ctt.get("gaz_car_relevee") or ctt.get("gaz_car_declaree") or 0)
        puissance = int(ctt.get("elec_puissance") or 0)
        # Reconstruit la chaine d'options pour les regles ENI >= 2026-05-01
        opt_parts = []
        if ctt.get("opt_rib"): opt_parts.append("RIB//")
        if ctt.get("opt_mail"): opt_parts.append("MAIL//")
        if ctt.get("opt_maint"): opt_parts.append("MAINT//")
        if ctt.get("notation"): opt_parts.append(f"NOTE:{ctt.get('notation')}//")
        info_cplt = "".join(opt_parts)
        palier = car if "GAZ" in sf else puissance
        palier2 = puissance if "GAZ" in sf else car
        nb_points = calcul_point_contrat(
            fam="ENI", ss_fam=sous_fam or "",
            palier=palier, date_sign=str(ctt["date_signature"]),
            info_cplt=info_cplt, palier2=palier2,
        )

    try:
        db.query(
            """UPDATE adv.pgt_eni_contrat
                  SET gaz_actif = ?, elec_actif = ?, nb_points = ?,
                      modif_date = NOW(), modif_op = ?, modif_elem = 'modif'
                WHERE id_contrat = ?""",
            (gaz_actif, elec_actif, nb_points, int(op_id), int(id_contrat)),
        )
    except Exception:
        pass
