"""Service Gestion Code OHM (intranet Vendeur).

Portage de Page_GestionCodeOhm WinDev :
  - Liste des demandes de code (partenaire OHM = 562949953421321) :
      * Onglet 1 (encours)   : id_tk_type_demande = 38, cloturee = FALSE
      * Onglet 2 (a desactiver) : id_tk_type_demande = 39, cloturee = FALSE
  - Voir contenu d'une demande (code, login, mdp + fichiers)
  - Enregistrer (modif code/login/mdp)
  - Rejet manque documents (passe le statut a 38 + envoi mail au BO ou RH)

Codes statut :
  1  : nouveau (blanc)
  35 : envoye (bleu)
  36 : recu/traite (vert)
  38 : rejet manque documents (violet)

TypeOri :
  'TK'    : origine ticket distributeur -> infos via pgt_tk_demande_dpae_distrib
  'DPAE'  : origine embauche interne    -> infos via pgt_salarie + pgt_salarie_coordonnees
"""

from __future__ import annotations

import io
import logging
import os
import zipfile
from datetime import datetime
from typing import Any
from urllib.request import Request, urlopen

from app.core.config import MAIL_BO, MAIL_RH
from app.core.database.pg import get_pg_connection
from app.shared.notifications.mail import envoi_mail

logger = logging.getLogger(__name__)

# ID partenaire OHM (constante WinDev 562949953421321)
ID_PART_OHM = 562949953421321
TYPE_DEMANDE_ENCOURS = 38
TYPE_DEMANDE_A_DESACTIVER = 39


def _str_id(v: Any) -> str:
    if v is None:
        return ""
    try:
        n = int(v)
        return str(n) if n else ""
    except (TypeError, ValueError):
        return ""


def _capitalise(v: str) -> str:
    return v[:1].upper() + v[1:].lower() if v else ""


def _iso_datetime(v: Any) -> str:
    if not v:
        return ""
    if isinstance(v, str):
        return v
    try:
        return v.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(v)


def liste_demandes(type_demande: int) -> list[dict]:
    """Liste des demandes de code (encours=38 ou desactivation=39).

    Retourne chaque demande enrichie de : nom, prenom, num_tel + libelle
    statut. Tri par statut ASC puis date crea ASC.
    """
    if type_demande not in (TYPE_DEMANDE_ENCOURS, TYPE_DEMANDE_A_DESACTIVER):
        return []
    db_bo = get_pg_connection("ticket_bo")
    db_tk = get_pg_connection("ticket")
    db_rh = get_pg_connection("rh")

    try:
        rows = db_bo.query(
            """SELECT d.id_tk_liste, d.id_tk_demande_code_vendeur,
                      d.id_elem, d.type_ori, d.id_partenaire,
                      d.code, d.login, d.mdp,
                      d.modif_date, d.modif_elem
                 FROM ticket_bo.pgt_tk_demande_code_vendeur d
                WHERE (d.modif_elem IS NULL OR d.modif_elem NOT LIKE '%suppr%')
                  AND d.id_partenaire = ?""",
            (int(ID_PART_OHM),),
        ) or []
    except Exception:
        logger.exception("liste_demandes: fetch demande code vendeur")
        return []

    if not rows:
        return []

    # JOIN cote pgt_tk_liste (schema ticket)
    ids_liste = [int(r.get("id_tk_liste") or 0) for r in rows]
    ids_liste = [i for i in ids_liste if i]
    if not ids_liste:
        return []
    ids_sql = ",".join(str(i) for i in ids_liste)
    try:
        liste_rows = db_tk.query(
            f"""SELECT id_tk_liste, date_crea, id_tk_statut,
                      id_tk_type_demande, cloturee
                 FROM ticket.pgt_tk_liste
                WHERE id_tk_liste IN ({ids_sql})
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')
                  AND id_tk_type_demande = ?
                  AND COALESCE(cloturee, FALSE) = FALSE""",
            (int(type_demande),),
        ) or []
    except Exception:
        logger.exception("liste_demandes: fetch tk_liste")
        return []
    liste_map = {int(r.get("id_tk_liste") or 0): r for r in liste_rows}
    if not liste_map:
        return []

    # Libelles statuts
    statuts_ids = {int(r.get("id_tk_statut") or 0)
                    for r in liste_rows} - {0}
    lib_statut_map: dict[int, str] = {}
    if statuts_ids:
        try:
            libs = db_tk.query(
                f"""SELECT id_tk_statut, lib_statut
                     FROM ticket.pgt_tk_statut
                    WHERE id_tk_statut IN ({','.join(str(i) for i in statuts_ids)})""",
            ) or []
            for l in libs:
                lib_statut_map[int(l.get("id_tk_statut") or 0)] = l.get("lib_statut") or ""
        except Exception:
            logger.exception("liste_demandes: fetch lib_statut")

    # Resolution nom/prenom/tel selon TypeOri
    ids_dpae_distrib = [int(r.get("id_elem") or 0) for r in rows
                        if (r.get("type_ori") or "").upper() == "TK"]
    ids_salarie = [int(r.get("id_elem") or 0) for r in rows
                   if (r.get("type_ori") or "").upper() != "TK"]
    ids_dpae_distrib = [i for i in ids_dpae_distrib if i]
    ids_salarie = [i for i in ids_salarie if i]

    dpae_map: dict[int, dict] = {}
    if ids_dpae_distrib:
        ids_dp = ",".join(str(i) for i in ids_dpae_distrib)
        try:
            drows = db_bo.query(
                f"""SELECT id_tk_liste, nom, prenom, gsm
                     FROM ticket_bo.pgt_tk_demande_dpae_distrib
                    WHERE id_tk_liste IN ({ids_dp})""",
            ) or []
            for d in drows:
                dpae_map[int(d.get("id_tk_liste") or 0)] = d
        except Exception:
            logger.exception("liste_demandes: fetch dpae_distrib")

    sal_map: dict[int, dict] = {}
    if ids_salarie:
        ids_s = ",".join(str(i) for i in ids_salarie)
        try:
            srows = db_rh.query(
                f"""SELECT s.id_salarie, s.nom, s.prenom, sc.tel_mob
                     FROM rh.pgt_salarie s
                     LEFT JOIN rh.pgt_salarie_coordonnees sc
                            ON sc.id_salarie = s.id_salarie
                    WHERE s.id_salarie IN ({ids_s})""",
            ) or []
            for s in srows:
                sal_map[int(s.get("id_salarie") or 0)] = s
        except Exception:
            logger.exception("liste_demandes: fetch salarie")

    out = []
    for r in rows:
        id_l = int(r.get("id_tk_liste") or 0)
        if id_l not in liste_map:
            continue  # exclu par le filtre type/cloture
        l = liste_map[id_l]
        type_ori = (r.get("type_ori") or "").upper()
        id_elem = int(r.get("id_elem") or 0)
        nom = prenom = num_tel = ""
        if type_ori == "TK":
            info = dpae_map.get(id_l, {})
            nom = (info.get("nom") or "").strip()
            prenom = _capitalise((info.get("prenom") or "").strip())
            num_tel = (info.get("gsm") or "").strip()
        else:
            info = sal_map.get(id_elem, {})
            nom = (info.get("nom") or "").strip()
            prenom = _capitalise((info.get("prenom") or "").strip())
            num_tel = (info.get("tel_mob") or "").strip()

        out.append({
            "IDTK_Liste": _str_id(id_l),
            "IDTK_DemandeCodeVendeur": _str_id(r.get("id_tk_demande_code_vendeur")),
            "IdElem": _str_id(id_elem) if id_elem else "",
            "TypeOri": type_ori,
            "IDPartenaire": _str_id(r.get("id_partenaire")),
            "Code": r.get("code") or "",
            "Login": r.get("login") or "",
            "MDP": r.get("mdp") or "",
            "DateCrea": _iso_datetime(l.get("date_crea")),
            "IDTKStatut": int(l.get("id_tk_statut") or 0),
            "LibStatut": lib_statut_map.get(int(l.get("id_tk_statut") or 0), ""),
            "Nom": nom,
            "Prenom": prenom,
            "NomPrenom": f"{nom} {prenom}".strip(),
            "NumTel": num_tel,
        })
    # Tri : statut ASC puis date crea ASC
    out.sort(key=lambda x: (x["IDTKStatut"], x["DateCrea"]))
    return out


def fichiers_demande(id_tk_liste: int) -> list[dict]:
    """Retourne les fichiers d'une demande. Chaque fichier a :
    NomFichier, LienFichier (chemin relatif), Url (URL publique construite
    depuis DOCS_URL, pattern DocOmaya/PhotoDPAE/{lien_fichier}).
    """
    if not id_tk_liste:
        return []
    db = get_pg_connection("ticket_bo")
    try:
        rows = db.query(
            """SELECT id_tk_demande_code_vendeur_fichier AS id_fic,
                      nom_fichier, lien_fichier
                 FROM ticket_bo.pgt_tk_demandecodevendeur_fichier
                WHERE id_tk_liste = ?
                  AND (modif_elem IS NULL OR modif_elem NOT LIKE '%suppr%')""",
            (int(id_tk_liste),),
        ) or []
    except Exception:
        logger.exception("fichiers_demande id=%s", id_tk_liste)
        return []
    base = (os.environ.get("DOCS_URL", "") or "").rstrip("/")
    out = []
    for r in rows:
        lien = (r.get("lien_fichier") or "").strip()
        url = f"{base}/PhotoDPAE/{lien}" if base and lien else ""
        out.append({
            "IDFichier": _str_id(r.get("id_fic")),
            "NomFichier": r.get("nom_fichier") or "",
            "LienFichier": lien,
            "Url": url,
        })
    return out


def enregistrer_modif(id_tk_liste: int, code: str, login: str,
                       mdp: str, user_id: int) -> bool:
    """Update code/login/mdp sur pgt_tk_demande_code_vendeur (equivalent
    ÉcranVersFichier + HModifie du WinDev)."""
    if not id_tk_liste:
        return False
    db = get_pg_connection("ticket_bo")
    now = datetime.now()
    try:
        db.query(
            """UPDATE ticket_bo.pgt_tk_demande_code_vendeur
                  SET code = ?, login = ?, mdp = ?,
                      modif_date = ?, modif_op = ?, modif_elem = 'modif'
                WHERE id_tk_liste = ?""",
            (code or "", login or "", mdp or "",
             now, int(user_id), int(id_tk_liste)),
        )
        return True
    except Exception:
        logger.exception("enregistrer_modif id=%s", id_tk_liste)
        return False


def _info_user(user_id: int) -> dict:
    """Retourne nom / prenom / mail du user connecte pour la signature."""
    if not user_id:
        return {}
    rh = get_pg_connection("rh")
    try:
        r = rh.query_one(
            """SELECT s.nom, s.prenom, sc.mail
                 FROM rh.pgt_salarie s
                 LEFT JOIN rh.pgt_salarie_coordonnees sc
                        ON sc.id_salarie = s.id_salarie
                WHERE s.id_salarie = ? LIMIT 1""",
            (int(user_id),),
        )
        return r or {}
    except Exception:
        return {}


def rejet_manque_document(id_tk_liste: int, user_id: int) -> bool:
    """Passe le statut a 38 (Rejet Manque Documents) + envoi mail au
    BO (type_ori='TK') ou RH (autre)."""
    if not id_tk_liste:
        return False
    db_tk = get_pg_connection("ticket")
    db_bo = get_pg_connection("ticket_bo")
    now = datetime.now()
    try:
        db_tk.query(
            """UPDATE ticket.pgt_tk_liste
                  SET id_tk_statut = 38, modif_date = ?, modif_op = ?,
                      modif_elem = 'modif'
                WHERE id_tk_liste = ?""",
            (now, int(user_id), int(id_tk_liste)),
        )
    except Exception:
        logger.exception("rejet_manque_document: update statut id=%s", id_tk_liste)
        return False

    # Envoi mail
    try:
        dem = db_bo.query_one(
            """SELECT type_ori FROM ticket_bo.pgt_tk_demande_code_vendeur
                WHERE id_tk_liste = ? LIMIT 1""",
            (int(id_tk_liste),),
        )
        type_ori = ((dem or {}).get("type_ori") or "").upper()
        dest = MAIL_BO if type_ori == "TK" else MAIL_RH
        if dest:
            info = _info_user(user_id)
            nom_user = (info.get("nom") or "").strip()
            prenom_user = _capitalise((info.get("prenom") or "").strip())
            mail_user = info.get("mail") or ""

            # Nom du vendeur (via meme resolution que liste_demandes)
            nom_vendeur = _nom_vendeur(id_tk_liste)
            html = (
                "<font face='arial' style='font-size:10pt;'>"
                "<p>Bonjour,</p>"
                f"<p>Il manque des documents pour la demande de code OHM "
                f"Énergie pour le vendeur <b>{nom_vendeur}</b><br/>"
                "Merci de les rajouter sur le ticket.</p>"
                "<p>Cdt</p>"
                f"<p><b>{prenom_user} {nom_user}</b><br/>"
                f"<i>{mail_user}</i></p></font>"
            )
            envoi_mail(
                sujet=f"PROBLEME DEMANDE CODE OHM ENERGIE pour {nom_vendeur}",
                html=html, destinataires=[dest],
            )
    except Exception:
        logger.exception("rejet_manque_document: mail id=%s", id_tk_liste)
        # Non bloquant : le statut est deja mis a jour

    return True


def _nom_vendeur(id_tk_liste: int) -> str:
    """Nom prenom du vendeur associe a la demande (via type_ori)."""
    db_bo = get_pg_connection("ticket_bo")
    db_rh = get_pg_connection("rh")
    try:
        d = db_bo.query_one(
            """SELECT id_elem, type_ori FROM ticket_bo.pgt_tk_demande_code_vendeur
                WHERE id_tk_liste = ? LIMIT 1""",
            (int(id_tk_liste),),
        )
    except Exception:
        return ""
    if not d:
        return ""
    type_ori = (d.get("type_ori") or "").upper()
    id_elem = int(d.get("id_elem") or 0)
    try:
        if type_ori == "TK":
            r = db_bo.query_one(
                """SELECT nom, prenom
                     FROM ticket_bo.pgt_tk_demande_dpae_distrib
                    WHERE id_tk_liste = ? LIMIT 1""",
                (int(id_tk_liste),),
            )
        else:
            r = db_rh.query_one(
                """SELECT nom, prenom FROM rh.pgt_salarie
                    WHERE id_salarie = ? LIMIT 1""",
                (id_elem,),
            )
    except Exception:
        return ""
    if not r:
        return ""
    return f"{(r.get('nom') or '').strip()} {_capitalise((r.get('prenom') or '').strip())}".strip()


# ---------------------------------------------------------------------------
#  Export XLSX + ZIP FTP
# ---------------------------------------------------------------------------

def export_selection(ids_tk_liste: list[int], user_id: int) -> bytes:
    """Genere un ZIP contenant :
    - un XLSX 'Demande Accréditation' avec NOM/PRENOM/NUM TEL des demandes
    - les documents associes (nom = 'NomFichier NOM PRENOM.ext') telecharges
      depuis DOCS_URL + /PhotoDPAE/<lien_fichier>

    Passe le statut de chaque demande a 35 (Envoye) apres export.

    Retourne les bytes du ZIP (l'appelant fait le download HTTP).
    """
    if not ids_tk_liste:
        return b""
    # Lookup infos demandes via liste_demandes (deja resolu nom/prenom/num)
    demandes = liste_demandes(TYPE_DEMANDE_ENCOURS) + liste_demandes(TYPE_DEMANDE_A_DESACTIVER)
    dmap = {int(d["IDTK_Liste"]): d for d in demandes if d.get("IDTK_Liste")}
    to_export = [dmap[i] for i in ids_tk_liste if i in dmap]
    if not to_export:
        return b""

    # 1. Genere le XLSX
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Demande Accréditation"
    ws.append(["NOM", "PRENOM", "NUM TEL"])
    for d in to_export:
        ws.append([d["Nom"], d["Prenom"], d["NumTel"]])
    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_bytes = xlsx_buf.getvalue()

    # 2. Genere le ZIP avec XLSX + docs
    docs_url = (os.environ.get("DOCS_URL", "") or "").rstrip("/")
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        zf.writestr(f"Demandes_Accreditations_{stamp}.xlsx", xlsx_bytes)
        for d in to_export:
            fichiers = fichiers_demande(int(d["IDTK_Liste"]))
            for f in fichiers:
                lien = f.get("LienFichier") or ""
                if not lien or not docs_url:
                    continue
                url = f"{docs_url}/PhotoDPAE/{lien}"
                try:
                    with urlopen(Request(url), timeout=15) as resp:
                        content = resp.read()
                    ext = os.path.splitext(lien)[1]
                    name = f"{f.get('NomFichier') or 'doc'} {d['Nom']} {d['Prenom']}{ext}"
                    zf.writestr(name, content)
                except Exception:
                    logger.exception("export_selection: fetch %s", url)

    # 3. Passe le statut a 35 pour chaque demande
    db_tk = get_pg_connection("ticket")
    now = datetime.now()
    for d in to_export:
        try:
            db_tk.query(
                """UPDATE ticket.pgt_tk_liste
                      SET id_tk_statut = 35, modif_date = ?, modif_op = ?,
                          modif_elem = 'modif'
                    WHERE id_tk_liste = ?""",
                (now, int(user_id), int(d["IDTK_Liste"])),
            )
        except Exception:
            logger.exception("export_selection: update statut id=%s", d["IDTK_Liste"])

    return zip_buf.getvalue()


# ---------------------------------------------------------------------------
#  Import XLSX (codes remplis en masse)
# ---------------------------------------------------------------------------

def import_codes(xlsx_bytes: bytes, col_code: str, col_mdp: str,
                  col_nom: str, col_prenom: str,
                  user_id: int) -> dict:
    """Parse le XLSX + met a jour les demandes matchees (par code puis
    par nom+prenom).

    Colonnes : lettres Excel A/B/C/D (comme WinDev Asc(SAI_xxx)-64).

    Renvoie {'lignes_lues': N, 'maj_effectuees': N, 'mails_envoyes': N}.
    """
    from openpyxl import load_workbook

    def _col_idx(letter: str) -> int:
        letter = (letter or "").strip().upper()
        if not letter or not letter.isalpha():
            return 0
        # A=1, B=2, etc. — Asc(letter) - 64 en WinDev
        return ord(letter[0]) - 64

    ic = _col_idx(col_code)
    im = _col_idx(col_mdp)
    inom = _col_idx(col_nom)
    ipre = _col_idx(col_prenom)
    if not (ic and im and inom and ipre):
        raise ValueError("Mapping colonnes A/B/C/D invalide")

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return {"lignes_lues": 0, "maj_effectuees": 0, "mails_envoyes": 0}

    # Cache demandes en cours par code + par nom+prenom
    demandes = liste_demandes(TYPE_DEMANDE_ENCOURS)
    by_code = {(d.get("Code") or "").upper(): d for d in demandes if d.get("Code")}
    by_name = {
        f"{(d.get('Nom') or '').upper()} {(d.get('Prenom') or '').upper()}": d
        for d in demandes
    }

    db_bo = get_pg_connection("ticket_bo")
    db_tk = get_pg_connection("ticket")
    db_rh = get_pg_connection("rh")
    now = datetime.now()

    lignes = 0
    maj = 0
    mails = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        lignes += 1

        def _cell(i: int) -> str:
            if i < 1 or i > len(row):
                return ""
            v = row[i - 1]
            return str(v).strip() if v is not None else ""

        code_val = _cell(ic)
        mdp_val = _cell(im)
        nom_val = _cell(inom)
        pre_val = _cell(ipre)
        if not code_val or "Power_groupe".lower() not in code_val.lower():
            continue

        # 1. Match par code
        target = by_code.get(code_val.upper())
        if not target:
            # 2. Fallback nom+prenom (capitalise dans le cache)
            key = f"{nom_val.upper()} {pre_val.upper()}"
            target = by_name.get(key)
        if not target:
            # Deja renseigne mais statut 35 -> passer directement en 36
            for d in demandes:
                if (d.get("Code") or "").upper() == code_val.upper() \
                        and d.get("IDTKStatut") == 35:
                    try:
                        db_tk.query(
                            """UPDATE ticket.pgt_tk_liste
                                  SET id_tk_statut = 36, modif_date = ?,
                                      modif_op = ?, modif_elem = 'modif'
                                WHERE id_tk_liste = ?""",
                            (now, int(user_id), int(d["IDTK_Liste"])),
                        )
                    except Exception:
                        pass
            continue

        # Update TK_DemandeCodeVendeur + TK_Liste statut = 36
        try:
            db_bo.query(
                """UPDATE ticket_bo.pgt_tk_demande_code_vendeur
                      SET code = ?, login = ?, mdp = ?,
                          modif_date = ?, modif_op = ?, modif_elem = 'modif'
                    WHERE id_tk_liste = ?""",
                (code_val, code_val, mdp_val,
                 now, int(user_id), int(target["IDTK_Liste"])),
            )
            db_tk.query(
                """UPDATE ticket.pgt_tk_liste
                      SET id_tk_statut = 36, modif_date = ?, modif_op = ?,
                          modif_elem = 'modif'
                    WHERE id_tk_liste = ?""",
                (now, int(user_id), int(target["IDTK_Liste"])),
            )
            maj += 1
        except Exception:
            logger.exception("import_codes: update id=%s", target.get("IDTK_Liste"))
            continue

        # Si TypeOri DPAE : upsert salarie_partenaire + envoi mail vendeur
        if (target.get("TypeOri") or "").upper() != "TK":
            id_elem = int(target.get("IdElem") or 0)
            id_part = int(target.get("IDPartenaire") or 0)
            if id_elem and id_part:
                _upsert_salarie_partenaire(id_elem, id_part, code_val, mdp_val,
                                            int(user_id))
                # Mail vendeur
                sal = db_rh.query_one(
                    """SELECT s.nom, s.prenom, sc.mail
                         FROM rh.pgt_salarie s
                         LEFT JOIN rh.pgt_salarie_coordonnees sc
                                ON sc.id_salarie = s.id_salarie
                        WHERE s.id_salarie = ? LIMIT 1""",
                    (id_elem,),
                ) or {}
                mail_vend = sal.get("mail") or ""
                if mail_vend:
                    nom = (sal.get("nom") or "").strip()
                    prenom = _capitalise((sal.get("prenom") or "").strip())
                    html = (
                        "<font face='arial' style='font-size:10pt;'>"
                        "<p>Bonjour,</p>"
                        f"<p>{nom} {prenom} :<br/><ul>"
                        "<li><b>Partenaire :</b> OHM ENERGIE</li>"
                        f"<li><b>Code :</b> {code_val}</li>"
                        f"<li><b>Login :</b> {code_val}</li>"
                        f"<li><b>Mdp :</b> {mdp_val}</li>"
                        "</ul></p><br/>---"
                        "Cdt.<br/>"
                        "<p><i>PS : Ceci est un mail automatique, ne pas répondre. Merci.</i></p></font>"
                    )
                    try:
                        if envoi_mail(
                            sujet=f"Enregistrement Code partenaire {nom} {prenom}",
                            html=html, destinataires=[mail_vend],
                        ):
                            mails += 1
                            # Cloture ticket
                            try:
                                db_tk.query(
                                    """UPDATE ticket.pgt_tk_liste
                                          SET cloturee = TRUE, date_cloture = ?,
                                              modif_date = ?, modif_op = ?, modif_elem = 'modif'
                                        WHERE id_tk_liste = ?""",
                                    (now, now, int(user_id),
                                     int(target["IDTK_Liste"])),
                                )
                            except Exception:
                                pass
                    except Exception:
                        logger.exception("import_codes: envoi mail vendeur")

    return {"lignes_lues": lignes, "maj_effectuees": maj, "mails_envoyes": mails}


def _upsert_salarie_partenaire(id_salarie: int, id_partenaire: int,
                                 code: str, mdp: str, user_id: int) -> None:
    """Cree ou met a jour l'entree salarie_partenaire pour le user + OHM."""
    db = get_pg_connection("rh")
    now = datetime.now()
    try:
        existing = db.query_one(
            """SELECT id_salarie_partenaire
                 FROM rh.pgt_salarie_partenaire
                WHERE id_salarie = ? AND id_partenaire = ?
                LIMIT 1""",
            (int(id_salarie), int(id_partenaire)),
        )
        if existing:
            db.query(
                """UPDATE rh.pgt_salarie_partenaire
                      SET code = ?, login = ?, mdp = ?,
                          modif_date = ?, modif_op = ?, modif_elem = 'modif'
                    WHERE id_salarie_partenaire = ?""",
                (code, code, mdp, now, int(user_id),
                 int(existing.get("id_salarie_partenaire") or 0)),
            )
        else:
            new_id = int(datetime.now().strftime("%Y%m%d%H%M%S%f")[:17])
            db.query(
                """INSERT INTO rh.pgt_salarie_partenaire
                     (id_salarie_partenaire, id_partenaire, id_salarie,
                      code, login, mdp,
                      modif_date, modif_op, modif_elem)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'new')""",
                (new_id, int(id_partenaire), int(id_salarie),
                 code, code, mdp, now, int(user_id)),
            )
    except Exception:
        logger.exception("_upsert_salarie_partenaire id_sal=%s id_part=%s",
                         id_salarie, id_partenaire)
