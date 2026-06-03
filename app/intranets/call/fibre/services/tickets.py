"""
Service Call Fibre - tickets a traiter / traites du jour / stats agences.

Transposition de la fenetre principale WinDev Call SFR.

ETAPE 1 = lecture+ecriture sur HFSQL via get_connection().
Au cutover (octobre 2026), remplacer par get_pg_connection() + reecrire les
queries (pgt_*, snake_case, LIMIT au lieu de TOP, etc.).

Cross-base SQL impossible via le pont : TK_Liste/TK_Statut sont dans 'ticket'
et TK_CallSFR/TK_CallSFR_Panier dans 'ticket_bo'. On fait 2 queries puis on
joint en Python sur IDTK_Liste.
"""

import base64
import time
from datetime import date as _date, datetime
from typing import Any

from app.core.database import get_connection


# --- Constantes business (hardcodes comme WinDev) -------------------------

IDTK_TYPE_DEMANDE_CALL_FIBRE = 20  # IDTK_TypeDemande pour les Call Fibre
IDTK_STATUT_BLEU = 34  # statut affiche en bleu dans le tableau du haut

# Statuts du tableau du BAS (tickets "traites" du jour).
# Source : code WinDev `AfficherTkTraitésDiff` : BETWEEN 14 AND 19 sauf 18 et 28.
STATUTS_TRAITES = (14, 15, 16, 17, 19)  # 18 et 28 exclus

# Statut special "tk en cours mais a afficher en bleu" (cf. WinDev).
# Inclu dans le tableau du haut via OR explicite.
IDTK_STATUT_BLEU_EN_COURS = 34

# Agences "internes" (codees en dur cote WinDev).
# Tuple (id_orga_racine, libelle_affichage).
AGENCES_INTERNES: list[tuple[int, str]] = [
    (64, "Agence CD"),
    (20260402170805658, "Agence Duval Caen"),
    (20191203164626234, "Agence JR"),
    (20210906121249525, "Agence Le Mans"),
    (20260402142812484, "Agence Brosset Tours"),
    (20260402165637765, "Agence Poitiers"),
]

# Agences externes (reseau distrib).
ID_ORGA_POWER = 20180131091629815
ID_ORGA_FOX = 20230105145730716

# Filtre special "formation" : si OPCrea=6 et user_id != 6, on masque.
ID_OPE_FORMATION = 6

# Filtre special "TicketDiff masque pour le poste 20" (a confirmer cote metier).
ID_POSTE_MASQUE_DIFF = 20


# --- Helpers ---------------------------------------------------------------

def _civilite_to_prefix(civ: int) -> str:
    """Civilite 1 = M., autre = Mme (transposition exacte du code WinDev)."""
    return "M." if (civ or 0) <= 1 else "Mme"


def _capitalize(s: str) -> str:
    """Capitalise comme WinDev Capitalise() : 1ere lettre uppercase, reste lowercase."""
    if not s:
        return ""
    return s[0].upper() + s[1:].lower()


def _format_nom_client(civ: int, nom: str, nom_marital: str, prenom: str) -> str:
    """'M. NOM ép MARITAL Prenom' (transposition exacte WinDev)."""
    parts = [_civilite_to_prefix(civ), (nom or "").strip()]
    if (nom_marital or "").strip():
        parts.append(f"ép {nom_marital.strip()}")
    parts.append(_capitalize((prenom or "").strip()))
    return " ".join(p for p in parts if p)


def _format_ville(ville: str) -> str:
    """Coupe la partie '(...)' a la fin si presente."""
    if not ville:
        return ""
    if "(" in ville:
        return ville.split("(")[0].strip()
    return ville.strip()


def _iso(v) -> str:
    """datetime ou chaine HFSQL compact -> 'YYYY-MM-DD HH:MM:SS'."""
    if not v:
        return ""
    if isinstance(v, (datetime,)):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    s = str(v).strip()
    if not s or s.startswith("0000"):
        return ""
    # Compact HFSQL "20260602093015000"
    if len(s) >= 14 and s[:8].isdigit() and s[8:14].isdigit():
        return f"{s[0:4]}-{s[4:6]}-{s[6:8]} {s[8:10]}:{s[10:12]}:{s[12:14]}"
    # ISO deja
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:19]
    return s


def _to_int(v) -> int:
    try:
        return int(v or 0)
    except (TypeError, ValueError):
        return 0


def _parse_dt(v) -> datetime | None:
    """Parse robuste d'une date(time) HFSQL sous plusieurs formats.

    Retourne None si non parsable. Gere : datetime natif, ISO avec espace ou
    T, format compact HFSQL "YYYYMMDDHHMMSS...", date seule, etc.
    """
    if v is None or v == "":
        return v if isinstance(v, datetime) else None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s or s.startswith("0000"):
        return None
    s = s.replace("T", " ")
    # Compact HFSQL "20260602100000000"
    if len(s) >= 14 and s[:14].isdigit():
        try:
            return datetime.strptime(s[:14], "%Y%m%d%H%M%S")
        except ValueError:
            pass
    # ISO avec heure+sec. ATTENTION : len("%Y-%m-%d %H:%M:%S") == 17 mais
    # la chaine rendue fait 19 chars -> on hardcode la longueur de slice.
    for fmt, length in (("%Y-%m-%d %H:%M:%S", 19), ("%Y-%m-%d %H:%M", 16)):
        try:
            return datetime.strptime(s[:length], fmt)
        except ValueError:
            continue
    # Compact date seule "20260602"
    if len(s) == 8 and s.isdigit():
        try:
            return datetime.strptime(s, "%Y%m%d")
        except ValueError:
            pass
    # ISO date seule
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError:
            pass
    return None


def _str_id(v) -> str:
    """ID 8 octets -> str (cf. feedback_ids_8octets_string)."""
    return str(_to_int(v))


def _today_compact() -> str:
    """Date du jour au format HFSQL YYYYMMDD (pour LEFT(Datecrea, 8) = ?)."""
    return _date.today().strftime("%Y%m%d")


# --- Caches en memoire (vendeur, orga, contrats) --------------------------
# Ces helpers sont appeles plusieurs fois par appel. On cache localement
# pour eviter de re-requeter HFSQL.

def _load_salaries(db_rh, ids: set[int]) -> dict[int, dict]:
    """Charge nom/prenom pour une liste d'IDs salaries.

    Note : VendeurDistrib n'est PAS une colonne de la table Salarie. C'est
    une donnee derivee (calculee depuis l'orga d'affectation = descendant
    de Power/Fox). Voir _enrich_vendeur_distrib() pour l'enrichissement.
    """
    if not ids:
        return {}
    ids_sql = ",".join(str(i) for i in ids if i)
    if not ids_sql:
        return {}
    rows = db_rh.query(
        f"""SELECT IDSalarie, Nom, Prenom
        FROM Salarie
        WHERE IDSalarie IN ({ids_sql})"""
    )
    out: dict[int, dict] = {}
    for r in rows:
        out[_to_int(r.get("IDSalarie"))] = {
            "nom": (r.get("Nom") or "").strip(),
            "prenom": (r.get("Prenom") or "").strip(),
            "vendeur_distrib": False,  # rempli par _enrich_vendeur_distrib
        }
    return out


def _enrich_vendeur_distrib(
    salaries: dict[int, dict],
    affectations: dict[int, dict],
    db_rh,
) -> None:
    """Calcule en place le flag vendeur_distrib pour chaque salarie.

    Un vendeur est "distrib" (= externe) si son orga d'affectation est un
    descendant de Power ou Fox. Cette info etait portee par la colonne
    Salarie.VendeurDistrib en WinDev (peut-etre denormalisee via la
    procedure DonneInfoSalarié).
    """
    power_set = _orga_descendants(db_rh, ID_ORGA_POWER)
    fox_set = _orga_descendants(db_rh, ID_ORGA_FOX)
    distrib_orgas = power_set | fox_set
    for id_sal, sal in salaries.items():
        id_orga = affectations.get(id_sal, {}).get("id_orga", 0)
        sal["vendeur_distrib"] = id_orga in distrib_orgas


def _load_affectations(db_rh, id_salaries: set[int]) -> dict[int, dict]:
    """Pour chaque salarie, retourne l'orga d'affectation actuelle (lib + parent.lib).

    Note : on prend l'affectation la plus recente sans condition de date_fin.
    Au plus proche du `affectationVendeur` WinDev (mais simplifie).
    """
    if not id_salaries:
        return {}
    ids_sql = ",".join(str(i) for i in id_salaries if i)
    if not ids_sql:
        return {}
    # IMPORTANT : on filtre WHERE so.IDSalarie IN (...) pour ne charger que les
    # salaries qui nous interessent. Sans ce filtre, un TOP 1000 pourrait
    # tronquer arbitrairement et laisser certains salaries sans affectation
    # (bug : equipe vide dans le tableau du bas, stats agences faussees).
    rows = db_rh.query(
        f"""SELECT
            so.IDSalarie, so.IDOrganigramme, o.Lib_ORGA AS lib_orga,
            o.IdPARENT AS id_orga_parent
        FROM Salarie_Organigramme so
        INNER JOIN Organigramme o ON o.IDOrganigramme = so.IDOrganigramme
        WHERE so.IDSalarie IN ({ids_sql})
          AND so.ModifELEM NOT LIKE '%suppr%'
          AND (so.DateFin = '' OR so.DateFin >= ?)
        ORDER BY so.DateDébut DESC""",
        (_today_compact(),),
    )
    # Recuperer en plus le lib_orga des parents
    parent_ids = {_to_int(r.get("id_orga_parent")) for r in rows}
    parent_ids.discard(0)
    parents: dict[int, str] = {}
    if parent_ids:
        parent_ids_sql = ",".join(str(i) for i in parent_ids)
        prows = db_rh.query(
            f"""SELECT IDOrganigramme, Lib_ORGA FROM Organigramme
            WHERE IDOrganigramme IN ({parent_ids_sql})"""
        )
        for pr in prows:
            parents[_to_int(pr.get("IDOrganigramme"))] = (pr.get("Lib_ORGA") or "").strip()

    # On garde la 1re ligne par salarie (la plus recente grace au ORDER BY DESC)
    out: dict[int, dict] = {}
    for r in rows:
        id_sal = _to_int(r.get("IDSalarie"))
        if id_sal in out:
            continue
        parent_lib = parents.get(_to_int(r.get("id_orga_parent")), "")
        lib = (r.get("lib_orga") or "").strip()
        out[id_sal] = {
            "id_orga": _to_int(r.get("IDOrganigramme")),
            "lib_orga": lib,
            "parent_lib": parent_lib,
            "lib_equipe": f"{parent_lib} => {lib}" if parent_lib else lib,
        }
    return out


def _check_premier_contrat(db_adv, id_salarie: int, date_avant: str) -> bool:
    """True si le vendeur n'a pas de SFR_contrat signe avant date_avant.

    Equivalent du test WinDev "1er contrat -> coloration verte".
    `date_avant` doit etre au format HFSQL compact YYYYMMDD.

    DEPRECATED : 1 query par appel -> N+1 si beaucoup de tickets.
    Preferer _load_premiers_contrats() qui batch toutes les verifications.
    """
    rows = db_adv.query(
        """SELECT TOP 1 IDcontrat FROM SFR_contrat
        WHERE IDSalarie = ?
          AND LEFT(datesignature, 8) < ?""",
        (id_salarie, date_avant),
    )
    return not rows


def _load_premiers_contrats(db_adv, id_salaries: set[int]) -> dict[int, str]:
    """Pour chaque vendeur, retourne la date de son 1er contrat SFR
    (MIN(datesignature), format compact YYYYMMDD). Vide si jamais signe.

    1 seule query au lieu de N. Utiliser ensuite avec :
        is_premier = (date_ticket < date_premier_contrat) or pas dans le dict
    """
    if not id_salaries:
        return {}
    ids_sql = ",".join(str(i) for i in id_salaries if i)
    if not ids_sql:
        return {}
    rows = db_adv.query(
        f"""SELECT IDSalarie, MIN(datesignature) AS min_date
        FROM SFR_contrat
        WHERE IDSalarie IN ({ids_sql})
        GROUP BY IDSalarie"""
    )
    out: dict[int, str] = {}
    for r in rows:
        d = str(r.get("min_date") or "")[:8]
        out[_to_int(r.get("IDSalarie"))] = d
    return out


# --- Tableau du HAUT : tickets a traiter ----------------------------------

def _parse_id_encours(s: str) -> list[int]:
    """Parse SuiviTicketCall.IdEncours : chaine d'IDs separes par RC, format
    'id-xxx' (on prend la partie avant le '-').

    WinDev : `pour toute chaîne IDTk de ListeTkEncoursTesté séparée par RC`
    + `ExtraitChaîne(IDTk, 1, "-")`.
    """
    if not s:
        return []
    out: list[int] = []
    for line in s.replace("\r\n", "\n").split("\n"):
        line = line.strip()
        if not line:
            continue
        # Prend la partie avant le '-' si present, sinon l'ID complet.
        if "-" in line:
            line = line.split("-", 1)[0]
        try:
            out.append(int(line))
        except ValueError:
            continue
    return out


def list_tickets_en_cours(user_id: int, user_id_poste: int) -> list[dict]:
    """Liste les tickets Call Fibre a traiter (= tableau du haut).

    HFSQL n'a pas d'index utilisable sur TK_Liste(TypeDemande, Datecrea) ->
    n'importe quelle query WHERE timeout 60s. Workaround : on utilise le
    cache `SuiviTicketCall` (table maintenue par l'exe externe) qui contient
    la liste des IDs des tickets en cours. On SELECT ensuite par cle
    primaire (ultra rapide).

    Filtres business (transposition exacte WinDev) :
    - OPCrea = 6 masque sauf si user_id = 6 (formation)
    - Datecrea > today 23:59:59 masque sauf user_id = 6 (tickets futurs)
    - TicketDiff = True masque si user_id_poste = 20
    """
    db_ticket = get_connection("ticket")
    db_bo = get_connection("ticket_bo")
    db_rh = get_connection("rh")
    db_adv = get_connection("adv")
    db_divers = get_connection("divers")

    # 1. Lit le cache SuiviTicketCall pour TypeCall='FIBRE'
    rows_cache = db_divers.query(
        """SELECT IdEncours, IdAppelEnCours FROM SuiviTicketCall
        WHERE TypeCall = 'FIBRE'"""
    )
    if not rows_cache:
        return []
    id_encours_raw = (rows_cache[0] or {}).get("IdEncours") or ""
    id_appel_raw = (rows_cache[0] or {}).get("IdAppelEnCours") or ""
    ids = _parse_id_encours(id_encours_raw)
    if not ids:
        return []

    # 2. SELECT par cle primaire dans TK_Liste (rapide car IDTK_Liste est PK).
    # On re-applique les filtres business car le cache SuiviTicketCall peut etre
    # obsolete (l'exe externe tourne periodiquement, un ticket peut avoir change
    # de statut entre 2 runs).
    # Filtre Datecrea > today_00 (transposition WinDev TK_Liste.Datecrea > {ParamdateCrea}).
    # SI user != formation (id 6) : borne sup aussi -> seulement aujourd'hui.
    today_00 = _date.today().strftime("%Y%m%d000000000")
    today_compact = _date.today().strftime("%Y%m%d")
    ids_sql = ",".join(str(i) for i in ids)
    if user_id == ID_OPE_FORMATION:
        rows_liste_raw = db_ticket.query(
            f"""SELECT
                IDTK_Liste     AS id_tk_liste,
                Datecrea       AS date_crea,
                OPCrea         AS op_crea,
                IDTK_Statut    AS id_tk_statut,
                Cloturée       AS cloturee,
                ModifELEM      AS modif_elem
            FROM TK_Liste
            WHERE IDTK_Liste IN ({ids_sql})
              AND Datecrea > ?""",
            (today_00,),
        )
    else:
        # User normal : strict aujourd'hui (LEFT(Datecrea, 8) = today_compact).
        rows_liste_raw = db_ticket.query(
            f"""SELECT
                IDTK_Liste     AS id_tk_liste,
                Datecrea       AS date_crea,
                OPCrea         AS op_crea,
                IDTK_Statut    AS id_tk_statut,
                Cloturée       AS cloturee,
                ModifELEM      AS modif_elem
            FROM TK_Liste
            WHERE IDTK_Liste IN ({ids_sql})
              AND LEFT(Datecrea, 8) = ?""",
            (today_compact,),
        )
    # Filtres business (transposition exacte WinDev) :
    rows_liste = [
        r for r in rows_liste_raw
        if not bool(r.get("cloturee"))
        and "suppr" not in (r.get("modif_elem") or "").lower()
        and _to_int(r.get("id_tk_statut")) not in (18, 28)
        and (
            _to_int(r.get("id_tk_statut")) < 14
            or _to_int(r.get("id_tk_statut")) == 34
        )
    ]
    # Chargement TK_Statut (referentiel statique, cache module 10min)
    statuts = _load_tk_statut(db_ticket)
    if not rows_liste:
        return []

    # Map IDTK_Liste -> infos TK_Liste (avec Lib_Statut depuis cache)
    by_id: dict[int, dict] = {}
    for r in rows_liste:
        id_tk = _to_int(r.get("id_tk_liste"))
        id_statut = _to_int(r.get("id_tk_statut"))
        by_id[id_tk] = {
            "id_tk_liste": id_tk,
            "date_crea": r.get("date_crea") or "",
            "op_crea": _to_int(r.get("op_crea")),
            "id_tk_statut": id_statut,
            "lib_statut": statuts.get(id_statut, ""),
        }

    # 2. TK_CallSFR pour ces IDs
    ids_sql = ",".join(str(i) for i in by_id.keys())
    rows_call = db_bo.query(
        f"""SELECT
            IDTK_Liste, IDSalarie, CivilitéClient, NomClient, NomMaritalClient,
            PrenomClient, CP, VILLE, AppelEnCours, OpéAppel, TicketDiff
        FROM TK_CallSFR
        WHERE IDTK_Liste IN ({ids_sql})
          AND ModifELEM NOT LIKE '%suppr%'"""
    )

    # 3. Filtres business + format
    today_2359 = datetime.combine(_date.today(), datetime.max.time())
    salaries_to_load: set[int] = set()
    enriched: list[dict] = []

    for r in rows_call:
        id_tk = _to_int(r.get("IDTK_Liste"))
        tk = by_id.get(id_tk)
        if not tk:
            continue
        op_crea = tk["op_crea"]

        # Filtre formation (OPCrea = 6)
        if op_crea == ID_OPE_FORMATION and user_id != ID_OPE_FORMATION:
            continue
        # Filtre tickets futurs
        date_crea_raw = tk["date_crea"]
        date_crea_iso = _iso(date_crea_raw)
        if date_crea_iso:
            try:
                dt = datetime.strptime(date_crea_iso[:19], "%Y-%m-%d %H:%M:%S")
                if dt > today_2359 and user_id != ID_OPE_FORMATION:
                    continue
            except ValueError:
                pass
        # Filtre TicketDiff masque pour idPoste = 20
        ticket_diff = bool(r.get("TicketDiff"))
        if ticket_diff and user_id_poste == ID_POSTE_MASQUE_DIFF:
            continue

        id_salarie = _to_int(r.get("IDSalarie"))
        salaries_to_load.add(id_salarie)
        ope_appel_id = _to_int(r.get("OpéAppel"))
        if ope_appel_id:
            salaries_to_load.add(ope_appel_id)

        enriched.append({
            **tk,
            "id_salarie": id_salarie,
            "civilite": _to_int(r.get("CivilitéClient")),
            "nom": r.get("NomClient") or "",
            "nom_marital": r.get("NomMaritalClient") or "",
            "prenom": r.get("PrenomClient") or "",
            "cp": (r.get("CP") or "").strip(),
            "ville": _format_ville(r.get("VILLE") or ""),
            "appel_en_cours": bool(r.get("AppelEnCours")),
            "ope_appel_id": ope_appel_id,
            "ticket_diff": ticket_diff,
        })

    # 4. Resolve noms vendeur + affectations en PARALLELE (2 queries sur rh).
    from concurrent.futures import ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=2) as pool:
        f_sal = pool.submit(_load_salaries, db_rh, salaries_to_load)
        f_aff = pool.submit(_load_affectations, db_rh, salaries_to_load)
        salaries = f_sal.result()
        affectations = f_aff.result()
    _enrich_vendeur_distrib(salaries, affectations, db_rh)
    # Batch des premiers contrats SFR (1 query au lieu de N)
    sal_non_distrib_ids = {
        tk["id_salarie"] for tk in enriched
        if not salaries.get(tk["id_salarie"], {}).get("vendeur_distrib")
    }
    premiers_contrats = _load_premiers_contrats(db_adv, sal_non_distrib_ids)

    # 5. Construire la liste finale + non_prod
    out: list[dict] = []
    for tk in enriched:
        sal = salaries.get(tk["id_salarie"], {"nom": "", "prenom": "", "vendeur_distrib": False})
        aff = affectations.get(tk["id_salarie"], {"lib_equipe": ""})
        ope_appel_sal = salaries.get(tk["ope_appel_id"], {"nom": "", "prenom": ""})

        # non_prod = 1er contrat (= aucun SFR_contrat signe avant date_ticket)
        non_prod = False
        if not sal["vendeur_distrib"]:
            date_avant = _iso(tk["date_crea"])[:10].replace("-", "")
            min_date = premiers_contrats.get(tk["id_salarie"], "")
            # non_prod = True si pas de contrat OU 1er contrat est apres/egal a date_avant
            non_prod = (not min_date) or (min_date >= date_avant)

        out.append({
            "id": _str_id(tk["id_tk_liste"]),
            "date_crea": _iso(tk["date_crea"]),
            "nom_client": _format_nom_client(
                tk["civilite"], tk["nom"], tk["nom_marital"], tk["prenom"]
            ),
            "cp": tk["cp"],
            "ville": tk["ville"],
            "nom_vendeur": f"{sal['nom']} {_capitalize(sal['prenom'])}".strip(),
            "lib_equipe": aff["lib_equipe"],
            "lib_statut": tk["lib_statut"],
            "id_tk_statut": tk["id_tk_statut"],
            "fdv_interne": not sal["vendeur_distrib"],
            "non_prod": non_prod,
            "appel_en_cours": tk["appel_en_cours"],
            "ope_appel_nom": (
                f"{ope_appel_sal['nom']} {_capitalize(ope_appel_sal['prenom'])}".strip()
                if tk["appel_en_cours"] else ""
            ),
            "ticket_diff": tk["ticket_diff"],
        })

    # Tri DateCrea ASC (comme WinDev)
    out.sort(key=lambda t: t["date_crea"])
    return out


# --- Tableau du BAS : tickets traites du jour -----------------------------

def list_tickets_traites(jour: str | None = None) -> list[dict]:
    """Liste les tickets Call Fibre traites a la date donnee (defaut = aujourd'hui).

    `jour` : format ISO 'YYYY-MM-DD' ou compact 'YYYYMMDD'. None = today.
    """
    if jour:
        j = jour.replace("-", "")
    else:
        j = _today_compact()

    db_ticket = get_connection("ticket")
    db_bo = get_connection("ticket_bo")
    db_rh = get_connection("rh")
    db_adv = get_connection("adv")

    statuts_traites_sql = ",".join(str(s) for s in STATUTS_TRAITES)
    rows_liste = db_ticket.query(
        f"""SELECT
            IDTK_Liste     AS id_tk_liste,
            Datecrea       AS date_crea,
            IDTK_Statut    AS id_tk_statut
        FROM TK_Liste
        WHERE IDTK_TypeDemande = ?
          AND ModifELEM NOT LIKE '%suppr%'
          AND IDTK_Statut IN ({statuts_traites_sql})
          AND LEFT(Datecrea, 8) = ?
        ORDER BY Datecrea ASC""",
        (IDTK_TYPE_DEMANDE_CALL_FIBRE, j),
    )
    if not rows_liste:
        return []

    # TK_Statut (referentiel statique, cache module 10min)
    statuts = _load_tk_statut(db_ticket)

    by_id: dict[int, dict] = {}
    for r in rows_liste:
        id_tk = _to_int(r.get("id_tk_liste"))
        id_statut = _to_int(r.get("id_tk_statut"))
        by_id[id_tk] = {
            "id_tk_liste": id_tk,
            "date_crea": r.get("date_crea") or "",
            "id_tk_statut": id_statut,
            "lib_statut": statuts.get(id_statut, ""),
        }

    # TK_CallSFR pour ces IDs
    ids_sql = ",".join(str(i) for i in by_id.keys())
    rows_call = db_bo.query(
        f"""SELECT
            IDtk_CallSFR, IDTK_Liste, IDSalarie, CivilitéClient, NomClient,
            NomMaritalClient, PrenomClient, CP, VILLE, RefAppel
        FROM TK_CallSFR
        WHERE IDTK_Liste IN ({ids_sql})
          AND ModifELEM NOT LIKE '%suppr%'"""
    )

    # Map id_call -> tk_liste + collect salarie ids
    call_by_id_tk: dict[int, dict] = {}
    id_calls: list[int] = []
    salaries_to_load: set[int] = set()
    for r in rows_call:
        id_tk = _to_int(r.get("IDTK_Liste"))
        id_call = _to_int(r.get("IDtk_CallSFR"))
        if id_tk not in by_id:
            continue
        salaries_to_load.add(_to_int(r.get("IDSalarie")))
        id_calls.append(id_call)
        call_by_id_tk[id_tk] = {
            "id_call": id_call,
            "id_salarie": _to_int(r.get("IDSalarie")),
            "civilite": _to_int(r.get("CivilitéClient")),
            "nom": r.get("NomClient") or "",
            "nom_marital": r.get("NomMaritalClient") or "",
            "prenom": r.get("PrenomClient") or "",
            "cp": (r.get("CP") or "").strip(),
            "ville": _format_ville(r.get("VILLE") or ""),
            "ref_appel": (r.get("RefAppel") or "").strip(),
        }

    # Panier + Salaries + Affectations en PARALLELE (3 queries independantes).
    from concurrent.futures import ThreadPoolExecutor
    paniers: dict[int, list[dict]] = {}
    rows_p: list[dict] = []
    with ThreadPoolExecutor(max_workers=3) as pool:
        f_panier = pool.submit(
            db_bo.query,
            f"""SELECT IDtk_CallSFR, IDOffres_SFR, TYPE, TypeVente, NUM,
                Num_DateSaisie, StatutProd
            FROM TK_CallSFR_Panier
            WHERE IDtk_CallSFR IN ({",".join(str(i) for i in id_calls)})
              AND ModifELEM NOT LIKE '%suppr%'""",
        ) if id_calls else None
        f_sal = pool.submit(_load_salaries, db_rh, salaries_to_load)
        f_aff = pool.submit(_load_affectations, db_rh, salaries_to_load)
        rows_p = f_panier.result() if f_panier else []
        salaries = f_sal.result()
        affectations = f_aff.result()
    if id_calls and rows_p:
        # Lib_Offre via SFR_OffresProvad (cross-base avec ticket_bo donc 2e query)
        offre_ids = {_to_int(p.get("IDOffres_SFR")) for p in rows_p}
        offre_ids.discard(0)
        offre_libs: dict[int, str] = {}
        if offre_ids:
            oids = ",".join(str(i) for i in offre_ids)
            rows_off = db_adv.query(
                f"""SELECT IDOffres_SFR, Lib_Offre FROM SFR_OffresProvad
                WHERE IDOffres_SFR IN ({oids})"""
            )
            for o in rows_off:
                offre_libs[_to_int(o.get("IDOffres_SFR"))] = (o.get("Lib_Offre") or "").strip()
        for p in rows_p:
            id_call = _to_int(p.get("IDtk_CallSFR"))
            paniers.setdefault(id_call, []).append({
                "type": (p.get("TYPE") or "").strip(),
                "type_vente": _to_int(p.get("TypeVente")),
                "num": (p.get("NUM") or "").strip(),
                "num_date_saisie": _iso(p.get("Num_DateSaisie")),
                "_num_date_saisie_raw": p.get("Num_DateSaisie"),
                "statut_prod": _to_int(p.get("StatutProd")),
                "lib_offre": offre_libs.get(_to_int(p.get("IDOffres_SFR")), ""),
            })

    _enrich_vendeur_distrib(salaries, affectations, db_rh)
    # Batch premiers contrats (1 query au lieu de N)
    sal_non_distrib_ids = {
        sid for sid in salaries_to_load
        if not salaries.get(sid, {}).get("vendeur_distrib")
    }
    premiers_contrats = _load_premiers_contrats(db_adv, sal_non_distrib_ids)

    # Construction finale
    out: list[dict] = []
    for id_tk, tk in by_id.items():
        call = call_by_id_tk.get(id_tk)
        if not call:
            continue
        sal = salaries.get(call["id_salarie"], {"nom": "", "prenom": "", "vendeur_distrib": False})
        aff = affectations.get(call["id_salarie"], {"lib_equipe": "", "lib_orga": ""})
        panier = paniers.get(call["id_call"], [])

        nb_fibre_valide = 0
        nb_mobile_valide = 0
        offres_fibre_txt = ""
        lib_statut_force = tk["lib_statut"]
        delai_depasse = False

        for off in panier:
            if off["statut_prod"] in (1, 3):
                if off["type"] == "FIBRE":
                    nb_fibre_valide += 1
                    offres_fibre_txt += f"\n{off['lib_offre']}"
                    if off["num"]:
                        lib_statut_force = "Tk Call - Num BS SFR renseigné"
                    if off["type_vente"] in (1, 2):
                        offres_fibre_txt += " - CQ"
                    else:
                        offres_fibre_txt += " - Mig"
                elif off["type"] == "MOBILE":
                    nb_mobile_valide += 1
            # Delai prise num : si Num_DateSaisie >= date_crea + 1h
            # (transposition exacte WinDev : DateHeureDifference -> Milieu(9,2) >= 1)
            dt_saisie = _parse_dt(off.get("_num_date_saisie_raw"))
            if dt_saisie is not None:
                dt_crea = _parse_dt(tk["date_crea"])
                if dt_crea is not None and (dt_saisie - dt_crea).total_seconds() >= 3600:
                    delai_depasse = True

        premier_contrat = False
        if not sal["vendeur_distrib"]:
            date_avant = _iso(tk["date_crea"])[:10].replace("-", "")
            min_date = premiers_contrats.get(call["id_salarie"], "")
            premier_contrat = (not min_date) or (min_date >= date_avant)

        out.append({
            "id": _str_id(id_tk),
            "date_crea": _iso(tk["date_crea"]),
            "nom_client": _format_nom_client(
                call["civilite"], call["nom"], call["nom_marital"], call["prenom"]
            ),
            "cp": call["cp"],
            "ville": call["ville"],
            "nom_vendeur": f"{sal['nom']} {_capitalize(sal['prenom'])}".strip(),
            # "agence" = libelle complet "PARENT_Lib => Lib" (transposition WinDev :
            # affectationVendeurByDate renvoie le format hierarchique).
            "agence": aff.get("lib_equipe") or aff.get("lib_orga", ""),
            "lib_statut": lib_statut_force,
            "ref_appel": call["ref_appel"],
            "nb_offres": len(panier),
            "nb_fibre_valide": nb_fibre_valide,
            "nb_mobile_valide": nb_mobile_valide,
            "col_offres_fibre": offres_fibre_txt.strip(),
            "vendeur_distrib": sal["vendeur_distrib"],
            "premier_contrat": premier_contrat,
            "delai_depasse": delai_depasse,
            # On expose le panier brut pour le calcul de stats par agence dans la
            # foulee (eviter de re-quetter).
            "_panier": panier,
            "_id_salarie": call["id_salarie"],
            "_lib_orga": aff["lib_orga"],
        })

    return out


# --- Stats : 4 compteurs globaux + stats par agence -----------------------

def compute_stats(tickets_traites: list[dict], db_rh=None) -> dict:
    """Calcule les 4 compteurs principaux + les stats par agence interne/externe.

    Prend en entree la liste des tickets traites (telle que retournee par
    list_tickets_traites). Itere une fois, ne refait pas de query.
    """
    if db_rh is None:
        db_rh = get_connection("rh")

    paniers_valides = 0
    offres_fibre_thd = 0
    cq_fibre_valides = 0
    mobiles_valides = 0

    # Charger tous les organigrammes "descendants" des agences hardcodees pour
    # savoir si un vendeur appartient a l'une d'elles. On charge largement.
    agence_to_descendants: dict[int, set[int]] = {}
    all_agence_ids = {a[0] for a in AGENCES_INTERNES} | {ID_ORGA_POWER, ID_ORGA_FOX}
    for id_orga in all_agence_ids:
        agence_to_descendants[id_orga] = _orga_descendants(db_rh, id_orga)

    nb_fibre_internes: dict[int, int] = {a[0]: 0 for a in AGENCES_INTERNES}
    nb_mobile_internes: dict[int, int] = {a[0]: 0 for a in AGENCES_INTERNES}
    nb_fibre_power = 0
    nb_mobile_power = 0
    nb_fibre_fox = 0
    nb_mobile_fox = 0

    # Resolver orga d'affectation par vendeur
    aff = _load_affectations(db_rh, {t["_id_salarie"] for t in tickets_traites})

    for t in tickets_traites:
        id_orga_vendeur = aff.get(t["_id_salarie"], {}).get("id_orga", 0)
        for off in t.get("_panier", []):
            if off["statut_prod"] not in (1, 3):
                continue
            # WinDev : ComptPanierValidtest++ est UNIQUEMENT dans le cas FIBRE
            # (pas pour MOBILE). C'est compteur des "paniers Fibre valides".
            if off["type"] == "FIBRE":
                paniers_valides += 1
                if off["num"]:
                    offres_fibre_thd += 1
                    # Attribution par agence
                    for id_agence in nb_fibre_internes:
                        if id_orga_vendeur in agence_to_descendants[id_agence]:
                            nb_fibre_internes[id_agence] += 1
                            break
                    else:
                        if id_orga_vendeur in agence_to_descendants[ID_ORGA_POWER]:
                            nb_fibre_power += 1
                        elif id_orga_vendeur in agence_to_descendants[ID_ORGA_FOX]:
                            nb_fibre_fox += 1
                if off["type_vente"] in (1, 2):
                    cq_fibre_valides += 1
            elif off["type"] == "MOBILE":
                mobiles_valides += 1
                for id_agence in nb_mobile_internes:
                    if id_orga_vendeur in agence_to_descendants[id_agence]:
                        nb_mobile_internes[id_agence] += 1
                        break
                else:
                    if id_orga_vendeur in agence_to_descendants[ID_ORGA_POWER]:
                        nb_mobile_power += 1
                    elif id_orga_vendeur in agence_to_descendants[ID_ORGA_FOX]:
                        nb_mobile_fox += 1

    # Enrichit avec lib_orga reel + gimmick_url (Societe.GUIMMICK)
    agences_meta = _load_agences_meta(db_rh)

    return {
        "paniers_valides": paniers_valides,
        "offres_fibre_thd": offres_fibre_thd,
        "cq_fibre_valides": cq_fibre_valides,
        "mobiles_valides": mobiles_valides,
        "agences_internes": [
            {
                "id_orga": str(id_orga),
                "lib_orga": agences_meta.get(id_orga, {}).get("lib_orga") or lib_default,
                "nb_fibre": nb_fibre_internes.get(id_orga, 0),
                "nb_mobile": nb_mobile_internes.get(id_orga, 0),
                "gimmick_url": agences_meta.get(id_orga, {}).get("gimmick_url", ""),
            }
            for (id_orga, lib_default) in AGENCES_INTERNES
        ],
        "nb_fibre_power": nb_fibre_power,
        "nb_mobile_power": nb_mobile_power,
        "nb_fibre_fox": nb_fibre_fox,
        "nb_mobile_fox": nb_mobile_fox,
    }


# Cache module : organigramme complet (parents -> enfants) charge 1 fois,
# expire au bout de 5 minutes. Evite le N+1 BFS pour chaque appel a la page.
_ORGA_CHILDREN_CACHE: dict[int, list[int]] | None = None
_ORGA_CACHE_AT: float = 0.0
_ORGA_CACHE_TTL = 300.0  # 5 minutes

# Cache module : TK_Statut (referentiel statique, 39 rows).
_TK_STATUT_CACHE: dict[int, str] | None = None
_TK_STATUT_CACHE_AT: float = 0.0
_TK_STATUT_CACHE_TTL = 600.0  # 10 minutes

# Cache module : libelles + gimmicks (logos Societe) des agences internes.
# Les ids sont hardcodes (AGENCES_INTERNES), les libelles changent rarement.
_AGENCES_META_CACHE: dict[int, dict] | None = None
_AGENCES_META_AT: float = 0.0
_AGENCES_META_TTL = 600.0  # 10 minutes


def _load_agences_meta(db_rh) -> dict[int, dict]:
    """Charge {id_orga: {"lib_orga", "gimmick_url"}} pour les agences internes.

    Reprend la logique WinDev : Organigramme.IdSte -> Societe.GUIMMICK
    (image binaire) encode en base64 -> data URL utilisable cote frontend.
    Cache 10min.
    """
    global _AGENCES_META_CACHE, _AGENCES_META_AT
    now = time.monotonic()
    if _AGENCES_META_CACHE is not None and now - _AGENCES_META_AT < _AGENCES_META_TTL:
        return _AGENCES_META_CACHE

    ids = [a[0] for a in AGENCES_INTERNES]
    ids_sql = ",".join(str(i) for i in ids)
    orga_rows = db_rh.query(
        f"SELECT IDOrganigramme, Lib_ORGA, IdSte FROM Organigramme WHERE IDOrganigramme IN ({ids_sql})"
    )
    orga_to_lib: dict[int, str] = {}
    ste_to_orga: dict[int, int] = {}
    for r in orga_rows:
        id_orga = _to_int(r.get("IDOrganigramme"))
        id_ste = _to_int(r.get("IdSte"))
        orga_to_lib[id_orga] = (r.get("Lib_ORGA") or "").strip()
        if id_ste:
            ste_to_orga[id_ste] = id_orga

    gimmicks_by_orga: dict[int, str] = {}
    if ste_to_orga:
        ste_ids_sql = ",".join(str(i) for i in ste_to_orga.keys())
        ste_rows = db_rh.query(
            f"SELECT IdSte, GUIMMICK FROM Societe WHERE IdSte IN ({ste_ids_sql})"
        )
        for s in ste_rows:
            id_ste = _to_int(s.get("IdSte"))
            raw = s.get("GUIMMICK")
            if not raw:
                continue
            # Le pont HFSQL renvoie en general des binaires deja en string b64
            # (json-friendly). Mais on couvre aussi bytes/memoryview au cas ou.
            if isinstance(raw, memoryview):
                raw = bytes(raw)
            if isinstance(raw, (bytes, bytearray)):
                b64 = base64.b64encode(raw).decode("ascii")
            else:
                b64 = str(raw)
            id_orga = ste_to_orga.get(id_ste, 0)
            if id_orga:
                # Type d'image inconnu (BMP, PNG, JPG selon stockage WinDev).
                # 'image/*' marche dans le navigateur grace au content sniffing.
                gimmicks_by_orga[id_orga] = f"data:image/png;base64,{b64}"

    _AGENCES_META_CACHE = {
        id_orga: {
            "lib_orga": orga_to_lib.get(id_orga, ""),
            "gimmick_url": gimmicks_by_orga.get(id_orga, ""),
        }
        for id_orga in ids
    }
    _AGENCES_META_AT = now
    return _AGENCES_META_CACHE


def _load_tk_statut(db_ticket) -> dict[int, str]:
    """Charge TK_Statut (referentiel statique). Cache 10 minutes."""
    global _TK_STATUT_CACHE, _TK_STATUT_CACHE_AT
    now = time.monotonic()
    if _TK_STATUT_CACHE is not None and now - _TK_STATUT_CACHE_AT < _TK_STATUT_CACHE_TTL:
        return _TK_STATUT_CACHE
    rows = db_ticket.query("SELECT IDTK_Statut, Lib_Statut FROM TK_Statut")
    _TK_STATUT_CACHE = {
        _to_int(r.get("IDTK_Statut")): (r.get("Lib_Statut") or "").strip()
        for r in rows
    }
    _TK_STATUT_CACHE_AT = now
    return _TK_STATUT_CACHE


def _load_orga_children_map(db_rh) -> dict[int, list[int]]:
    """Charge la table Organigramme complete et construit la map parent->enfants.

    Cache 5 minutes. Pour 1 appel a la page on charge 1 fois maxi.
    """
    global _ORGA_CHILDREN_CACHE, _ORGA_CACHE_AT
    now = time.monotonic()
    if (
        _ORGA_CHILDREN_CACHE is not None
        and now - _ORGA_CACHE_AT < _ORGA_CACHE_TTL
    ):
        return _ORGA_CHILDREN_CACHE
    rows = db_rh.query(
        "SELECT IDOrganigramme, IdPARENT FROM Organigramme"
    )
    children: dict[int, list[int]] = {}
    for r in rows:
        parent = _to_int(r.get("IdPARENT"))
        child = _to_int(r.get("IDOrganigramme"))
        if parent and child:
            children.setdefault(parent, []).append(child)
    _ORGA_CHILDREN_CACHE = children
    _ORGA_CACHE_AT = now
    return children


def _orga_descendants(db_rh, id_orga_racine: int) -> set[int]:
    """Retourne l'ID racine + tous ses descendants dans Organigramme.

    Equivalent de la fonction WinDev `ListeOrgaComplet(id, Vrai)`.
    Utilise un cache module pour eviter N queries par appel.
    """
    children = _load_orga_children_map(db_rh)
    out: set[int] = {id_orga_racine}
    to_visit: list[int] = [id_orga_racine]
    while to_visit:
        cur = to_visit.pop(0)
        for c in children.get(cur, []):
            if c not in out:
                out.add(c)
                to_visit.append(c)
    return out


# --- Endpoint unifie : tout en 1 seul appel -------------------------------

def get_last_modif_call_fibre() -> str:
    """Token de change-detection sur les tickets Call Fibre.

    Concatene :
    - Max(ModifDate) sur TK_Liste + TK_CallSFR (7 derniers jours) -> couvre
      creation / cloture / changement de statut.
    - Hash du contenu de `SuiviTicketCall` (IdEncours + IdAppelEnCours) ->
      couvre le verrou ope (l'exe externe MAJ uniquement cette table quand
      un ope prend / lache un appel, sans toucher au ModifDate de TK_Liste).

    Format : "MAX_MODIF#HASH". On compare avec `!=` cote wait_for_change
    (pas `>`) car le hash change non-monotonique.
    """
    import hashlib
    from datetime import timedelta
    cutoff = (datetime.now() - timedelta(days=7)).strftime("%Y%m%d000000000")
    db_ticket = get_connection("ticket")
    db_bo = get_connection("ticket_bo")
    db_divers = get_connection("divers")
    r1 = db_ticket.query(
        """SELECT MAX(ModifDate) AS max_modif FROM TK_Liste
        WHERE IDTK_TypeDemande = ?
          AND ModifDate > ?""",
        (IDTK_TYPE_DEMANDE_CALL_FIBRE, cutoff),
    )
    r2 = db_bo.query(
        """SELECT MAX(ModifDate) AS max_modif FROM TK_CallSFR
        WHERE ModifDate > ?""",
        (cutoff,),
    )
    m1 = _iso((r1[0] or {}).get("max_modif")) if r1 else ""
    m2 = _iso((r2[0] or {}).get("max_modif")) if r2 else ""
    max_modif = max(m1, m2)

    # Capture le verrou ope (IdEncours/IdAppelEnCours change quand un ope
    # prend ou lache un appel).
    r3 = db_divers.query(
        "SELECT IdEncours, IdAppelEnCours FROM SuiviTicketCall WHERE TypeCall = 'FIBRE'"
    )
    suivi_blob = ""
    if r3:
        s = r3[0] or {}
        suivi_blob = (s.get("IdEncours") or "") + "|" + (s.get("IdAppelEnCours") or "")
    suivi_hash = hashlib.md5(suivi_blob.encode("utf-8", errors="ignore")).hexdigest()[:12]
    return f"{max_modif}#{suivi_hash}"


def wait_for_change(since: str, timeout_seconds: float = 25, poll_interval: float = 1.0) -> tuple[bool, str]:
    """Long polling : attend qu'un changement survienne sur Call Fibre.

    Renvoie (changed, latest_token).
    - changed = True si le token (max_modif#hash_suivi) a change avant timeout.
    - changed = False si timeout atteint.

    On compare avec `!=` (pas `>`) car la partie hash n'est pas monotone.

    Si `since` est vide, retourne immediatement (changed=True) pour forcer
    un chargement initial.
    """
    if not since:
        return True, get_last_modif_call_fibre()

    deadline = time.monotonic() + timeout_seconds
    latest = ""
    while True:
        latest = get_last_modif_call_fibre()
        if latest and latest != since:
            return True, latest
        if time.monotonic() >= deadline:
            return False, latest
        time.sleep(poll_interval)


def load_page_en_cours(user_id: int, user_id_poste: int) -> dict:
    """Charge UNIQUEMENT le tableau du haut + token last_modif.

    Rapide (~5 queries) -> affichage immediat de la page.
    Le tableau du bas + stats sont charges separement (load_page_traites).
    """
    en_cours = list_tickets_en_cours(user_id, user_id_poste)
    return {
        "tickets_en_cours": en_cours,
        "serveur_now": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "last_modif": get_last_modif_call_fibre(),
    }


def export_traites_xlsx(jour: str | None = None, traites: list[dict] | None = None) -> bytes:
    """Genere un fichier .xlsx du tableau des tickets traites du jour.

    Si `traites` est fourni (par exemple recu en POST depuis le frontend qui
    a deja les donnees en cache), on saute la requete HFSQL (~5s).
    Sinon on recharge tout via list_tickets_traites().

    Coloration des lignes identique a l'UI :
    - delai_depasse : ROUGE (priorite max, match WinDev)
    - vendeur_distrib : GRIS
    - premier_contrat : VERT
    - default : blanc
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    if traites is None:
        traites = list_tickets_traites(jour)
    j_label = (jour or _date.today().isoformat()).replace("-", "/")

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets traites"

    headers = [
        "Commande faite le", "Client", "CP", "Ville", "Commercial",
        "Agence", "Etat", "NB Offres", "NB Fibre Valide", "NB Mobile Valide",
        "Ref Appel", "Offres Fibre",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Couleurs (match UI bg-red-100 / bg-gray-100 / bg-green-100)
    fill_red = PatternFill("solid", fgColor="FECACA")    # ~ red-200
    fill_gray = PatternFill("solid", fgColor="E5E7EB")   # ~ gray-200
    fill_green = PatternFill("solid", fgColor="BBF7D0")  # ~ green-200

    for t in traites:
        row = [
            _iso(t.get("date_crea"))[:16] if t.get("date_crea") else "",
            t.get("nom_client", ""),
            t.get("cp", ""),
            t.get("ville", ""),
            t.get("nom_vendeur", ""),
            t.get("agence", ""),
            t.get("lib_statut", ""),
            t.get("nb_offres", 0),
            t.get("nb_fibre_valide", 0),
            t.get("nb_mobile_valide", 0),
            t.get("ref_appel", ""),
            t.get("col_offres_fibre", ""),
        ]
        ws.append(row)
        # Coloration ligne (priorite WinDev : rouge > gris > vert)
        if t.get("delai_depasse"):
            fill = fill_red
        elif t.get("vendeur_distrib"):
            fill = fill_gray
        elif t.get("premier_contrat"):
            fill = fill_green
        else:
            fill = None
        if fill is not None:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = fill

    # Largeurs de colonnes approximatives
    widths = [17, 28, 8, 22, 22, 50, 28, 11, 16, 18, 14, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Titre dans la ligne 0 (insertion)
    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=f"Tickets Call Fibre traites du {j_label}").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.freeze_panes = "A3"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def load_page_traites(jour: str | None = None) -> dict:
    """Charge le tableau du bas + stats agences.

    Plus lent (paniers, offres, agences, ~10 queries). Appele en background
    par le frontend apres l'affichage des en cours.
    """
    traites = list_tickets_traites(jour)
    stats = compute_stats(traites)
    traites_clean = [{k: v for k, v in t.items() if not k.startswith("_")} for t in traites]
    return {
        "tickets_traites": traites_clean,
        "stats": stats,
    }


def load_page(user_id: int, user_id_poste: int, jour: str | None = None) -> dict:
    """Charge tout (en cours + traites + stats). Garde pour compatibilite.

    Pref : load_page_en_cours() + load_page_traites() pour servir en 2 temps.
    """
    p1 = load_page_en_cours(user_id, user_id_poste)
    p2 = load_page_traites(jour)
    return {**p1, **p2}
