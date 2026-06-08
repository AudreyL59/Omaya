"""
SDTC : export XLS (openpyxl).

Une feuille 'Contrats' qui reprend la TableContratTOT WinDev :
concatenation des contrats deja traites coches + selection SDTC, avec
flag STC sur la selection.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="17494E", end_color="17494E", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_STC_FILL = PatternFill(start_color="EFE9E7", end_color="EFE9E7", fill_type="solid")

COLUMNS = [
    ("stc", "STC", 6),
    ("partenaire", "Partenaire", 12),
    ("client_nom", "Client", 28),
    ("client_ville", "Ville", 18),
    ("client_cp", "CP", 8),
    ("lib_produit", "Lib Produit", 30),
    ("type_prod", "Type", 14),
    ("num_bs", "N° BS", 16),
    ("date_signature", "Date sign.", 12),
    ("mois_paiement", "Mois P.", 10),
    ("type_etat_lib", "Type État", 18),
    ("etat_contrat_lib", "État contrat", 20),
    ("nb_points", "Points", 8),
]


def _date_fr(iso: str) -> str:
    if not iso or len(iso) < 10:
        return ""
    return f"{iso[8:10]}/{iso[5:7]}/{iso[0:4]}"


def _mois_fr(iso: str) -> str:
    if not iso or len(iso) < 7:
        return ""
    return f"{iso[5:7]}/{iso[0:4]}"


def build_workbook(
    contrats_traites: list[dict],
    contrats_sdtc: list[dict],
    lib_salarie: str = "",
) -> bytes:
    """Genere le workbook XLSX et retourne les bytes.

    Les contrats SDTC ont le flag STC=Oui, les traites STC=Non.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Contrats"

    # En-tete
    for i, (_, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    if lib_salarie:
        ws.oddHeader.center.text = f"Solde de tout compte - {lib_salarie}"

    # Lignes
    def _write_row(row_idx: int, c: dict, is_stc: bool) -> None:
        for col_idx, (key, _, _w) in enumerate(COLUMNS, start=1):
            if key == "stc":
                v = "Oui" if is_stc else "Non"
            elif key == "date_signature":
                v = _date_fr(str(c.get(key) or ""))
            elif key == "mois_paiement":
                v = _mois_fr(str(c.get(key) or ""))
            else:
                v = c.get(key) if c.get(key) is not None else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            if is_stc:
                cell.fill = _STC_FILL

    row = 2
    for c in contrats_traites:
        _write_row(row, c, is_stc=False)
        row += 1
    for c in contrats_sdtc:
        _write_row(row, c, is_stc=True)
        row += 1

    # Freeze + filtre
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
